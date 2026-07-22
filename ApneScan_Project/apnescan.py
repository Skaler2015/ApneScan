# -*- coding: utf-8 -*-
"""
ApneScan - v18  [31 naye public features: bundle auto-split, auto-orient, auto-colour,
                 scanner discovery, book/business-card/photo-restore modes, PDF tools
                 (editor, sign/stamp, Word/Excel/JPG, watermark, unlock, archival),
                 tags + instant OCR-index search, self-installing auto-update,
                 portable mode, settings export/import, crash reporter, touch mode,
                 4 nayi bhashayein (partial)]
=========================================================
Run with 32-bit Python (HP TWAIN driver is 32-bit):
    py -3.12-32 scanner_app_v10.py
One-click scan shortcut:
    py -3.12-32 scanner_app_v10.py --scan --profile "Documents"
"""

import io
import os
import re
import sys
import json
import shutil
import socket
import tempfile
import traceback
import datetime

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog

try:
    from PyQt5.QtNetwork import QNetworkAccessManager, QNetworkRequest
    HAS_QTNET = True
except Exception:
    HAS_QTNET = False

try:
    from PyQt5.QtMultimedia import QCamera, QCameraInfo, QCameraImageCapture
    from PyQt5.QtMultimediaWidgets import QCameraViewfinder
    HAS_CAMERA = True
except Exception:
    HAS_CAMERA = False

try:
    import twain
    HAS_TWAIN = True
except Exception:
    HAS_TWAIN = False

from PIL import Image, ImageChops, ImageEnhance, ImageOps, ImageDraw, ImageFont, ImageFilter

try:
    import numpy as np
    HAS_NUMPY = True
except Exception:
    HAS_NUMPY = False

try:
    import pytesseract
    from pypdf import PdfReader, PdfWriter
    HAS_OCR_LIBS = True
except Exception:
    HAS_OCR_LIBS = False

try:
    import fitz  # PyMuPDF — best PDF page rendering (optional)
    HAS_FITZ = True
except Exception:
    HAS_FITZ = False


def pdf_to_images(pdf_path, tmpdir, dpi=200):
    """Turn each page of a PDF into a PNG (for import). Uses PyMuPDF if installed
    (renders any PDF), else falls back to extracting embedded images with pypdf
    (works for scanned PDFs). Returns a list of PNG paths."""
    out = []
    if HAS_FITZ:
        try:
            doc = fitz.open(pdf_path)
            zoom = dpi / 72.0
            mat = fitz.Matrix(zoom, zoom)
            for page in doc:
                pix = page.get_pixmap(matrix=mat)
                fd, png = tempfile.mkstemp(suffix=".png", dir=tmpdir); os.close(fd)
                pix.save(png)
                out.append(png)
            doc.close()
            if out:
                return out
        except Exception:
            out = []
    # fallback: pull the biggest embedded image from each page (scanned PDFs)
    try:
        from pypdf import PdfReader as _PR
        reader = _PR(pdf_path)
        for page in reader.pages:
            try:
                imgs = list(getattr(page, "images", []))
            except Exception:
                imgs = []
            if not imgs:
                continue
            big = max(imgs, key=lambda im: len(getattr(im, "data", b"")))
            fd, png = tempfile.mkstemp(suffix=".png", dir=tmpdir); os.close(fd)
            with open(png, "wb") as fh:
                fh.write(big.data)
            try:
                with Image.open(png) as im2:
                    im2.convert("RGB").save(png, "PNG")
            except Exception:
                pass
            out.append(png)
    except Exception:
        pass
    return out


def _configure_tesseract():
    """Point pytesseract at a Tesseract install even if it's not on PATH."""
    if not HAS_OCR_LIBS:
        return
    try:
        home = os.path.expanduser("~")
        for c in (r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                  r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
                  os.path.join(home, "AppData", "Local", "Tesseract-OCR", "tesseract.exe"),
                  os.path.join(home, "AppData", "Local", "Programs", "Tesseract-OCR", "tesseract.exe")):
            if os.path.exists(c):
                pytesseract.pytesseract.tesseract_cmd = c
                break
    except Exception:
        pass


_TESS_OK = None      # cache: get_tesseract_version() har baar tesseract.exe ko
                     # subprocess me chalata hai (~0.5-2s) — isse UI atak jati
                     # thi (khaaskar rename ke baad). Ek baar check, phir yaad.
def tesseract_available():
    global _TESS_OK
    if _TESS_OK is not None:
        return _TESS_OK
    if not HAS_OCR_LIBS:
        _TESS_OK = False
        return False
    try:
        pytesseract.get_tesseract_version()
        _TESS_OK = True
    except Exception:
        _TESS_OK = False
    return _TESS_OK


_configure_tesseract()

try:
    import openpyxl
    HAS_XLSX = True
except Exception:
    HAS_XLSX = False

try:
    from pyzbar.pyzbar import decode as _zbar_decode
    HAS_ZBAR = True
except Exception:
    HAS_ZBAR = False

try:
    import win32com.client as _w32
    import pythoncom as _pythoncom
    HAS_W32 = True
except Exception:
    HAS_W32 = False


APP_NAME = "ApneScan"
VERSION = "97"
UPDATE_API = "https://api.github.com/repos/Skaler2015/ApneScan/releases/latest"
DOWNLOAD_PAGE = "https://github.com/Skaler2015/ApneScan/releases/latest"
# App ko phailane (share/QR/poster) ke liye
WEBSITE_URL = "https://apnescan.apnesoft.com"
GITHUB_URL = "https://github.com/Skaler2015/ApneScan"
SHARE_TEXT = ("ApneScan — bilkul FREE document scanner software (Windows). "
              "Scan to PDF, Hindi+English OCR, PDF compress (200KB), WhatsApp share. "
              "No ads. Download: " + WEBSITE_URL)
def _portable_dir():
    """PORTABLE MODE: exe ke saath 'portable.txt' naam ki khaali file rakh do —
    saari settings wahi folder me rahengi (pen-drive se chalao, settings saath)."""
    try:
        base = os.path.dirname(sys.executable if getattr(sys, "frozen", False)
                               else os.path.abspath(__file__))
        if os.path.exists(os.path.join(base, "portable.txt")):
            return base
    except Exception:
        pass
    return None


_PORTABLE = _portable_dir()
CONFIG_PATH = (os.path.join(_PORTABLE, "apnescan_config.json") if _PORTABLE
               else os.path.join(os.path.expanduser("~"), ".apnescan.json"))
CRASH_PATH = os.path.join(os.path.expanduser("~"), "apnescan_crash.txt")
TRASH_DIR = os.path.join(os.path.expanduser("~"), ".apnescan_trash")   # Recycle Bin
PSTATS_PATH = os.path.join(os.path.expanduser("~"), ".apnescan_pstats.json")
_OLD_CONFIG = os.path.join(os.path.expanduser("~"), ".noble_doc_scanner.json")
if not os.path.exists(CONFIG_PATH) and os.path.exists(_OLD_CONFIG):
    try:
        import shutil as _sh
        _sh.copy2(_OLD_CONFIG, CONFIG_PATH)
    except Exception:
        pass
SCANNER_PORTS = (80, 443, 8080, 9100)

# Built-in ApneScan worldwide-stats endpoint (Google Apps Script). Every install
# reports scan COUNTS here (never any document/patient data).
DEFAULT_STATS_URL = "https://script.google.com/macros/s/AKfycbzwh2HQHXKe_09wgXpSh-NZQu-GbR7N-NVmtUEyRpl8oOa-MdoJ9ShqHB_i0QCcuqe2_w/exec"

_TESS_GUESSES = [
    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
]
if HAS_OCR_LIBS:
    for _p in _TESS_GUESSES:
        if os.path.exists(_p):
            pytesseract.pytesseract.tesseract_cmd = _p
            break

def resource_path(name):
    """Path to a bundled data file, works both from source and PyInstaller .exe."""
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(sys.argv[0])))
    p = os.path.join(base, name)
    if os.path.exists(p):
        return p
    return os.path.join(os.path.dirname(os.path.abspath(sys.argv[0])), name)


# Embedded app icon (base64 PNG) so the ApneScan icon always shows in the title
# bar + taskbar even if the icon file wasn't bundled with the build.
_EMBEDDED_ICON_B64 = """\
iVBORw0KGgoAAAANSUhEUgAAAgAAAAIACAYAAAD0eNT6AAAQoklEQVR4nO3dS48c5d3G4Rqrl0gzlpA/iXdss3bWwQokMZFickAG
EnbhsCOBOOTEJuEYJ+v4O7DjkyAktyX2k4XduMeemT5V1XO4r0tqvYr0ClU/7q7/r57q6T4aqMrx735zWvoYAKbw8A9/OSp9DDzh
H2Nmx7/9tQEPcI6Hf/yrmTQjiz0hwx7gMKJgOhZ2RAY+wLQEwXgs5AEMfICyBMH+LNyODH2AOomB3VisLRj6AG0RA5tZoEsY/ABt
EwIXszBPMfQB+iQGzrIYjxn8ABmEwCPxi2DwA2RKD4HYJ2/wAzAMuSEQ96QNfgDOkxYCMU/W4AdgGykh0P2TNPgB2EfvIdDtkzP4
ARhDryFwpfQBTMHwB2Asvc6Urqqm138kAOrQ025AF0/E4AdgTj2EQPO3AAx/AObWw+xpOgB6+AcAoE2tz6AmtzBaX3QA+tLiLYHm
Dvj4TcMfgPo8/KCtCGjqFoDhD0CtWptRTdRKa4sKQLYWdgOqP8DjN39l+APQnIcf/K3qGVv1LQDDH4BW1T7Dqg2A2hcOADapeZZV
GQA1LxgA7KLWmVZdANS6UACwrxpnWzUfUKhxcQBgbLV8OLCKHQDDH4AUtcy84gFQy0IAwFxqmH1FA6CGBQCAEkrPwGIBUPqJA0Bp
JWdh8VsAAMD8igSAq38AeKTUTJw9AAx/ADirxGycNQAMfwA439wzcrYAMPwB4HJzzspZAsDwB4DtzDUzJw8Awx8AdjPH7PRngAAQ
aNIfJDh+w9U/AOzr4YfT/XDQZDsAhj8AHGbKWTpJABj+ADCOqWaqzwAAQKDRA8DVPwCMa4rZOmoAGP4AMI2xZ+xony48fuOXhj8A
TOzhh38fZXb7DAAABBolAFz9A8A8xpq5dgAAINDBAeDqHwDmNcbsPSgADH8AKOPQGewWAAAE2jsAXP0DQFmHzGI7AAAQaK8AcPUP
AHXYdybbAQCAQDsHgKt/AKjLPrPZDgAABNopAFz9A0Cddp3RdgAAINDWAeDqHwDqtsusXmz9XzX+AaAbW+0AHL/u6h8AWrDtzPYZ
AAAItDEAXP0DQFu2md12AAAgkAAAgECXBoDtfwBo06YZbgcAAAJdGACu/gGgbZfN8ku+CMj8B4BeuQUAAIHODYDj1191+Q8AHbho
ptsBAIBAAgAAAj0TALb/AaAv5812OwAAEEgAAEAgAQAAgc4EgPv/ANCnp2e8HQAACCQAACDQ2d8CcAMAACJ8vwNwfMf9fwDo2fqs
dwsAAAIJAAAIJAAAIJAAAIBAV4bBBwABIMVq5tsBAIBAAgAAAgkAAAgkAAAgkAAAgECPfwvAHwEAQBI7AAAQ6Oj4zm2X/wAQxg4A
AAQSAAAQSAAAQCABAACBBAAABBIAABBo4TuAACCPHQAACCQAACCQAACAQAIAAAIJAAAIJAAAIJAAAIBAAgAAAgkAAAgkAAAgkAAA
gEACAAACLUofAKxb3v1H6UOgIV8/+Kb0IczmB++9U/oQ6IwAoDhDH2B+i8HvAVPI8u7HpQ8BGuJczbh8BoAiDH/GcP3qtdKHAM0S
AMzO8GdMIgD2IwCYleHPFEQA7E4AMBvDnymJANiNAGAWhj9zEAGwPQHA5Ax/5iQCYDsLf1nClJZ/NvyZ3/Wr1/r7kiDnakZmBwDo
kp0AuJwAYDKu/ilNBMDFBADQNREA5xMAQPdEADxLADAJ2//URgTAWQIAiCEC4AkBAEQRAfCIAADiiAAQAEAoEUA6AQDEEgEkEwBA
NBFAKgEAxBMBJBIAAIMIII8AAHhMBJBEAACsEQGkEAAATxEBJBAAAOcQAfRuMQynpY8BoErXr14bvn7wTenDeMy5mnHZAQC4hJ0A
eiUAADYQAfRoYVcJYLPitwOcqxmZHQCALdkJoCcCAGAHIoBeCACAHYkAeiAAAPYgAmidAADYkwigZQIA4AAigFYJAIADiQBaJAAA
RiACaI0AABiJCKAlAgBgRCKAVggAgJGJAFogAAAmIAKonQAAmIgIoGYCAGBCIoBaCQCAiYkAaiQAAGYgAqiNAACYiQigJgIAYEYi
gFoIAICZiQBqIAAAChABlLYYTk9LHwNApOtXrw1fP/hmu/9n52pGZgcAoCA7AZQiAAAKEwGUIAAAKiACmJsAAKiECGBOAgCgIiKA
uQgAgMqIAOYgAAAqJAKY2qL0AQBwPhHAlOwAAEAgAQAAgQQAAAQSAAAQSAAAQCABAACBBAAABBIAABBIAABAIAEAAIEEAAAEEgAA
EEgAAECgxXBa+hAA2Mi5mpHZAQCAQAIAAAIt7CsBtMC5mnHZAQCAQAIAAAIJAAAItCh9AFDKyVuv3Sx9DLRh+f5H90ofA4xNABDH
4GdXq9eMEKAnbgEQxfDnEF4/9EQAEMPJmzF4HdELAUAEJ23G5PVEDwQAAAQSAHTP1RpT8LqidQIAAAIJAAAIJAAAIJAAAIBAAoDu
+fY2puB1ResEAAAEEgBEcLXGmLye6MGV4XQYPDxGf1TISZsxFHsdlX5Pe3T3sANAFBHAIbx+6ImfAybO6iTum9zYlsFPjwQAsZzU
gWRuAQBAIAEAAIEWjz4OCEDdnKsZlx0AAAgkAAAgkAAAgEACAAACCQAACCQAACCQAACAQAIAAAIJAAAIJAAAIJAAAIBAAgAAAgkA
AAgkAAAg0MIvTAI0wLmakdkBAIBAAgAAAgkAAAgkAAAg0KL0AUApJ2+9drP0MdCG5fsf3St9DDA2AUAcg59drV4zQoCeuAVAFMOf
Q3j90BMBQAwnb8bgdUQvrjz6dgkPj7EfdXHSZkxlXk+l39MevT3sAABAIAFA91z9MwWvK1onAAAgkAAAgEACAAACHR2/euu09EFQ
hx++c+fb0scATO9/79x9vvQxUJ4ACGfoQzYxkMstgGCGP+A8kMsOQCBveOA8dgOy2AEIY/gDF3F+yCIAgnhzA5s4T+RYHJU+AgCq
Yi5kODq57TMACW68q+qB7d1/2+cBeucWQADDH9iV80b/FqUPgPm88NyJogc2+uq7peEfwA5ACMMf2JbzRQYB0Lkb79751psZ2NUL
z5087zZA3wRA5wx/YF/OH30TAAAQSAAAQCABAACBBAAABBIAnfP3vMC+nD/6JgA6d//tu897EwO7+uq75be+Drhvi2HwUwAJvvpu
6fsAgK08uWgwH3rmq4CD2AkAYMUtgAD33/6TK39gJ84b/RMAIbyZgW05X2QQAAAQSAAEUfXAJs4TOQRAGG9u4CLOD1mOTm7/zN95
hLrx7uv+KgAw+EPZAQjmTQ84D+Q6OvmFHQAeufGeHQFIcP/3hj4CgIk8+PhfpQ8BunL19q3Sh0BnfBMgsU7eeu1m6WPoyfL9j+6V
PgZgewKAOAb/NFbrKgSgDT4ESBTDf3rWGNogAIhhMM3HWkP9BAARDKT5WXOomwAAgEACgO65Ei3H2kO9BAAABBIAABBIAABAIAEA
AIEEAN3zzXTlWHuolwAAgEBXhuF08PAY/1EXV6Lzs+ZjK/2e9ujtYQeAGAbSfKw11E8AEMVgmp41hjb4OWDirAaUb6kbl8EPbREA
xDKwgGRuAQBAoMVwWvoQANjIuZqR2QEAgEACAAACCQAACCQAACCQAACAQAIAAAIJAAAIJAAAIJAAAIBAAgAAAgkAAAgkAAAgkAAA
gEACAAACCQAACCQAACCQAACAQIthOC19DABs5FzNuOwAAEAgAQAAgQQAAARauK0E0ADnakZmBwAAAgkAAAgkAAAgkAAAgEACAAAC
CQAACCQAACCQAACAQAIAAAIJAAAIJAAAIJAAAIBAAgAAAgkAAAgkAAAgkAAAgEACAAACCQAACCQAACCQAACAQIvh9LT0MQCwiXM1
I7MDAACBBAAABBIAABBIAABAIAEAAIEEAAAEEgAAEEgAAEAgAQAAgQQAAAQSAAAQSAAAQCABAACBBAAABBIAABBIAABAIAEAAIEE
AAAEEgAAEEgAAECgxXBa+hAA2Mi5mpHZAQCAQAIAAAIt7CsBtMC5mnHZAQCAQAIAAAIJAAAIJAAAIJAAAIBAAgAAAgkAAAgkAAAg
kAAAgEACAAACCQAACCQAACCQAGASV3/+culDgG54PzEFAQAAgQQAAAQSAEzGtiUczvuIqQgAAAgkAJiUqxfYn/cPUzo6ufXSaemD
oH8P/vl56UOAplx9xfBnWnYAmIWTGWzP+4U5CABm46QGm3mfMBcBwKyc3OBi3h/MSQAwOyc5eJb3BXMTABThZAdPeD9QwtHJrR/7
KwCKevDPL0ofAhRx9ZWXSh8CwQQAVRED9M7QpxYCAAAC+QwAAAQSAAAQSAAAQCABAACBBAAABBIAABBIAABAIAEAAIEEAAAEEgAA
EEgAAEAgAQAAgRaDnwICgDh2AAAgkAAAgEACAAACCQAACCQAACDQleUnXx6VPggAYD7LT748sgMAAIEEAAAEWjz6P74NCACS2AEA
gEACAAACCQAACCQAACDQlWEYhuUn//ZdAAAQYDXz7QAAQCABAACBBAAABBIAABDo+wDwQUAA6Nv6rLcDAACBFmf+l58EAIAIdgAA
INCZAFh+6nMAANCjp2e8HQAACCQAACCQAACAQM8EgM8BAEBfzpvtdgAAIJAAAIBA5waA2wAA0IeLZrodAAAIJAAAINClW/0nP73p
1wEAoFHLT+9dOOftAABAoEsD4LJyAADqtWmG2wEAgEACAAACbQwAtwEAoC3bzG47AAAQaKsAsAsAAG3YdmbbAQCAQFsHgF0AAKjb
LrN6sdN/2fcCAkAXdroFsPzMLgAA1GjXGe0zAAAQaOcAsAsAAHXZZzbbAQCAQHsFgF0AAKjDvjPZDgAABNo7AOwCAEBZh8xiOwAA
EOigALALAABlHDqDD94BEAEAMK8xZq9bAAAQaJQAsAsAAPMYa+baAQCAQKMFgF0AAJjWmLN29KF98pMX/WgwAIxs+dl/Rp3Zo98C
GPsAASDdFLPVZwAAINAkAWAXAADGMdVMnWwHQAQAwGGmnKWT3gIQAQCwn6ln6GLK//gwDMPgbwIAoDqTfwhw+bldAADYxRyzc5a/
AhABALCduWbmbH8GKAIA4HJzzspZvwdABADA+eaekbN/EZAIAICzSszGIt8EKAIA4JFSM9FXAQNAoGIBYBcAgHQlZ2HRHQARAECq
0jOw+C2A0gsAAHOrYfYVD4BhqGMhAGAOtcy8Kg5i3cnLL/r1AAC6U8vgX6liB2BdbQsEAIeqcbZVFwDDUOdCAcA+ap1pVQbAMNS7
YACwrZpnWbUBMAx1LxwAXKb2GVZ1AAxD/QsIAE9rYXZVf4DrTl7+kb8QAKBay8//28xcrX4HYF1LCwtAltZmVFMBMAztLTAA/Wtx
NjV3wOtOXnJLAIByll+0N/hXmtsBWNfywgPQttZnUNMBMAzt/wMA0J4eZk/zT2CdWwIATKmHwb/SzRNZJwQAGFNPg3+l+VsA5+nx
HwqAMnqdKV0+qXV2AwDYR6+Df6XrJ7dOCACwjd4H/0rEk1wnBAA4T8rgX4l6suuEAADDkDf4VyKf9DohAJApdfCvRD/5dUIAIEP6
4F+xCE8RAgB9MvjPshiXEAMAbTP0L2ZhtiAEANpi8G9mgXYkBgDqZOjvxmIdQAwAlGXo78/CjUgQAEzLwB+PhZyQIAA4jIE/HQs7
M1EAcD7Dfl4WuzICAeiVAV+X/wMZzgP3ZyE7mwAAAABJRU5ErkJggg==
"""


def app_icon():
    for n in ("apnescan.ico", "apnescan_icon.png"):
        p = resource_path(n)
        if os.path.exists(p):
            return QtGui.QIcon(p)
    # fallback: decode the embedded icon
    try:
        import base64 as _b64
        pm = QtGui.QPixmap()
        pm.loadFromData(_b64.b64decode(_EMBEDDED_ICON_B64))
        if not pm.isNull():
            return QtGui.QIcon(pm)
    except Exception:
        pass
    return QtGui.QIcon()


COLOUR_MODES = {"Colour (rang)": "color", "Grayscale": "gray", "Black & White": "bw"}
RESOLUTIONS = ["150", "200", "300", "600"]

# (id, label, default key) — all reassignable via Settings -> Keyboard Shortcuts
SHORTCUTS = [
    ("scan", "Scan", "Return"),
    ("rename", "Rename page", "F2"),
    ("dpi_150", "Scan at 150 dpi", "F3"),
    ("dpi_200", "Scan at 200 dpi", "F4"),
    ("dpi_300", "Scan at 300 dpi", "F5"),
    ("dpi_600", "Scan at 600 dpi", "F6"),
    ("dpi_custom", "Scan at custom dpi", "F7"),
    ("import", "Import images / PDF", "Ctrl+I"),
    ("save_all", "Save all as PDF", "Ctrl+S"),
    ("save_sel", "Save selected as PDF", "Space"),
    ("save_pw", "Save PDF (password)", ""),
    ("save_img", "Save images", "Ctrl+Shift+S"),
    ("ocr_text", "Export OCR text", "Ctrl+Alt+O"),
    ("print", "Print", "Ctrl+P"),
    ("rotate_left", "Rotate left", "Ctrl+Left"),
    ("rotate_right", "Rotate right", "Ctrl+Right"),
    ("bright_up", "Brightness +", ""),
    ("bright_dn", "Brightness -", ""),
    ("contrast_up", "Contrast +", ""),
    ("contrast_dn", "Contrast -", ""),
    ("autocrop", "Auto-crop page", ""),
    ("autoname", "Auto-name pages (OCR)", "Ctrl+Alt+N"),
    ("undo", "Undo delete", "Ctrl+Z"),
    ("delete", "Delete page", "Delete"),
    ("clear", "Clear all", "Ctrl+Shift+Delete"),
    ("move_up", "Move page up", "Ctrl+Up"),
    ("move_down", "Move page down", "Ctrl+Down"),
    ("search", "Search past PDFs", "Ctrl+F"),
    ("merge", "Merge PDFs", ""),
    ("split", "Split into PDFs", ""),
    ("report", "Monthly report", ""),
    ("zoom_in", "Zoom in (thumbnails)", "Ctrl+="),
    ("zoom_out", "Zoom out (thumbnails)", "Ctrl+-"),
    ("options", "Options / Settings", "Ctrl+,"),
    ("profiles", "Profiles", "Ctrl+L"),
]

DEFAULT_OPTIONS = {
    "auto_save": False,
    "save_folder": os.path.join(os.path.expanduser("~"), "Documents", "NobleScans"),
    "filename_template": "{claim}_{date}_{seq}",
    "make_claim_folder": False,
    "year_month_folders": False,
    "remove_blank": False,
    "blank_sensitivity": "normal",   # kam / normal / zyada
    "auto_crop": False,
    "deskew": False,
    "quality_enhance": False,
    "clean_edges": False,        # scan ki kaali border / kinare ke chhed saaf karo
    "split_two_page": False,     # ek glass par do page → apne aap alag karo
    "searchable_pdf": False,     # har save par PDF ke andar OCR text (Ctrl+F se dhoondo)
    "compress": False,
    "jpeg_quality": 60,
    "watermark": False,
    "watermark_text": "Noble Care Hospital",
    "batch_mode": False,
    "validate_claim": False,
    "claim_pattern": r"^[A-Za-z0-9\-]{4,}$",
    "barcode_autofill": False,
    "duplicate_check": False,
    "excel_log": False,
    "activity_log": False,
    "backup": False,
    "backup_folder": os.path.join(os.path.expanduser("~"), "Documents", "NobleScans_Backup"),
    "scanner_method": "twain",   # "twain" ya "wia"
    "twain_file_xfer": False,    # experimental: continuous ADF feed (TWAIN file transfer)
    "auto_orient": False,        # ulta page OCR se seedha karo
    "auto_colour": False,        # rangeen page colour me, baki gray me (chhoti file)
    "custom_page_mm": 600,       # "custom" page size ki lambai (mm)
    "touch_mode": False,         # bade buttons/font (touch / buzurg mode)
    "sign_image": "",            # sign/stamp wali image ka path
    "tags": {},                  # pdf path -> [tags]
    "show_files_panel": True,    # daayan "Meri Files" panel
    "show_left_panel": True,     # baayan scan-settings panel
    "fav_folders": [],           # panel ke ⭐ favourite folders
    "pinned_files": [],          # panel ke 📌 pin ki hui files (upar dikhti hain)
    "files_grid": False,         # panel results grid (badi thumbnail) view
    "files_panel_root": "",      # panel me khola gaya doosra folder (khaali = save-folder)
    "files_sort": "name_asc",    # panel ki list kis tarah sort ho (user pasand)
    "sidebar_stats": [],         # khaali = default set dikhega
    "name_append_number": False, # auto-naam me document number bhi jodo
    "name_append_date": False,   # auto-naam me date bhi jodo
    "ui_dashboard": True,        # khaali screen par bade action-cards
    "ui_header": True,           # toolbar ke neeche status-patti
    "ui_graph": True,            # sidebar me 7-din ka graph
    "ui_preview": False,         # daayan preview panel
    "ui_jobs": False,            # job-chips patti
    "ui_kiosk": False,           # kiosk overlay
    "jobs": [],                  # job-chips: [{name,icon,profile,folder,template}]
    "ui_ribbon": False,          # ribbon toolbar (classic toolbar ki jagah)
    "wia_device_id": None,
    "scanner_name": "",          # abhi chuni gayi scanner ka naam (dikhane ke liye)
    "language": "en",            # "hi" ya "en" (default English)
    "simple_mode": False,
    "feedback_email": "",
    "fast_mode": False,
    "after_save": "nothing",   # nothing / open / folder
    "naps2_path": "",
    "naps2_profile": "",
    "scanner_ip": "",
    "theme": "light",             # light / dark
    "show_page_numbers": True,
    "save_images_too": False,     # also save each page as an image
    "auto_name": False,           # OCR the document title as the page label
    "shortcuts": {},              # id -> custom key (overrides default)
    # ---- v73: 40+ UI-customize settings (live-preview supported) ----
    "accent": "teal",             # primary button colour (teal/blue/green/purple/orange/rose)
    "font_scale": 100,            # app-wide font size percent (70-180)
    "ui_animations": True,        # small animations/transitions
    "ui_tooltips": True,          # hover help-tips
    "start_maximized": False,     # app open hote hi poori screen
    "toolbar_style": "icon_text", # icon_only / text_only / icon_text
    "thumb_size": 150,            # page thumbnail width (px)
    "ui_spacing": "normal",       # compact / normal / roomy (page grid gap)
    "left_panel_w": 252,          # baayan scan-settings panel width
    "files_panel_w": 250,         # daayan Meri Files panel width
    "preview_panel_w": 310,       # preview panel width
    "ui_filmstrip": True,         # preview panel me niche filmstrip
    "ui_analytics": True,         # sidebar me analytics card
    "ui_confirm_delete": True,    # delete se pehle pucho
    "footer_folder": True,        # status-bar: folder
    "footer_pages": True,         # status-bar: pages/selection
    "footer_last": True,          # status-bar: last saved file
    "footer_disk": True,          # status-bar: free disk
    "footer_version": True,       # status-bar: version
    "footer_scanner": True,       # status-bar: scanner/busy
    # ---- v74: aur customize (presets, branding, footer-extra, window…) ----
    "ui_corners": "rounded",      # rounded / sharp
    "high_contrast": False,       # saaf gaadhe rang (kam nazar ke liye)
    "window_opacity": 100,        # poori window ki paardarshita (60-100)
    "brand_name": "",             # header me hospital/clinic ka naam
    "brand_logo": "",             # header me logo image ka path
    "remember_window": True,      # window ka size/jagah yaad rakho
    "win_geometry": "",           # (auto) saved window geometry
    "auto_update_check": True,    # apne aap update jaanchna
    "dbl_action": "edit",         # thumbnail double-click: edit / preview
    "files_panel_side": "right",  # 'Meri Files' panel right / left
    "sound_on_done": False,       # scan poora hote hi 'ting'
    "confirm_exit": False,        # band karte samay pucho
    "footer_clock": False,        # status-bar: ghadi/date
    "footer_today": False,        # status-bar: aaj ke scan
    "footer_online": False,       # status-bar: online log
    "footer_msg": "",             # status-bar: apna sandesh
}


def load_config():
    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_config(cfg):
    try:
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(cfg, f)
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Image helpers
# ---------------------------------------------------------------------------

def is_blank_page(img, dark_fraction_threshold=0.0008, dark_level=75):
    # Count only VERY DARK ink (< dark_level, default 75). Real printed text / stamps
    # are near-black; fold-line shadows and faint bleed-through from the front are
    # medium/light grey (>75) and are therefore IGNORED. So a blank back side (only
    # creases + bleed-through) reads ~0% and gets dropped, while a page with real
    # (even sparse) dark text is kept. No erosion (that was thinning real text).
    try:
        g = img.convert("L")
        ink = g.point(lambda p: 255 if p < dark_level else 0)
        hist = ink.histogram()
        total = (g.width * g.height) or 1
        return (hist[-1] / total) < dark_fraction_threshold
    except Exception:
        return False


def whiten_dark_background(img, bright_thresh=160):
    """Scanner ki DARK backing (kala / navy / GRAY — sab) jo page ke chaaron
    taraf dikhti hai use WHITE karo — pixel-level par.

    Har row/column me KINARE se shuru karke pehle BRIGHT (paper) pixel tak ke
    pixels hi white hote hain. Isliye:
      - tedha page ho to bhi uska kona/kinara kabhi nahi 'katta' (purana code
        poori patti white pot deta tha jisme page ka kinara bhi chala jata
        tha — 'side se kata hua' wala bug),
      - gray backing bhi pakdi jaati hai (purana threshold sirf <90 tha,
        gray 90-160 wali backing reh jaati thi),
      - page ke ANDAR ka content (text, X-ray, stamp) kabhi nahi chhedha
        jaata — run pehle bright pixel par ruk jaata hai."""
    try:
        import numpy as _np
        g = _np.asarray(img.convert("L"))
        # fast early-out: kinare pehle se bright hain to backing hai hi nahi
        if (g[0].mean() > 150 and g[-1].mean() > 150
                and g[:, 0].mean() > 150 and g[:, -1].mean() > 150):
            return img
        bright = g >= bright_thresh
        H, W = g.shape
        arr = _np.asarray(img.convert("RGB")).copy()
        cols = _np.arange(W)[None, :]
        rows = _np.arange(H)[:, None]
        # Left/Right: har row me pehla/aakhri bright pixel — uske bahar sab white.
        any_row = bright.any(axis=1)
        first = _np.where(any_row, bright.argmax(axis=1), W).astype(_np.int64)[:, None]
        last = _np.where(any_row, W - 1 - bright[:, ::-1].argmax(axis=1), -1).astype(_np.int64)[:, None]
        mask = (cols < first) | (cols > last)
        # Poori-dark rows (koi bright pixel nahi) sirf tab white hon jab wo
        # upar/neeche ke KINARE se lagatar judi hon (backing band). Beech ki
        # dark rows (jaise X-ray film) kabhi nahi.
        no_bright_row = ~any_row
        top_run = _np.cumprod(no_bright_row).astype(bool)
        bot_run = _np.cumprod(no_bright_row[::-1]).astype(bool)[::-1]
        interior_dark_rows = no_bright_row & ~top_run & ~bot_run
        mask[interior_dark_rows, :] = False
        # Top/Bottom: har column me bhi wahi — kinare se pehle bright pixel tak.
        any_col = bright.any(axis=0)
        firstc = _np.where(any_col, bright.argmax(axis=0), H).astype(_np.int64)[None, :]
        lastc = _np.where(any_col, H - 1 - bright[::-1, :].argmax(axis=0), -1).astype(_np.int64)[None, :]
        cmask = (rows < firstc) | (rows > lastc)
        no_bright_col = ~any_col
        left_run = _np.cumprod(no_bright_col).astype(bool)
        right_run = _np.cumprod(no_bright_col[::-1]).astype(bool)[::-1]
        interior_dark_cols = no_bright_col & ~left_run & ~right_run
        cmask[:, interior_dark_cols] = False
        mask |= cmask
        arr[mask] = 255
        return Image.fromarray(arr)
    except Exception:
        return img


def trim_dark_borders(img, bright_thresh=120, min_paper_frac=0.04, pad=6):
    """Remove wide DARK margins (the scanner's backing — black OR dark-blue —
    showing around a narrow sheet like an ECG strip) so the page comes out to its
    real paper. A row/column is "paper" if enough of its pixels are BRIGHT
    (paper background), so dark backing is trimmed. Safe: only crops when a real
    dark border exists; leaves normal white pages and fully-dark pages untouched."""
    try:
        import numpy as _np
        g = img.convert("L")
        a = _np.asarray(g)
        bright = a > bright_thresh                       # paper background is bright
        row_has = bright.mean(axis=1) > min_paper_frac
        col_has = bright.mean(axis=0) > min_paper_frac
        rows = _np.where(row_has)[0]
        cols = _np.where(col_has)[0]
        if len(rows) == 0 or len(cols) == 0:
            return img                                   # no bright paper found
        t, b = int(rows[0]), int(rows[-1])
        l, r = int(cols[0]), int(cols[-1])
        t = max(0, t - pad); l = max(0, l - pad)
        b = min(a.shape[0] - 1, b + pad); r = min(a.shape[1] - 1, r + pad)
        # only crop if it removes a meaningful border (>1.5% on some side)
        if (r - l) < img.width * 0.985 or (b - t) < img.height * 0.985:
            return img.crop((l, t, r + 1, b + 1))
        return img
    except Exception:
        return img


def autocrop(img, border=20):
    try:
        rgb = img.convert("RGB")
        bg = Image.new("RGB", rgb.size, (255, 255, 255))
        bbox = ImageChops.difference(rgb, bg).getbbox()
        if not bbox:
            return img
        l, t, r, b = bbox
        l = max(0, l - border); t = max(0, t - border)
        r = min(img.width, r + border); b = min(img.height, b + border)
        return img.crop((l, t, r, b))
    except Exception:
        return img


def deskew(img):
    if not HAS_NUMPY:
        return img
    try:
        g = img.convert("L")
        small = g.resize((max(1, g.width // 4), max(1, g.height // 4)))
        best_angle, best_score = 0.0, -1.0
        for i in range(-10, 11):
            angle = i * 0.5
            rot = small.rotate(angle, resample=Image.BILINEAR, fillcolor=255)
            arr = np.asarray(rot, dtype=np.float32)
            score = float(np.var(np.diff(arr.sum(axis=1))))
            if score > best_score:
                best_score, best_angle = score, angle
        if abs(best_angle) < 0.25:
            return img
        return img.rotate(best_angle, resample=Image.BICUBIC, fillcolor=(255, 255, 255), expand=True)
    except Exception:
        return img


def auto_enhance(img):
    """Clean up faded / dull documents automatically."""
    try:
        rgb = img.convert("RGB")
        rgb = ImageOps.autocontrast(rgb, cutoff=1)
        rgb = ImageEnhance.Sharpness(rgb).enhance(1.4)
        rgb = ImageEnhance.Contrast(rgb).enhance(1.06)
        return rgb
    except Exception:
        return img


def clean_edges(img, margin_frac=0.018, dark_level=90):
    """Scan ke kinaron par aane wali KAALI border aur kinare ke chhed (punch-hole)
    ke nishan saaf karo. Sirf bahari patti (margin) dekhi jaati hai — beech ka
    text/stamp bilkul nahi chhua jaata. Margin ke andar ke gehre (dark) pixel
    safed kar diye jaate hain."""
    try:
        rgb = img.convert("RGB")
        w, h = rgb.size
        m = max(2, int(min(w, h) * margin_frac))
        px = rgb.load()
        white = (255, 255, 255)

        def _row(y):
            for x in range(w):
                r, g, b = px[x, y][:3]
                if (r + g + b) / 3 < dark_level:
                    px[x, y] = white

        def _col(x):
            for y in range(h):
                r, g, b = px[x, y][:3]
                if (r + g + b) / 3 < dark_level:
                    px[x, y] = white
        for y in list(range(0, m)) + list(range(h - m, h)):
            _row(y)
        for x in list(range(0, m)) + list(range(w - m, w)):
            _col(x)
        return rgb
    except Exception:
        return img


def split_two_pages(img, min_aspect=1.15):
    """Ek glass par do page rakhe ho (chaudi/landscape scan) to unhe do alag
    page me kaato. Beech me sabse safed (khaali) vertical patti dhoondh kar wahi
    se kaatte hain. Do image ki list lautata hai; kaatne layak na ho to [img]."""
    try:
        rgb = img.convert("RGB")
        w, h = rgb.size
        if w < h * min_aspect:      # chaudi nahi — split ki zaroorat nahi
            return [rgb]
        gray = rgb.convert("L")
        if HAS_NUMPY:
            import numpy as _np
            a = _np.asarray(gray, dtype=_np.float32)
            col_mean = a.mean(axis=0)                 # har column ki roshni
            lo, hi = int(w * 0.35), int(w * 0.65)     # beech ke 30% me hi gutter
            band = col_mean[lo:hi]
            cut = lo + int(band.argmax())             # sabse safed column
            if col_mean[cut] < 200:                   # itni safed patti nahi mili
                cut = w // 2
        else:
            cut = w // 2
        left = rgb.crop((0, 0, cut, h))
        right = rgb.crop((cut, 0, w, h))
        return [left, right]
    except Exception:
        return [img]


def flatten_photo_shadows(img):
    """Phone-photo ke shadow / roshni ke dabbe hatao: background ko blur karke
    usse divide karo — paper flat white ho jata hai, text/stamp waise hi rehte
    hain. numpy na ho to sirf autocontrast."""
    if not HAS_NUMPY:
        return ImageOps.autocontrast(img.convert("RGB"), cutoff=1)
    try:
        import numpy as _np
        rgb = img.convert("RGB")
        bg = rgb.convert("L").filter(ImageFilter.GaussianBlur(radius=40))
        b = _np.asarray(bg, dtype=_np.float32)
        b[b < 1] = 1
        a = _np.asarray(rgb, dtype=_np.float32)
        flat = _np.clip(a / b[..., None] * 230.0, 0, 255).astype("uint8")
        out = Image.fromarray(flat, "RGB")
        return ImageOps.autocontrast(out, cutoff=1)
    except Exception:
        return img


def clean_photo(img):
    """Phone se kheenchi photo ko scan-jaisa banao: EXIF ke hisaab se seedha,
    bahut badi ho to chhota, shadow hatao."""
    try:
        img = ImageOps.exif_transpose(img)
    except Exception:
        pass
    img = img.convert("RGB")
    m = max(img.width, img.height)
    if m > 2600:
        s = 2600.0 / m
        img = img.resize((max(1, int(img.width * s)), max(1, int(img.height * s))),
                         Image.LANCZOS)
    return flatten_photo_shadows(img)


def _largest_gap(flags, min_gap):
    """Bool array me sabse badi False-run (start,end) jo kinaron ko na chhue."""
    best = None
    run = 0
    for i, v in enumerate(flags):
        if not v:
            run += 1
        else:
            if run >= min_gap and i - run > 0:
                if best is None or run > (best[1] - best[0]):
                    best = (i - run, i)
            run = 0
    return best


def detect_content_boxes(img, bg_thresh=235):
    """White background par rakhe alag-alag cards/documents ke boxes dhoondo
    (2-3 ID ek saath scan karne par). Bade khaali gap par recursively todta hai."""
    if not HAS_NUMPY:
        return [(0, 0, img.width, img.height)]
    import numpy as _np
    g = _np.asarray(img.convert("L"))
    mask = g < bg_thresh
    boxes = []

    def _split(t, b, l, r, depth):
        sub = mask[t:b, l:r]
        if sub.size == 0 or not sub.any():
            return
        rows = sub.any(axis=1)
        cols = sub.any(axis=0)
        t2 = t + int(_np.argmax(rows))
        b2 = b - int(_np.argmax(rows[::-1]))
        l2 = l + int(_np.argmax(cols))
        r2 = r - int(_np.argmax(cols[::-1]))
        if depth < 6:
            rows2 = mask[t2:b2, l2:r2].any(axis=1)
            gap = _largest_gap(rows2, max(14, (b2 - t2) // 20))
            if gap:
                _split(t2, t2 + gap[0], l2, r2, depth + 1)
                _split(t2 + gap[1], b2, l2, r2, depth + 1)
                return
            cols2 = mask[t2:b2, l2:r2].any(axis=0)
            gap = _largest_gap(cols2, max(14, (r2 - l2) // 20))
            if gap:
                _split(t2, b2, l2, l2 + gap[0], depth + 1)
                _split(t2, b2, l2 + gap[1], r2, depth + 1)
                return
        boxes.append((l2, t2, r2, b2))

    _split(0, 0 + mask.shape[0], 0, mask.shape[1], 0)
    return boxes


def auto_orient(img):
    """Ulta (90/180/270 ghuma hua) page OCR (Tesseract OSD) se pehchan kar
    seedha karo. Tesseract na ho ya samajh na aaye to page waise hi rehta hai."""
    try:
        osd = pytesseract.image_to_osd(img.convert("RGB"))
        m = re.search(r"Rotate:\s*(\d+)", osd or "")
        rot = int(m.group(1)) if m else 0
        if rot in (90, 180, 270):
            return img.rotate(-rot, expand=True)
    except Exception:
        pass
    return img


def colorfulness(img):
    """Page kitna rangeen hai (0 = pura B&W jaisa). Auto colour-detect ke liye."""
    try:
        import numpy as _np
        small = img.convert("RGB").resize(
            (max(1, img.width // 8), max(1, img.height // 8)))
        a = _np.asarray(small, dtype=_np.int16)
        rg = _np.abs(a[..., 0] - a[..., 1])
        yb = _np.abs((a[..., 0] + a[..., 1]) // 2 - a[..., 2])
        return float(rg.mean() + yb.mean())
    except Exception:
        return 999.0


def restore_photo(img):
    """Purani dhundhli/feeki photo ko sudharo: contrast, rang, sharpness."""
    try:
        rgb = img.convert("RGB")
        rgb = rgb.filter(ImageFilter.MedianFilter(3))          # halka noise saaf
        rgb = ImageOps.autocontrast(rgb, cutoff=2)
        rgb = ImageEnhance.Color(rgb).enhance(1.25)
        rgb = ImageEnhance.Sharpness(rgb).enhance(1.3)
        return rgb
    except Exception:
        return img


def save_image_keep_ext(img, path, quality=92):
    """Image ko uske extension ke hisaab se sahi format me save karo."""
    if path.lower().endswith((".jpg", ".jpeg")):
        img.convert("RGB").save(path, "JPEG", quality=quality)
    else:
        img.save(path, "PNG")


def apply_watermark(img, text):
    try:
        base = img.convert("RGBA")
        overlay = Image.new("RGBA", base.size, (255, 255, 255, 0))
        draw = ImageDraw.Draw(overlay)
        size = max(14, base.width // 45)
        try:
            font = ImageFont.truetype("arial.ttf", size)
        except Exception:
            font = ImageFont.load_default()
        stamp = "%s  |  %s" % (text, datetime.datetime.now().strftime("%d-%m-%Y"))
        try:
            bbox = draw.textbbox((0, 0), stamp, font=font)
            tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
        except Exception:
            tw, th = draw.textsize(stamp, font=font)
        x = (base.width - tw) // 2
        y = base.height - th - max(10, base.height // 60)
        draw.rectangle([x - 8, y - 6, x + tw + 8, y + th + 6], fill=(255, 255, 255, 160))
        draw.text((x, y), stamp, fill=(120, 120, 120, 200), font=font)
        return Image.alpha_composite(base, overlay).convert("RGB")
    except Exception:
        return img


def enhance_image(path, brightness=1.0, contrast=1.0):
    try:
        with Image.open(path) as im:
            im = im.convert("RGB")
            if brightness != 1.0:
                im = ImageEnhance.Brightness(im).enhance(brightness)
            if contrast != 1.0:
                im = ImageEnhance.Contrast(im).enhance(contrast)
            im.save(path, "PNG")
        return True
    except Exception:
        return False


def read_barcode(path):
    """Return first barcode/QR text found in the image, else None."""
    if not HAS_ZBAR:
        return None
    try:
        with Image.open(path) as im:
            results = _zbar_decode(im.convert("RGB"))
        for r in results:
            try:
                val = r.data.decode("utf-8").strip()
            except Exception:
                val = str(r.data).strip()
            if val:
                return val
        return None
    except Exception:
        return None


def sanitize(text, fallback="scan"):
    safe = "".join(c for c in (text or "") if c.isalnum() or c in "-_")
    return safe or fallback


# ---------------------------------------------------------------------------
# Network + TWAIN backend
# ---------------------------------------------------------------------------

def tcp_reachable(ip, ports=SCANNER_PORTS, timeout=1.2):
    for p in ports:
        try:
            with socket.create_connection((ip, p), timeout=timeout):
                return True
        except Exception:
            continue
    return False


class ConnectionChecker(QtCore.QThread):
    result = QtCore.pyqtSignal(bool, str)

    def __init__(self, ip):
        super().__init__()
        self.ip = (ip or "").strip()

    def run(self):
        if not self.ip:
            self.result.emit(False, 'Scanner IP address is not set')
            return
        if tcp_reachable(self.ip):
            self.result.emit(True, 'Connected — scanner found on the network (%s)' % self.ip)
        else:
            self.result.emit(False, 'Not connected — %s is unreachable. Is the scanner ON and on the same WiFi/LAN?' % self.ip)


class ScannerStateChecker(QtCore.QThread):
    """Ask an eSCL scanner whether it's Idle (free) or Processing (busy)."""
    state = QtCore.pyqtSignal(str)   # "free" / "busy" / "unknown"

    def __init__(self, ip):
        super().__init__()
        self.ip = (ip or "").strip()

    def run(self):
        if not self.ip:
            self.state.emit("unknown"); return
        try:
            import urllib.request as _u
            r = _u.urlopen("http://%s/eSCL/ScannerStatus" % self.ip, timeout=5)
            xml = r.read().decode("utf-8", "ignore")
            m = re.search(r"State>\s*([A-Za-z]+)", xml)
            st = (m.group(1).lower() if m else "")
            if st == "idle":
                self.state.emit("free")
            elif st:
                self.state.emit("busy")
            else:
                self.state.emit("unknown")
        except Exception:
            self.state.emit("unknown")


class EsclTestWorker(QtCore.QThread):
    """Step-by-step eSCL probe that reports the ACTUAL response at each stage,
    so the real cause of a scan failure (busy job / bad settings / offline) is
    visible instead of just a generic error."""
    done = QtCore.pyqtSignal(str)

    def __init__(self, ip):
        super().__init__()
        self.ip = (ip or "").strip()

    def run(self):
        import urllib.request as U
        import urllib.error as UE
        lines = []
        def add(t=""):
            lines.append(t)
        add("===== ApneScan eSCL Test =====")
        add("Time: %s" % datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        add("Scanner IP: %s" % (self.ip or "(not set)"))
        base = "http://%s/eSCL" % self.ip

        add("")
        add("[1] TCP connect (port 80):")
        try:
            add("    %s" % ("OK - reachable" if tcp_reachable(self.ip) else "FAIL - not reachable (scanner ON? same WiFi/LAN?)"))
        except Exception as e:
            add("    ERROR: %s" % e)

        add("")
        add("[2] GET /eSCL/ScannerCapabilities:")
        try:
            r = U.urlopen(base + "/ScannerCapabilities", timeout=8)
            body = r.read().decode("utf-8", "ignore")
            add("    HTTP %s, %d bytes" % (getattr(r, "status", 200), len(body)))
            mw = re.findall(r"MaxWidth>\s*(\d+)", body)
            mh = re.findall(r"MaxHeight>\s*(\d+)", body)
            if mw and mh:
                add("    Max scan area: W=%s H=%s (1/300 inch)" % (max(mw), max(mh)))
            if "Duplex" in body:
                add("    Duplex: supported (eSCL)")
        except UE.HTTPError as e:
            add('    HTTP ERROR %s (eSCL may not be supported)' % e.code)
        except Exception as e:
            add("    ERROR: %s" % e)

        add("")
        add("[3] GET /eSCL/ScannerStatus (free/busy + open jobs):")
        try:
            r = U.urlopen(base + "/ScannerStatus", timeout=8)
            body = r.read().decode("utf-8", "ignore")
            st = re.search(r"State>\s*([A-Za-z]+)", body)
            add("    HTTP %s, State=%s" % (getattr(r, "status", 200), st.group(1) if st else "?"))
            jobs = re.findall(r"JobState>\s*([A-Za-z]+)", body)
            if jobs:
                add("    Open/last jobs: %s" % ", ".join(jobs))
                add("    (Agar koi 'Processing'/'Pending' job hai to scanner busy rahega.)")
            else:
                add("    No open jobs listed.")
        except UE.HTTPError as e:
            add("    HTTP ERROR %s" % e.code)
        except Exception as e:
            add("    ERROR: %s" % e)

        add("")
        add("[4] POST /eSCL/ScanJobs  (THE REAL TEST - minimal A4 simplex gray):")
        settings = (
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            '<scan:ScanSettings xmlns:scan="http://schemas.hp.com/imaging/escl/2011/05/03" '
            'xmlns:pwg="http://www.pwg.org/schemas/2010/12/sm">\n'
            '<pwg:Version>2.6</pwg:Version>\n'
            '<pwg:ScanRegions pwg:MustHonor="false"><pwg:ScanRegion>\n'
            '<pwg:Height>3550</pwg:Height><pwg:Width>2480</pwg:Width>\n'
            '<pwg:XOffset>0</pwg:XOffset><pwg:YOffset>0</pwg:YOffset>\n'
            '</pwg:ScanRegion></pwg:ScanRegions>\n'
            '<pwg:InputSource>Feeder</pwg:InputSource>\n'
            '<scan:ColorMode>Grayscale8</scan:ColorMode>\n'
            '<scan:XResolution>200</scan:XResolution>\n'
            '<scan:YResolution>200</scan:YResolution>\n'
            '<pwg:DocumentFormat>image/jpeg</pwg:DocumentFormat>\n'
            '</scan:ScanSettings>'
        )
        job = None
        try:
            req = U.Request(base + "/ScanJobs", data=settings.encode("utf-8"),
                            headers={"Content-Type": "text/xml"}, method="POST")
            r = U.urlopen(req, timeout=15)
            job = r.headers.get("Location")
            add("    HTTP %s  ->  JOB CREATED" % getattr(r, "status", 201))
            add("    Location: %s" % job)
        except UE.HTTPError as e:
            add("    HTTP ERROR %s" % e.code)
            try:
                body = e.read().decode("utf-8", "ignore")
                if body.strip():
                    add("    Response body: %s" % body.replace("\n", " ")[:300])
            except Exception:
                pass
            if e.code == 503:
                add('    => 503 = scanner BUSY / an old job is still open.')
                add('       Fix: turn the scanner off/on; close any other scan app.')
            elif e.code in (400, 415, 409):
                add('    => %s = scan settings/format problem (needs a tweak for this scanner).' % e.code)
        except Exception as e:
            add("    ERROR: %s" % e)

        if job:
            if job.startswith("/"):
                job = "http://%s%s" % (self.ip, job)
            add("")
            add("[5] GET job/NextDocument (1 page fetch):")
            try:
                r = U.urlopen(job + "/NextDocument", timeout=30)
                d = r.read()
                add("    HTTP %s, %d bytes  ->  PAGE RECEIVED (scan works!)" % (getattr(r, "status", 200), len(d)))
            except UE.HTTPError as e:
                add('    HTTP %s (no page / done)' % e.code)
            except Exception as e:
                add("    ERROR: %s" % e)
            add("")
            add("[6] DELETE job (release):")
            try:
                U.urlopen(U.Request(job, method="DELETE"), timeout=8)
                add("    OK - job released (scanner free).")
            except Exception as e:
                add("    ERROR: %s" % e)

        add("")
        add("===== RESULT =====")
        add("If [4] shows 'JOB CREATED' -> scanning should work.")
        add('If [4] shows 503 -> scanner busy: turn it off/on and close other scan apps.')
        add('If [1] FAILs -> scanner not found on the network (check IP/WiFi).')
        add("==================")
        self.done.emit("\n".join(lines))


def _urlopen_safe(req, timeout=20):
    """HTTPS open jo Windows/PyInstaller build me bhi chale. Kai baar frozen
    .exe ko CA-certificate store nahi milta -> SSL verify fail hota hai aur
    HTTPS request chup-chaap error de deti hai (isi wajah se sidebar ke
    worldwide stats '...' par atak jaate the). Aise me — sirf public ginti/
    version ke liye — bina-verify dobara try karte hain, taaki numbers aur
    update-check hamesha chalein."""
    import urllib.request as U
    import urllib.error as E
    import ssl
    try:
        return U.urlopen(req, timeout=timeout)
    except E.URLError as ex:
        if isinstance(getattr(ex, "reason", None), ssl.SSLError):
            return U.urlopen(req, timeout=timeout,
                             context=ssl._create_unverified_context())
        raise


class StatsWorker(QtCore.QThread):
    """Talk to the ApneScan stats server (Google Apps Script). action:
    'ping' (mark online + fetch), 'scan' (add scans + fetch), 'stats' (fetch).
    App version + country (sirf 2-akshar ka code) bhi jaata hai — kabhi koi
    document/naam nahi."""
    got = QtCore.pyqtSignal(int, int, int)   # total, today, online
    got_full = QtCore.pyqtSignal(dict)       # poora server payload
    failed = QtCore.pyqtSignal()

    def __init__(self, url, client, action="ping", n=0, imp=0, prt=0):
        super().__init__()
        self.url = url; self.client = client; self.action = action; self.n = n
        self.imp = imp; self.prt = prt          # import / print ginti
        try:
            self.country = (QtCore.QLocale().name().split("_") + [""])[1][:2]
        except Exception:
            self.country = ""

    def run(self):
        try:
            import urllib.request as U
            import urllib.parse as P
            import json as J
            import ssl
            q = {"action": self.action, "client": self.client,
                 "v": VERSION, "c": self.country}
            if self.action == "scan":
                q["n"] = str(self.n)
            if self.imp:
                q["imp"] = str(self.imp)
            if self.prt:
                q["prt"] = str(self.prt)
            full = self.url + ("&" if "?" in self.url else "?") + P.urlencode(q)
            req = U.Request(full, headers={"User-Agent": "ApneScan/%s" % VERSION})
            data = None
            try:
                r = _urlopen_safe(req, timeout=20)
                data = J.loads(r.read().decode("utf-8", "ignore"))
            except Exception:
                # ping/stats sirf GINTI padhte hain (idempotent) — isliye kisi
                # bhi dikkat (SSL/proxy/redirect) par bina-verify dobara try
                # karo, taaki worldwide stats hamesha aayein. 'scan' ko dobara
                # nahi bhejte (double-count na ho).
                if self.action != "scan":
                    try:
                        r = U.urlopen(req, timeout=25,
                                      context=ssl._create_unverified_context())
                        data = J.loads(r.read().decode("utf-8", "ignore"))
                    except Exception:
                        # aakhri koshish: system-proxy ko BYPASS karke (kabhi
                        # office/hospital ka galat proxy app ko rok deta hai,
                        # jabki browser chal jata hai)
                        try:
                            opener = U.build_opener(
                                U.ProxyHandler({}),
                                U.HTTPSHandler(context=ssl._create_unverified_context()))
                            r = opener.open(req, timeout=25)
                            data = J.loads(r.read().decode("utf-8", "ignore"))
                        except Exception:
                            data = None
            if data and data.get("ok"):
                self.got.emit(int(data.get("total", 0)),
                              int(data.get("today", 0)),
                              int(data.get("online", 0)))
                try:
                    self.got_full.emit(dict(data))
                except Exception:
                    pass
            else:
                self.failed.emit()
        except Exception:
            self.failed.emit()


class UpdateChecker(QtCore.QThread):
    """Latest version ka pata karo — pehle WEBSITE se (version.txt, jo har
    deploy par apne aap banta hai), na mile to GitHub Releases se."""
    result = QtCore.pyqtSignal(str, str)   # latest_tag ("v19"), download_url

    def run(self):
        import urllib.request as U
        try:
            req = U.Request("https://apnescan.apnesoft.com/version.txt",
                            headers={"User-Agent": "ApneScan"})
            tag = _urlopen_safe(req, timeout=10).read().decode("utf-8", "ignore").strip()[:20]
            if re.search(r"\d", tag or ""):
                self.result.emit(tag, DOWNLOAD_PAGE)
                return
        except Exception:
            pass
        try:
            import json as J
            req = U.Request(UPDATE_API, headers={"User-Agent": "ApneScan"})
            data = J.loads(_urlopen_safe(req, timeout=10).read().decode("utf-8", "ignore"))
            self.result.emit(str(data.get("tag_name") or ""),
                             str(data.get("html_url") or DOWNLOAD_PAGE))
        except Exception:
            self.result.emit("", "")


class ScannerFinder(QtCore.QThread):
    """Network par eSCL scanner khud dhoondo — poore /24 subnet ko jaanchta hai
    (IP hath se daalne ki zaroorat nahi)."""
    found = QtCore.pyqtSignal(list)   # [(ip, model), ...]

    def run(self):
        import socket as _s
        import concurrent.futures as _cf
        import urllib.request as _u
        results = []
        try:
            s = _s.socket(_s.AF_INET, _s.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            my_ip = s.getsockname()[0]
            s.close()
        except Exception:
            self.found.emit([])
            return
        base = my_ip.rsplit(".", 1)[0]

        def probe(i):
            ip = "%s.%d" % (base, i)
            for port in (80, 8080):
                try:
                    c = _s.create_connection((ip, port), timeout=0.25)
                    c.close()
                except Exception:
                    continue
                try:
                    host = ip if port == 80 else "%s:%d" % (ip, port)
                    r = _u.urlopen("http://%s/eSCL/ScannerCapabilities" % host, timeout=1.5)
                    xml = r.read(4000).decode("utf-8", "ignore")
                    m = re.search(r"MakeAndModel>\s*([^<]+)", xml)
                    return (ip, (m.group(1).strip() if m else "eSCL scanner"))
                except Exception:
                    pass
            return None

        try:
            with _cf.ThreadPoolExecutor(max_workers=64) as ex:
                for res in ex.map(probe, range(1, 255)):
                    if res:
                        results.append(res)
        except Exception:
            pass
        self.found.emit(results)


class ScannerError(Exception):
    pass


def list_sources(hwnd):
    if not HAS_TWAIN:
        raise ScannerError('TWAIN (pytwain) is not installed.')
    sm = twain.SourceManager(hwnd)
    try:
        try:
            return list(sm.source_list)
        except Exception:
            return list(sm.GetSourceList())
    finally:
        try:
            sm.close()
        except Exception:
            try:
                sm.destroy()
            except Exception:
                pass


class _FileXferUnsupported(Exception):
    pass


def _open_twain_source(sm, source_name):
    try:
        src = sm.open_source(source_name) if source_name else sm.open_source()
    except Exception:
        try:
            src = sm.OpenSource(source_name) if source_name else sm.OpenSource()
        except Exception:
            src = None
    if src is None:
        try:
            src = sm.open_source()
        except Exception:
            try:
                src = sm.OpenSource()
            except Exception:
                src = None
    return src


def _common_caps(src, dpi, pixel_type, duplex):
    def _try(fn):
        try:
            fn()
        except Exception:
            pass
    _try(lambda: src.set_capability(twain.ICAP_XRESOLUTION, twain.TWTY_FIX32, float(dpi)))
    _try(lambda: src.set_capability(twain.ICAP_YRESOLUTION, twain.TWTY_FIX32, float(dpi)))
    pt = {"color": twain.TWPT_RGB, "gray": twain.TWPT_GRAY, "bw": twain.TWPT_BW}[pixel_type]
    _try(lambda: src.set_capability(twain.ICAP_PIXELTYPE, twain.TWTY_UINT16, pt))
    _try(lambda: src.set_capability(twain.CAP_FEEDERENABLED, twain.TWTY_BOOL, True))
    _try(lambda: src.set_capability(twain.CAP_AUTOFEED, twain.TWTY_BOOL, True))
    _try(lambda: src.set_capability(twain.CAP_AUTOSCAN, twain.TWTY_BOOL, True))
    _try(lambda: src.set_capability(twain.CAP_XFERCOUNT, twain.TWTY_INT16, -1))
    _try(lambda: src.set_capability(twain.CAP_DUPLEXENABLED, twain.TWTY_BOOL, bool(duplex)))


def _scan_file(hwnd, source_name, dpi, pixel_type, duplex, on_page=None, should_stop=None):
    """File-transfer acquire: the driver writes each page to a file and (on most
    ADF scanners) keeps the feeder moving continuously. If the source doesn't
    support it, raises _FileXferUnsupported so the caller falls back to native.
    Only raises _FileXferUnsupported when ZERO pages were produced (so we never
    double-feed paper)."""
    count = 0
    MAX_PAGES = 1000   # hard guard: never loop forever
    sm = twain.SourceManager(hwnd)
    src = None
    try:
        src = _open_twain_source(sm, source_name)
        if src is None:
            raise _FileXferUnsupported()
        _common_caps(src, dpi, pixel_type, duplex)
        # Switch to FILE transfer mechanism. If unsupported -> fall back.
        try:
            src.set_capability(twain.ICAP_XFERMECH, twain.TWTY_UINT16, twain.TWSX_FILE)
        except Exception:
            raise _FileXferUnsupported()
        try:
            src.request_acquire(show_ui=False, modal_ui=False)
        except Exception:
            raise _FileXferUnsupported()

        while count < MAX_PAGES:
            if should_stop and should_stop():
                break
            fd, bmp = tempfile.mkstemp(suffix=".bmp")
            os.close(fd)
            try:
                os.remove(bmp)
            except Exception:
                pass
            # point the driver at our file
            try:
                src.file_xfer_params(bmp, twain.TWFF_BMP)
            except Exception:
                if count == 0:
                    raise _FileXferUnsupported()
                break
            # transfer one page to the file
            try:
                more = src.xfer_image_by_file()
            except Exception:
                # feeder empty / done (or unsupported on first try)
                if count == 0:
                    raise _FileXferUnsupported()
                break
            if not os.path.exists(bmp):
                if count == 0:
                    raise _FileXferUnsupported()
                break
            try:
                with Image.open(bmp) as im:
                    img = im.convert("RGB").copy()
            finally:
                try:
                    os.remove(bmp)
                except Exception:
                    pass
            count += 1
            if on_page:
                on_page(img)
            # 'more' may be a bool/int/pending-count depending on pytwain; stop when falsy
            if more in (None, 0, False):
                break
        return count
    finally:
        try:
            if src is not None:
                src.close()
        except Exception:
            pass
        try:
            sm.close()
        except Exception:
            try:
                sm.destroy()
            except Exception:
                pass


def scan_pages(hwnd, source_name, dpi, pixel_type, duplex, on_page=None, should_stop=None,
               file_xfer=False):
    """TWAIN scan. When file_xfer=True (Settings -> experimental continuous ADF
    feed) try the file-transfer path first and fall back to the proven native
    path only when zero pages were produced."""
    if file_xfer:
        try:
            n = _scan_file(hwnd, source_name, dpi, pixel_type, duplex, on_page, should_stop)
            if n > 0:
                return n
            raise _FileXferUnsupported()
        except _FileXferUnsupported:
            pass
        except ScannerError:
            raise
        except Exception:
            pass
    return _scan_native(hwnd, source_name, dpi, pixel_type, duplex, on_page, should_stop)


def _scan_native(hwnd, source_name, dpi, pixel_type, duplex, on_page=None, should_stop=None):
    if not HAS_TWAIN:
        raise ScannerError('TWAIN (pytwain) is not installed.')
    count = 0
    sm = twain.SourceManager(hwnd)
    src = None
    try:
        try:
            src = sm.open_source(source_name) if source_name else sm.open_source()
        except Exception:
            try:
                src = sm.OpenSource(source_name) if source_name else sm.OpenSource()
            except Exception:
                src = None
        if src is None:
            # Stored name might be a WIA name (when switching method) -> open default TWAIN source.
            try:
                src = sm.open_source()
            except Exception:
                try:
                    src = sm.OpenSource()
                except Exception:
                    src = None
        if src is None:
            raise ScannerError('Scanner did not open. Choose a device in the profile.')

        def _try(fn):
            try:
                fn()
            except Exception:
                pass

        _try(lambda: src.set_capability(twain.ICAP_XRESOLUTION, twain.TWTY_FIX32, float(dpi)))
        _try(lambda: src.set_capability(twain.ICAP_YRESOLUTION, twain.TWTY_FIX32, float(dpi)))
        pt = {"color": twain.TWPT_RGB, "gray": twain.TWPT_GRAY, "bw": twain.TWPT_BW}[pixel_type]
        _try(lambda: src.set_capability(twain.ICAP_PIXELTYPE, twain.TWTY_UINT16, pt))
        # Continuous ADF feed: pull EVERY page in a single acquire (no per-page gap).
        _try(lambda: src.set_capability(twain.CAP_FEEDERENABLED, twain.TWTY_BOOL, True))
        _try(lambda: src.set_capability(twain.CAP_AUTOFEED, twain.TWTY_BOOL, True))
        _try(lambda: src.set_capability(twain.CAP_AUTOSCAN, twain.TWTY_BOOL, True))
        _try(lambda: src.set_capability(twain.CAP_XFERCOUNT, twain.TWTY_INT16, -1))
        _try(lambda: src.set_capability(twain.CAP_DUPLEXENABLED, twain.TWTY_BOOL, bool(duplex)))
        _try(lambda: src.request_acquire(show_ui=False, modal_ui=False))

        while True:
            if should_stop and should_stop():
                break
            try:
                rv = src.xfer_image_natively()
            except Exception as exc:
                if count:
                    break
                raise ScannerError("Scan transfer fail: %s" % exc)
            if rv is None:
                break
            handle, more = rv
            fd, bmp = tempfile.mkstemp(suffix=".bmp")
            os.close(fd)
            try:
                twain.dib_to_bm_file(handle, bmp)
            finally:
                try:
                    twain.global_handle_free(handle)
                except Exception:
                    pass
            with Image.open(bmp) as im:
                img = im.convert("RGB").copy()
            try:
                os.remove(bmp)
            except Exception:
                pass
            count += 1
            if on_page:
                on_page(img)
            if not more:
                break
    finally:
        try:
            if src is not None:
                src.close()
        except Exception:
            pass
        try:
            sm.close()
        except Exception:
            try:
                sm.destroy()
            except Exception:
                pass

    if count == 0:
        raise ScannerError('No page was scanned. Is a document in the feeder? Is the scanner ON?')
    return count


# ---------------------------------------------------------------------------
# WIA backend (alternative to TWAIN, for other users' scanners)
# ---------------------------------------------------------------------------

WIA_ERROR_PAPER_EMPTY = -2145320957   # 0x80210003


def list_wia_sources():
    if not HAS_W32:
        raise ScannerError('WIA (pywin32) is not installed.')
    dm = _w32.Dispatch("WIA.DeviceManager")
    out = []
    for i in range(1, dm.DeviceInfos.Count + 1):
        info = dm.DeviceInfos.Item(i)
        name = None
        try:
            for p in info.Properties:
                if p.Name == "Name":
                    name = p.Value
                    break
        except Exception:
            pass
        out.append((info.DeviceID, name or ("Scanner %d" % i)))
    return out


def wia_scan_pages(device_id, dpi, pixel_type, duplex, on_page=None, should_stop=None,
                   source="auto"):
    wia_scan_pages.last_duplex_value = getattr(wia_scan_pages, "last_duplex_value", None)
    """Scan via Windows WIA. Kept simple: no forced properties (some drivers
    throw otherwise). Runs inside a thread that has called CoInitialize."""
    if not HAS_W32:
        raise ScannerError('WIA (pywin32) is not installed.')
    dm = _w32.Dispatch("WIA.DeviceManager")
    device = None
    for i in range(1, dm.DeviceInfos.Count + 1):
        info = dm.DeviceInfos.Item(i)
        if device_id is None or info.DeviceID == device_id:
            device = info.Connect()
            break
    if device is None:
        raise ScannerError('No WIA scanner found. Choose a scanner in Settings.')

    def _wia_set(props, pid, value):
        # Set one WIA property; True if it took, False if the driver rejected it.
        try:
            for p in props:
                if int(p.PropertyID) == pid:
                    p.Value = value
                    return True
        except Exception:
            return False
        return False

    def _wia_valid_dpi(props, want):
        # Pick a resolution the device actually supports (avoids "parameter incorrect").
        try:
            for p in props:
                if int(p.PropertyID) == 6147:
                    vals = None
                    try:
                        vals = [int(x) for x in p.SubType.Values]   # WIA_PROP_LIST
                    except Exception:
                        vals = None
                    if vals:
                        return min(vals, key=lambda v: abs(v - want))
        except Exception:
            pass
        return want

    def _has_feeder(dev):
        # WIA_DPS_DOCUMENT_HANDLING_CAPABILITIES (3086): bit 0x01 = FEEDER.
        # M1136 jaise flatbed-only MFP me ye bit nahi hota.
        try:
            for p in dev.Properties:
                if int(p.PropertyID) == 3086:
                    return bool(int(p.Value) & 0x01)
        except Exception:
            pass
        return None   # pata nahi

    def _apply_flatbed(dev):
        # FLATBED (glass) — ek hi page scan hota hai.
        try:
            _wia_set(dev.Properties, 3088, 2)     # DOCUMENT_HANDLING_SELECT = FLATBED
        except Exception:
            pass

    def _apply_feeder(dev, want_duplex=None):
        # Feeder + (optional) DUPLEX + all pages. HP network drivers use different
        # duplex "select" values, so try a few and keep the first the driver accepts.
        wd = duplex if want_duplex is None else want_duplex
        try:
            dprops = dev.Properties
            if wd:
                for val in (5, 7, 0x8005, 4, 3):   # FEEDER|DUPLEX variants (0x8005=32773 HP)
                    if _wia_set(dprops, 3088, val):
                        wia_scan_pages.last_duplex_value = val
                        break
                else:
                    wia_scan_pages.last_duplex_value = None
            else:
                _wia_set(dprops, 3088, 1)          # feeder only (single side)
            _wia_set(dprops, 3096, 0)              # all pages
        except Exception:
            pass

    # FLATBED ya FEEDER? User ne "Glass" chuna, ya device me feeder hai hi nahi
    # (jaise HP LaserJet M1136 MFP) -> flatbed: SIRF 1 page scan karke ruko.
    _feeder_cap = _has_feeder(device)
    flatbed_only = (str(source).lower().startswith("glass")
                    or _feeder_cap is False)

    def _apply_quality(it):
        try:
            props = it.Properties
            _wia_set(props, 4104, {"bw": 0, "gray": 2, "color": 3}.get(pixel_type, 3))
            dv = _wia_valid_dpi(props, int(dpi))
            _wia_set(props, 6147, dv)
            _wia_set(props, 6148, dv)
        except Exception:
            pass

    if flatbed_only:
        _apply_flatbed(device)     # glass: 1 page only
    else:
        _apply_feeder(device)      # set FEEDER|DUPLEX on the DEVICE first
    item = device.Items[1]         # enumerate the page item AFTER duplex is active
    _apply_quality(item)
    _props_ok = True
    # BMP format id (widely accepted by HP WIA drivers).
    WIA_FMT_BMP = "{B96B3CAB-0728-11D3-9D7B-0000F81EF32E}"

    def _transfer_one(it):
        # Try Transfer(format) first (robust on HP), then plain Transfer().
        try:
            return it.Transfer(WIA_FMT_BMP)
        except Exception:
            return it.Transfer()

    count = 0
    retry_stage = 0
    while True:
        if should_stop and should_stop():
            break
        try:
            image = _transfer_one(item)
        except Exception as exc:
            hres = exc.args[0] if getattr(exc, "args", None) else None
            # End-of-feeder conditions -> stop cleanly.
            if hres in (WIA_ERROR_PAPER_EMPTY, -2145320957) or count > 0:
                break
            # First page failed: retry once with a fresh item and NO custom props.
            if retry_stage < 2:
                stage = retry_stage
                retry_stage += 1
                try:
                    device2 = None
                    dm2 = _w32.Dispatch("WIA.DeviceManager")
                    for i in range(1, dm2.DeviceInfos.Count + 1):
                        info = dm2.DeviceInfos.Item(i)
                        if device_id is None or info.DeviceID == device_id:
                            device2 = info.Connect(); break
                    if device2 is not None:
                        if stage == 0:
                            if flatbed_only:
                                _apply_flatbed(device2)
                            else:
                                _apply_feeder(device2)      # retry: duplex, no quality props
                        # stage 1: apply nothing -> guaranteed to scan (single side)
                        item = device2.Items[1]
                except Exception:
                    pass
                continue
            raise ScannerError("WIA scan fail: %s" % exc)
        fd, raw = tempfile.mkstemp(suffix=".bmp")
        os.close(fd)
        try:
            os.remove(raw)
        except Exception:
            pass
        try:
            image.SaveFile(raw)
        except Exception as exc:
            raise ScannerError("WIA save fail: %s" % exc)
        with Image.open(raw) as im:
            img = im.convert("RGB").copy()
        try:
            os.remove(raw)
        except Exception:
            pass
        count += 1
        if on_page:
            on_page(img)
        # FLATBED (glass) me feeder khaali hone ka signal nahi aata — isliye
        # 1 page ke baad KHUD ruk jao, warna ye baar-baar scan karta rahega.
        if flatbed_only:
            break
    if count == 0:
        raise ScannerError('No page was scanned (WIA).')
    return count


# ---------------------------------------------------------------------------
# Friendly error messages (technical -> simple)
# ---------------------------------------------------------------------------

def friendly_error(text, lang="hi"):
    t = str(text)
    low = t.lower()
    hi = {
        "not_found": "Scanner nahi mila. Scanner ON karein aur PC se juda hai ya nahi check karein.",
        "busy": "Scanner abhi busy hai. Thodi der baad dobara try karein.",
        "empty": "Feeder me koi document nahi hai. Page rakh kar dobara scan karein.",
        "jam": "Paper jam hua lagta hai. Feeder check karein.",
        "driver": "Scanner driver ki dikkat. Doosri scan-method (Settings me TWAIN/WIA) try karein.",
        "twain": "TWAIN scanner nahi khula. Settings me WIA method try karein.",
        "generic": "Scan nahi ho paya. Scanner ON hai aur juda hai? Dobara try karein.",
    }
    en = {
        "not_found": "Scanner not found. Turn it on and check it's connected.",
        "busy": "Scanner is busy. Please try again in a moment.",
        "empty": "No document in the feeder. Add a page and scan again.",
        "jam": "Looks like a paper jam. Check the feeder.",
        "driver": "Scanner driver issue. Try the other method (TWAIN/WIA) in Settings.",
        "twain": "TWAIN scanner didn't open. Try WIA method in Settings.",
        "generic": "Scanning failed. Is the scanner on and connected? Try again.",
    }
    m = hi if lang == "hi" else en
    if "paper_empty" in low or "0x80210003" in low or "-2145320957" in low or "feeder me" in low:
        return m["empty"]
    if "busy" in low or "0x80210006" in low:
        return m["busy"]
    if "jam" in low:
        return m["jam"]
    if "device driver threw" in low or "exception_in_driver" in low or "-2145320946" in low:
        return m["driver"]
    if "nahi mila" in low or "not found" in low or "no such" in low or "device" in low and "nahi" in low:
        return m["not_found"]
    if "twain" in low and ("open" in low or "khula" in low or "install" in low):
        return m["twain"]
    return m["generic"] + "\n\n(" + t[:200] + ")"


# ---------------------------------------------------------------------------
# Small translation layer (English / Hindi) for the main interface
# ---------------------------------------------------------------------------

T = {
    "scan": {"hi": "Scan", "en": "Scan"},
    "profile": {"hi": "Profile:", "en": "Profile:"},
    "profiles": {"hi": "Profiles…", "en": "Profiles…"},
    "claim_no": {"hi": "Claim No.:", "en": "Claim No.:"},
    "import": {"hi": "Import", "en": "Import"},
    "save_pdf": {"hi": "Save PDF", "en": "Save PDF"},
    "check": {"hi": "Check", "en": "Check"},
    "up": {"hi": "Upar", "en": "Up"},
    "down": {"hi": "Neeche", "en": "Down"},
    "delete": {"hi": "Delete", "en": "Delete"},
    "clear": {"hi": "Clear", "en": "Clear"},
    "rotate_l": {"hi": "Rotate ⟲", "en": "Rotate ⟲"},
    "rotate_r": {"hi": "Rotate ⟳", "en": "Rotate ⟳"},
    "menu_file": {"hi": "File", "en": "File"},
    "menu_edit": {"hi": "Edit", "en": "Edit"},
    "menu_tools": {"hi": "Tools", "en": "Tools"},
    "menu_settings": {"hi": "Settings", "en": "Settings"},
    "menu_help": {"hi": "Help", "en": "Help"},
    "help_guide": {"hi": "Guide / Madad", "en": "Guide / Help"},
    "about": {"hi": "About", "en": "About"},
    "whatsnew": {"hi": "Naya kya hai", "en": "What's new"},
    "feedback": {"hi": "Feedback / Sujhav", "en": "Feedback"},
    "language": {"hi": "Language / Bhasha", "en": "Language"},
    "scan_method": {"hi": "Scan method (TWAIN/WIA)", "en": "Scan method (TWAIN/WIA)"},
    "options": {"hi": "Options…", "en": "Options…"},
    "pages_hint": {"hi": "Pages (drag se order badlein)", "en": "Pages (drag to reorder)"},
    "simple_on": {"hi": "Simple mode (aasan)", "en": "Simple mode"},
    "scanner_ip": {"hi": "Scanner IP:", "en": "Scanner IP:"},
    "st_pages": {"hi": "Pages", "en": "Pages"},
    "st_profile": {"hi": "Profile", "en": "Profile"},
    "none_profile": {"hi": "koi profile nahi", "en": "no profile"},
    "checking": {"hi": "Connection check ho raha hai...", "en": "Checking connection..."},
    "select_first": {"hi": "Pehle koi page select karein.", "en": "Select a page first."},
    "scan_first": {"hi": "Pehle koi page scan/import karein.", "en": "Scan or import a page first."},
    "fast": {"hi": "Fast", "en": "Fast"},
    "save_all": {"hi": "Sab pages ki PDF", "en": "Save all pages (PDF)"},
    "save_sel": {"hi": "Selected pages ki PDF", "en": "Save selected pages (PDF)"},
    "no_selection": {"hi": "Koi page select nahi hai. Thumbnails me Ctrl/Shift se pages chuno.", "en": "No pages selected. Use Ctrl/Shift to select pages."},
}

# Aur bhashayein (adhuri — jo key translate nahi hui wo Hindi me dikhegi)
_EXTRA_LANGS = {
    "mr": {"up": "वर", "down": "खाली", "help_guide": "मदत / Guide", "whatsnew": "नवीन काय आहे",
           "none_profile": "प्रोफाइल नाही", "checking": "कनेक्शन तपासत आहे...",
           "select_first": "आधी एखादे पान निवडा.", "scan_first": "आधी पान scan/import करा.",
           "save_all": "सर्व पानांची PDF", "save_sel": "निवडलेल्या पानांची PDF",
           "pages_hint": "Pages (क्रम बदलण्यासाठी drag करा)",
           "no_selection": "कोणतेही पान निवडलेले नाही. Ctrl/Shift ने निवडा."},
    "gu": {"up": "ઉપર", "down": "નીચે", "help_guide": "મદદ / Guide", "whatsnew": "નવું શું છે",
           "none_profile": "કોઈ પ્રોફાઇલ નથી", "checking": "કનેક્શન તપાસી રહ્યાં છીએ...",
           "select_first": "પહેલા કોઈ પાનું પસંદ કરો.", "scan_first": "પહેલા પાનું scan/import કરો.",
           "save_all": "બધા પાનાની PDF", "save_sel": "પસંદ કરેલા પાનાની PDF",
           "pages_hint": "Pages (ક્રમ બદલવા drag કરો)",
           "no_selection": "કોઈ પાનું પસંદ નથી. Ctrl/Shift થી પસંદ કરો."},
    "pa": {"up": "ਉੱਪਰ", "down": "ਹੇਠਾਂ", "help_guide": "ਮਦਦ / Guide", "whatsnew": "ਨਵਾਂ ਕੀ ਹੈ",
           "none_profile": "ਕੋਈ ਪ੍ਰੋਫਾਈਲ ਨਹੀਂ", "checking": "ਕਨੈਕਸ਼ਨ ਚੈੱਕ ਹੋ ਰਿਹਾ ਹੈ...",
           "select_first": "ਪਹਿਲਾਂ ਕੋਈ ਪੰਨਾ ਚੁਣੋ।", "scan_first": "ਪਹਿਲਾਂ ਪੰਨਾ scan/import ਕਰੋ।",
           "save_all": "ਸਾਰੇ ਪੰਨਿਆਂ ਦੀ PDF", "save_sel": "ਚੁਣੇ ਪੰਨਿਆਂ ਦੀ PDF",
           "pages_hint": "Pages (ਕ੍ਰਮ ਬਦਲਣ ਲਈ drag ਕਰੋ)",
           "no_selection": "ਕੋਈ ਪੰਨਾ ਨਹੀਂ ਚੁਣਿਆ। Ctrl/Shift ਨਾਲ ਚੁਣੋ।"},
    "ta": {"up": "மேலே", "down": "கீழே", "help_guide": "உதவி / Guide", "whatsnew": "புதிதாக என்ன",
           "none_profile": "சுயவிவரம் இல்லை", "checking": "இணைப்பு சரிபார்க்கப்படுகிறது...",
           "select_first": "முதலில் ஒரு பக்கத்தைத் தேர்வு செய்யவும்.",
           "scan_first": "முதலில் பக்கத்தை scan/import செய்யவும்.",
           "save_all": "எல்லா பக்கங்களின் PDF", "save_sel": "தேர்ந்த பக்கங்களின் PDF",
           "pages_hint": "Pages (வரிசை மாற்ற drag)",
           "no_selection": "எந்தப் பக்கமும் தேர்வில்லை. Ctrl/Shift பயன்படுத்தவும்."},
}
for _lc, _tab in _EXTRA_LANGS.items():
    for _k, _v in _tab.items():
        T.setdefault(_k, {})[_lc] = _v


def tr(key, lang="hi"):
    e = T.get(key)
    if not e:
        return key
    return e.get(lang, e.get("hi", key))


HELP_TEXT_HI = """<h3>ApneScan — Madad</h3>
<b>Shuru kaise karein</b><br>
1. <b>Settings → Profiles…</b> me ek profile banayein: <i>Choose device</i> se apna
scanner chuno, DPI/Colour/Duplex set karo.<br>
2. Upar profile chuno aur <b>Scan</b> dabao (ya F5).<br>
3. Pages left me dikhenge — drag se order badlo, rotate/delete karo.<br>
4. <b>Save PDF</b> se PDF banao (OCR ka option milega).<br><br>
<b>Scanner nahi mil raha?</b><br>
Settings → <i>Scan method</i> me TWAIN aur WIA dono try karo. Scanner ka driver
install hona chahiye, aur woh ON + connected ho.<br><br>
<b>Fast scan ke liye:</b> Profile me DPI 200, Mode Black & White.<br><br>
<b>Auto-save, register, backup, watermark, split, merge, password PDF</b> —
sab <b>Settings → Options</b> aur <b>Tools</b> menu me hain.<br><br>
Keyboard: F5=Scan, Ctrl+S=Save PDF, Ctrl+P=Print, Ctrl+F=Search, Delete=Delete page.
"""

CHANGELOG_HTML = """<h3>ApneScan — Naya kya hai / What's new</h3>
<b>Version 1.0</b><br>
- TWAIN + WIA dono scanners support<br>
- Setup wizard (pehli baar aasan setup)<br>
- Profiles, auto-save PDF, auto file-naming, claim folders<br>
- OCR searchable PDF (Hindi+English), OCR text export<br>
- Blank-page removal, auto-crop, deskew, quality enhance<br>
- PDF compress, password PDF, split, merge, print<br>
- Excel register, activity log, duplicate check, daily backup<br>
- Barcode/QR se claim number, monthly report<br>
- Hindi / English interface, Simple mode, Feedback, in-app Help<br>
"""


HELP_TEXT_EN = """<h3>ApneScan — Help</h3>
<b>Getting started</b><br>
1. Go to <b>Settings → Profiles…</b> and create a profile: click <i>Choose device</i>
to pick your scanner, then set DPI/Colour/Duplex.<br>
2. Select the profile at the top and click <b>Scan</b> (or press F5).<br>
3. Pages appear on the left — drag to reorder, rotate/delete.<br>
4. Click <b>Save PDF</b> (you'll be asked about OCR).<br><br>
<b>Scanner not detected?</b><br>
In Settings → <i>Scan method</i>, try both TWAIN and WIA. The scanner's driver
must be installed, and the device must be on and connected.<br><br>
<b>For faster scans:</b> set DPI 200 and Black & White in the profile.<br><br>
<b>Auto-save, register, backup, watermark, split, merge, password PDF</b> are all
under <b>Settings → Options</b> and the <b>Tools</b> menu.<br><br>
Shortcuts: F5=Scan, Ctrl+S=Save PDF, Ctrl+P=Print, Ctrl+F=Search, Delete=Delete page.
"""



# ---------------------------------------------------------------------------
# NAPS2 engine backend (uses NAPS2's console to scan — reliable fast + duplex)
# ---------------------------------------------------------------------------

NAPS2_GUESSES = [
    r"C:\\Program Files\\NAPS2\\NAPS2.Console.exe",
    r"C:\\Program Files (x86)\\NAPS2\\NAPS2.Console.exe",
    r"C:\\Program Files\\NAPS2\\NAPS2.exe",
    r"C:\\Program Files (x86)\\NAPS2\\NAPS2.exe",
]


def find_naps2():
    for p in NAPS2_GUESSES:
        if os.path.exists(p):
            return p
    return ""


def scan_via_naps2(naps2_exe, profile, tmpdir, duplex, on_page=None, should_stop=None):
    import subprocess
    import glob as _glob
    import re as _re
    if not naps2_exe or not os.path.exists(naps2_exe):
        raise ScannerError('NAPS2 not found. Set the NAPS2 path in Settings.')
    # clean old temp scans
    for f in _glob.glob(os.path.join(tmpdir, "naps2_*.jpg")):
        try:
            os.remove(f)
        except Exception:
            pass
    out_pattern = os.path.join(tmpdir, "naps2_$(n).jpg")
    cmd = [naps2_exe, "-o", out_pattern, "--split", "-f", "-v"]
    if profile:
        cmd += ["-p", profile]
    if duplex:
        cmd += ["--source", "duplex"]
    try:
        proc = subprocess.run(cmd, capture_output=True, text=True, timeout=600,
                              creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0))
    except subprocess.TimeoutExpired:
        raise ScannerError('NAPS2 scan took too long (timeout).')
    except Exception as exc:
        raise ScannerError("NAPS2 chalane me dikkat: %s" % exc)

    files = _glob.glob(os.path.join(tmpdir, "naps2_*.jpg"))

    def _num(p):
        m = _re.search(r"naps2_(\d+)\.jpg$", os.path.basename(p))
        return int(m.group(1)) if m else 0
    files.sort(key=_num)

    if not files:
        msg = (proc.stderr or proc.stdout or "").strip()
        raise ScannerError('No page came from NAPS2.\nIs the profile name correct? Is the scanner ON?\n\n%s'
                           % msg[:300])
    count = 0
    for f in files:
        if should_stop and should_stop():
            break
        try:
            with Image.open(f) as im:
                img = im.convert("RGB").copy()
        except Exception:
            continue
        try:
            os.remove(f)
        except Exception:
            pass
        count += 1
        if on_page:
            on_page(img)
    if count == 0:
        raise ScannerError('No page came from NAPS2.')
    return count


# ---------------------------------------------------------------------------
# eSCL (AirScan) network backend — pure Python, talks to the scanner's IP over
# HTTP. Supports DUPLEX over the network with NO other software needed.
# ---------------------------------------------------------------------------

def scan_via_escl(ip, dpi, color, duplex, on_page=None, should_stop=None, page_size="auto"):
    import urllib.request as _u
    import urllib.error as _ue
    import io as _io
    import time as _t

    if not ip:
        raise ScannerError('Scanner IP is not set. Enter it in Settings → Scanner IP (e.g. 192.168.1.8).')
    ip = ip.strip()
    base = "http://%s/eSCL" % ip
    # NOTE: "BlackAndWhite1" + image/jpeg ek saath kaam NAHI karta (JPEG 1-bit
    # ho hi nahi sakta). HP N4000 aisi job me pages kheenchta rehta hai par data
    # kabhi taiyar nahi karta — page feeder me atak jata tha aur app ko 0 page
    # milte the (Fast/B&W mode ka jam wala bug). Isliye B&W bhi Grayscale8 me
    # scan hota hai; asli 1-bit conversion app me hota hai (ScanWorker save step).
    cmode = {"color": "RGB24", "gray": "Grayscale8", "bw": "Grayscale8"}.get(color, "RGB24")

    # eSCL region units are 1/300 inch. Fixed sizes request that exact sheet.
    # "auto" requests the FULL BED WIDTH (so sides are never cut) with a Legal-length
    # height, and MustHonor="false" so the ADF stops at each sheet's real trailing
    # edge (no forced long scan = no feeder jam). The dark backing that appears
    # around a narrow sheet is painted white by whiten_dark_background.
    ps = (page_size or "auto").lower()
    if ps.startswith("a4"):
        w, h = 2480, 3550
    elif ps.startswith("letter"):
        w, h = 2550, 3300
    elif ps.startswith("legal"):
        w, h = 2550, 4200
    elif ps.startswith("a5"):
        w, h = 1748, 2480
    elif ps.startswith("custom"):
        # "custom:800" => 800mm lambi parchi (receipt/bahi-khata). Width full bed.
        mm = 600
        _m = re.findall(r"(\d+)", ps)
        if _m:
            mm = max(100, min(3000, int(_m[0])))
        w, h = 2550, int(round(mm * 300 / 25.4))
    else:  # auto -> full width, Legal length, let ADF stop at paper end
        w, h = 2550, 4200
        try:
            _rr = _u.urlopen(base + "/ScannerCapabilities", timeout=8)
            _cap = _rr.read().decode("utf-8", "ignore")
            _mw = re.findall(r"MaxWidth>\s*(\d+)", _cap)
            if _mw:
                w = max(int(x) for x in _mw)   # full bed width (no side cut)
        except Exception:
            pass

    settings = (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<scan:ScanSettings xmlns:scan="http://schemas.hp.com/imaging/escl/2011/05/03"'
        ' xmlns:pwg="http://www.pwg.org/schemas/2010/12/sm">\n'
        '<pwg:Version>2.6</pwg:Version>\n'
        '<scan:Intent>Document</scan:Intent>\n'
        '<pwg:ScanRegions pwg:MustHonor="false"><pwg:ScanRegion>\n'
        '<pwg:Height>%d</pwg:Height><pwg:Width>%d</pwg:Width>\n'
        '<pwg:XOffset>0</pwg:XOffset><pwg:YOffset>0</pwg:YOffset>\n'
        '</pwg:ScanRegion></pwg:ScanRegions>\n'
        '<pwg:InputSource>Feeder</pwg:InputSource>\n'
        '<scan:Duplex>%s</scan:Duplex>\n'
        '<scan:ColorMode>%s</scan:ColorMode>\n'
        '<scan:XResolution>%d</scan:XResolution>\n'
        '<scan:YResolution>%d</scan:YResolution>\n'
        '<pwg:DocumentFormat>image/jpeg</pwg:DocumentFormat>\n'
        '</scan:ScanSettings>'
    ) % (h, w, "true" if duplex else "false", cmode, dpi, dpi)

    # Before starting, clear any STUCK job (e.g. a "Processing" job left over from
    # an earlier interrupted scan). Such a job makes every new ScanJobs POST return
    # 503 even though ScannerStatus shows Idle. We read ScannerStatus, find job
    # URIs/UUIDs whose state is not finished, and DELETE them.
    def _clear_stuck_jobs():
        try:
            rr = _u.urlopen(base + "/ScannerStatus", timeout=8)
            xml = rr.read().decode("utf-8", "ignore")
        except Exception:
            return
        blocks = re.findall(r"<[a-zA-Z]*:?JobInfo>.*?</[a-zA-Z]*:?JobInfo>", xml, re.S)
        for blk in blocks:
            stm = re.search(r"JobState>\s*([A-Za-z]+)", blk)
            state = (stm.group(1).lower() if stm else "")
            if state in ("completed", "canceled", "aborted", ""):
                continue                          # already finished -> leave it
            uri = re.search(r"JobUri>\s*([^<\s]+)", blk)
            uuid = re.search(r"JobUuid>\s*([^<\s]+)", blk)
            target = None
            if uri:
                target = uri.group(1)
                if target.startswith("/"):
                    target = "http://%s%s" % (ip, target)
            elif uuid:
                target = "%s/ScanJobs/%s" % (base, uuid.group(1))
            if target:
                try:
                    _u.urlopen(_u.Request(target, method="DELETE"), timeout=8)
                    _t.sleep(0.5)
                except Exception:
                    pass

    _clear_stuck_jobs()

    # Create the scan job. Try immediately first (this worked before); only if the
    # scanner reports busy (503/409) do we wait briefly and retry a couple of times.
    def _post_job():
        req = _u.Request(base + "/ScanJobs", data=settings.encode("utf-8"),
                         headers={"Content-Type": "text/xml"}, method="POST")
        return _u.urlopen(req, timeout=45)

    resp = None
    last_code = None
    for attempt in range(10):                    # be patient: scanner is briefly busy at start
        if should_stop and should_stop():
            break
        try:
            resp = _post_job()
            break
        except _ue.HTTPError as e:
            last_code = e.code
            if e.code in (503, 409, 429):        # busy -> wait and keep trying
                _t.sleep(1.5)
                continue
            raise ScannerError("eSCL job error: HTTP %s. Scanner network-scan (eSCL) support karta hai?" % e.code)
        except Exception as e:
            last_code = str(e)
            _t.sleep(1.2)
            continue
    if resp is None:
        raise ScannerError(
            "The scanner has been busy for a while (HTTP %s). If another scan app "
            "(NAPS2 / HP Scan / an old ApneScan window) is open, close it, "
            "or turn the scanner off/on once." % last_code)

    job = resp.headers.get("Location")
    if not job:
        raise ScannerError('eSCL job location not found (the scanner did not start the job).')
    if job.startswith("/"):
        job = "http://%s%s" % (ip, job)

    count = 0
    empties = 0
    hiccups = 0
    try:
        while True:
            if should_stop and should_stop():
                break
            try:
                r = _u.urlopen(job + "/NextDocument", timeout=90)
            except _ue.HTTPError as e:
                if e.code in (404, 410):
                    break                       # no more pages -> done
                if e.code == 503:
                    _t.sleep(1); empties += 1
                    if empties > 30:
                        break
                    continue
                break
            except Exception:
                # network hiccup (timeout/reset) — turant give-up mat karo, warna
                # job DELETE ho jata hai aur beech ka page feeder me atak jata hai
                hiccups += 1
                if hiccups > 3:
                    break
                _t.sleep(1.0)
                continue
            data = r.read()
            if not data:
                break
            try:
                with Image.open(_io.BytesIO(data)) as im:
                    img = im.convert("RGB").copy()
            except Exception:
                break
            count += 1
            empties = 0
            hiccups = 0
            if on_page:
                on_page(img)
    finally:
        # ALWAYS release the scan job (even on error/interrupt) so the scanner is
        # free for the next scan. A job left open is the #1 cause of later "busy"/503.
        try:
            _u.urlopen(_u.Request(job, method="DELETE"), timeout=10)
        except Exception:
            pass

    if count == 0:
        # Sach-sach batao kya hua: jam / khaali feeder / kuch aur. (Pehle har
        # zero-page case par "feeder khaali" dikhta tha — jam me bhi.)
        st = ""
        try:
            rr = _u.urlopen(base + "/ScannerStatus", timeout=8)
            st = rr.read().decode("utf-8", "ignore").lower()
        except Exception:
            pass
        if "jam" in st:
            raise ScannerError(
                'Paper jam: a page is stuck inside the scanner. Gently remove it, place the pages straight in the feeder, and scan again.')
        if "adfempty" in st.replace(" ", ""):
            raise ScannerError('PAPER_EMPTY: the feeder is empty.')
        raise ScannerError(
            'eSCL: the scanner produced no page. Turn it off/on and try again; if it persists, run Help > eSCL Test.')
    return count


# Role for storing a custom (document-name) label on a thumbnail item.
TITLE_ROLE = int(QtCore.Qt.UserRole) + 1
NAMEKEY_ROLE = int(QtCore.Qt.UserRole) + 2   # normalized key for remembering saved names


def underscore_name(s):
    """Filename-friendly: spaces -> underscore, drop odd chars, collapse repeats."""
    s = re.sub(r"[^\w\s.\-]", "", s or "")
    s = re.sub(r"\s+", "_", s.strip())
    s = re.sub(r"_+", "_", s)
    return s.strip("_.")


def name_key(s):
    """Normalized key to match the 'same' document title across scans."""
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


# Document TYPE classifier — recognise WHAT the document is (by tell-tale keywords)
# instead of naming it after whatever text is printed on top (which is usually just
# the hospital name and would be the same for every page). Tuned for ECHS/RGHS
# hospital paperwork. Order = priority when tie-broken.
DOC_TYPES = [
    # Only STRONG, distinctive phrases (avoid generic words that appear on every
    # hospital form like self/spouse/patient/hospital). If nothing here matches
    # clearly, the page stays "Page N" (better than a wrong name).
    ("Discharge_Summary", ["discharge summary"]),
    ("Bill",              ["cash receipt", "final bill", "bill of supply", "tax invoice",
                           "bill no.", "receipt no.", "total amount", "net payable"]),
    ("Referral",          ["referral form", "referral slip", "individual referred",
                           "referred hospital", "empanelled hospital"]),
    ("ECHS_Card",         ["semi-pvt", "echs card", "echs smart card", "ex-serviceman card",
                           "esm :"]),
    ("Aadhaar",           ["unique identification authority", "aadhaar", "आधार",
                           "government of india"]),
    ("Prescription",      ["provisional diagnosis", "chief complaint", "on examination",
                           "treatment advised", "diagnosis :", "complaints :"]),
    ("Lab_Report",        ["laboratory report", "pathology report", "haematology",
                           "biochemistry", "specimen", "sample collected", "test result"]),
    ("Xray_Radiology",    ["x-ray", "radiology report", "mri scan", "ct scan",
                           "ultrasound report", "sonography", "impression :"]),
    ("Registration",      ["registration no.", "uhid no", "op registration"]),
    ("Consent_Form",      ["consent form", "i hereby give my consent", "informed consent"]),
    ("Discharge_Card",    ["discharge card"]),
]


def page_ocr_text(path, frac=0.55):
    """OCR the top `frac` of a page and return the raw text (used by both the
    document-type classifier and the learned-name matcher)."""
    if not HAS_OCR_LIBS:
        return ""
    try:
        with Image.open(path) as im:
            w, h = im.size
            crop = im.crop((0, 0, w, max(1, int(h * frac)))).convert("L")
            if crop.width > 1500:
                r = 1500.0 / crop.width
                crop = crop.resize((1500, max(1, int(crop.height * r))))
        try:
            return pytesseract.image_to_string(crop, lang="eng+hin")
        except Exception:
            return pytesseract.image_to_string(crop)
    except Exception:
        return ""


def sig_words(text):
    """Kisi form ki FIXED chhapi layout (letterhead, field-labels, title) ko
    pehchanne wale khaas shabd. Patient/customer-specific data (naam, number)
    zyadatar nikal jata hai, isliye SAME form ke do scan inme se zyadatar shabd
    share karte hain.

    Ab ENGLISH + HINDI (Devanagari) DONO pakadta hai — pehle sirf English tha,
    isliye Hindi forms ka naam theek se yaad nahi rehta tha."""
    low = (text or "").lower()
    en = re.findall(r"[a-z]{3,}", low)          # 3+ (pehle 4+ tha)
    hi = re.findall(r"[ऀ-ॿ]{2,}", text or "")   # Devanagari shabd
    stop = {"name", "date", "time", "age", "years", "year", "male", "female",
            "address", "sign", "signature", "page", "self", "spouse", "mobile",
            "phone", "number", "the", "and", "for", "with", "this", "that",
            "नाम", "दिनांक", "पता", "उम्र", "हस्ताक्षर", "मोबाइल", "फोन"}
    return set(w for w in (en + hi) if w not in stop)


def _jaccard(a, b):
    if not a or not b:
        return 0.0
    inter = len(a & b)
    union = len(a | b)
    return inter / union if union else 0.0


def _containment(sig, stored):
    """Stored form-signature ka kitna hissa NAYE page me maujood hai (0..1).
    Ye Jaccard se behtar hai jab naye scan me zyada/kam text ho — kyunki ek hi
    form dobara scan karne par uska poora signature naye me aa hi jata hai."""
    if not sig or not stored:
        return 0.0
    return len(sig & stored) / float(len(stored))


def learned_name_for(text, learned, thr=0.45):
    """learned = list of (wordset, name). Jis stored signature se ye page sabse
    zyada milta hai uska naam do (agar kaafi milta ho). Jaccard YA containment —
    dono me se jo behtar ho use lete hain (thr thoda narm bhi kiya hai)."""
    sig = sig_words(text)
    if not sig:
        return None
    best_name, best_score = None, 0.0
    for words, name in learned:
        ws = words if isinstance(words, set) else set(words)
        # containment ko halka zyada wazan (same form dobara aane par ~1.0 hota hai)
        sc = max(_jaccard(sig, ws), _containment(sig, ws) * 0.95)
        if sc > best_score:
            best_score, best_name = sc, name
    # thr=0.40 (narm) — par containment 0.6+ ho to bhi maan lo
    return best_name if best_score >= 0.40 else None


def extract_doc_number(text):
    """Page ke text me se sabse sambhavit 'document number' (invoice/bill/claim/
    receipt no.) nikaalo. 'No/No./Number/Bill/Invoice/Claim/Receipt' ke aage wala
    number pehle dekha jata hai; na mile to koi lamba alphanumeric code."""
    t = text or ""
    m = re.search(r"(?:invoice|bill|claim|receipt|ref(?:erence)?|no|number|"
                  r"बिल|रसीद|क्लेम|संख्या|नंबर)\.?\s*(?:संख्या|नंबर|no|number)?\.?"
                  r"\s*[:#-]?\s*([A-Za-z]{0,4}\d[\w\-/]{2,20})",
                  t, re.IGNORECASE)
    if m:
        return re.sub(r"[^A-Za-z0-9\-]", "", m.group(1))[:24]
    m = re.search(r"\b([A-Z]{2,5}[-/]?\d{3,})\b", t)   # jaise INV-1234, ABC/2026/45
    if m:
        return re.sub(r"[^A-Za-z0-9\-]", "", m.group(1))[:24]
    return ""


def classify_from_text(text):
    t = " " + re.sub(r"\s+", " ", (text or "").lower()) + " "
    scores = []
    for name, kws in DOC_TYPES:
        n = sum(1 for kw in kws if kw in t)
        if n > 0:
            scores.append((n, name))
    if not scores:
        return None
    scores.sort(reverse=True)
    # Accept only a CLEAR winner (top strictly beats runner-up); ties -> None.
    if len(scores) >= 2 and scores[0][0] == scores[1][0]:
        return None
    return scores[0][1]


def classify_document(path):
    """Return a clean document-TYPE name (e.g. Discharge_Summary, Bill, ECHS_Card)
    by matching keywords in the page text. None if no type is confidently found."""
    if not HAS_OCR_LIBS:
        return None
    return classify_from_text(page_ocr_text(path, 0.6))


def ocr_page_title(path):
    """Read the top of a scanned page and return a likely PRINTED document title
    (e.g. 'DISCHARGE SUMMARY'). Uses OCR word-confidence so low-confidence
    handwriting/garbage is rejected. None if nothing reliable is found."""
    if not HAS_OCR_LIBS:
        return None

    def _clean(ln):
        ln = re.sub(r"[^A-Za-z0-9 &.\-/,]", "", ln)
        return re.sub(r"\s+", " ", ln).strip()

    def _looks_garbage(ln):
        # a single alphabetic run longer than 15 chars with no space = OCR garbage
        for tok in ln.split():
            if tok.isalpha() and len(tok) > 15:
                return True
        return False

    def _title_from(fraction):
        try:
            with Image.open(path) as im:
                w, h = im.size
                strip = im.crop((0, 0, w, max(1, int(h * fraction)))).convert("L")
                if strip.width > 1400:
                    r = 1400.0 / strip.width
                    strip = strip.resize((1400, max(1, int(strip.height * r))))
            # Confidence-aware: build lines only from words OCR is confident about.
            try:
                try:
                    data = pytesseract.image_to_data(strip, lang="eng+hin",
                                                     output_type=pytesseract.Output.DICT)
                except Exception:
                    data = pytesseract.image_to_data(strip,
                                                     output_type=pytesseract.Output.DICT)
                lines = {}
                for i in range(len(data.get("text", []))):
                    txt = (data["text"][i] or "").strip()
                    try:
                        conf = float(data["conf"][i])
                    except Exception:
                        conf = -1
                    if not txt or conf < 55:
                        continue
                    if sum(c.isalnum() for c in txt) < 2:
                        continue
                    key = (data["block_num"][i], data["par_num"][i], data["line_num"][i])
                    lines.setdefault(key, []).append((txt, conf))
                cands = []
                for words in lines.values():
                    text = _clean(" ".join(wtc[0] for wtc in words))
                    if len(text) < 4 or _looks_garbage(text):
                        continue
                    avg = sum(wtc[1] for wtc in words) / len(words)
                    cands.append((text, avg))
                if cands:
                    def score(tc):
                        ln, cf = tc
                        letters = sum(c.isalpha() for c in ln)
                        caps = sum(1 for c in ln if c.isupper())
                        return cf * 0.5 + letters + caps * 0.4 - abs(len(ln) - 22) * 0.05
                    best = max(cands, key=score)[0]
                    if sum(c.isalpha() for c in best) >= 3:
                        return best[:38]
            except Exception:
                pass
            return None
        except Exception:
            return None

    return _title_from(0.22) or _title_from(0.45)


class NameWorker(QtCore.QThread):
    named = QtCore.pyqtSignal(int, str)   # (row, title)
    finished_all = QtCore.pyqtSignal()

    def __init__(self, items, learned=None):
        super().__init__()
        self.items = items                # list of (row, path)
        self.learned = learned or []      # list of (wordset, name)

    def run(self):
        for row, path in self.items:
            if self.isInterruptionRequested():
                break
            text = page_ocr_text(path, 0.6)
            # priority: name YOU taught it > document-type > printed heading
            title = (learned_name_for(text, self.learned)
                     or classify_from_text(text)
                     or ocr_page_title(path))
            if title:
                self.named.emit(row, title)
        self.finished_all.emit()


class FuncWorker(QtCore.QThread):
    """Koi bhi bhaari kaam (OCR/network) background me chala kar result de —
    UI kabhi nahi rukti. Result ya Exception, dono `done` signal me aate hain."""
    done = QtCore.pyqtSignal(object)

    def __init__(self, fn):
        super().__init__()
        self.fn = fn

    def run(self):
        try:
            self.done.emit(self.fn())
        except Exception as e:
            self.done.emit(e)


class LearnWorker(QtCore.QThread):
    """Rename ke baad naam 'seekhne' ka OCR background me — pehle ye UI thread
    par chalta tha aur har rename par app 3-10 second jam jaati thi."""
    got = QtCore.pyqtSignal(str, list)   # (name, words)

    def __init__(self, path, name):
        super().__init__()
        self.path = path
        self.name = name

    def run(self):
        try:
            # Zyada text (0.75) padho taaki form ka signature bharpoor bane —
            # naam yaad rakhna jitna bharpoor, utna reliable.
            words = sorted(sig_words(page_ocr_text(self.path, 0.75)))
        except Exception:
            words = []
        self.got.emit(self.name, words)


class ImportWorker(QtCore.QThread):
    """Import (images/PDF/photo-cleanup) background me — UI kabhi nahi rukti.
    Thumbnail bhi yahin chhota (QImage worker thread me safe hai)."""
    page_ready = QtCore.pyqtSignal(str, QtGui.QImage)
    done = QtCore.pyqtSignal(int)

    def __init__(self, files, tmpdir, mode="normal", thumb_w=360, thumb_h=480):
        super().__init__()
        self.files = list(files)
        self.tmpdir = tmpdir
        self.mode = mode          # "normal" ya "photo" (phone-photo cleanup)
        self.thumb_w = thumb_w
        self.thumb_h = thumb_h

    def _thumb(self, path):
        r = QtGui.QImageReader(path)
        r.setAutoTransform(True)
        sz = r.size()
        if sz.isValid() and sz.width() > 0 and sz.height() > 0:
            s = min(float(self.thumb_w) / sz.width(),
                    float(self.thumb_h) / sz.height(), 1.0)
            r.setScaledSize(QtCore.QSize(max(1, int(sz.width() * s)),
                                         max(1, int(sz.height() * s))))
        return r.read()

    def run(self):
        count = 0
        for f in self.files:
            if self.isInterruptionRequested():
                break
            try:
                if self.mode == "photo":
                    with Image.open(f) as im:
                        img = clean_photo(im)
                    fd, out = tempfile.mkstemp(suffix=".jpg", dir=self.tmpdir)
                    os.close(fd)
                    img.save(out, "JPEG", quality=90)
                    outs = [out]
                elif f.lower().endswith(".pdf"):
                    outs = pdf_to_images(f, self.tmpdir) or []
                else:
                    with Image.open(f) as im:
                        fd, out = tempfile.mkstemp(suffix=".png", dir=self.tmpdir)
                        os.close(fd)
                        im.convert("RGB").save(out, "PNG")
                    outs = [out]
                for p in outs:
                    count += 1
                    self.page_ready.emit(p, self._thumb(p))
            except Exception:
                pass
        self.done.emit(count)


class ScanWorker(QtCore.QThread):
    page_done = QtCore.pyqtSignal(str)
    done = QtCore.pyqtSignal(int, int)
    failed = QtCore.pyqtSignal(str)

    def __init__(self, hwnd, source_name, dpi, pixel_type, duplex, tmpdir, opts):
        super().__init__()
        self.hwnd = hwnd
        self.source_name = source_name
        self.dpi = dpi
        self.pixel_type = pixel_type
        self.duplex = duplex
        self.tmpdir = tmpdir
        self.opts = opts or {}
        self.kept = 0
        self.skipped = 0

    def run(self):
        import queue as _q, threading as _th
        pageq = _q.Queue()

        def _consumer():
            # Runs in parallel: saves/processes each page while the scanner
            # is already pulling the NEXT one (this removes the per-page pause).
            while True:
                img = pageq.get()
                if img is None:
                    break
                try:
                    if self.opts.get("remove_blank"):
                        _sens = {"kam": 0.0003, "normal": 0.0008, "zyada": 0.004}
                        _thr = _sens.get(self.opts.get("blank_sensitivity", "normal"), 0.0008)
                        if is_blank_page(img, _thr):
                            self.skipped += 1
                            continue
                    # Backing (kala/navy/gray) HAMESHA white karo — kisi bhi page
                    # size me sheet chhoti ho to backing dikhti hai; print me
                    # kala bilkul nahi aana chahiye.
                    img = whiten_dark_background(img)
                    if self.opts.get("clean_edges"):
                        img = clean_edges(img)
                    if self.opts.get("auto_crop"):
                        img = autocrop(img)
                    if self.opts.get("deskew"):
                        img = deskew(img)
                    if self.opts.get("quality_enhance"):
                        img = auto_enhance(img)
                    if self.opts.get("auto_orient") and tesseract_available():
                        img = auto_orient(img)
                    # Rang-heen page ko gray bana do (chhoti file) — sirf colour
                    # scan me, aur sirf jab option ON ho.
                    if (self.opts.get("auto_colour") and self.pixel_type == "color"
                            and colorfulness(img) < 6.0):
                        img = img.convert("L")
                    # Ek glass par do page → do alag page (option ON ho to)
                    if self.opts.get("split_two_page"):
                        parts = split_two_pages(img)
                    else:
                        parts = [img]
                    for _pi in parts:
                        self._save_and_emit(_pi)
                except Exception:
                    pass

        def _save_and_emit(im):
            # Ek image ko temp file me save karke UI ko bhejo. JPEG encode ~5x
            # tez (bade scan me speed) — colour/grey ke liye JPEG, 1-bit B&W ke
            # liye PNG.
            if self.pixel_type == "bw":
                fd, out = tempfile.mkstemp(suffix=".png", dir=self.tmpdir)
                os.close(fd)
                try:
                    if im.mode != "1":
                        im = im.convert("L").point(
                            lambda v: 255 if v >= 160 else 0, mode="1")
                except Exception:
                    pass
                try:
                    im.save(out, "PNG", compress_level=1)
                except Exception:
                    im.save(out, "PNG")
            else:
                fd, out = tempfile.mkstemp(suffix=".jpg", dir=self.tmpdir)
                os.close(fd)
                try:
                    if im.mode == "L":
                        im.save(out, "JPEG", quality=88)
                    else:
                        im.convert("RGB").save(out, "JPEG", quality=88)
                except Exception:
                    im.convert("RGB").save(out, "JPEG", quality=85)
            self.kept += 1
            self.page_done.emit(out)
        self._save_and_emit = _save_and_emit

        consumer = _th.Thread(target=_consumer, daemon=True)
        consumer.start()

        def _on_page(img):
            # Producer: hand the page off instantly and let the scanner keep going.
            pageq.put(img)

        err = None
        method = self.opts.get("scanner_method", "twain")
        try:
            _ps = str(self.opts.get("page_size", "auto"))
            if _ps.startswith("custom"):
                _ps = "custom:%d" % int(self.opts.get("custom_page_mm", 600) or 600)
            if method == "escl":
                scan_via_escl(self.opts.get("scanner_ip"), self.dpi, self.pixel_type,
                              self.duplex, _on_page, should_stop=self.isInterruptionRequested,
                              page_size=_ps)
            elif method == "naps2":
                scan_via_naps2(self.opts.get("naps2_path") or find_naps2(),
                               self.opts.get("naps2_profile"), self.tmpdir,
                               self.duplex, _on_page,
                               should_stop=self.isInterruptionRequested)
            elif method == "wia":
                try:
                    _pythoncom.CoInitialize()
                except Exception:
                    pass
                try:
                    wia_scan_pages(self.opts.get("wia_device_id"), self.dpi,
                                   self.pixel_type, self.duplex, _on_page,
                                   should_stop=self.isInterruptionRequested,
                                   source=self.opts.get("paper_source", "auto"))
                finally:
                    try:
                        _pythoncom.CoUninitialize()
                    except Exception:
                        pass
            else:
                scan_pages(self.hwnd, self.source_name, self.dpi,
                           self.pixel_type, self.duplex, _on_page,
                           should_stop=self.isInterruptionRequested,
                           file_xfer=bool(self.opts.get("twain_file_xfer")))
        except ScannerError as exc:
            err = str(exc)
        except Exception:
            err = "Scan error:\n%s" % traceback.format_exc()

        pageq.put(None)            # tell the saver thread to finish
        consumer.join(timeout=60)  # wait until all pages are saved

        if err is not None and self.kept == 0:
            self.failed.emit(err)
            return
        if self.kept == 0 and self.skipped == 0:
            self.failed.emit(err or 'No page was scanned.')
            return
        self.done.emit(self.kept, self.skipped)


# ---------------------------------------------------------------------------
# Dialogs
# ---------------------------------------------------------------------------

class EditProfileDialog(QtWidgets.QDialog):
    def __init__(self, parent, profile=None):
        super().__init__(parent)
        self.setWindowTitle("Profile settings")
        self.setMinimumWidth(420)
        self.profile = dict(profile) if profile else {
            "name": "", "source_name": None, "dpi": 200, "color": "gray", "duplex": False}
        form = QtWidgets.QFormLayout(self)

        def qh(text, tip):
            # "?" help-icon wala row-label (Hindi + English hover)
            h = QtWidgets.QLabel("?")
            h.setToolTip(tip); h.setCursor(QtCore.Qt.WhatsThisCursor)
            h.setStyleSheet("QLabel{color:#0f766e;border:1px solid #0f766e;border-radius:9px;"
                            "min-width:18px;max-width:18px;min-height:18px;max-height:18px;"
                            "font-weight:bold;qproperty-alignment:AlignCenter;}")
            r = QtWidgets.QHBoxLayout(); r.setContentsMargins(0, 0, 0, 0); r.setSpacing(6)
            r.addWidget(QtWidgets.QLabel(text)); r.addWidget(h); r.addStretch(1)
            w = QtWidgets.QWidget(); w.setLayout(r); return w
        self.name_edit = QtWidgets.QLineEdit(self.profile.get("name", ""))
        self.name_edit.setPlaceholderText("e.g. Documents fast")
        self.name_edit.setToolTip("A name for this profile (e.g. 'Fast B&W'); you'll pick it by this name.")
        form.addRow(qh("Display Name:", self.name_edit.toolTip()), self.name_edit)
        dev_row = QtWidgets.QHBoxLayout()
        self.device_label = QtWidgets.QLabel(self.profile.get("source_name") or "(no device)")
        btn_dev = QtWidgets.QPushButton("Choose device")
        btn_dev.clicked.connect(self._choose_device)
        dev_row.addWidget(self.device_label, 1); dev_row.addWidget(btn_dev)
        dw = QtWidgets.QWidget(); dw.setLayout(dev_row)
        form.addRow(qh("Device:", "हिन्दी: स्कैनर प्रोफ़ाइल में कौन-सा स्कैनर इस्तेमाल होगा (TWAIN डिवाइस चुनें)।\nEnglish: Which scanner this profile uses (choose a TWAIN device)."), dw)
        self.cmb_dpi = QtWidgets.QComboBox(); self.cmb_dpi.setEditable(True)
        self.cmb_dpi.setValidator(QtGui.QIntValidator(50, 1200, self.cmb_dpi))
        self.cmb_dpi.addItems(RESOLUTIONS)
        self.cmb_dpi.setToolTip('You can type any DPI here (e.g. 250).')
        self.cmb_dpi.setCurrentText(str(self.profile.get("dpi", 200)))
        form.addRow(qh("Resolution (DPI):", "हिन्दी: स्कैन की सफ़ाई (रेज़ॉल्यूशन)। 200 = तेज़ और टेक्स्ट के लिए ठीक; 300 = बेहतर (धीमा); 600 = फ़ोटो के लिए।\nEnglish: Scan sharpness. 200 = fast, fine for text; 300 = better (slower); 600 = for photos."), self.cmb_dpi)
        self.cmb_color = QtWidgets.QComboBox(); self.cmb_color.addItems(list(COLOUR_MODES.keys()))
        for label, code in COLOUR_MODES.items():
            if code == self.profile.get("color", "gray"):
                self.cmb_color.setCurrentText(label)
        form.addRow(qh("Colour mode:", "हिन्दी: रंगीन = colour, Grayscale = काले-सफ़ेद shades, B&W = सिर्फ़ काला/सफ़ेद (सबसे छोटी फ़ाइल, तेज़)।\nEnglish: Colour, Grayscale, or pure Black & White (smallest & fastest)."), self.cmb_color)
        self.chk_duplex = QtWidgets.QCheckBox("Duplex (both sides)")
        self.chk_duplex.setChecked(bool(self.profile.get("duplex")))
        self.chk_duplex.setToolTip("ON: scan both sides of the paper automatically (duplex ADF).")
        _dxr = QtWidgets.QHBoxLayout(); _dxr.setContentsMargins(0, 0, 0, 0); _dxr.setSpacing(6)
        _dxh = QtWidgets.QLabel("?"); _dxh.setToolTip(self.chk_duplex.toolTip())
        _dxh.setCursor(QtCore.Qt.WhatsThisCursor)
        _dxh.setStyleSheet("QLabel{color:#0f766e;border:1px solid #0f766e;border-radius:9px;"
                           "min-width:18px;max-width:18px;min-height:18px;max-height:18px;"
                           "font-weight:bold;qproperty-alignment:AlignCenter;}")
        _dxr.addWidget(self.chk_duplex); _dxr.addWidget(_dxh); _dxr.addStretch(1)
        _dxw = QtWidgets.QWidget(); _dxw.setLayout(_dxr)
        form.addRow("", _dxw)
        self.cmb_psize = QtWidgets.QComboBox()
        self._PSIZES = [("Auto (detect each page's size)", "auto"),
                        ("A4 (210x297 mm)", "a4"), ("Letter", "letter"),
                        ("Legal", "legal"), ("A5", "a5"),
                        ("Long receipt / custom (length in Settings)", "custom")]
        self.cmb_psize.addItems([t for t, _c in self._PSIZES])
        _ps = (self.profile.get("page_size") or "auto").lower()
        _idx = next((k for k, (_t, c) in enumerate(self._PSIZES) if _ps.startswith(c[:4])), 0)
        self.cmb_psize.setCurrentIndex(_idx)
        form.addRow(qh("Page size:", "हिन्दी: Auto = हर पेज का असली आकार खुद पहचाने (मिली-जुली/ID/आधा पेज भी पूरा)। A4/Letter/Legal/A5 = तय आकार। Custom = लंबी पर्ची (लंबाई Settings में)।\nEnglish: Auto = detect each page's real size; A4/Letter/Legal/A5 = fixed; Custom = long receipts."), self.cmb_psize)
        btns = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        btns.accepted.connect(self._ok); btns.rejected.connect(self.reject)
        form.addRow(btns)

    def _main_window(self):
        w = self.parent()
        while w is not None and not hasattr(w, "pick_scanner_dialog"):
            w = w.parent()
        return w

    def _choose_device(self):
        """Sabhi scanner (LAN + USB + TWAIN) KHUD dhoondh kar list dikhao —
        koi TWAIN-error nahi. Chuna hua device is profile me lag jayega."""
        win = self._main_window()
        if win is None:
            QtWidgets.QMessageBox.warning(self, "Error", "Cannot detect scanners here."); return
        win.pick_scanner_dialog(on_done=self._device_picked)

    def _device_picked(self, name, kind, value):
        self.profile["source_name"] = name
        self.device_label.setText(name)

    def _ok(self):
        if not self.name_edit.text().strip():
            QtWidgets.QMessageBox.warning(self, "Error", "Enter a Display Name."); return
        self.accept()

    def get_profile(self):
        self.profile["name"] = self.name_edit.text().strip() or "Profile"
        try:
            self.profile["dpi"] = max(50, min(1200, int(self.cmb_dpi.currentText().strip())))
        except Exception:
            self.profile["dpi"] = 200
        self.profile["color"] = COLOUR_MODES[self.cmb_color.currentText()]
        self.profile["duplex"] = self.chk_duplex.isChecked()
        self.profile["page_size"] = self._PSIZES[self.cmb_psize.currentIndex()][1]
        return self.profile


class ProfileManagerDialog(QtWidgets.QDialog):
    def __init__(self, parent, profiles):
        super().__init__(parent)
        self.setWindowTitle("Profiles"); self.resize(420, 320)
        self.profiles = [dict(p) for p in profiles]
        lay = QtWidgets.QVBoxLayout(self)
        lay.addWidget(QtWidgets.QLabel("Scan profiles:"))
        self.listw = QtWidgets.QListWidget()
        self.listw.itemDoubleClicked.connect(lambda *_: self._edit())
        lay.addWidget(self.listw, 1); self._refresh()
        row = QtWidgets.QHBoxLayout()
        for t, s in [("New", self._new), ("Edit", self._edit), ("Delete", self._delete)]:
            b = QtWidgets.QPushButton(t); b.clicked.connect(s); row.addWidget(b)
        row.addStretch(1)
        done = QtWidgets.QPushButton("Done"); done.setObjectName("primary")
        done.clicked.connect(self.accept); row.addWidget(done)
        lay.addLayout(row)

    def _refresh(self):
        self.listw.clear()
        for p in self.profiles:
            self.listw.addItem(p.get("name", "Profile"))

    def _new(self):
        dlg = EditProfileDialog(self)
        if dlg.exec_() == QtWidgets.QDialog.Accepted:
            self.profiles.append(dlg.get_profile()); self._refresh()
            self.listw.setCurrentRow(len(self.profiles) - 1)

    def _edit(self):
        i = self.listw.currentRow()
        if i < 0:
            return
        dlg = EditProfileDialog(self, self.profiles[i])
        if dlg.exec_() == QtWidgets.QDialog.Accepted:
            self.profiles[i] = dlg.get_profile(); self._refresh(); self.listw.setCurrentRow(i)

    def _delete(self):
        i = self.listw.currentRow()
        if i < 0:
            return
        del self.profiles[i]; self._refresh()


class OptionsDialog(QtWidgets.QDialog):
    def __init__(self, parent, opts):
        super().__init__(parent)
        self.setWindowTitle("Options / Settings")
        self.setMinimumWidth(560)
        self.opts = dict(DEFAULT_OPTIONS); self.opts.update(opts or {})

        outer = QtWidgets.QVBoxLayout(self)
        scroll = QtWidgets.QScrollArea(); scroll.setWidgetResizable(True)
        inner = QtWidgets.QWidget(); form = QtWidgets.QFormLayout(inner)
        scroll.setWidget(inner); outer.addWidget(scroll, 1)

        def header(t):
            lbl = QtWidgets.QLabel("<b>%s</b>" % t); form.addRow(lbl)

        def qhelp(tip):
            h = QtWidgets.QLabel("?")
            h.setToolTip(tip); h.setCursor(QtCore.Qt.WhatsThisCursor)
            h.setStyleSheet("QLabel{color:#0f766e; border:1px solid #0f766e; border-radius:9px;"
                            "min-width:18px; max-width:18px; min-height:18px; max-height:18px;"
                            "font-weight:bold; qproperty-alignment:AlignCenter;}")
            return h

        def chkrow(chk, tip):
            chk.setToolTip(tip)
            row = QtWidgets.QHBoxLayout(); row.setContentsMargins(0, 0, 0, 0); row.setSpacing(6)
            row.addWidget(chk); row.addWidget(qhelp(tip)); row.addStretch(1)
            w = QtWidgets.QWidget(); w.setLayout(row); return w

        def lblhelp(text, tip):
            row = QtWidgets.QHBoxLayout(); row.setContentsMargins(0, 0, 0, 0); row.setSpacing(6)
            row.addWidget(QtWidgets.QLabel(text)); row.addWidget(qhelp(tip)); row.addStretch(1)
            w = QtWidgets.QWidget(); w.setLayout(row); return w

        header("Interface")
        self.cmb_theme = QtWidgets.QComboBox(); self.cmb_theme.addItems(["Light", "Dark", "Dark Pro"])
        self.cmb_theme.setCurrentIndex({"light": 0, "dark": 1, "darkpro": 2}.get(self.opts.get("theme"), 0))
        form.addRow(lblhelp("Theme:", 'हिन्दी: ऐप का रंग-रूप। Dark = काली थीम (आँखों को आराम, रात में अच्छी)। Light = सफ़ेद।\nEnglish: App look. Dark = dark theme (easy on eyes at night). Light = white.'), self.cmb_theme)
        self.chk_pagenum = QtWidgets.QCheckBox("Show page numbers under thumbnails")
        self.chk_pagenum.setChecked(self.opts.get("show_page_numbers", True)); form.addRow(chkrow(self.chk_pagenum, "हिन्दी: हर थंबनेल के नीचे 'Page 1, 2…' दिखाना। बंद करने पर नंबर नहीं दिखेंगे।\nEnglish: Show 'Page 1, 2..' under each thumbnail. Off = no numbers."))
        self.chk_autoname = QtWidgets.QCheckBox("Auto-read the document name (OCR)")
        self.chk_autoname.setChecked(self.opts.get("auto_name", False))
        form.addRow(chkrow(self.chk_autoname, "हिन्दी: चालू करने पर: स्कैन के बाद हर पेज का नाम (document का शीर्षक, जैसे DISCHARGE SUMMARY / RECEIPT) खुद पढ़कर लिख देगा — 'Page 1,2' के बजाय। (OCR/Tesseract चाहिए।)\nEnglish: ON: after scanning, auto-labels each page with its document title (needs OCR)."))
        self.chk_name_num = QtWidgets.QCheckBox("Also add a document number to the name")
        self.chk_name_num.setChecked(self.opts.get("name_append_number", False))
        form.addRow(chkrow(self.chk_name_num, "हिन्दी: चालू करने पर: अपने-आप बने नाम में bill/invoice/claim नंबर भी जुड़ जाएगा (जैसे Bill_INV1234)। नंबर पेज के टेक्स्ट से पढ़ा जाता है।\nEnglish: ON: appends the detected bill/invoice/claim number to the auto name (e.g. Bill_INV1234)."))
        self.chk_name_date = QtWidgets.QCheckBox("Also add today's date to the name")
        self.chk_name_date.setChecked(self.opts.get("name_append_date", False))
        form.addRow(chkrow(self.chk_name_date, "हिन्दी: चालू करने पर: अपने-आप बने नाम के आगे आज की तारीख़ लग जाएगी (जैसे Bill_2026-07-21)।\nEnglish: ON: appends today's date to the auto name (e.g. Bill_2026-07-21)."))

        header("Save")
        self.chk_autosave = QtWidgets.QCheckBox("Save PDF automatically after scanning")
        self.chk_autosave.setChecked(self.opts["auto_save"]); form.addRow(chkrow(self.chk_autosave, 'हिन्दी: चालू करने पर: स्कैन ख़त्म होते ही PDF अपने-आप सेव हो जाएगी (हर बार Save दबाना नहीं पड़ेगा)। रोज़ बहुत स्कैन करते हों तो चालू रखें।\nEnglish: ON: PDF saves automatically after each scan. Handy if you scan a lot.'))
        fr = QtWidgets.QHBoxLayout()
        self.folder_edit = QtWidgets.QLineEdit(self.opts["save_folder"])
        b = QtWidgets.QPushButton("…"); b.setFixedWidth(36); b.clicked.connect(self._pick_folder)
        fr.addWidget(self.folder_edit, 1); fr.addWidget(b)
        fw = QtWidgets.QWidget(); fw.setLayout(fr); form.addRow(lblhelp("Save folder:", 'हिन्दी: PDF कहाँ सेव हों वह फ़ोल्डर।\nEnglish: Folder where PDFs get saved.'), fw)
        self.tmpl_edit = QtWidgets.QLineEdit(self.opts["filename_template"])
        form.addRow(lblhelp("Filename template:", 'हिन्दी: फ़ाइल का नाम कैसे बने। {claim}=claim नंबर, {date}=तारीख़, {time}=समय, {seq}=क्रम संख्या।\nEnglish: How filenames are built: {claim} {date} {time} {seq}.'), self.tmpl_edit)
        form.addRow("", QtWidgets.QLabel("Tags: {claim} {date} {time} {seq}"))
        self.chk_claimfolder = QtWidgets.QCheckBox("Separate folder per claim number")
        self.chk_claimfolder.setChecked(self.opts["make_claim_folder"]); form.addRow(chkrow(self.chk_claimfolder, 'हिन्दी: चालू करने पर: हर claim नंबर का अलग फ़ोल्डर बनेगा (एक claim के सारे पेज एक जगह)।\nEnglish: ON: a separate folder per claim number.'))
        self.chk_ymfolder = QtWidgets.QCheckBox("Year/Month folders (2026/07/...)")
        self.chk_ymfolder.setChecked(self.opts["year_month_folders"]); form.addRow(chkrow(self.chk_ymfolder, 'हिन्दी: चालू करने पर: साल/महीने के फ़ोल्डर (2026/07/…) — पुराने स्कैन आसानी से मिलें।\nEnglish: ON: year/month folders (2026/07/...) for easy filing.'))
        self.cmb_after = QtWidgets.QComboBox(); self.cmb_after.addItems(["Do nothing", "Open the PDF", "Open the folder"])
        self.cmb_after.setCurrentIndex({"nothing": 0, "open": 1, "folder": 2}.get(self.opts.get("after_save", "nothing"), 0))
        form.addRow(lblhelp("After save:", 'हिन्दी: सेव के बाद क्या हो — कुछ नहीं / PDF खुले / फ़ोल्डर खुले।\nEnglish: After save — do nothing / open the PDF / open the folder.'), self.cmb_after)
        self.chk_imgtoo = QtWidgets.QCheckBox("Also save a separate image (JPG) for each page")
        self.chk_imgtoo.setChecked(self.opts.get("save_images_too", False)); form.addRow(chkrow(self.chk_imgtoo, 'हिन्दी: चालू करने पर: PDF के साथ हर पेज की अलग JPG इमेज भी बनेगी।\nEnglish: ON: also save each page as a separate JPG image.'))
        self.chk_searchable = QtWidgets.QCheckBox("Searchable PDF (find text with Ctrl+F)")
        self.chk_searchable.setChecked(bool(self.opts.get("searchable_pdf")))
        if not HAS_OCR_LIBS:
            self.chk_searchable.setEnabled(False)
            self.chk_searchable.setText("Searchable PDF (needs OCR libraries)")
        form.addRow(chkrow(self.chk_searchable, 'हिन्दी: चालू करने पर: हर सेव पर PDF के अंदर छुपा हुआ OCR टेक्स्ट भी सेव होगा — बाद में किसी भी PDF रीडर में Ctrl+F से मरीज़ का नाम/claim नंबर ढूँढ सकेंगे। (Tesseract चाहिए; सेव थोड़ा धीमा होगा।)\nEnglish: ON: every save embeds a hidden OCR text layer so you can later find a patient name/claim number with Ctrl+F in any PDF reader. (Needs Tesseract; save is a little slower.)'))

        header("Image cleanup")
        self.chk_blank = QtWidgets.QCheckBox("Remove blank pages")
        self.chk_blank.setChecked(self.opts["remove_blank"]); form.addRow(chkrow(self.chk_blank, 'हिन्दी: चालू करने पर: खाली (blank) पेज अपने-आप हट जाएँगे — जैसे duplex में पीछे का खाली हिस्सा। (NAPS2 जैसा।)\nEnglish: ON: blank pages are removed automatically (e.g. blank back side in duplex).'))
        self.cmb_blank_sens = QtWidgets.QComboBox(); self.cmb_blank_sens.addItems(["Kam (safe)", "Normal", 'High (aggressive)'])
        self.cmb_blank_sens.setCurrentIndex({"kam": 0, "normal": 1, "zyada": 2}.get(self.opts.get("blank_sensitivity", "normal"), 1))
        form.addRow(lblhelp("Blank hatane ki sensitivity:", 'हिन्दी: कितनी सख़्ती से खाली पेज हटाए। "ज़्यादा" = मोड़ की लकीर/हल्के स्टांप वाले पीछे के खाली पेज भी हट जाएँगे (पर कभी-कभी कम सामग्री वाला असली पेज भी हट सकता है)। "कम" = सिर्फ़ बिल्कुल खाली। Normal बीच का।\nEnglish: How aggressively to drop blanks. Zyada = also removes back sides with fold lines/faint marks; Kam = only truly empty; Normal = balanced.'), self.cmb_blank_sens)
        self.chk_crop = QtWidgets.QCheckBox("Border auto-crop")
        self.chk_crop.setChecked(self.opts["auto_crop"]); form.addRow(chkrow(self.chk_crop, 'हिन्दी: चालू करने पर: पेज के आस-पास की खाली सफ़ेद border अपने-आप कट जाएगी।\nEnglish: ON: auto-trims the empty white border around the page.'))
        self.chk_deskew = QtWidgets.QCheckBox("Auto-deskew (straighten)")
        self.chk_deskew.setChecked(self.opts["deskew"])
        if not HAS_NUMPY:
            self.chk_deskew.setEnabled(False)
            self.chk_deskew.setText("Auto-deskew (install numpy)")
        form.addRow(chkrow(self.chk_deskew, 'हिन्दी: चालू करने पर: टेढ़ा स्कैन हुआ पेज अपने-आप सीधा हो जाएगा।\nEnglish: ON: straightens a tilted/skewed page automatically.'))
        self.chk_enhance = QtWidgets.QCheckBox("Auto quality improvement (clean faded documents)")
        self.chk_enhance.setChecked(self.opts["quality_enhance"]); form.addRow(chkrow(self.chk_enhance, 'हिन्दी: चालू करने पर: फीके/हल्के document साफ़ और गहरे दिखेंगे।\nEnglish: ON: brightens & sharpens faded documents.'))
        self.chk_clean_edges = QtWidgets.QCheckBox("Clean scan edges (black border / punch-holes)")
        self.chk_clean_edges.setChecked(bool(self.opts.get("clean_edges"))); form.addRow(chkrow(self.chk_clean_edges, 'हिन्दी: चालू करने पर: स्कैन के किनारों की काली border और किनारे के छेद (punch-hole) के निशान अपने-आप सफ़ेद हो जाएँगे। बीच का टेक्स्ट/स्टांप नहीं छुआ जाता।\nEnglish: ON: whitens the black scan border and edge punch-hole marks. The middle content is never touched.'))
        self.chk_split2 = QtWidgets.QCheckBox("Split two pages on one glass into two")
        self.chk_split2.setChecked(bool(self.opts.get("split_two_page"))); form.addRow(chkrow(self.chk_split2, 'हिन्दी: चालू करने पर: एक glass पर दो छोटे पेज एक साथ रखें — स्कैन होते ही वे दो अलग पेज बन जाएँगे (बीच की सफ़ेद पट्टी से काटकर)।\nEnglish: ON: put two small pages together on the glass — the scan is auto-split into two separate pages at the white gutter.'))
        self.chk_orient = QtWidgets.QCheckBox("Auto-rotate upside-down pages (via OCR)")
        self.chk_orient.setChecked(bool(self.opts.get("auto_orient")))
        form.addRow(chkrow(self.chk_orient, 'हिन्दी: चालू करने पर: उल्टा (90/180/270) स्कैन हुआ पेज OCR से पहचानकर अपने-आप सीधा हो जाएगा (Tesseract चाहिए; स्कैन थोड़ा धीमा होता है)।\nEnglish: ON: auto-rotates upside-down/sideways pages using OCR (needs Tesseract; slightly slower).'))
        self.chk_autocolour = QtWidgets.QCheckBox("Auto colour-detect (colourless pages to gray)")
        self.chk_autocolour.setChecked(bool(self.opts.get("auto_colour")))
        form.addRow(chkrow(self.chk_autocolour, 'हिन्दी: चालू करने पर (सिर्फ़ Colour स्कैन में): जिस पेज पर रंग नहीं है वह अपने-आप ग्रे में सेव होगा — छोटी फ़ाइल, साफ़ प्रिंट।\nEnglish: ON (colour scans only): pages with no real colour are saved as grayscale — smaller files.'))
        self.spin_custlen = QtWidgets.QSpinBox(); self.spin_custlen.setRange(100, 3000)
        self.spin_custlen.setValue(int(self.opts.get("custom_page_mm", 600) or 600))
        self.spin_custlen.setSuffix(" mm")
        form.addRow(lblhelp("Custom page ki lambai:", 'हिन्दी: प्रोफ़ाइल में पेज साइज़ "लंबी पर्ची/custom" चुनने पर इतनी लंबी पट्टी स्कैन होगी (receipt/बही-खाता)।\nEnglish: Length used when the profile page size is "custom" (long receipts).'), self.spin_custlen)
        self.chk_filexfer = QtWidgets.QCheckBox("TWAIN continuous feed (experimental)")
        self.chk_filexfer.setChecked(bool(self.opts.get("twain_file_xfer")))
        form.addRow(chkrow(self.chk_filexfer, 'हिन्दी: चालू करने पर (सिर्फ़ TWAIN तरीके में): स्कैनर का feeder बिना रुके चलता रहेगा (file-transfer mode)। अगर स्कैनर समर्थन न करे तो ऐप खुद पुराने तरीके पर आ जाएगी।\nEnglish: ON (TWAIN method only): keeps the ADF feeding continuously via file-transfer mode; falls back to the normal path automatically if unsupported.'))

        header("Output")
        self.chk_compress = QtWidgets.QCheckBox("PDF compress (smaller file)")
        self.chk_compress.setChecked(self.opts["compress"]); form.addRow(chkrow(self.chk_compress, 'हिन्दी: चालू करने पर: PDF की फ़ाइल छोटी बनेगी (email/upload आसान); quality थोड़ी कम।\nEnglish: ON: smaller PDF (easier to email/upload); slightly lower quality.'))
        self.q_spin = QtWidgets.QSpinBox(); self.q_spin.setRange(20, 95)
        self.q_spin.setValue(int(self.opts["jpeg_quality"]))
        form.addRow(lblhelp("Compress quality:", 'हिन्दी: Compress की quality (ज़्यादा = बेहतर तस्वीर पर बड़ी फ़ाइल)। 60-80 ठीक है।\nEnglish: Compression quality (higher = better image, larger file). 60-80 is fine.'), self.q_spin)
        self.chk_wm = QtWidgets.QCheckBox("Watermark/stamp on every page")
        self.chk_wm.setChecked(self.opts["watermark"]); form.addRow(chkrow(self.chk_wm, 'हिन्दी: चालू करने पर: हर पेज पर आपका टेक्स्ट/स्टांप (जैसे अस्पताल का नाम) छपेगा।\nEnglish: ON: stamps your text (e.g. hospital name) on every page.'))
        self.wm_edit = QtWidgets.QLineEdit(self.opts["watermark_text"])
        form.addRow(lblhelp("Watermark text:", 'हिन्दी: Watermark में क्या लिखा हो।\nEnglish: The watermark text.'), self.wm_edit)

        header("Workflow")
        self.chk_batch = QtWidgets.QCheckBox("Batch mode (ready for next claim after save)")
        self.chk_batch.setChecked(self.opts["batch_mode"]); form.addRow(chkrow(self.chk_batch, 'हिन्दी: चालू करने पर: एक claim सेव होते ही अगला claim स्कैन करने के लिए स्क्रीन साफ़ हो जाएगी (तेज़ी से एक के बाद एक)।\nEnglish: ON: after saving, screen clears for the next claim (fast back-to-back).'))
        self.chk_validate = QtWidgets.QCheckBox("Validate claim number")
        self.chk_validate.setChecked(self.opts["validate_claim"]); form.addRow(chkrow(self.chk_validate, 'हिन्दी: चालू करने पर: ग़लत/अधूरा claim नंबर डालने पर चेतावनी देगा।\nEnglish: ON: warns if the claim number looks wrong/incomplete.'))
        self.pat_edit = QtWidgets.QLineEdit(self.opts["claim_pattern"])
        form.addRow(lblhelp("Claim pattern (regex):", 'हिन्दी: claim नंबर का सही रूप (regex)। आम तौर पर बदलने की ज़रूरत नहीं।\nEnglish: The valid claim-number pattern (regex). Usually leave as is.'), self.pat_edit)
        self.chk_barcode = QtWidgets.QCheckBox("Auto-fill claim number from Barcode/QR")
        self.chk_barcode.setChecked(self.opts["barcode_autofill"])
        if not HAS_ZBAR:
            self.chk_barcode.setEnabled(False)
            self.chk_barcode.setText("Barcode/QR (install pyzbar)")
        form.addRow(chkrow(self.chk_barcode, 'हिन्दी: चालू करने पर: पेज पर barcode/QR हो तो claim नंबर अपने-आप भर जाएगा।\nEnglish: ON: auto-fills the claim number from a barcode/QR on the page.'))
        self.chk_dup = QtWidgets.QCheckBox("Warn on duplicate claim number")
        self.chk_dup.setChecked(self.opts["duplicate_check"]); form.addRow(chkrow(self.chk_dup, 'हिन्दी: चालू करने पर: वही claim नंबर दोबारा हो तो चेतावनी (दोहरी एंट्री से बचाव)।\nEnglish: ON: warns on a duplicate claim number.'))

        header("Records & safety")
        self.chk_excel = QtWidgets.QCheckBox("Entry in Excel register (register.xlsx)")
        self.chk_excel.setChecked(self.opts["excel_log"])
        if not HAS_XLSX:
            self.chk_excel.setEnabled(False)
            self.chk_excel.setText("Excel register (install openpyxl)")
        form.addRow(chkrow(self.chk_excel, 'हिन्दी: चालू करने पर: हर स्कैन की एंट्री एक Excel रजिस्टर (register.xlsx) में जुड़ेगी — रिकॉर्ड के लिए।\nEnglish: ON: logs each scan into an Excel register (register.xlsx).'))
        self.chk_log = QtWidgets.QCheckBox("Activity log (activity_log.txt)")
        self.chk_log.setChecked(self.opts["activity_log"]); form.addRow(chkrow(self.chk_log, 'हिन्दी: चालू करने पर: कब क्या स्कैन/सेव हुआ इसका टेक्स्ट लॉग रखेगा।\nEnglish: ON: keeps a text activity log.'))
        self.chk_backup = QtWidgets.QCheckBox("Backup copy of every save")
        self.chk_backup.setChecked(self.opts["backup"]); form.addRow(chkrow(self.chk_backup, 'हिन्दी: चालू करने पर: हर सेव की एक बैकअप कॉपी अलग फ़ोल्डर में भी रखेगा (सुरक्षा)।\nEnglish: ON: keeps a backup copy of every save in another folder.'))
        br = QtWidgets.QHBoxLayout()
        self.backup_edit = QtWidgets.QLineEdit(self.opts["backup_folder"])
        bb = QtWidgets.QPushButton("…"); bb.setFixedWidth(36); bb.clicked.connect(self._pick_backup)
        br.addWidget(self.backup_edit, 1); br.addWidget(bb)
        bw = QtWidgets.QWidget(); bw.setLayout(br); form.addRow(lblhelp("Backup folder:", 'हिन्दी: बैकअप कहाँ रखे वह फ़ोल्डर।\nEnglish: Folder for backups.'), bw)

        btns = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept); btns.rejected.connect(self.reject)
        outer.addWidget(btns)

    def _pick_folder(self):
        d = QtWidgets.QFileDialog.getExistingDirectory(self, "Folder", self.folder_edit.text())
        if d:
            self.folder_edit.setText(d)

    def _pick_backup(self):
        d = QtWidgets.QFileDialog.getExistingDirectory(self, "Backup folder", self.backup_edit.text())
        if d:
            self.backup_edit.setText(d)

    def get_opts(self):
        o = self.opts
        o["theme"] = ["light", "dark", "darkpro"][self.cmb_theme.currentIndex()]
        o["show_page_numbers"] = self.chk_pagenum.isChecked()
        o["auto_name"] = self.chk_autoname.isChecked()
        o["name_append_number"] = self.chk_name_num.isChecked()
        o["name_append_date"] = self.chk_name_date.isChecked()
        o["save_images_too"] = self.chk_imgtoo.isChecked()
        o["searchable_pdf"] = self.chk_searchable.isChecked()
        o["auto_save"] = self.chk_autosave.isChecked()
        o["save_folder"] = self.folder_edit.text().strip()
        o["filename_template"] = self.tmpl_edit.text().strip() or "{date}_{seq}"
        o["make_claim_folder"] = self.chk_claimfolder.isChecked()
        o["after_save"] = ["nothing", "open", "folder"][self.cmb_after.currentIndex()]
        o["year_month_folders"] = self.chk_ymfolder.isChecked()
        o["remove_blank"] = self.chk_blank.isChecked()
        o["blank_sensitivity"] = {0: "kam", 1: "normal", 2: "zyada"}.get(self.cmb_blank_sens.currentIndex(), "normal")
        o["auto_crop"] = self.chk_crop.isChecked()
        o["deskew"] = self.chk_deskew.isChecked()
        o["quality_enhance"] = self.chk_enhance.isChecked()
        o["clean_edges"] = self.chk_clean_edges.isChecked()
        o["split_two_page"] = self.chk_split2.isChecked()
        o["twain_file_xfer"] = self.chk_filexfer.isChecked()
        o["auto_orient"] = self.chk_orient.isChecked()
        o["auto_colour"] = self.chk_autocolour.isChecked()
        o["custom_page_mm"] = int(self.spin_custlen.value())
        o["compress"] = self.chk_compress.isChecked()
        o["jpeg_quality"] = int(self.q_spin.value())
        o["watermark"] = self.chk_wm.isChecked()
        o["watermark_text"] = self.wm_edit.text().strip() or "Noble Care Hospital"
        o["batch_mode"] = self.chk_batch.isChecked()
        o["validate_claim"] = self.chk_validate.isChecked()
        o["claim_pattern"] = self.pat_edit.text().strip() or r"^[A-Za-z0-9\-]{4,}$"
        o["barcode_autofill"] = self.chk_barcode.isChecked()
        o["duplicate_check"] = self.chk_dup.isChecked()
        o["excel_log"] = self.chk_excel.isChecked()
        o["activity_log"] = self.chk_log.isChecked()
        o["backup"] = self.chk_backup.isChecked()
        o["backup_folder"] = self.backup_edit.text().strip()
        return o


# ---------------------------------------------------------------------------
# Main Window
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Setup wizard + Help (public-friendly)
# ---------------------------------------------------------------------------

class ScanProgressDialog(QtWidgets.QDialog):
    """NAPS2-style scanning box: page counter + Run in Background + Cancel."""
    cancelled = QtCore.pyqtSignal()

    def __init__(self, parent, title, lang="hi"):
        super().__init__(parent)
        self._lang = lang
        self.setWindowTitle(title or "Scanning")
        self.setModal(False)
        self.setMinimumWidth(430)
        lay = QtWidgets.QVBoxLayout(self)
        self.lbl = QtWidgets.QLabel(self._page_text(1))
        lay.addWidget(self.lbl)
        self.bar = QtWidgets.QProgressBar(); self.bar.setRange(0, 0)
        lay.addWidget(self.bar)
        row = QtWidgets.QHBoxLayout(); row.addStretch(1)
        self.btn_bg = QtWidgets.QPushButton("Run in Background" if lang == "en" else "Run in Background")
        self.btn_bg.clicked.connect(self.hide)
        self.btn_cancel = QtWidgets.QPushButton("Cancel")
        self.btn_cancel.clicked.connect(lambda: self.cancelled.emit())
        row.addWidget(self.btn_bg); row.addWidget(self.btn_cancel)
        lay.addLayout(row)
        self.show()

    def _page_text(self, n):
        return ("Scanning page %d..." % n) if self._lang == "en" else ('Scanning page %d...' % n)

    def set_page(self, n):
        self.lbl.setText(self._page_text(n))

    def closeEvent(self, e):
        e.accept()


class SetupWizard(QtWidgets.QWizard):
    def __init__(self, parent, lang="hi"):
        super().__init__(parent)
        self.lang = lang
        self.result_method = "twain"
        self.result_device_name = None
        self.result_wia_id = None
        self.setWindowTitle("Setup — ApneScan")
        self.resize(580, 440)
        self.addPage(self._welcome())
        self.addPage(self._device_page())
        self.addPage(self._finish())

    def _welcome(self):
        p = QtWidgets.QWizardPage()
        p.setTitle("Swagat hai" if self.lang == "hi" else "Welcome")
        lay = QtWidgets.QVBoxLayout(p)
        txt = ('This short setup will connect your scanner.\nClick Next.'
               if self.lang == "hi"
               else "This quick setup will connect your scanner.\nClick Next.")
        lbl = QtWidgets.QLabel(txt); lbl.setWordWrap(True); lay.addWidget(lbl)
        return p

    def _device_page(self):
        p = QtWidgets.QWizardPage()
        p.setTitle('Choose scanner' if self.lang == "hi" else "Choose scanner")
        lay = QtWidgets.QVBoxLayout(p)
        self.rb_twain = QtWidgets.QRadioButton("TWAIN (most scanners)")
        self.rb_wia = QtWidgets.QRadioButton("WIA (Windows built-in)")
        self.rb_twain.setChecked(True)
        lay.addWidget(self.rb_twain); lay.addWidget(self.rb_wia)
        row = QtWidgets.QHBoxLayout()
        self.dev_lbl = QtWidgets.QLabel("(no scanner chosen)")
        btn = QtWidgets.QPushButton("Choose scanner"); btn.clicked.connect(self._choose)
        row.addWidget(self.dev_lbl, 1); row.addWidget(btn)
        w = QtWidgets.QWidget(); w.setLayout(row); lay.addWidget(w)
        hint = QtWidgets.QLabel(
            "If your scanner isn't listed, try the other option (TWAIN/WIA)."
            if self.lang == "hi"
            else "If your scanner isn't listed, try the other option (TWAIN/WIA).")
        hint.setWordWrap(True); lay.addWidget(hint)
        return p

    def _choose(self):
        if self.rb_wia.isChecked():
            self.result_method = "wia"
            try:
                devs = list_wia_sources()
            except Exception as exc:
                QtWidgets.QMessageBox.warning(self, "Error", friendly_error(exc, self.lang)); return
            if not devs:
                QtWidgets.QMessageBox.warning(self, "Error", "No WIA scanner found."); return
            names = [n for _i, n in devs]
            name, ok = QtWidgets.QInputDialog.getItem(self, "Scanner", "Scanner:", names, 0, False)
            if ok and name:
                for _id, n in devs:
                    if n == name:
                        self.result_wia_id = _id; self.result_device_name = n; break
                self.dev_lbl.setText(name)
        else:
            self.result_method = "twain"
            try:
                names = list_sources(int(self.winId()))
            except Exception as exc:
                QtWidgets.QMessageBox.warning(self, "Error", friendly_error(exc, self.lang)); return
            if not names:
                QtWidgets.QMessageBox.warning(self, "Error", "No TWAIN scanner found."); return
            name, ok = QtWidgets.QInputDialog.getItem(self, "Scanner", "Scanner:", names, 0, False)
            if ok and name:
                self.result_device_name = name; self.dev_lbl.setText(name)

    def _finish(self):
        p = QtWidgets.QWizardPage()
        p.setTitle('Done!' if self.lang == "hi" else "Done!")
        lay = QtWidgets.QVBoxLayout(p)
        lbl = QtWidgets.QLabel(
            "Click Finish, then press 'Scan'. That's it!"
            if self.lang == "hi" else "Click Finish, then press 'Scan'. That's it!")
        lbl.setWordWrap(True); lay.addWidget(lbl)
        return p


class HelpDialog(QtWidgets.QDialog):
    def __init__(self, parent, lang="hi"):
        super().__init__(parent)
        self.setWindowTitle("Guide / Help"); self.resize(620, 520)
        lay = QtWidgets.QVBoxLayout(self)
        br = QtWidgets.QTextBrowser()
        br.setHtml(HELP_TEXT_HI if lang == "hi" else HELP_TEXT_EN)
        lay.addWidget(br)
        b = QtWidgets.QPushButton("Close" if lang == "hi" else "Close")
        b.clicked.connect(self.accept); lay.addWidget(b)



def _make_icon(kind, color="#0f766e"):
    """Draw a small flat icon (28x28) for the toolbar. No external files needed."""
    from PyQt5 import QtGui, QtCore
    pm = QtGui.QPixmap(30, 30); pm.fill(QtCore.Qt.transparent)
    p = QtGui.QPainter(pm); p.setRenderHint(QtGui.QPainter.Antialiasing)
    pen = QtGui.QPen(QtGui.QColor(color)); pen.setWidth(2); pen.setJoinStyle(QtCore.Qt.RoundJoin)
    pen.setCapStyle(QtCore.Qt.RoundCap); p.setPen(pen)
    br = QtGui.QColor(color)
    def doc():
        p.drawRoundedRect(8, 5, 14, 20, 2, 2)
        p.drawLine(11, 11, 19, 11); p.drawLine(11, 15, 19, 15); p.drawLine(11, 19, 16, 19)
    if kind == "scan":
        doc(); p.setBrush(QtGui.QColor("#5eead4")); p.drawRoundedRect(5, 13, 20, 4, 2, 2)
    elif kind == "profiles":
        p.drawRoundedRect(6, 7, 18, 5, 2, 2); p.drawRoundedRect(6, 14, 18, 5, 2, 2); p.drawRoundedRect(6, 21, 12, 5, 2, 2)
    elif kind == "fast":
        p.setBrush(QtGui.QColor("#f59e0b")); p.setPen(QtGui.QColor("#f59e0b"))
        p.drawPolygon(QtGui.QPolygon([QtCore.QPoint(17,4),QtCore.QPoint(9,17),QtCore.QPoint(14,17),QtCore.QPoint(12,26),QtCore.QPoint(21,12),QtCore.QPoint(15,12)]))
    elif kind == "ocr":
        f = p.font(); f.setBold(True); f.setPixelSize(11); p.setFont(f); p.drawText(pm.rect(), QtCore.Qt.AlignCenter, "OCR")
    elif kind == "import":
        doc(); p.drawLine(15, 26, 15, 19); p.drawLine(12, 22, 15, 19); p.drawLine(18, 22, 15, 19)
    elif kind == "savepdf":
        doc(); f=p.font(); f.setBold(True); f.setPixelSize(7); p.setFont(f); p.drawText(QtCore.QRect(8,17,14,8), QtCore.Qt.AlignCenter, "PDF")
    elif kind == "images":
        p.drawRoundedRect(6, 8, 18, 14, 2, 2); p.setBrush(br); p.drawEllipse(10, 12, 3, 3)
        p.drawLine(8, 20, 13, 15); p.drawLine(13, 20, 18, 14)
    elif kind == "print":
        p.drawRoundedRect(8, 5, 14, 7, 1, 1); p.drawRoundedRect(5, 12, 20, 9, 2, 2); p.drawRoundedRect(9, 19, 12, 7, 1, 1)
    elif kind == "rotate":
        p.drawArc(7, 7, 16, 16, 30*16, 280*16); p.drawLine(22, 8, 22, 13); p.drawLine(22, 13, 17, 13)
    elif kind == "up":
        p.drawLine(15, 22, 15, 9); p.drawLine(15, 9, 10, 14); p.drawLine(15, 9, 20, 14)
    elif kind == "down":
        p.drawLine(15, 8, 15, 21); p.drawLine(15, 21, 10, 16); p.drawLine(15, 21, 20, 16)
    elif kind == "delete":
        p.drawLine(9, 9, 21, 21); p.drawLine(21, 9, 9, 21)
    elif kind == "clear":
        p.drawRoundedRect(8, 8, 14, 16, 2, 2); p.drawLine(6, 8, 24, 8); p.drawLine(13, 5, 17, 5)
    elif kind == "language":
        p.drawEllipse(6, 6, 18, 18); p.drawLine(6, 15, 24, 15); p.drawArc(11, 6, 8, 18, 0, 360*16)
    elif kind == "about":
        p.drawEllipse(6, 6, 18, 18); f=p.font(); f.setBold(True); f.setPixelSize(13); p.setFont(f); p.drawText(pm.rect(), QtCore.Qt.AlignCenter, "i")
    elif kind == "guide":
        # document with a "?" — Complete Guide / help
        p.drawRoundedRect(7, 4, 16, 22, 2, 2)
        f = p.font(); f.setBold(True); f.setPixelSize(14); p.setFont(f)
        p.drawText(QtCore.QRect(7, 4, 16, 22), QtCore.Qt.AlignCenter, "?")
    else:
        doc()
    p.end()
    return QtGui.QIcon(pm)


class PreviewDialog(QtWidgets.QDialog):
    """NAPS2-style full preview: navigate pages, zoom, rotate, rename, delete,
    save — all from one window."""
    def __init__(self, win, row):
        super().__init__(win)
        self.win = win
        self.row = row
        self.zoom = 1.0
        self.fit = True
        self.setWindowTitle("Preview")
        self.resize(940, 900)
        v = QtWidgets.QVBoxLayout(self)
        tb = QtWidgets.QHBoxLayout()

        def b(txt, fn, tip=""):
            btn = QtWidgets.QToolButton()
            btn.setText(txt); btn.setToolTip(tip)
            btn.setToolButtonStyle(QtCore.Qt.ToolButtonTextOnly)
            btn.clicked.connect(fn)
            btn.setStyleSheet("QToolButton{padding:5px 9px; font-size:13px;}")
            tb.addWidget(btn)
            return btn

        b("\u25c0", self.prev, "Pichhla page (Left arrow)")
        self.lbl_count = QtWidgets.QLabel(); self.lbl_count.setStyleSheet("font-weight:bold;")
        tb.addWidget(self.lbl_count)
        b("\u25b6", self.next, "Agla page (Right arrow)")
        tb.addSpacing(14)
        b("\u2796", self.zoom_out, "Zoom out ( - )")
        b("Fit", self.fit_view, 'Fit whole page')
        b("\u2795", self.zoom_in, "Zoom in ( + )")
        b("100%", self.actual_size, "Asli size")
        tb.addSpacing(14)
        b("\u21ba", lambda: self.rotate(-90), 'Rotate left')
        b("\u21bb", lambda: self.rotate(90), 'Rotate right')
        b("\u2712 Rename", self.rename, 'Rename (F2)')
        b("\U0001f5d1 Delete", self.delete, 'Delete this page')
        b("\U0001f4be Save", self.save_one, 'Save only this page as PDF')
        tb.addStretch(1)
        b("\u2715 Close", self.accept, 'Close (Esc)')
        v.addLayout(tb)

        self.lbl = QtWidgets.QLabel(); self.lbl.setAlignment(QtCore.Qt.AlignCenter)
        self.scr = QtWidgets.QScrollArea(); self.scr.setWidgetResizable(False)
        self.scr.setAlignment(QtCore.Qt.AlignCenter); self.scr.setWidget(self.lbl)
        v.addWidget(self.scr, 1)
        QtCore.QTimer.singleShot(0, self._load)

    def _path(self):
        it = self.win.list.item(self.row)
        return it.data(QtCore.Qt.UserRole) if it else None

    def _load(self):
        n = self.win.list.count()
        if n == 0:
            self.accept(); return
        self.row = max(0, min(self.row, n - 1))
        it = self.win.list.item(self.row)
        title = it.data(TITLE_ROLE) or it.text() or ("Page %d" % (self.row + 1))
        self.lbl_count.setText("  %d / %d  \u2014 %s  " % (self.row + 1, n, title))
        self.win.list.setCurrentRow(self.row)
        pix = QtGui.QPixmap(self._path())
        if pix.isNull():
            self.lbl.clear(); return
        if self.fit:
            area = self.scr.viewport().size()
            pix = pix.scaled(max(50, area.width() - 6), max(50, area.height() - 6),
                             QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)
        else:
            pix = pix.scaled(int(pix.width() * self.zoom), int(pix.height() * self.zoom),
                             QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)
        self.lbl.setPixmap(pix); self.lbl.resize(pix.size())

    def prev(self):
        self.row -= 1; self.fit = True; self._load()

    def next(self):
        self.row += 1; self.fit = True; self._load()

    def zoom_in(self):
        self.fit = False; self.zoom = min(6.0, self.zoom * 1.25); self._load()

    def zoom_out(self):
        self.fit = False; self.zoom = max(0.1, self.zoom / 1.25); self._load()

    def actual_size(self):
        self.fit = False; self.zoom = 1.0; self._load()

    def fit_view(self):
        self.fit = True; self._load()

    def rotate(self, ang):
        p = self._path()
        try:
            with Image.open(p) as im:
                im.rotate(-ang, expand=True).save(p)
            self.win._refresh_item(self.win.list.item(self.row))
            self.win._dirty = True
        except Exception:
            pass
        self.fit = True; self._load()

    def rename(self):
        self.win.list.setCurrentRow(self.row)
        self.win.rename_current_page()
        self._load()

    def delete(self):
        it = self.win.list.item(self.row)
        if it is None:
            return
        self.win.list.clearSelection(); it.setSelected(True)
        self.win.list.setCurrentRow(self.row)
        self.win.delete_page()
        self._load()

    def save_one(self):
        p = self._path()
        if p:
            self.win.save_pdf([p])

    def keyPressEvent(self, e):
        k = e.key()
        if k == QtCore.Qt.Key_Left:
            self.prev()
        elif k == QtCore.Qt.Key_Right:
            self.next()
        elif k in (QtCore.Qt.Key_Plus, QtCore.Qt.Key_Equal):
            self.zoom_in()
        elif k == QtCore.Qt.Key_Minus:
            self.zoom_out()
        elif k == QtCore.Qt.Key_F2:
            self.rename()
        else:
            super().keyPressEvent(e)


class _EditorCanvas(QtWidgets.QLabel):
    """Editor ka canvas — mouse editor ko deta hai, overlay khud kheenchta hai."""
    def __init__(self, editor):
        super().__init__()
        self.ed = editor
        self.setMouseTracking(True)
        self.setStyleSheet("background:#ffffff;")

    def mousePressEvent(self, e):
        self.ed._canvas_press(e.pos())

    def mouseMoveEvent(self, e):
        self.ed._canvas_move(e.pos())

    def mouseReleaseEvent(self, e):
        self.ed._canvas_release(e.pos())

    def wheelEvent(self, e):
        if e.modifiers() & QtCore.Qt.ControlModifier:
            self.ed._wheel_zoom(1 if e.angleDelta().y() > 0 else -1)
            e.accept()
        else:
            super().wheelEvent(e)

    def paintEvent(self, e):
        super().paintEvent(e)
        try:
            self.ed._overlay(self)
        except Exception:
            pass


class ImageEditor(QtWidgets.QDialog):
    """Poora Document Editor (attractive) — left tool-rail, right adjust-panel,
    filmstrip, zoom/pan, auto-fix, sabhi tools + sliders. Har cheez auto BHI aur
    haath me BHI."""

    _FONTS = ("nirmala.ttf", "arial.ttf", "DejaVuSans.ttf", "Arial.ttf")
    _QSS = """
    QDialog { background:#eef2f6; }
    QLabel#hdr { color:#0f172a; font-weight:700; font-size:13px; }
    QFrame#rail { background:#ffffff; border-radius:12px; }
    QFrame#side { background:#ffffff; border-radius:12px; }
    QFrame#strip { background:#ffffff; border-radius:12px; }
    QLabel.grp { color:#64748b; font-size:10px; font-weight:700; letter-spacing:1px; }
    QToolButton.tool { border:none; border-radius:10px; padding:7px 3px; color:#334155;
                       font-size:10px; background:transparent; }
    QToolButton.tool:hover { background:#e6eef0; }
    QToolButton.tool:checked { background:#0f766e; color:#ffffff; }
    QToolButton.act { border:1px solid #e2e8f0; border-radius:9px; padding:5px 7px;
                      color:#334155; font-size:11px; background:#fff; }
    QToolButton.act:hover { border-color:#0f766e; color:#0f766e; background:#f0fdfa; }
    QPushButton#primary { background:#0f766e; color:#fff; border:none; border-radius:9px;
                          padding:8px 16px; font-weight:700; }
    QPushButton#primary:hover { background:#0d5f58; }
    QPushButton#ghost { background:#fff; color:#334155; border:1px solid #cbd5e1;
                        border-radius:9px; padding:8px 14px; }
    QPushButton#ghost:hover { border-color:#0f766e; color:#0f766e; }
    QPushButton#accent { background:#f59e0b; color:#1f2937; border:none; border-radius:9px;
                         padding:7px 14px; font-weight:700; }
    QPushButton#accent:hover { background:#d97706; color:#fff; }
    QSlider::groove:horizontal { height:5px; background:#e2e8f0; border-radius:3px; }
    QSlider::handle:horizontal { background:#0f766e; width:14px; margin:-5px 0; border-radius:7px; }
    QScrollArea#cv { background:#334155; border:none; border-radius:12px; }
    QListWidget { border:none; background:transparent; }
    """

    def __init__(self, win, path, on_saved=None, tool=None):
        super().__init__(win)
        self.win = win
        self.path = path
        self.on_saved = on_saved
        self._start_tool = tool
        L = win.L
        self.L = L
        self.setWindowTitle(L("🎨 Document editor", "🎨 Document editor"))
        self.resize(1180, 900)
        self.setStyleSheet(self._QSS)
        try:
            self.row = self.win.list.currentRow()
        except Exception:
            self.row = -1
        # state
        self.tool = None
        self.zoom = 1.0
        self.crop_ratio = None
        self.pen_color = (220, 20, 20)
        self.text_size = 40
        self.show_original = False
        self._start = self._cur = None
        self._erasing = self._penning = self._panning = False
        self._pen_last = None
        self._pan_ref = None
        self._corners = []
        self._disp_scale = 1.0
        self._preview_img = None
        self._rot_base = None
        self._dirty_any = False

        root = QtWidgets.QVBoxLayout(self)
        root.setContentsMargins(10, 8, 10, 8); root.setSpacing(8)

        # ---------- TOP HEADER ----------
        top = QtWidgets.QHBoxLayout(); top.setSpacing(6)
        self.lbl_hdr = QtWidgets.QLabel(""); self.lbl_hdr.setObjectName("hdr")
        top.addWidget(self.lbl_hdr)
        top.addStretch(1)
        bfix = QtWidgets.QPushButton("✨ " + L("Auto-Fix", "Auto-Fix")); bfix.setObjectName("accent")
        bfix.setToolTip(L("Seedha + crop + saaf + whiten — sab ek saath",
                          "Straighten + crop + enhance + whiten — all at once"))
        bfix.clicked.connect(self._auto_fix)
        bmagic = QtWidgets.QPushButton("🪄 " + L("Magic", "Magic")); bmagic.setObjectName("ghost")
        bmagic.setToolTip(L("Scanner jaisa: background pura safed, akshar gehre-saaf",
                            "Scanner-clean: pure white background, crisp dark text"))
        bmagic.clicked.connect(self._magic_color)
        bup = QtWidgets.QPushButton("🔤 " + L("Auto-seedha", "Auto-upright")); bup.setObjectName("ghost")
        bup.setToolTip(L("Text padhkar ulta/tirchha page seedha karo",
                         "Use OCR to turn a sideways page upright"))
        bup.clicked.connect(self._auto_upright)
        top.addWidget(bfix); top.addWidget(bmagic); top.addWidget(bup)
        top.addSpacing(10)
        for ic, tip, fn in (("➖", L("Zoom kam", "Zoom out"), lambda: self._set_zoom(self.zoom / 1.25)),
                            ("🔳", L("Fit", "Fit"), lambda: self._set_zoom(1.0)),
                            ("➕", L("Zoom zyada", "Zoom in"), lambda: self._set_zoom(self.zoom * 1.25)),
                            ("⛶", L("Poori screen", "Fullscreen"), self._toggle_full)):
            b = QtWidgets.QToolButton(); b.setText(ic); b.setToolTip(tip)
            b.setProperty("class", "act"); b.clicked.connect(fn); top.addWidget(b)
        self.btn_ba = QtWidgets.QToolButton(); self.btn_ba.setText("👁 " + L("Pehle", "Before"))
        self.btn_ba.setCheckable(True); self.btn_ba.setProperty("class", "act")
        self.btn_ba.setToolTip(L("Dabao — original dikhega (pehle/baad)", "Toggle to see the original"))
        self.btn_ba.toggled.connect(self._toggle_before)
        top.addWidget(self.btn_ba)
        root.addLayout(top)

        # ---------- MIDDLE: rail | canvas | side ----------
        mid = QtWidgets.QHBoxLayout(); mid.setSpacing(8)

        # left tool rail
        rail = QtWidgets.QFrame(); rail.setObjectName("rail"); rail.setFixedWidth(78)
        rl = QtWidgets.QVBoxLayout(rail); rl.setContentsMargins(6, 8, 6, 8); rl.setSpacing(2)
        self._tool_btns = {}
        for key, icon, name, tip in (
            ("pan", "✋", L("Ghumao", "Pan"), L("Khali haath — page khisko", "Drag to pan")),
            ("crop", "✂", L("Crop", "Crop"), L("Area chuno", "Select area")),
            ("erase", "🧽", L("Miṭao", "Erase"), L("Safed karo", "Paint white")),
            ("pen", "✏", L("Pen", "Pen"), L("Haath se", "Freehand")),
            ("text", "🔤", L("Text", "Text"), L("Likho", "Type")),
            ("line", "／", L("Line", "Line"), L("Seedhi line", "Straight line")),
            ("arrow", "➡", L("Teer", "Arrow"), L("Teer", "Arrow")),
            ("rect", "▭", L("Aayat", "Rect"), L("Aayat", "Rectangle")),
            ("circle", "◯", L("Gola", "Circle"), L("Gola", "Circle")),
            ("box", "🖍", L("Mark", "Mark"), L("Peela highlight", "Highlight")),
            ("redact", "🖤", L("Chhupao", "Redact"), L("Kaala box (jaankari sach me mit jaati hai)", "Black box (data truly erased)")),
            ("blur", "🌫", L("Dhundhla", "Blur"), L("Chehra/jaankari dhundhli", "Blur an area")),
            ("note", "🗒", L("Note", "Note"), L("Chipakne wala peela note", "Sticky note")),
            ("stamp", "✔", L("Mohar", "Stamp"), L("Tick / cross / star / date", "Tick / cross / star / date")),
            ("sign", "✍", L("Sign", "Sign"), L("Sign/mohar", "Signature/stamp")),
            ("persp", "📐", L("Kone", "Corners"), L("4 kone se seedha", "4-corner straighten")),
        ):
            b = QtWidgets.QToolButton(); b.setText(icon + "\n" + name)
            b.setToolTip(tip); b.setCheckable(True)
            b.setToolButtonStyle(QtCore.Qt.ToolButtonTextUnderIcon)
            b.setProperty("class", "tool"); b.setFixedWidth(64)
            b.clicked.connect(lambda _c, t=key: self._set_tool(t))
            rl.addWidget(b); self._tool_btns[key] = b
        rl.addStretch(1)
        mid.addWidget(rail)

        # canvas
        self.canvas = _EditorCanvas(self)
        self.canvas.setAlignment(QtCore.Qt.AlignCenter)
        self.scroll = QtWidgets.QScrollArea(); self.scroll.setObjectName("cv")
        self.scroll.setWidgetResizable(False); self.scroll.setAlignment(QtCore.Qt.AlignCenter)
        self.scroll.setWidget(self.canvas)
        mid.addWidget(self.scroll, 1)

        # right side panel (scrollable)
        side = QtWidgets.QFrame(); side.setObjectName("side"); side.setFixedWidth(268)
        sv = QtWidgets.QVBoxLayout(side); sv.setContentsMargins(10, 10, 10, 10); sv.setSpacing(6)

        def grp(txt):
            l = QtWidgets.QLabel(txt); l.setProperty("class", "grp"); sv.addWidget(l)

        # tool options (contextual)
        grp(L("TOOL", "TOOL"))
        self.opt_wrap = QtWidgets.QWidget(); ow = QtWidgets.QHBoxLayout(self.opt_wrap)
        ow.setContentsMargins(0, 0, 0, 0); ow.setSpacing(6)
        self.sl_brush = QtWidgets.QSlider(QtCore.Qt.Horizontal)
        self.sl_brush.setRange(3, 80); self.sl_brush.setValue(16); self.sl_brush.setFixedWidth(110)
        self.btn_col = QtWidgets.QToolButton(); self.btn_col.setText("🎨"); self.btn_col.setProperty("class", "act")
        self.btn_col.clicked.connect(self._pick_colour)
        self.cmb_ratio = QtWidgets.QComboBox()
        for _v, _t in (("free", L("Free", "Free")), ("a4", "A4"), ("id", "ID"), ("sq", "1:1")):
            self.cmb_ratio.addItem(_t, _v)
        self.cmb_ratio.currentIndexChanged.connect(self._ratio_changed)
        self.sp_tsize = QtWidgets.QSpinBox(); self.sp_tsize.setRange(10, 200); self.sp_tsize.setValue(40)
        self.sp_tsize.valueChanged.connect(lambda v: setattr(self, "text_size", v))
        self.cmb_stamp = QtWidgets.QComboBox()
        for _g, _t in (("✔", L("Tick ✔", "Tick ✔")), ("✘", L("Cross ✘", "Cross ✘")),
                       ("★", L("Star ★", "Star ★")), ("date", L("Aaj ki date", "Today's date")),
                       ("approved", L("APPROVED", "APPROVED")), ("paid", L("PAID", "PAID"))):
            self.cmb_stamp.addItem(_t, _g)
        for w in (QtWidgets.QLabel("○"), self.sl_brush, self.btn_col, self.cmb_ratio, self.cmb_stamp, self.sp_tsize):
            ow.addWidget(w)
        ow.addStretch(1)
        sv.addWidget(self.opt_wrap)

        # adjust sliders
        grp(L("SUDHAAR (haath me)", "ADJUST (live)"))
        self._sliders = {}
        for key, name in (("bright", L("Roshni", "Brightness")), ("contrast", "Contrast"),
                          ("sharp", L("Dhaar", "Sharpness")), ("satur", L("Rang", "Saturation")),
                          ("temp", L("Warm/Cool", "Temperature"))):
            r = QtWidgets.QHBoxLayout(); r.setSpacing(6)
            lb = QtWidgets.QLabel(name); lb.setFixedWidth(90); lb.setStyleSheet("font-size:11px;color:#475569;")
            s = QtWidgets.QSlider(QtCore.Qt.Horizontal); s.setRange(20, 200); s.setValue(100)
            s.valueChanged.connect(self._adj_changed)
            r.addWidget(lb); r.addWidget(s); self._sliders[key] = s
            w = QtWidgets.QWidget(); w.setLayout(r); sv.addWidget(w)
        self.sl_bright = self._sliders["bright"]; self.sl_contrast = self._sliders["contrast"]
        rr = QtWidgets.QHBoxLayout()
        lbt = QtWidgets.QLabel("B&W"); lbt.setFixedWidth(90); lbt.setStyleSheet("font-size:11px;color:#475569;")
        self.sl_bw = QtWidgets.QSlider(QtCore.Qt.Horizontal); self.sl_bw.setRange(60, 240); self.sl_bw.setValue(160)
        rr.addWidget(lbt); rr.addWidget(self.sl_bw)
        rw = QtWidgets.QWidget(); rw.setLayout(rr); sv.addWidget(rw)
        breset = QtWidgets.QPushButton("⟲ " + L("Slider reset", "Reset sliders")); breset.setObjectName("ghost")
        breset.clicked.connect(self._reset_sliders_btn); sv.addWidget(breset)

        # fix / geometry
        grp(L("SEEDHA / AAKAAR", "FIX / SHAPE"))
        gg = QtWidgets.QGridLayout(); gg.setSpacing(4)

        def gbtn(r, c, icon, name, fn):
            b = QtWidgets.QToolButton(); b.setText(icon + " " + name)
            b.setProperty("class", "act"); b.setToolButtonStyle(QtCore.Qt.ToolButtonTextBesideIcon)
            b.clicked.connect(fn); gg.addWidget(b, r, c)
        gbtn(0, 0, "↺", L("Baayen", "Left"), lambda: self._op(lambda im: im.rotate(90, expand=True)))
        gbtn(0, 1, "↻", L("Dayen", "Right"), lambda: self._op(lambda im: im.rotate(-90, expand=True)))
        gbtn(1, 0, "↔", "Flip H", lambda: self._op(lambda im: im.transpose(Image.FLIP_LEFT_RIGHT)))
        gbtn(1, 1, "↕", "Flip V", lambda: self._op(lambda im: im.transpose(Image.FLIP_TOP_BOTTOM)))
        gbtn(2, 0, "📐", L("Seedha", "Straighten"), lambda: self._op(lambda im: deskew(im).convert("RGB")))
        gbtn(2, 1, "✂", L("Auto-crop", "Auto-crop"), lambda: self._op(lambda im: autocrop(im).convert("RGB")))
        sv.addLayout(gg)
        # live rotate slider
        rrow = QtWidgets.QHBoxLayout()
        rrow.addWidget(QtWidgets.QLabel("🎯"))
        self.sl_angle = QtWidgets.QSlider(QtCore.Qt.Horizontal); self.sl_angle.setRange(-45, 45); self.sl_angle.setValue(0)
        self.sl_angle.sliderPressed.connect(self._angle_begin)
        self.sl_angle.valueChanged.connect(self._angle_preview)
        self.sl_angle.sliderReleased.connect(self._angle_commit)
        self.lbl_angle = QtWidgets.QLabel("0°"); self.lbl_angle.setFixedWidth(34)
        rrow.addWidget(self.sl_angle); rrow.addWidget(self.lbl_angle)
        rw2 = QtWidgets.QWidget(); rw2.setLayout(rrow); sv.addWidget(rw2)

        # colour / cleanup
        grp(L("RANG / SAFAI", "COLOUR / CLEAN"))
        cg = QtWidgets.QGridLayout(); cg.setSpacing(4)

        def cbtn(r, c, icon, name, fn):
            b = QtWidgets.QToolButton(); b.setText(icon + " " + name)
            b.setProperty("class", "act"); b.setToolButtonStyle(QtCore.Qt.ToolButtonTextBesideIcon)
            b.clicked.connect(fn); cg.addWidget(b, r, c)
        cbtn(0, 0, "✨", L("Saaf", "Enhance"), lambda: self._op(lambda im: auto_enhance(im).convert("RGB")))
        cbtn(0, 1, "⬜", "Whiten", lambda: self._op(lambda im: whiten_dark_background(im).convert("RGB")))
        cbtn(1, 0, "🌓", L("Chhaya", "De-shadow"), lambda: self._op(lambda im: flatten_photo_shadows(im).convert("RGB")))
        cbtn(1, 1, "•", L("Dhabbe", "Despeckle"), lambda: self._op(lambda im: im.filter(ImageFilter.MedianFilter(3))))
        cbtn(2, 0, "⬛", "B&W", lambda: self._op(self._bw_fn))
        cbtn(2, 1, "🩶", "Gray", lambda: self._op(lambda im: im.convert("L").convert("RGB")))
        cbtn(3, 0, "🔄", L("Ulta", "Invert"), lambda: self._op(lambda im: ImageOps.invert(im.convert("RGB"))))
        cbtn(3, 1, "🔢", "Page #", self._add_page_number)
        cbtn(4, 0, "🟤", L("Purani", "Sepia"), lambda: self._op(self._sepia_fn))
        cbtn(4, 1, "❄", L("Denoise", "Denoise"), lambda: self._op(lambda im: im.filter(ImageFilter.MedianFilter(5))))
        sv.addLayout(cg)

        # ---- SMART (ek click) ----
        grp(L("SMART", "SMART"))
        sg = QtWidgets.QGridLayout(); sg.setSpacing(4)

        def sgbtn(r, c, icon, name, fn, tip=""):
            b = QtWidgets.QToolButton(); b.setText(icon + " " + name)
            b.setProperty("class", "act"); b.setToolButtonStyle(QtCore.Qt.ToolButtonTextBesideIcon)
            if tip:
                b.setToolTip(tip)
            b.clicked.connect(fn); sg.addWidget(b, r, c)
        sgbtn(0, 0, "🔲", L("Auto kinaara", "Auto edges"), self._smart_perspective,
              L("Kagaz ke 4 kone khud dhoondh kar seedha aayat banao",
                "Auto-detect the paper's 4 corners and flatten to a rectangle"))
        sgbtn(0, 1, "🧾", L("Naam sujhao", "Suggest name"), self._suggest_name,
              L("Andar ka text padhkar file ka naam sujhao", "Read the text and suggest a filename"))
        sgbtn(1, 0, "🗑", L("Khaali hatao", "Drop blank"), self._remove_blank_pages,
              L("Khaali pages dhoondh kar hatao (sabhi pages me)", "Find & remove blank pages (all pages)"))
        sgbtn(1, 1, "⚖", L("Quality", "Quality"), self._quality_check,
              L("Page ki quality jaancho + turant sudhaar", "Check page quality + one-click fix"))
        sv.addLayout(sg)

        # ---- PRIVACY ----
        grp(L("NIJTA (privacy)", "PRIVACY"))
        pg = QtWidgets.QGridLayout(); pg.setSpacing(4)

        def pgbtn(r, c, icon, name, fn, tip=""):
            b = QtWidgets.QToolButton(); b.setText(icon + " " + name)
            b.setProperty("class", "act"); b.setToolButtonStyle(QtCore.Qt.ToolButtonTextBesideIcon)
            if tip:
                b.setToolTip(tip)
            b.clicked.connect(fn); pg.addWidget(b, r, c)
        pgbtn(0, 0, "🙈", L("Auto chhupao", "Auto-hide"), self._auto_redact,
              L("Mobile/Aadhaar/UHID jaisi jaankari khud kaali karo",
                "Auto-black mobile/Aadhaar/UHID-like info"))
        pgbtn(0, 1, "💧", L("Watermark", "Watermark"), self._add_watermark,
              L("Poore page par halka COPY/naam", "Faint COPY/name across the page"))
        sv.addLayout(pg)

        # ---- CROP / LAYOUT ----
        grp(L("CROP / LAYOUT", "CROP / LAYOUT"))
        lg = QtWidgets.QGridLayout(); lg.setSpacing(4)

        def lgbtn(r, c, icon, name, fn, tip=""):
            b = QtWidgets.QToolButton(); b.setText(icon + " " + name)
            b.setProperty("class", "act"); b.setToolButtonStyle(QtCore.Qt.ToolButtonTextBesideIcon)
            if tip:
                b.setToolTip(tip)
            b.clicked.connect(fn); lg.addWidget(b, r, c)
        lgbtn(0, 0, "✂", L("Beech se kaato", "Split page"), self._split_page,
              L("Ek page ko beech se do pages me kaato", "Cut one page into two"))
        lgbtn(0, 1, "➕", L("Hashiya", "Margins"), self._add_margin,
              L("Chaaro taraf safed border jodo", "Add a white border around the page"))
        lgbtn(1, 0, "▦", L("Grid", "Grid"), self._toggle_grid,
              L("Seedha karne ke liye jaali dikhao/chhupao", "Show/hide a straightening grid"))
        lgbtn(1, 1, "🪪", L("ID 2-side", "ID 2-side"), self._id_two_side,
              L("Is page ko agla page ke saath ek A4 par", "This + next page together on one A4"))
        sv.addLayout(lg)

        # ---- TEXT / OCR ----
        grp(L("TEXT / OCR", "TEXT / OCR"))
        tg = QtWidgets.QGridLayout(); tg.setSpacing(4)

        def tgbtn(r, c, icon, name, fn, tip=""):
            b = QtWidgets.QToolButton(); b.setText(icon + " " + name)
            b.setProperty("class", "act"); b.setToolButtonStyle(QtCore.Qt.ToolButtonTextBesideIcon)
            if tip:
                b.setToolTip(tip)
            b.clicked.connect(fn); tg.addWidget(b, r, c)
        tgbtn(0, 0, "📋", L("Text copy", "Copy text"), self._ocr_copy,
              L("Page ka text padhkar clipboard me", "OCR the page text to the clipboard"))
        tgbtn(0, 1, "🔍", L("Text dikhao", "Show text"), self._ocr_show,
              L("Nikala hua text ek box me dikhao", "Show the extracted text in a box"))
        sv.addLayout(tg)
        self.chk_searchable = QtWidgets.QCheckBox(L("PDF me text bhi (khoja ja sake)",
                                                    "Searchable PDF (embed text)"))
        self.chk_searchable.setStyleSheet("font-size:11px;color:#475569;")
        sv.addWidget(self.chk_searchable)

        # ---- kis-kis page par lage ----
        grp(L("KIS PAGE PAR?", "APPLY TO"))
        self.cmb_scope = QtWidgets.QComboBox()
        for _v, _t in (("one", L("Sirf yeh page", "This page only")),
                       ("sel", L("Chune hue pages (filmstrip)", "Selected pages (filmstrip)")),
                       ("all", L("SABHI pages", "ALL pages"))):
            self.cmb_scope.addItem(_t, _v)
        self.cmb_scope.setStyleSheet("font-size:11px;")
        sv.addWidget(self.cmb_scope)
        sv.addStretch(1)
        side_scroll = QtWidgets.QScrollArea(); side_scroll.setWidgetResizable(True)
        side_scroll.setWidget(side); side_scroll.setFixedWidth(288)
        side_scroll.setStyleSheet("QScrollArea{border:none;background:transparent;}")
        mid.addWidget(side_scroll)
        root.addLayout(mid, 1)

        # ---------- FILMSTRIP ----------
        strip = QtWidgets.QFrame(); strip.setObjectName("strip"); strip.setFixedHeight(84)
        stl = QtWidgets.QHBoxLayout(strip); stl.setContentsMargins(6, 4, 6, 4)
        self.film = QtWidgets.QListWidget()
        self.film.setViewMode(QtWidgets.QListView.IconMode); self.film.setFlow(QtWidgets.QListView.LeftToRight)
        self.film.setWrapping(False); self.film.setFixedHeight(72); self.film.setIconSize(QtCore.QSize(48, 60))
        self.film.setMovement(QtWidgets.QListView.Static)
        self.film.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.film.setToolTip(self.L(
            "Click = us page par jao. Ctrl+click = kai pages chuno ('Chune hue pages' scope ke liye).",
            "Click = go to that page. Ctrl+click = pick several ('Selected pages' scope)."))
        self.film.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        self.film.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.film.itemClicked.connect(self._film_click)
        stl.addWidget(self.film)
        root.addWidget(strip)

        # ---------- BOTTOM BAR ----------
        bot = QtWidgets.QHBoxLayout(); bot.setSpacing(5)
        self.btn_prev = QtWidgets.QToolButton(); self.btn_prev.setText("◀"); self.btn_prev.setProperty("class", "act")
        self.btn_prev.clicked.connect(lambda: self._go_page(-1))
        self.lbl_page = QtWidgets.QLabel(""); self.lbl_page.setMinimumWidth(72); self.lbl_page.setAlignment(QtCore.Qt.AlignCenter)
        self.btn_next = QtWidgets.QToolButton(); self.btn_next.setText("▶"); self.btn_next.setProperty("class", "act")
        self.btn_next.clicked.connect(lambda: self._go_page(1))
        for ic, tip, fn in (("➕", L("Page jodo (import)", "Add page (import)"), self._add_page),
                            ("▧", L("Khaali page", "Blank page"), self._blank_page),
                            ("⧉", L("Nakal", "Duplicate"), self._dup_page),
                            ("🗑", L("Hatao", "Delete"), self._del_page)):
            b = QtWidgets.QToolButton(); b.setText(ic); b.setToolTip(tip); b.setProperty("class", "act")
            b.clicked.connect(fn); setattr(self, "_pm_%s" % ic, b)
        bot.addWidget(self.btn_prev); bot.addWidget(self.lbl_page); bot.addWidget(self.btn_next)
        for ic in ("➕", "▧", "⧉", "🗑"):
            bot.addWidget(getattr(self, "_pm_%s" % ic))
        bot.addSpacing(8)
        for ic, name, fn in (("↶", "Undo", self._undo_last), ("↷", "Redo", self._redo_last),
                            ("⟲", L("Reset", "Reset"), self._reset_original)):
            b = QtWidgets.QToolButton(); b.setText(ic + " " + name); b.setProperty("class", "act")
            b.clicked.connect(fn); bot.addWidget(b)
        bot.addStretch(1)
        bmore = QtWidgets.QPushButton("⋯ " + L("Aur", "More")); bmore.setObjectName("ghost")
        bmore.clicked.connect(self._more_menu)
        bot.addWidget(bmore)
        bprint = QtWidgets.QPushButton("🖨 " + L("Print", "Print")); bprint.setObjectName("ghost")
        bprint.clicked.connect(self._print)
        bpdf = QtWidgets.QPushButton("📄 " + L("Poora PDF", "Save all PDF")); bpdf.setObjectName("ghost")
        bpdf.clicked.connect(self._save_all_pdf)
        bsaveas = QtWidgets.QPushButton("⇩ " + L("Alag", "Save as")); bsaveas.setObjectName("ghost")
        bsaveas.clicked.connect(self._save_as)
        bsave = QtWidgets.QPushButton("💾 " + L("Save", "Save")); bsave.setObjectName("primary")
        bsave.clicked.connect(self._save)
        bclose = QtWidgets.QPushButton("✖"); bclose.setObjectName("ghost"); bclose.clicked.connect(self.reject)
        for b in (bprint, bpdf, bsaveas, bsave, bclose):
            bot.addWidget(b)
        root.addLayout(bot)

        self._load_current(initial=True)
        self._build_film()
        QtCore.QTimer.singleShot(0, self._render)
        QtCore.QTimer.singleShot(0, lambda: self._set_tool(self._start_tool or "pan"))

    # ---------- helpers ----------
    def _font(self, size):
        for fn in self._FONTS:
            try:
                return ImageFont.truetype(fn, size)
            except Exception:
                pass
        try:
            return ImageFont.load_default()
        except Exception:
            return None

    def _pil_to_qpix(self, im):
        im = im.convert("RGB"); data = im.tobytes("raw", "RGB")
        qim = QtGui.QImage(data, im.width, im.height, 3 * im.width, QtGui.QImage.Format_RGB888)
        return QtGui.QPixmap.fromImage(qim.copy())

    def _sv(self, key):
        return self._sliders[key].value() / 100.0

    def _composited(self):
        im = self.base
        try:
            if abs(self._sv("bright") - 1) > 0.001:
                im = ImageEnhance.Brightness(im).enhance(self._sv("bright"))
            if abs(self._sv("contrast") - 1) > 0.001:
                im = ImageEnhance.Contrast(im).enhance(self._sv("contrast"))
            if abs(self._sv("sharp") - 1) > 0.001:
                im = ImageEnhance.Sharpness(im).enhance(self._sv("sharp"))
            if abs(self._sv("satur") - 1) > 0.001:
                im = ImageEnhance.Color(im).enhance(self._sv("satur"))
            t = self._sv("temp") - 1.0
            if abs(t) > 0.01:
                r, g, b = im.convert("RGB").split()
                sh = int(t * 30)
                r = r.point(lambda v: max(0, min(255, v + sh)))
                b = b.point(lambda v: max(0, min(255, v - sh)))
                im = Image.merge("RGB", (r, g, b))
        except Exception:
            pass
        return im

    def _bw_fn(self, im):
        thr = self.sl_bw.value()
        return im.convert("L").point(lambda v: 255 if v >= thr else 0).convert("RGB")

    def _render(self):
        img = self.original if self.show_original else (self._preview_img or self._composited())
        w, h = img.size
        fitw, fith = 720, 560
        s = min(fitw / float(w), fith / float(h), 1.0) * self.zoom
        dw, dh = max(1, int(w * s)), max(1, int(h * s))
        self._disp_scale = w / float(dw)
        self.canvas.setPixmap(self._pil_to_qpix(img.resize((dw, dh))))
        self.canvas.setFixedSize(dw, dh)
        self.canvas.update()

    def _set_zoom(self, z):
        self.zoom = max(0.2, min(6.0, z)); self._render()

    def _wheel_zoom(self, d):
        self._set_zoom(self.zoom * (1.2 if d > 0 else 0.83))

    def _toggle_full(self):
        self.showNormal() if self.isFullScreen() else self.showFullScreen()

    def _toggle_before(self, on):
        self.show_original = on; self._render()
        self.btn_ba.setText(("👁 " + self.L("Baad", "After")) if on else ("👁 " + self.L("Pehle", "Before")))

    def _adj_changed(self, _v):
        self._render()

    def _reset_sliders_btn(self):
        for s in self._sliders.values():
            s.blockSignals(True); s.setValue(100); s.blockSignals(False)
        self._render()

    def _update_tool_opts(self):
        t = self.tool
        self.sl_brush.setVisible(t in ("erase", "pen", "blur"))
        self.btn_col.setVisible(t in ("pen", "text", "line", "arrow", "rect", "circle", "note"))
        self.cmb_ratio.setVisible(t == "crop")
        self.cmb_stamp.setVisible(t == "stamp")
        self.sp_tsize.setVisible(t in ("text", "note", "stamp"))

    def _set_tool(self, t):
        self.tool = t
        for k, b in self._tool_btns.items():
            b.setChecked(k == t)
        self._corners = []
        self._update_tool_opts()
        cur = {"pan": QtCore.Qt.OpenHandCursor, "text": QtCore.Qt.IBeamCursor}.get(t, QtCore.Qt.CrossCursor)
        self.canvas.setCursor(cur)

    def _pick_colour(self):
        c = QtWidgets.QColorDialog.getColor(QtGui.QColor(*self.pen_color), self)
        if c.isValid():
            self.pen_color = (c.red(), c.green(), c.blue())

    def _ratio_changed(self, _i):
        self.crop_ratio = {"free": None, "a4": 210.0 / 297.0, "id": 54.0 / 85.6, "sq": 1.0}.get(
            self.cmb_ratio.currentData())

    def _bake(self):
        self._undo.append(self.base.copy())
        if len(self._undo) > 18:
            self._undo.pop(0)
        self._redo = []
        comp = self._composited()
        if comp is not self.base:
            self.base = comp.convert("RGB")
        for s in self._sliders.values():
            s.blockSignals(True); s.setValue(100); s.blockSignals(False)
        self._dirty_any = True

    def _d2i(self, pt):
        x = int(max(0, min(self.base.width, pt.x() * self._disp_scale)))
        y = int(max(0, min(self.base.height, pt.y() * self._disp_scale)))
        return x, y

    # ---- mouse ----
    def _canvas_press(self, pos):
        t = self.tool
        if t == "pan":
            self._panning = True; self._pan_ref = pos
            self.canvas.setCursor(QtCore.Qt.ClosedHandCursor)
        elif t in ("crop", "arrow", "line", "box", "redact", "rect", "circle", "blur"):
            self._start = pos; self._cur = pos
        elif t == "erase":
            self._bake(); self._erasing = True; self._erase_at(pos)
        elif t == "pen":
            self._bake(); self._penning = True; self._pen_last = pos; self._pen_to(pos)
        elif t == "text":
            self._place_text(pos)
        elif t == "note":
            self._place_note(pos)
        elif t == "stamp":
            self._place_stamp(pos)
        elif t == "sign":
            self._place_sign(pos)
        elif t == "persp":
            self._corners.append(pos)
            if len(self._corners) >= 4:
                self._apply_perspective()
            self.canvas.update()

    def _canvas_move(self, pos):
        t = self.tool
        if self._panning and t == "pan":
            d = pos - self._pan_ref
            hb = self.scroll.horizontalScrollBar(); vb = self.scroll.verticalScrollBar()
            hb.setValue(hb.value() - d.x()); vb.setValue(vb.value() - d.y())
        elif self._start and t in ("crop", "arrow", "line", "box", "redact", "rect", "circle", "blur"):
            self._cur = pos; self.canvas.update()
        elif self._erasing and t == "erase":
            self._erase_at(pos)
        elif self._penning and t == "pen":
            self._pen_to(pos)

    def _canvas_release(self, pos):
        t = self.tool
        if self._start:
            if t == "crop":
                self._apply_crop(self._start, pos)
            elif t == "arrow":
                self._apply_arrow(self._start, pos)
            elif t == "line":
                self._apply_line(self._start, pos)
            elif t == "box":
                self._apply_rect(self._start, pos, (255, 235, 59, 90), fill=True)
            elif t == "redact":
                self._apply_rect(self._start, pos, (0, 0, 0, 255), fill=True)
            elif t == "rect":
                self._apply_rect(self._start, pos, self.pen_color, fill=False)
            elif t == "circle":
                self._apply_circle(self._start, pos)
            elif t == "blur":
                self._apply_blur(self._start, pos)
        if self._panning:
            self.canvas.setCursor(QtCore.Qt.OpenHandCursor)
        self._start = self._cur = None
        self._erasing = self._penning = self._panning = False
        self._pen_last = None
        self.canvas.update()

    def _overlay(self, canvas):
        grid_on = getattr(self, "_grid_on", False)
        if not (self._start and self._cur) and not self._corners and not grid_on:
            return
        p = QtGui.QPainter(canvas)
        if grid_on:
            gp = QtGui.QPen(QtGui.QColor(15, 118, 110, 80)); gp.setWidth(1); p.setPen(gp)
            w = canvas.width(); h = canvas.height(); step = max(24, w // 12)
            x = step
            while x < w:
                p.drawLine(x, 0, x, h); x += step
            y = step
            while y < h:
                p.drawLine(0, y, w, y); y += step
        if self._start and self._cur:
            col = "#111827" if self.tool == "redact" else "#0f766e"
            pen = QtGui.QPen(QtGui.QColor(col)); pen.setWidth(2); p.setPen(pen)
            if self.tool in ("arrow", "line"):
                p.drawLine(self._start, self._cur)
            elif self.tool == "circle":
                p.drawEllipse(QtCore.QRect(self._start, self._cur).normalized())
            else:
                p.drawRect(QtCore.QRect(self._start, self._cur).normalized())
        for c in self._corners:
            p.setBrush(QtGui.QColor("#22c55e")); p.setPen(QtCore.Qt.NoPen); p.drawEllipse(c, 5, 5)
        p.end()

    # ---- pixel operations ----
    def _erase_at(self, pos):
        x, y = self._d2i(pos); r = max(3, int(self.sl_brush.value() * self._disp_scale))
        ImageDraw.Draw(self.base).ellipse([x - r, y - r, x + r, y + r], fill=(255, 255, 255))
        self._render()

    def _pen_to(self, pos):
        x, y = self._d2i(pos); w = max(2, int(self.sl_brush.value() * self._disp_scale * 0.5))
        d = ImageDraw.Draw(self.base)
        if self._pen_last is not None:
            lx, ly = self._d2i(self._pen_last); d.line([lx, ly, x, y], fill=self.pen_color, width=w)
        d.ellipse([x - w // 2, y - w // 2, x + w // 2, y + w // 2], fill=self.pen_color)
        self._pen_last = pos; self._render()

    def _apply_crop(self, a, b):
        x0, y0 = self._d2i(a); x1, y1 = self._d2i(b)
        x0, x1 = sorted((x0, x1)); y0, y1 = sorted((y0, y1))
        if x1 - x0 < 8 or y1 - y0 < 8:
            return
        if self.crop_ratio:
            w = x1 - x0; h = y1 - y0
            if w / float(h) > self.crop_ratio:
                x1 = x0 + int(h * self.crop_ratio)
            else:
                y1 = y0 + int(w / self.crop_ratio)
        self._bake(); self.base = self.base.crop((x0, y0, x1, y1)); self._render()

    def _apply_arrow(self, a, b):
        if (a - b).manhattanLength() < 6:
            return
        self._bake()
        x0, y0 = self._d2i(a); x1, y1 = self._d2i(b); w = max(3, int(self.base.width / 300))
        d = ImageDraw.Draw(self.base); col = self.pen_color
        d.line([x0, y0, x1, y1], fill=col, width=w)
        import math
        ang = math.atan2(y1 - y0, x1 - x0); hl = w * 6
        for da in (math.pi - 0.5, math.pi + 0.5):
            d.line([x1, y1, x1 + int(hl * math.cos(ang + da)), y1 + int(hl * math.sin(ang + da))],
                   fill=col, width=w)
        self._render()

    def _apply_line(self, a, b):
        if (a - b).manhattanLength() < 6:
            return
        self._bake()
        x0, y0 = self._d2i(a); x1, y1 = self._d2i(b)
        ImageDraw.Draw(self.base).line([x0, y0, x1, y1], fill=self.pen_color,
                                       width=max(2, int(self.base.width / 350)))
        self._render()

    def _apply_rect(self, a, b, rgba, fill):
        x0, y0 = self._d2i(a); x1, y1 = self._d2i(b)
        x0, x1 = sorted((x0, x1)); y0, y1 = sorted((y0, y1))
        if x1 - x0 < 5 or y1 - y0 < 5:
            return
        self._bake()
        if fill and rgba[3] >= 255:
            ImageDraw.Draw(self.base).rectangle([x0, y0, x1, y1], fill=rgba[:3])
        elif fill:
            ov = Image.new("RGBA", self.base.size, (0, 0, 0, 0))
            ImageDraw.Draw(ov).rectangle([x0, y0, x1, y1], fill=rgba)
            self.base = Image.alpha_composite(self.base.convert("RGBA"), ov).convert("RGB")
        else:
            ImageDraw.Draw(self.base).rectangle([x0, y0, x1, y1], outline=rgba[:3],
                                                width=max(2, int(self.base.width / 350)))
        self._render()

    def _apply_circle(self, a, b):
        x0, y0 = self._d2i(a); x1, y1 = self._d2i(b)
        x0, x1 = sorted((x0, x1)); y0, y1 = sorted((y0, y1))
        if x1 - x0 < 5 or y1 - y0 < 5:
            return
        self._bake()
        ImageDraw.Draw(self.base).ellipse([x0, y0, x1, y1], outline=self.pen_color,
                                          width=max(2, int(self.base.width / 350)))
        self._render()

    def _apply_blur(self, a, b):
        x0, y0 = self._d2i(a); x1, y1 = self._d2i(b)
        x0, x1 = sorted((x0, x1)); y0, y1 = sorted((y0, y1))
        if x1 - x0 < 6 or y1 - y0 < 6:
            return
        self._bake()
        reg = self.base.crop((x0, y0, x1, y1))
        small = reg.resize((max(1, reg.width // 14), max(1, reg.height // 14)))
        self.base.paste(small.resize(reg.size), (x0, y0))
        self._render()

    def _place_text(self, pos):
        txt, ok = QtWidgets.QInputDialog.getText(self, self.L("Text", "Add text"),
                                                 self.L("Kya likhein?", "Text:"))
        if not ok or not txt.strip():
            return
        self._bake(); x, y = self._d2i(pos)
        ImageDraw.Draw(self.base).text((x, y), txt, fill=self.pen_color, font=self._font(self.text_size))
        self._render()

    def _place_sign(self, pos):
        sp = self.win._opts.get("sign_image") or ""
        if not sp or not os.path.exists(sp):
            QtWidgets.QMessageBox.information(self, APP_NAME,
                self.L("Pehle Settings me sign image chuno.", "Choose a signature image in Settings first."))
            return
        try:
            sig = Image.open(sp).convert("RGBA")
        except Exception:
            return
        self._bake(); tw = max(60, int(self.base.width * 0.22)); sig.thumbnail((tw, tw * 3))
        try:
            sig.putdata([(r, g, b, 0) if r > 235 and g > 235 and b > 235 else (r, g, b, a)
                         for r, g, b, a in sig.getdata()])
        except Exception:
            pass
        x, y = self._d2i(pos); self.base = self.base.convert("RGBA")
        self.base.alpha_composite(sig, (max(0, x - sig.width // 2), max(0, y - sig.height // 2)))
        self.base = self.base.convert("RGB"); self._render()

    def _add_page_number(self):
        self._bake()
        try:
            label = "%d / %d" % (self.row + 1, self.win.list.count())
        except Exception:
            label = "1"
        size = max(16, int(self.base.width / 32))
        ImageDraw.Draw(self.base).text(
            (self.base.width - int(size * 5), self.base.height - int(size * 2)),
            label, fill=(60, 60, 60), font=self._font(size))
        self._render()

    def _apply_perspective(self):
        pts = [self._d2i(c) for c in self._corners[:4]]; self._corners = []
        try:
            import math
            (tl, tr, br, bl) = pts
            def dist(a, b):
                return math.hypot(a[0] - b[0], a[1] - b[1])
            ow = int(max(dist(tl, tr), dist(bl, br))); oh = int(max(dist(tl, bl), dist(tr, br)))
            if ow < 10 or oh < 10:
                self._render(); return
            self._bake()
            self.base = self.base.transform((ow, oh), Image.QUAD,
                (tl[0], tl[1], bl[0], bl[1], br[0], br[1], tr[0], tr[1]), Image.BILINEAR)
        except Exception:
            pass
        self._render()

    # ---- live rotate slider ----
    def _angle_begin(self):
        self._rot_base = self._composited().convert("RGB")

    def _angle_preview(self, v):
        self.lbl_angle.setText("%d°" % v)
        if self._rot_base is None:
            self._rot_base = self._composited().convert("RGB")
        self._preview_img = self._rot_base.rotate(-v, expand=True, fillcolor=(255, 255, 255))
        self._render()

    def _angle_commit(self):
        v = self.sl_angle.value()
        if v != 0 and self._rot_base is not None:
            self._bake()
            self.base = self._rot_base.rotate(-v, expand=True, fillcolor=(255, 255, 255)).convert("RGB")
        self._preview_img = None; self._rot_base = None
        self.sl_angle.blockSignals(True); self.sl_angle.setValue(0); self.sl_angle.blockSignals(False)
        self.lbl_angle.setText("0°"); self._render()

    # ---- generic op (yeh page / chune hue / sabhi) ----
    def _scope(self):
        try:
            return self.cmb_scope.currentData()
        except Exception:
            return "one"

    def _op(self, fn):
        sc = self._scope()
        if sc == "all":
            self._apply_all_pages(fn)
        elif sc == "sel":
            rows = sorted({self.film.row(it) for it in self.film.selectedItems()})
            rows = [r for r in rows if r >= 0]
            if len(rows) <= 1:
                self._op_one(fn)
            else:
                self._apply_rows(fn, rows)
        else:
            self._op_one(fn)

    def _op_one(self, fn):
        self._bake()
        try:
            self.base = fn(self.base).convert("RGB")
        except Exception:
            pass
        self._render()

    def _apply_rows(self, fn, rows):
        self._persist()
        try:
            for i in rows:
                if i == self.row:
                    continue
                p = self.win.list.item(i).data(QtCore.Qt.UserRole)
                try:
                    im = Image.open(p).convert("RGB")
                    fn(im).convert("RGB").save(p, "PNG")
                except Exception:
                    pass
            # current page bhi (agar chuna hai)
            if self.row in rows:
                self._op_one(fn)
                self._persist()
            if callable(self.on_saved):
                self.on_saved()
        except Exception:
            pass
        self._dirty_any = False
        self._load_current(); self._build_film()

    def _apply_all_pages(self, fn):
        self._persist()
        try:
            for i in range(self.win.list.count()):
                p = self.win.list.item(i).data(QtCore.Qt.UserRole)
                try:
                    im = Image.open(p).convert("RGB")
                    fn(im).convert("RGB").save(p, "PNG")
                except Exception:
                    pass
            if callable(self.on_saved):
                self.on_saved()
        except Exception:
            pass
        self._dirty_any = False
        self._load_current(); self._build_film()

    def _auto_fix(self):
        def fn(im):
            try:
                im = deskew(im)
                im = autocrop(im)
                im = whiten_dark_background(im)
                im = auto_enhance(im)
            except Exception:
                pass
            return im.convert("RGB")
        self._op(fn)

    def _auto_upright(self):
        if not tesseract_available():
            QtWidgets.QMessageBox.information(self, APP_NAME,
                self.L("Iske liye Tesseract OCR chahiye.", "This needs Tesseract OCR.")); return
        self.setCursor(QtCore.Qt.WaitCursor)
        try:
            osd = pytesseract.image_to_osd(self.base)
            import re as _re
            m = _re.search(r"Rotate: (\d+)", osd)
            rot = int(m.group(1)) if m else 0
        except Exception:
            rot = 0
        self.unsetCursor()
        if rot:
            self._op(lambda im: im.rotate(-rot, expand=True))

    # ---- undo / redo / reset ----
    def _undo_last(self):
        if not self._undo:
            return
        self._redo.append(self.base.copy()); self.base = self._undo.pop(); self._reset_sliders_btn()

    def _redo_last(self):
        if not self._redo:
            return
        self._undo.append(self.base.copy()); self.base = self._redo.pop(); self._reset_sliders_btn()

    def _reset_original(self):
        self._undo.append(self.base.copy()); self._redo = []
        self.base = self.original.copy(); self._reset_sliders_btn()

    # ---- pages ----
    def _load_current(self, initial=False):
        try:
            self.base = Image.open(self.path).convert("RGB")
        except Exception:
            self.base = Image.new("RGB", (827, 1169), "white")
        self.original = self.base.copy()
        self._undo = []; self._redo = []; self._preview_img = None
        for s in self._sliders.values():
            s.blockSignals(True); s.setValue(100); s.blockSignals(False)
        try:
            n = self.win.list.count()
            self.lbl_page.setText("Page %d / %d" % (self.row + 1, n) if self.row >= 0 else "")
            self.lbl_hdr.setText("📄 " + (os.path.basename(self.path) if self.path else ""))
            self.btn_prev.setEnabled(self.row > 0); self.btn_next.setEnabled(0 <= self.row < n - 1)
        except Exception:
            pass
        if not initial:
            self._render()
        self._sync_film()

    def _build_film(self):
        try:
            self.film.clear()
            for i in range(self.win.list.count()):
                src = self.win.list.item(i)
                it = QtWidgets.QListWidgetItem(src.icon(), str(i + 1))
                it.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignBottom)
                self.film.addItem(it)
            self._sync_film()
        except Exception:
            pass

    def _sync_film(self):
        try:
            if 0 <= self.row < self.film.count():
                self.film.setCurrentRow(self.row)
        except Exception:
            pass

    def _film_click(self, it):
        try:
            self._go_row(self.film.row(it))
        except Exception:
            pass

    def _persist(self):
        if not self._dirty_any and not self._undo:
            return
        try:
            self._composited().save(self.path, "PNG")
            if callable(self.on_saved):
                self.on_saved()
        except Exception:
            pass

    def _go_row(self, new):
        try:
            n = self.win.list.count()
        except Exception:
            return
        if new < 0 or new >= n or new == self.row:
            return
        self._persist()
        self.row = new; self.win.list.setCurrentRow(new)
        self.path = self.win.list.item(new).data(QtCore.Qt.UserRole)
        self._dirty_any = False; self._load_current()

    def _go_page(self, delta):
        self._go_row(self.row + delta)

    def _add_page(self):
        self._persist()
        try:
            self.win.import_images()
        except Exception:
            pass
        self._build_film()
        try:
            self.row = self.win.list.currentRow()
            self.path = self.win.list.item(self.row).data(QtCore.Qt.UserRole)
            self._load_current()
        except Exception:
            pass

    def _blank_page(self):
        self._persist()
        try:
            im = Image.new("RGB", self.base.size, "white")
            fd, out = tempfile.mkstemp(suffix=".png", dir=self.win._tmpdir); os.close(fd)
            im.save(out, "PNG")
            self.win._add_item_for_path(out, at=self.row + 1)
            self.row += 1; self.win.list.setCurrentRow(self.row); self.path = out
            self._dirty_any = False; self._load_current(); self._build_film()
        except Exception:
            pass

    def _dup_page(self):
        self._persist()
        try:
            self.win.list.setCurrentRow(self.row); self.win.duplicate_current_page()
            self.row = self.win.list.currentRow()
            self.path = self.win.list.item(self.row).data(QtCore.Qt.UserRole)
            self._dirty_any = False; self._load_current(); self._build_film()
        except Exception:
            pass

    def _del_page(self):
        try:
            if self.win.list.count() <= 1:
                QtWidgets.QMessageBox.information(self, APP_NAME,
                    self.L("Ye aakhri page hai.", "This is the only page.")); return
            self.win.list.setCurrentRow(self.row); self.win.delete_page()
            self.row = min(self.row, self.win.list.count() - 1)
            self.win.list.setCurrentRow(self.row)
            self.path = self.win.list.item(self.row).data(QtCore.Qt.UserRole)
            self._dirty_any = False; self._load_current(); self._build_film()
        except Exception:
            pass

    # ---- save / export ----
    def _save(self):
        try:
            self._composited().save(self.path, "PNG")
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, APP_NAME, str(e)); return
        if callable(self.on_saved):
            try:
                self.on_saved()
            except Exception:
                pass
        self.accept()

    def _save_as(self):
        out, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, self.L("Alag save", "Save as"),
            os.path.join(self.win._opts.get("save_folder", os.path.expanduser("~")), "page.png"),
            "PNG (*.png);;JPEG (*.jpg);;PDF (*.pdf)")
        if not out:
            return
        try:
            img = self._composited().convert("RGB")
            img.save(out, "PDF", resolution=200) if out.lower().endswith(".pdf") else img.save(out, quality=92)
            self.win.status.showMessage(self.L("⇩ Save: ", "⇩ Saved: ") + os.path.basename(out), 6000)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, APP_NAME, str(e))

    def _save_all_pdf(self):
        self._persist()
        out, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, self.L("Poora doc PDF", "Save whole document as PDF"),
            os.path.join(self.win._opts.get("save_folder", os.path.expanduser("~")), "document.pdf"),
            "PDF (*.pdf)")
        if not out:
            return
        try:
            paths = [self.win.list.item(i).data(QtCore.Qt.UserRole) for i in range(self.win.list.count())]
            want_txt = False
            try:
                want_txt = self.chk_searchable.isChecked()
            except Exception:
                pass
            if want_txt and tesseract_available():
                self.setCursor(QtCore.Qt.WaitCursor)
                try:
                    self.win._save_ocr_pdf(paths, out)
                finally:
                    self.unsetCursor()
                self.win.status.showMessage(
                    self.L("📄 Searchable PDF ban gaya: ", "📄 Searchable PDF saved: ") + os.path.basename(out), 7000)
            else:
                self.win._pages_as_pdf(paths, out)
                self.win.status.showMessage(self.L("📄 PDF ban gaya: ", "📄 PDF saved: ") + os.path.basename(out), 7000)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, APP_NAME, str(e))

    def _print(self):
        self._persist()
        try:
            self.win.list.setCurrentRow(self.row)
            self.win._do_print([self.path], per_page=1)
        except Exception:
            pass

    # ================= SMART / COLOUR extras =================
    def _sepia_fn(self, im):
        try:
            g = im.convert("L")
            return ImageOps.colorize(g, black=(35, 22, 10), white=(255, 240, 205)).convert("RGB")
        except Exception:
            return im.convert("RGB")

    def _magic_color(self):
        def fn(im):
            try:
                im = whiten_dark_background(im)
                im = auto_enhance(im).convert("RGB")
                im = ImageEnhance.Contrast(im).enhance(1.15)
                im = ImageEnhance.Sharpness(im).enhance(1.3)
            except Exception:
                pass
            return im.convert("RGB")
        self._op(fn)

    def _smart_perspective(self):
        def fn(im):
            try:
                im = deskew(im)
                im = autocrop(im)
            except Exception:
                pass
            return im.convert("RGB")
        self._op(fn)

    def _suggest_name(self):
        if not tesseract_available():
            QtWidgets.QMessageBox.information(self, APP_NAME,
                self.L("Iske liye Tesseract OCR chahiye.", "This needs Tesseract OCR.")); return
        self.setCursor(QtCore.Qt.WaitCursor)
        txt = self._ocr_text(self.base)
        self.unsetCursor()
        import re as _re
        cand = ""
        for line in (txt or "").splitlines():
            s = line.strip()
            if len(s) >= 4 and any(c.isalpha() for c in s):
                cand = s; break
        cand = _re.sub(r"[^A-Za-z0-9 \-]", "", cand).strip().replace(" ", "_")[:40] or "document"
        name, ok = QtWidgets.QInputDialog.getText(
            self, self.L("Naam sujhao", "Suggest name"),
            self.L("Document ka naam:", "Document name:"), text=cand)
        if ok and name.strip():
            try:
                self.win.claim_edit.setText(name.strip())
                self.win.status.showMessage(self.L("Naam set: ", "Name set: ") + name.strip(), 5000)
            except Exception:
                QtWidgets.QApplication.clipboard().setText(name.strip())

    def _remove_blank_pages(self):
        n = self.win.list.count()
        if n <= 1:
            QtWidgets.QMessageBox.information(self, APP_NAME,
                self.L("Sirf ek page hai.", "Only one page.")); return
        self._persist()
        self.setCursor(QtCore.Qt.WaitCursor)
        blanks = []
        for i in range(n):
            p = self.win.list.item(i).data(QtCore.Qt.UserRole)
            try:
                if is_blank_page(Image.open(p)):
                    blanks.append(i)
            except Exception:
                pass
        self.unsetCursor()
        if not blanks:
            QtWidgets.QMessageBox.information(self, APP_NAME,
                self.L("Koi khaali page nahi mila.", "No blank pages found.")); return
        if len(blanks) >= n:
            blanks = blanks[1:]        # kam se kam ek page bacha rahe
        if not blanks:
            return
        if QtWidgets.QMessageBox.question(
                self, APP_NAME,
                self.L("%d khaali page mile. Hataayein?" % len(blanks),
                       "Found %d blank page(s). Remove them?" % len(blanks)),
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No) != QtWidgets.QMessageBox.Yes:
            return
        try:
            self.win.list.clearSelection()
            for i in blanks:
                it = self.win.list.item(i)
                if it:
                    it.setSelected(True)
            self.win.delete_page()
            self.row = min(self.row, self.win.list.count() - 1)
            self.win.list.setCurrentRow(self.row)
            self.path = self.win.list.item(self.row).data(QtCore.Qt.UserRole)
            self._dirty_any = False
            self._load_current(); self._build_film()
            if callable(self.on_saved):
                self.on_saved()
        except Exception:
            pass

    def _quality_check(self):
        im = self.base.convert("L")
        small = im.resize((min(420, im.width), min(420, im.height)))
        px = list(small.getdata()); mean = (sum(px) / len(px)) if px else 128
        issues = []
        if mean < 90:
            issues.append(self.L("Page bahut gehra/kaala hai — 'Ujla' ya Auto-Fix karein.",
                                 "Page is too dark — try brighten or Auto-Fix."))
        elif mean > 212:
            issues.append(self.L("Page bahut halka/feeka hai — Contrast badhaayein.",
                                 "Page is very faint — increase contrast."))
        if HAS_NUMPY:
            try:
                import numpy as _np
                a = _np.asarray(small, dtype=_np.float32)
                sharp = _np.abs(_np.diff(a, axis=1)).mean()
                if sharp < 4:
                    issues.append(self.L("Page dhundhla lag raha hai — 'Dhaar' slider badhaayein.",
                                         "Page looks blurry — raise the sharpness slider."))
            except Exception:
                pass
        if not issues:
            QtWidgets.QMessageBox.information(self, APP_NAME,
                self.L("Page theek dikh raha hai ✔", "This page looks good ✔")); return
        msg = self.L("Sujhaav:", "Suggestions:") + "\n\n• " + "\n• ".join(issues) + \
            "\n\n" + self.L("Abhi Auto-Fix laga dein?", "Apply Auto-Fix now?")
        if QtWidgets.QMessageBox.question(self, APP_NAME, msg,
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No) == QtWidgets.QMessageBox.Yes:
            self._auto_fix()

    # ================= PRIVACY =================
    def _auto_redact(self):
        if not tesseract_available():
            QtWidgets.QMessageBox.information(self, APP_NAME,
                self.L("Iske liye Tesseract OCR chahiye.", "This needs Tesseract OCR.")); return
        import re as _re
        self.setCursor(QtCore.Qt.WaitCursor)
        try:
            data = pytesseract.image_to_data(self.base.convert("RGB"),
                                             output_type=pytesseract.Output.DICT)
        except Exception:
            self.unsetCursor()
            QtWidgets.QMessageBox.warning(self, APP_NAME,
                self.L("OCR nahi ho paaya.", "OCR failed.")); return
        self.unsetCursor()
        boxes = []
        words = data.get("text", [])
        for i, wtok in enumerate(words):
            digits = _re.sub(r"\D", "", (wtok or ""))
            if len(digits) >= 8:      # mobile(10) / Aadhaar(12) / UHID jaise
                try:
                    boxes.append((data["left"][i], data["top"][i],
                                  data["width"][i], data["height"][i]))
                except Exception:
                    pass
        if not boxes:
            QtWidgets.QMessageBox.information(self, APP_NAME,
                self.L("Koi mobile/number jaisi jaankari nahi mili.",
                       "No phone/number-like info found.")); return
        self._bake()
        d = ImageDraw.Draw(self.base)
        for (x, y, w, h) in boxes:
            d.rectangle([x - 2, y - 2, x + w + 2, y + h + 2], fill=(0, 0, 0))
        self._render()
        self.win.status.showMessage(
            self.L("%d jagah chhupa di gayi." % len(boxes),
                   "Hid %d spot(s)." % len(boxes)), 6000)

    def _add_watermark(self):
        txt, ok = QtWidgets.QInputDialog.getText(
            self, self.L("Watermark", "Watermark"),
            self.L("Kya likhein? (poore page par halka)", "Text (faint, across the page):"),
            text="COPY")
        if not ok or not txt.strip():
            return
        txt = txt.strip()

        def fn(im):
            im = im.convert("RGBA"); W, H = im.size
            lay = Image.new("RGBA", (W, H), (0, 0, 0, 0))
            d = ImageDraw.Draw(lay); f = self._font(max(28, int(W / 12)))
            step = max(140, int(W / 3)); y = -H // 3
            while y < int(H * 1.4):
                x = -W // 3
                while x < int(W * 1.4):
                    d.text((x, y), txt, font=f, fill=(120, 120, 120, 55))
                    x += step * 2
                y += step
            try:
                lay = lay.rotate(30, expand=False)
            except Exception:
                pass
            return Image.alpha_composite(im, lay).convert("RGB")
        self._op(fn)

    # ================= CROP / LAYOUT =================
    def _split_page(self):
        box = QtWidgets.QMessageBox(self)
        box.setWindowTitle(APP_NAME)
        box.setText(self.L("Page ko kaise kaatein?", "How to split the page?"))
        b_lr = box.addButton(self.L("Baayen | Dayen", "Left | Right"), QtWidgets.QMessageBox.AcceptRole)
        b_tb = box.addButton(self.L("Upar | Neeche", "Top | Bottom"), QtWidgets.QMessageBox.AcceptRole)
        box.addButton(self.L("Rehne do", "Cancel"), QtWidgets.QMessageBox.RejectRole)
        box.exec_()
        clicked = box.clickedButton()
        if clicked not in (b_lr, b_tb):
            return
        self._persist()
        im = self.base.convert("RGB"); W, H = im.size
        if clicked is b_lr:
            p1 = im.crop((0, 0, W // 2, H)); p2 = im.crop((W // 2, 0, W, H))
        else:
            p1 = im.crop((0, 0, W, H // 2)); p2 = im.crop((0, H // 2, W, H))
        try:
            p1.save(self.path, "PNG")
            fd, out = tempfile.mkstemp(suffix=".png", dir=self.win._tmpdir); os.close(fd)
            p2.save(out, "PNG")
            self.win._add_item_for_path(out, at=self.row + 1)
            self._dirty_any = False
            self._load_current(); self._build_film()
            if callable(self.on_saved):
                self.on_saved()
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, APP_NAME, str(e))

    def _add_margin(self):
        def fn(im):
            im = im.convert("RGB")
            m = max(12, int(im.width * 0.05))
            return ImageOps.expand(im, border=m, fill="white")
        self._op(fn)

    def _toggle_grid(self):
        self._grid_on = not getattr(self, "_grid_on", False)
        self.canvas.update()

    def _id_two_side(self):
        n = self.win.list.count()
        if self.row >= n - 1:
            QtWidgets.QMessageBox.information(self, APP_NAME,
                self.L("Iske aage koi page nahi. ID ke dono side do alag page par scan karein.",
                       "No page after this. Scan both ID sides as two pages first.")); return
        self._persist()
        try:
            a = Image.open(self.path).convert("RGB")
            b = Image.open(self.win.list.item(self.row + 1).data(QtCore.Qt.UserRole)).convert("RGB")
            A4 = (1240, 1754)
            canvas = Image.new("RGB", A4, "white")
            half = A4[1] // 2

            def fit(im, bw, bh):
                im = im.copy(); im.thumbnail((bw - 60, bh - 60)); return im
            fa = fit(a, A4[0], half); fb = fit(b, A4[0], half)
            canvas.paste(fa, ((A4[0] - fa.width) // 2, (half - fa.height) // 2))
            canvas.paste(fb, ((A4[0] - fb.width) // 2, half + (half - fb.height) // 2))
            fd, out = tempfile.mkstemp(suffix=".png", dir=self.win._tmpdir); os.close(fd)
            canvas.save(out, "PNG")
            self.win._add_item_for_path(out, at=self.row + 2)
            self._build_film()
            if callable(self.on_saved):
                self.on_saved()
            self.win.status.showMessage(
                self.L("🪪 ID dono side ek A4 par ban gaya (naya page).",
                       "🪪 Both ID sides placed on one A4 (new page)."), 6000)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, APP_NAME, str(e))

    # ================= TEXT / OCR =================
    def _ocr_text(self, im):
        try:
            return pytesseract.image_to_string(im.convert("RGB"), lang="eng+hin")
        except Exception:
            try:
                return pytesseract.image_to_string(im.convert("RGB"))
            except Exception:
                return ""

    def _ocr_copy(self):
        if not tesseract_available():
            QtWidgets.QMessageBox.information(self, APP_NAME,
                self.L("Iske liye Tesseract OCR chahiye.", "This needs Tesseract OCR.")); return
        self.setCursor(QtCore.Qt.WaitCursor)
        t = self._ocr_text(self.base)
        self.unsetCursor()
        QtWidgets.QApplication.clipboard().setText(t or "")
        if (t or "").strip():
            self.win.status.showMessage(self.L("📋 Text copy ho gaya.", "📋 Text copied."), 5000)
        else:
            QtWidgets.QMessageBox.information(self, APP_NAME,
                self.L("Is page par koi text nahi mila.", "No text found on this page."))

    def _ocr_show(self):
        if not tesseract_available():
            QtWidgets.QMessageBox.information(self, APP_NAME,
                self.L("Iske liye Tesseract OCR chahiye.", "This needs Tesseract OCR.")); return
        self.setCursor(QtCore.Qt.WaitCursor)
        t = self._ocr_text(self.base)
        self.unsetCursor()
        dlg = QtWidgets.QDialog(self); dlg.setWindowTitle(self.L("Page ka text", "Page text"))
        dlg.resize(560, 600)
        v = QtWidgets.QVBoxLayout(dlg)
        te = QtWidgets.QTextEdit(); te.setPlainText(t or ""); v.addWidget(te)
        h = QtWidgets.QHBoxLayout(); h.addStretch(1)
        bc = QtWidgets.QPushButton(self.L("Copy", "Copy")); bc.setObjectName("ghost")
        bc.clicked.connect(lambda: QtWidgets.QApplication.clipboard().setText(te.toPlainText()))
        bx = QtWidgets.QPushButton(self.L("Band karo", "Close")); bx.setObjectName("primary")
        bx.clicked.connect(dlg.accept)
        h.addWidget(bc); h.addWidget(bx); v.addLayout(h)
        dlg.exec_()

    # ================= NOTE / STAMP tools =================
    def _place_note(self, pos):
        txt, ok = QtWidgets.QInputDialog.getMultiLineText(
            self, self.L("Note", "Sticky note"), self.L("Note likhein:", "Note text:"))
        if not ok or not txt.strip():
            return
        self._bake(); x, y = self._d2i(pos)
        f = self._font(self.text_size)
        d = ImageDraw.Draw(self.base)
        try:
            bb = d.multiline_textbbox((0, 0), txt, font=f); tw = bb[2] - bb[0]; th = bb[3] - bb[1]
        except Exception:
            tw = len(txt) * self.text_size // 2; th = self.text_size * (txt.count("\n") + 1)
        pad = max(6, int(self.text_size * 0.4))
        d.rectangle([x, y, x + tw + 2 * pad, y + th + 2 * pad],
                    fill=(255, 245, 157), outline=(251, 192, 45), width=2)
        d.multiline_text((x + pad, y + pad), txt, fill=(60, 50, 10), font=f)
        self._render()

    def _place_stamp(self, pos):
        g = self.cmb_stamp.currentData()
        self._bake(); x, y = self._d2i(pos)
        d = ImageDraw.Draw(self.base)
        if g == "date":
            import datetime as _dt
            s = _dt.datetime.now().strftime("%d-%b-%Y")
            d.text((x, y), s, fill=(200, 20, 20), font=self._font(max(20, self.text_size)))
        elif g in ("approved", "paid"):
            s = "APPROVED" if g == "approved" else "PAID"
            f = self._font(max(24, int(self.text_size * 1.2)))
            try:
                bb = d.textbbox((0, 0), s, font=f); tw = bb[2] - bb[0]; th = bb[3] - bb[1]
            except Exception:
                tw = len(s) * self.text_size // 2; th = self.text_size
            pad = max(6, int(self.text_size * 0.4))
            d.rectangle([x, y, x + tw + 2 * pad, y + th + 2 * pad], outline=(200, 20, 20), width=4)
            d.text((x + pad, y + pad), s, fill=(200, 20, 20), font=f)
        else:
            col = (22, 160, 80) if g == "✔" else ((200, 20, 20) if g == "✘" else (240, 170, 20))
            d.text((x, y), g, fill=col, font=self._font(max(30, int(self.text_size * 1.6))))
        self._render()

    # ================= MORE menu (reorder / preset / share / password) =================
    def _more_menu(self):
        m = QtWidgets.QMenu(self)
        m.addAction("◀ " + self.L("Page baayen le jao", "Move page left"), lambda: self._move_page(-1))
        m.addAction("▶ " + self.L("Page dayen le jao", "Move page right"), lambda: self._move_page(1))
        m.addAction("⭜ " + self.L("Agle page se jodo", "Merge with next page"), self._merge_next)
        m.addSeparator()
        pm = m.addMenu("⭐ " + self.L("Preset (slider yaad)", "Preset (saved sliders)"))
        pm.addAction(self.L("Abhi ke slider save karo…", "Save current sliders…"), self._save_preset)
        presets = self.win._opts.get("editor_presets", {}) or {}
        if presets:
            pm.addSeparator()
            for nm in list(presets.keys()):
                pm.addAction(nm, lambda _c=False, n=nm: self._apply_preset(n))
        m.addSeparator()
        m.addAction("🔒 " + self.L("Password-wali PDF…", "Password-protect PDF…"), self._password_pdf)
        sm = m.addMenu("📤 " + self.L("Bhejo", "Share"))
        sm.addAction("🟢 WhatsApp", lambda: self._share("wa"))
        sm.addAction("✉ Email", lambda: self._share("mail"))
        m.addSeparator()
        m.addAction("⌨ " + self.L("Shortcut list", "Keyboard shortcuts"), self._shortcuts_help)
        m.exec_(QtGui.QCursor.pos())

    def _move_page(self, delta):
        n = self.win.list.count(); new = self.row + delta
        if new < 0 or new >= n:
            return
        self._persist()
        try:
            it = self.win.list.takeItem(self.row)
            self.win.list.insertItem(new, it)
            self.row = new; self.win.list.setCurrentRow(new)
            self.path = self.win.list.item(new).data(QtCore.Qt.UserRole)
            try:
                self.win._renumber_pages()
            except Exception:
                pass
            self._dirty_any = False
            self._load_current(); self._build_film()
            if callable(self.on_saved):
                self.on_saved()
        except Exception:
            pass

    def _merge_next(self):
        n = self.win.list.count()
        if self.row >= n - 1:
            QtWidgets.QMessageBox.information(self, APP_NAME,
                self.L("Iske aage koi page nahi.", "There is no next page.")); return
        self._persist()
        try:
            nxt = self.win.list.item(self.row + 1).data(QtCore.Qt.UserRole)
            a = Image.open(self.path).convert("RGB")
            b = Image.open(nxt).convert("RGB")
            W = max(a.width, b.width)

            def rw(im):
                return im if im.width == W else im.resize((W, max(1, int(im.height * W / im.width))))
            a2 = rw(a); b2 = rw(b)
            combo = Image.new("RGB", (W, a2.height + b2.height), "white")
            combo.paste(a2, (0, 0)); combo.paste(b2, (0, a2.height))
            combo.save(self.path, "PNG")
            self.win.list.clearSelection()
            self.win.list.item(self.row + 1).setSelected(True)
            self.win.delete_page()
            self.win.list.setCurrentRow(self.row)
            self.path = self.win.list.item(self.row).data(QtCore.Qt.UserRole)
            self._dirty_any = False
            self._load_current(); self._build_film()
            if callable(self.on_saved):
                self.on_saved()
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, APP_NAME, str(e))

    def _save_preset(self):
        name, ok = QtWidgets.QInputDialog.getText(
            self, self.L("Preset", "Preset"), self.L("Preset ka naam:", "Preset name:"))
        if not ok or not name.strip():
            return
        d = {k: self._sliders[k].value() for k in self._sliders}
        d["bw"] = self.sl_bw.value()
        presets = self.win._opts.setdefault("editor_presets", {})
        presets[name.strip()] = d
        try:
            self.win._save_opts()
        except Exception:
            pass
        self.win.status.showMessage(self.L("⭐ Preset save: ", "⭐ Preset saved: ") + name.strip(), 5000)

    def _apply_preset(self, name):
        d = (self.win._opts.get("editor_presets", {}) or {}).get(name)
        if not d:
            return
        for k, s in self._sliders.items():
            if k in d:
                s.blockSignals(True); s.setValue(int(d[k])); s.blockSignals(False)
        if "bw" in d:
            self.sl_bw.setValue(int(d["bw"]))
        self._render()

    def _password_pdf(self):
        pw, ok = QtWidgets.QInputDialog.getText(
            self, self.L("Password PDF", "Password PDF"),
            self.L("PDF ka password:", "PDF password:"), QtWidgets.QLineEdit.Normal, "")
        if not ok or not pw:
            return
        out, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, self.L("Password-wali PDF", "Password PDF"),
            os.path.join(self.win._opts.get("save_folder", os.path.expanduser("~")), "document.pdf"),
            "PDF (*.pdf)")
        if not out:
            return
        self._persist()
        try:
            paths = [self.win.list.item(i).data(QtCore.Qt.UserRole) for i in range(self.win.list.count())]
            self.win._pages_as_pdf(paths, out, password=pw)
            self.win.status.showMessage(self.L("🔒 Password PDF ban gaya: ", "🔒 Password PDF saved: ")
                                        + os.path.basename(out), 7000)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, APP_NAME, str(e))

    def _share(self, how):
        self._persist()
        try:
            paths = [self.win.list.item(i).data(QtCore.Qt.UserRole) for i in range(self.win.list.count())]
            fd, out = tempfile.mkstemp(suffix=".pdf", dir=self.win._tmpdir); os.close(fd)
            self.win._pages_as_pdf(paths, out)
            if how == "wa":
                self.win.share_whatsapp(out)
            else:
                self.win.share_email(out)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, APP_NAME, str(e))

    def _shortcuts_help(self):
        QtWidgets.QMessageBox.information(
            self, self.L("Shortcut", "Keyboard shortcuts"),
            "C — Crop\nE — Erase\nP — Pen\nT — Text\nA — Arrow\nH — Highlight\n"
            "B — Blur\nV — Pan\n[  — " + self.L("Baayen ghumao", "Rotate left") + "\n"
            "]  — " + self.L("Dayen ghumao", "Rotate right") + "\n"
            "Ctrl+Z — Undo\nCtrl+Shift+Z / Ctrl+Y — Redo\nCtrl+S — " + self.L("Save", "Save") + "\n"
            "Ctrl+" + self.L("scroll", "scroll") + " — Zoom\nEsc — " + self.L("Band karo", "Close"))

    def keyPressEvent(self, e):
        k = e.key(); mod = e.modifiers()
        keys = {QtCore.Qt.Key_C: "crop", QtCore.Qt.Key_E: "erase", QtCore.Qt.Key_P: "pen",
                QtCore.Qt.Key_T: "text", QtCore.Qt.Key_A: "arrow", QtCore.Qt.Key_H: "box",
                QtCore.Qt.Key_B: "blur", QtCore.Qt.Key_V: "pan"}
        if k == QtCore.Qt.Key_Escape:
            self.reject()
        elif k == QtCore.Qt.Key_Z and (mod & QtCore.Qt.ControlModifier):
            (self._redo_last if (mod & QtCore.Qt.ShiftModifier) else self._undo_last)()
        elif k == QtCore.Qt.Key_Y and (mod & QtCore.Qt.ControlModifier):
            self._redo_last()
        elif k == QtCore.Qt.Key_S and (mod & QtCore.Qt.ControlModifier):
            self._save()
        elif (not mod) and k in keys:
            self._set_tool(keys[k])
        elif k == QtCore.Qt.Key_BracketRight:
            self._op(lambda im: im.rotate(-90, expand=True))
        elif k == QtCore.Qt.Key_BracketLeft:
            self._op(lambda im: im.rotate(90, expand=True))
        else:
            super().keyPressEvent(e)


class CameraDialog(QtWidgets.QDialog):
    """Webcam/USB-camera se document capture — scanner na ho to bhi PDF banao.
    Capture ki gayi har photo apne aap saaf (shadow hatana/seedha) hoti hai."""

    def __init__(self, parent, tmpdir):
        super().__init__(parent)
        self.tmpdir = tmpdir
        self.captured = []          # saved temp image paths
        self.setWindowTitle("📷 Camera se scan")
        self.resize(720, 560)
        v = QtWidgets.QVBoxLayout(self)
        self.viewf = QCameraViewfinder()
        v.addWidget(self.viewf, 1)
        row = QtWidgets.QHBoxLayout()
        self.cmb = QtWidgets.QComboBox()
        self._cams = QCameraInfo.availableCameras()
        for c in self._cams:
            self.cmb.addItem(c.description())
        row.addWidget(self.cmb, 1)
        self.lbl = QtWidgets.QLabel("0 captured")
        row.addWidget(self.lbl)
        v.addLayout(row)
        brow = QtWidgets.QHBoxLayout()
        self.b_shot = QtWidgets.QPushButton("📸 Capture (Space)")
        self.b_shot.setObjectName("primary"); self.b_shot.setMinimumHeight(40)
        self.b_shot.clicked.connect(self._capture)
        b_done = QtWidgets.QPushButton("✔ Done — add pages")
        b_done.clicked.connect(self.accept)
        b_cancel = QtWidgets.QPushButton("Cancel")
        b_cancel.clicked.connect(self.reject)
        brow.addWidget(self.b_shot, 2); brow.addWidget(b_done, 1); brow.addWidget(b_cancel)
        v.addLayout(brow)
        QtWidgets.QShortcut(QtGui.QKeySequence("Space"), self, self._capture)
        self.camera = None
        self.cap = None
        self.cmb.currentIndexChanged.connect(self._start_camera)
        if self._cams:
            self._start_camera(0)
        else:
            self.b_shot.setEnabled(False)
            self.viewf.setStyleSheet("background:#111;")

    def _start_camera(self, idx):
        try:
            if self.camera:
                self.camera.stop()
            self.camera = QCamera(self._cams[idx])
            self.camera.setViewfinder(self.viewf)
            self.cap = QCameraImageCapture(self.camera)
            self.cap.imageCaptured.connect(self._on_captured)
            self.camera.start()
        except Exception:
            self.b_shot.setEnabled(False)

    def _capture(self):
        if not self.cap:
            return
        try:
            self.cap.capture()          # imageCaptured signal me frame aayega
        except Exception:
            pass

    def _on_captured(self, _id, qimg):
        try:
            fd, png = tempfile.mkstemp(suffix=".jpg", dir=self.tmpdir)
            os.close(fd)
            qimg.save(png, "JPG", 92)
            # scan-jaisa saaf: shadow hatana + seedha (clean_photo module-level hai)
            try:
                with Image.open(png) as im:
                    clean_photo(im).save(png, "JPEG", quality=90)
            except Exception:
                pass
            self.captured.append(png)
            self.lbl.setText("%d captured" % len(self.captured))
        except Exception:
            pass

    def closeEvent(self, e):
        try:
            if self.camera:
                self.camera.stop()
        except Exception:
            pass
        super().closeEvent(e)


class SparkBars(QtWidgets.QWidget):
    """Chhota bar-chart (bina kisi library ke) — dashboard ke liye."""

    def __init__(self, values, labels=None, color="#0f766e"):
        super().__init__()
        self.v = [max(0, int(x or 0)) for x in (values or [0])]
        self.labels = labels
        self.c = QtGui.QColor(color)
        self.setMinimumHeight(96)

    def set_values(self, values, labels=None):
        self.v = [max(0, int(x or 0)) for x in (values or [0])]
        if labels is not None:
            self.labels = labels
        self.update()

    def paintEvent(self, _e):
        p = QtGui.QPainter(self)
        try:
            W, H = self.width(), self.height() - 18
            mx = max(self.v) or 1
            n = max(1, len(self.v))
            gap = W // n
            bw = max(6, int(gap * 0.6))
            p.setPen(QtCore.Qt.NoPen)
            p.setBrush(self.c)
            for i, val in enumerate(self.v):
                h = int((H - 14) * val / mx)
                x = i * gap + (gap - bw) // 2
                p.drawRect(x, H - h, bw, h)
            p.setPen(QtGui.QColor("#64748b"))
            f = p.font(); f.setPointSize(7); p.setFont(f)
            for i, val in enumerate(self.v):
                p.drawText(i * gap, 0, gap, 12, QtCore.Qt.AlignCenter, str(val))
            if self.labels:
                for i, lb in enumerate(self.labels):
                    p.drawText(i * gap, H + 2, gap, 14, QtCore.Qt.AlignCenter, str(lb)[:3])
        finally:
            p.end()


class LibraryModel(QtWidgets.QFileSystemModel):
    """'Meri Files' panel ka model — AAJ banayi files hari dikhti hain, aur
    har file ke naam ke aage () me uska size bhi dikhta hai."""

    @staticmethod
    def _fmt_size(nbytes):
        try:
            nbytes = float(nbytes)
        except Exception:
            return ""
        if nbytes < 1024:
            return "%d B" % int(nbytes)
        if nbytes < 1024 * 1024:
            return "%d KB" % round(nbytes / 1024.0)
        if nbytes < 1024 * 1024 * 1024:
            v = nbytes / (1024.0 * 1024.0)
            return ("%.1f MB" % v) if v < 10 else ("%d MB" % round(v))
        return "%.1f GB" % (nbytes / (1024.0 * 1024.0 * 1024.0))

    def data(self, index, role=QtCore.Qt.DisplayRole):
        if index.column() == 0:
            if role == QtCore.Qt.DisplayRole:
                try:
                    fi = self.fileInfo(index)
                    if fi.isFile():
                        # QFileSystemModel ka cached size kabhi-kabhi 0/stale hota
                        # hai (file abhi bani/likhi gayi ho). Isliye asli size
                        # seedha disk se padho — taaki HAR file ka size sahi dikhe.
                        sz = 0
                        try:
                            sz = os.path.getsize(self.filePath(index))
                        except Exception:
                            sz = 0
                        if not sz:
                            try:
                                sz = int(fi.size())
                            except Exception:
                                sz = 0
                        return "(%s) %s" % (self._fmt_size(sz), fi.fileName())
                except Exception:
                    pass
            elif role == QtCore.Qt.ForegroundRole:
                try:
                    fi = self.fileInfo(index)
                    if fi.isFile() and fi.lastModified().date() == QtCore.QDate.currentDate():
                        return QtGui.QBrush(QtGui.QColor("#16a34a"))
                except Exception:
                    pass
        return super().data(index, role)


class FilesTree(QtWidgets.QTreeView):
    """Folder tree jo pages ka DROP le leta hai — page ko kheench kar folder
    par chhodo = seedha wahan save."""

    def __init__(self, on_drop):
        super().__init__()
        self._on_drop = on_drop
        self._on_activate = None      # Enter dabane par folder/file kholo
        self.setAcceptDrops(True)
        # Files ko yahan se KHEENCH kar doc-area me drop karke import karo
        self.setDragEnabled(True)
        self.setDragDropMode(QtWidgets.QAbstractItemView.DragDrop)
        self.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)

    def keyPressEvent(self, e):
        # Enter = chuna hua folder kholo / file dikhaao (SCAN nahi hona chahiye)
        if e.key() in (QtCore.Qt.Key_Return, QtCore.Qt.Key_Enter):
            idx = self.currentIndex()
            if self._on_activate is not None and idx.isValid():
                self._on_activate(idx)
                e.accept()
                return
        super().keyPressEvent(e)

    def dragEnterEvent(self, e):
        e.acceptProposedAction()

    def dragMoveEvent(self, e):
        e.acceptProposedAction()

    def dropEvent(self, e):
        try:
            self._on_drop(self.indexAt(e.pos()))
            e.acceptProposedAction()
        except Exception:
            e.ignore()


class PagesList(QtWidgets.QListWidget):
    """Pages ki central list. Do kaam karti hai:
      1. Andar-hi-andar pages ko kheench kar aage-peeche karo (reorder).
      2. Bahar se (Explorer/desktop) image/PDF ko SEEDHA is list par
         drag-drop karke import karo.
    Bahar ki files ka drop 'on_files' ko jaata hai; baaki sab (reorder) Qt
    ka apna InternalMove sambhalta hai."""

    IMPORT_EXTS = (".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".pdf")

    def __init__(self, on_files):
        super().__init__()
        self._on_files = on_files
        self.setAcceptDrops(True)

    def _dropped_files(self, e):
        md = e.mimeData()
        if not md.hasUrls():
            return []
        out = []
        for u in md.urls():
            p = u.toLocalFile()
            if p and p.lower().endswith(self.IMPORT_EXTS):
                out.append(p)
        return out

    def dragEnterEvent(self, e):
        if self._dropped_files(e):
            e.acceptProposedAction()
        else:
            super().dragEnterEvent(e)

    def dragMoveEvent(self, e):
        if self._dropped_files(e):
            e.acceptProposedAction()
        else:
            super().dragMoveEvent(e)

    def dropEvent(self, e):
        files = self._dropped_files(e)
        if files:
            e.acceptProposedAction()
            try:
                self._on_files(files)
            except Exception:
                pass
        else:
            super().dropEvent(e)      # andar-hi-andar reorder


class UrlListWidget(QtWidgets.QListWidget):
    """Search-results list — items ko FILE ki tarah kheencha ja sake, taaki
    search me mili file ko seedha doc-area me drag karke import kar sakein.
    Sirf asli files (folder-header nahi) drag hoti hain."""

    def __init__(self):
        super().__init__()
        self.setDragEnabled(True)
        self.setDragDropMode(QtWidgets.QAbstractItemView.DragOnly)
        self.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)

    def mimeData(self, items):
        md = QtCore.QMimeData()
        urls = []
        for it in items:
            p = it.data(QtCore.Qt.UserRole)
            if p and os.path.isfile(p):
                urls.append(QtCore.QUrl.fromLocalFile(p))
        if urls:
            md.setUrls(urls)
        return md


class ScannerWindow(QtWidgets.QMainWindow):
    THUMB_W = 150
    THUMB_H = 200

    def __init__(self, auto_scan_profile=None):
        super().__init__()
        self.setWindowTitle(APP_NAME)
        try:
            self.setWindowIcon(app_icon())
        except Exception:
            pass
        self.resize(1220, 820)
        self._tmpdir = tempfile.mkdtemp(prefix="nobledoc_")
        self._config = load_config()
        self._profiles = self._config.get("profiles", [])
        self._opts = dict(DEFAULT_OPTIONS)
        self._opts.update(self._config.get("options", {}))
        self._recent = self._config.get("recent", [])
        self._used_claims = set(self._config.get("used_claims", []))
        self._barcode_tried = False
        self._dirty = False
        self._undo_stack = []
        # Startup guard: app khulte hi (jaise app ko Enter dabakar launch karne par
        # wahi Enter naye window me 'Enter = Scan' chala deta tha) galti se scan na
        # ho — pehle ~900ms tak do_scan ignore hota hai.
        self._start_timer = QtCore.QElapsedTimer(); self._start_timer.start()
        self._lang = self._opts.get("language", "en")
        # One-time migration: the app now ships as an English UI. Older installs
        # may still carry a saved Hindi/Hinglish setting — flip them to English
        # once. Users can always switch back via Settings -> Language.
        if not self._opts.get("_english_migrated"):
            self._opts["_english_migrated"] = True
            self._opts["language"] = "en"
            self._lang = "en"
            try:
                self._save_opts()
            except Exception:
                pass
        # v94: naye keyboard shortcuts (F2=Rename, F3=150, F4=200, F5=300,
        # F6=600, F7=Custom). Purane saved keys hata do taaki naya default lage.
        if not self._opts.get("_sc_reset_v94"):
            self._opts["_sc_reset_v94"] = True
            scm = dict(self._opts.get("shortcuts", {}) or {})
            for _sid in ("scan", "rename", "dpi_150", "dpi_200", "dpi_300",
                         "dpi_600", "dpi_custom", "save_all", "save_sel"):
                scm.pop(_sid, None)
            self._opts["shortcuts"] = scm
            try:
                self._save_opts()
            except Exception:
                pass

        # Install ke baad ek default profile khud ban jaaye (DPI 150, colour, auto).
        self._ensure_default_profile()

        self._build_menu()
        self._build_ui()
        self._apply_style()
        self._refresh_profile_combo()
        self._load_profile_to_panel(self._selected_profile())
        self._refresh_method_label()
        self._refresh_recent_menu()
        self._update_status()

        QtCore.QTimer.singleShot(400, self.check_connection)
        self._conn_timer = QtCore.QTimer(self)
        self._conn_timer.setInterval(30000)
        self._conn_timer.timeout.connect(self.check_connection)
        self._conn_timer.start()

        if not self._config.get("setup_done"):
            QtCore.QTimer.singleShot(600, self._run_wizard)

        self._apply_simple_mode()

        if auto_scan_profile:
            i = self.cmb_profile.findText(auto_scan_profile)
            if i >= 0:
                self.cmb_profile.setCurrentIndex(i)
            QtCore.QTimer.singleShot(800, self.do_scan)

    # ---- config helpers ----
    def _save_opts(self):
        self._config["options"] = self._opts; save_config(self._config)

    def _save_profiles(self):
        self._config["profiles"] = self._profiles; save_config(self._config)

    def _add_recent(self, path):
        self._recent = [path] + [r for r in self._recent if r != path]
        self._recent = self._recent[:12]
        self._config["recent"] = self._recent; save_config(self._config)
        self._refresh_recent_menu()

    # ---- menu + shortcuts ----
    def _ma(self, menu, text, slot, tip="", sc=None):
        # add a menu action with a "?"-prefixed label + a Hindi/English hover tip
        a = menu.addAction(("\u2753 " + text) if tip else text, slot)
        if tip:
            a.setToolTip(tip); a.setStatusTip(tip)
        return a

    def _build_menu(self):
        mb = self.menuBar()
        mf = mb.addMenu(tr("menu_file", self._lang)); mf.setToolTipsVisible(True)
        self._ma(mf, tr("save_all", self._lang), self.save_pdf_all,
                 "हिन्दी: सभी पेजों की एक PDF बनाओ।\nEnglish: Save all pages as one PDF.", "Ctrl+S")
        self._ma(mf, tr("save_sel", self._lang), self.save_pdf_selected,
                 "हिन्दी: सिर्फ़ चुने हुए (Ctrl/Shift से) पेजों की PDF।\nEnglish: PDF of only the selected pages.")
        self._ma(mf, "Save PDF (password)…", self.save_pdf_password,
                 "हिन्दी: पासवर्ड से सुरक्षित PDF (खोलने के लिए पासवर्ड लगेगा)।\nEnglish: Password-protected PDF.")
        self._ma(mf, "Save Images…", self.save_images,
                 "हिन्दी: पेजों को JPG/PNG इमेज में सेव करो (PDF के बजाय)।\nEnglish: Save pages as JPG/PNG images.", "Ctrl+Shift+S")
        self._ma(mf, "Export OCR text…", self.export_ocr_text,
                 "हिन्दी: पेज के टेक्स्ट को पढ़कर .txt में निकालो (OCR)।\nEnglish: Extract page text to .txt via OCR.")
        pmenu = mf.addMenu("Print")
        pmenu.setToolTipsVisible(True)
        self._ma(pmenu, "All print (all pages)", self.print_all,
                 "हिन्दी: सारे पेज प्रिंट करो।\nEnglish: Print all pages.")
        self._ma(pmenu, "Selected print", self.print_selected,
                 "हिन्दी: सिर्फ़ चुने हुए (Ctrl/Shift से) पेज प्रिंट करो।\nEnglish: Print only the selected pages.")
        self._ma(pmenu, "ID print (2 IDs per page)", self.print_ids,
                 "हिन्दी: ID कार्ड को एक A4 पेज पर 2-2 करके प्रिंट करो (काग़ज़ बचेगा)।\nEnglish: Print ID cards two-per-A4-sheet to save paper.")
        self._ma(pmenu, "ID print - selected only", self.print_ids_selected,
                 "हिन्दी: सिर्फ़ चुने हुए ID को 2-per-page प्रिंट करो।\nEnglish: Print only selected IDs, two per sheet.")
        shmenu = mf.addMenu('Share')
        shmenu.setToolTipsVisible(True)
        self._ma(shmenu, "Send via WhatsApp…", self.share_whatsapp,
                 "हिन्दी: आख़िरी सेव की हुई PDF WhatsApp पर भेजो (फ़ाइल कॉपी हो जाती है, चैट में Ctrl+V या drag से attach)।\nEnglish: Send the last saved PDF via WhatsApp (file is copied; paste or drag into the chat).")
        self._ma(shmenu, "Send via Email…", self.share_email,
                 "हिन्दी: PDF को Email से भेजो (Outlook हो तो attachment के साथ draft खुलता है)।\nEnglish: Send the PDF by Email (opens an Outlook draft with the attachment if available).")
        self.recent_menu = mf.addMenu("Recent PDFs")
        self.recent_menu.setToolTipsVisible(True)
        mf.addSeparator()
        self._ma(mf, "Exit", self.close, "हिन्दी: ऐप बंद करो। (बिना सेव किए पेज हों तो चेतावनी आएगी।)\nEnglish: Close the app. (Warns if there are unsaved pages.)")

        me = mb.addMenu(tr("menu_edit", self._lang)); me.setToolTipsVisible(True)
        self._ma(me, "Rotate left", self.rotate_left, "हिन्दी: चुने हुए पेज को बाएँ घुमाओ।\nEnglish: Rotate the selected page left.")
        self._ma(me, "Rotate right", self.rotate_right, "हिन्दी: चुने हुए पेज को दाएँ घुमाओ।\nEnglish: Rotate the selected page right.")
        self._ma(me, "Brightness +", lambda: self._enhance_current(1.12, 1.0), "हिन्दी: पेज को हल्का (bright) करो।\nEnglish: Make the page brighter.")
        self._ma(me, "Brightness -", lambda: self._enhance_current(0.9, 1.0), "हिन्दी: पेज को गहरा (dim) करो।\nEnglish: Make the page darker.")
        self._ma(me, "Contrast +", lambda: self._enhance_current(1.0, 1.15), "हिन्दी: टेक्स्ट और साफ़/गहरा दिखे।\nEnglish: Increase contrast (sharper text).")
        self._ma(me, "Contrast -", lambda: self._enhance_current(1.0, 0.88), "हिन्दी: Contrast कम करो।\nEnglish: Decrease contrast.")
        self._ma(me, "Auto-crop page", self.autocrop_current, "हिन्दी: पेज के आस-पास की खाली border काटो।\nEnglish: Trim the empty border around the page.")
        self._ma(me, "Copy page text", self.copy_page_text, "हिन्दी: इस पेज का पूरा टेक्स्ट पढ़कर कॉपी कर लो (कहीं भी paste करो)।\nEnglish: OCR this page's text to the clipboard.")
        self._ma(me, "Translate page (Hindi ↔ English)…", self.translate_page, "हिन्दी: पेज का टेक्स्ट पढ़कर हिंदी/English में translate करो (इंटरनेट चाहिए)।\nEnglish: Translate the page's text between Hindi and English (needs internet).")
        self._ma(me, "Undo delete", self.undo_delete, "हिन्दी: ग़लती से delete हुआ पेज वापस लाओ।\nEnglish: Restore a deleted page.", "Ctrl+Z")
        me.addSeparator()
        self._ma(me, "Delete page", self.delete_page, "हिन्दी: चुने हुए पेज को हटाओ।\nEnglish: Delete the selected page.", "Delete")
        self._ma(me, "Clear all", self.clear_all, "हिन्दी: सारे पेज हटाओ (खाली करो)।\nEnglish: Remove all pages.")

        mt = mb.addMenu(tr("menu_tools", self._lang)); mt.setToolTipsVisible(True)
        self._ma(mt, "Bundle → separate PDFs (blank separator)…", self.bundle_split_save, "हिन्दी: पूरा bundle एक साथ feeder में डालो, documents के बीच खाली पेज रखो — यह खुद अलग-अलग PDF बना देगा। (Settings में 'blank हटाओ' बंद रखें।)\nEnglish: Scan a whole bundle with a blank page between documents; this splits and saves separate PDFs automatically.")
        self._ma(mt, "Book page → split into 2 pages", self.split_book_page, "हिन्दी: खुली किताब के स्कैन को बीच से काटकर दो अलग पेज बना दो।\nEnglish: Split an open-book scan into left and right pages.")
        self._ma(mt, "Business cards → contacts…", self.business_cards, "हिन्दी: विज़िटिंग कार्ड के स्कैन से नाम/फ़ोन/email पढ़कर contact फ़ाइलें (.vcf) + Excel बनाओ।\nEnglish: Read visiting cards into contact (.vcf) files and an Excel sheet.")
        self._ma(mt, "Restore old photo", self.restore_photo_current, "हिन्दी: फीकी/धुँधली पुरानी फ़ोटो का रंग-रूप सुधारो।\nEnglish: Restore faded/dull old photos.")
        self._ma(mt, "Scan History…", self.show_history, "हिन्दी: अब तक की सारी सेव की हुई PDF — नई से पुरानी, filter के साथ।\nEnglish: All saved PDFs, newest first, with quick filter.")
        self._ma(mt, "📊 Analytics…", self.show_analytics, "हिन्दी: आपकी + दुनिया भर की गिनती — कितने स्कैन, import, print (सिर्फ़ गिनती, कोई document नहीं)।\nEnglish: Your + worldwide counts — scans, imports, prints (counts only, no documents).")
        self._ma(mt, "📷 Scan from camera (webcam)…", self.scan_from_camera, "हिन्दी: स्कैनर न हो तो भी — webcam/USB कैमरा से document capture करके PDF बनाओ (फ़ोटो अपने-आप साफ़ होती है)।\nEnglish: No scanner? Capture documents with a webcam/USB camera (auto-cleaned).")
        self._ma(mt, "Phone photo to PDF (photo import)…", self.import_photos, "हिन्दी: फ़ोन से खींची document-फ़ोटो को साफ़ करके पेज बनाओ (परछाई हटाना, सीधा करना) — फिर PDF सेव करो।\nEnglish: Clean up phone photos of documents (remove shadows, straighten) and add them as pages.")
        self._ma(mt, "Split ID cards (from this page)…", self.split_id_cards, "हिन्दी: एक पेज पर 2-3 ID कार्ड स्कैन किए हैं? यह उन्हें अलग-अलग पेजों में काट देगा।\nEnglish: Scanned 2-3 ID cards on one page? This splits them into separate pages.")
        self._ma(mt, "Search past PDFs…", self.search_pdfs, "हिन्दी: पुरानी सेव की हुई PDF ढूँढो (claim/नाम/tag से, या PDF के अंदर के टेक्स्ट से)।\nEnglish: Search your saved PDFs by name, tag or text content.", "Ctrl+F")
        self._ma(mt, "Build/refresh search index…", self.build_search_index, "हिन्दी: सारी PDF का टेक्स्ट एक बार पढ़कर index बना लो — फिर अंदर-के-टेक्स्ट वाली search तुरंत होगी।\nEnglish: Build a one-time text index so in-PDF search becomes instant.")
        self._ma(mt, "Add tag (to a PDF)…", self.tag_pdf, "हिन्दी: PDF पर अपने tags लगाओ (जैसे Aadhaar, School, बिजली-बिल) — बाद में tag से तुरंत ढूँढो।\nEnglish: Put your own tags on a PDF for quick finding later.")
        self._ma(mt, "Find by tag…", self.search_by_tag, "हिन्दी: लगाए हुए tag से फ़ाइलों की सूची देखो और खोलो।\nEnglish: List and open files by tag.")
        self._ma(mt, "Merge PDFs…", self.merge_pdfs, "हिन्दी: कई PDF को जोड़कर एक PDF बनाओ।\nEnglish: Merge several PDFs into one.")
        self._ma(mt, "🗑 Recycle Bin…", self.show_recycle_bin, "हिन्दी: Delete की हुई फ़ाइलें यहाँ आती हैं — ग़लती से हटी फ़ाइल को वापस लाओ, या हमेशा के लिए हटाओ।\nEnglish: Deleted files go here — restore anything you removed by mistake, or delete forever.")
        self._ma(mt, "Split into multiple PDFs…", self.split_pdfs, "हिन्दी: एक स्कैन को कई अलग PDF में बाँटो।\nEnglish: Split into multiple PDFs.")
        self._ma(mt, "Compress PDF…", self.compress_pdf_tool, "हिन्दी: अभी के पेज या कोई पुरानी PDF को 200KB/500KB/1MB/2MB तक छोटा करो (portal upload के लिए)।\nEnglish: Shrink current pages or any PDF to a 200KB/500KB/1MB/2MB target for portal uploads.")
        pdft = mt.addMenu("PDF Tools")
        pdft.setToolTipsVisible(True)
        self._ma(pdft, "PDF page editor (reorder/rotate/delete)…", self.pdf_page_editor,
                 "हिन्दी: किसी भी PDF के पेजों का क्रम बदलो, घुमाओ या हटाओ — बिना quality ख़राब किए (lossless)।\nEnglish: Reorder, rotate or remove pages of any PDF, losslessly.")
        self._ma(pdft, "Place Sign/Stamp (on this page)…", self.place_sign,
                 "हिन्दी: अपने signature/मोहर की इमेज मौजूदा पेज पर लगाओ (सफ़ेद background अपने-आप पारदर्शी)।\nEnglish: Place your signature/stamp image on the current page (white background auto-transparent).")
        self._ma(pdft, "Add page numbers (on all pages)…", self.add_page_numbers,
                 "हिन्दी: हर पेज पर 'Page 1/5' और चाहें तो ऊपर अपना header छापो।\nEnglish: Print 'Page 1/5' on every page, with an optional header.")
        self._ma(pdft, "Watermark a PDF…", self.watermark_pdf_tool,
                 "हिन्दी: किसी पुरानी PDF पर अपना text-watermark/स्टांप छापो।\nEnglish: Stamp a text watermark onto any existing PDF.")
        self._ma(pdft, "Remove PDF password…", self.remove_pdf_password,
                 "हिन्दी: पासवर्ड पता हो तो उस PDF की बिना-पासवर्ड कॉपी बनाओ।\nEnglish: If you know the password, make a password-free copy of the PDF.")
        self._ma(pdft, "PDF → Word (.docx)…", self.pdf_to_word,
                 "हिन्दी: PDF/पेजों का टेक्स्ट OCR करके Word फ़ाइल बनाओ (edit करने लायक)।\nEnglish: OCR the text into an editable Word document.")
        self._ma(pdft, "PDF → Excel (.xlsx)…", self.pdf_to_excel,
                 "हिन्दी: bill/table वाले पेजों को OCR करके Excel में निकालो (best-effort)।\nEnglish: Extract bill/table pages into an Excel sheet (best-effort).")
        self._ma(pdft, "PDF → JPG images…", self.pdf_to_jpgs,
                 "हिन्दी: किसी PDF के हर पेज को अलग JPG इमेज में निकालो।\nEnglish: Export each page of a PDF as a separate JPG image.")
        self._ma(pdft, "Folder images → one PDF…", self.folder_to_pdf,
                 "हिन्दी: एक फ़ोल्डर की सारी images (नाम के क्रम में) एक PDF में जोड़ो।\nEnglish: Combine all images in a folder (name order) into one PDF.")
        self._ma(pdft, "Archival PDF (300dpi + metadata)…", self.save_archival_pdf,
                 "हिन्दी: High-quality PDF (300dpi) पूरे metadata (title/date) के साथ — लंबे समय तक सँभालने के लिए।\nEnglish: High-quality 300dpi PDF with full metadata for long-term archiving.")
        self._ma(mt, "Monthly report…", self.monthly_report, "हिन्दी: महीने का स्कैन/claim report बनाओ।\nEnglish: Generate a monthly report.")
        self._ma(mt, "Create desktop shortcut…", self.create_shortcut, "हिन्दी: Desktop पर एक-क्लिक स्कैन का shortcut बनाओ।\nEnglish: Make a one-click desktop scan shortcut.")
        self._ma(mt, "Auto-name pages (document name)", self.auto_name_pages, "हिन्दी: हर पेज को पढ़कर उसका नाम (जैसे DISCHARGE SUMMARY, RECEIPT) थंबनेल के नीचे लिखे। 'Page 1,2' के बजाय असली नाम।\nEnglish: Read each page and label it with its document title instead of 'Page 1,2'.")
        self._ma(mt, "Learned names (manage)…", self.manage_learned_names, "हिन्दी: आपने F2 से जो नाम सिखाए हैं उन्हें देखो/बदलो/हटाओ। एक बार नाम सिखाने पर अगली बार वही document अपने-आप उसी नाम से आता है।\nEnglish: View/edit/remove the names you taught with F2. Once taught, the same document auto-names itself next time.")

        ms = mb.addMenu(tr("menu_settings", self._lang)); ms.setToolTipsVisible(True)
        self._ma(ms, tr("options", self._lang), self.open_options, "हिन्दी: ऐप की सारी settings (auto-save, blank हटाओ, backup, वग़ैरह)।\nEnglish: All app settings.")
        self._ma(ms, tr("profiles", self._lang), self.open_profiles, "हिन्दी: स्कैन profiles बनाओ/बदलो (device, dpi, colour, duplex)।\nEnglish: Create/edit scan profiles.")
        self._ma(ms, tr("scan_method", self._lang) + "…", self.choose_scan_method, "हिन्दी: स्कैन का तरीका: escl (network duplex), twain (USB duplex), या wia।\nEnglish: Scan method: escl (network duplex), twain (USB), or wia.")
        self._ma(ms, tr("language", self._lang) + "…", self.choose_language, "हिन्दी: ऐप की भाषा बदलो (Hindi/English)।\nEnglish: Change the app language.")
        self._ma(ms, "🔍 Scanner auto-detect (LAN + USB)…", self.auto_detect_scanner, "हिन्दी: स्कैनर खुद पहचानो — LAN (network) पर है या USB पर, दोनों ढूँढकर सबसे बेहतर चुन लेता है। कुछ सोचना नहीं पड़ता।\nEnglish: Auto-detect the scanner — finds it on LAN or USB automatically and picks the best.")
        self._ma(ms, "Find scanner (network only)…", self.find_scanners, "हिन्दी: सिर्फ़ network (eSCL) पर स्कैनर ढूँढो।\nEnglish: Discover only network (eSCL) scanners.")
        self._ma(ms, "Scanner IP…", self.set_scanner_ip, "हिन्दी: network स्कैनर का IP सेट करो (जैसे 192.168.1.8)।\nEnglish: Set the network scanner IP.")
        self._ma(ms, "Keyboard Shortcuts…", self.show_shortcuts, "हिन्दी: कीबोर्ड के shortcuts की सूची देखो।\nEnglish: View keyboard shortcuts.")
        self.act_simple = self._ma(ms, tr("simple_on", self._lang), self.toggle_simple_mode, "हिन्दी: Simple mode: सिर्फ़ ज़रूरी buttons दिखें (नए users के लिए आसान)।\nEnglish: Simple mode: show only the essential buttons.")
        self.act_simple.setCheckable(True)
        self.act_simple.setChecked(bool(self._opts.get("simple_mode")))
        self._ma(ms, self.L("Left sidebar dikhao/chhupao", "Show/hide left sidebar"), self.toggle_left_panel, "हिन्दी: बाईं तरफ़ का scan-settings panel चालू/बंद (ज़्यादा जगह के लिए)।\nEnglish: Show/hide the left scan-settings sidebar for more space.", "F9")
        self.act_files_panel = self._ma(ms, self.L("Right sidebar (Meri Files) dikhao/chhupao", "Show/hide right sidebar (My Files)"), self.toggle_files_panel, "हिन्दी: दाईं तरफ़ का folders वाला panel चालू/बंद करो।\nEnglish: Show/hide the right-side files panel.", "F10")
        self._ma(ms, "🎨 Customize UI…", self.customize_ui, "हिन्दी: ऐप का लुक अपने हिसाब से: dashboard, status-पट्टी, sidebar graph, Dark Pro theme — जो चाहो चालू/बंद करो।\nEnglish: Customize the UI: dashboard, status bar, sidebar graph, Dark Pro theme.")
        self.act_touch = self._ma(ms, "Touch / large-button mode", self.toggle_touch_mode, "हिन्दी: buttons/लिखाई बड़ी हो जाएगी — touch screen या बुज़ुर्गों के लिए आसान।\nEnglish: Bigger buttons and text for touch screens or elderly users.")
        self.act_touch.setCheckable(True)
        self.act_touch.setChecked(bool(self._opts.get("touch_mode")))
        ms.addSeparator()
        self._ma(ms, "Export settings…", self.export_settings, "हिन्दी: सारी settings एक फ़ाइल में — नए PC पर ले जाने के लिए।\nEnglish: Export all settings to a file for another PC.")
        self._ma(ms, "Import settings…", self.import_settings, "हिन्दी: Export की हुई settings फ़ाइल से सब वापस ले आओ।\nEnglish: Import settings from an exported file.")

        mh = mb.addMenu(tr("menu_help", self._lang)); mh.setToolTipsVisible(True)
        self._ma(mh, "📖 Complete Guide (all options)…", self.show_guide, "हिन्दी: पूरे software की complete guide — हर option कहाँ है और क्या करता है (Hindi + English)। यह सूची ऐप के menus से खुद बनती है, इसलिए हर update में अपने-आप up-to-date रहती है।\nEnglish: The complete guide — every option, where it is and what it does (Hindi + English). Built automatically from the app's menus, so it stays up to date on every update.", "F1")
        self._ma(mh, tr("help_guide", self._lang), self.show_help, "हिन्दी: ऐप इस्तेमाल करने की guide।\nEnglish: How-to guide.")
        self._ma(mh, "Setup wizard", self._run_wizard, "हिन्दी: पहली बार वाला setup दोबारा चलाओ।\nEnglish: Re-run the first-time setup.")
        self._ma(mh, tr("whatsnew", self._lang), self.show_whatsnew, "हिन्दी: नए बदलाव/features।\nEnglish: What's new.")
        self._ma(mh, "Test / Diagnostics", self.run_diagnostics, "हिन्दी: स्कैनर/ऐप की जानकारी + error report (share करने के लिए)।\nEnglish: Scanner/app info + error report.")
        self._ma(mh, "Duplex Test (both-side)", self.run_duplex_test, "हिन्दी: जाँचो कि दोनों तरफ़ (duplex) स्कैन हो रहा है या नहीं।\nEnglish: Test whether both-side (duplex) scanning works.")
        self._ma(mh, "eSCL Test (network scan check)", self.run_escl_test, "हिन्दी: network scan (eSCL) को step-by-step जाँचकर असली समस्या बताता है (connect / status / job)।\nEnglish: Step-by-step eSCL network-scan test that shows the real problem.")
        self._ma(mh, "Check for updates…", lambda: self.check_updates(False), "हिन्दी: नया version आया हो तो ऐप उसे खुद download करके install कर लेगी।\nEnglish: If a newer version exists the app downloads and installs it itself.")
        self._ma(mh, "View error report…", self.open_crash_report, "हिन्दी: अगर ऐप कभी crash हुई हो तो उसकी report खोलो (feedback में भेजने के लिए)।\nEnglish: Open the saved crash report, if any.")
        self._ma(mh, tr("feedback", self._lang), self.send_feedback, "हिन्दी: सुझाव/शिकायत भेजो।\nEnglish: Send feedback.")
        mh.addSeparator()
        self._ma(mh, "📣 Share ApneScan (tell friends)…", self.share_app, "हिन्दी: इस free ऐप को दोस्तों/customers तक पहुँचाओ — WhatsApp, link copy, QR code या poster/pamphlet (दुकान/अस्पताल में लगाने लायक)। जितने ज़्यादा लोग, उतना अच्छा।\nEnglish: Spread this free app — WhatsApp, copy link, QR code, or a printable poster for your shop/clinic.")
        self._ma(mh, "⭐ Review / Star (GitHub)…", self.ask_review, "हिन्दी: पसंद आया तो GitHub पर ⭐ star या review दें — इससे ऐप और लोगों तक पहुँचेगा।\nEnglish: Like it? Give a ⭐ or a review on GitHub — it helps others find ApneScan.")
        self._ma(mh, tr("about", self._lang), self.show_about, "हिन्दी: ऐप के बारे में।\nEnglish: About this app.")

        # AUTOMATIC: har menu/submenu ke har option par "?" + Hindi/English hover
        # PAKKA karo — aage koi naya option seedhe addAction se bhi joda jaye to
        # bhi ye use chhoot-ne nahi dega (chhupa hua safety-jaal).
        self._finalize_menu_help()
        self._build_shortcuts()

    def _finalize_menu_help(self):
        """Poore menu-tree ko scan karke jis bhi option par abhi tak '?' /
        tooltip nahi hai, uspar khud laga do. Ye function baar-baar chalana
        safe hai (dobara prefix nahi lagta)."""
        MARK = "❓ "

        def walk(menu):
            try:
                menu.setToolTipsVisible(True)
            except Exception:
                pass
            for act in menu.actions():
                if act.isSeparator():
                    continue
                sub = act.menu()
                if sub is not None:
                    walk(sub)
                    continue
                t = act.text()
                if not t:
                    continue
                clean = t[len(MARK):] if t.startswith(MARK) else t
                if not t.startswith(MARK):
                    act.setText(MARK + t)
                # tooltip na ho (ya sirf label ki copy ho) to ek default bilingual
                # explanation laga do
                if not act.toolTip() or act.toolTip() in (t, clean):
                    label = clean.replace("&", "").replace("…", "").strip()
                    act.setToolTip("%s — runs this menu option.\n"
                                   "Runs the \"%s\" action." % (label, label))
        try:
            for act in self.menuBar().actions():
                if act.menu() is not None:
                    walk(act.menu())
        except Exception:
            pass

    def _refresh_recent_menu(self):
        self.recent_menu.clear()
        if not self._recent:
            a = self.recent_menu.addAction("(none)"); a.setEnabled(False); return
        for path in self._recent:
            self.recent_menu.addAction(os.path.basename(path),
                                       lambda checked=False, p=path: self._open_path(p))

    def _open_path(self, path):
        try:
            os.startfile(path)
        except Exception as exc:
            self._warn("Could not open file:\n%s" % exc)

    # ---- profiles / options ----
    def _ensure_default_profile(self):
        """Install ke baad (jab koi profile na ho) ek 'Default' profile khud
        bana do: DPI 150, 24-bit colour, page-size Auto, single side. Device
        pehli-baar auto-detect se aata hai (auto_detect_scanner)."""
        if self._profiles:
            return
        self._profiles = [{
            "name": "Default",
            "dpi": 150,
            "color": "color",
            "duplex": False,
            "page_size": "auto",
            "paper_source": "feeder",
            "source_name": self._opts.get("scanner_name", ""),
        }]
        self._config["profiles"] = self._profiles
        self._config["selected_profile"] = "Default"
        try:
            save_config(self._config)
        except Exception:
            pass

    def _refresh_profile_combo(self):
        self.cmb_profile.blockSignals(True); self.cmb_profile.clear()
        for p in self._profiles:
            self.cmb_profile.addItem(p.get("name", "Profile"))
        sel = self._config.get("selected_profile")
        if sel:
            i = self.cmb_profile.findText(sel)
            if i >= 0:
                self.cmb_profile.setCurrentIndex(i)
        self.cmb_profile.blockSignals(False)

    def _selected_profile(self):
        name = self.cmb_profile.currentText()
        for p in self._profiles:
            if p.get("name") == name:
                return p
        return None

    def _on_fast_toggled(self, on):
        self._opts["fast_mode"] = bool(on); self._save_opts()
        if on:
            # reflect fast settings in the panel so the user sees them
            self.cmb_dpi.setCurrentText("200 dpi")
            self.cmb_depth.setCurrentText("Black & White")
        self.status.showMessage("Fast mode ON (200 dpi, B&W)" if on else "Fast mode OFF", 3000)

    def _on_profile_changed(self, name):
        self._config["selected_profile"] = name; save_config(self._config)
        self._load_profile_to_panel(self._selected_profile())


    def open_profiles(self):
        dlg = ProfileManagerDialog(self, self._profiles)
        if dlg.exec_() == QtWidgets.QDialog.Accepted:
            self._profiles = dlg.profiles; self._save_profiles(); self._refresh_profile_combo()

    def open_options(self):
        dlg = OptionsDialog(self, self._opts)
        if dlg.exec_() == QtWidgets.QDialog.Accepted:
            self._opts = dlg.get_opts(); self._save_opts(); self._update_status()
            self._apply_style(); self._refresh_page_numbers()
            self._refresh_files_root()   # save-folder badla ho to panel bhi

    def _refresh_page_numbers(self):
        show = self._opts.get("show_page_numbers", True)
        for i in range(self.list.count()):
            it = self.list.item(i)
            title = it.data(TITLE_ROLE)
            if title:
                it.setText(title)
            else:
                it.setText(("Page %d" % (i + 1)) if show else "")

    def _vsep(self):
        f = QtWidgets.QFrame(); f.setObjectName("hr"); f.setFrameShape(QtWidgets.QFrame.VLine); return f

    def _on_panel_dpi_changed(self, txt):
        """'Custom…' chunne par apni dpi pucho aur list me jod do."""
        if not txt.startswith("Custom"):
            return
        n, ok = QtWidgets.QInputDialog.getInt(
            self, self.L("Apni DPI", "Custom DPI"),
            self.L("DPI likho (50–1200):", "Enter DPI (50–1200):"), 300, 50, 1200, 50)
        item = ("%d dpi" % n)
        self.cmb_dpi.blockSignals(True)
        if ok:
            if self.cmb_dpi.findText(item) < 0:
                self.cmb_dpi.insertItem(self.cmb_dpi.count() - 1, item)
            self.cmb_dpi.setCurrentText(item)
        else:
            self.cmb_dpi.setCurrentText("200 dpi")
        self.cmb_dpi.blockSignals(False)

    def _panel_scan_params(self, prof):
        try:
            dpi = int(self.cmb_dpi.currentText().split()[0])
        except Exception:
            dpi = int(prof.get("dpi", 200))
        d = self.cmb_depth.currentText()
        color = "color" if d.startswith("24") else ("gray" if d.startswith("Gray") else "bw")
        duplex = "Both" in self.cmb_sides.currentText()
        return dpi, color, duplex

    def _profile_scan_params(self, prof, dpi_override=None):
        """Scan settings SIRF profile se — panel ke manual badlaav se nahi. Isi
        liye Enter/Scan hamesha profile ki saved setting par scan karta hai.
        dpi_override diya ho (DPI shortcut se) to sirf us ek scan ke liye dpi
        badal jata hai — panel/profile nahi badalta."""
        prof = prof or {}
        try:
            dpi = int(dpi_override) if dpi_override else int(prof.get("dpi", 150))
        except Exception:
            dpi = 150
        color = prof.get("color", "color")
        if color not in ("color", "gray", "bw"):
            color = "color"
        duplex = bool(prof.get("duplex"))
        page_size = (prof.get("page_size") or "auto").lower()
        source = (prof.get("paper_source") or "feeder").lower()
        return dpi, color, duplex, page_size, source

    def _refresh_method_label(self):
        if not hasattr(self, "method_lbl"):
            return
        m = self._opts.get("scanner_method", "twain")
        color = "#0f766e" if m in ("wia", "naps2", "escl") else "#b45309"
        self.method_lbl.setText('<b>Connected via:</b> <span style="color:%s">%s</span>'
                                % (color, m.upper()))

    def _load_profile_to_panel(self, prof):
        if prof is None or not hasattr(self, "cmb_dpi"):
            return
        self.dev_lbl.setText(prof.get("source_name")
                             or self._opts.get("scanner_name") or "(no device)")
        _dpitem = str(prof.get("dpi", 200)) + " dpi"
        if self.cmb_dpi.findText(_dpitem) < 0:           # custom dpi — list me jodo
            self.cmb_dpi.blockSignals(True)
            self.cmb_dpi.insertItem(max(0, self.cmb_dpi.count() - 1), _dpitem)
            self.cmb_dpi.blockSignals(False)
        self.cmb_dpi.setCurrentText(_dpitem)
        c = prof.get("color", "gray")
        self.cmb_depth.setCurrentText("24-bit Colour" if c == "color"
                                      else ("Grayscale" if c == "gray" else "Black & White"))
        self.cmb_source.setCurrentText("Feeder (ADF)")
        self.cmb_sides.setCurrentText("Both sides (duplex)" if prof.get("duplex") else "Single side")
        # page size (per-profile)
        if hasattr(self, "cmb_pagesize"):
            _ps = (prof.get("page_size") or "auto").lower()
            for k in range(self.cmb_pagesize.count()):
                if self.cmb_pagesize.itemText(k).lower().startswith(_ps[:4]):
                    self.cmb_pagesize.setCurrentIndex(k); break

    def _quick_new_profile(self):
        dlg = EditProfileDialog(self)
        if dlg.exec_() == QtWidgets.QDialog.Accepted:
            self._profiles.append(dlg.get_profile()); self._save_profiles(); self._refresh_profile_combo()
            self.cmb_profile.setCurrentIndex(self.cmb_profile.count() - 1)

    def _quick_edit_profile(self):
        name = self.cmb_profile.currentText()
        idx = next((k for k, p in enumerate(self._profiles) if p.get("name") == name), -1)
        if idx < 0:
            self._quick_new_profile(); return
        dlg = EditProfileDialog(self, self._profiles[idx])
        if dlg.exec_() == QtWidgets.QDialog.Accepted:
            self._profiles[idx] = dlg.get_profile(); self._save_profiles(); self._refresh_profile_combo()
            i = self.cmb_profile.findText(self._profiles[idx].get("name"))
            if i >= 0:
                self.cmb_profile.setCurrentIndex(i)
            self._load_profile_to_panel(self._selected_profile())

    def _open_preview_dialog(self, item):
        if item is None:
            return
        row = self.list.row(item)
        PreviewDialog(self, row).exec_()

    def auto_name_pages(self):
        if not tesseract_available():
            self._warn(
                "To read a document's name you must install the 'Tesseract OCR' program.\n\n"
                "Download (free): https://github.com/UB-Mannheim/tesseract/wiki\n"
                "Install it and reopen ApneScan. Then Tools \u2192 Auto-name pages will work."); return
        items = [(i, self.list.item(i).data(QtCore.Qt.UserRole)) for i in range(self.list.count())]
        if not items:
            self._warn("Scan or import a page first."); return
        self._named_count = 0
        self._named_total = len(items)
        self.status.showMessage("Reading document names... (please wait)", 0)
        self._namer = NameWorker(items, learned=self._learned_names())
        self._namer.named.connect(self._apply_page_title)
        self._namer.finished_all.connect(self._on_naming_done)
        self._namer.start()

    def _name_one_page(self, row, path, force=False):
        """Ek page ka naam TURANT background me padho (scan/import ke saath-saath).
        Isse sab pages ke baad naming ka intezaar nahi karna padta."""
        if not path or not tesseract_available():
            return
        if not force and not self._opts.get("auto_name"):
            return
        self._page_namers = getattr(self, "_page_namers", [])
        w = NameWorker([(row, path)], learned=self._learned_names())
        w.named.connect(self._apply_page_title)
        w.finished_all.connect(
            lambda w=w: self._page_namers.remove(w) if w in self._page_namers else None)
        self._page_namers.append(w)
        w.start()

    def _learned_names(self):
        out = []
        for e in (self._config.get("learned_names", []) or []):
            try:
                out.append((set(e.get("words", [])), e.get("name", "")))
            except Exception:
                pass
        return out

    def _apply_page_title(self, row, title):
        if 0 <= row < self.list.count():
            it = self.list.item(row)
            key = name_key(title)
            # if this same document was saved before, reuse THAT saved name
            remembered = (self._config.get("doc_names", {}) or {}).get(key)
            label = remembered or underscore_name(title)
            # #5: naam me document-number aur/ya date apne aap jodo
            try:
                path = it.data(QtCore.Qt.UserRole)
                extra = []
                if self._opts.get("name_append_number"):
                    num = extract_doc_number(page_ocr_text(path, 0.75))
                    if num:
                        extra.append(num)
                if self._opts.get("name_append_date"):
                    extra.append(datetime.datetime.now().strftime("%Y-%m-%d"))
                if extra:
                    label = underscore_name(label + "_" + "_".join(extra))
            except Exception:
                pass
            it.setData(TITLE_ROLE, label)
            it.setData(NAMEKEY_ROLE, key)
            # naam khaali reh gaya to bhi kuch dikhe — 'Page N' fallback
            it.setText(label if (label and label.strip())
                       else ("Page %d" % (row + 1)))
            self._named_count = getattr(self, "_named_count", 0) + 1
            self._pstats_bump(ocr_named=1)
            # #2: is naam ke liye folder yaad ho to save wahin default ho
            self._apply_folder_hint_for(title)

    def _apply_folder_hint_for(self, title):
        """Seekhe naam ke saath jo folder yaad hai (agar hai) use save-hint bana do."""
        try:
            base = underscore_name(title).lower()
            for e in (self._config.get("learned_names", []) or []):
                if e.get("folder") and underscore_name(e.get("name", "")).lower() in base:
                    if os.path.isdir(e["folder"]):
                        self._doc_folder_hint = e["folder"]
                        return
        except Exception:
            pass

    def _on_naming_done(self):
        got = getattr(self, "_named_count", 0)
        total = getattr(self, "_named_total", 0)
        if got == 0:
            self.status.clearMessage()
            self._warn(
                "OCR ran, but no page name could be found.\n\n"
                "There can be 2 reasons:\n"
                "1) Tesseract was not found properly \u2014 open Help \u2192 Test/Diagnostics and "
                "check the 'Tesseract version' line (says WORKS or NOT FOUND).\n"
                "2) There is no clearly printed heading at the TOP of the page (e.g. handwritten pages) \u2014 "
                "such pages don't get a name, which is normal.")
        else:
            self.status.showMessage("Found names for %d / %d pages." % (got, total), 6000)

    def _build_shortcuts(self):
        methods = {
            "scan": self.do_scan,
            "dpi_150": lambda: self._scan_at_dpi(150),
            "dpi_200": lambda: self._scan_at_dpi(200),
            "dpi_300": lambda: self._scan_at_dpi(300),
            "dpi_600": lambda: self._scan_at_dpi(600),
            "dpi_custom": self._scan_at_dpi_custom,
            "import": self.import_images,
            "rename": self.rename_current_page,
            "save_all": self.save_pdf_all,
            "save_sel": self.save_pdf_selected,
            "save_pw": self.save_pdf_password,
            "save_img": self.save_images,
            "ocr_text": self.export_ocr_text,
            "print": self.print_pages,
            "rotate_left": self.rotate_left,
            "rotate_right": self.rotate_right,
            "bright_up": lambda: self._enhance_current(1.12, 1.0),
            "bright_dn": lambda: self._enhance_current(0.9, 1.0),
            "contrast_up": lambda: self._enhance_current(1.0, 1.15),
            "contrast_dn": lambda: self._enhance_current(1.0, 0.88),
            "autocrop": self.autocrop_current,
            "autoname": self.auto_name_pages,
            "undo": self.undo_delete,
            "delete": self.delete_page,
            "clear": self.clear_all,
            "move_up": self.move_up,
            "move_down": self.move_down,
            "search": self.search_pdfs,
            "merge": self.merge_pdfs,
            "split": self.split_pdfs,
            "report": self.monthly_report,
            "zoom_in": lambda: self._zoom_thumbs(1.15),
            "zoom_out": lambda: self._zoom_thumbs(0.87),
            "options": self.open_options,
            "profiles": self.open_profiles,
        }
        self._sc = {}
        custom = self._opts.get("shortcuts", {}) or {}
        for sid, label, default in SHORTCUTS:
            fn = methods.get(sid)
            if fn is None:
                continue
            scut = QtWidgets.QShortcut(self)
            scut.activated.connect(fn)
            key = custom.get(sid, default)
            if key:
                scut.setKey(QtGui.QKeySequence(key))
            self._sc[sid] = scut
        # Enter/Return se Scan — LEKIN 'Meri Files' panel me focus ho to nahi
        # (wahan Enter = folder kholo). Focus badalte hi scan-shortcut on/off.
        try:
            app = QtWidgets.QApplication.instance()
            if app is not None:
                app.focusChanged.connect(self._on_focus_changed_scan)
        except Exception:
            pass
        QtCore.QTimer.singleShot(0, self._refresh_shortcut_line)

    def _on_focus_changed_scan(self, _old, now):
        """Meri Files panel me focus ho to Enter=Scan band, warna chalu."""
        sc = self._sc.get("scan")
        if sc is None:
            return
        try:
            fp = getattr(self, "files_panel", None)
            in_panel = bool(fp and now is not None and (now is fp or fp.isAncestorOf(now)))
            sc.setEnabled(not in_panel)
        except Exception:
            pass

    def _set_panel_dpi(self, n):
        """Resolution combo ko is dpi par set karo (shortcut se)."""
        if not hasattr(self, "cmb_dpi"):
            return
        item = "%d dpi" % n
        self.cmb_dpi.blockSignals(True)
        if self.cmb_dpi.findText(item) < 0:
            self.cmb_dpi.insertItem(max(0, self.cmb_dpi.count() - 1), item)
        self.cmb_dpi.setCurrentText(item)
        self.cmb_dpi.blockSignals(False)
        try:
            self.status.showMessage(self.L("Resolution: %d dpi", "Resolution: %d dpi") % n, 3000)
        except Exception:
            pass

    def _set_panel_dpi_custom(self):
        n, ok = QtWidgets.QInputDialog.getInt(
            self, self.L("Apni DPI", "Custom DPI"),
            self.L("DPI likho (50–1200):", "Enter DPI (50–1200):"), 300, 50, 1200, 50)
        if ok:
            self._set_panel_dpi(n)

    def _scan_at_dpi(self, n):
        """DPI shortcut: profile ki baaki setting + is dpi par ek scan.
        Panel/profile me kuch nahi badalta (sirf yeh ek scan us dpi par)."""
        self.do_scan(dpi_override=n)

    def _scan_at_dpi_custom(self):
        """F7: apni DPI poochho, phir usi par ek scan (panel nahi badalta)."""
        n, ok = QtWidgets.QInputDialog.getInt(
            self, self.L("Apni DPI", "Custom DPI"),
            self.L("DPI likho (50–1200):", "Enter DPI (50–1200):"), 300, 50, 1200, 50)
        if ok:
            self.do_scan(dpi_override=n)

    def _refresh_shortcut_line(self):
        """Toolbar ke neeche wali line — abhi ke (user ke) shortcuts dikhao."""
        if not hasattr(self, "lbl_shortcuts"):
            return
        custom = self._opts.get("shortcuts", {}) or {}

        def k(sid, default):
            key = custom.get(sid, default)
            return "—" if not key else key.replace("Return", "Enter")
        left = (("rename", "F2", "Rename"), ("dpi_150", "F3", "Scan 150dpi"),
                ("dpi_200", "F4", "Scan 200dpi"), ("dpi_300", "F5", "Scan 300dpi"),
                ("dpi_600", "F6", "Scan 600dpi"), ("dpi_custom", "F7", "Scan Custom"))
        right = (("scan", "Return", "Scan"), ("save_sel", "Space", "Selected save"),
                 ("save_all", "Ctrl+S", "Save all"))
        p1 = ["<b>%s</b> = %s" % (k(sid, dflt), lbl) for sid, dflt, lbl in left]
        p2 = ["<b>%s</b> = %s" % (k(sid, dflt), lbl) for sid, dflt, lbl in right]
        self.lbl_shortcuts.setText(
            "⌨&nbsp; " + "  ·  ".join(p1)
            + "&nbsp;&nbsp;|&nbsp;&nbsp;" + "  ·  ".join(p2))

    def _apply_shortcut(self, sid, key):
        scut = self._sc.get(sid)
        if scut is not None:
            scut.setKey(QtGui.QKeySequence(key) if key else QtGui.QKeySequence())
        self._refresh_shortcut_line()      # neeche wali line bhi update

    def _cur_shortcut(self, sid, default):
        return (self._opts.get("shortcuts", {}) or {}).get(sid, default)

    def show_shortcuts(self):
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Keyboard Shortcuts"); dlg.resize(560, 560)
        lay = QtWidgets.QVBoxLayout(dlg)
        lay.addWidget(QtWidgets.QLabel(
            "To change a shortcut: pick a row → press the new key in the 'New' box → Assign."))
        tbl = QtWidgets.QTableWidget(len(SHORTCUTS), 2)
        tbl.setHorizontalHeaderLabels(["Action", "Shortcut"])
        tbl.verticalHeader().setVisible(False)
        tbl.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        tbl.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        tbl.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        tbl.horizontalHeader().setStretchLastSection(True); tbl.setColumnWidth(0, 300)

        def refill():
            for r, (sid, label, default) in enumerate(SHORTCUTS):
                tbl.setItem(r, 0, QtWidgets.QTableWidgetItem(label))
                tbl.setItem(r, 1, QtWidgets.QTableWidgetItem(self._cur_shortcut(sid, default) or "—"))
        refill()
        lay.addWidget(tbl, 1)

        row = QtWidgets.QHBoxLayout()
        kseq = QtWidgets.QKeySequenceEdit()
        row.addWidget(QtWidgets.QLabel("New:")); row.addWidget(kseq, 1)
        b_assign = QtWidgets.QPushButton("Assign"); b_unassign = QtWidgets.QPushButton("Unassign")
        row.addWidget(b_assign); row.addWidget(b_unassign)
        lay.addLayout(row)

        def sel():
            r = tbl.currentRow()
            return (SHORTCUTS[r][0], r) if r >= 0 else (None, -1)

        def do_assign():
            sid, r = sel()
            if sid is None:
                return
            key = kseq.keySequence().toString()
            if not key:
                return
            sm = self._opts.setdefault("shortcuts", {})
            for osid, olabel, odef in SHORTCUTS:      # clear conflicts
                if osid != sid and self._cur_shortcut(osid, odef) == key:
                    sm[osid] = ""; self._apply_shortcut(osid, "")
            sm[sid] = key; self._apply_shortcut(sid, key); self._save_opts()
            kseq.clear(); refill()

        def do_unassign():
            sid, r = sel()
            if sid is None:
                return
            self._opts.setdefault("shortcuts", {})[sid] = ""
            self._apply_shortcut(sid, ""); self._save_opts(); refill()

        b_assign.clicked.connect(do_assign); b_unassign.clicked.connect(do_unassign)

        bottom = QtWidgets.QHBoxLayout()
        b_reset = QtWidgets.QPushButton("Restore Defaults")

        def do_reset():
            self._opts["shortcuts"] = {}
            for sid, label, default in SHORTCUTS:
                self._apply_shortcut(sid, default)
            self._save_opts(); refill()
        b_reset.clicked.connect(do_reset)
        b_ok = QtWidgets.QPushButton("OK"); b_ok.clicked.connect(dlg.accept)
        bottom.addWidget(b_reset); bottom.addStretch(1); bottom.addWidget(b_ok)
        lay.addLayout(bottom)
        dlg.exec_()

    def _stats_url(self):
        if "stats_url" in self._config:
            return self._config["stats_url"]      # user override (may be "" to disable)
        return DEFAULT_STATS_URL

    def _get_client_id(self):
        cid = self._config.get("client_id")
        if not cid:
            import uuid
            cid = uuid.uuid4().hex[:16]
            self._config["client_id"] = cid
            try:
                save_config(self._config)
            except Exception:
                pass
        return cid

    def _set_stats_display(self, stats):
        if stats and isinstance(stats, tuple) and len(stats) == 3:
            ws = getattr(self, "_world_stats", {}) or {}
            ws.update({"total": stats[0], "today": stats[1], "online": stats[2]})
            self._world_stats = ws
        self._update_sidebar_stats()

    def _stats_failed(self):
        # network na ho to bhi personal stats dikhti rahein
        self._update_sidebar_stats()

    # ================= STATS (naya, seedha-saada) =================
    # NOTE: Pehle alag 'StatsWorker' (QThread + custom signals) istemal hota
    # tha — kisi wajah se uska signal sidebar tak nahi pahunchta tha aur
    # worldwide numbers '...' par atke reh jate the (jabki network theek tha).
    # Ab wahi PROVEN background system (FuncWorker, jo save/edit me chalta hai)
    # se fetch karte hain — ye bharosemand hai.

    def _run_bg_quiet(self, fn, on_done):
        """_run_bg jaisa hi — par bina 'busy' pill ke (background stats/poll ke
        liye). FuncWorker ko list me rakhta hai taaki GC na ho."""
        self._bg_workers = getattr(self, "_bg_workers", [])
        w = FuncWorker(fn)

        def _fin(res, w=w):
            try:
                self._bg_workers.remove(w)
            except ValueError:
                pass
            try:
                on_done(res)
            except Exception:
                pass
        w.done.connect(_fin)
        self._bg_workers.append(w)
        w.start()

    # =================================================================
    #  ANALYTICS (NAYA — bilkul fresh; personal + worldwide)
    #  Networking ka asli fix: background thread me SIRF unverified SSL
    #  context (ssl._create_unverified_context) — ye Windows ka cert-store
    #  LOAD hi nahi karta, isliye thread me kabhi hang/fail nahi hota
    #  (yahi purani dikkat thi: verified context thread me atak jata tha).
    # =================================================================
    def _an_url(self):
        u = self._config.get("stats_url")
        if u is None:
            u = DEFAULT_STATS_URL
        return u or ""

    def _an_fetch(self, params, on_data):
        """Server se ek request (background, UI nahi rukti). on_data(dict|None)."""
        url = self._an_url()
        if not url:
            on_data(None); return
        ver = VERSION

        def fn():
            import urllib.request as U
            import urllib.parse as P
            import json as J
            import ssl
            ctx = ssl._create_unverified_context()   # thread-safe (cert-store load nahi)
            full = url + ("&" if "?" in url else "?") + P.urlencode(params)
            req = U.Request(full, headers={"User-Agent": "ApneScan/%s" % ver})
            try:
                r = U.urlopen(req, timeout=15, context=ctx)
                return J.loads(r.read().decode("utf-8", "ignore"))
            except Exception:
                return None

        def done(res):
            on_data(res if isinstance(res, dict) else None)
        self._run_bg_quiet(fn, done)

    def _an_apply(self, data):
        if data and data.get("ok"):
            self._an_world = data
        self._an_update_box()

    def _an_refresh(self):
        """Worldwide numbers laao + sidebar/box taaza karo."""
        self._an_fetch({"action": "stats"}, self._an_apply)

    def _an_report(self, action, n=0, imp=0, prt=0):
        """scan / ping / event (import-print) server ko bhejo + display taaza."""
        p = {"action": action, "client": self._get_client_id(), "v": VERSION}
        if n:
            p["n"] = str(n)
        if imp:
            p["imp"] = str(imp)
        if prt:
            p["prt"] = str(prt)
        self._an_fetch(p, self._an_apply)

    def _an_wv(self, key):
        w = getattr(self, "_an_world", {}) or {}
        v = w.get(key)
        try:
            return "{:,}".format(int(v))
        except Exception:
            return "…"

    def _an_update_box(self):
        if not hasattr(self, "an_box"):
            return
        st = self._pstats()
        t = st.get("totals", {})
        day = (st.get("days") or {}).get(datetime.datetime.now().strftime("%Y-%m-%d"), {})
        lines = [
            '<b>📊 ApneScan</b> <span style="color:#94a3b8;font-size:10px;">(click = detail)</span>',
            self.L("🌍 Duniya bhar ke scans: <b>%s</b>", "🌍 World scans: <b>%s</b>") % self._an_wv("total"),
            self.L("📅 Aaj (duniya): <b>%s</b>", "📅 Today (world): <b>%s</b>") % self._an_wv("today"),
            self.L("🟢 Abhi online: <b>%s</b>", "🟢 Online now: <b>%s</b>") % self._an_wv("online"),
            self.L("📄 Mere aaj: <b>%d</b> pages", "📄 My today: <b>%d</b> pages") % day.get("pages", 0),
            self.L("🔥 Streak: <b>%d din</b>", "🔥 Streak: <b>%d days</b>") % self._pstats_streak(),
        ]
        try:
            self.an_box.setText("<br>".join(lines))
        except Exception:
            pass

    def show_analytics(self):
        """Poora analytics — personal + worldwide (naya, saaf)."""
        self._an_refresh()
        st = self._pstats()
        t = st.get("totals", {})
        w = getattr(self, "_an_world", {}) or {}
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle(self.L("📊 Analytics", "📊 Analytics"))
        dlg.resize(480, 520)
        v = QtWidgets.QVBoxLayout(dlg)

        def _card(title, rows):
            box = QtWidgets.QGroupBox(title)
            gl = QtWidgets.QVBoxLayout(box)
            lbl = QtWidgets.QLabel("<br>".join(rows))
            lbl.setTextFormat(QtCore.Qt.RichText); lbl.setWordWrap(True)
            gl.addWidget(lbl)
            return box, lbl

        day = (st.get("days") or {}).get(datetime.datetime.now().strftime("%Y-%m-%d"), {})
        my_rows = [
            self.L("📄 Aaj: <b>%d</b> pages · <b>%d</b> PDF", "📄 Today: <b>%d</b> pages · <b>%d</b> PDF")
            % (day.get("pages", 0), day.get("pdfs", 0)),
            self.L("🗓 Is hafte: <b>%d</b> pages", "🗓 This week: <b>%d</b> pages") % self._pstats_sum(7, "pages"),
            self.L("📚 Kul: <b>%s</b> pages · <b>%s</b> PDF", "📚 Total: <b>%s</b> pages · <b>%s</b> PDF")
            % ("{:,}".format(t.get("pages", 0)), "{:,}".format(t.get("pdfs", 0))),
            self.L("📥 Import: <b>%s</b> · 🖨 Print: <b>%s</b>", "📥 Imports: <b>%s</b> · 🖨 Prints: <b>%s</b>")
            % ("{:,}".format(t.get("imports", 0)), "{:,}".format(t.get("prints", 0))),
            self.L("📤 Share: <b>%d</b> · 🔥 Streak: <b>%d din</b>", "📤 Shared: <b>%d</b> · 🔥 Streak: <b>%d days</b>")
            % (t.get("shared", 0), self._pstats_streak()),
        ]
        box1, _ = _card(self.L("🙋 Meri stats (is PC par)", "🙋 My stats (this PC)"), my_rows)
        v.addWidget(box1)

        wbox, wlbl = _card(self.L("🌍 Worldwide (sab users)", "🌍 Worldwide (all users)"), ["…"])
        v.addWidget(wbox)

        def _wfill():
            w2 = getattr(self, "_an_world", {}) or {}
            rows = [
                "🌍 Total scans: <b>%s</b>" % self._an_wv("total"),
                "📅 Aaj: <b>%s</b> · 🟢 Online: <b>%s</b>" % (self._an_wv("today"), self._an_wv("online")),
                "👥 Users: <b>%s</b>" % self._an_wv("users"),
                "📥 Import: <b>%s</b> · 🖨 Print: <b>%s</b>" % (self._an_wv("imports"), self._an_wv("prints")),
            ]
            wlbl.setText("<br>".join(rows))
        _wfill()

        note = QtWidgets.QLabel(self.L(
            "<span style='color:#64748b;font-size:11px;'>Privacy: server par sirf GINTI jaati hai — "
            "kabhi koi document/naam/file nahi.</span>",
            "<span style='color:#64748b;font-size:11px;'>Privacy: only counts are sent — "
            "never any document, name or file.</span>"))
        note.setTextFormat(QtCore.Qt.RichText); note.setWordWrap(True)
        v.addWidget(note)
        v.addStretch(1)
        row = QtWidgets.QHBoxLayout()
        bref = QtWidgets.QPushButton(self.L("🔄 Refresh", "🔄 Refresh"))
        bref.clicked.connect(lambda: (self._an_fetch({"action": "stats"},
                                                     lambda d: (self._an_apply(d), _wfill()))))
        bok = QtWidgets.QPushButton("OK"); bok.clicked.connect(dlg.accept)
        row.addWidget(bref); row.addStretch(1); row.addWidget(bok)
        v.addLayout(row)
        # thodi der baad worldwide bhar do (fetch aane par)
        QtCore.QTimer.singleShot(1500, _wfill)
        QtCore.QTimer.singleShot(3500, _wfill)
        dlg.exec_()

    def _stats_nam(self):
        """Qt ka apna network manager (main event-loop par — koi thread nahi).
        Redirect follow karta hai aur SSL-error ko ignore (sirf public ginti)."""
        if getattr(self, "_nam", None) is None:
            self._nam = QNetworkAccessManager(self)
            try:    # HTTPS cert dikkat ho to bhi ruke nahi (yahan sirf ginti hai)
                self._nam.sslErrors.connect(lambda reply, errs: reply.ignoreSslErrors())
            except Exception:
                pass
            try:    # Apps Script 302 redirect follow karo (Qt 5.9+)
                self._nam.setRedirectPolicy(QNetworkRequest.NoLessSafeRedirectPolicy)
            except Exception:
                pass
        return self._nam

    def _stats_fetch(self, action="stats", n=0, imp=0, prt=0, want_display=True):
        """Stats server se ek request — Qt network se (UI nahi rukti, thread
        nahi). Naye sire se banaya — pehle urllib+thread wala rasta kuch
        machino par kaam nahi karta tha."""
        url = self._stats_url()
        if not url:
            return
        if not HAS_QTNET:      # Qt-network na ho to seedha urllib fallback
            self._stats_fetch_urllib(action, n, imp, prt, want_display)
            return
        import urllib.parse as P
        try:
            country = (QtCore.QLocale().name().split("_") + [""])[1][:2]
        except Exception:
            country = ""
        q = {"action": action, "client": self._get_client_id(),
             "v": VERSION, "c": country}
        if action == "scan" and n:
            q["n"] = str(n)
        if imp:
            q["imp"] = str(imp)
        if prt:
            q["prt"] = str(prt)
        full = url + ("&" if "?" in url else "?") + P.urlencode(q)
        req = QNetworkRequest(QtCore.QUrl(full))
        req.setHeader(QNetworkRequest.UserAgentHeader, "ApneScan/%s" % VERSION)
        try:    # purane Qt (5.6-5.14) me redirect isse follow hota hai
            req.setAttribute(QNetworkRequest.FollowRedirectsAttribute, True)
        except Exception:
            pass
        reply = self._stats_nam().get(req)
        self._stat_replies = getattr(self, "_stat_replies", [])
        self._stat_replies.append(reply)

        def _done():
            try:
                self._stat_replies.remove(reply)
            except Exception:
                pass
            data = None
            try:
                raw = bytes(reply.readAll()).decode("utf-8", "ignore")
                import json
                data = json.loads(raw)
            except Exception:
                data = None
            try:
                reply.deleteLater()
            except Exception:
                pass
            if not want_display:
                return
            if isinstance(data, dict) and data.get("ok"):
                self._on_world_stats(data)
                self._set_stats_display((int(data.get("total", 0)),
                                         int(data.get("today", 0)),
                                         int(data.get("online", 0))))
            else:
                # Qt se na aaya to purana (urllib+thread) rasta ek baar aajma lo
                self._stats_fetch_urllib(action, n, imp, prt, want_display)
        reply.finished.connect(_done)

    def _stats_fetch_urllib(self, action="stats", n=0, imp=0, prt=0, want_display=True):
        """Fallback: agar Qt-network kaam na kare to urllib se (background thread).
        FuncWorker ka proven rasta — GC-safe."""
        url = self._stats_url()
        if not url:
            self._stats_failed(); return
        client = self._get_client_id(); ver = VERSION
        try:
            country = (QtCore.QLocale().name().split("_") + [""])[1][:2]
        except Exception:
            country = ""

        def fn():
            import urllib.request as U
            import urllib.parse as P
            import json as J
            import ssl
            q = {"action": action, "client": client, "v": ver, "c": country}
            if action == "scan" and n:
                q["n"] = str(n)
            if imp:
                q["imp"] = str(imp)
            if prt:
                q["prt"] = str(prt)
            full = url + ("&" if "?" in url else "?") + P.urlencode(q)
            req = U.Request(full, headers={"User-Agent": "ApneScan/%s" % ver})
            for mk in (lambda: U.urlopen(req, timeout=20),
                       lambda: U.urlopen(req, timeout=20, context=ssl._create_unverified_context())):
                try:
                    r = mk()
                    return J.loads(r.read().decode("utf-8", "ignore"))
                except Exception:
                    continue
            return None

        def on_done(data):
            if not want_display:
                return
            if isinstance(data, dict) and data.get("ok"):
                self._on_world_stats(data)
                self._set_stats_display((int(data.get("total", 0)),
                                         int(data.get("today", 0)),
                                         int(data.get("online", 0))))
            else:
                self._stats_failed()
        self._run_bg_quiet(fn, on_done)

    # Analytics/stats HATA diya gaya — ye sab ab kuch nahi karte (no-op).
    def _refresh_stats(self, action="ping"):
        return

    def _report_scan_stat(self, n):
        return

    def _report_event(self, imports=0, prints=0):
        return

    def set_stats_url(self):
        cur = self._stats_url()
        url, ok = QtWidgets.QInputDialog.getText(
            self, "Stats server URL",
            'Google Apps Script web app URL (…/exec):\n(leave empty to turn stats off)',
            text=cur)
        if not ok:
            return
        self._config["stats_url"] = url.strip()
        try:
            save_config(self._config)
        except Exception:
            pass
        self._set_stats_display(None)
        self._refresh_stats()

    def test_stats_connection(self):
        """Stats server se connection ka LIVE test — exact galti dikhata hai
        (taaki 'worldwide stats aa nahi rahe' ki asli wajah pata chale)."""
        import urllib.request as U
        import urllib.parse as P
        import ssl
        url = self._stats_url()
        if not url:
            self._warn(self.L("Stats URL khaali hai (Settings → Stats server URL).",
                              "Stats URL is empty (Settings → Stats server URL)."))
            return
        full = url + ("&" if "?" in url else "?") + P.urlencode({"action": "stats"})
        lines = ["App version: v%s" % VERSION, "URL:", full, ""]
        QtWidgets.QApplication.setOverrideCursor(QtCore.Qt.WaitCursor)
        try:
            for label, ctx in (("1) Normal (verified)", None),
                               ("2) Bina-verify (unverified)", ssl._create_unverified_context())):
                try:
                    req = U.Request(full, headers={"User-Agent": "ApneScan/%s" % VERSION})
                    if ctx is None:
                        r = U.urlopen(req, timeout=15)
                    else:
                        r = U.urlopen(req, timeout=15, context=ctx)
                    body = r.read().decode("utf-8", "ignore")[:400]
                    lines.append("%s: ✅ OK (HTTP %s)" % (label, getattr(r, "status", "?")))
                    lines.append("   " + body)
                except Exception as e:
                    lines.append("%s: ❌ FAIL" % label)
                    lines.append("   " + repr(e)[:400])
                lines.append("")
        finally:
            QtWidgets.QApplication.restoreOverrideCursor()
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle(self.L("Stats connection test", "Stats connection test"))
        dlg.resize(620, 380)
        v = QtWidgets.QVBoxLayout(dlg)
        v.addWidget(QtWidgets.QLabel(self.L(
            "Ye nateeja copy karke bhej dein — isse asli galti pata chal jayegi:",
            "Copy this result and send it — it pinpoints the exact problem:")))
        te = QtWidgets.QPlainTextEdit("\n".join(lines))
        te.setReadOnly(True)
        v.addWidget(te, 1)
        bb = QtWidgets.QHBoxLayout()
        bcopy = QtWidgets.QPushButton(self.L("📋 Copy", "📋 Copy"))
        bcopy.clicked.connect(lambda: QtWidgets.QApplication.clipboard().setText("\n".join(lines)))
        bok = QtWidgets.QPushButton("OK"); bok.clicked.connect(dlg.accept)
        bb.addWidget(bcopy); bb.addStretch(1); bb.addWidget(bok)
        v.addLayout(bb)
        dlg.exec_()

    def _paste_from_clipboard(self):
        """Ctrl+V: clipboard se image/PDF ya copy ki hui file(s) app me laao.
        Agar kisi text-box me likh rahe hain to wahan normal paste ho."""
        fw = QtWidgets.QApplication.focusWidget()
        if isinstance(fw, (QtWidgets.QLineEdit, QtWidgets.QPlainTextEdit, QtWidgets.QTextEdit)):
            try:
                fw.paste()
            except Exception:
                pass
            return
        cb = QtWidgets.QApplication.clipboard()
        md = cb.mimeData()
        exts = (".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".pdf")
        if md is not None and md.hasUrls():
            files = [u.toLocalFile() for u in md.urls()
                     if u.toLocalFile().lower().endswith(exts)]
            if files:
                self._start_import(files, "normal")
                self.status.showMessage(self.L("📋 Paste se %d file aayi" % len(files),
                                               "📋 Pasted %d file(s)" % len(files)), 4000)
                return
        img = cb.image()
        if img is not None and not img.isNull():
            try:
                fd, tmp = tempfile.mkstemp(suffix=".png")
                os.close(fd)
                img.save(tmp, "PNG")
                self._start_import([tmp], "normal")
                self.status.showMessage(self.L("📋 Clipboard ki image aa gayi",
                                               "📋 Pasted image from clipboard"), 4000)
                return
            except Exception:
                pass
        self.status.showMessage(self.L("Clipboard me image/PDF nahi mila",
                                       "No image/PDF found in clipboard"), 3000)

    def set_scanner_ip(self):
        ip, ok = QtWidgets.QInputDialog.getText(self, "Scanner IP", "Scanner IP:", text=self.ip_field.text())
        if ok:
            self.ip_field.setText(ip.strip()); self.check_connection()

    def undo_delete(self):
        if not self._undo_stack:
            self.status.showMessage("Nothing to undo", 3000); return
        path, row = self._undo_stack.pop()
        if not os.path.exists(path):
            return
        icon = QtGui.QIcon(self._make_thumb(path))
        item = QtWidgets.QListWidgetItem(icon, "Page")
        item.setData(QtCore.Qt.UserRole, path); item.setTextAlignment(QtCore.Qt.AlignHCenter)
        self.list.insertItem(min(row, self.list.count()), item)
        self.list.setCurrentItem(item); self.list.clearSelection()
        self._renumber_pages(); self._dirty = True
        self._update_status(); self._update_empty_state()

    def _renumber_pages(self):
        show = self._opts.get("show_page_numbers", True)
        for i in range(self.list.count()):
            it = self.list.item(i)
            title = it.data(TITLE_ROLE)
            if title:
                it.setText(title)
            else:
                it.setText(("Page %d" % (i + 1)) if show else "")

    def _update_empty_state(self):
        empty = self.list.count() == 0
        use_dash = empty and bool(self._opts.get("ui_dashboard", True)) and hasattr(self, "_dash")
        if hasattr(self, "_dash"):
            self._dash.setVisible(use_dash)
        self._empty_lbl.setVisible(empty and not use_dash)
        if empty:
            r = self.list.viewport().rect()
            self._empty_lbl.setGeometry(r)
            if hasattr(self, "_dash"):
                self._dash.setGeometry(r)
                try:
                    names = [os.path.basename(p) for p in (self._recent or [])[:3]]
                    self._dash_recent.setText(("Recent: " + "  ·  ".join(names)) if names else "")
                except Exception:
                    pass

    def _apply_thumb_zoom(self, w):
        w = max(80, min(340, int(w)))
        h = int(w * 4 / 3)
        self._thumb_w, self._thumb_h = w, h
        self.list.setIconSize(QtCore.QSize(w, h))
        # naam ke liye neeche zyada jagah (2 line ka naam bhi poora dikhe)
        self.list.setGridSize(QtCore.QSize(w + 24, h + 52))
        for _i in range(self.list.count()):
            self.list.item(_i).setSizeHint(QtCore.QSize(w + 24, h + 52))

    def _zoom_thumbs(self, factor):
        self._apply_thumb_zoom(self._thumb_w * factor)

    def eventFilter(self, obj, ev):
        if obj is self.list.viewport():
            if ev.type() == QtCore.QEvent.Resize:
                self._empty_lbl.setGeometry(self.list.viewport().rect())
                if hasattr(self, "_dash"):
                    self._dash.setGeometry(self.list.viewport().rect())
            elif ev.type() == QtCore.QEvent.Wheel and (ev.modifiers() & QtCore.Qt.ControlModifier):
                # Ctrl + mouse scroll -> zoom thumbnails in/out (like NAPS2)
                self._apply_thumb_zoom(self._thumb_w * (1.15 if ev.angleDelta().y() > 0 else 0.87))
                return True
        # Preview panel: Ctrl+scroll = zoom
        if (hasattr(self, "pv_scroll") and obj is self.pv_scroll.viewport()
                and ev.type() == QtCore.QEvent.Wheel
                and (ev.modifiers() & QtCore.Qt.ControlModifier)):
            self._pv_do_zoom(1.2 if ev.angleDelta().y() > 0 else 0.83)
            return True
        # Loupe (kaanch): preview image par maus le jao to zoom-jhalak
        if (getattr(self, "_pv_loupe_on", False) and hasattr(self, "pv_img")
                and obj is self.pv_img):
            if ev.type() == QtCore.QEvent.MouseMove:
                self._pv_loupe_at(ev)
                return False
            elif ev.type() == QtCore.QEvent.Leave:
                if hasattr(self, "_pv_loupe"):
                    self._pv_loupe.hide()
        return super().eventFilter(obj, ev)

    def _pv_loupe_at(self, ev):
        """Maus ke neeche ke hisse ko bada karke dikhao (loupe)."""
        try:
            pm = self.pv_img.pixmap()
            if pm is None or pm.isNull():
                return
            pos = ev.pos()
            src = 60          # kitna area uthana (px)
            x = max(0, min(pm.width() - src, pos.x() - src // 2))
            y = max(0, min(pm.height() - src, pos.y() - src // 2))
            crop = pm.copy(x, y, src, src).scaled(180, 180,
                       QtCore.Qt.KeepAspectRatio, QtCore.Qt.FastTransformation)
            self._pv_loupe.setPixmap(crop)
            self._pv_loupe.resize(crop.size())
            gp = self.pv_img.mapToGlobal(pos)
            self._pv_loupe.move(gp.x() + 16, gp.y() + 16)
            self._pv_loupe.show()
        except Exception:
            pass

    def dragEnterEvent(self, e):
        if e.mimeData().hasUrls():
            e.acceptProposedAction()

    def dropEvent(self, e):
        exts = (".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".pdf")
        files = [url.toLocalFile() for url in e.mimeData().urls()
                 if url.toLocalFile().lower().endswith(exts)]
        if files:
            self._start_import(files, "normal")

    def _after_save_action(self, out):
        act = self._opts.get("after_save", "nothing")
        try:
            if act == "open":
                os.startfile(out)
            elif act == "folder":
                os.startfile(os.path.dirname(out))
        except Exception:
            pass

    # ---- Update check ----
    def check_updates(self, silent=False):
        self._upd_worker = UpdateChecker()
        self._upd_worker.result.connect(
            lambda tag, url, s=silent: self._on_update_result(tag, url, s))
        self._upd_worker.start()

    def _on_update_result(self, tag, url, silent):
        def _n(s):
            m = re.findall(r"\d+", s or "")
            return int(m[0]) if m else -1
        if not tag:
            if not silent:
                self._warn("Could not check for updates. Is the internet working?")
            return
        if _n(tag) > _n(VERSION):
            # Sidebar me banner dikhao — wahi se EK click me update
            self._show_update_banner(tag)
            if not silent:
                r = QtWidgets.QMessageBox.question(
                    self, "New version available",
                    "A new version %s of ApneScan is available (you have v%s).\n\n"
                    "Download and update now? (The app will close and reopen the new version.)"
                    % (tag, VERSION),
                    QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
                if r == QtWidgets.QMessageBox.Yes:
                    self._start_self_update()
        elif not silent:
            QtWidgets.QMessageBox.information(
                self, "Update", "You are already on the latest version (v%s)." % VERSION)

    def _show_update_banner(self, tag):
        self._latest_tag = tag
        try:
            self.update_box.setText(self.L(
                "🔔 Naya update %s aa gaya!\n⬇ Ek click me update karo" % tag,
                "🔔 Update %s is available!\n⬇ One-click update" % tag))
            self.update_box.setEnabled(True)
            self.update_box.show()
        except Exception:
            pass

    def _sidebar_update_clicked(self):
        try:
            self.update_box.setEnabled(False)
            self.update_box.setText("⬇ Downloading…\n(you can keep using the app)")
        except Exception:
            pass
        self._start_self_update()

    def _reset_update_banner(self):
        """Download fail hone par banner wapas clickable karo."""
        try:
            if getattr(self, "_latest_tag", ""):
                self._show_update_banner(self._latest_tag)
        except Exception:
            pass

    def _start_self_update(self):
        """Naya version website se download karke KHUD install karo (feature 36)."""
        url = "https://apnescan.apnesoft.com/ApneScan.exe"

        def job():
            import urllib.request as U
            fd, tmp = tempfile.mkstemp(suffix=".exe")
            os.close(fd)
            req = U.Request(url, headers={"User-Agent": "ApneScan"})
            expected = 0
            with _urlopen_safe(req, timeout=300) as r, open(tmp, "wb") as fh:
                try:
                    expected = int(r.headers.get("Content-Length") or 0)
                except Exception:
                    expected = 0
                shutil.copyfileobj(r, fh)
            got = os.path.getsize(tmp)
            # ADHOORA download hi asli dikkat thi: aadhi-likhi onefile exe
            # 'python312.dll load fail' deti hai. Isliye pakka karo ki poori
            # file aayi (Content-Length se), aur ye sach me ek Windows exe hai.
            with open(tmp, "rb") as fh:
                head = fh.read(2)
            if head != b"MZ":
                raise RuntimeError('The download was not a valid exe (the server may have sent the wrong file). Try from the website.')
            if expected and got < expected:
                raise RuntimeError('The download was incomplete (only %d of %d bytes). Check your internet and try again.'
                                   % (expected, got))
            if got < 20_000_000:
                raise RuntimeError('The download did not complete (only %d bytes). Try again or get it from the website.' % got)
            return tmp

        def done(res):
            if isinstance(res, Exception):
                self._reset_update_banner()
                self._warn("Could not download the update:\n%s\n\n"
                           "Please download it yourself from the website: apnescan.apnesoft.com" % res)
                try:
                    import webbrowser
                    webbrowser.open(DOWNLOAD_PAGE)
                except Exception:
                    pass
                return
            self._apply_downloaded_update(res)
        self._run_bg(job, done, "Downloading new version… (the app keeps running)")

    def _apply_downloaded_update(self, tmp_exe):
        if not getattr(sys, "frozen", False):
            self._reveal_in_explorer(tmp_exe)
            QtWidgets.QMessageBox.information(
                self, "Download complete", "The new exe is here:\n%s" % tmp_exe)
            return
        cur = os.path.abspath(sys.executable)
        bat = os.path.join(tempfile.gettempdir(), "apnescan_update.bat")
        # App band hone ka intezaar karo, phir purani exe ko nayi se badlo.
        # Kabhi purani exe abhi bhi lock hoti hai -> copy fail; isliye kai baar
        # koshish karte hain. Agar phir bhi na ho, to (verified) nayi exe ko
        # SEEDHA temp se hi chala dete hain — taaki adhoori-copy exe kabhi na
        # chale (wahi 'python DLL load fail' deti thi).
        body = (
            "@echo off\r\n"
            "ping 127.0.0.1 -n 4 > nul\r\n"
            "set SRC=__SRC__\r\n"
            "set DST=__DST__\r\n"
            "set OK=0\r\n"
            "for /L %%i in (1,1,15) do (\r\n"
            "  copy /y \"%SRC%\" \"%DST%\" > nul 2>&1\r\n"
            "  if not errorlevel 1 ( set OK=1 & goto launch )\r\n"
            "  ping 127.0.0.1 -n 2 > nul\r\n"
            ")\r\n"
            ":launch\r\n"
            "if \"%OK%\"==\"1\" ( start \"\" \"%DST%\" ) else ( start \"\" \"%SRC%\" )\r\n"
            "del \"%~f0\"\r\n"
        ).replace("__SRC__", tmp_exe).replace("__DST__", cur)
        try:
            with open(bat, "w") as fh:
                fh.write(body)
        except Exception as exc:
            self._warn("Update script failed: %s" % exc)
            return
        r = QtWidgets.QMessageBox.question(
            self, "Update ready",
            "The new version has been downloaded.\n\nThe app will now close and the new version "
            "will open by itself. OK?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
        if r == QtWidgets.QMessageBox.Yes:
            try:
                os.startfile(bat)
            except Exception as exc:
                self._warn("Update did not start: %s" % exc)
                return
            QtWidgets.QApplication.quit()

    def export_settings(self):
        """Saari settings ek file me — naye PC par le jaane ke liye."""
        out, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, 'Export settings',
            os.path.join(os.path.expanduser("~"), "apnescan_settings.json"),
            "Settings (*.json)")
        if not out:
            return
        try:
            shutil.copy2(CONFIG_PATH, out)
        except Exception as exc:
            self._warn("Export failed: %s" % exc); return
        QtWidgets.QMessageBox.information(
            self, "Done", "Settings exported:\n%s\n\nOn a new PC: Settings → "
            "Import settings… to bring them back." % out)

    def import_settings(self):
        f, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, 'Choose a settings file', os.path.expanduser("~"), "Settings (*.json)")
        if not f:
            return
        try:
            with open(f, "r", encoding="utf-8") as fh:
                json.load(fh)          # valid json hai?
            shutil.copy2(f, CONFIG_PATH)
        except Exception as exc:
            self._warn("Import failed (file does not look valid): %s" % exc); return
        QtWidgets.QMessageBox.information(
            self, "Done", "Settings imported. Close and reopen the app.")

    def toggle_touch_mode(self):
        self._opts["touch_mode"] = bool(self.act_touch.isChecked())
        self._save_opts()
        self._apply_style()
        self.status.showMessage(
            "Touch mode %s — full effect after reopening the app." %
            ("ON" if self._opts["touch_mode"] else "OFF"), 5000)

    def open_crash_report(self):
        if os.path.exists(CRASH_PATH):
            self._open_path(CRASH_PATH)
        else:
            QtWidgets.QMessageBox.information(
                self, "Report", "No error report — everything is running fine! 🙂")

    # ---- Share (WhatsApp / Email) ----
    def _pick_share_pdf(self):
        """Sabse aakhri saved PDF; koi na ho to user se file chunwao."""
        for p in (self._recent or []):
            if p and os.path.exists(p):
                return p
        start = self._opts.get("save_folder", os.path.expanduser("~"))
        f, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, 'Which PDF to send?', start, "PDF (*.pdf)")
        return f or None

    def _reveal_in_explorer(self, path):
        try:
            import subprocess
            subprocess.Popen(["explorer", "/select,", os.path.normpath(path)])
        except Exception:
            try:
                os.startfile(os.path.dirname(path))
            except Exception:
                pass

    def _copy_file_to_clipboard(self, path):
        """File ko clipboard par daalo taaki WhatsApp/Email me Ctrl+V se attach ho."""
        try:
            md = QtCore.QMimeData()
            md.setUrls([QtCore.QUrl.fromLocalFile(os.path.abspath(path))])
            QtWidgets.QApplication.clipboard().setMimeData(md)
            return True
        except Exception:
            return False

    def _open_whatsapp(self):
        """WhatsApp Desktop SEEDHA kholo (whatsapp://). Na ho to WhatsApp Web."""
        try:
            os.startfile("whatsapp://")     # Desktop app (protocol)
            return True
        except Exception:
            pass
        try:
            import webbrowser
            webbrowser.open("https://web.whatsapp.com/")
            return True
        except Exception:
            return False

    def share_whatsapp(self, pdf=None):
        if not isinstance(pdf, str) or not pdf:
            pdf = self._pick_share_pdf()
        if not pdf:
            return
        # 1) file clipboard par (Ctrl+V se attach)  2) WhatsApp seedha khol do
        self._copy_file_to_clipboard(pdf)
        self._open_whatsapp()
        self._pstats_bump(shared=1)
        # bina roke chhoti si hint (koi blocking dialog / Explorer popup nahi)
        self.status.showMessage(self.L(
            "WhatsApp khul gaya — contact chuno, phir Ctrl+V (file copy ho chuki hai)",
            "WhatsApp opened — pick a contact, then press Ctrl+V (the file is already copied)"), 9000)

    def share_email(self, pdf=None):
        if not isinstance(pdf, str) or not pdf:
            pdf = self._pick_share_pdf()
        if not pdf:
            return
        self._pstats_bump(shared=1)
        # Pehle Outlook try karo (attachment ke saath draft khulta hai)
        if HAS_W32:
            try:
                _pythoncom.CoInitialize()
                try:
                    ol = _w32.Dispatch("Outlook.Application")
                    mail = ol.CreateItem(0)
                    mail.Subject = os.path.splitext(os.path.basename(pdf))[0]
                    mail.Attachments.Add(os.path.abspath(pdf))
                    mail.Display(True)
                    return
                finally:
                    try:
                        _pythoncom.CoUninitialize()
                    except Exception:
                        pass
            except Exception:
                pass
        # Fallback: default mail app (mailto attachment support nahi karta,
        # isliye file clipboard par + Explorer me dikha do)
        self._copy_file_to_clipboard(pdf)
        self._reveal_in_explorer(pdf)
        try:
            import webbrowser
            import urllib.parse as _up
            webbrowser.open("mailto:?subject=%s"
                            % _up.quote(os.path.splitext(os.path.basename(pdf))[0]))
        except Exception:
            pass
        QtWidgets.QMessageBox.information(
            self, "Send via Email",
            "An email draft is opening. To attach the PDF:\n\n"
            "1) Press Ctrl+V in the email (the file is already copied)\n"
            "2) Or drag the PDF from Explorer onto the email.\n\n"
            "File: %s" % pdf)

    # ---- App phailao: Share / QR / Poster / Review (growth) ----
    def _app_qr_image(self, url=None, box=8):
        """Download-link ka QR (PIL image). qrcode na ho to None."""
        try:
            import qrcode
            qr = qrcode.QRCode(border=2, box_size=box,
                               error_correction=qrcode.constants.ERROR_CORRECT_M)
            qr.add_data(url or WEBSITE_URL); qr.make(fit=True)
            return qr.make_image(fill_color="#0f766e", back_color="white").convert("RGB")
        except Exception:
            return None

    def share_app(self):
        """Is app ko doosron tak pahuchao — WhatsApp / link copy / QR / poster."""
        L = self.L
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle(L("📣 ApneScan doosron ko batao", "📣 Share ApneScan"))
        dlg.setMinimumWidth(440)
        v = QtWidgets.QVBoxLayout(dlg)
        top = QtWidgets.QLabel(L(
            "<b>ApneScan free hai — jitne zyada log istemal karenge, utna achha!</b><br>"
            "Neeche se ek tarika chuno. (Message + download-link tayaar hai.)",
            "<b>ApneScan is free — the more people use it, the better!</b><br>"
            "Pick a way below. (A ready message + download link is included.)"))
        top.setTextFormat(QtCore.Qt.RichText); top.setWordWrap(True)
        v.addWidget(top)
        msg = QtWidgets.QPlainTextEdit(SHARE_TEXT); msg.setFixedHeight(76)
        v.addWidget(msg)

        def _wa():
            try:
                import urllib.parse as _up
                os.startfile("whatsapp://send?text=" + _up.quote(msg.toPlainText()))
            except Exception:
                try:
                    import webbrowser, urllib.parse as _up
                    webbrowser.open("https://wa.me/?text=" + _up.quote(msg.toPlainText()))
                except Exception:
                    pass
            self.status.showMessage(L("WhatsApp khul gaya — contact chuno aur bhejo",
                                      "WhatsApp opened — pick a contact and send"), 6000)

        def _copy():
            QtWidgets.QApplication.clipboard().setText(msg.toPlainText())
            self.status.showMessage(L("📋 Message copy ho gaya — kahin bhi paste karo",
                                      "📋 Message copied — paste it anywhere"), 5000)

        def _copylink():
            QtWidgets.QApplication.clipboard().setText(WEBSITE_URL)
            self.status.showMessage(L("📋 Link copy: " + WEBSITE_URL, "📋 Link copied: " + WEBSITE_URL), 5000)

        def _email():
            try:
                import webbrowser, urllib.parse as _up
                webbrowser.open("mailto:?subject=%s&body=%s"
                                % (_up.quote("ApneScan — free scanner software"),
                                   _up.quote(msg.toPlainText())))
            except Exception:
                pass
        grid = QtWidgets.QGridLayout()
        for i, (ic, lb, fn) in enumerate((
                ("🟢", L("WhatsApp par bhejo", "Send on WhatsApp"), _wa),
                ("📋", L("Message copy", "Copy message"), _copy),
                ("🔗", L("Sirf link copy", "Copy link only"), _copylink),
                ("✉", L("Email se bhejo", "Send by Email"), _email),
                ("🔳", L("QR code dikhao/save", "Show/save QR code"), self.show_app_qr),
                ("🖼", L("Poster/pamphlet banao", "Make a poster"), self.make_poster))):
            b = QtWidgets.QPushButton("%s  %s" % (ic, lb)); b.setMinimumHeight(38)
            b.clicked.connect(fn)
            grid.addWidget(b, i // 2, i % 2)
        v.addLayout(grid)
        note = QtWidgets.QLabel(L(
            "<span style='color:#64748b;font-size:11px;'>💡 Dukaan/hospital me QR-poster "
            "laga do — log mobile se scan karke seedha download kar lenge.</span>",
            "<span style='color:#64748b;font-size:11px;'>💡 Put a QR poster up at your "
            "shop/hospital — people can scan it and download directly.</span>"))
        note.setTextFormat(QtCore.Qt.RichText); note.setWordWrap(True)
        v.addWidget(note)
        bb = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Close)
        bb.rejected.connect(dlg.reject); bb.accepted.connect(dlg.reject)
        v.addWidget(bb)
        dlg.exec_()

    def show_app_qr(self):
        """Download-link ka QR — screen par + save karne ka option."""
        img = self._app_qr_image(box=10)
        if img is None:
            self._warn(self.L("QR banane ke liye 'qrcode' library chahiye (agle build me).",
                              "The 'qrcode' library is needed for QR (in the next build).")); return
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle(self.L("🔳 ApneScan QR", "🔳 ApneScan QR"))
        v = QtWidgets.QVBoxLayout(dlg)
        lbl = QtWidgets.QLabel(); lbl.setAlignment(QtCore.Qt.AlignCenter)
        data = img.tobytes("raw", "RGB")
        qim = QtGui.QImage(data, img.width, img.height, 3 * img.width, QtGui.QImage.Format_RGB888)
        lbl.setPixmap(QtGui.QPixmap.fromImage(qim.copy()))
        v.addWidget(lbl)
        cap = QtWidgets.QLabel(self.L("Mobile camera se scan karo → seedha download",
                                      "Scan with a phone camera → direct download"))
        cap.setAlignment(QtCore.Qt.AlignCenter); cap.setStyleSheet("color:#475569;")
        v.addWidget(cap)
        row = QtWidgets.QHBoxLayout()
        bsave = QtWidgets.QPushButton(self.L("💾 Image save karo", "💾 Save image"))

        def _save():
            out, _ = QtWidgets.QFileDialog.getSaveFileName(
                self, self.L("QR save", "Save QR"),
                os.path.join(os.path.expanduser("~"), "ApneScan_QR.png"), "PNG (*.png)")
            if out:
                try:
                    img.save(out); self.status.showMessage(self.L("QR save ho gaya", "QR saved"), 4000)
                except Exception as e:
                    self._warn(str(e))
        bsave.clicked.connect(_save)
        bclose = QtWidgets.QPushButton(self.L("Band", "Close")); bclose.clicked.connect(dlg.accept)
        row.addWidget(bsave); row.addStretch(1); row.addWidget(bclose)
        v.addLayout(row)
        dlg.exec_()

    def make_poster(self):
        """Dukaan/hospital me lagane layak poster (QR + features) — PNG + PDF."""
        L = self.L

        def job():
            W, H = 1240, 1754      # A4 @150dpi (portrait)
            im = Image.new("RGB", (W, H), "white")
            d = ImageDraw.Draw(im)

            def _font(sz, bold=False):
                for fn in (("arialbd.ttf", "Arial_Bold.ttf") if bold else ("arial.ttf", "Arial.ttf")):
                    try:
                        return ImageFont.truetype(fn, sz)
                    except Exception:
                        pass
                try:
                    return ImageFont.truetype("DejaVuSans%s.ttf" % ("-Bold" if bold else ""), sz)
                except Exception:
                    return ImageFont.load_default()
            teal = (15, 118, 110)
            d.rectangle([0, 0, W, 210], fill=teal)
            d.text((60, 55), "ApneScan", fill="white", font=_font(96, True))
            d.text((64, 165), "FREE Document Scanner  •  Windows", fill=(209, 234, 231), font=_font(30))
            feats = [
                "Scan to PDF  —  Hindi + English OCR (searchable)",
                "PDF compress  —  200KB / 500KB / 1MB (portal upload)",
                "Phone photo / webcam  →  clean PDF",
                "WhatsApp & Email share  •  Merge / Split / Sign",
                "No ads  •  No cloud  •  Works fully offline  •  100% FREE",
            ]
            y = 300
            for f in feats:
                d.ellipse([60, y + 10, 84, y + 34], fill=teal)
                d.text((110, y), f, fill=(31, 41, 51), font=_font(38))
                y += 84
            qr = self._app_qr_image(box=12)
            if qr is not None:
                qr = qr.resize((520, 520))
                im.paste(qr, ((W - 520) // 2, 820))
            d.text((W // 2, 1380), "Scan this QR with your phone", fill=(71, 85, 105),
                   font=_font(40), anchor="mm")
            d.text((W // 2, 1440), "or visit", fill=(71, 85, 105), font=_font(34), anchor="mm")
            d.text((W // 2, 1500), WEBSITE_URL, fill=teal, font=_font(46, True), anchor="mm")
            d.rectangle([0, H - 70, W, H], fill=teal)
            d.text((W // 2, H - 35), "ApneSoft  •  Free & Open-Source", fill="white",
                   font=_font(28), anchor="mm")
            base = os.path.join(self._opts.get("save_folder", os.path.expanduser("~")),
                                "ApneScan_Poster")
            png = base + ".png"; pdf = base + ".pdf"
            im.save(png)
            im.save(pdf, "PDF", resolution=150)
            return (png, pdf)

        def done(res):
            if isinstance(res, Exception):
                self._warn(str(res)); return
            png, pdf = res
            self.status.showMessage(L("🖼 Poster ban gaya: ", "🖼 Poster created: ") + pdf, 8000)
            try:
                self._open_path(png)
            except Exception:
                pass
        self._run_bg(job, done, L("Poster bana rahe…", "Making the poster…"))

    def ask_review(self):
        """User se GitHub par ⭐ / review maango (aasaan link)."""
        r = QtWidgets.QMessageBox.question(
            self, self.L("⭐ Pasand aaya?", "⭐ Enjoying ApneScan?"),
            self.L("ApneScan free hai. Agar pasand aaya to GitHub par ⭐ star dein ya "
                   "review likhein — isse aur logon tak pahunchega. Ab kholu?",
                   "ApneScan is free. If you like it, please give it a ⭐ on GitHub or "
                   "write a review — it helps others find it. Open now?"),
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No, QtWidgets.QMessageBox.Yes)
        if r == QtWidgets.QMessageBox.Yes:
            try:
                import webbrowser
                webbrowser.open(GITHUB_URL)
            except Exception:
                pass

    def _maybe_growth_nudge(self):
        """Kabhi-kabhi (bahut halke se) share/review ki yaad dilao — kabhi
        pareshan na kare: 30 din me sirf ek baar, aur kaafi kaam ke baad."""
        try:
            cfg = self._config
            import time as _t
            now = int(_t.time())
            last = int(cfg.get("nudge_last", 0) or 0)
            if now - last < 30 * 86400:       # 30 din me ek baar hi
                return
            total_pdfs = int((self._pstats().get("totals", {}) or {}).get("pdfs", 0))
            if total_pdfs < 25:               # thoda istemal ho jaane ke baad hi
                return
            cfg["nudge_last"] = now
            save_config(cfg)
            self.status.showMessage(self.L(
                "🙏 ApneScan pasand aaya? Help → 'ApneScan share karo' se doston ko batayein.",
                "🙏 Liking ApneScan? Tell friends via Help → 'Share ApneScan'."), 12000)
        except Exception:
            pass

    # ---- PDF compress tool ----
    def compress_pdf_tool(self, src_pdf=None):
        """Abhi ke pages (ya koi bhi purani PDF) ko chhota karke alag PDF banao —
        portal ki 200KB/500KB/1MB/2MB limit ke hisaab se."""
        if not isinstance(src_pdf, str) or not src_pdf:
            src_pdf = None
        paths = self._ordered_paths()
        if src_pdf is None and not paths:
            start = self._opts.get("save_folder", os.path.expanduser("~"))
            src_pdf, _ = QtWidgets.QFileDialog.getOpenFileName(
                self, 'Which PDF to shrink?', start, "PDF (*.pdf)")
            if not src_pdf:
                return
        targets = ["200 KB (portal upload)", "500 KB", "1 MB", "2 MB",
                   'Only reduce quality (no size limit)']
        choice, ok = QtWidgets.QInputDialog.getItem(
            self, "Compress PDF", "How small should the file be?", targets, 2, False)
        if not ok:
            return
        idx = targets.index(choice)
        limit = {0: 200 * 1024, 1: 500 * 1024, 2: 1024 * 1024, 3: 2 * 1024 * 1024}.get(idx)
        if src_pdf:
            pages = pdf_to_images(src_pdf, self._tmpdir)
            if not pages:
                self._warn("Could not extract pages from this PDF.\nFor better results install 'PyMuPDF':\n  py -3.12-32 -m pip install PyMuPDF")
                return
            default = os.path.splitext(src_pdf)[0] + "_small.pdf"
        else:
            pages = paths
            base = self._build_filename(".pdf", paths=paths)
            default = (base[:-4] if base.lower().endswith(".pdf") else base) + "_small.pdf"
        out, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, 'Save the smaller PDF', default, "PDF (*.pdf)")
        if not out:
            return
        if not out.lower().endswith(".pdf"):
            out += ".pdf"

        def job():
            size = self._write_compressed_pdf(pages, out, limit)
            saved = 0
            if src_pdf:
                try:
                    saved = os.path.getsize(src_pdf) - size
                except Exception:
                    saved = 0
            return (size, saved)

        def done(res):
            if isinstance(res, Exception):
                self._warn("Compress failed:\n%s" % res); return
            size, saved = res
            self._remember_save_dir(out)
            if saved > 0:
                self._pstats_bump(saved_bytes=saved)
            kb = size / 1024.0
            pretty = ("%.0f KB" % kb) if kb < 1024 else ("%.1f MB" % (kb / 1024.0))
            note = ""
            if limit and size > limit:
                note = '\n\n(Could not get this small — too many pages. Choose fewer pages or a larger limit.)'
            QtWidgets.QMessageBox.information(
                self, "Done", "Smaller PDF saved (%s):\n%s%s" % (pretty, out, note))
        self._run_bg(job, done, "Shrinking PDF (compress)…")

    def _write_compressed_pdf(self, img_paths, out, limit_bytes=None):
        """JPEG quality/size ghata-ghata kar PDF banao jab tak limit ke andar na aaye.
        Returns final size in bytes."""
        combos = [(75, 1.0), (60, 1.0), (45, 1.0), (60, 0.8), (45, 0.8),
                  (35, 0.8), (45, 0.65), (35, 0.65), (30, 0.5), (25, 0.4)]
        if limit_bytes is None:
            combos = [(60, 1.0)]
        data = None
        for q, s in combos:
            buf = io.BytesIO()
            imgs = []
            try:
                for p in img_paths:
                    im = Image.open(p).convert("RGB")
                    if s < 1.0:
                        im = im.resize((max(1, int(im.width * s)),
                                        max(1, int(im.height * s))), Image.LANCZOS)
                    b2 = io.BytesIO()
                    im.save(b2, "JPEG", quality=q)
                    im.close()
                    b2.seek(0)
                    im2 = Image.open(b2)
                    im2.load()
                    imgs.append(im2)
                imgs[0].save(buf, "PDF", save_all=True, append_images=imgs[1:], resolution=200.0)
            finally:
                for im in imgs:
                    try:
                        im.close()
                    except Exception:
                        pass
            data = buf.getvalue()
            if limit_bytes is None or len(data) <= limit_bytes:
                break
        with open(out, "wb") as fh:
            fh.write(data)
        return len(data)

    def _run_wizard(self):
        wiz = SetupWizard(self, self._opts.get("language", "en"))
        if wiz.exec_() == QtWidgets.QDialog.Accepted:
            self._opts["scanner_method"] = wiz.result_method
            if wiz.result_method == "wia":
                self._opts["wia_device_id"] = wiz.result_wia_id
            self._save_opts()
            prof = self._selected_profile()
            if prof is None:
                prof = {"name": "Default", "source_name": wiz.result_device_name,
                        "dpi": 200, "color": "gray", "duplex": False}
                self._profiles.append(prof)
            elif wiz.result_method == "twain" and wiz.result_device_name:
                prof["source_name"] = wiz.result_device_name
            self._save_profiles(); self._refresh_profile_combo()
            i = self.cmb_profile.findText(prof.get("name"))
            if i >= 0:
                self.cmb_profile.setCurrentIndex(i)
        self._config["setup_done"] = True
        save_config(self._config)
        self._update_status()

    def choose_scan_method(self):
        cur = self._opts.get("scanner_method", "twain")
        labels = [self.L("🔍 Auto-detect (recommended) — LAN ya USB khud pehchano",
                         "🔍 Auto-detect (recommended) — find LAN or USB automatically"),
                  "escl (network duplex - no extra software)",
                  "twain (USB duplex)", "wia (USB)"]
        keys = ["auto", "escl", "twain", "wia"]
        idx = keys.index(cur) if cur in keys else 0
        choice, ok = QtWidgets.QInputDialog.getItem(
            self, "Scan method", "Method:", labels, idx, False)
        if not ok:
            return
        method = keys[labels.index(choice)]
        if method == "auto":
            self.auto_detect_scanner()
            return
        self._opts["scanner_method"] = method
        if method == "escl":
            ip = self.ip_field.text().strip()
            if not ip:
                ip, oki = QtWidgets.QInputDialog.getText(
                    self, "Scanner IP", "Enter the scanner's network IP (e.g. 192.168.1.8):")
                if oki:
                    ip = ip.strip(); self.ip_field.setText(ip)
            self._opts["scanner_ip"] = ip
        elif method == "wia":
            try:
                devs = list_wia_sources()
            except Exception as exc:
                self._warn(friendly_error(exc, self._opts.get("language", "en"))); devs = []
            if devs:
                names = [n for _i, n in devs]
                name, ok2 = QtWidgets.QInputDialog.getItem(self, "WIA scanner", "Scanner:", names, 0, False)
                if ok2 and name:
                    for _id, n in devs:
                        if n == name:
                            self._opts["wia_device_id"] = _id; break
        else:
            # TWAIN: let the user pick the real driver source (e.g. "HP TWAIN USB"),
            # which supports duplex — unlike the WIA-bridge sources.
            try:
                names = list_sources(int(self.winId()))
            except Exception as exc:
                self._warn(friendly_error(exc, self._opts.get("language", "en"))); names = []
            if names:
                name, ok2 = QtWidgets.QInputDialog.getItem(
                    self, "TWAIN scanner",
                    "Choose a scanner (for duplex pick a real driver like 'HP TWAIN USB',\nnot the WIA-... one):",
                    names, 0, False)
                if ok2 and name:
                    prof = self._selected_profile()
                    if prof is not None:
                        prof["source_name"] = name
                        self._save_profiles()
                        self._load_profile_to_panel(prof)
        self._save_opts(); self._update_status(); self._refresh_method_label()

    def choose_language(self):
        cur = self._opts.get("language", "en")
        LANGS = [("Hindi", "hi"), ("English", "en"),
                 ("मराठी Marathi (adhuri)", "mr"), ("ગુજરાતી Gujarati (adhuri)", "gu"),
                 ("ਪੰਜਾਬੀ Punjabi (adhuri)", "pa"), ("தமிழ் Tamil (adhuri)", "ta")]
        idx = next((i for i, (_n, c) in enumerate(LANGS) if c == cur), 0)
        lang, ok = QtWidgets.QInputDialog.getItem(
            self, "Language", "Language:", [n for n, _c in LANGS], idx, False)
        if not ok:
            return
        self._opts["language"] = dict((n, c) for n, c in LANGS).get(lang, "hi")
        self._save_opts()
        QtWidgets.QMessageBox.information(
            self, "Language",
            "Language changed. Close and reopen the app to fully apply."
            if self._opts["language"] == "hi"
            else "Language changed. Close and reopen the app to fully apply.")

    def show_help(self):
        HelpDialog(self, self._opts.get("language", "en")).exec_()

    def run_duplex_test(self):
        """Scan with Both-side ON, then report whether the scanner really
        returned 2 sides — with the duplex value used and any error."""
        if self.list.count() > 0:
            if QtWidgets.QMessageBox.question(
                    self, "Duplex Test",
                    "This clears current pages. Continue?"
                    if self._lang == "hi" else "This clears current pages. Continue?",
                    QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                    QtWidgets.QMessageBox.No) != QtWidgets.QMessageBox.Yes:
                return
            for p in self._ordered_paths():
                try:
                    os.remove(p)
                except Exception:
                    pass
            self.list.clear(); self._update_empty_state()

        QtWidgets.QMessageBox.information(
            self, "Duplex Test",
            "Put 1 sheet (printed on both sides) in the feeder, then press OK.\n"
            "The test scans that sheet and tells you whether both sides came through."
            if self._lang == "hi" else
            "Put ONE double-sided sheet in the feeder, then press OK.")

        prof = self._selected_profile()
        if prof is None:
            self._warn("Select a profile/scanner first."); return
        method = self._opts.get("scanner_method", "twain")
        dpi = 200
        pixel = "gray"
        got = {"pages": 0, "err": None, "trace": []}
        WIA_FMT_BMP = "{B96B3CAB-0728-11D3-9D7B-0000F81EF32E}"

        QtWidgets.QApplication.setOverrideCursor(QtCore.Qt.WaitCursor)
        try:
            if method == "wia":
                try:
                    _pythoncom.CoInitialize()
                except Exception:
                    pass
                try:
                    tr = got["trace"]
                    wid = self._opts.get("wia_device_id")

                    def _connect():
                        dm = _w32.Dispatch("WIA.DeviceManager")
                        for i in range(1, dm.DeviceInfos.Count + 1):
                            info = dm.DeviceInfos.Item(i)
                            if wid is None or info.DeviceID == wid:
                                return info.Connect()
                        return None

                    def _setp(obj, pid, val):
                        try:
                            for p in obj.Properties:
                                if int(p.PropertyID) == pid:
                                    p.Value = val
                                    return True
                        except Exception:
                            pass
                        return False

                    def _valid_dpi(it, want):
                        try:
                            for p in it.Properties:
                                if int(p.PropertyID) == 6147:
                                    vals = [int(x) for x in p.SubType.Values]
                                    if vals:
                                        return min(vals, key=lambda v: abs(v - want))
                        except Exception:
                            pass
                        return want

                    def _try_transfers(item, n=3):
                        pages = 0; err = None
                        for k in range(n):
                            try:
                                try:
                                    item.Transfer(WIA_FMT_BMP)
                                except Exception:
                                    item.Transfer()
                                pages += 1
                            except Exception as e:
                                err = str(e); break
                        return pages, err

                    # Each strategy: (label, set-3088?, set-3096?, set item props?, plain-only?)
                    strategies = [
                        ("A duplex=5, no pages, no item-props", 5, False, False),
                        ("B duplex=5 + pages=0", 5, True, False),
                        ("C duplex=5 + item props(res/depth)", 5, False, True),
                        ("D duplex=7", 7, False, False),
                        ("E duplex=0x8005(32773)", 0x8005, False, False),
                    ]
                    best = 0
                    for label, dupval, setpages, setitem in strategies:
                        dev = None
                        try:
                            dev = _connect()
                            if dev is None:
                                tr.append("%s -> connect fail" % label); continue
                            ok88 = _setp(dev, 3088, dupval)
                            if setpages:
                                _setp(dev, 3096, 0)
                            it = dev.Items[1]
                            if setitem:
                                _setp(it, 4104, 2)  # grayscale
                                dv = _valid_dpi(it, 200)
                                _setp(it, 6147, dv); _setp(it, 6148, dv)
                            pages, err = _try_transfers(it, 3)
                            tr.append("%s -> set88=%s pages=%d err=%s"
                                      % (label, ok88, pages, err or "none"))
                            if pages > best:
                                best = pages
                            if pages >= 1:
                                tr.append(">>> Ye tarika chala! (pages=%d). Baaki skip." % pages)
                                break
                        except Exception as e:
                            tr.append("%s -> EXC %s" % (label, e))
                        finally:
                            try:
                                if dev is not None:
                                    dev = None
                            except Exception:
                                pass
                    got["pages"] = best
                finally:
                    try:
                        _pythoncom.CoUninitialize()
                    except Exception:
                        pass
            else:
                def _count(img):
                    got["pages"] += 1
                try:
                    scan_pages(int(self.winId()), prof.get("source_name"), dpi, pixel, True, _count)
                except Exception as exc:
                    got["err"] = str(exc)
        except Exception as exc:
            got["err"] = str(exc)
        finally:
            QtWidgets.QApplication.restoreOverrideCursor()

        lines = []
        lines.append("===== ApneScan Duplex Test =====")
        lines.append("Time: %s" % datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        lines.append("Scan method: %s" % method)
        lines.append("Device: %s" % (prof.get("source_name") or "(none)"))
        lines.append("Test settings: 1 sheet, Both-side ON, 200 dpi")
        lines.append("")
        lines.append("Pages returned: %d" % got["pages"])
        lines.append("Error: %s" % (got["err"] or "(none)"))
        if got["trace"]:
            lines.append("")
            lines.append("--- Detail (step by step) ---")
            for t in got["trace"]:
                lines.append("  " + t)
        lines.append("")
        if got["pages"] >= 2:
            lines.append("RESULT: DUPLEX WORKING \u2705  (scanner ne dono taraf diye)")
        elif got["pages"] == 1:
            lines.append("RESULT: SINGLE SIDE ONLY \u274c  (scanner ne sirf ek taraf diya)")
            lines.append('Meaning: duplex is not available on this method/driver.')
        else:
            lines.append('RESULT: no page came through (check the scanner/feeder)')
        lines.append("================================")
        report = "\n".join(lines)

        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Duplex Test Result"); dlg.resize(560, 420)
        lay = QtWidgets.QVBoxLayout(dlg)
        lay.addWidget(QtWidgets.QLabel(
            'Copy and share this result:' if self._lang == "hi" else "Copy & share this result:"))
        box = QtWidgets.QPlainTextEdit(); box.setPlainText(report); box.setReadOnly(True)
        box.setStyleSheet("font-family: Consolas, monospace; font-size: 12px;")
        lay.addWidget(box, 1)
        row = QtWidgets.QHBoxLayout()
        b_copy = QtWidgets.QPushButton("Copy" if self._lang == "hi" else "Copy")
        b_copy.clicked.connect(lambda: (QtWidgets.QApplication.clipboard().setText(report),
                                        self.status.showMessage("Copied", 3000)))
        b_close = QtWidgets.QPushButton("Close" if self._lang == "hi" else "Close")
        b_close.clicked.connect(dlg.accept)
        row.addWidget(b_copy); row.addStretch(1); row.addWidget(b_close)
        lay.addLayout(row)
        dlg.exec_()

    def _show_report(self, title, report):
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle(title); dlg.resize(640, 560)
        lay = QtWidgets.QVBoxLayout(dlg)
        lay.addWidget(QtWidgets.QLabel("Copy and share this report:"))
        box = QtWidgets.QPlainTextEdit(); box.setPlainText(report); box.setReadOnly(True)
        box.setStyleSheet("font-family: Consolas, monospace; font-size: 12px;")
        lay.addWidget(box, 1)
        row = QtWidgets.QHBoxLayout()
        b_copy = QtWidgets.QPushButton("Copy")
        def _copy():
            QtWidgets.QApplication.clipboard().setText(report)
            self.status.showMessage("Report copied", 3000)
        b_copy.clicked.connect(_copy)
        b_save = QtWidgets.QPushButton("Save to file")
        def _save():
            p, _ = QtWidgets.QFileDialog.getSaveFileName(
                dlg, "Save report", os.path.join(os.path.expanduser("~"), "ApneScan_escl_test.txt"), "Text (*.txt)")
            if p:
                try:
                    with open(p, "w", encoding="utf-8") as fh:
                        fh.write(report)
                    self.status.showMessage("Saved: %s" % p, 5000)
                except Exception as exc:
                    self._warn("Save failed: %s" % exc)
        b_save.clicked.connect(_save)
        b_close = QtWidgets.QPushButton("Close"); b_close.clicked.connect(dlg.accept)
        row.addWidget(b_copy); row.addWidget(b_save); row.addStretch(1); row.addWidget(b_close)
        lay.addLayout(row)
        dlg.exec_()

    def run_escl_test(self):
        ip = self.ip_field.text().strip() or self._config.get("scanner_ip", "")
        if not ip:
            self._warn("Scanner IP is not set. Enter it in Settings \u2192 Scanner IP (e.g. 192.168.1.8).")
            return
        try:
            self._state_timer.stop()
        except Exception:
            pass
        self.status.showMessage("Running eSCL test... (a few seconds, talking to the scanner)", 0)
        self._escl_tester = EsclTestWorker(ip)
        self._escl_tester.done.connect(self._on_escl_test_done)
        self._escl_tester.start()

    def _on_escl_test_done(self, report):
        try:
            self._state_timer.start()
        except Exception:
            pass
        self.status.clearMessage()
        self._show_report("eSCL Test (network scan jaanch)", report)

    def run_diagnostics(self):
        """Collect scanner + app info and any errors into one report the user
        can copy and share. Runs a safe non-scanning probe of TWAIN and WIA."""
        import platform
        lines = []
        def add(t=""):
            lines.append(t)

        add("===== ApneScan Diagnostics =====")
        add("Version: %s" % VERSION)
        add("Time: %s" % datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        try:
            add("Windows: %s" % platform.platform())
            add("Python: %s (%d-bit)" % (platform.python_version(),
                                         64 if sys.maxsize > 2**32 else 32))
        except Exception:
            pass
        add("Libs: TWAIN=%s  WIA(pywin32)=%s  OCR=%s  Excel=%s  Barcode=%s"
            % (HAS_TWAIN, HAS_W32, HAS_OCR_LIBS, HAS_XLSX, HAS_ZBAR))
        # Tesseract (needed for document-name OCR) — show path + version so we can
        # tell "not installed" apart from "installed but no heading found".
        try:
            tcmd = ""
            try:
                tcmd = pytesseract.pytesseract.tesseract_cmd
            except Exception:
                tcmd = "(default: tesseract)"
            add("Tesseract cmd: %s" % tcmd)
            add("Tesseract cmd exists: %s" % (os.path.exists(tcmd) if (tcmd and tcmd != "tesseract") else "unknown"))
            try:
                add("Tesseract version: %s  -> WORKS" % pytesseract.get_tesseract_version())
            except Exception as _te:
                add("Tesseract version: NOT FOUND (%s)" % type(_te).__name__)
        except Exception:
            add("Tesseract: (pytesseract missing)")
        add("")
        add("Scan method: %s" % self._opts.get("scanner_method", "twain"))
        prof = self._selected_profile()
        add("Profile: %s" % (prof.get("name") if prof else "(none)"))
        add("Device (stored): %s" % (prof.get("source_name") if prof else "(none)"))
        try:
            add("Panel -> Source: %s | Sides: %s | DPI: %s | Depth: %s"
                % (self.cmb_source.currentText(), self.cmb_sides.currentText(),
                   self.cmb_dpi.currentText(), self.cmb_depth.currentText()))
            add("Fast mode: %s" % self.chk_fast.isChecked())
        except Exception:
            pass
        add("Scanner IP: %s" % self.ip_field.text())
        add("")

        # ---- TWAIN probe ----
        add("----- TWAIN sources -----")
        if HAS_TWAIN:
            try:
                names = list_sources(int(self.winId()))
                if names:
                    for n in names:
                        add("  * %s" % n)
                else:
                    add('  (no TWAIN source found)')
            except Exception as exc:
                add("  TWAIN error: %s" % exc)
        else:
            add('  TWAIN (pytwain) is not installed')
        add("")

        # ---- WIA probe + duplex capability ----
        add("----- WIA devices -----")
        if HAS_W32:
            try:
                try:
                    _pythoncom.CoInitialize()
                except Exception:
                    pass
                devs = list_wia_sources()
                if not devs:
                    add('  (no WIA device found)')
                for dev_id, dev_name in devs:
                    add("  * %s" % dev_name)
                    try:
                        dm = _w32.Dispatch("WIA.DeviceManager")
                        device = None
                        for i in range(1, dm.DeviceInfos.Count + 1):
                            info = dm.DeviceInfos.Item(i)
                            if info.DeviceID == dev_id:
                                device = info.Connect(); break
                        if device is not None:
                            for p in device.Properties:
                                try:
                                    pid = int(p.PropertyID)
                                except Exception:
                                    continue
                                if pid == 3086:
                                    add("      Duplex capable (handling caps value): %s" % p.Value)
                                if pid == 3088:
                                    add("      Current handling select: %s" % p.Value)
                    except Exception as exc:
                        add("      (device probe error: %s)" % exc)
            except Exception as exc:
                add("  WIA error: %s" % exc)
            finally:
                try:
                    _pythoncom.CoUninitialize()
                except Exception:
                    pass
        else:
            add('  WIA (pywin32) is not installed')
        add("")
        add("Last scan error: %s" % getattr(self, "_last_error", "(none)"))
        add("================================")

        report = "\n".join(lines)
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Test / Diagnostics"); dlg.resize(640, 560)
        lay = QtWidgets.QVBoxLayout(dlg)
        lay.addWidget(QtWidgets.QLabel(
            'Copy and share this report (scanner + error info):'
            if self._lang == "hi" else "Copy and share this report:"))
        box = QtWidgets.QPlainTextEdit(); box.setPlainText(report); box.setReadOnly(True)
        box.setStyleSheet("font-family: Consolas, monospace; font-size: 12px;")
        lay.addWidget(box, 1)
        row = QtWidgets.QHBoxLayout()
        b_copy = QtWidgets.QPushButton("Copy" if self._lang == "hi" else "Copy")
        def _copy():
            QtWidgets.QApplication.clipboard().setText(report)
            self.status.showMessage("Report copied", 3000)
        b_copy.clicked.connect(_copy)
        b_save = QtWidgets.QPushButton("Save to file" if self._lang == "hi" else "Save to file")
        def _save():
            p, _ = QtWidgets.QFileDialog.getSaveFileName(
                dlg, "Save report", os.path.join(os.path.expanduser("~"), "ApneScan_diagnostics.txt"),
                "Text (*.txt)")
            if p:
                try:
                    with open(p, "w", encoding="utf-8") as fh:
                        fh.write(report)
                    self.status.showMessage("Saved: %s" % p, 5000)
                except Exception as exc:
                    self._warn("Save failed: %s" % exc)
        b_save.clicked.connect(_save)
        b_close = QtWidgets.QPushButton("Close" if self._lang == "hi" else "Close")
        b_close.clicked.connect(dlg.accept)
        row.addWidget(b_copy); row.addWidget(b_save); row.addStretch(1); row.addWidget(b_close)
        lay.addLayout(row)
        dlg.exec_()

    def show_guide(self):
        """Complete guide — app ke SAARE menus ko khud padhkar har option ki
        list banata hai (kahan hai + kya karta hai, Hindi + English). Ye menus
        se auto-ban-ti hai, isliye har update me apne aap up-to-date rehti hai."""
        import html as _html
        rows = []   # (menu_path, label, hi, en)

        def _split_tip(tip):
            hi = en = ""
            for line in (tip or "").split("\n"):
                s = line.strip()
                if s.startswith("हिन्दी:"):
                    hi = s.split(":", 1)[1].strip()
                elif s.startswith("English:"):
                    en = s.split(":", 1)[1].strip()
            if not hi and not en:
                en = (tip or "").strip()
            return hi, en

        def walk(menu, path):
            for act in menu.actions():
                if act.isSeparator():
                    continue
                label = act.text().replace("&", "").replace("❓ ", "").strip()
                sub = act.menu()
                if sub is not None:
                    walk(sub, path + " › " + label)
                    continue
                if not label:
                    continue
                hi, en = _split_tip(act.toolTip())
                rows.append((path, label, hi, en))

        for act in self.menuBar().actions():
            m = act.menu()
            if m is not None:
                walk(m, act.text().replace("&", "").strip())

        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("📖 Complete Guide — ApneScan v%s" % VERSION)
        dlg.resize(780, 660)
        v = QtWidgets.QVBoxLayout(dlg)
        search = QtWidgets.QLineEdit()
        search.setPlaceholderText(self.L("🔍 Koi bhi option dhoondo…", "🔍 Search any option…"))
        search.setClearButtonEnabled(True)
        v.addWidget(search)
        view = QtWidgets.QTextBrowser()
        view.setOpenExternalLinks(True)
        v.addWidget(view, 1)

        def render(q=""):
            q = (q or "").strip().lower()
            out = ["<style>h3{color:#0f766e;margin:16px 0 6px;font-size:15px;}"
                   ".opt{margin:7px 0;padding:6px 10px;border-left:3px solid #99f6e4;background:#f8fafc;}"
                   ".lbl{font-weight:700;color:#0f172a;}.loc{color:#94a3b8;font-size:11px;}"
                   ".hi{color:#334155;}.en{color:#475569;}</style>"]
            cur = None
            cnt = 0
            for path, label, hi, en in rows:
                hay = (path + " " + label + " " + hi + " " + en).lower()
                if q and q not in hay:
                    continue
                top = path.split(" › ")[0]
                if top != cur:
                    cur = top
                    out.append("<h3>📂 %s</h3>" % _html.escape(top))
                desc = (en if self._lang == "en" else hi) or en or hi or "—"
                out.append("<div class='opt'><span class='lbl'>%s</span> "
                           "<span class='loc'>— %s</span><br>"
                           "<span class='en'>%s</span></div>"
                           % (_html.escape(label), _html.escape(path),
                              _html.escape(desc)))
                cnt += 1
            out.insert(1, "<p style='color:#64748b;'>%d options</p>" % cnt)
            if not cnt:
                out.append("<p>Nothing found.</p>")
            view.setHtml("".join(out))
        search.textChanged.connect(render)
        render()
        v.addWidget(QtWidgets.QLabel(
            "<span style='color:#94a3b8;font-size:11px;'>This guide is built "
            "automatically from the app's menus — it stays up to date on every "
            "update. / यह सूची menus से खुद बनती है, हर update में up-to-date रहती है।</span>"))
        brow = QtWidgets.QHBoxLayout()
        bpdf = QtWidgets.QPushButton(self.L("📄 PDF में सेव करें", "📄 Save as PDF"))
        bpdf.setToolTip(self.L("Poori guide ki ek PDF banao (print/share ke liye)",
                               "Make a PDF of the whole guide (to print or share)"))
        bpdf.clicked.connect(lambda: self._export_doc_to_pdf(
            view.document(), "ApneScan_Complete_Guide.pdf",
            self.L("ApneScan — पूरी गाइड", "ApneScan — Complete Guide")))
        b = QtWidgets.QPushButton(self.L("बंद करें", "Close")); b.clicked.connect(dlg.accept)
        brow.addWidget(bpdf); brow.addStretch(1); brow.addWidget(b)
        v.addLayout(brow)
        dlg.exec_()

    def _export_doc_to_pdf(self, doc, default_name, title=""):
        """Kisi bhi guide (QTextDocument) ko PDF me save karo — Windows ke Nirmala
        font se Hindi sahi aati hai, koi extra font chahiye nahi."""
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, self.L("गाइड PDF में सेव करें", "Save guide as PDF"),
            os.path.join(self._opts.get("save_folder", os.path.expanduser("~")), default_name),
            "PDF (*.pdf)")
        if not path:
            return
        if not path.lower().endswith(".pdf"):
            path += ".pdf"
        try:
            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFormat(QPrinter.PdfFormat)
            printer.setOutputFileName(path)
            try:
                printer.setPageSize(QtGui.QPageSize(QtGui.QPageSize.A4))
                printer.setPageMargins(QtCore.QMarginsF(14, 14, 14, 16),
                                       QtGui.QPageLayout.Millimeter)
            except Exception:
                pass
            if title:
                printer.setDocName(title)
            # ek clone par title jodkar print karo (asli view badle bina)
            d = doc.clone()
            if title:
                cur = QtGui.QTextCursor(d)
                cur.setPosition(0)
                cur.insertHtml("<h1 style='color:#0f766e'>%s</h1>"
                               "<p style='color:#64748b;font-size:11px'>ApneScan v%s "
                               "· apnescan.apnesoft.com</p><hr>" % (title, VERSION))
            d.print_(printer)
            self.status.showMessage(self.L("📄 गाइड PDF सेव हो गई: ", "📄 Guide PDF saved: ")
                                    + os.path.basename(path), 8000)
            try:
                self._open_path(path)
            except Exception:
                pass
        except Exception as e:
            self._warn(str(e))

    def show_about(self):
        QtWidgets.QMessageBox.information(
            self, "About ApneScan",
            "ApneScan\nFree document scanning + PDF tool.\n"
            "TWAIN + WIA scanners support.\nVersion " + VERSION)

    # ---- Panel guides ("?" buttons on Meri Files / Preview) ----
    _GUIDE_CSS = ("<style>h2{color:#0f766e;font-size:17px;margin:2px 0 4px;}"
                  "h3{color:#0f172a;font-size:14px;margin:14px 0 4px;}"
                  "p{color:#334155;margin:4px 0;}"
                  "b{color:#0f172a;} li{color:#334155;margin:3px 0;}"
                  "ul{margin:4px 0 4px 4px;padding-left:18px;}"
                  ".t{color:#64748b;font-size:12px;}</style>")

    def _show_panel_guide(self, title, hi_html, en_html):
        """Ek panel ki poori guide — 2 tabs: हिंदी aur English."""
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle(title)
        dlg.resize(620, 680)
        v = QtWidgets.QVBoxLayout(dlg)
        tabs = QtWidgets.QTabWidget()
        for lang, htmlc in (("हिंदी", hi_html), ("English", en_html)):
            br = QtWidgets.QTextBrowser()
            br.setOpenExternalLinks(True)
            br.setHtml(self._GUIDE_CSS + htmlc)
            tabs.addTab(br, lang)
        # user ki abhi ki bhasha wala tab pehle khol do
        tabs.setCurrentIndex(1 if self._lang == "en" else 0)
        v.addWidget(tabs, 1)
        row = QtWidgets.QHBoxLayout()
        bpdf = QtWidgets.QPushButton(self.L("📄 PDF में सेव करें", "📄 Save as PDF"))
        _fname = "ApneScan_Guide.pdf"
        bpdf.clicked.connect(lambda: self._export_doc_to_pdf(
            tabs.currentWidget().document(), _fname, title))
        b = QtWidgets.QPushButton(self.L("बंद करें", "Close"))
        b.clicked.connect(dlg.accept)
        row.addWidget(bpdf); row.addStretch(1); row.addWidget(b)
        v.addLayout(row)
        dlg.exec_()

    def show_files_guide(self):
        hi = """
<h2>📁 मेरी फ़ाइलें (My Files) — पूरी गाइड</h2>
<p class='t'>दाईं तरफ़ का यह पैनल आपके save-फ़ोल्डर की सभी फ़ोल्डर और फ़ाइलें दिखाता है।
यहीं से आप सीधे किसी फ़ोल्डर में सेव कर सकते हैं, ढूँढ सकते हैं, और फ़ाइलों पर काम कर सकते हैं।</p>

<h3>🧭 इधर-उधर जाना (Navigation)</h3>
<ul>
<li><b>फ़ोल्डर खोलना:</b> किसी फ़ोल्डर पर <b>डबल-क्लिक</b> करें — आप उसके अंदर चले जाएँगे।</li>
<li><b>⬅ पीछे:</b> ऊपर वाले फ़ोल्डर में वापस जाने के लिए।</li>
<li><b>ब्रेडक्रम्ब:</b> पैनल में हरे रंग में लिखा रहता है कि आप अभी किस फ़ोल्डर में हैं।</li>
<li><b>📦 फ़ोल्डर जानकारी:</b> नीचे दिखता है इस फ़ोल्डर में कितनी फ़ाइलें और कुल कितनी जगह।</li>
</ul>

<h3>🔍 ढूँढना (Search) — तुरंत</h3>
<ul>
<li>ऊपर search बॉक्स में <b>2 अक्षर</b> लिखते ही नतीजे तुरंत आ जाते हैं (नाम बीच से भी मिलता है)।</li>
<li>कई शब्द लिखें (जैसे <b>राम बिल</b>) तो सब मिलने चाहिए।</li>
<li><b>📄 बटन</b> दबाकर search करें तो PDF के <b>अंदर के टेक्स्ट</b> में भी ढूँढेगा (मरीज़ का नाम/claim नंबर)।</li>
</ul>

<h3>🎛 छाँटना और दिखाना</h3>
<ul>
<li><b>फ़िल्टर ड्रॉपडाउन:</b> सब / 📕 PDF / 🖼 फ़ोटो / 📅 आज / 📆 हफ़्ता / 📌 Pinned।</li>
<li><b>🕘 Recent:</b> हाल में बनी/बदली फ़ाइलें सबसे ऊपर।</li>
<li><b>⧉ Grid:</b> बड़ी thumbnail view (PDF का पहला पेज दिखता है)।</li>
<li><b>⇅ Sort:</b> नाम/तारीख़/आकार से क्रम — आपकी पसंद सेव रहती है।</li>
<li><b>⭐ Favourites:</b> ड्रॉपडाउन से पसंदीदा फ़ोल्डर पर एक क्लिक में जाएँ; ⭐ बटन से फ़ोल्डर जोड़ें/हटाएँ।</li>
</ul>

<h3>💾 सेव करना और लाना</h3>
<ul>
<li><b>💾 यहाँ सेव करें:</b> अभी स्कैन/import किए पेजों की PDF सीधे <b>इसी खुले फ़ोल्डर</b> में सेव।</li>
<li><b>📥 फ़ाइलें लाओ (Import):</b> कंप्यूटर से PDF/फ़ोटो चुनकर सीधे इस फ़ोल्डर में कॉपी।</li>
<li><b>➕ नया फ़ोल्डर:</b> यहीं नया फ़ोल्डर बनाएँ (Space/Ctrl+S से सेव भी इसी खुले फ़ोल्डर में जाता है)।</li>
<li><b>Drag करके लाना:</b> किसी फ़ाइल को पैनल से खींचकर बीच वाली पेज-लिस्ट में डालें — वह import हो जाएगी।</li>
</ul>

<h3>🖱 राइट-क्लिक मेन्यू (फ़ाइल पर)</h3>
<ul>
<li><b>खोलें</b>, <b>📌 Pin</b> (ऊपर रखो), <b>🟢 WhatsApp</b>, <b>✉ Email</b>, <b>🗜 Compress</b>, <b>🏷 Tag</b></li>
<li><b>🖨 Print</b>, <b>📄 Copy</b> / <b>📁 Move</b> दूसरे फ़ोल्डर में, <b>✏ Rename</b>, <b>🗑 Delete</b></li>
</ul>
<h3>🖱 राइट-क्लिक (फ़ोल्डर पर)</h3>
<ul>
<li><b>💾 यहाँ सेव</b>, <b>📥 Import</b>, <b>➕ नया फ़ोल्डर</b>, <b>🧩 सारी PDF → एक PDF</b>, <b>🗜 ZIP बनाओ</b>, <b>📂 Explorer में खोलो</b>, <b>⭐ Favourite</b></li>
</ul>
<h3>🖱 कई फ़ाइलें चुनकर (Ctrl/Shift से)</h3>
<ul>
<li><b>🧩 एक PDF में जोड़ो</b>, <b>📁 Move</b>, <b>📄 Copy</b>, <b>✏ Bulk rename</b> (naam_001, 002…), <b>🗑 Delete</b></li>
</ul>

<h3>🧹 और सुविधाएँ</h3>
<ul>
<li><b>🔁 Duplicate finder:</b> खाली जगह पर राइट-क्लिक → एक जैसी नक़ल फ़ाइलें ढूँढो।</li>
<li><b>🗑 Recycle Bin:</b> ग़लती से हटी फ़ाइल वापस लाओ या हमेशा के लिए हटाओ।</li>
<li><b>क्लिक करके झलक:</b> किसी फ़ाइल पर एक क्लिक = Preview पैनल में जल्दी झलक।</li>
</ul>
"""
        en = """
<h2>📁 My Files panel — complete guide</h2>
<p class='t'>This right-hand panel shows every folder and document in your save
folder. From here you can save straight into a folder, search, and act on files.</p>

<h3>🧭 Navigation</h3>
<ul>
<li><b>Open a folder:</b> <b>double-click</b> it to go inside.</li>
<li><b>⬅ Back:</b> go up one folder.</li>
<li><b>Breadcrumb:</b> the green line shows which folder you are in now.</li>
<li><b>📦 Folder info:</b> shows the file count and total size of the folder.</li>
</ul>

<h3>🔍 Search — instant</h3>
<ul>
<li>Type <b>2 letters</b> in the search box — results appear instantly (matches anywhere in the name).</li>
<li>Type several words (e.g. <b>ram bill</b>) — all must match.</li>
<li>Turn on the <b>📄 button</b> to also search <b>inside PDF text</b> (a patient name / claim number).</li>
</ul>

<h3>🎛 Filter & view</h3>
<ul>
<li><b>Filter dropdown:</b> All / 📕 PDF / 🖼 Images / 📅 Today / 📆 Week / 📌 Pinned.</li>
<li><b>🕘 Recent:</b> newest files first.</li>
<li><b>⧉ Grid:</b> big-thumbnail view (shows a PDF's first page).</li>
<li><b>⇅ Sort:</b> by name/date/size — your choice is saved.</li>
<li><b>⭐ Favourites:</b> jump to a favourite folder from the dropdown; the ⭐ button adds/removes a folder.</li>
</ul>

<h3>💾 Saving & importing</h3>
<ul>
<li><b>💾 Save here:</b> saves the currently scanned/imported pages as a PDF into <b>the open folder</b>.</li>
<li><b>📥 Import files:</b> pick PDFs/photos from your PC straight into this folder.</li>
<li><b>➕ New folder:</b> create one here (Space/Ctrl+S also saves into the open folder).</li>
<li><b>Drag in:</b> drag a file from the panel onto the middle page-list to import it.</li>
</ul>

<h3>🖱 Right-click on a file</h3>
<ul>
<li><b>Open</b>, <b>📌 Pin</b>, <b>🟢 WhatsApp</b>, <b>✉ Email</b>, <b>🗜 Compress</b>, <b>🏷 Tag</b></li>
<li><b>🖨 Print</b>, <b>📄 Copy</b> / <b>📁 Move</b> to another folder, <b>✏ Rename</b>, <b>🗑 Delete</b></li>
</ul>
<h3>🖱 Right-click on a folder</h3>
<ul>
<li><b>💾 Save here</b>, <b>📥 Import</b>, <b>➕ New folder</b>, <b>🧩 All PDFs → one PDF</b>, <b>🗜 Make ZIP</b>, <b>📂 Open in Explorer</b>, <b>⭐ Favourite</b></li>
</ul>
<h3>🖱 Several files selected (Ctrl/Shift)</h3>
<ul>
<li><b>🧩 Merge into one PDF</b>, <b>📁 Move</b>, <b>📄 Copy</b>, <b>✏ Bulk rename</b> (name_001, 002…), <b>🗑 Delete</b></li>
</ul>

<h3>🧹 More</h3>
<ul>
<li><b>🔁 Find duplicates:</b> right-click empty space → find identical copies.</li>
<li><b>🗑 Recycle Bin:</b> restore a deleted file or remove it for good.</li>
<li><b>Click to preview:</b> one click on a file shows a quick preview in the Preview panel.</li>
</ul>
"""
        self._show_panel_guide(self.L("📁 मेरी फ़ाइलें — गाइड", "📁 My Files — Guide"), hi, en)

    def show_preview_guide(self):
        hi = """
<h2>👁 Preview पैनल — पूरी गाइड</h2>
<p class='t'>बीच की पेज-लिस्ट में किसी पेज पर क्लिक करें — उसकी बड़ी झलक यहाँ आती है,
और आप एक-क्लिक में उसे सुधार सकते हैं। नीचे के बटन उसी पेज पर लगते हैं (या 'सभी पेज' पर, अगर टिक हो)।</p>

<h3>🧭 पेज बदलना और देखना</h3>
<ul>
<li><b>◀ ▶</b> से पिछला/अगला पेज; ऊपर <b>Page x/N</b> दिखता है।</li>
<li><b>Filmstrip:</b> नीचे सभी पेजों की छोटी झलक — किसी पर क्लिक = वही पेज।</li>
<li><b>Zoom:</b> ➖ छोटा · 🔳 Fit · ➕ बड़ा · ⛶ पूरी स्क्रीन (या <b>Ctrl+scroll</b>)।</li>
<li><b>टैब:</b> 👁 झलक · 🔤 Text (OCR)। पूरी जानकारी (आकार, DPI, तारीख़) सीधे preview के नीचे दिखती है।</li>
</ul>

<h3>🔧 एक-क्लिक सुधार</h3>
<ul>
<li><b>घुमाना:</b> ↩️ बाएँ · ↪️ दाएँ · 🎯 किसी भी कोण पर सीधा।</li>
<li><b>✂️ Crop</b> (border अपने-आप) · <b>📐 सीधा</b> (deskew)।</li>
<li><b>रंग/रोशनी:</b> ☀️ उजला · 🌙 गहरा · 🌗 Contrast · ✨ Auto-साफ़ · ⬜ Whiten (backing सफ़ेद)।</li>
<li><b>⬛ B&W</b> · <b>🩶 Gray</b> · <b>🖼️ Restore</b> (पुरानी फ़ोटो) · <b>✍️ Sign/मोहर</b> · <b>🆔 ID अलग</b>।</li>
</ul>

<h3>🎨 Editor (crop/मिटाओ/text/तीर) — सबसे ताक़तवर</h3>
<p>🎨 <b>Editor</b> बटन एक पूरा canvas खोलता है — माउस से सीधे पेज पर:</p>
<ul>
<li><b>✂ Crop</b> (घसीट कर हिस्सा रखो) · <b>🧽 मिटाओ</b> (दाग सफ़ेद) · <b>✍ Text</b> (क्लिक करके लिखो)</li>
<li><b>➡ तीर</b> · <b>🖍 Highlight</b> · <b>📐 Perspective</b> (4 कोने क्लिक → तिरछा सीधा)</li>
<li><b>🔢 Page number</b> · <b>☀◐ Brightness/Contrast slider</b> · <b>↶ Undo</b> · <b>💾 Save</b> (पेज पर लग जाता है)</li>
</ul>

<h3>🛠 और काम (⋯ More मेन्यू)</h3>
<ul>
<li><b>↕ Compare:</b> दो पेज साथ-साथ (पहले–बाद)।</li>
<li><b>🔎 Loupe:</b> माउस के नीचे ज़ूम-कांच।</li>
<li><b>▶ Slideshow:</b> अपने-आप एक-एक पेज।</li>
<li><b>⬆⬇ Move:</b> पेज ऊपर-नीचे (क्रम बदलो)।</li>
<li><b>📤 Save-as</b> इस पेज को अलग JPG/PDF · <b>📋 Copy image</b> (WhatsApp/Word में paste)।</li>
<li><b>🖨 Print</b> सिर्फ़ यह पेज · <b>🟢 Share</b> यह पेज · <b>📭 खाली है क्या?</b></li>
</ul>

<h3>🔤 Text टैब</h3>
<ul>
<li><b>🔤 Text पढ़ो</b> (OCR) · <b>📋 Copy</b> · <b>🌐 Translate</b> (हिंदी/English)।</li>
</ul>

<h3>📌 ज़रूरी बात</h3>
<ul>
<li><b>"सभी पेज पर लगाओ"</b> टिक करने पर नीचे का कोई भी सुधार <b>सारे पेजों</b> पर एक साथ लगेगा।</li>
<li><b>↶ Undo</b> से आख़िरी सुधार वापस; <b>🗑 Delete</b> से यह पेज हटेगा।</li>
</ul>
"""
        en = """
<h2>👁 Preview panel — complete guide</h2>
<p class='t'>Click a page in the middle list — its large preview appears here and
you can fix it in one click. The buttons below act on that page (or on ALL pages
if the toggle is ticked).</p>

<h3>🧭 Move & view</h3>
<ul>
<li><b>◀ ▶</b> previous/next page; <b>Page x/N</b> shows at the top.</li>
<li><b>Filmstrip:</b> thumbnails of every page below — click one to jump.</li>
<li><b>Zoom:</b> ➖ out · 🔳 Fit · ➕ in · ⛶ full screen (or <b>Ctrl+scroll</b>).</li>
<li><b>Tabs:</b> 👁 Preview · 🔤 Text (OCR). Full info (size, DPI, date) shows right below the preview.</li>
</ul>

<h3>🔧 One-click fixes</h3>
<ul>
<li><b>Rotate:</b> ↩️ left · ↪️ right · 🎯 any angle.</li>
<li><b>✂️ Crop</b> (auto border) · <b>📐 Straighten</b> (deskew).</li>
<li><b>Colour/light:</b> ☀️ brighter · 🌙 darker · 🌗 contrast · ✨ auto-enhance · ⬜ whiten backing.</li>
<li><b>⬛ B&W</b> · <b>🩶 Gray</b> · <b>🖼️ Restore</b> (old photo) · <b>✍️ Sign/stamp</b> · <b>🆔 Split ID</b>.</li>
</ul>

<h3>🎨 Editor (crop/erase/text/arrow) — the most powerful</h3>
<p>The 🎨 <b>Editor</b> button opens a full canvas — work directly on the page with the mouse:</p>
<ul>
<li><b>✂ Crop</b> (drag to keep) · <b>🧽 Erase</b> (paint white) · <b>✍ Text</b> (click to type)</li>
<li><b>➡ Arrow</b> · <b>🖍 Highlight</b> · <b>📐 Perspective</b> (click 4 corners → fix skew)</li>
<li><b>🔢 Page number</b> · <b>☀◐ Brightness/Contrast sliders</b> · <b>↶ Undo</b> · <b>💾 Save</b> (applies to the page)</li>
</ul>

<h3>🛠 More (⋯ menu)</h3>
<ul>
<li><b>↕ Compare:</b> two pages side by side.</li>
<li><b>🔎 Loupe:</b> a magnifier under the mouse.</li>
<li><b>▶ Slideshow:</b> auto-advance through pages.</li>
<li><b>⬆⬇ Move:</b> move the page up/down (reorder).</li>
<li><b>📤 Save-as</b> this page as JPG/PDF · <b>📋 Copy image</b> (paste into WhatsApp/Word).</li>
<li><b>🖨 Print</b> only this page · <b>🟢 Share</b> this page · <b>📭 Is it blank?</b></li>
</ul>

<h3>🔤 Text tab</h3>
<ul>
<li><b>🔤 Read text</b> (OCR) · <b>📋 Copy</b> · <b>🌐 Translate</b> (Hindi/English).</li>
</ul>

<h3>📌 Important</h3>
<ul>
<li>Tick <b>"Apply edits to ALL pages"</b> and any fix below is applied to <b>every page</b> at once.</li>
<li><b>↶ Undo</b> reverts the last edit; <b>🗑 Delete</b> removes this page.</li>
</ul>
"""
        self._show_panel_guide(self.L("👁 Preview — गाइड", "👁 Preview — Guide"), hi, en)

    def toggle_simple_mode(self):
        self._opts["simple_mode"] = bool(self.act_simple.isChecked())
        self._save_opts()
        self._apply_simple_mode()

    def _apply_simple_mode(self):
        simple = bool(self._opts.get("simple_mode"))
        for w in getattr(self, "_adv_btns", []):
            try:
                w.setVisible(not simple)
            except Exception:
                pass


    def show_whatsnew(self):
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle(tr("whatsnew", self._lang)); dlg.resize(560, 460)
        lay = QtWidgets.QVBoxLayout(dlg)
        br = QtWidgets.QTextBrowser(); br.setHtml(CHANGELOG_HTML); lay.addWidget(br)
        b = QtWidgets.QPushButton("Close" if self._lang == "en" else "Close")
        b.clicked.connect(dlg.accept); lay.addWidget(b)
        dlg.exec_()

    def send_feedback(self):
        title = tr("feedback", self._lang)
        text, ok = QtWidgets.QInputDialog.getMultiLineText(
            self, title,
            ("Aapka sujhav ya problem likhein:" if self._lang == "hi"
             else "Write your feedback or problem:"))
        if not ok or not text.strip():
            return
        email = self._opts.get("feedback_email", "").strip()
        if not email:
            email, ok2 = QtWidgets.QInputDialog.getText(
                self, title,
                ("Which email should feedback go to?"
                 if self._lang == "hi" else "Which email should feedback go to?"))
            if not ok2 or not email.strip():
                QtWidgets.QMessageBox.information(self, title, text)
                return
            self._opts["feedback_email"] = email.strip(); self._save_opts()
            email = email.strip()
        try:
            from PyQt5.QtGui import QDesktopServices
            body = QtCore.QUrl.toPercentEncoding(text).data().decode()
            url = QtCore.QUrl("mailto:%s?subject=ApneScan%%20Feedback&body=%s" % (email, body))
            QDesktopServices.openUrl(url)
        except Exception:
            QtWidgets.QMessageBox.information(self, title,
                                              "Email: %s\n\n%s" % (email, text))

    # ---- UI ----
    def _build_ui(self):
        central = QtWidgets.QWidget(); self.setCentralWidget(central)
        outer = QtWidgets.QVBoxLayout(central); outer.setContentsMargins(0, 0, 0, 0); outer.setSpacing(0)

        # ---------- Top toolbar ----------
        tbwrap = QtWidgets.QWidget(); tbwrap.setObjectName("toolbar")
        tb = QtWidgets.QHBoxLayout(tbwrap); tb.setContentsMargins(8, 4, 8, 4); tb.setSpacing(2)
        self._adv_btns = []

        tips = {
            "scan": 'Start scan (F5)', "profiles": 'Create/edit scan profiles',
            "ocr": 'OCR: searchable PDF (click to turn on)', "fast": "Fast: 200 dpi + B&W, sabse tez",
            "import": 'Import an image from the computer', "savepdf": 'Save PDF (arrow: all / selected)',
            "images": 'Save pages as JPG/PNG', "print": 'Print (Ctrl+P)',
            "rotate": 'Rotate the selected page', "up": 'Move page up', "down": 'Move page down',
            "delete": 'Delete the selected page (Ctrl+Z to undo)', "clear": 'Remove all pages',
            "language": 'Change language (Hindi/English)', "about": 'About the app',
            "guide": self.L("📖 Complete Guide — har option kya karta hai (F1)",
                            "📖 Complete Guide — what every option does (F1)"),
        }
        def tbtn(kind, label, fn, advanced=False, checkable=False):
            b = QtWidgets.QToolButton()
            b.setToolTip(tips.get(kind, label))
            b.setToolButtonStyle(QtCore.Qt.ToolButtonTextUnderIcon)
            b.setIcon(_make_icon(kind)); b.setIconSize(QtCore.QSize(28, 28))
            b.setText(label); b.setAutoRaise(True); b.setMinimumWidth(62)
            if checkable:
                b.setCheckable(True)
                b.toggled.connect(fn) if fn else None
            else:
                b.clicked.connect(fn)
            tb.addWidget(b)
            if advanced:
                self._adv_btns.append(b)
            return b

        self.btn_scan_top = tbtn("scan", tr("scan", self._lang), self.do_scan)
        self.btn_profiles = tbtn("profiles", tr("profiles", self._lang).replace("…", ""), self.open_profiles, advanced=True)
        self.chk_ocr = tbtn("ocr", "OCR", None, checkable=True)
        if not HAS_OCR_LIBS:
            self.chk_ocr.setEnabled(False)
        self.chk_fast = tbtn("fast", tr("fast", self._lang), None, checkable=True)
        self.chk_fast.setToolTip("Fast: 200 dpi + Black & White, no extra processing")
        self.chk_fast.setChecked(bool(self._opts.get("fast_mode")))
        self.chk_fast.toggled.connect(self._on_fast_toggled)
        tb.addWidget(self._vsep())
        self.btn_import = tbtn("import", tr("import", self._lang), self.import_images, advanced=True)
        tbtn("import", self.L("Camera", "Camera"), self.scan_from_camera, advanced=True)
        self.btn_save_pdf = tbtn("savepdf", tr("save_pdf", self._lang), self.save_pdf)
        tbtn("images", "Save Images", self.save_images, advanced=True)
        self.btn_print = tbtn("print", "Print", self.print_all, advanced=True)
        tb.addWidget(self._vsep())
        tbtn("rotate", "Rotate", self.rotate_right, advanced=True)
        tbtn("up", tr("up", self._lang), self.move_up, advanced=True)
        tbtn("down", tr("down", self._lang), self.move_down, advanced=True)
        tbtn("delete", tr("delete", self._lang), self.delete_page, advanced=True)
        tbtn("clear", tr("clear", self._lang), self.clear_all, advanced=True)
        tb.addStretch(1)
        tbtn("guide", self.L("Guide", "Guide"), self.show_guide)
        tbtn("language", "Language", self.choose_language)
        tbtn("about", tr("about", self._lang), self.show_about)
        pdfmenu = QtWidgets.QMenu(self.btn_save_pdf); pdfmenu.setToolTipsVisible(True)
        self._ma(pdfmenu, tr("save_all", self._lang), self.save_pdf_all,
                 "हिन्दी: सभी पेजों की एक PDF बनाओ।\nEnglish: Save all pages as one PDF.")
        self._ma(pdfmenu, tr("save_sel", self._lang), self.save_pdf_selected,
                 "हिन्दी: सिर्फ़ चुने हुए पेजों की PDF।\nEnglish: PDF of only the selected pages.")
        self.btn_save_pdf.setMenu(pdfmenu)
        self.btn_save_pdf.setPopupMode(QtWidgets.QToolButton.MenuButtonPopup)
        printmenu = QtWidgets.QMenu(self.btn_print); printmenu.setToolTipsVisible(True)
        self._ma(printmenu, "All print (sabhi pages)", self.print_all,
                 "हिन्दी: सारे पेज प्रिंट करो।\nEnglish: Print all pages.")
        self._ma(printmenu, "Selected print (chune hue)", self.print_selected,
                 "हिन्दी: सिर्फ़ चुने हुए पेज प्रिंट करो।\nEnglish: Print only the selected pages.")
        self._ma(printmenu, "ID print (2 ID ek page par)", self.print_ids,
                 "हिन्दी: ID कार्ड 2-per-A4 प्रिंट (काग़ज़ बचेगा)।\nEnglish: Print IDs two per A4 sheet.")
        self._ma(printmenu, "ID print - sirf selected", self.print_ids_selected,
                 "हिन्दी: सिर्फ़ चुने हुए ID 2-per-page।\nEnglish: Selected IDs, two per sheet.")
        self.btn_print.setMenu(printmenu)
        self.btn_print.setPopupMode(QtWidgets.QToolButton.MenuButtonPopup)
        self._classic_toolbar = tbwrap
        outer.addWidget(tbwrap)
        # ---- UI #1: Ribbon toolbar (MS-Office jaisa — tabs me buttons) ----
        self.ribbon = QtWidgets.QTabWidget()
        self.ribbon.setObjectName("ribbon")
        self.ribbon.setMaximumHeight(96)

        def _ribbon_tab(pairs):
            w = QtWidgets.QWidget()
            h = QtWidgets.QHBoxLayout(w)
            h.setContentsMargins(8, 4, 8, 4); h.setSpacing(4)
            for icon, text, fn in pairs:
                b = QtWidgets.QToolButton()
                b.setToolButtonStyle(QtCore.Qt.ToolButtonTextUnderIcon)
                b.setText("%s\n%s" % (icon, text))
                b.setAutoRaise(True); b.setMinimumWidth(66); b.setMinimumHeight(56)
                b.clicked.connect(fn)
                h.addWidget(b)
            h.addStretch(1)
            return w
        L = self.L
        self.ribbon.addTab(_ribbon_tab([
            ("🖨", L("Scan", "Scan"), self.do_scan),
            ("📷", "Camera", self.scan_from_camera),
            ("📥", "Import", self.import_images),
            ("🖼", "Photo", self.import_photos),
            ("💾", "Save PDF", self.save_pdf_all),
            ("🖨️", "Print", self.print_pages)]), L("🏠 Home", "🏠 Home"))
        self.ribbon.addTab(_ribbon_tab([
            ("↺", L("Ghumao", "Rotate"), self.rotate_left),
            ("✂", "Crop", self.autocrop_current),
            ("✒", "Sign", self.place_sign),
            ("🔤", "OCR text", self.copy_page_text),
            ("🗑", "Delete", self.delete_page),
            ("↩", "Undo", self.undo_delete)]), L("✏ Edit", "✏ Edit"))
        self.ribbon.addTab(_ribbon_tab([
            ("🗜", "Compress", self.compress_pdf_tool),
            ("🧩", "Merge", self.merge_pdfs),
            ("✂", "Split", self.split_pdfs),
            ("🔍", "Search", self.search_pdfs),
            ("📷", "Cards", self.business_cards)]), L("🧰 Tools", "🧰 Tools"))
        self.ribbon.addTab(_ribbon_tab([
            ("🟢", "WhatsApp", self.share_whatsapp),
            ("✉", "Email", self.share_email)]), L("📤 Share", "📤 Share"))
        self.ribbon.setVisible(bool(self._opts.get("ui_ribbon", False)))
        if self._opts.get("ui_ribbon"):
            tbwrap.hide()
        outer.addWidget(self.ribbon)
        # ---- Toolbar ke NEECHE ek patli shortcut-line (DPI F-keys + Enter/Space) ----
        self.lbl_shortcuts = QtWidgets.QLabel("")
        self.lbl_shortcuts.setObjectName("scutline")
        self.lbl_shortcuts.setTextFormat(QtCore.Qt.RichText)
        self.lbl_shortcuts.setStyleSheet(
            "#scutline{background:#f1f5f9;border-bottom:1px solid #e2e8f0;"
            "color:#475569;font-size:11px;padding:3px 12px;}")
        self.lbl_shortcuts.setToolTip(self.L(
            "Ye shortcuts Settings → Keyboard Shortcuts se badal sakte ho — yahan "
            "wahi dikhega jo aapne set kiya.",
            "Change these in Settings → Keyboard Shortcuts — this line always shows "
            "your current keys."))
        outer.addWidget(self.lbl_shortcuts)
        # ---- UI #7: Status-header card (toolbar ke neeche patli smart patti) ----
        self.ui_header = QtWidgets.QWidget()
        self.ui_header.setObjectName("uiheader")
        self.ui_header.setStyleSheet(
            "#uiheader{background:#f0fdfa;border-bottom:1px solid #ccfbf1;}"
            "#uiheader QLabel{font-size:12px;color:#115e59;}")
        _hb = QtWidgets.QHBoxLayout(self.ui_header)
        _hb.setContentsMargins(10, 3, 10, 3)
        # Branding: hospital/clinic ka logo + naam (customize se set hota hai)
        self.hdr_logo = QtWidgets.QLabel("")
        self.hdr_brand = QtWidgets.QLabel("")
        self.hdr_brand.setStyleSheet("font-weight:700;font-size:13px;")
        self.hdr_scanner = QtWidgets.QLabel("●")
        self.hdr_profile = QtWidgets.QLabel("")
        self.hdr_today = QtWidgets.QLabel("")
        _hb.addWidget(self.hdr_logo)
        _hb.addWidget(self.hdr_brand)
        _hb.addWidget(self.hdr_scanner)
        _hb.addWidget(self.hdr_profile)
        _hb.addStretch(1)
        _hb.addWidget(self.hdr_today)
        self._apply_branding()
        self.ui_header.setVisible(bool(self._opts.get("ui_header", True)))
        outer.addWidget(self.ui_header)
        # ---- UI #10: Job-chips bar (ek click me profile+folder+naming set) ----
        self.jobs_bar = QtWidgets.QWidget()
        self.jobs_bar.setObjectName("jobsbar")
        self.jobs_bar.setStyleSheet("#jobsbar{background:#fffbeb;border-bottom:1px solid #fde68a;}")
        self._jobs_lay = QtWidgets.QHBoxLayout(self.jobs_bar)
        self._jobs_lay.setContentsMargins(10, 4, 10, 4)
        self._jobs_lay.setSpacing(6)
        self.jobs_bar.setVisible(bool(self._opts.get("ui_jobs", False)))
        outer.addWidget(self.jobs_bar)
        self._rebuild_jobs_bar()
        hr = QtWidgets.QFrame(); hr.setObjectName("hr"); hr.setFrameShape(QtWidgets.QFrame.HLine); outer.addWidget(hr)

        # ---------- Body: left settings panel | thumbnails ----------
        body = QtWidgets.QHBoxLayout(); body.setContentsMargins(0, 0, 0, 0); body.setSpacing(0)
        self._body_layout = body   # files-panel side swap ke liye reference
        panel = QtWidgets.QWidget(); panel.setObjectName("panel"); panel.setFixedWidth(252)
        self.left_panel = panel
        if not self._opts.get("show_left_panel", True):
            panel.hide()
        pl = QtWidgets.QVBoxLayout(panel); pl.setContentsMargins(14, 14, 14, 14); pl.setSpacing(5)

        pl.addWidget(QtWidgets.QLabel(tr("profile", self._lang)))
        prow = QtWidgets.QHBoxLayout(); prow.setSpacing(4)
        self.cmb_profile = QtWidgets.QComboBox(); self.cmb_profile.currentTextChanged.connect(self._on_profile_changed)
        prow.addWidget(self.cmb_profile, 1)
        bnew = QtWidgets.QPushButton("+"); bnew.setFixedWidth(30); bnew.clicked.connect(self._quick_new_profile); prow.addWidget(bnew)
        bedit = QtWidgets.QPushButton("✎"); bedit.setFixedWidth(30); bedit.clicked.connect(self._quick_edit_profile); prow.addWidget(bedit)
        pw = QtWidgets.QWidget(); pw.setLayout(prow); pl.addWidget(pw)

        pl.addSpacing(4); pl.addWidget(QtWidgets.QLabel("Device:"))
        self.dev_lbl = QtWidgets.QLabel(self._opts.get("scanner_name") or "(no device)")
        self.dev_lbl.setObjectName("dev"); self.dev_lbl.setWordWrap(True)
        pl.addWidget(self.dev_lbl)
        # 🔄 Scanner badlo (kai scanner ho to) — auto-detect list dikhata hai
        self.btn_change_dev = QtWidgets.QPushButton(
            self.L("🔄 Scanner badlo / dhoondo", "🔄 Change / find scanner"))
        self.btn_change_dev.setToolTip(self.L(
            "Sabhi scanner (LAN + USB) KHUD dhoondh kar list dikhata hai — kai "
            "scanner ho to yahan se badlo.",
            "Auto-detects all scanners (LAN + USB) and lists them — switch here "
            "if you use more than one."))
        self.btn_change_dev.clicked.connect(lambda: self.pick_scanner_dialog())
        pl.addWidget(self.btn_change_dev)
        self.method_lbl = QtWidgets.QLabel(""); self.method_lbl.setObjectName("dev")
        pl.addWidget(self.method_lbl)

        pl.addSpacing(4); pl.addWidget(QtWidgets.QLabel("Claim No.:"))
        self.claim_edit = QtWidgets.QLineEdit(); self.claim_edit.setPlaceholderText("optional"); pl.addWidget(self.claim_edit)

        pl.addWidget(QtWidgets.QLabel("Paper source:"))
        self.cmb_source = QtWidgets.QComboBox(); self.cmb_source.addItems(["Feeder (ADF)", "Glass (Flatbed)"]); pl.addWidget(self.cmb_source)
        pl.addWidget(QtWidgets.QLabel("Scan sides:"))
        self.cmb_sides = QtWidgets.QComboBox()
        self.cmb_sides.addItems(["Single side", "Both sides (duplex)"])
        self.cmb_sides.setToolTip("Both side = scan both sides of the paper (duplex)")
        pl.addWidget(self.cmb_sides)
        pl.addWidget(QtWidgets.QLabel("Page size:"))
        self.cmb_pagesize = QtWidgets.QComboBox(); self.cmb_pagesize.addItems(["Auto (detect each page's size)", "A4 (210x297 mm)", "Letter", "Legal", "A5"]); pl.addWidget(self.cmb_pagesize)
        self.cmb_pagesize.setToolTip("Auto = detect each page's real size (mixed sizes / ID card / half page too). A4/Letter/Legal = fixed size.")
        pl.addWidget(QtWidgets.QLabel("Resolution:"))
        self.cmb_dpi = QtWidgets.QComboBox(); self.cmb_dpi.addItems([d + " dpi" for d in RESOLUTIONS])
        self.cmb_dpi.addItem(self.L("Custom… (apni dpi)", "Custom…"))
        self.cmb_dpi.setCurrentText("200 dpi")
        self.cmb_dpi.currentTextChanged.connect(self._on_panel_dpi_changed)
        pl.addWidget(self.cmb_dpi)
        pl.addWidget(QtWidgets.QLabel("Bit depth:"))
        self.cmb_depth = QtWidgets.QComboBox(); self.cmb_depth.addItems(["24-bit Colour", "Grayscale", "Black & White"]); self.cmb_depth.setCurrentText("Grayscale"); pl.addWidget(self.cmb_depth)

        pl.addStretch(1)
        # UPDATE BANNER: naya version website par aate hi yahan dikhta hai —
        # ek click me download + install + restart (chhupa rehta hai warna).
        self.update_box = QtWidgets.QPushButton()
        self.update_box.setObjectName("updatebox")
        self.update_box.setStyleSheet(
            "#updatebox{background:#f59e0b; color:#1f2937; font-weight:700;"
            " border:none; border-radius:8px; padding:10px; font-size:12px;}"
            "#updatebox:hover{background:#d97706; color:#fff;}"
            "#updatebox:disabled{background:#fbbf24; color:#6b7280;}")
        self.update_box.clicked.connect(self._sidebar_update_clicked)
        self.update_box.hide()
        pl.addWidget(self.update_box)
        self.stats_box = QtWidgets.QLabel()
        self.stats_box.setTextFormat(QtCore.Qt.RichText)
        self.stats_box.setObjectName("statsbox")
        self.stats_box.setStyleSheet(
            "#statsbox{border:1px solid #cbd5e1; border-radius:8px; padding:8px;"
            " color:#334155; font-size:12px; background:#f8fafc;}")
        self.stats_box.setToolTip("Click \u2014 opens the full Stats Dashboard.\n"
                                  "What to show: Settings \u2192 Choose sidebar stats")
        self.stats_box.setCursor(QtCore.Qt.PointingHandCursor)
        self.stats_box.hide()                 # purana box band
        self.side_graph = SparkBars([0] * 7)
        self.side_graph.hide()
        # NAYA analytics card (personal + worldwide)
        self.an_box = QtWidgets.QLabel()
        self.an_box.setTextFormat(QtCore.Qt.RichText)
        self.an_box.setStyleSheet(
            "border:1px solid #cbd5e1;border-radius:8px;padding:8px;"
            "color:#334155;font-size:12px;background:#f8fafc;")
        self.an_box.setToolTip(self.L("Click karo — poora Analytics khulega",
                                      "Click for full Analytics"))
        self.an_box.setCursor(QtCore.Qt.PointingHandCursor)
        self.an_box.mousePressEvent = lambda _e: self.show_analytics()
        pl.addWidget(self.an_box)
        self._an_world = {}
        self._an_update_box()
        # ---- Speed presets: 1 click me poori scan-setting ----
        _spd = QtWidgets.QHBoxLayout(); _spd.setSpacing(4)
        for _sk, _st, _tip in (
                ("fast", self.L("⚡ Jaldi", "⚡ Fast"),
                 self.L("150 dpi · Black & White — sabse tez",
                        "150 dpi · Black & White — fastest")),
                ("normal", self.L("⚖ Normal", "⚖ Normal"),
                 self.L("200 dpi · Grayscale — roz ke liye",
                        "200 dpi · Grayscale — everyday")),
                ("best", self.L("★ Badhiya", "★ Best"),
                 self.L("300 dpi · Colour + saaf — sabse achhi quality",
                        "300 dpi · Colour + clean — best quality"))):
            _b = QtWidgets.QPushButton(_st); _b.setToolTip(_tip)
            _b.setStyleSheet("QPushButton{font-size:11px;padding:5px 2px;border:1px solid "
                             "#cbd5e1;border-radius:8px;background:#fff;}"
                             "QPushButton:hover{border-color:#0f766e;color:#0f766e;}")
            _b.clicked.connect(lambda _c, k=_sk: self._apply_speed_preset(k))
            _spd.addWidget(_b)
        _spdw = QtWidgets.QWidget(); _spdw.setLayout(_spd); pl.addWidget(_spdw)
        self.btn_scan = QtWidgets.QPushButton("▶  " + tr("scan", self._lang)); self.btn_scan.setObjectName("primary")
        self.btn_scan.setMinimumHeight(38); self.btn_scan.clicked.connect(self.do_scan); pl.addWidget(self.btn_scan)
        self.btn_scan.setToolTip("Start scan (F5)")
        self.claim_edit.setToolTip("Claim/Patient number (appears in the file name)")
        self.cmb_dpi.setToolTip("Resolution: lower dpi = faster scan")
        self.cmb_depth.setToolTip("Black & White is fastest, Colour is slower")
        self.setAcceptDrops(True)
        body.addWidget(panel)

        vline = QtWidgets.QFrame(); vline.setObjectName("hr"); vline.setFrameShape(QtWidgets.QFrame.VLine); body.addWidget(vline)

        self.list = PagesList(lambda files: self._start_import(files, "normal"))
        self.list.setViewMode(QtWidgets.QListView.IconMode)
        self.list.setIconSize(QtCore.QSize(self.THUMB_W, self.THUMB_H))
        self.list.setResizeMode(QtWidgets.QListView.Adjust)
        self.list.setMovement(QtWidgets.QListView.Snap)
        self.list.setDragDropMode(QtWidgets.QAbstractItemView.InternalMove)
        self.list.setSpacing(10)
        # uniformItemSizes OFF — warna ek chhota (jaise ultrasound) document aane
        # par Qt SAB thumbnails ko chhota kar deta tha. Ab har thumbnail apni
        # natural size me (fixed cell me) dikhega: chhota chhota, poora poora,
        # aur naam bhi neeche saaf dikhega.
        self.list.setUniformItemSizes(False)
        self.list.setWordWrap(True)          # naam poora dikhe (kate nahi, 2 line me)
        self.list.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self._thumb_w, self._thumb_h = self.THUMB_W, self.THUMB_H   # zoomable display size
        self._apply_thumb_zoom(self.THUMB_W)
        # empty-state hint (shown when no pages)
        self._empty_lbl = QtWidgets.QLabel(
            '⬇  Press Scan, or add an image via Import\n\n(or drag an image here)'
            if self._lang == "hi" else
            "\u2b07  Press Scan, or Import an image\n\n(or drag an image here)",
            self.list.viewport())
        self._empty_lbl.setAlignment(QtCore.Qt.AlignCenter)
        self._empty_lbl.setStyleSheet("color:#94a3b8; font-size:15px;")
        # ---- UI #2: Start-dashboard (khaali screen par bade action-cards) ----
        self._dash = QtWidgets.QWidget(self.list.viewport())
        _dv = QtWidgets.QVBoxLayout(self._dash)
        _dv.addStretch(1)
        _dt = QtWidgets.QLabel(self.L("Kya karna hai?", "What would you like to do?"))
        _dt.setAlignment(QtCore.Qt.AlignCenter)
        _dt.setStyleSheet("color:#64748b;font-size:17px;font-weight:600;")
        _dv.addWidget(_dt)
        _dr = QtWidgets.QHBoxLayout()
        _dr.addStretch(1)

        def _dbtn(icon, text, slot):
            b = QtWidgets.QPushButton("%s\n%s" % (icon, text))
            b.setMinimumSize(118, 84)
            b.setStyleSheet("QPushButton{font-size:13px;font-weight:700;border:1px solid "
                            "#cbd5e1;border-radius:12px;background:#fff;padding:8px;}"
                            "QPushButton:hover{border-color:#0f766e;color:#0f766e;}")
            b.clicked.connect(slot)
            _dr.addWidget(b)
        _dbtn("🖨", self.L("Scan karo", "Scan"), self.do_scan)
        _dbtn("📥", "Import", self.import_images)
        _dbtn("📷", "Photo→PDF", self.import_photos)
        _dbtn("🕘", "History", self.show_history)
        _dr.addStretch(1)
        _dv.addLayout(_dr)
        self._dash_recent = QtWidgets.QLabel("")
        self._dash_recent.setAlignment(QtCore.Qt.AlignCenter)
        self._dash_recent.setStyleSheet("color:#94a3b8;font-size:11px;")
        _dv.addWidget(self._dash_recent)
        _dv.addStretch(1)
        self._dash.hide()
        self.list.viewport().installEventFilter(self)
        self.list.itemDoubleClicked.connect(self._on_thumb_dblclick)
        self.list.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.list.customContextMenuRequested.connect(self._list_context_menu)
        body.addWidget(self.list, 1)

        # ---------- Right sidebar: "Meri Files" (folder list + save-here) ----------
        self.files_panel = QtWidgets.QWidget()
        self.files_panel.setObjectName("panel")
        fp = QtWidgets.QVBoxLayout(self.files_panel)
        fp.setContentsMargins(8, 8, 8, 8)
        _hdr = QtWidgets.QLabel(self.L("📁 <b>Meri Files</b>", "📁 <b>My Files</b>"))
        _hdr.setTextFormat(QtCore.Qt.RichText)
        _hdr.setToolTip("Folders and documents in the save folder — create a new folder here "
                        "and save scanned PDFs straight into it.")
        _hdrrow = QtWidgets.QHBoxLayout(); _hdrrow.setContentsMargins(0, 0, 0, 0)
        _hdrrow.addWidget(_hdr); _hdrrow.addStretch(1)
        _bfhelp = QtWidgets.QToolButton(); _bfhelp.setText("❓")
        _bfhelp.setAutoRaise(True); _bfhelp.setCursor(QtCore.Qt.PointingHandCursor)
        _bfhelp.setToolTip(self.L("Meri Files ki poori guide (Hindi + English)",
                                  "Full My Files guide (Hindi + English)"))
        _bfhelp.clicked.connect(self.show_files_guide)
        _hdrrow.addWidget(_bfhelp)
        fp.addLayout(_hdrrow)
        # Search: kisi bhi folder ke andar naam se dhoondo (folder chuna ho to
        # usi ke andar, warna poore save-folder me)
        _srow = QtWidgets.QHBoxLayout(); _srow.setSpacing(4)
        self.files_search = QtWidgets.QLineEdit()
        self.files_search.setPlaceholderText(
            self.L("🔍 Dhoondo… (chune folder ke andar)", "🔍 Search… (inside selected folder)"))
        self.files_search.setClearButtonEnabled(True)
        _srow.addWidget(self.files_search, 1)
        # 📄 = PDF ke ANDAR ke text me bhi dhoondo (on/off)
        self.btn_search_text = QtWidgets.QToolButton()
        self.btn_search_text.setText("📄")
        self.btn_search_text.setCheckable(True)
        self.btn_search_text.setToolTip(self.L(
            "PDF ke ANDAR likhe text me bhi dhoondo (naam ke alawa) — jaise mareez\n"
            "ka naam ya claim number. On karke phir se search karein.",
            "Also search INSIDE PDF text (not just the name) — e.g. a patient name\n"
            "or claim number. Turn on, then search again."))
        self.btn_search_text.toggled.connect(lambda _c: self._files_search_timer.start())
        _srow.addWidget(self.btn_search_text)
        fp.addLayout(_srow)
        # ⬅ Peeche · ⭐ Favourites (dropdown) · ⇅ Sort
        self.fav_bar = QtWidgets.QHBoxLayout()
        self.fav_bar.setSpacing(4)
        self.btn_panel_back = QtWidgets.QToolButton()
        self.btn_panel_back.setText("⬅")
        self.btn_panel_back.setToolTip(self.L("Peeche (upar wale folder me) jao",
                                              "Back (up one folder)"))
        self.btn_panel_back.clicked.connect(self._panel_back)
        self.btn_panel_back.setEnabled(False)
        self.fav_combo = QtWidgets.QComboBox()
        self.fav_combo.setToolTip(self.L(
            "⭐ Favourite folders — chuno aur foran us folder par jao.\n"
            "Kisi folder ko favourite banane ke liye use chuno phir ⭐ dabao.",
            "⭐ Favourite folders — pick one to jump straight there.\n"
            "To favourite a folder, select it then press ⭐."))
        self.fav_combo.activated.connect(self._on_fav_selected)
        self.fav_star = QtWidgets.QToolButton()
        self.fav_star.setText("⭐")
        self.fav_star.setToolTip(self.L(
            "Chune folder ko favourite banao / hatao",
            "Add / remove the selected folder as a favourite"))
        self.fav_star.clicked.connect(self._fav_star_clicked)
        self.btn_panel_sort = QtWidgets.QToolButton()
        self.btn_panel_sort.setText("⇅")
        self.btn_panel_sort.setToolTip(self.L(
            "List ko apne hisaab se sort karo (naam/date/size). Aapki pasand save rahegi.",
            "Sort the list your way (name/date/size). Your choice is saved."))
        self.btn_panel_sort.setPopupMode(QtWidgets.QToolButton.InstantPopup)
        self.btn_panel_sort.setMenu(self._build_sort_menu())
        self.fav_bar.addWidget(self.btn_panel_back)
        self.fav_bar.addWidget(self.fav_combo, 1)
        self.fav_bar.addWidget(self.fav_star)
        self.fav_bar.addWidget(self.btn_panel_sort)
        fp.addLayout(self.fav_bar)
        # ---- Filter · Recent · Grid (quick controls) ----
        _qrow = QtWidgets.QHBoxLayout(); _qrow.setSpacing(4)
        self.files_filter = QtWidgets.QComboBox()
        for _fv, _ft in (("all", self.L("Sab", "All")),
                         ("pdf", "📕 PDF"),
                         ("img", self.L("🖼 Photo", "🖼 Images")),
                         ("today", self.L("📅 Aaj", "📅 Today")),
                         ("week", self.L("📆 Hafta", "📆 Week")),
                         ("pinned", self.L("📌 Pinned", "📌 Pinned"))):
            self.files_filter.addItem(_ft, _fv)
        self.files_filter.setToolTip(self.L(
            "Dikhaao: sab / sirf PDF / sirf photo / aaj / is hafte / pin ki hui",
            "Show: all / PDF only / images / today / this week / pinned"))
        self.files_filter.currentIndexChanged.connect(lambda _i: self._apply_files_filter())
        _qrow.addWidget(self.files_filter, 1)
        self.btn_recent = QtWidgets.QToolButton(); self.btn_recent.setText("🕘")
        self.btn_recent.setToolTip(self.L("Haal me bani/badli files (naye sabse upar)",
                                          "Recent files (newest first)"))
        self.btn_recent.clicked.connect(self._show_recent_files)
        _qrow.addWidget(self.btn_recent)
        self.btn_grid = QtWidgets.QToolButton(); self.btn_grid.setText("⧉")
        self.btn_grid.setCheckable(True)
        self.btn_grid.setChecked(bool(self._opts.get("files_grid")))
        self.btn_grid.setToolTip(self.L("Grid view — badi thumbnail (PDF ka pehla page)",
                                        "Grid view — big thumbnails (PDF first page)"))
        self.btn_grid.toggled.connect(self._toggle_files_grid)
        _qrow.addWidget(self.btn_grid)
        fp.addLayout(_qrow)
        # Abhi kis folder me ho — chhoti si patti (breadcrumb)
        self.lbl_panel_cwd = QtWidgets.QLabel("")
        self.lbl_panel_cwd.setStyleSheet("color:#0f766e;font-size:11px;font-weight:600;")
        self.lbl_panel_cwd.setWordWrap(True)
        fp.addWidget(self.lbl_panel_cwd)
        # Is folder me kitni files + kul size
        self.lbl_folder_info = QtWidgets.QLabel("")
        self.lbl_folder_info.setStyleSheet("color:#64748b;font-size:11px;")
        fp.addWidget(self.lbl_folder_info)
        _root = self._files_root()
        self.files_model = LibraryModel(self)
        self.files_model.setRootPath(_root)
        self.files_model.setNameFilters(["*.pdf", "*.jpg", "*.jpeg", "*.png",
                                         "*.tif", "*.tiff", "*.docx", "*.xlsx"])
        self.files_model.setNameFilterDisables(False)
        self.files_tree = FilesTree(self._on_pages_dropped)
        self.files_tree.setModel(self.files_model)
        self.files_tree.setRootIndex(self.files_model.index(_root))
        for _c in (1, 2, 3):
            self.files_tree.hideColumn(_c)
        self.files_tree.setHeaderHidden(True)
        self.files_tree.setAnimated(True)
        # Ek baar me sirf ek folder — andar jaane ke liye folder par 2x click
        # (poori tree ek saath nahi khulti). Wapas ke liye ⬅ button.
        self.files_tree.setItemsExpandable(False)
        self.files_tree.setRootIsDecorated(False)
        self.files_tree.setExpandsOnDoubleClick(False)
        self.files_tree.doubleClicked.connect(self._files_tree_open)
        self.files_tree._on_activate = self._files_tree_open   # Enter = folder kholo
        self.files_tree.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.files_tree.customContextMenuRequested.connect(self._files_tree_menu)
        self.files_tree.selectionModel().currentChanged.connect(self._files_sel_changed)
        self._apply_files_sort()          # user ki saved sort-pasand lagao
        self._update_panel_nav()          # breadcrumb + back-button
        self.files_tree.setToolTip("Double-click a folder = go INSIDE it (shows only\n"
                                   "that). Use the ⬅ button to go back.\n"
                                   "Double-click a file = open. Single click = choose for\n"
                                   "save/favourite. Drop pages on a folder = save there.\n"
                                   "Drag a file into the doc area = import. Green = today's file.")
        fp.addWidget(self.files_tree, 1)
        # search ke results (search karte hi tree ki jagah dikhte hain)
        self.files_results = UrlListWidget()      # yahan se file drag = import
        self.files_results.setWordWrap(True)      # naam kate nahi, agli line me
        self.files_results.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.files_results.setUniformItemSizes(False)
        self.files_results.setStyleSheet(
            "QListWidget{outline:0;}"
            "QListWidget::item{padding:5px 4px;border-bottom:1px solid #eef2f7;}"
            "QListWidget::item:selected{background:#e0f2f1;color:#0f172a;}")
        self.files_results.itemDoubleClicked.connect(self._files_result_activated)
        self.files_results.itemActivated.connect(self._files_result_activated)  # Enter = kholo
        self.files_results.itemClicked.connect(self._on_result_clicked)  # 1-click = preview
        self.files_results.hide()
        fp.addWidget(self.files_results, 1)
        self._files_search_timer = QtCore.QTimer(self)
        self._files_search_timer.setSingleShot(True)
        self._files_search_timer.setInterval(60)   # index memory me — lagbhag turant
        self._files_search_timer.timeout.connect(self._run_files_search)
        self.files_search.textChanged.connect(lambda _t: self._files_search_timer.start())
        self._rebuild_fav_bar()
        _row = QtWidgets.QHBoxLayout()
        _bopen = QtWidgets.QPushButton(self.L("📂 Kholo", "📂 Open"))
        _bopen.setToolTip(self.L("Kahin bhi rakha koi bhi folder is panel me kholo (uske andar ke documents dikhenge).",
                                 "Open any folder on your PC in this panel."))
        _bopen.clicked.connect(self.open_existing_folder)
        _bhome = QtWidgets.QPushButton("🏠")
        _bhome.setFixedWidth(34)
        _bhome.setToolTip(self.L("Wapas save-folder par", "Back to the save folder"))
        _bhome.clicked.connect(self.reset_panel_folder)
        _bnew = QtWidgets.QPushButton(self.L("➕ Naya", "➕ New"))
        _bnew.setToolTip(self.L("Chune folder me naya folder banao", "Create a new folder inside the selected one"))
        _bnew.clicked.connect(self.new_library_folder)
        _row.addWidget(_bopen); _row.addWidget(_bhome); _row.addWidget(_bnew)
        fp.addLayout(_row)
        # 📥 Bahar se files SEEDHA chune folder me laao (copy)
        self.btn_import_here = QtWidgets.QPushButton(self.L("📥 Files laao (import)", "📥 Import files"))
        self.btn_import_here.setMinimumHeight(32)
        self.btn_import_here.setToolTip(self.L(
            "PC se PDF/photo/Word/Excel chuno — wo SEEDHA is (chune) folder me aa jaayengi.",
            "Pick PDF/photo/Word/Excel from your PC — they copy straight into the selected folder."))
        self.btn_import_here.clicked.connect(self.import_into_selected_folder)
        fp.addWidget(self.btn_import_here)
        self.btn_save_here = QtWidgets.QPushButton(self.L("💾 Yahan save karo", "💾 Save here"))
        self.btn_save_here.setObjectName("primary")
        self.btn_save_here.setMinimumHeight(34)
        self.btn_save_here.setToolTip(self.L(
            "Abhi jo pages scan/import kiye hain, unki PDF SEEDHA is chune folder me save karo.",
            "Save the currently scanned/imported pages as a PDF straight into the selected folder."))
        self.btn_save_here.clicked.connect(self.save_into_selected_folder)
        fp.addWidget(self.btn_save_here)
        self.files_panel.setFixedWidth(250)
        body.addWidget(self.files_panel)
        if not self._opts.get("show_files_panel", True):
            self.files_panel.hide()

        # ---- UI #3: Preview panel (page click → badi jhalak + quick-edit) ----
        self.preview_panel = QtWidgets.QWidget()
        self.preview_panel.setObjectName("panel")
        self.preview_panel.setFixedWidth(310)
        self._pv_zoom = 1.0
        pv = QtWidgets.QVBoxLayout(self.preview_panel)
        pv.setContentsMargins(8, 8, 8, 8); pv.setSpacing(4)
        # header: ◀ Page x/N ▶
        _nav = QtWidgets.QHBoxLayout()
        _bprev = QtWidgets.QPushButton("◀"); _bprev.setFixedWidth(30)
        _bprev.setToolTip(self.L("Pichhla page", "Previous page"))
        _bprev.clicked.connect(lambda: self._pv_step(-1))
        _bnext = QtWidgets.QPushButton("▶"); _bnext.setFixedWidth(30)
        _bnext.setToolTip(self.L("Agla page", "Next page"))
        _bnext.clicked.connect(lambda: self._pv_step(1))
        self.pv_title = QtWidgets.QLabel(self.L("👁 Preview", "👁 Preview"))
        self.pv_title.setAlignment(QtCore.Qt.AlignCenter)
        self.pv_title.setStyleSheet("font-weight:700;")
        _bpvhelp = QtWidgets.QPushButton("❓"); _bpvhelp.setFixedWidth(30)
        _bpvhelp.setToolTip(self.L("Preview panel ki poori guide (Hindi + English)",
                                   "Full Preview panel guide (Hindi + English)"))
        _bpvhelp.clicked.connect(self.show_preview_guide)
        _nav.addWidget(_bprev); _nav.addWidget(self.pv_title, 1)
        _nav.addWidget(_bnext); _nav.addWidget(_bpvhelp)
        pv.addLayout(_nav)
        # tabs: Preview | Text | Info
        self.pv_tabs = QtWidgets.QTabWidget()
        _p1 = QtWidgets.QWidget(); _p1l = QtWidgets.QVBoxLayout(_p1)
        _p1l.setContentsMargins(0, 0, 0, 0)
        self.pv_scroll = QtWidgets.QScrollArea()
        self.pv_scroll.setWidgetResizable(False)
        self.pv_scroll.setAlignment(QtCore.Qt.AlignCenter)
        self.pv_img = QtWidgets.QLabel()
        self.pv_img.setAlignment(QtCore.Qt.AlignCenter)
        self.pv_img.setStyleSheet("background:#fff;")
        self.pv_scroll.setWidget(self.pv_img)
        self.pv_scroll.setStyleSheet("border:1px solid #cbd5e1;border-radius:8px;background:#fff;")
        _p1l.addWidget(self.pv_scroll, 1)
        # ---------- Attractive, RANG-GROUP wale buttons ----------
        def _grp_qss(border, hov_bg, hov_txt):
            return ("QPushButton{border:1px solid %s;border-radius:10px;"
                    "background:#ffffff;color:#334155;font-size:10px;padding:2px 0;}"
                    "QPushButton:hover{border-color:%s;background:%s;color:%s;}"
                    "QPushButton:pressed{background:%s;}"
                    % (border, hov_txt, hov_bg, hov_txt, hov_bg))
        _GRP = {
            "blue":   _grp_qss("#bfdbfe", "#eff6ff", "#1d4ed8"),
            "green":  _grp_qss("#bbf7d0", "#f0fdf4", "#15803d"),
            "purple": _grp_qss("#e9d5ff", "#faf5ff", "#7e22ce"),
            "slate":  _grp_qss("#e2e8f0", "#f1f5f9", "#475569"),
        }

        def _mkbtn(icon, label, tip, fn, h=44, grp="slate"):
            b = QtWidgets.QPushButton(icon + "\n" + label)
            b.setToolTip(tip); b.setCursor(QtCore.Qt.PointingHandCursor)
            b.setMinimumHeight(h)
            b.setStyleSheet(_GRP.get(grp, _GRP["slate"]))
            b.clicked.connect(fn)
            return b
        self._mk_pv_btn = _mkbtn

        # zoom row
        _zr = QtWidgets.QHBoxLayout(); _zr.setSpacing(4)
        for _ic, _lb, _tip, _fn in (
                ("➖", self.L("Chhota", "Zoom −"), self.L("Zoom kam", "Zoom out"),
                 lambda: self._pv_do_zoom(0.8)),
                ("🔳", self.L("Fit", "Fit"), self.L("Panel me fit", "Fit to panel"),
                 self._pv_fit),
                ("➕", self.L("Bada", "Zoom +"), self.L("Zoom zyada", "Zoom in"),
                 lambda: self._pv_do_zoom(1.25)),
                ("⛶", self.L("Screen", "Full"), self.L("Poori screen", "Full screen"),
                 self._pv_fullscreen)):
            _zr.addWidget(_mkbtn(_ic, _lb, _tip, _fn, h=38, grp="slate"))
        _p1l.addLayout(_zr)
        self.pv_tabs.addTab(_p1, self.L("👁 Jhalak", "👁 Preview"))
        # Text tab
        _p2 = QtWidgets.QWidget(); _p2l = QtWidgets.QVBoxLayout(_p2)
        self.pv_text = QtWidgets.QPlainTextEdit()
        self.pv_text.setReadOnly(True)
        self.pv_text.setPlaceholderText(self.L("Is page ka text yahan padha jayega (OCR)…",
                                               "This page's text (OCR) appears here…"))
        _p2l.addWidget(self.pv_text, 1)
        _tb = QtWidgets.QHBoxLayout()
        _breadt = QtWidgets.QPushButton(self.L("🔤 Text padho", "🔤 Read text"))
        _breadt.clicked.connect(self._pv_read_text)
        _bcopyt = QtWidgets.QPushButton(self.L("📋 Copy", "📋 Copy"))
        _bcopyt.clicked.connect(lambda: QtWidgets.QApplication.clipboard().setText(self.pv_text.toPlainText()))
        _btr = QtWidgets.QPushButton(self.L("🌐 Translate", "🌐 Translate"))
        _btr.clicked.connect(self._pv_translate)
        _tb.addWidget(_breadt); _tb.addWidget(_bcopyt); _tb.addWidget(_btr)
        _p2l.addLayout(_tb)
        self.pv_tabs.addTab(_p2, self.L("🔤 Text", "🔤 Text"))
        pv.addWidget(self.pv_tabs, 1)

        # ---- Filmstrip: sabhi pages ki chhoti jhalak (click = wo page) ----
        self.pv_strip = QtWidgets.QListWidget()
        self.pv_strip.setViewMode(QtWidgets.QListView.IconMode)
        self.pv_strip.setFlow(QtWidgets.QListView.LeftToRight)
        self.pv_strip.setWrapping(False); self.pv_strip.setMovement(QtWidgets.QListView.Static)
        self.pv_strip.setFixedHeight(60); self.pv_strip.setIconSize(QtCore.QSize(38, 48))
        self.pv_strip.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        self.pv_strip.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.pv_strip.setStyleSheet("QListWidget{border:1px solid #e2e8f0;border-radius:6px;"
                                    "background:#fff;}QListWidget::item:selected{"
                                    "background:#e0f2f1;border:1px solid #0f766e;border-radius:4px;}")
        self.pv_strip.itemClicked.connect(self._pv_strip_click)
        pv.addWidget(self.pv_strip)

        # ---- Ek hi "Open editor" button — saare edit tools ab naye
        #      document editor me hain (double-click ya isse khulta hai). ----
        _edit_btn = QtWidgets.QPushButton(self.L("🎨 Open editor", "🎨 Open editor"))
        _edit_btn.setToolTip(self.L(
            "Poora document editor kholo — crop, seedha, saaf, whiten, sign, text… sab kuch.",
            "Open the full document editor — crop, straighten, clean, whiten, sign, text… everything."))
        _edit_btn.setCursor(QtCore.Qt.PointingHandCursor)
        _edit_btn.setStyleSheet(
            "QPushButton{font-size:12px;font-weight:700;padding:7px;border-radius:8px;"
            "border:1px solid #0f766e;background:#0f766e;color:#fff;}"
            "QPushButton:hover{background:#0d5f58;}")
        _edit_btn.clicked.connect(lambda: self._pv_open_image_editor())
        pv.addWidget(_edit_btn)
        # ---- Complete info seedha preview ke NEECHE (alag 'Info' tab nahi) ----
        _infhdr = QtWidgets.QLabel(self.L("ℹ <b>Poori jaankari</b>", "ℹ <b>File info</b>"))
        _infhdr.setStyleSheet("color:#0f766e;font-size:11px;font-weight:700;margin-top:2px;")
        pv.addWidget(_infhdr)
        self.pv_info = QtWidgets.QLabel("")
        self.pv_info.setTextFormat(QtCore.Qt.RichText)
        self.pv_info.setWordWrap(True)
        self.pv_info.setTextInteractionFlags(QtCore.Qt.TextSelectableByMouse)
        self.pv_info.setStyleSheet(
            "color:#475569;font-size:11px;padding:6px 8px;background:#f8fafc;"
            "border:1px solid #e2e8f0;border-radius:8px;")
        pv.addWidget(self.pv_info)
        # 'Info' ab alag tab me nahi — saari jaankari isi label me (upar wale
        # sabhi pv_info2.setText(...) ab seedha isi complete-info par jaate hain).
        self.pv_info2 = self.pv_info
        self.preview_panel.setVisible(bool(self._opts.get("ui_preview", False)))
        body.addWidget(self.preview_panel)
        self.list.currentItemChanged.connect(lambda cur, prev: self._update_preview_panel())
        self.pv_scroll.viewport().installEventFilter(self)   # Ctrl+scroll zoom

        outer.addLayout(body, 1)

        # ---- UI #8: Kiosk overlay (bade buttons — dukaan/touch) ----
        self.kiosk = QtWidgets.QWidget(self)
        self.kiosk.setStyleSheet("background:#f4f6f8;")
        _kg = QtWidgets.QGridLayout(self.kiosk)
        _kg.setContentsMargins(30, 30, 30, 30)
        _kg.setSpacing(16)
        _kbtns = [("🖨", self.L("SCAN", "SCAN"), self.do_scan),
                  ("💾", self.L("SAVE PDF", "SAVE PDF"), self.save_pdf_all),
                  ("🟢", "WHATSAPP", self.share_whatsapp),
                  ("📷", self.L("PHOTO→PDF", "PHOTO→PDF"), self.import_photos),
                  ("🗜", "COMPRESS", self.compress_pdf_tool),
                  ("🖨️", "PRINT", self.print_pages),
                  ("🕘", "HISTORY", self.show_history),
                  ("🚪", self.L("BAND KARO", "EXIT KIOSK"), self.toggle_kiosk)]
        for i, (ic, tx, fn) in enumerate(_kbtns):
            b = QtWidgets.QPushButton("%s\n%s" % (ic, tx))
            b.setMinimumHeight(120)
            b.setStyleSheet("QPushButton{font-size:20px;font-weight:800;border:2px solid "
                            "#0f766e;border-radius:18px;background:#fff;color:#0f766e;}"
                            "QPushButton:hover{background:#0f766e;color:#fff;}")
            b.clicked.connect(fn)
            _kg.addWidget(b, i // 3, i % 3)
        self.kiosk.hide()

        # ---------- Status bar: connection ----------
        self.status = self.statusBar()
        self.lbl_conn = QtWidgets.QLabel(); self.lbl_conn.setTextFormat(QtCore.Qt.RichText)
        self._set_conn_display(None, tr("checking", self._lang))
        self.status.addWidget(self.lbl_conn)

        def _sep():
            s = QtWidgets.QLabel("│"); s.setStyleSheet("color:#cbd5e1;")
            return s

        def _foot(clickfn=None, tip=""):
            l = QtWidgets.QLabel("")
            l.setTextFormat(QtCore.Qt.RichText)
            if tip:
                l.setToolTip(tip)
            if clickfn:
                l.setCursor(QtCore.Qt.PointingHandCursor)
                l.mousePressEvent = lambda _e: clickfn()
            return l

        # LEFT side: folder (click→kholo) · pages/selection
        self.status.addWidget(_sep())
        self.foot_folder = _foot(
            lambda: self._open_path(self._files_root()),
            self.L("Save folder — click karke kholo", "Save folder — click to open"))
        self.status.addWidget(self.foot_folder)
        self.status.addWidget(_sep())
        self.foot_pages = _foot(None, self.L("Pages / selected ki ginti + banne wali PDF ka size",
                                             "Pages / selected count + size of the resulting PDF"))
        self.status.addWidget(self.foot_pages)
        self.foot_last = _foot(
            lambda: self._recent and self._open_path(self._recent[0]),
            self.L("Aakhri save ki hui file — click karke kholo",
                   "Last saved file — click to open"))
        self.status.addWidget(self.foot_last)
        # v74 extra (customize se on/off): aaj ke scan · online · apna sandesh
        self.foot_today = _foot(None, self.L("Aaj kitne pages scan hue",
                                             "Pages scanned today"))
        self.status.addWidget(self.foot_today)
        self.foot_online = _foot(None, self.L("Abhi kitne log online (poori duniya)",
                                              "People online now (worldwide)"))
        self.status.addWidget(self.foot_online)
        self.foot_msg = _foot(None, self.L("Aapka apna sandesh", "Your custom message"))
        self.status.addWidget(self.foot_msg)

        # RIGHT side (permanent): disk · version · busy · scanner · ghadi
        # (Analytics hata diya gaya — 'aaj/streak' footer element bhi hata diya)
        self.foot_disk = _foot(None, self.L("Save-drive par kitni jagah bachi hai",
                                            "Free space on the save drive"))
        self.status.addPermanentWidget(self.foot_disk)
        self.status.addPermanentWidget(_sep())
        self.foot_ver = _foot(lambda: self.check_updates(False),
                              self.L("App version — click = update check",
                                     "App version — click to check for updates"))
        self.foot_ver.setText("v%s" % VERSION)
        self.status.addPermanentWidget(self.foot_ver)
        self.status.addPermanentWidget(_sep())
        self.lbl_busy = QtWidgets.QLabel(); self.lbl_busy.setTextFormat(QtCore.Qt.RichText)
        self.status.addPermanentWidget(self.lbl_busy)
        self._set_busy_display("unknown")
        self.status.addPermanentWidget(_sep())
        self.foot_clock = _foot(None, self.L("Ghadi aur tareekh", "Clock and date"))
        self.status.addPermanentWidget(self.foot_clock)
        # footer ko har 30 sec + zaroori events par refresh karo (ghadi har 20 sec)
        self._foot_timer = QtCore.QTimer(self)
        self._foot_timer.setInterval(20000)
        self._foot_timer.timeout.connect(self._update_footer)
        self._foot_timer.start()
        try:
            self.list.itemSelectionChanged.connect(self._update_footer)
        except Exception:
            pass
        QtCore.QTimer.singleShot(1200, self._update_footer)
        self._state_timer = QtCore.QTimer(self)
        self._state_timer.setInterval(6000)
        self._state_timer.timeout.connect(self._tick_scanner_state)
        self._state_timer.start()
        QtCore.QTimer.singleShot(800, self._tick_scanner_state)
        QtCore.QTimer.singleShot(1400, self._detect_device_name)   # Device me naam dikhao
        # Ctrl+O = koi folder kholo (sidebar me) · Ctrl+V = clipboard se paste
        QtWidgets.QShortcut(QtGui.QKeySequence("Ctrl+O"), self, self.open_existing_folder)
        QtWidgets.QShortcut(QtGui.QKeySequence("Ctrl+V"), self, self._paste_from_clipboard)
        # Tesseract check ko startup par BACKGROUND me warm kar do (ye tesseract.exe
        # subprocess chalata hai ~1-2s) — taaki baad me koi bhi kaam (rename, scan)
        # is check par UI thread par ATKE nahi.
        self._run_bg_quiet(lambda: tesseract_available(), lambda _r: None)
        # NAYA analytics: startup par + har 2 min me worldwide numbers laao,
        # aur online-ping bhejo
        QtCore.QTimer.singleShot(1500, self._an_refresh)
        QtCore.QTimer.singleShot(2500, lambda: self._an_report("ping"))
        self._an_timer = QtCore.QTimer(self)
        self._an_timer.setInterval(120000)
        self._an_timer.timeout.connect(lambda: (self._an_report("ping"), self._an_refresh()))
        self._an_timer.start()
        # Naya version aaya ho to sidebar me banner dikhao — startup par aur
        # phir har 6 ghante (lambi chalti app bhi update dekh legi).
        # (customize se 'auto update-check' band ho to nahi jaanchega)
        QtCore.QTimer.singleShot(4000, lambda: self._opts.get("auto_update_check", True)
                                 and self.check_updates(True))
        self._upd_timer = QtCore.QTimer(self)
        self._upd_timer.setInterval(6 * 3600 * 1000)
        self._upd_timer.timeout.connect(lambda: self._opts.get("auto_update_check", True)
                                        and self.check_updates(True))
        self._upd_timer.start()
        # Mahine ki pehli baar kholne par "Aapka Mahina" summary
        QtCore.QTimer.singleShot(6000, self._maybe_month_wrap)
        # Pehli baar (koi scanner set nahi) — auto-detect khud chala do
        if not self._config.get("auto_detect_done"):
            self._config["auto_detect_done"] = True
            save_config(self._config)
            QtCore.QTimer.singleShot(1500, self.auto_detect_scanner)
        # Kiosk mode band karke exit kiya tha to wapas usi me kholo
        if self._opts.get("ui_kiosk"):
            QtCore.QTimer.singleShot(300, lambda: (self.kiosk.setGeometry(
                self.centralWidget().rect()), self.kiosk.show(), self.kiosk.raise_()))
        # kept (hidden) for the connection feature; IP set via Settings menu
        self.ip_field = QtWidgets.QLineEdit(self._config.get("scanner_ip", "")); self.ip_field.hide()
        self.btn_check = QtWidgets.QPushButton(); self.btn_check.hide()
        self._pb_holder = None
        # v73: saari UI-customize settings (accent, font, panel-widths, thumb,
        # footer, spacing…) startup par ek baar laga do taaki restart ke baad
        # bhi user ki pasand yaad rahe.
        QtCore.QTimer.singleShot(0, self._apply_ui_live)
        # Window ka size/jagah pichhli baar jaisa (agar yaad rakhna on ho)
        if self._opts.get("remember_window", True) and self._opts.get("win_geometry"):
            try:
                self.restoreGeometry(QtCore.QByteArray.fromBase64(
                    self._opts["win_geometry"].encode("ascii")))
            except Exception:
                pass
        if self._opts.get("start_maximized"):
            QtCore.QTimer.singleShot(50, self.showMaximized)


    def _apply_style(self):
        self._apply_base_style()
        # Touch / buzurg mode: sab kuch bada — buttons, text, thumbnails
        if self._opts.get("touch_mode"):
            self.setStyleSheet(self.styleSheet() + """
                QToolButton { font-size:14px; padding:10px 12px; }
                QPushButton { font-size:15px; padding:12px 18px; }
                QMenuBar, QMenu, QLabel, QLineEdit, QComboBox, QSpinBox { font-size:14px; }
                QListWidget { font-size:14px; }
            """)
            try:
                self.list.setIconSize(QtCore.QSize(190, 250))
            except Exception:
                pass

    def _apply_base_style(self):
        if self._opts.get("theme") == "darkpro":
            # UI #4: Dark Pro — gehra premium look, teal glow
            self.setStyleSheet("""
                QMainWindow, QWidget { background:#0b1220; color:#e2e8f0; }
                QMenuBar { background:#0b1220; color:#e2e8f0; }
                QMenuBar::item:selected { background:#132036; }
                QMenu { background:#111c30; color:#e2e8f0; border:1px solid #1e3a5f; }
                QMenu::item:selected { background:#134e4a; color:#5eead4; }
                #toolbar { background:#0d1526; border-bottom:1px solid #1e3a5f; }
                QToolButton { border:1px solid transparent; border-radius:9px; padding:5px 7px; color:#94a3b8; font-size:11px; }
                QToolButton:hover { background:#132036; border-color:#2dd4bf; color:#5eead4; }
                QToolButton:checked { background:#134e4a; border-color:#2dd4bf; color:#5eead4; }
                #panel { background:#0d1526; }
                #uiheader { background:#0d1f1d; border-bottom:1px solid #134e4a; }
                #uiheader QLabel { color:#5eead4; }
                #dev { color:#64748b; font-size:12px; }
                #hr { color:#1e3a5f; }
                QLabel { color:#cbd5e1; }
                QLineEdit, QComboBox, QSpinBox, QPlainTextEdit { background:#111c30; border:1px solid #1e3a5f; border-radius:8px; padding:5px 9px; color:#e2e8f0; }
                QLineEdit:focus, QComboBox:focus { border-color:#2dd4bf; }
                QPushButton { background:#111c30; border:1px solid #1e3a5f; border-radius:8px; padding:6px 12px; color:#e2e8f0; }
                QPushButton:hover { background:#132036; border-color:#2dd4bf; }
                QPushButton#primary { background:qlineargradient(x1:0,y1:0,x2:0,y2:1,stop:0 #14b8a6,stop:1 #0f766e); border:1px solid #2dd4bf; color:#fff; font-weight:700; }
                QPushButton#primary:hover { background:#0d9488; }
                QListWidget { background:#080e1a; border:none; }
                QListWidget::item:selected { background:#134e4a; color:#e2e8f0; }
                QTreeView { background:#0d1526; color:#cbd5e1; border:none; }
                QTreeView::item:selected { background:#134e4a; color:#5eead4; }
                QScrollArea { background:#0d1526; }
                #statsbox { border:1px solid #1e3a5f; border-radius:10px; padding:8px; color:#94a3b8; background:#0d1526; }
            """)
            return
        if self._opts.get("theme") == "dark":
            self.setStyleSheet("""
                QMainWindow, QWidget { background:#0f172a; color:#e2e8f0; }
                QMenuBar { background:#0f172a; color:#e2e8f0; }
                QMenuBar::item:selected { background:#1e293b; }
                QMenu { background:#1e293b; color:#e2e8f0; border:1px solid #334155; }
                QMenu::item:selected { background:#334155; }
                #toolbar { background:#111827; }
                QToolButton { border:1px solid transparent; border-radius:8px; padding:4px 6px; color:#cbd5e1; font-size:11px; }
                QToolButton:hover { background:#1e293b; border-color:#334155; }
                QToolButton:checked { background:#134e4a; border-color:#2dd4bf; color:#5eead4; }
                #panel { background:#111827; }
                #dev { color:#94a3b8; font-size:12px; }
                #hr { color:#334155; }
                QLabel { color:#cbd5e1; }
                QLineEdit, QComboBox, QSpinBox, QPlainTextEdit { background:#1e293b; border:1px solid #475569; border-radius:6px; padding:4px 8px; color:#e2e8f0; }
                QPushButton { background:#1e293b; border:1px solid #475569; border-radius:6px; padding:6px 12px; color:#e2e8f0; }
                QPushButton:hover { background:#334155; }
                QPushButton#primary { background:#0f766e; border:1px solid #0f766e; color:#fff; font-weight:700; }
                QPushButton#primary:hover { background:#115e59; }
                QListWidget { background:#0b1220; border:none; }
                QListWidget::item:selected { background:#134e4a; color:#e2e8f0; }
                QScrollArea { background:#111827; }
            """)
        else:
            self.setStyleSheet("""
                QMainWindow, QWidget { background:#f4f6f8; color:#1f2933; }
                QMenuBar { background:#f4f6f8; }
                #toolbar { background:#ffffff; }
                QToolButton { border:1px solid transparent; border-radius:8px; padding:4px 6px; color:#334155; font-size:11px; }
                QToolButton:hover { background:#eef2f6; border-color:#e2e8f0; }
                QToolButton:checked { background:#d1eae7; border-color:#0f766e; color:#0f766e; }
                #panel { background:#ffffff; }
                #dev { color:#475569; font-size:12px; }
                #hr { color:#e2e8f0; }
                QLabel { color:#334155; }
                QLineEdit, QComboBox, QSpinBox { background:#fff; border:1px solid #cbd5e1; border-radius:6px; padding:4px 8px; }
                QPushButton { background:#fff; border:1px solid #cbd5e1; border-radius:6px; padding:6px 12px; color:#1f2933; }
                QPushButton:hover { background:#eef2f6; }
                QPushButton#primary { background:#0f766e; border:1px solid #0f766e; color:#fff; font-weight:700; }
                QPushButton#primary:hover { background:#115e59; }
                QListWidget { background:#fbfcfd; border:none; }
                QListWidget::item:selected { background:#d1eae7; color:#0f172a; }
            """)


    def _set_conn_display(self, state, message):
        colour = {True: "#16a34a", False: "#dc2626", None: "#9ca3af"}[state]
        dot = '<span style="color:%s; font-size:16px;">&#9679;</span>' % colour
        self.lbl_conn.setText("%s&nbsp; %s" % (dot, message))

    def check_connection(self):
        ip = self.ip_field.text().strip()
        self._config["scanner_ip"] = ip; save_config(self._config)
        if not ip:
            self._set_conn_display(False, self.L("Scanner IP daalein", "Enter scanner IP")); return
        self._set_conn_display(None, self.L("Check ho raha hai… (%s)", "Checking… (%s)") % ip)
        self.btn_check.setEnabled(False)
        self._checker = ConnectionChecker(ip)
        self._checker.result.connect(self._on_conn_result)
        self._checker.finished.connect(lambda: self.btn_check.setEnabled(True))
        self._checker.start()

    def _on_conn_result(self, ok, message):
        self._set_conn_display(bool(ok), message)

    def _set_busy_display(self, kind):
        if kind == "free":
            txt, col = self.L("Scanner FREE (taiyar)", "Scanner FREE (ready)"), "#16a34a"
        elif kind == "busy":
            txt, col = self.L("Scanner BUSY (vyast)", "Scanner BUSY"), "#dc2626"
        else:
            txt, col = "Scanner: --", "#9ca3af"
        self.lbl_busy.setText(
            '<span style="color:%s; font-size:15px;">&#9679;</span>&nbsp;<b>%s</b>' % (col, txt))
        # UI header ki scanner-pill bhi yahi dikhaye
        try:
            self.hdr_scanner.setText(
                '<span style="color:%s;">●</span> <b>%s</b>' % (col, txt))
        except Exception:
            pass

    def _tick_scanner_state(self):
        # IMPORTANT: do NOT poll the scanner over the network here. This HP scanner
        # allows only one eSCL connection at a time, so a background status poll
        # collides with the real scan and makes it fail with HTTP 503. The indicator
        # simply reflects OUR own activity: busy only while we are scanning.
        if getattr(self, "_worker", None) is not None and self._worker.isRunning():
            self._set_busy_display("busy")
        elif getattr(self, "_scanning", False):
            self._set_busy_display("busy")
        else:
            self._set_busy_display("free")

    # ---- list ----
    def _add_item_for_path(self, path, qimg=None, at=None):
        # qimg: background worker se pehle se bana-banaya thumbnail (QImage) —
        # UI thread par bhaari decode/scale nahi karna padta.
        # at: agar diya ho to us position par INSERT karo (warna aakhir me jodo)
        if qimg is not None and not qimg.isNull():
            icon = QtGui.QIcon(QtGui.QPixmap.fromImage(qimg))
        else:
            icon = QtGui.QIcon(self._make_thumb(path))
        # Har page ke neeche kuch na kuch dikhe — jab tak naam nahi aata,
        # 'Page N' dikhta hai (khaali kabhi nahi).
        _lbl = "Page %d" % (self.list.count() + 1)
        item = QtWidgets.QListWidgetItem(icon, _lbl)
        item.setData(QtCore.Qt.UserRole, path); item.setTextAlignment(QtCore.Qt.AlignHCenter)
        # har cell ek jaisa (icon + naam ke liye poori jagah) — chhota doc aane
        # par baaki thumbnails chhote na ho, aur naam hamesha dikhe
        item.setSizeHint(QtCore.QSize(self._thumb_w + 24, self._thumb_h + 52))
        if at is not None and 0 <= at <= self.list.count():
            self.list.insertItem(at, item)
        else:
            self.list.addItem(item)
        self.list.setCurrentItem(item)
        self.list.clearSelection()  # nothing "selected" by default; user picks with Ctrl/Shift
        self._dirty = True
        self._update_status(); self._update_empty_state()
        try:
            self._pv_build_filmstrip()
        except Exception:
            pass
        return item

    THUMB_HI_W = 360
    THUMB_HI_H = 480

    def _make_thumb(self, path):
        # QImageReader ka scaled decode (khaaskar JPEG me) poori image decode
        # karke SmoothTransformation se sikodne se KAI GUNA tez hai — isi se
        # scan/import ke time UI ka atakna band hota hai.
        r = QtGui.QImageReader(path)
        r.setAutoTransform(True)
        sz = r.size()
        if sz.isValid() and sz.width() > 0 and sz.height() > 0:
            s = min(float(self.THUMB_HI_W) / sz.width(),
                    float(self.THUMB_HI_H) / sz.height(), 1.0)
            r.setScaledSize(QtCore.QSize(max(1, int(sz.width() * s)),
                                         max(1, int(sz.height() * s))))
        img = r.read()
        if img.isNull():
            return QtGui.QPixmap(self.THUMB_HI_W, self.THUMB_HI_H)
        return QtGui.QPixmap.fromImage(img)

    def _refresh_item(self, item):
        path = item.data(QtCore.Qt.UserRole)
        item.setIcon(QtGui.QIcon(self._make_thumb(path)))
        if item is self.list.currentItem():
            self._show_preview(item, None)

    def _ordered_paths(self):
        return [self.list.item(i).data(QtCore.Qt.UserRole) for i in range(self.list.count())]

    def _selected_paths(self):
        # selected items, kept in the on-screen (list) order
        sel = set(self.list.row(it) for it in self.list.selectedItems())
        return [self.list.item(i).data(QtCore.Qt.UserRole) for i in range(self.list.count()) if i in sel]

    def _show_preview(self, current, _prev=None):
        # NAPS2-style: no side preview pane; double-click opens a viewer.
        return


    def resizeEvent(self, event):
        super().resizeEvent(event)


    def _apply_speed_preset(self, name):
        """Ek click me poori scan-setting badlo. Kyunki scan ab PROFILE ki
        setting se hota hai, ye preset chune hue profile me hi save hota hai
        (dpi + rang), taaki Enter bhi isi par scan kare."""
        presets = {
            "fast":   (150, "bw",    "150 dpi", "Black & White", False, False),
            "normal": (200, "gray",  "200 dpi", "Grayscale",     False, False),
            "best":   (300, "color", "300 dpi", "24-bit Colour", True,  True),
        }
        p = presets.get(name)
        if not p:
            return
        dpi_n, color, dpi, depth, enhance, clean = p
        try:
            self.chk_fast.setChecked(False)
        except Exception:
            pass
        # profile me save karo (yahi Enter/Scan par lagta hai)
        prof = self._selected_profile()
        if prof is not None:
            prof["dpi"] = dpi_n
            prof["color"] = color
            self._save_profiles()
            self._load_profile_to_panel(prof)     # panel ko profile jaisa dikhao
        self._opts["quality_enhance"] = enhance
        self._opts["clean_edges"] = clean
        self._save_opts()
        self.status.showMessage(self.L(
            "Profile setting: %s · %s" % (dpi, depth),
            "Profile settings: %s · %s" % (dpi, depth)), 4000)

    def do_scan(self, _checked=False, dpi_override=None):
        # Startup guard: agar window abhi-abhi khuli hai to scan mat karo. Isse
        # app ko Enter dabakar launch karne par startup par galat scan-box nahi aata.
        try:
            if self._start_timer.elapsed() < 900:
                return
        except Exception:
            pass
        # Rescan/Insert ka pending placement — sirf tab lagega jab scan sach me
        # shuru ho (guards me return ho gaya to chhup-chaap chhoot jayega).
        _place = getattr(self, "_pending_place", None)
        self._pending_place = None
        self._doc_folder_hint = None   # naya scan = folder-hint reset
        method = self._opts.get("scanner_method", "twain")
        prof = self._selected_profile()
        if method == "escl":
            ip = self.ip_field.text().strip()
            self._opts["scanner_ip"] = ip
            if not ip:
                self._warn("Scanner IP is not set. Enter the IP in Settings \u2192 Scanner IP (e.g. 192.168.1.8)."); return
        elif method == "naps2":
            if not (self._opts.get("naps2_path") or find_naps2()):
                self._warn("NAPS2 not found. Set its path in Settings \u2192 Scan method \u2192 naps2."); return
            if not self._opts.get("naps2_profile"):
                self._warn("Set the NAPS2 profile name in Settings \u2192 Scan method \u2192 naps2."); return
        else:
            if not HAS_TWAIN and method == "twain":
                self._warn("TWAIN (pytwain) is not installed."); return
            if prof is None:
                QtWidgets.QMessageBox.information(self, APP_NAME, "Create a profile first.\n'Profiles…' → New.")
                self.open_profiles(); prof = self._selected_profile()
                if prof is None:
                    return
            if method == "twain" and not prof.get("source_name"):
                self._warn("No scanner is set in this profile. 'Profiles…' → Edit → Choose device."); return
        self._barcode_tried = False
        self._scan_count = 0
        self._progress = ScanProgressDialog(self, prof.get("source_name") or self.L("Scan ho raha hai…", "Scanning…"), self._lang)
        self._progress.cancelled.connect(self._cancel_scan)
        # Scan settings sirf PROFILE se (panel ke manual badlaav se nahi).
        # dpi_override sirf DPI-shortcut wale ek scan ke liye.
        dpi, color, duplex, page_size, source = self._profile_scan_params(prof or {}, dpi_override)
        opts = self._opts
        if self.chk_fast.isChecked():
            dpi, color = 200, "bw"     # keep the profile's duplex (both-side) choice
            opts = dict(self._opts)    # lean copy: skip heavy per-page processing
            for k in ("remove_blank", "auto_crop", "deskew", "quality_enhance",
                      "clean_edges", "split_two_page"):
                opts[k] = False
        opts = dict(opts)
        opts["page_size"] = page_size
        opts["paper_source"] = source
        self._worker = ScanWorker(int(self.winId()), (prof or {}).get("source_name"),
                                  dpi, color, duplex, self._tmpdir, opts)
        self._worker.page_done.connect(self._on_page_scanned)
        self._worker.done.connect(self._on_scan_done)
        self._worker.failed.connect(self._on_scan_failed)
        self.btn_scan.setEnabled(False)
        try:
            self._conn_timer.stop()
        except Exception:
            pass
        self._scanning = True
        self._scan_place = _place        # rescan/insert ab pakka lagega
        try:
            self._state_timer.stop()
        except Exception:
            pass
        self._worker.start(); self._progress.show()
        self._set_busy_display("busy")

    def _on_page_scanned(self, path):
        # Rescan / Insert mode: page ko us hi jagah lagao (aakhir me nahi)
        place = getattr(self, "_scan_place", None)
        if place:
            mode, idx = place
            if mode == "replace":
                old = self.list.item(idx)
                if old is not None:
                    old.setData(QtCore.Qt.UserRole, path)
                    self._refresh_item(old)
                    self.list.setCurrentItem(old)
                    self._dirty = True
                    row = idx
                else:
                    self._add_item_for_path(path, at=idx); row = idx
                # is page ke baad wale extra pages usi ke aage insert ho
                self._scan_place = ("insert", idx + 1)
            else:  # insert
                self._add_item_for_path(path, at=idx); row = idx
                self._scan_place = ("insert", idx + 1)
            self._scan_count += 1
            self._name_one_page(row, path)
            try:
                self._pv_build_filmstrip()
            except Exception:
                pass
            if self._progress:
                self._progress.set_page(self._scan_count)
            return
        self._add_item_for_path(path)
        self._scan_count += 1
        # NAAM turant padhna shuru (per-page) — sab pages ke baad delay na ho
        self._name_one_page(self.list.count() - 1, path)
        if self._progress:
            self._progress.set_page(self._scan_count)
        if (self._opts.get("barcode_autofill") and not self._barcode_tried
                and not self.claim_edit.text().strip()):
            self._barcode_tried = True
            code = read_barcode(path)
            if code:
                self.claim_edit.setText(code)
                self.status.showMessage("Claim number found from barcode: %s" % code, 5000)

    def _on_scan_done(self, kept, skipped):
        self._scan_place = None          # rescan/insert mode khatam
        if kept:
            self._pstats_bump(scan_ok=1)
        if self._progress:
            self._progress.close(); self._progress = None
        self._scanning = False
        self.btn_scan.setEnabled(True)
        try:
            self._state_timer.start()
        except Exception:
            pass
        QtCore.QTimer.singleShot(2500, self._tick_scanner_state)
        try:
            self._conn_timer.start()
        except Exception:
            pass
        msg = '%d page(s) scanned.' % kept
        if skipped:
            msg += " (%d blank hataye)" % skipped
        self.status.showMessage(msg, 5000)
        if kept and self._opts.get("sound_on_done"):
            try:
                QtWidgets.QApplication.beep()
            except Exception:
                pass
        self._an_report("scan", n=kept)      # worldwide scan-count + refresh
        # (naam har page ke scan hote hi background me padh liya gaya —
        #  isliye yahan bulk naming ka intezaar nahi)
        if self._opts.get("auto_save") and kept:
            saved = self._auto_save_pdf()
            if saved and self._opts.get("batch_mode"):
                self._start_next_batch()

    def _on_scan_failed(self, msg):
        self._scan_place = None          # rescan/insert mode khatam
        self._pstats_bump(scan_fail=1)
        self._last_error = msg
        if self._progress:
            self._progress.close(); self._progress = None
        self._scanning = False
        self.btn_scan.setEnabled(True)
        try:
            self._state_timer.start()
        except Exception:
            pass
        QtCore.QTimer.singleShot(2500, self._tick_scanner_state)
        try:
            self._conn_timer.start()
        except Exception:
            pass
        self._warn(friendly_error(msg, self._opts.get("language", "en")))

    def _cancel_scan(self):
        self._scan_place = None          # rescan/insert mode khatam
        try:
            if getattr(self, "_worker", None):
                self._worker.requestInterruption()
        except Exception:
            pass
        if self._progress:
            self._progress.close(); self._progress = None
        self.btn_scan.setEnabled(True)
        try:
            self._conn_timer.start()
        except Exception:
            pass

    def _start_next_batch(self):
        for p in self._ordered_paths():
            try:
                os.remove(p)
            except Exception:
                pass
        self.list.clear(); self._show_preview(None)
        text, ok = QtWidgets.QInputDialog.getText(self, "Batch mode", "Next document's Claim No.:")
        if ok:
            self.claim_edit.setText(text.strip())
        self._update_status()

    def import_images(self):
        start = self._config.get("last_import_dir") or self._config.get("last_save_dir") or ""
        files, _ = QtWidgets.QFileDialog.getOpenFileNames(
            self, 'Choose images / PDF', start,
            "Images & PDF (*.png *.jpg *.jpeg *.bmp *.tif *.tiff *.pdf);;PDF (*.pdf);;Images (*.png *.jpg *.jpeg *.bmp *.tif *.tiff)")
        if not files:
            return
        try:
            self._config["last_import_dir"] = os.path.dirname(files[0])
            save_config(self._config)
        except Exception:
            pass
        self._start_import(files, "normal")

    def _start_import(self, files, mode):
        """Import BACKGROUND me chalta hai — app kabhi nahi rukti/jam hoti."""
        if getattr(self, "_importer", None) is not None and self._importer.isRunning():
            self._warn("A previous import is still running — please wait a moment and try again.")
            return
        self.status.showMessage("Importing… (you can keep using the app)", 0)
        self._importer = ImportWorker(files, self._tmpdir, mode,
                                      self.THUMB_HI_W, self.THUMB_HI_H)
        self._importer.page_ready.connect(self._on_import_page)
        self._importer.done.connect(self._on_import_done)
        self._importer.start()

    def _on_import_page(self, path, qimg):
        self._add_item_for_path(path, qimg)
        # naam turant (per-page) — import ke baad alag se naming ka wait nahi
        self._name_one_page(self.list.count() - 1, path, force=True)

    def _on_import_done(self, count):
        if count:
            self.status.showMessage("Imported %d page(s)." % count, 4000)
            self._pstats_bump(imports=count)     # personal
            self._an_report("event", imp=count)  # worldwide + refresh
            self._an_update_box()
        else:
            self.status.clearMessage()
            self._warn("Import failed. If it is a PDF, for better results "
                       "install 'PyMuPDF':\n  py -3.12-32 -m pip install PyMuPDF")

    def scan_from_camera(self):
        """Webcam/USB-camera se seedha document capture — scanner na ho to bhi."""
        if not HAS_CAMERA:
            self._warn("Camera support is not in this build (PyQt5 QtMultimedia).\n"
                       "For now use 'Phone photo to PDF' or Import.")
            return
        if not QCameraInfo.availableCameras():
            self._warn("No camera found. Is a webcam connected and turned on?")
            return
        dlg = CameraDialog(self, self._tmpdir)
        if dlg.exec_() == QtWidgets.QDialog.Accepted and dlg.captured:
            for p in dlg.captured:
                self._add_item_for_path(p)
            self.status.showMessage("Added %d page(s) from the camera." % len(dlg.captured), 4000)
            if tesseract_available():
                self.auto_name_pages()

    def import_photos(self):
        """Phone se kheenchi photos ko saaf karke pages banao (shadow hatana,
        seedha karna, chhota karna) — background me, app nahi rukti."""
        start = self._config.get("last_import_dir") or ""
        files, _ = QtWidgets.QFileDialog.getOpenFileNames(
            self, 'Choose phone photos', start,
            "Photos (*.jpg *.jpeg *.png *.bmp *.tif *.tiff *.webp)")
        if not files:
            return
        try:
            self._config["last_import_dir"] = os.path.dirname(files[0])
            save_config(self._config)
        except Exception:
            pass
        self._start_import(files, "photo")

    def split_id_cards(self):
        """Selected page par rakhe 2-3 ID cards ko alag-alag pages me kaato."""
        item = self._current_item_or_warn()
        if not item:
            return
        if not HAS_NUMPY:
            self._warn("This feature needs numpy (pip install numpy).")
            return
        path = item.data(QtCore.Qt.UserRole)
        try:
            with Image.open(path) as im:
                img = im.convert("RGB").copy()
        except Exception as exc:
            self._warn("Could not open page:\n%s" % exc)
            return
        boxes = detect_content_boxes(img)
        area = img.width * img.height
        keep = []
        for (l, t, r, b) in boxes:
            if (r - l) * (b - t) >= area * 0.01 and (r - l) > 60 and (b - t) > 60:
                keep.append((max(0, l - 10), max(0, t - 10),
                             min(img.width, r + 10), min(img.height, b + 10)))
        keep.sort(key=lambda x: (x[1], x[0]))
        if not keep:
            self._warn("No separate cards were found on this page.")
            return
        added = 0
        for box in keep:
            try:
                card = img.crop(box)
                fd, png = tempfile.mkstemp(suffix=".png", dir=self._tmpdir)
                os.close(fd)
                card.save(png, "PNG")
                self._add_item_for_path(png)
                added += 1
            except Exception:
                pass
        if added:
            self._dirty = True
            QtWidgets.QMessageBox.information(
                self, "Done",
                "Separated %d card(s)/piece(s) and added them.\n(The original page is unchanged — "
                "delete it if you don't need it.)" % added)

    def find_scanners(self):
        """Network par eSCL scanners khud dhoondo aur chun kar IP set karo."""
        self.status.showMessage("Searching for scanners on the network… (10-20 sec)", 0)
        self._finder = ScannerFinder()
        self._finder.found.connect(self._on_scanners_found)
        self._finder.start()

    def auto_detect_scanner(self):
        """Scanner KHUD pehchano — LAN (eSCL) par hai ya USB (WIA) par, dono
        dhoondh kar sabse behtar chun lo. User ko kuch nahi sochna padta."""
        self.status.showMessage(
            self.L("Scanner dhoondh rahe hain — LAN aur USB dono… (10-20 sec)",
                   "Detecting scanner — LAN and USB… (10-20 sec)"), 0)
        # USB/WIA list turant (halka kaam)
        try:
            self._auto_usb = list_wia_sources() or []
        except Exception:
            self._auto_usb = []
        # network eSCL background me
        self._finder = ScannerFinder()
        self._finder.found.connect(self._on_auto_detect_done)
        self._finder.start()

    def _on_auto_detect_done(self, net):
        self.status.clearMessage()
        usb = getattr(self, "_auto_usb", []) or []
        # combined choices banao — LAN (eSCL) + USB (WIA) + TWAIN
        choices = []   # (label, kind, value)
        for ip, model in (net or []):
            choices.append(("🌐 LAN: %s  (%s)" % (model, ip), "escl", ip))
        for _id, name in usb:
            choices.append(("🔌 USB: %s" % name, "wia", _id))
        if HAS_TWAIN:
            try:
                for nm in (list_sources(int(self.winId())) or []):
                    choices.append(("🖭 TWAIN: %s" % nm, "twain", nm))
            except Exception:
                pass

        def _apply(kind, value, label=""):
            disp = (label.split(":", 1)[1].strip() if ":" in label else label) or value
            if kind == "escl":
                self._opts["scanner_method"] = "escl"
                self._opts["scanner_ip"] = value
                self._config["scanner_ip"] = value
                self.ip_field.setText(value)
                msg = self.L("Scanner set ho gaya (LAN / network): %s" % value,
                             "Scanner set (LAN / network): %s" % value)
            elif kind == "twain":
                self._opts["scanner_method"] = "twain"
                # TWAIN device ka naam abhi ke profile me bhi rakh do
                try:
                    prof = self._selected_profile()
                    if prof is not None:
                        prof["source_name"] = value
                        self._config["profiles"] = self._profiles
                except Exception:
                    pass
                msg = self.L("Scanner set ho gaya (TWAIN): %s" % value,
                             "Scanner set (TWAIN): %s" % value)
            else:
                self._opts["scanner_method"] = "wia"
                self._opts["wia_device_id"] = value
                msg = self.L("Scanner set ho gaya (USB): %s" % disp,
                             "Scanner set (USB): %s" % disp)
            self._opts["scanner_name"] = disp
            self._save_opts(); save_config(self._config)
            try:
                self._refresh_conn_and_method()
            except Exception:
                pass
            cb = getattr(self, "_auto_detect_cb", None)
            self._auto_detect_cb = None
            if callable(cb):
                try:
                    cb(disp, kind, value)
                except Exception:
                    pass
            else:
                QtWidgets.QMessageBox.information(self, self.L("Ho gaya", "Done"), msg)

        if not choices:
            r = QtWidgets.QMessageBox.question(
                self, self.L("Koi scanner nahi mila", "No scanner found"),
                self.L("LAN ya USB par koi scanner nahi mila.\n\nScanner ON hai aur "
                       "juda/isi WiFi par hai?\n\nNetwork scanner ka IP hath se daalna hai?",
                       "No scanner found on LAN or USB.\n\nIs it on and connected/on the "
                       "same WiFi?\n\nEnter a network IP manually?"),
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
            self._auto_detect_cb = None
            if r == QtWidgets.QMessageBox.Yes:
                self.set_scanner_ip()
            return
        if len(choices) == 1:
            lbl, kind, value = choices[0]
            _apply(kind, value, lbl)
            return
        # ek se zyada mile — user chune (LAN pehle)
        labels = [c[0] for c in choices]
        pick, ok = QtWidgets.QInputDialog.getItem(
            self, self.L("Scanner mil gaye", "Scanners found"),
            self.L("%d scanner mile — kaunsa use karein?" % len(choices),
                   "%d scanners found — which to use?" % len(choices)),
            labels, 0, False)
        if not ok or not pick:
            self._auto_detect_cb = None
            return
        c = choices[labels.index(pick)]
        _apply(c[1], c[2], c[0])

    def pick_scanner_dialog(self, on_done=None):
        """Sabhi scanner (LAN eSCL + USB WIA + TWAIN) auto-detect karke list
        dikhao aur chuna hua set kar do. on_done(name, kind, value) callback."""
        self._auto_detect_cb = on_done
        self.auto_detect_scanner()

    def _detect_device_name(self):
        """Startup par abhi JUDA scanner ka naam dhoondh kar 'Device' me dikhao
        (WIA/USB — tez). Background me, taaki UI atke nahi. '(no device)' ki
        jagah asli scanner ka naam dikhega."""
        if self._opts.get("scanner_name"):
            try:
                self.dev_lbl.setText(self._opts["scanner_name"])
            except Exception:
                pass
            return
        if not HAS_W32:
            return

        def job():
            try:
                _pythoncom.CoInitialize()
            except Exception:
                pass
            try:
                return list_wia_sources() or []
            except Exception:
                return []
            finally:
                try:
                    _pythoncom.CoUninitialize()
                except Exception:
                    pass

        def done(devs):
            if not isinstance(devs, list) or not devs:
                return
            wid = self._opts.get("wia_device_id")
            name = None
            for _id, nm in devs:
                if wid and _id == wid:
                    name = nm; break
            if name is None:
                name = devs[0][1]
            if name:
                self._opts["scanner_name"] = name
                self._save_opts()
                try:
                    self.dev_lbl.setText(name)
                except Exception:
                    pass
        self._run_bg_quiet(job, done)

    def _refresh_conn_and_method(self):
        try:
            self.method_lbl.setText("Connected via: %s" % self._opts.get("scanner_method", "").upper())
        except Exception:
            pass
        try:
            nm = self._opts.get("scanner_name") or ""
            if nm and hasattr(self, "dev_lbl"):
                self.dev_lbl.setText(nm)
        except Exception:
            pass

    def _on_scanners_found(self, results):
        self.status.clearMessage()
        if not results:
            self._warn("No eSCL scanner found on the network.\n\n"
                       "Is the scanner ON and on the same WiFi/LAN? You can also enter the IP "
                       "manually: Settings → Scanner IP.")
            return
        items = ["%s  —  %s" % (ip, model) for ip, model in results]
        pick, ok = QtWidgets.QInputDialog.getItem(
            self, "Scanners found", "Found %d scanner(s) — which one to use?" % len(results),
            items, 0, False)
        if not ok or not pick:
            return
        ip = pick.split()[0]
        self._config["scanner_ip"] = ip
        self._opts["scanner_ip"] = ip
        self._opts["scanner_method"] = "escl"
        self.ip_field.setText(ip)
        self._save_opts(); save_config(self._config)
        QtWidgets.QMessageBox.information(
            self, "Done", "Scanner set: %s\n(Method: eSCL network scan)" % ip)

    def bundle_split_save(self):
        """Bundle scan → BLANK page ko separator maan kar alag-alag PDFs banao.
        (Scan karte waqt 'blank hatao' OFF rakhein aur documents ke beech ek
        khaali page daalein.)"""
        paths = self._ordered_paths()
        if not paths:
            self._warn(tr("scan_first", self._lang)); return
        groups, cur = [], []
        for p in paths:
            blank = False
            try:
                with Image.open(p) as im:
                    blank = is_blank_page(im.convert("RGB"))
            except Exception:
                pass
            if blank:
                if cur:
                    groups.append(cur); cur = []
            else:
                cur.append(p)
        if cur:
            groups.append(cur)
        if len(groups) <= 1:
            self._warn("No blank separator page found — everything looks like one document "
                       ".\n(Put a blank page between documents when scanning, "
                       "and keep 'Remove blank pages' OFF in Settings.)")
            return
        folder = QtWidgets.QFileDialog.getExistingDirectory(
            self, "PDFs kis folder me banayein?", self._opts.get("save_folder", ""))
        if not folder:
            return
        os.makedirs(folder, exist_ok=True)
        made = []
        for i, grp in enumerate(groups, 1):
            it0 = None
            for r in range(self.list.count()):
                if self.list.item(r).data(QtCore.Qt.UserRole) == grp[0]:
                    it0 = self.list.item(r); break
            name = (it0.data(TITLE_ROLE) if it0 else "") or ("Document_%d" % i)
            out = os.path.join(folder, sanitize(underscore_name(name)) + ".pdf")
            n = 2
            while os.path.exists(out):
                out = os.path.join(folder, "%s_%d.pdf" % (sanitize(underscore_name(name)), n)); n += 1
            try:
                self._pages_as_pdf(grp, out)
                made.append(os.path.basename(out))
            except Exception:
                pass
        QtWidgets.QMessageBox.information(
            self, "Done", "Created %d separate PDF(s):\n\n%s" % (len(made), "\n".join(made[:15])))

    def split_book_page(self):
        """Khuli kitab ka scan → do alag pages (baayan + daayan)."""
        items = self.list.selectedItems() or ([self.list.currentItem()] if self.list.currentItem() else [])
        if not items:
            self._warn("Select a page first."); return
        added = 0
        for it in items:
            path = it.data(QtCore.Qt.UserRole)
            try:
                with Image.open(path) as im:
                    img = im.convert("RGB").copy()
                mid = img.width // 2
                for half in (img.crop((0, 0, mid, img.height)),
                             img.crop((mid, 0, img.width, img.height))):
                    fd, png = tempfile.mkstemp(suffix=".png", dir=self._tmpdir)
                    os.close(fd); half.save(png, "PNG")
                    self._add_item_for_path(png); added += 1
            except Exception:
                pass
        if added:
            self._dirty = True
            self.status.showMessage("Created %d pages (you can delete the original)." % added, 5000)

    def business_cards(self):
        """Visiting cards ke scan se contacts nikaalo: har card alag + naam/phone/email
        padh kar .vcf (contact file) aur contacts.xlsx me save."""
        item = self._current_item_or_warn()
        if not item:
            return
        if not tesseract_available():
            self._warn("This needs Tesseract OCR."); return
        path = item.data(QtCore.Qt.UserRole)
        with Image.open(path) as im:
            img = im.convert("RGB").copy()
        boxes = detect_content_boxes(img)
        area = img.width * img.height
        cards = [b for b in boxes if (b[2]-b[0])*(b[3]-b[1]) >= area*0.01
                 and (b[2]-b[0]) > 80 and (b[3]-b[1]) > 50]
        if not cards:
            cards = [(0, 0, img.width, img.height)]
        folder = QtWidgets.QFileDialog.getExistingDirectory(
            self, 'Where to save contacts?', self._opts.get("save_folder", ""))
        if not folder:
            return
        os.makedirs(folder, exist_ok=True)
        got = 0
        rows = []
        for i, b in enumerate(sorted(cards, key=lambda x: (x[1], x[0])), 1):
            card = img.crop(b)
            try:
                text = pytesseract.image_to_string(card, lang="eng+hin")
            except Exception:
                text = ""
            lines = [l.strip() for l in (text or "").splitlines() if l.strip()]
            name = lines[0] if lines else ("Card_%d" % i)
            phones = re.findall(r"(?:\+?\d[\d\s\-]{8,14}\d)", text or "")
            emails = re.findall(r"[\w.+-]+@[\w-]+\.[\w.]+", text or "")
            phone = (phones[0].replace(" ", "").replace("-", "") if phones else "")
            email = emails[0] if emails else ""
            vcf = os.path.join(folder, sanitize(underscore_name(name))[:40] + ".vcf")
            try:
                with open(vcf, "w", encoding="utf-8") as fh:
                    fh.write("BEGIN:VCARD\nVERSION:3.0\nFN:%s\n" % name)
                    if phone:
                        fh.write("TEL;TYPE=CELL:%s\n" % phone)
                    if email:
                        fh.write("EMAIL:%s\n" % email)
                    fh.write("END:VCARD\n")
                got += 1
            except Exception:
                pass
            rows.append((name, phone, email))
            fd, png = tempfile.mkstemp(suffix=".png", dir=self._tmpdir)
            os.close(fd); card.save(png, "PNG")
            self._add_item_for_path(png)
        if HAS_XLSX and rows:
            try:
                xp = os.path.join(folder, "contacts.xlsx")
                wb = openpyxl.load_workbook(xp) if os.path.exists(xp) else openpyxl.Workbook()
                ws = wb.active
                if ws.max_row == 1 and not ws.cell(1, 1).value:
                    ws.append(["Naam", "Phone", "Email", "Date"])
                for nm, ph, em in rows:
                    ws.append([nm, ph, em, datetime.datetime.now().strftime("%Y-%m-%d")])
                wb.save(xp)
            except Exception:
                pass
        QtWidgets.QMessageBox.information(
            self, "Done",
            "Found %d card(s).\n%d contact (.vcf) files created + entries in contacts.xlsx.\n\n"
            "(Send the .vcf to your phone — opening it saves the contact.)" % (len(rows), got))

    def restore_photo_current(self):
        """Selected page/photo ko sudharo — BACKGROUND me, app nahi rukti."""
        items = self.list.selectedItems() or ([self.list.currentItem()] if self.list.currentItem() else [])
        if not items:
            self._warn("Select a page first."); return
        paths = [it.data(QtCore.Qt.UserRole) for it in items]

        def job():
            for path in paths:
                try:
                    with Image.open(path) as im:
                        restore_photo(im).save(path, "PNG")
                except Exception:
                    pass
            return True

        def on_done(_res):
            for it in items:
                try:
                    self._refresh_item(it)
                except Exception:
                    pass
            self._update_preview_panel()
            self._dirty = True
            self.status.showMessage("Photo restored.", 4000)
        self._run_bg(job, on_done, self.L("Photo sudhaar rahe hain…", "Restoring photo…"))

    def show_history(self):
        """Save folder ki saari PDFs — nayi se purani, search ke saath."""
        root = self._opts.get("save_folder", os.path.expanduser("~"))
        files = []
        try:
            for dp, _dn, fn in os.walk(root):
                for f in fn:
                    if f.lower().endswith(".pdf"):
                        p = os.path.join(dp, f)
                        try:
                            st = os.stat(p)
                            files.append((st.st_mtime, st.st_size, p))
                        except Exception:
                            pass
        except Exception:
            pass
        files.sort(reverse=True)
        files = files[:300]
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Scan History (%d files)" % len(files))
        dlg.resize(720, 480)
        v = QtWidgets.QVBoxLayout(dlg)
        ed = QtWidgets.QLineEdit(); ed.setPlaceholderText("Filter by name…")
        v.addWidget(ed)
        tree = QtWidgets.QTreeWidget()
        tree.setHeaderLabels(["File", "Kab", "Size"])
        tree.setRootIsDecorated(False)
        tree.setColumnWidth(0, 380); tree.setColumnWidth(1, 150)
        for mt, sz, p in files:
            when = datetime.datetime.fromtimestamp(mt).strftime("%d-%m-%Y %H:%M")
            kb = sz / 1024.0
            pretty = ("%.0f KB" % kb) if kb < 1024 else ("%.1f MB" % (kb / 1024))
            it = QtWidgets.QTreeWidgetItem([os.path.basename(p), when, pretty])
            it.setData(0, QtCore.Qt.UserRole, p)
            tree.addTopLevelItem(it)
        v.addWidget(tree, 1)

        def _filter(txt):
            t = txt.strip().lower()
            for i in range(tree.topLevelItemCount()):
                it = tree.topLevelItem(i)
                it.setHidden(bool(t) and t not in it.text(0).lower())
        ed.textChanged.connect(_filter)
        tree.itemDoubleClicked.connect(
            lambda it, col: self._open_path(it.data(0, QtCore.Qt.UserRole)))
        row = QtWidgets.QHBoxLayout()
        b1 = QtWidgets.QPushButton("Open")
        b1.clicked.connect(lambda: tree.currentItem() and self._open_path(
            tree.currentItem().data(0, QtCore.Qt.UserRole)))
        b2 = QtWidgets.QPushButton("Open folder")
        b2.clicked.connect(lambda: tree.currentItem() and self._reveal_in_explorer(
            tree.currentItem().data(0, QtCore.Qt.UserRole)))
        bc = QtWidgets.QPushButton("Close"); bc.clicked.connect(dlg.accept)
        row.addWidget(b1); row.addWidget(b2); row.addStretch(1); row.addWidget(bc)
        v.addLayout(row)
        dlg.exec_()

    # ---- Personal stats (sab kuch aapke PC par — kahin nahi jaata) ----
    MILESTONES = [10, 50, 100, 250, 500, 1000, 2500, 5000, 10000]

    def _pstats(self):
        if not hasattr(self, "_pstats_cache"):
            try:
                with open(PSTATS_PATH, "r", encoding="utf-8") as fh:
                    self._pstats_cache = json.load(fh)
            except Exception:
                self._pstats_cache = {}
        return self._pstats_cache

    def _pstats_save(self):
        try:
            with open(PSTATS_PATH, "w", encoding="utf-8") as fh:
                json.dump(self._pstats(), fh)
        except Exception:
            pass

    def _pstats_bump(self, pages=0, pdfs=0, shared=0, ocr_named=0,
                     scan_ok=0, scan_fail=0, saved_bytes=0, doc_type=None,
                     imports=0, prints=0):
        st = self._pstats()
        day = datetime.datetime.now().strftime("%Y-%m-%d")
        d = st.setdefault("days", {}).setdefault(day, {})
        for k, v in (("pages", pages), ("pdfs", pdfs), ("shared", shared),
                     ("imports", imports), ("prints", prints)):
            if v:
                d[k] = d.get(k, 0) + v
        t = st.setdefault("totals", {})
        for k, v in (("pages", pages), ("pdfs", pdfs), ("shared", shared),
                     ("ocr_named", ocr_named), ("scan_ok", scan_ok),
                     ("scan_fail", scan_fail), ("saved_bytes", saved_bytes),
                     ("imports", imports), ("prints", prints)):
            if v:
                t[k] = t.get(k, 0) + v
        if pages:
            hh = st.setdefault("hours", {})
            h = datetime.datetime.now().strftime("%H")
            hh[h] = hh.get(h, 0) + pages
        if doc_type:
            ty = st.setdefault("types", {})
            key = str(doc_type).strip()[:24] or "Anya"
            ty[key] = ty.get(key, 0) + 1
        self._pstats_save()
        if pdfs:
            self._check_milestones()
        self._update_sidebar_stats()

    def _pstats_sum(self, days=7, key="pages"):
        st = self._pstats().get("days", {})
        now = datetime.datetime.now()
        total = 0
        for i in range(days):
            k = (now - datetime.timedelta(days=i)).strftime("%Y-%m-%d")
            total += (st.get(k) or {}).get(key, 0)
        return total

    def _pstats_streak(self):
        st = self._pstats().get("days", {})
        now = datetime.datetime.now()
        streak = 0
        i = 0
        # aaj kaam na hua ho to kal se ginti shuru (streak abhi tuta nahi)
        if not (st.get(now.strftime("%Y-%m-%d")) or {}).get("pages", 0):
            i = 1
        while True:
            k = (now - datetime.timedelta(days=i)).strftime("%Y-%m-%d")
            if (st.get(k) or {}).get("pages", 0) > 0:
                streak += 1
                i += 1
            else:
                break
        return streak

    def _pstats_best_day(self):
        days = self._pstats().get("days", {})
        if not days:
            return ("—", 0)
        k, v = max(days.items(), key=lambda kv: kv[1].get("pages", 0))
        return (k, v.get("pages", 0))

    def _check_milestones(self):
        st = self._pstats()
        t = st.get("totals", {})
        shown = st.setdefault("milestones_shown", [])
        for kind, label in (("pdfs", "PDFs"), ("pages", "pages")):
            val = t.get(kind, 0)
            for m in self.MILESTONES:
                key = "%s_%d" % (kind, m)
                if val >= m and key not in shown:
                    shown.append(key)
                    self._pstats_save()
                    self.status.showMessage(
                        "🏆 Congratulations! You have reached %s %s!" % ("{:,}".format(m), label), 9000)

    def _maybe_month_wrap(self):
        """Mahine ki 1 tarikh ko pichhle mahine ka 'Aapka Mahina' summary."""
        st = self._pstats()
        now = datetime.datetime.now()
        this_month = now.strftime("%Y-%m")
        if st.get("month_shown") == this_month:
            return
        prev = (now.replace(day=1) - datetime.timedelta(days=1)).strftime("%Y-%m")
        days = {k: v for k, v in (st.get("days") or {}).items() if k.startswith(prev)}
        st["month_shown"] = this_month
        self._pstats_save()
        if not days:
            return
        pages = sum(d.get("pages", 0) for d in days.values())
        pdfs = sum(d.get("pdfs", 0) for d in days.values())
        busy_k, busy_v = max(days.items(), key=lambda kv: kv[1].get("pages", 0))
        QtWidgets.QMessageBox.information(
            self, "📊 Your month (%s)" % prev,
            "Last month you:\n\n📄 scanned %s pages\n🗂 created %s PDFs\n"
            "🔥 Busiest day: %s (%d pages)\n\nGreat work! 🎉"
            % ("{:,}".format(pages), "{:,}".format(pdfs), busy_k, busy_v))

    # ---- Sidebar stats (aap chunte ho kya-kya dikhe) ----
    def _sidebar_stat_items(self):
        """(key, label, value) — sidebar aur chooser dono iski list se chalte hain."""
        w = getattr(self, "_world_stats", {}) or {}
        st = self._pstats()
        t = st.get("totals", {})
        day = (st.get("days") or {}).get(datetime.datetime.now().strftime("%Y-%m-%d"), {})
        L = self.L
        return [
            ("my_today", L("📄 Mere aaj ke pages", "📄 My pages today"), day.get("pages", 0)),
            ("my_today_pdfs", L("🗂 Meri aaj ki PDFs", "🗂 My PDFs today"), day.get("pdfs", 0)),
            ("my_week", L("🗓 Is hafte (pages)", "🗓 This week (pages)"), self._pstats_sum(7, "pages")),
            ("my_total", L("📚 Mere kul pages", "📚 My total pages"), t.get("pages", 0)),
            ("my_pdfs", L("🗂 Meri kul PDFs", "🗂 My total PDFs"), t.get("pdfs", 0)),
            ("my_imports_today", L("📥 Aaj import kiye", "📥 Imported today"), day.get("imports", 0)),
            ("my_imports", L("📥 Kul import", "📥 Total imports"), t.get("imports", 0)),
            ("my_prints_today", L("🖨 Aaj print kiye", "🖨 Printed today"), day.get("prints", 0)),
            ("my_prints", L("🖨 Kul print", "🖨 Total prints"), t.get("prints", 0)),
            ("streak", L("🔥 Streak (din)", "🔥 Streak (days)"), self._pstats_streak()),
            ("shared", L("📤 Share ki hui", "📤 Shared"), t.get("shared", 0)),
            ("saved_mb", L("🗜 Compress se bachaya (MB)", "🗜 Saved by compress (MB)"),
             int(t.get("saved_bytes", 0) / 1048576)),
        ]

    DEFAULT_SIDEBAR_STATS = ["my_today", "my_today_pdfs", "my_week",
                             "streak", "my_imports", "my_prints"]

    def _update_sidebar_stats(self):
        # Analytics हटा दिया गया — sirf header ka profile-naam taaza karte hain.
        if hasattr(self, "hdr_profile"):
            try:
                prof = self._selected_profile()
                self.hdr_profile.setText("〔 %s 〕" % (prof.get("name") if prof else "—"))
            except Exception:
                pass

    def choose_sidebar_stats(self):
        """Aap khud chuno sidebar me kaun-kaun si stats dikhein."""
        items = self._sidebar_stat_items()
        sel = set(self._opts.get("sidebar_stats") or self.DEFAULT_SIDEBAR_STATS)
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Choose sidebar stats")
        v = QtWidgets.QVBoxLayout(dlg)
        v.addWidget(QtWidgets.QLabel("Tick what to show in the sidebar:"))
        lw = QtWidgets.QListWidget()
        for k, lbl, _val in items:
            it = QtWidgets.QListWidgetItem(lbl)
            it.setData(QtCore.Qt.UserRole, k)
            it.setFlags(it.flags() | QtCore.Qt.ItemIsUserCheckable)
            it.setCheckState(QtCore.Qt.Checked if k in sel else QtCore.Qt.Unchecked)
            lw.addItem(it)
        v.addWidget(lw, 1)
        bb = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        bb.accepted.connect(dlg.accept); bb.rejected.connect(dlg.reject)
        v.addWidget(bb)
        if dlg.exec_() != QtWidgets.QDialog.Accepted:
            return
        chosen = [lw.item(i).data(QtCore.Qt.UserRole) for i in range(lw.count())
                  if lw.item(i).checkState() == QtCore.Qt.Checked]
        self._opts["sidebar_stats"] = chosen
        self._save_opts()
        self._update_sidebar_stats()

    def _on_world_stats(self, data):
        try:
            self._world_stats = dict(data or {})
        except Exception:
            self._world_stats = {}
        self._update_sidebar_stats()
        try:
            self._update_footer()
        except Exception:
            pass

    # ---- Stats Dashboard ----
    def show_stats_dashboard(self):
        st = self._pstats()
        t = st.get("totals", {})
        w = getattr(self, "_world_stats", {}) or {}
        now = datetime.datetime.now()
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("📊 Stats Dashboard")
        dlg.resize(560, 560)
        v = QtWidgets.QVBoxLayout(dlg)
        tabs = QtWidgets.QTabWidget()
        v.addWidget(tabs, 1)

        # --- Meri Stats tab ---
        p1 = QtWidgets.QWidget()
        l1 = QtWidgets.QVBoxLayout(p1)
        days = st.get("days", {})
        last7, labels = [], []
        for i in range(6, -1, -1):
            dt = now - datetime.timedelta(days=i)
            last7.append((days.get(dt.strftime("%Y-%m-%d")) or {}).get("pages", 0))
            labels.append(["So", "Mo", "Tu", "We", "Th", "Fr", "Sa"][int(dt.strftime("%w"))])
        l1.addWidget(QtWidgets.QLabel("<b>Last 7 days (pages):</b>"))
        l1.addWidget(SparkBars(last7, labels))
        best_k, best_v = self._pstats_best_day()
        ok, fail = t.get("scan_ok", 0), t.get("scan_fail", 0)
        rate = ("%.0f%%" % (100.0 * ok / (ok + fail))) if (ok + fail) else "—"
        types = sorted((st.get("types") or {}).items(), key=lambda kv: -kv[1])[:5]
        tys = "<br>".join("&nbsp;&nbsp;%s: <b>%d</b>" % (k, n) for k, n in types) or '&nbsp;&nbsp;(not yet)'
        hours = st.get("hours") or {}
        peak_h = max(hours.items(), key=lambda kv: kv[1])[0] + ":00 baje" if hours else "—"
        info = QtWidgets.QLabel(
            "📄 Today: <b>%d</b> pages &nbsp;|&nbsp; 🗓 Week: <b>%d</b> &nbsp;|&nbsp; "
            "Month: <b>%d</b><br>"
            "📚 Total: <b>%s pages</b>, <b>%s PDFs</b><br>"
            "📥 Import: <b>%s</b> &nbsp;|&nbsp; 🖨 Print: <b>%s</b><br>"
            "🔥 Streak: <b>%d days</b> &nbsp;|&nbsp; 🏅 Best day: <b>%s (%d)</b><br>"
            "📤 Share: <b>%d</b> &nbsp;|&nbsp; 🗜 Saved: <b>%.1f MB</b><br>"
            "🩺 Scan success: <b>%s</b> &nbsp;|&nbsp; 🔤 OCR names found: <b>%d</b><br>"
            "⏰ Busiest time: <b>%s</b><br><br><b>Document types (top 5):</b><br>%s"
            % (self._pstats_sum(1), self._pstats_sum(7), self._pstats_sum(30),
               "{:,}".format(t.get("pages", 0)), "{:,}".format(t.get("pdfs", 0)),
               "{:,}".format(t.get("imports", 0)), "{:,}".format(t.get("prints", 0)),
               self._pstats_streak(), best_k, best_v,
               t.get("shared", 0), t.get("saved_bytes", 0) / 1048576.0,
               rate, t.get("ocr_named", 0), peak_h, tys))
        info.setTextFormat(QtCore.Qt.RichText)
        info.setWordWrap(True)
        l1.addWidget(info)
        l1.addStretch(1)
        bexp = QtWidgets.QPushButton("📥 Export to Excel")
        bexp.clicked.connect(self._export_pstats_excel)
        l1.addWidget(bexp)
        tabs.addTab(p1, "🙋 Meri Stats")

        bcl = QtWidgets.QPushButton("Close")
        bcl.clicked.connect(dlg.accept)
        v.addWidget(bcl)
        dlg.exec_()

    def _export_pstats_excel(self):
        if not HAS_XLSX:
            self._warn("openpyxl is not installed."); return
        out, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Stats export", os.path.join(os.path.expanduser("~"), "apnescan_stats.xlsx"),
            "Excel (*.xlsx)")
        if not out:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Date", "Pages", "PDFs", "Shared"])
            for k in sorted(self._pstats().get("days", {})):
                d = self._pstats()["days"][k]
                ws.append([k, d.get("pages", 0), d.get("pdfs", 0), d.get("shared", 0)])
            wb.save(out)
        except Exception as exc:
            self._warn("Export failed: %s" % exc); return
        QtWidgets.QMessageBox.information(self, "Done", "Stats export:\n%s" % out)

    # ---- "Meri Files" right panel ----
    def _files_root(self):
        # Agar user ne panel me koi doosra folder "kholo" kiya hai to wahi;
        # warna save-folder.
        ov = self._opts.get("files_panel_root")
        if ov and os.path.isdir(ov):
            return ov
        root = (self._opts.get("save_folder")
                or os.path.join(os.path.expanduser("~"), "Documents", "NobleScans"))
        try:
            os.makedirs(root, exist_ok=True)
        except Exception:
            pass
        return root

    def _refresh_files_root(self):
        try:
            root = self._files_root()
            self.files_model.setRootPath(root)
            self.files_tree.setRootIndex(self.files_model.index(root))
            self._apply_files_sort()
            self._update_panel_nav()
            if hasattr(self, "foot_folder"):
                self._update_footer()
        except Exception:
            pass

    def open_existing_folder(self):
        """Panel me kahin bhi rakha koi bhi folder kholo (uske andar ke documents
        dikhenge, wahin save bhi kar sakte ho)."""
        start = self._opts.get("files_panel_root") or self._opts.get("save_folder") or os.path.expanduser("~")
        f = QtWidgets.QFileDialog.getExistingDirectory(
            self, self.L("Kaunsa folder kholna hai?", "Which folder to open?"), start)
        if not f:
            return
        self._opts["files_panel_root"] = f
        self._save_opts()
        self._refresh_files_root()
        self.status.showMessage(
            self.L("📂 Folder khul gaya: %s   (🏠 se wapas save-folder)" % os.path.basename(f),
                   "📂 Opened: %s   (🏠 to go back to the save folder)" % os.path.basename(f)), 6000)

    def reset_panel_folder(self):
        """Panel ko wapas save-folder par le jao."""
        self._opts["files_panel_root"] = ""
        self._save_opts()
        self._refresh_files_root()

    def _files_tree_open(self, index):
        try:
            path = self.files_model.filePath(index)
        except Exception:
            return
        if not path:
            return
        if os.path.isdir(path):
            self._panel_show(path)      # folder ke ANDAR chalo (sirf uski cheezein)
        elif os.path.isfile(path):
            self._open_path(path)

    # ---- Panel navigation: ek baar me ek hi folder (drill-in / back) ----
    def _panel_show(self, folder):
        try:
            self.files_model.setRootPath(folder)
            self.files_tree.setRootIndex(self.files_model.index(folder))
        except Exception:
            return
        self._update_panel_nav()
        # naya folder khula — search index dobara banao (background me pehle se
        # taiyaar rakho taaki pehla akshar likhte hi natije turant aayein)
        self._invalidate_files_index()
        try:
            if folder and os.path.isdir(folder):
                self._ensure_files_index_async(folder)
        except Exception:
            pass
        try:
            self._update_folder_info()
        except Exception:
            pass

    def _panel_current_dir(self):
        try:
            p = self.files_model.filePath(self.files_tree.rootIndex())
            if p and os.path.isdir(p):
                return p
        except Exception:
            pass
        return self._files_root()

    def _panel_back(self):
        cur = self._panel_current_dir()
        base = self._files_root()
        if os.path.normpath(cur) == os.path.normpath(base):
            return                       # base par hain — isse upar nahi
        self._panel_show(os.path.dirname(os.path.normpath(cur)))

    def _update_panel_nav(self):
        cur = self._panel_current_dir()
        base = self._files_root()
        at_top = os.path.normpath(cur) == os.path.normpath(base)
        if hasattr(self, "btn_panel_back"):
            self.btn_panel_back.setEnabled(not at_top)
        if hasattr(self, "lbl_panel_cwd"):
            self.lbl_panel_cwd.setText("📂 " + (os.path.basename(cur) or cur))

    # ---- List sort (user ki pasand save rehti hai) ----
    SORT_MODES = [
        ("name_asc",  "Naam: A → Z",        "Name: A → Z",        0, QtCore.Qt.AscendingOrder),
        ("name_desc", "Naam: Z → A",        "Name: Z → A",        0, QtCore.Qt.DescendingOrder),
        ("date_desc", 'Date: newest first',   "Date: newest first", 3, QtCore.Qt.DescendingOrder),
        ("date_asc",  'Date: oldest first', "Date: oldest first", 3, QtCore.Qt.AscendingOrder),
        ("size_desc", 'Size: largest first',   "Size: largest first",1, QtCore.Qt.DescendingOrder),
        ("size_asc",  'Size: smallest first', "Size: smallest first",1, QtCore.Qt.AscendingOrder),
    ]

    def _build_sort_menu(self):
        menu = QtWidgets.QMenu(self)
        grp = QtWidgets.QActionGroup(menu); grp.setExclusive(True)
        cur = self._opts.get("files_sort", "name_asc")
        for key, hi, en, _col, _order in self.SORT_MODES:
            act = menu.addAction(self.L(hi, en))
            act.setCheckable(True)
            act.setChecked(key == cur)
            act.triggered.connect(lambda _c=False, k=key: self._set_files_sort(k))
            grp.addAction(act)
        return menu

    def _set_files_sort(self, key):
        self._opts["files_sort"] = key
        self._save_opts()
        self._apply_files_sort()

    def _apply_files_sort(self):
        key = self._opts.get("files_sort", "name_asc")
        col, order = 0, QtCore.Qt.AscendingOrder
        for k, _hi, _en, c, o in self.SORT_MODES:
            if k == key:
                col, order = c, o
                break
        try:
            self.files_model.sort(col, order)
        except Exception:
            pass

    def _selected_library_folder(self):
        # HAMESHA wahi folder jo panel me ABHI KHULA (view) hai — highlighted
        # subfolder nahi. (New folder / save / import isi khule folder me ho.)
        # Kisi doosre subfolder me daalna ho to us par 2x click karke us me
        # jaao (drill-in), phir wahan banao/save karo.
        return self._panel_current_dir()

    def new_library_folder(self):
        base = self._selected_library_folder()
        name, ok = QtWidgets.QInputDialog.getText(
            self, "New folder",
            "Folder name:\n(will be created inside: '%s')" % (os.path.basename(base) or "Meri Files"))
        if not ok or not name.strip():
            return
        p = os.path.join(base, sanitize(underscore_name(name.strip())))
        try:
            os.makedirs(p, exist_ok=True)
        except Exception as exc:
            self._warn("Could not create folder:\n%s" % exc)
            return
        try:
            self.files_tree.setCurrentIndex(self.files_model.index(p))
            self.files_tree.expand(self.files_model.index(base))
        except Exception:
            pass
        self.status.showMessage("Folder created: %s" % os.path.basename(p), 4000)

    def save_into_selected_folder(self):
        """Panel me chune folder me SEEDHA save — bas naam confirm karo, Enter."""
        paths = self._ordered_paths()
        if not paths:
            self._warn(tr("scan_first", self._lang))
            return
        self._save_pages_to_folder(self._selected_library_folder(), paths, ask_name=True)

    def import_into_selected_folder(self):
        """Panel ke ⬇ Import button — chune folder me bahar se files laao."""
        folder = self._selected_library_folder()
        if not folder or not os.path.isdir(folder):
            folder = self._files_root()
        self._import_files_dialog(folder)

    def _import_files_dialog(self, folder):
        files, _ = QtWidgets.QFileDialog.getOpenFileNames(
            self,
            self.L("Files chuno — ye is folder me aa jaayengi",
                   "Choose files — they'll be copied into this folder"),
            os.path.expanduser("~"),
            "Documents (*.pdf *.jpg *.jpeg *.png *.tif *.tiff *.docx *.xlsx);;"
            "All files (*.*)")
        if files:
            self._import_files_to_folder(folder, files)

    def _import_files_to_folder(self, folder, files):
        """Chuni gayi files ko folder me copy karo (naam takraye to _2, _3…);
        background me, UI nahi rukti."""
        def job():
            done = 0
            for src in files:
                try:
                    stem, ext = os.path.splitext(os.path.basename(src))
                    dst = os.path.join(folder, stem + ext)
                    k = 2
                    while os.path.exists(dst):
                        dst = os.path.join(folder, "%s_%d%s" % (stem, k, ext))
                        k += 1
                    shutil.copy2(src, dst)
                    done += 1
                except Exception:
                    pass
            return done

        def on_done(n):
            if isinstance(n, Exception):
                self._warn(self.L("Import fail ho gaya", "Import failed"))
                return
            try:
                idx = self.files_model.index(folder)
                self.files_tree.setCurrentIndex(idx)
                self.files_tree.expand(idx)
                self.files_tree.scrollTo(idx)
            except Exception:
                pass
            self.status.showMessage(self.L(
                "📥 %d file is folder me aa gayi" % n,
                "📥 Imported %d file(s)" % n), 4000)
        self._run_bg(job, on_done, self.L("Laa rahe hain…", "Importing…"))

    def _save_pages_to_folder(self, folder, paths, ask_name=True, merge_same=False,
                              forced_name=None):
        if not paths:
            self._warn(tr("scan_first", self._lang))
            return
        if forced_name:
            default = forced_name
        else:
            default = os.path.basename(self._build_filename(".pdf", paths=paths))
            if default.lower().endswith(".pdf"):
                default = default[:-4]
        if ask_name:
            name, ok = QtWidgets.QInputDialog.getText(
                self, "Save",
                "File name:   (folder: %s)" % (os.path.basename(folder) or folder),
                text=default)
            if not ok or not name.strip():
                return
            default = name.strip()
        base = sanitize(underscore_name(default))
        out = os.path.join(folder, base + ".pdf")
        # merge_same (drag-drop save): usi naam ki PDF ho to naye pages usi me
        # jod do (ek hi PDF). Warna duplicate par _2, _3… numbering.
        merge_into = out if (merge_same and os.path.exists(out) and HAS_OCR_LIBS) else None
        if not merge_into:
            n = 2
            while os.path.exists(out):
                out = os.path.join(folder, "%s_%d.pdf" % (base, n))
                n += 1
        if not self._validate_claim_ok():
            return
        if not merge_into and not self._duplicate_ok():
            return
        # PDF banana/merge karna BACKGROUND me — app ek pal ke liye bhi na ruke
        npages = len(paths)
        target = merge_into or out

        def job():
            if merge_into:
                self._append_pages_to_pdf(merge_into, paths)
            else:
                self._pages_as_pdf(paths, out)
            return target

        def on_done(res):
            if isinstance(res, Exception):
                self._warn("Save failed:\n%s" % res)
                return
            saved = res
            self._remember_save_dir(saved)
            self._remember_doc_name(saved)
            self._record_save(saved, npages)
            self._dirty = False
            self._after_save_action(saved)
            # Save hote hi wahi folder sidebar me khol do (aur file select)
            try:
                self._panel_show(folder)
                self.files_tree.setCurrentIndex(self.files_model.index(saved))
            except Exception:
                pass
            m = ('✔ Added to this PDF: %s' if merge_into else '✔ Saved: %s') % saved
            self.status.showMessage(m, 7000)
        self._run_bg(job, on_done, self.L("Save ho raha hai… (app chalti rahegi)",
                                          "Saving… (app stays usable)"))

    def _append_pages_to_pdf(self, pdf_path, paths):
        """Naye pages ko ek maujooda PDF ke aage jodo (same-naam merge)."""
        fd, tmp = tempfile.mkstemp(suffix=".pdf")
        os.close(fd)
        try:
            self._pages_as_pdf(paths, tmp)
            writer = PdfWriter()
            for src in (pdf_path, tmp):
                for pg in PdfReader(src).pages:
                    writer.add_page(pg)
            with open(pdf_path, "wb") as fh:
                writer.write(fh)
        finally:
            try:
                os.remove(tmp)
            except Exception:
                pass

    def _on_pages_dropped(self, idx):
        """Drop pages on a folder = save there instantly. Pages are grouped by
        their NAME (title): same-named pages -> one PDF (named after the title),
        differently-named pages -> separate PDFs."""
        try:
            p = self.files_model.filePath(idx) if idx.isValid() else self._panel_current_dir()
        except Exception:
            p = self._panel_current_dir()
        folder = p if os.path.isdir(p) else os.path.dirname(p)
        paths = self._selected_paths() or self._ordered_paths()
        self._save_pages_grouped(folder, paths)

    def _save_pages_grouped(self, folder, paths):
        """Selected pages ko unke NAAM (title) ke hisaab se group karke save karo:
        ek jaise naam wale pages ki EK hi PDF; alag naam wale alag PDF. Jinka
        naam nahi hai unke liye auto-naam."""
        if not paths:
            self._warn("Scan or import a page first."); return
        title_of = {}
        for i in range(self.list.count()):
            it = self.list.item(i)
            title_of[it.data(QtCore.Qt.UserRole)] = it.data(TITLE_ROLE)
        groups, order = [], {}
        for p in paths:
            t = title_of.get(p) or ""
            key = t or "__auto__"
            if key not in order:
                order[key] = len(groups)
                groups.append([t, []])
            groups[order[key]][1].append(p)
        for name, gpaths in groups:
            self._save_pages_to_folder(folder, gpaths, ask_name=False,
                                       merge_same=True, forced_name=(name or None))
        if len(groups) > 1:
            self.status.showMessage("Saved %d documents (grouped by name)" % len(groups), 5000)

    def _files_sel_changed(self, cur, _prev):
        # Folder chunte hi status me hisaab; FILE chunte hi uska preview
        try:
            p = self.files_model.filePath(cur)
            if not p:
                return
            if os.path.isdir(p):
                files = [f for f in os.listdir(p)
                         if os.path.isfile(os.path.join(p, f))]
                sz = sum(os.path.getsize(os.path.join(p, f)) for f in files)
                self.status.showMessage(
                    "📁 %s — %d files, %.1f MB" %
                    (os.path.basename(p) or p, len(files), sz / 1048576.0), 4000)
            elif os.path.isfile(p):
                self._preview_file_in_panel(p)   # PDF/image ka preview panel me
        except Exception:
            pass

    def _files_tree_menu(self, pos):
        idx = self.files_tree.indexAt(pos)
        menu = QtWidgets.QMenu(self)
        # Kai files ek saath chuni hui? -> bulk actions
        sel_files = self._selected_library_files()
        if len(sel_files) > 1:
            menu.addAction(self.L("🧩 %d files → ek PDF…" % len(sel_files),
                                  "🧩 Merge %d files → one PDF…" % len(sel_files)),
                           lambda: self._bulk_merge(sel_files))
            menu.addAction(self.L("📁 Dusre folder me le jao…", "📁 Move to another folder…"),
                           lambda: self._bulk_move(sel_files, copy=False))
            menu.addAction(self.L("📄 Dusre folder me copy…", "📄 Copy to another folder…"),
                           lambda: self._bulk_move(sel_files, copy=True))
            menu.addAction(self.L("✏ %d files ka bulk rename…" % len(sel_files),
                                  "✏ Bulk rename %d files…" % len(sel_files)),
                           lambda: self._bulk_rename(sel_files))
            menu.addSeparator()
            menu.addAction(self.L("🗑 %d files delete (Recycle Bin)" % len(sel_files),
                                  "🗑 Delete %d files (Recycle Bin)" % len(sel_files)),
                           lambda: self._bulk_delete(sel_files))
            menu.addSeparator()
            menu.addAction(self.L("🗑 Recycle Bin…", "🗑 Recycle Bin…"), self.show_recycle_bin)
            menu.exec_(self.files_tree.viewport().mapToGlobal(pos))
            return
        if idx.isValid():
            path = self.files_model.filePath(idx)
            if os.path.isdir(path):
                menu.addAction("💾 Save here",
                               lambda: self._save_pages_to_folder(path, self._ordered_paths()))
                menu.addAction("📥 Import files here…",
                               lambda: self._import_files_dialog(path))
                menu.addAction("➕ New folder here…", lambda: self._new_folder_in(path))
                menu.addAction("🧩 All PDFs in this folder → one PDF…",
                               lambda: self._merge_folder_pdfs(path))
                menu.addAction(self.L("🗜 Is folder ki ZIP banao…",
                                      "🗜 Make a ZIP of this folder…"),
                               lambda: self._zip_folder(path))
                menu.addAction("📂 Open in Explorer", lambda: self._open_path(path))
                favs = self._opts.get("fav_folders") or []
                menu.addAction("⭐ Remove favourite" if path in favs else "⭐ Add favourite",
                               lambda: self._toggle_fav(path))
            else:
                menu.addAction("📖 Open", lambda: self._open_path(path))
                pinned = self._opts.get("pinned_files") or []
                menu.addAction(self.L("📌 Pin hatao", "📌 Unpin") if path in pinned
                               else self.L("📌 Pin karo (upar rakho)", "📌 Pin (keep on top)"),
                               lambda: self._toggle_pin(path))
                if path.lower().endswith(".pdf"):
                    menu.addAction("🟢 Send via WhatsApp", lambda: self.share_whatsapp(path))
                    menu.addAction("✉ Send via Email", lambda: self.share_email(path))
                    menu.addAction("🗜 Compress…",
                                   lambda: self.compress_pdf_tool(path))
                    menu.addAction("🏷 Add tag…", lambda: self.tag_pdf(path))
                menu.addAction(self.L("🖨 Print", "🖨 Print"), lambda: self._print_library_file(path))
                menu.addSeparator()
                menu.addAction(self.L("📄 Dusre folder me copy…", "📄 Copy to another folder…"),
                               lambda: self._bulk_move([path], copy=True))
                menu.addAction(self.L("📁 Dusre folder me le jao…", "📁 Move to another folder…"),
                               lambda: self._bulk_move([path], copy=False))
                menu.addAction("✏ Rename…", lambda: self._rename_library_file(path))
                menu.addAction("🗑 Delete…", lambda: self._delete_library_file(path))
        else:
            menu.addAction(self.L("📂 Koi folder kholo…", "📂 Open a folder…"), self.open_existing_folder)
            menu.addAction(self.L("🏠 Wapas save-folder", "🏠 Back to save folder"), self.reset_panel_folder)
            menu.addAction("➕ New folder", self.new_library_folder)
        menu.addSeparator()
        menu.addAction(self.L("🔁 Nakli (duplicate) files dhoondo…",
                              "🔁 Find duplicate files…"), self._find_duplicates)
        menu.addAction(self.L("🗑 Recycle Bin…", "🗑 Recycle Bin…"), self.show_recycle_bin)
        menu.exec_(self.files_tree.viewport().mapToGlobal(pos))

    def _new_folder_in(self, base):
        name, ok = QtWidgets.QInputDialog.getText(
            self, "New folder",
            "Folder name:\n(will be created inside: '%s')" % (os.path.basename(base) or "Meri Files"))
        if not ok or not name.strip():
            return
        p = os.path.join(base, sanitize(underscore_name(name.strip())))
        try:
            os.makedirs(p, exist_ok=True)
        except Exception as exc:
            self._warn("Could not create folder:\n%s" % exc)
            return
        try:
            self.files_tree.expand(self.files_model.index(base))
            self.files_tree.setCurrentIndex(self.files_model.index(p))
        except Exception:
            pass
        self.status.showMessage("Folder created: %s" % os.path.basename(p), 4000)

    def _toggle_fav(self, path):
        favs = self._opts.setdefault("fav_folders", [])
        if path in favs:
            favs.remove(path)
            msg = self.L("⭐ Favourite hata diya", "⭐ Removed from favourites")
        else:
            favs.append(path)
            while len(favs) > 12:
                favs.pop(0)
            msg = self.L("⭐ Favourite me jud gaya", "⭐ Added to favourites")
        self._save_opts()
        self._rebuild_fav_bar()
        try:
            self.status.showMessage(msg, 3000)
        except Exception:
            pass

    def _toggle_pin(self, path):
        """File ko 📌 pin/unpin — pinned files '📌 Pinned' filter me upar dikhti hain."""
        pins = self._opts.setdefault("pinned_files", [])
        if path in pins:
            pins.remove(path); msg = self.L("📌 Pin hataya", "📌 Unpinned")
        else:
            pins.append(path)
            while len(pins) > 50:
                pins.pop(0)
            msg = self.L("📌 Pin kar diya (📌 Pinned filter me dekho)",
                         "📌 Pinned (see the 📌 Pinned filter)")
        self._save_opts()
        self.status.showMessage(msg, 3000)
        try:
            if hasattr(self, "files_filter") and self.files_filter.currentData() == "pinned":
                self._apply_files_filter()
        except Exception:
            pass

    def _print_library_file(self, path):
        """Saved file ko seedha print par bhejo (default app se)."""
        try:
            os.startfile(path, "print")
            self.status.showMessage(self.L("🖨 Print par bhej diya", "🖨 Sent to printer"), 4000)
        except Exception:
            try:
                self._open_path(path)
            except Exception as e:
                self._warn(str(e))

    def _zip_folder(self, folder):
        """Poore folder ki ek ZIP banao (backup / bhejne ke liye) — background me."""
        if not folder or not os.path.isdir(folder):
            return
        default = os.path.join(os.path.dirname(folder),
                               (os.path.basename(folder) or "files") + ".zip")
        out, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, self.L("ZIP save karein", "Save ZIP"), default, "ZIP (*.zip)")
        if not out:
            return
        if not out.lower().endswith(".zip"):
            out += ".zip"

        def job():
            import zipfile
            base = os.path.dirname(folder)
            with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as z:
                for dp, _dns, fn in os.walk(folder):
                    for f in fn:
                        fp = os.path.join(dp, f)
                        if os.path.abspath(fp) == os.path.abspath(out):
                            continue
                        z.write(fp, os.path.relpath(fp, base))
            return out

        def done(res):
            if isinstance(res, Exception):
                self._warn(self.L("ZIP nahi bani:\n", "ZIP failed:\n") + str(res)); return
            self.status.showMessage(self.L("🗜 ZIP ban gayi: ", "🗜 ZIP created: ")
                                    + os.path.basename(res), 8000)
            try:
                self._open_path(os.path.dirname(res))
            except Exception:
                pass
        self._run_bg(job, done, self.L("ZIP bana rahe…", "Zipping…"))

    def _bulk_rename(self, files):
        """Kai files ko ek naam + number (naam_001, naam_002…) me rename karo."""
        if not files:
            return
        base, ok = QtWidgets.QInputDialog.getText(
            self, self.L("Bulk rename", "Bulk rename"),
            self.L("Sabhi ke liye ek naam (aage number apne aap lagega):",
                   "One base name (a number is appended automatically):"), text="doc")
        if not ok or not base.strip():
            return
        base = sanitize(underscore_name(base.strip())) or "doc"
        n = 0
        for i, p in enumerate(sorted(files, key=lambda x: os.path.basename(x).lower()), 1):
            try:
                ext = os.path.splitext(p)[1]
                new = os.path.join(os.path.dirname(p), "%s_%03d%s" % (base, i, ext))
                if os.path.abspath(new) == os.path.abspath(p):
                    n += 1; continue
                if os.path.exists(new):
                    continue
                os.rename(p, new)
                n += 1
            except Exception:
                pass
        self._invalidate_files_index()
        self.status.showMessage(self.L("✏ %d files rename ho gayi" % n,
                                       "✏ Renamed %d files" % n), 5000)

    def _find_duplicates(self):
        """Ek jaisi (nakli) files dhoondo — pehle size, phir content (md5) se."""
        scope = self._panel_current_dir()
        if not (scope and os.path.isdir(scope)):
            scope = self._files_root()

        def job():
            import hashlib
            by_size = {}
            for dp, _dns, fn in os.walk(scope):
                for f in fn:
                    if not f.lower().endswith(self._FILE_EXTS):
                        continue
                    fp = os.path.join(dp, f)
                    try:
                        by_size.setdefault(os.path.getsize(fp), []).append(fp)
                    except Exception:
                        pass
            dups = []
            for _sz, group in by_size.items():
                if len(group) < 2:
                    continue
                seen = {}
                for fp in group:
                    try:
                        h = hashlib.md5()
                        with open(fp, "rb") as fh:
                            for chunk in iter(lambda: fh.read(65536), b""):
                                h.update(chunk)
                        seen.setdefault(h.hexdigest(), []).append(fp)
                    except Exception:
                        pass
                for _k, same in seen.items():
                    if len(same) >= 2:
                        dups.extend(same)
            return [("file", p) for p in dups]

        def done(res):
            if isinstance(res, Exception):
                return
            if not res:
                self.status.showMessage(self.L("Koi nakli file nahi mili.",
                                               "No duplicate files found."), 4000)
                return
            try:
                self.files_filter.blockSignals(True)
                self.files_filter.setCurrentIndex(0)
                self.files_filter.blockSignals(False)
            except Exception:
                pass
            self._render_files_results(res)
            self.status.showMessage(self.L("🔁 %d nakli files mili (ek jaisi)" % len(res),
                                           "🔁 %d duplicate files found" % len(res)), 6000)
        self._run_bg(job, done, self.L("Nakli files dhoondh rahe…", "Finding duplicates…"))

    def _rebuild_fav_bar(self):
        """Favourites ko dropdown me bharo (galat/hata diye gaye folder chhod do)."""
        if not hasattr(self, "fav_combo"):
            return
        self.fav_combo.blockSignals(True)
        self.fav_combo.clear()
        favs = [p for p in (self._opts.get("fav_folders") or []) if os.path.isdir(p)]
        if favs:
            self.fav_combo.addItem(self.L("⭐ Favourites… (chuno)", "⭐ Favourites… (pick)"), "")
            for p in favs:
                self.fav_combo.addItem("⭐ " + (os.path.basename(p) or p), p)
            self.fav_combo.setEnabled(True)
        else:
            self.fav_combo.addItem(self.L("⭐ (koi favourite nahi)", "⭐ (no favourites yet)"), "")
            self.fav_combo.setEnabled(False)
        self.fav_combo.setCurrentIndex(0)
        self.fav_combo.blockSignals(False)

    def _on_fav_selected(self, _idx):
        p = self.fav_combo.currentData()
        if p:
            self._jump_to_folder(p)
        # wapas placeholder par le aao taaki dobara wahi chuna ja sake
        self.fav_combo.blockSignals(True)
        self.fav_combo.setCurrentIndex(0)
        self.fav_combo.blockSignals(False)

    def _fav_star_clicked(self):
        p = self._selected_library_folder()
        if p and os.path.isdir(p):
            self._toggle_fav(p)

    def _jump_to_folder(self, p):
        """Kisi folder me SEEDHA chalo (drill-in). Base ke andar ho to bas
        usme; bahar ho to use naya base bana do."""
        if not p or not os.path.isdir(p):
            return
        try:
            self.files_search.clear()
            self.files_results.hide(); self.files_tree.show()
        except Exception:
            pass
        base = os.path.normpath(self._files_root())
        npath = os.path.normpath(p)
        under = (npath == base) or npath.startswith(base + os.sep)
        if under:
            self._panel_show(p)
        else:
            self._opts["files_panel_root"] = p
            self._save_opts()
            self._refresh_files_root()

    def _files_result_activated(self, it):
        """Search-result par double-click: folder ho to panel me usme chale
        jao (uske andar ki files dikhein), file ho to use kholo."""
        p = it.data(QtCore.Qt.UserRole)
        if not p:
            return
        if os.path.isdir(p):
            self._jump_to_folder(p)
        else:
            self._open_path(p)

    def _on_result_clicked(self, it):
        """Search-result par EK click: document ho to uska preview panel me
        dikhao (PDF ka pehla page / image seedha). Folder par kuch nahi."""
        p = it.data(QtCore.Qt.UserRole)
        if p and os.path.isfile(p):
            self._preview_file_in_panel(p)

    def _pdf_text_cached(self, path):
        """PDF ke andar ka text (pehle 15 page) — content-search ke liye.
        Har file ka result yaad rakhte hain (path+time se) taaki dobara search
        turant ho. (Sirf digital text wali PDF me milta hai; poori scanned image
        PDF me embedded text nahi hota.)"""
        if not HAS_FITZ:
            return ""
        try:
            mt = os.path.getmtime(path)
        except Exception:
            return ""
        cache = getattr(self, "_pdf_text_cache", None)
        if cache is None:
            cache = self._pdf_text_cache = {}
        key = (path, mt)
        if key in cache:
            return cache[key]
        txt = ""
        try:
            doc = fitz.open(path)
            parts = []
            for i, page in enumerate(doc):
                if i >= 15:
                    break
                parts.append(page.get_text())
            txt = " ".join(parts).lower()
            doc.close()
        except Exception:
            txt = ""
        cache[key] = txt
        if len(cache) > 500:                 # cache bahut bada na ho
            try:
                cache.pop(next(iter(cache)))
            except Exception:
                pass
        return txt

    def _render_file_pixmap(self, path):
        """Kisi file ka QPixmap banao — image seedha, PDF ka pehla page fitz se."""
        ext = os.path.splitext(path)[1].lower()
        if ext in (".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff"):
            pm = QtGui.QPixmap(path)
            return pm if not pm.isNull() else None
        if ext == ".pdf" and HAS_FITZ:
            try:
                doc = fitz.open(path)
                page = doc.load_page(0)
                pix = page.get_pixmap(matrix=fitz.Matrix(1.6, 1.6), alpha=False)
                img = QtGui.QImage(pix.samples, pix.width, pix.height,
                                   pix.stride, QtGui.QImage.Format_RGB888)
                pm = QtGui.QPixmap.fromImage(img.copy())
                doc.close()
                return pm if not pm.isNull() else None
            except Exception:
                return None
        return None

    def _render_pdf_all_pages_qimage(self, path, cap=30, zoom=1.3):
        """PDF ke SAARE pages (cap tak) ko ek lambi image me jodo — user neeche
        scroll karke sab dekh sake. QImage/QPainter background thread me safe
        hai, isliye UI kabhi atkegi nahi. (qimage, total_pages) lautata hai."""
        try:
            doc = fitz.open(path)
        except Exception:
            return None, 0
        total = doc.page_count
        n = min(total, cap)
        imgs = []
        maxw = 1
        for i in range(n):
            try:
                pix = doc.load_page(i).get_pixmap(matrix=fitz.Matrix(zoom, zoom), alpha=False)
                qi = QtGui.QImage(pix.samples, pix.width, pix.height,
                                  pix.stride, QtGui.QImage.Format_RGB888).copy()
                imgs.append(qi); maxw = max(maxw, qi.width())
            except Exception:
                pass
        doc.close()
        if not imgs:
            return None, total
        gap = 14
        total_h = sum(im.height() for im in imgs) + gap * (len(imgs) + 1)
        canvas = QtGui.QImage(maxw, total_h, QtGui.QImage.Format_RGB888)
        canvas.fill(QtGui.QColor("#e5e7eb"))     # halka grey — pages ke beech gap dikhe
        p = QtGui.QPainter(canvas)
        y = gap
        for im in imgs:
            x = (maxw - im.width()) // 2
            p.fillRect(x, y, im.width(), im.height(), QtGui.QColor("#ffffff"))
            p.drawImage(x, y, im)
            y += im.height() + gap
        p.end()
        return canvas, total

    def _pv_file_date(self, path):
        """File ki modified date/time — complete info me dikhane ke liye."""
        try:
            return datetime.datetime.fromtimestamp(
                os.path.getmtime(path)).strftime("%d %b %Y, %I:%M %p")
        except Exception:
            return "-"

    def _pv_img_meta(self, path):
        """Image ki (DPI, colour-mode) — complete info me dikhane ke liye."""
        mode = "-"; dpi = None
        try:
            with Image.open(path) as im:
                mode = {"1": "Black & White", "L": "Grayscale", "RGB": "Colour",
                        "RGBA": "Colour"}.get(im.mode, im.mode)
                dpi = im.info.get("dpi")
        except Exception:
            pass
        dpitxt = ("%d dpi" % int(dpi[0])) if (dpi and dpi[0]) else "—"
        return dpitxt, mode

    def _preview_file_in_panel(self, path):
        """Preview panel me is file ki jhalak dikhao (band ho to khol do).
        Multi-page PDF ke SAARE pages neeche-neeche dikhte hain (scroll)."""
        if not getattr(self, "preview_panel", None):
            return
        if not self.preview_panel.isVisible():
            self.preview_panel.setVisible(True)
            self._opts["ui_preview"] = True
        name = os.path.basename(path)
        try:
            self.pv_tabs.setCurrentIndex(0)
        except Exception:
            pass
        self._pv_file_path = path            # race-guard (beech me doosri file to)
        try:
            kb = os.path.getsize(path) / 1024.0
            szt = ("%.0f KB" % kb) if kb < 1024 else ("%.1f MB" % (kb / 1024))
        except Exception:
            szt = "-"
        ext = os.path.splitext(path)[1].lower()

        # Multi-page PDF -> saare pages ek lambi image me (BACKGROUND me)
        if ext == ".pdf" and HAS_FITZ:
            self.pv_title.setText(self.L("👁 Khul raha hai…", "👁 Loading…"))

            def job():
                return self._render_pdf_all_pages_qimage(path)

            def done(res):
                if getattr(self, "_pv_file_path", None) != path:
                    return                    # tab tak user ne doosri file chun li
                qi, total = (res if isinstance(res, tuple) else (None, 0))
                if qi is None or qi.isNull():
                    self.pv_img.clear(); self._pv_pm = None
                    self.pv_title.setText(self.L("👁 Preview nahi bana", "👁 No preview"))
                    self.pv_info.setText(name); self.pv_text.setPlainText("")
                    return
                self._pv_pm = QtGui.QPixmap.fromImage(qi)
                self._pv_zoom = 1.0
                capd = self.L(" (pehle 30)", " (first 30)") if total > 30 else ""
                self.pv_title.setText("%s — %d %s%s"
                                      % (name, total, self.L("page", "pages"), capd))
                self._pv_render()
                QtCore.QTimer.singleShot(30, self._pv_render)
                self.pv_info.setText(
                    "<b>%s</b><br>🗂 PDF · 📄 %d %s<br>💾 %s<br>📅 %s<br>"
                    "<span style='color:#94a3b8'>%s</span>"
                    % (name, total, self.L("page", "pages"), szt,
                       self._pv_file_date(path), path))
                self.pv_text.setPlainText("")
            self._run_bg_quiet(job, done)
            return

        # Image / single -> seedha
        pm = self._render_file_pixmap(path)
        if pm is None:
            self.pv_img.clear(); self._pv_pm = None
            self.pv_title.setText(self.L("👁 Preview nahi bana", "👁 No preview"))
            self.pv_info.setText(name)
            self.pv_info2.setText("<b>%s</b><br><span style='color:#94a3b8'>%s</span>"
                                  % (name, path))
            self.pv_text.setPlainText("")
            return
        self._pv_pm = pm
        self._pv_zoom = 1.0
        self.pv_title.setText(name)
        self._pv_render()
        QtCore.QTimer.singleShot(30, self._pv_render)
        ext = (os.path.splitext(path)[1].lstrip(".").upper() or "FILE")
        dpitxt, mode = self._pv_img_meta(path)
        self.pv_info.setText(
            "<b>%s</b><br>🗂 %s · 📐 %d × %d px<br>🖨 %s · 🎨 %s<br>💾 %s<br>📅 %s<br>"
            "<span style='color:#94a3b8'>%s</span>"
            % (name, ext, pm.width(), pm.height(), dpitxt, mode, szt,
               self._pv_file_date(path), path))
        self.pv_text.setPlainText("")

    # ---- Fast in-memory index (turant search) ----
    _FILE_EXTS = (".pdf", ".jpg", ".jpeg", ".png", ".tif", ".tiff", ".docx", ".xlsx")

    def _invalidate_files_index(self):
        """Folder badla ya nayi file bani — index dobara banega."""
        self._files_index = None

    def _build_files_index(self, scope):
        """scope ke andar sabhi folder+file ka halka index (path + naam).
        Ye disk-walk sirf EK BAAR hota hai; uske baad har keystroke par sirf
        is list ko memory me chhaanto (bahut tez)."""
        idx = []
        try:
            for dp, dns, fn in os.walk(scope):
                for d in dns:
                    full = os.path.join(dp, d)
                    idx.append(("dir", full, d.lower(),
                                os.path.relpath(full, scope).lower()))
                for f in fn:
                    if f.lower().endswith(self._FILE_EXTS):
                        full = os.path.join(dp, f)
                        idx.append(("file", full, f.lower(),
                                    os.path.relpath(full, scope).lower()))
                if len(idx) >= 40000:
                    break
        except Exception:
            pass
        return idx

    def _ensure_files_index_async(self, scope):
        """scope ka index memory me tayyar rakho (background me). Jab tak na bane
        tab tak None rehta hai."""
        if getattr(self, "_files_index", None) is not None and \
                getattr(self, "_files_index_scope", None) == scope:
            return

        def _done(idx):
            self._files_index = idx
            self._files_index_scope = scope
            # index aate hi agar search-box me abhi bhi kuch likha hai to
            # turant natije dikha do
            try:
                if len(self.files_search.text().strip()) >= 2:
                    self._run_files_search()
            except Exception:
                pass
        self._run_bg_quiet(lambda: self._build_files_index(scope), _done)

    def _run_files_search(self):
        """Advanced search — POORE panel-folder (aur uske andar ke sabhi
        folders) me naam se dhoondo. Ab TURANT: disk baar-baar nahi padha jaata,
        ek in-memory index ko memory me chhaanta hai (2 akshar likhte hi natije).
        Kai shabd (jaise 'ram bill') = sabhi match; folder-naam bhi milta hai."""
        q = self.files_search.text().strip().lower()
        if len(q) < 2:
            self.files_results.hide()
            self.files_tree.show()
            return
        terms = [t for t in q.split() if t]
        scope = self._panel_current_dir()
        if not (scope and os.path.isdir(scope)):
            scope = self._files_root()
        search_text = self.btn_search_text.isChecked()   # PDF ke andar bhi?

        def _rank(name, by_content=False):
            if by_content:
                return (2, 0)
            in_name = all(t in name for t in terms)
            at_start = any(name.startswith(t) for t in terms)
            return (0 if in_name else 1, 0 if at_start else 1)

        # ---- TEZ RAASTA: naam/path search — memory index se turant ----
        if not search_text:
            idx = getattr(self, "_files_index", None)
            if idx is None or getattr(self, "_files_index_scope", None) != scope:
                # index abhi nahi bana — background me banwao, tab tak "…" dikhao
                self._ensure_files_index_async(scope)
                self.files_results.clear()
                _w = QtWidgets.QListWidgetItem(self.L("⏳ Taiyaari…", "⏳ Indexing…"))
                _w.setFlags(QtCore.Qt.NoItemFlags)
                self.files_results.addItem(_w)
                self.files_tree.hide(); self.files_results.show()
                return
            dir_hits, file_hits = [], []
            for kind, full, name, rel in idx:
                if kind == "dir":
                    if all(t in name for t in terms):
                        dir_hits.append((_rank(name), name, full))
                else:
                    if all(t in rel for t in terms):
                        file_hits.append((_rank(name), name, full))
                if len(dir_hits) + len(file_hits) >= 1000:
                    break
            dir_hits.sort(key=lambda h: (h[0], h[1]))
            file_hits.sort(key=lambda h: (h[0], h[1]))
            res = ([("dir", h[2]) for h in dir_hits] +
                   [("file", h[2]) for h in file_hits])
            self._render_files_results(res, q)
            return

        # ---- BHAARI RAASTA: PDF ke andar ka text — background me ----
        exts = self._FILE_EXTS

        def job():
            dir_hits, file_hits = [], []
            for dp, dns, fn in os.walk(scope):
                for d in dns:
                    name = d.lower()
                    if all(t in name for t in terms):
                        dir_hits.append((_rank(name), name, os.path.join(dp, d)))
                for f in fn:
                    if not f.lower().endswith(exts):
                        continue
                    full = os.path.join(dp, f)
                    name = f.lower()
                    rel = os.path.relpath(full, scope).lower()
                    if all(t in rel for t in terms):
                        file_hits.append((_rank(name), name, full))
                    elif full.lower().endswith(".pdf"):
                        txt = self._pdf_text_cached(full)
                        if txt and all(t in txt for t in terms):
                            file_hits.append((_rank(name, True), name, full))
                if len(dir_hits) + len(file_hits) >= 1000:
                    break
            dir_hits.sort(key=lambda h: (h[0], h[1]))
            file_hits.sort(key=lambda h: (h[0], h[1]))
            return ([("dir", h[2]) for h in dir_hits] +
                    [("file", h[2]) for h in file_hits])

        def done(res):
            if isinstance(res, Exception):
                return
            if self.files_search.text().strip().lower() != q:
                return
            self._render_files_results(res, q)
        self._run_bg(job, done, self.L("Andar ka text dhoondh rahe…",
                                       "Searching inside text…"))

    def _render_files_results(self, res, q=None):
        """Search/filter ke natije (list) ko panel me dikhao. q diya ho to sirf
        tabhi render karo jab search-box me abhi bhi wahi likha ho."""
        if q is not None and self.files_search.text().strip().lower() != q:
            return                          # tab tak nayi search shuru ho gayi
        self._last_files_res = res          # grid toggle par dobara render ke liye
        grid = bool(self._opts.get("files_grid"))
        self.files_results.clear()
        # FOLDER ke hisaab se group karo: upar folder ka naam (header),
        # neeche usi folder ke documents.
        groups, order = {}, []
        for kind, p in res:
            if kind == "file":
                par = os.path.dirname(p)
                if par not in groups:
                    groups[par] = []; order.append(par)
                groups[par].append(p)
        for kind, p in res:                     # naam-se-mile khali folder bhi
            if kind == "dir" and p not in groups:
                groups[p] = []; order.append(p)
        order.sort(key=lambda f: os.path.basename(f).lower())
        n_dir = len(order)
        n_file = sum(len(v) for v in groups.values())
        if grid:
            self.files_results.setViewMode(QtWidgets.QListView.IconMode)
            self.files_results.setIconSize(QtCore.QSize(96, 124))
            self.files_results.setGridSize(QtCore.QSize(108, 150))
            self.files_results.setResizeMode(QtWidgets.QListView.Adjust)
            self.files_results.setMovement(QtWidgets.QListView.Static)
        else:
            self.files_results.setViewMode(QtWidgets.QListView.ListMode)
            self.files_results.setGridSize(QtCore.QSize())
        head = QtWidgets.QListWidgetItem(
            self.L("🔎 %d folder · %d file" % (n_dir, n_file),
                   "🔎 %d folders · %d files" % (n_dir, n_file))
            + ("+" if len(res) >= 1000 else ""))
        head.setFlags(QtCore.Qt.NoItemFlags)
        _hf = head.font(); _hf.setBold(True); head.setFont(_hf)
        head.setForeground(QtGui.QColor("#0f766e"))
        if not grid:
            self.files_results.addItem(head)
        _icon = {".pdf": "📕", ".docx": "📘", ".xlsx": "📗"}
        pinned = set(self._opts.get("pinned_files") or [])
        grid_files = []
        for folder in order:
            if not grid:
                fh = QtWidgets.QListWidgetItem("📁  " + (os.path.basename(folder) or folder))
                _ff = fh.font(); _ff.setBold(True); fh.setFont(_ff)
                fh.setForeground(QtGui.QColor("#0f766e"))
                fh.setBackground(QtGui.QColor("#eef4f3"))
                fh.setToolTip(folder + self.L("   (folder — 2x click = kholo)",
                                              "   (folder — double-click to open)"))
                fh.setData(QtCore.Qt.UserRole, folder)
                self.files_results.addItem(fh)
            for p in sorted(groups[folder], key=lambda x: os.path.basename(x).lower()):
                ext = os.path.splitext(p)[1].lower()
                pin = "📌 " if p in pinned else ""
                if grid:
                    it = QtWidgets.QListWidgetItem(pin + os.path.basename(p))
                    it.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignTop)
                    it.setSizeHint(QtCore.QSize(108, 150))
                    grid_files.append((self.files_results.count(), p))
                else:
                    it = QtWidgets.QListWidgetItem("      " + pin + _icon.get(ext, "🖼")
                                                   + "  " + os.path.basename(p))
                it.setToolTip(p + self.L("   (click = preview · 2x = kholo · drag = import)",
                                         "   (click = preview · double-click = open · drag = import)"))
                it.setData(QtCore.Qt.UserRole, p)
                self.files_results.addItem(it)
        if not order:
            it = QtWidgets.QListWidgetItem(self.L("(kuch nahi mila)", "(nothing found)"))
            it.setFlags(QtCore.Qt.NoItemFlags)
            self.files_results.addItem(it)
        self.files_tree.hide()
        self.files_results.show()
        if grid and grid_files:
            self._build_grid_thumbs(grid_files[:60])

    def _build_grid_thumbs(self, items):
        """Grid view ke liye har file ki thumbnail (PDF ka pehla page ya image)
        background me banao, phir icon laga do — UI kabhi atke nahi."""
        rows = [(r, p) for r, p in items]

        def job():
            out = []
            for r, p in rows:
                png = None
                try:
                    ext = os.path.splitext(p)[1].lower()
                    if ext == ".pdf" and HAS_FITZ:
                        doc = fitz.open(p)
                        page = doc.load_page(0)
                        pix = page.get_pixmap(matrix=fitz.Matrix(0.5, 0.5), alpha=False)
                        fd, png = tempfile.mkstemp(suffix=".png", dir=self._tmpdir)
                        os.close(fd); pix.save(png); doc.close()
                    elif ext in (".jpg", ".jpeg", ".png", ".tif", ".tiff"):
                        png = p
                except Exception:
                    png = None
                out.append((r, png))
            return out

        def done(res):
            if isinstance(res, Exception):
                return
            for r, png in res:
                if not png:
                    continue
                try:
                    it = self.files_results.item(r)
                    if it is not None and it.data(QtCore.Qt.UserRole):
                        it.setIcon(QtGui.QIcon(self._make_thumb(png)))
                except Exception:
                    pass
        self._run_bg_quiet(job, done)

    def _toggle_files_grid(self, on):
        self._opts["files_grid"] = bool(on)
        self._save_opts()
        # jo abhi dikh raha hai use usi tarah dobara render karo
        if self.files_results.isVisible() and getattr(self, "_last_files_res", None) is not None:
            self._render_files_results(self._last_files_res)
        else:
            self._apply_files_filter()

    def _files_index_ready(self, scope):
        """scope ka index chahiye — na ho to background me banwao, ready hone par
        given callback ke bajaye seedha filter/recent dobara chala do."""
        idx = getattr(self, "_files_index", None)
        if idx is not None and getattr(self, "_files_index_scope", None) == scope:
            return idx
        self._ensure_files_index_async(scope)
        return None

    def _apply_files_filter(self):
        """Filter dropdown: sab / PDF / photo / aaj / hafta / pinned."""
        mode = self.files_filter.currentData() if hasattr(self, "files_filter") else "all"
        if mode in (None, "all"):
            # normal folder-browser wapas
            try:
                self.files_search.clear()
            except Exception:
                pass
            self.files_results.hide(); self.files_tree.show()
            return
        scope = self._panel_current_dir()
        if not (scope and os.path.isdir(scope)):
            scope = self._files_root()
        if mode == "pinned":
            res = [("file", p) for p in (self._opts.get("pinned_files") or [])
                   if os.path.isfile(p)]
            self._render_files_results(res)
            return
        idx = self._files_index_ready(scope)
        if idx is None:
            self.files_results.clear()
            _w = QtWidgets.QListWidgetItem(self.L("⏳ Taiyaari…", "⏳ Indexing…"))
            _w.setFlags(QtCore.Qt.NoItemFlags)
            self.files_results.addItem(_w)
            self.files_tree.hide(); self.files_results.show()
            QtCore.QTimer.singleShot(400, self._apply_files_filter)
            return
        files = [(full) for kind, full, name, rel in idx if kind == "file"]
        if mode == "pdf":
            res = [("file", p) for p in files if p.lower().endswith(".pdf")]
        elif mode == "img":
            res = [("file", p) for p in files
                   if p.lower().endswith((".jpg", ".jpeg", ".png", ".tif", ".tiff"))]
        elif mode in ("today", "week"):
            import time as _t
            now = _t.time()
            span = 86400 if mode == "today" else 7 * 86400
            hit = []
            for p in files:
                try:
                    if now - os.path.getmtime(p) <= span:
                        hit.append(p)
                except Exception:
                    pass
            res = [("file", p) for p in hit]
        else:
            res = [("file", p) for p in files]
        self._render_files_results(res)

    def _show_recent_files(self):
        """Haal me bani/badli files — naye sabse upar (poore scope me)."""
        scope = self._panel_current_dir()
        if not (scope and os.path.isdir(scope)):
            scope = self._files_root()
        idx = self._files_index_ready(scope)
        if idx is None:
            self.status.showMessage(self.L("Taiyaari… phir se dabao.",
                                           "Indexing… tap again."), 2000)
            return
        files = [full for kind, full, name, rel in idx if kind == "file"]

        def job():
            dated = []
            for p in files:
                try:
                    dated.append((os.path.getmtime(p), p))
                except Exception:
                    pass
            dated.sort(reverse=True)
            return [("file", p) for _m, p in dated[:60]]

        def done(res):
            if isinstance(res, Exception):
                return
            try:
                self.files_filter.blockSignals(True)
                self.files_filter.setCurrentIndex(0)
                self.files_filter.blockSignals(False)
            except Exception:
                pass
            self._render_files_results(res)
        self._run_bg(job, done, self.L("Haal ki files…", "Recent files…"))

    def _update_folder_info(self):
        """Is folder me kitni files + kul size — chhoti patti me dikhao."""
        scope = self._panel_current_dir()
        if not (scope and os.path.isdir(scope)):
            self.lbl_folder_info.setText("")
            return

        def job():
            n = 0; total = 0
            try:
                for dp, dns, fn in os.walk(scope):
                    for f in fn:
                        if f.lower().endswith(self._FILE_EXTS):
                            n += 1
                            try:
                                total += os.path.getsize(os.path.join(dp, f))
                            except Exception:
                                pass
                    if n >= 40000:
                        break
            except Exception:
                pass
            return (n, total)

        def done(res):
            if isinstance(res, Exception) or not isinstance(res, tuple):
                return
            n, total = res
            gb = total / (1024.0 ** 3)
            sz = ("%.1f GB" % gb) if gb >= 1 else ("%.0f MB" % (total / 1048576.0))
            self.lbl_folder_info.setText(self.L("📦 %d files · %s" % (n, sz),
                                                "📦 %d files · %s" % (n, sz)))
        self._run_bg_quiet(job, done)

    def _merge_folder_pdfs(self, folder):
        if not HAS_OCR_LIBS:
            self._warn("pypdf is not installed.")
            return
        pdfs = sorted(os.path.join(folder, f) for f in os.listdir(folder)
                      if f.lower().endswith(".pdf"))
        if len(pdfs) < 2:
            self._warn("This folder has fewer than 2 PDFs.")
            return
        out, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, 'Save the merged PDF',
            os.path.join(folder, (os.path.basename(folder) or "sab") + "_sab.pdf"),
            "PDF (*.pdf)")
        if not out:
            return
        try:
            writer = PdfWriter()
            used = 0
            for p in pdfs:
                if os.path.abspath(p) == os.path.abspath(out):
                    continue
                try:
                    for pg in PdfReader(p).pages:
                        writer.add_page(pg)
                    used += 1
                except Exception:
                    pass
            with open(out, "wb") as fh:
                writer.write(fh)
        except Exception:
            self._warn("Merge failed:\n%s" % traceback.format_exc())
            return
        QtWidgets.QMessageBox.information(
            self, "Done", "Merged %d PDFs into one:\n%s" % (used, out))

    def _rename_library_file(self, path):
        stem, ext = os.path.splitext(os.path.basename(path))
        name, ok = QtWidgets.QInputDialog.getText(self, "Rename", "New name:", text=stem)
        if not ok or not name.strip():
            return
        new = os.path.join(os.path.dirname(path),
                           sanitize(underscore_name(name.strip())) + ext)
        try:
            os.rename(path, new)
        except Exception as exc:
            self._warn("Rename failed: %s" % exc)

    def _delete_library_file(self, path):
        if QtWidgets.QMessageBox.question(
                self, "Delete",
                "Move '%s' to the Recycle Bin?\n(You can restore it later.)"
                % os.path.basename(path),
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.No) != QtWidgets.QMessageBox.Yes:
            return
        if self._trash_file(path):
            self.status.showMessage(self.L("🗑 Recycle Bin me daal diya — wapas la sakte hain",
                                           "🗑 Moved to Recycle Bin — you can restore it"), 5000)
        else:
            self._warn(self.L("Delete fail ho gaya", "Delete failed"))

    # ---------- Recycle Bin (kachra-peti) ----------
    def _trash_index_path(self):
        return os.path.join(TRASH_DIR, "index.json")

    def _load_trash_index(self):
        import json
        try:
            with open(self._trash_index_path(), "r", encoding="utf-8") as fh:
                return json.load(fh) or []
        except Exception:
            return []

    def _save_trash_index(self, items):
        import json
        try:
            os.makedirs(TRASH_DIR, exist_ok=True)
            with open(self._trash_index_path(), "w", encoding="utf-8") as fh:
                json.dump(items, fh, ensure_ascii=False)
        except Exception:
            pass

    def _trash_file(self, path):
        """File ko Recycle Bin me le jao (delete karne ke bajay) — wapas laai
        ja sake."""
        import uuid
        try:
            os.makedirs(TRASH_DIR, exist_ok=True)
            base = os.path.basename(path)
            trash_name = uuid.uuid4().hex[:12] + "__" + base
            shutil.move(path, os.path.join(TRASH_DIR, trash_name))
            idx = self._load_trash_index()
            idx.append({"trash": trash_name, "orig": path, "name": base,
                        "when": datetime.datetime.now().strftime("%Y-%m-%d %H:%M")})
            self._save_trash_index(idx)
            return True
        except Exception:
            return False

    def show_recycle_bin(self):
        """Delete ki hui files — wapas laao ya hamesha ke liye hatao."""
        idx = self._load_trash_index()
        # sirf wahi jo sach me bin me maujood hain
        idx = [e for e in idx if os.path.exists(os.path.join(TRASH_DIR, e.get("trash", "")))]
        self._save_trash_index(idx)
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle(self.L("🗑 Recycle Bin", "🗑 Recycle Bin"))
        dlg.resize(560, 460)
        v = QtWidgets.QVBoxLayout(dlg)
        v.addWidget(QtWidgets.QLabel(self.L(
            "Delete ki hui files yahan hain. Chuno phir 'Wapas laao' ya 'Hamesha ke liye hatao'.",
            "Deleted files are here. Select, then Restore or Delete forever.")))
        lw = QtWidgets.QListWidget()
        lw.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        for e in reversed(idx):     # naye upar
            it = QtWidgets.QListWidgetItem("📄 %s\n       %s · %s" %
                                           (e.get("name", "?"), e.get("when", ""),
                                            os.path.dirname(e.get("orig", ""))))
            it.setData(QtCore.Qt.UserRole, e)
            lw.addItem(it)
        if not idx:
            it = QtWidgets.QListWidgetItem(self.L("(Recycle Bin khaali hai)", "(Recycle Bin is empty)"))
            it.setFlags(QtCore.Qt.NoItemFlags)
            lw.addItem(it)
        v.addWidget(lw, 1)

        def _restore():
            picked = [i.data(QtCore.Qt.UserRole) for i in lw.selectedItems() if i.data(QtCore.Qt.UserRole)]
            if not picked:
                return
            cur = self._load_trash_index()
            done = 0
            for e in picked:
                src = os.path.join(TRASH_DIR, e["trash"])
                dst = e["orig"]
                try:
                    os.makedirs(os.path.dirname(dst), exist_ok=True)
                    n = 2
                    while os.path.exists(dst):
                        stem, ext = os.path.splitext(e["orig"])
                        dst = "%s_%d%s" % (stem, n, ext); n += 1
                    shutil.move(src, dst)
                    cur = [c for c in cur if c.get("trash") != e["trash"]]
                    done += 1
                except Exception:
                    pass
            self._save_trash_index(cur)
            self.status.showMessage(self.L("♻ %d file wapas aa gayi" % done,
                                           "♻ Restored %d file(s)" % done), 5000)
            self._refresh_files_root()
            dlg.accept()

        def _delete_forever():
            picked = [i.data(QtCore.Qt.UserRole) for i in lw.selectedItems() if i.data(QtCore.Qt.UserRole)]
            if not picked:
                return
            if QtWidgets.QMessageBox.question(
                    dlg, self.L("Pakka?", "Sure?"),
                    self.L("%d file HAMESHA ke liye hat jayengi (wapas nahi aayengi). Theek?",
                           "%d file(s) will be deleted FOREVER (cannot restore). OK?") % len(picked),
                    QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                    QtWidgets.QMessageBox.No) != QtWidgets.QMessageBox.Yes:
                return
            cur = self._load_trash_index()
            for e in picked:
                try:
                    os.remove(os.path.join(TRASH_DIR, e["trash"]))
                except Exception:
                    pass
                cur = [c for c in cur if c.get("trash") != e["trash"]]
            self._save_trash_index(cur)
            dlg.accept()

        def _empty():
            if not idx:
                return
            if QtWidgets.QMessageBox.question(
                    dlg, self.L("Bin khaali karein?", "Empty bin?"),
                    self.L("Saari files HAMESHA ke liye hat jayengi. Theek?",
                           "All files will be deleted FOREVER. OK?"),
                    QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                    QtWidgets.QMessageBox.No) != QtWidgets.QMessageBox.Yes:
                return
            for e in self._load_trash_index():
                try:
                    os.remove(os.path.join(TRASH_DIR, e["trash"]))
                except Exception:
                    pass
            self._save_trash_index([])
            dlg.accept()

        bb = QtWidgets.QHBoxLayout()
        b1 = QtWidgets.QPushButton(self.L("♻ Wapas laao", "♻ Restore")); b1.clicked.connect(_restore)
        b2 = QtWidgets.QPushButton(self.L("❌ Hamesha ke liye hatao", "❌ Delete forever")); b2.clicked.connect(_delete_forever)
        b3 = QtWidgets.QPushButton(self.L("🧹 Bin khaali karo", "🧹 Empty bin")); b3.clicked.connect(_empty)
        b4 = QtWidgets.QPushButton("OK"); b4.clicked.connect(dlg.accept)
        bb.addWidget(b1); bb.addWidget(b2); bb.addWidget(b3); bb.addStretch(1); bb.addWidget(b4)
        v.addLayout(bb)
        dlg.exec_()

    # ---------- Multi-select bulk (kai files ek saath) ----------
    def _selected_library_files(self):
        out, seen = [], set()
        try:
            for idx in self.files_tree.selectionModel().selectedIndexes():
                if idx.column() != 0:
                    continue
                p = self.files_model.filePath(idx)
                if p and os.path.isfile(p) and p not in seen:
                    seen.add(p); out.append(p)
        except Exception:
            pass
        return out

    def _bulk_delete(self, files):
        if QtWidgets.QMessageBox.question(
                self, "Delete",
                self.L("%d files Recycle Bin me daalein? (wapas la sakte hain)",
                       "Move %d files to Recycle Bin? (restorable)") % len(files),
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.No) != QtWidgets.QMessageBox.Yes:
            return
        n = sum(1 for f in files if self._trash_file(f))
        self.status.showMessage(self.L("🗑 %d file Recycle Bin me" % n,
                                       "🗑 %d file(s) to Recycle Bin" % n), 5000)

    def _bulk_merge(self, files):
        if not HAS_OCR_LIBS:
            self._warn("pypdf is not installed (required for merge)."); return
        pdfs = [f for f in files if f.lower().endswith(".pdf")]
        if len(pdfs) < 2:
            self._warn(self.L("Merge ke liye kam se kam 2 PDF chuno.",
                              "Pick at least 2 PDFs to merge.")); return
        folder = os.path.dirname(pdfs[0])
        out, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, self.L("Jodi hui PDF save karein", "Save merged PDF"),
            os.path.join(folder, "merged.pdf"), "PDF (*.pdf)")
        if not out:
            return

        def job():
            writer = PdfWriter()
            used = 0
            for p in pdfs:
                if os.path.abspath(p) == os.path.abspath(out):
                    continue
                try:
                    for pg in PdfReader(p).pages:
                        writer.add_page(pg)
                    used += 1
                except Exception:
                    pass
            with open(out, "wb") as fh:
                writer.write(fh)
            return used

        def on_done(res):
            if isinstance(res, Exception):
                self._warn("Merge failed:\n%s" % res); return
            self.status.showMessage(self.L("🧩 %d PDF jud kar ek ban gayi" % res,
                                           "🧩 Merged %d PDFs into one" % res), 6000)
            self._refresh_files_root()
        self._run_bg(job, on_done, self.L("PDFs jud rahi hain…", "Merging PDFs…"))

    def _bulk_move(self, files, copy=False):
        dest = QtWidgets.QFileDialog.getExistingDirectory(
            self, self.L("Kaunse folder me?", "Into which folder?"),
            self._panel_current_dir())
        if not dest:
            return

        def job():
            done = 0
            for f in files:
                try:
                    dst = os.path.join(dest, os.path.basename(f))
                    stem, ext = os.path.splitext(os.path.basename(f))
                    k = 2
                    while os.path.exists(dst):
                        dst = os.path.join(dest, "%s_%d%s" % (stem, k, ext)); k += 1
                    if copy:
                        shutil.copy2(f, dst)
                    else:
                        shutil.move(f, dst)
                    done += 1
                except Exception:
                    pass
            return done

        def on_done(res):
            if isinstance(res, Exception):
                self._warn("Failed:\n%s" % res); return
            self.status.showMessage(
                (self.L("📄 %d file copy ho gayi", "📄 Copied %d file(s)") if copy
                 else self.L("📁 %d file le jaayi gayi", "📁 Moved %d file(s)")) % res, 5000)
            self._refresh_files_root()
        self._run_bg(job, on_done, self.L("Ho raha hai…", "Working…"))

    def toggle_files_panel(self):
        vis = not self.files_panel.isVisible()
        self.files_panel.setVisible(vis)
        self._opts["show_files_panel"] = vis
        self._save_opts()

    def _update_preview_panel(self):
        if not getattr(self, "preview_panel", None) or not self.preview_panel.isVisible():
            return
        # filmstrip me current page highlight rakho
        try:
            if hasattr(self, "pv_strip"):
                cur = self.list.currentRow()
                if self.pv_strip.count() != self.list.count():
                    self._pv_build_filmstrip()
                if 0 <= cur < self.pv_strip.count():
                    self.pv_strip.blockSignals(True)
                    self.pv_strip.setCurrentRow(cur)
                    self.pv_strip.scrollToItem(self.pv_strip.item(cur))
                    self.pv_strip.blockSignals(False)
        except Exception:
            pass
        it = self.list.currentItem()
        if it is None:
            self.pv_img.clear(); self.pv_info.setText("")
            self.pv_title.setText(self.L("👁 Preview", "👁 Preview"))
            self.pv_info2.setText(""); self.pv_text.setPlainText("")
            self._pv_pm = None
            return
        path = it.data(QtCore.Qt.UserRole)
        row = self.list.row(it)
        n = self.list.count()
        try:
            self._pv_pm = QtGui.QPixmap(path)
            self.pv_title.setText("Page %d / %d" % (row + 1, n))
            self._pv_zoom = 1.0
            self._pv_render()
            name = it.data(TITLE_ROLE) or it.text() or "-"
            kb = os.path.getsize(path) / 1024.0
            szt = ("%.0f KB" % kb) if kb < 1024 else ("%.1f MB" % (kb / 1024))
            self.pv_info.setText("%s · %s · %dx%d" %
                                 (name, szt, self._pv_pm.width(), self._pv_pm.height()))
            # Info tab (mode · DPI · tareekh bhi)
            mode = "-"; dpi = None
            try:
                with Image.open(path) as im:
                    mode = {"1": "Black & White", "L": "Grayscale", "RGB": "Colour",
                            "RGBA": "Colour"}.get(im.mode, im.mode)
                    dpi = im.info.get("dpi")
            except Exception:
                pass
            try:
                mt = datetime.datetime.fromtimestamp(os.path.getmtime(path))
                dt = mt.strftime("%d %b %Y, %I:%M %p")
            except Exception:
                dt = "-"
            dpitxt = ("%d dpi" % int(dpi[0])) if (dpi and dpi[0]) else "—"
            self.pv_info2.setText(
                "<b>%s</b><br>📄 Page %d / %d<br>📐 %d × %d px<br>🖨 %s<br>🎨 %s<br>"
                "💾 %s<br>🕒 %s<br><span style='color:#94a3b8'>%s</span>"
                % (name, row + 1, n, self._pv_pm.width(), self._pv_pm.height(),
                   dpitxt, mode, szt, dt, path))
            # Text tab: purana text mita do (naya sirf 'Text padho' par)
            self.pv_text.setPlainText("")
        except Exception:
            pass

    def _pv_render(self):
        pm = getattr(self, "_pv_pm", None)
        if pm is None or pm.isNull():
            return
        if self._pv_zoom <= 0:
            self._pv_zoom = 1.0
        base_w = self.pv_scroll.viewport().width() - 6
        w = int(base_w * self._pv_zoom)
        scaled = pm.scaledToWidth(max(40, w), QtCore.Qt.SmoothTransformation)
        self.pv_img.setPixmap(scaled)
        self.pv_img.resize(scaled.size())

    def _pv_do_zoom(self, factor):
        self._pv_zoom = max(0.2, min(6.0, self._pv_zoom * factor))
        self._pv_render()

    def _pv_fit(self):
        self._pv_zoom = 1.0
        self._pv_render()

    def _pv_step(self, d):
        r = self.list.currentRow()
        if r < 0:
            r = 0
        nr = r + d
        if 0 <= nr < self.list.count():
            self.list.setCurrentRow(nr)

    def _pv_fullscreen(self):
        it = self.list.currentItem()
        if not it:
            return
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle(self.L("Poori-screen preview", "Full-screen preview"))
        dlg.resize(900, 720)
        v = QtWidgets.QVBoxLayout(dlg)
        sc = QtWidgets.QScrollArea(); sc.setWidgetResizable(True)
        lbl = QtWidgets.QLabel(); lbl.setAlignment(QtCore.Qt.AlignCenter)
        pm = QtGui.QPixmap(it.data(QtCore.Qt.UserRole))
        lbl.setPixmap(pm.scaled(880, 680, QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation))
        sc.setWidget(lbl)
        v.addWidget(sc, 1)
        b = QtWidgets.QPushButton(self.L("Band karo", "Close")); b.clicked.connect(dlg.accept)
        v.addWidget(b)
        dlg.exec_()

    def _pv_open_editor(self):
        it = self.list.currentItem()
        if it:
            PreviewDialog(self, self.list.row(it)).exec_()
            self._update_preview_panel()

    def _pv_read_text(self):
        it = self.list.currentItem()
        if not it:
            return
        if not tesseract_available():
            self.pv_text.setPlainText(self.L("Tesseract OCR install nahi hai.",
                                             "Tesseract OCR is not installed."))
            return
        path = it.data(QtCore.Qt.UserRole)
        self.pv_text.setPlainText(self.L("Padh rahe hain…", "Reading…"))

        def job():
            with Image.open(path) as im:
                return pytesseract.image_to_string(im, lang="hin+eng")

        def done(res):
            self.pv_text.setPlainText("" if isinstance(res, Exception) else (res or ""))
        self._run_bg(job, done, self.L("Text padh rahe hain…", "Reading text…"))

    def _pv_backup(self, paths):
        """Undo ke liye: edit se pehle in pages ki asli copy rakh lo (1 level)."""
        self._pv_undo_backup = {}
        for p in paths:
            try:
                fd, tmp = tempfile.mkstemp(suffix=os.path.splitext(p)[1] or ".png")
                os.close(fd)
                shutil.copy2(p, tmp)
                self._pv_undo_backup[p] = tmp
            except Exception:
                pass

    def _pv_undo(self):
        """Aakhri edit wapas lo — backup se pages bahal karo."""
        bak = getattr(self, "_pv_undo_backup", None)
        if not bak:
            self.status.showMessage(self.L("Undo ke liye kuch nahi", "Nothing to undo"), 3000)
            return
        n = 0
        for p, tmp in list(bak.items()):
            try:
                if os.path.exists(tmp):
                    shutil.copy2(tmp, p)
                    os.remove(tmp)
                    n += 1
            except Exception:
                pass
        self._pv_undo_backup = {}
        self._refresh_all_thumbs()
        self._update_preview_panel()
        self.status.showMessage(self.L("↶ %d page wapas aaye" % n,
                                       "↶ Reverted %d page(s)" % n), 3000)

    def _edit_targets(self):
        """Edit kispar lage: apply-all on ho to SAB pages, warna current."""
        if getattr(self, "pv_apply_all", None) and self.pv_apply_all.isChecked():
            paths = self._ordered_paths()
            return paths, True
        item = self._current_item_or_warn()
        if not item:
            return [], False
        return [item.data(QtCore.Qt.UserRole)], False

    def _refresh_all_thumbs(self):
        for i in range(self.list.count()):
            try:
                self._refresh_item(self.list.item(i))
            except Exception:
                pass
        self._pv_build_filmstrip()

    def _edit_current_bg(self, transform, busy_msg):
        """Image-transform BACKGROUND me — UI nahi rukti. Apply-all ho to sab
        pages par; undo ke liye pehle backup. transform: PIL.Image(RGB)->Image."""
        paths, allmode = self._edit_targets()
        if not paths:
            return
        self._pv_backup(paths)

        def job():
            for p in paths:
                try:
                    with Image.open(p) as im:
                        transform(im.convert("RGB")).save(p, "PNG")
                except Exception:
                    pass
            return True

        def on_done(res):
            if isinstance(res, Exception):
                self._warn(self.L("Edit fail: %s", "Edit failed: %s") % res)
                return
            if allmode:
                self._refresh_all_thumbs()
            else:
                try:
                    self._refresh_item(self.list.currentItem())
                except Exception:
                    pass
                self._pv_build_filmstrip()
            self._update_preview_panel()
            self._dirty = True
        self._run_bg(job, on_done, busy_msg)

    def deskew_current(self):
        self._edit_current_bg(lambda im: deskew(im),
                              self.L("Seedha kar rahe hain…", "Straightening…"))

    def enhance_current_page(self):
        self._edit_current_bg(lambda im: auto_enhance(im),
                              self.L("Saaf kar rahe hain…", "Enhancing…"))

    def whiten_current_page(self):
        self._edit_current_bg(lambda im: whiten_dark_background(im),
                              self.L("Backing safed kar rahe hain…", "Whitening…"))

    def rotate_any(self):
        """Kisi bhi angle (jaise 2°, -3°) par ghumakar seedha karo."""
        deg, ok = QtWidgets.QInputDialog.getDouble(
            self, self.L("Kitne degree?", "How many degrees?"),
            self.L("Ghumao (+ = daayein, − = baayein):", "Rotate (+ = right, − = left):"),
            0.0, -180.0, 180.0, 1)
        if not ok or abs(deg) < 0.01:
            return
        self._edit_current_bg(
            lambda im: im.rotate(-deg, expand=True, fillcolor=(255, 255, 255),
                                 resample=Image.BICUBIC),
            self.L("Ghuma rahe hain…", "Rotating…"))

    def _to_mode(self, mode):
        """Page ko B&W ('1') / Grayscale ('L') me badlo (RGB me wapas dikhega)."""
        def _t(im):
            if mode == "1":
                return im.convert("L").point(lambda x: 255 if x > 150 else 0, "1").convert("RGB")
            if mode == "L":
                return im.convert("L").convert("RGB")
            return im
        self._edit_current_bg(_t, self.L("Badal rahe hain…", "Converting…"))

    def duplicate_current_page(self):
        """Current page ki ek nakal uske theek baad jod do."""
        item = self._current_item_or_warn()
        if not item:
            return
        src = item.data(QtCore.Qt.UserRole)
        try:
            fd, dst = tempfile.mkstemp(suffix=os.path.splitext(src)[1] or ".png", dir=self._tmpdir)
            os.close(fd)
            shutil.copy2(src, dst)
        except Exception as exc:
            self._warn("Copy failed: %s" % exc); return
        row = self.list.row(item)
        self._add_item_for_path(dst)               # end me add hota hai
        # use theek current ke baad le aao
        last = self.list.count() - 1
        it2 = self.list.takeItem(last)
        self.list.insertItem(row + 1, it2)
        self.list.setCurrentItem(it2)
        self._renumber_pages()
        self._pv_build_filmstrip()
        self.status.showMessage(self.L("⧉ Page ki nakal ban gayi", "⧉ Page duplicated"), 3000)

    def _pv_translate(self):
        txt = self.pv_text.toPlainText().strip()
        if not txt:
            self._pv_read_text()
            QtCore.QTimer.singleShot(1200, self._pv_translate_now)
            return
        self._pv_translate_now()

    def _pv_translate_now(self):
        txt = self.pv_text.toPlainText().strip()
        if not txt:
            self.status.showMessage(self.L("Pehle 'Text padho' dabao", "Press 'Read text' first"), 3000)
            return

        def job():
            try:
                import urllib.request as U, urllib.parse as P, json as J, ssl
                ctx = ssl._create_unverified_context()
                q = P.urlencode({"client": "gtx", "sl": "auto", "tl": "en",
                                 "dt": "t", "q": txt[:1800]})
                url = "https://translate.googleapis.com/translate_a/single?" + q
                r = U.urlopen(U.Request(url, headers={"User-Agent": "Mozilla/5.0"}),
                              timeout=15, context=ctx)
                data = J.loads(r.read().decode("utf-8", "ignore"))
                return "".join(seg[0] for seg in data[0] if seg and seg[0])
            except Exception as e:
                return e

        def done(res):
            if isinstance(res, Exception) or not res:
                self.status.showMessage(self.L("Translate nahi ho paya", "Translate failed"), 3000)
                return
            self.pv_tabs.setCurrentIndex(1)
            self.pv_text.setPlainText(res)
        self._run_bg(job, done, self.L("Translate ho raha hai…", "Translating…"))

    def _pv_strip_click(self, item):
        r = item.data(QtCore.Qt.UserRole)
        if isinstance(r, int) and 0 <= r < self.list.count():
            self.list.setCurrentRow(r)

    def _pv_build_filmstrip(self):
        if not hasattr(self, "pv_strip"):
            return
        try:
            self.pv_strip.blockSignals(True)
            self.pv_strip.clear()
            for i in range(self.list.count()):
                src = self.list.item(i)
                it = QtWidgets.QListWidgetItem(src.icon(), str(i + 1))
                it.setData(QtCore.Qt.UserRole, i)
                it.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignBottom)
                self.pv_strip.addItem(it)
            cur = self.list.currentRow()
            if 0 <= cur < self.pv_strip.count():
                self.pv_strip.setCurrentRow(cur)
            self.pv_strip.blockSignals(False)
        except Exception:
            pass

    def _pv_more_menu(self):
        m = QtWidgets.QMenu(self)
        m.addAction(self.L("🎨 Sudhaar editor (crop/miṭao/text/teer…)",
                           "🎨 Edit (crop/erase/text/arrow…)"), self._pv_open_image_editor)
        m.addAction(self.L("🔍 Bade editor me kholo", "🔍 Open big editor"), self._pv_open_editor)
        m.addAction(self.L("🎨 Rang wapas (RGB)", "🎨 Back to colour (RGB)"),
                    lambda: self._to_mode("RGB"))
        m.addSeparator()
        m.addAction(self.L("↕ Do page saath-saath (compare)", "↕ Compare two pages side by side"),
                    self._pv_compare)
        m.addAction(self.L("🔎 Loupe (kaanch) on/off", "🔎 Loupe (magnifier) on/off"),
                    self._pv_toggle_loupe)
        m.addAction(self.L("▶ Slideshow (auto page)", "▶ Slideshow (auto-advance)"),
                    self._pv_slideshow)
        m.addSeparator()
        m.addAction(self.L("⬆ Ye page upar karo", "⬆ Move page up"), lambda: self._pv_move(-1))
        m.addAction(self.L("⬇ Ye page niche karo", "⬇ Move page down"), lambda: self._pv_move(1))
        m.addSeparator()
        m.addAction(self.L("📤 Ye page alag save (JPG/PDF)…", "📤 Save this page as (JPG/PDF)…"),
                    self._pv_save_as)
        m.addAction(self.L("📋 Is page ki image copy (paste ke liye)",
                           "📋 Copy this page's image (to paste)"), self._pv_copy_image)
        m.addAction(self.L("🖨 Sirf ye page print karo", "🖨 Print only this page"),
                    self._print_this_page)
        m.addAction(self.L("🟢 Ye page WhatsApp/Email (PDF banakar)", "🟢 Share this page (as PDF)"),
                    self._share_this_page)
        m.addSeparator()
        m.addAction(self.L("📭 Ye page khaali hai kya?", "📭 Is this page blank?"),
                    self._check_blank_page)
        m.exec_(QtGui.QCursor.pos())

    def _pv_open_image_editor(self, tool=None):
        item = self._current_item_or_warn()
        if not item:
            return
        path = item.data(QtCore.Qt.UserRole)
        self._pv_backup([path])

        def _saved():
            try:
                self._refresh_item(self.list.currentItem())
                self._pv_build_filmstrip()
                self._update_preview_panel()
                self._dirty = True
            except Exception:
                pass
        ImageEditor(self, path, on_saved=_saved, tool=tool).exec_()

    def crop_current_page(self):
        """Crop button — editor 'Crop' tool ke saath khol do: maus se area
        select karke crop karo (auto-crop bhi editor me maujood hai)."""
        self._pv_open_image_editor(tool="crop")

    def _pv_move(self, direction):
        """Current page ko upar/niche (list me) le jao — crop bigade bina."""
        row = self.list.currentRow()
        if row < 0:
            return
        new = row + direction
        if new < 0 or new >= self.list.count():
            return
        it = self.list.takeItem(row)
        self.list.insertItem(new, it)
        self.list.setCurrentItem(it)
        self._dirty = True
        try:
            self._pv_build_filmstrip()
            self._update_preview_panel()
        except Exception:
            pass

    def _pv_save_as(self):
        item = self._current_item_or_warn()
        if not item:
            return
        src = item.data(QtCore.Qt.UserRole)
        out, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, self.L("Ye page kaha save?", "Save this page as"),
            os.path.join(self._opts.get("save_folder", os.path.expanduser("~")), "page.pdf"),
            "PDF (*.pdf);;JPEG (*.jpg);;PNG (*.png)")
        if not out:
            return
        try:
            ext = os.path.splitext(out)[1].lower()
            if ext == ".pdf":
                self._pages_as_pdf([src], out)
            else:
                with Image.open(src) as im:
                    im.convert("RGB").save(out, quality=92)
            self.status.showMessage(self.L("📤 Save ho gaya: ", "📤 Saved: ")
                                    + os.path.basename(out), 6000)
        except Exception as e:
            self._warn(str(e))

    def _pv_copy_image(self):
        item = self._current_item_or_warn()
        if not item:
            return
        try:
            pm = QtGui.QPixmap(item.data(QtCore.Qt.UserRole))
            if not pm.isNull():
                QtWidgets.QApplication.clipboard().setPixmap(pm)
                self.status.showMessage(self.L("📋 Image copy ho gayi — kahin bhi paste karo",
                                               "📋 Image copied — paste it anywhere"), 5000)
        except Exception as e:
            self._warn(str(e))

    def _pv_compare(self):
        """Current + agla page saath-saath dikhao (pehle-baad milaan)."""
        row = self.list.currentRow()
        if row < 0 or self.list.count() < 2:
            self._warn(self.L("Kam se kam 2 page chahiye.", "Need at least 2 pages."))
            return
        other = row + 1 if row + 1 < self.list.count() else row - 1
        p1 = self.list.item(row).data(QtCore.Qt.UserRole)
        p2 = self.list.item(other).data(QtCore.Qt.UserRole)
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle(self.L("↕ Do page saath-saath", "↕ Compare pages"))
        dlg.resize(1000, 720)
        h = QtWidgets.QHBoxLayout(dlg)
        for pth, cap in ((p1, "Page %d" % (row + 1)), (p2, "Page %d" % (other + 1))):
            col = QtWidgets.QVBoxLayout()
            lb = QtWidgets.QLabel(cap); lb.setAlignment(QtCore.Qt.AlignCenter)
            lb.setStyleSheet("font-weight:700;")
            img = QtWidgets.QLabel(); img.setAlignment(QtCore.Qt.AlignCenter)
            pm = QtGui.QPixmap(pth)
            if not pm.isNull():
                img.setPixmap(pm.scaled(460, 640, QtCore.Qt.KeepAspectRatio,
                                        QtCore.Qt.SmoothTransformation))
            sc = QtWidgets.QScrollArea(); sc.setWidgetResizable(True); sc.setWidget(img)
            col.addWidget(lb); col.addWidget(sc, 1)
            h.addLayout(col, 1)
        dlg.exec_()

    def _pv_toggle_loupe(self):
        self._pv_loupe_on = not getattr(self, "_pv_loupe_on", False)
        if self._pv_loupe_on:
            self.pv_img.setMouseTracking(True)
            if not getattr(self, "_pv_loupe_installed", False):
                self.pv_img.installEventFilter(self)
                self._pv_loupe_installed = True
            if not hasattr(self, "_pv_loupe"):
                self._pv_loupe = QtWidgets.QLabel(self)
                self._pv_loupe.setWindowFlags(QtCore.Qt.ToolTip)
                self._pv_loupe.setStyleSheet("border:2px solid #0f766e;background:#fff;")
            self.status.showMessage(self.L("🔎 Loupe ON — image par maus le jao",
                                           "🔎 Loupe ON — hover over the image"), 4000)
        else:
            if hasattr(self, "_pv_loupe"):
                self._pv_loupe.hide()
            self.status.showMessage(self.L("🔎 Loupe OFF", "🔎 Loupe OFF"), 3000)

    def _pv_slideshow(self):
        if getattr(self, "_pv_slide_timer", None) and self._pv_slide_timer.isActive():
            self._pv_slide_timer.stop()
            self.status.showMessage(self.L("▶ Slideshow band", "▶ Slideshow stopped"), 3000)
            return
        self._pv_slide_timer = QtCore.QTimer(self)
        self._pv_slide_timer.setInterval(2000)

        def _tick():
            if self.list.count() < 2:
                self._pv_slide_timer.stop(); return
            r = (self.list.currentRow() + 1) % self.list.count()
            self.list.setCurrentRow(r)
        self._pv_slide_timer.timeout.connect(_tick)
        self._pv_slide_timer.start()
        self.status.showMessage(self.L("▶ Slideshow chalu (dobara dabao = band)",
                                       "▶ Slideshow running (tap again to stop)"), 4000)

    def _print_this_page(self):
        item = self._current_item_or_warn()
        if item:
            self._do_print([item.data(QtCore.Qt.UserRole)], per_page=1)

    def _share_this_page(self):
        item = self._current_item_or_warn()
        if not item:
            return
        try:
            out = os.path.join(self._tmpdir, "page_share.pdf")
            self._pages_as_pdf([item.data(QtCore.Qt.UserRole)], out)
        except Exception as exc:
            self._warn("Failed: %s" % exc); return
        self.share_whatsapp(out)

    def _check_blank_page(self):
        item = self._current_item_or_warn()
        if not item:
            return
        path = item.data(QtCore.Qt.UserRole)

        def job():
            try:
                with Image.open(path) as im:
                    g = im.convert("L")
                    hist = g.histogram()
                    total = sum(hist) or 1
                    dark = sum(hist[:200])          # kitne pixel likhai-jaise (dark)
                    return dark / float(total)
            except Exception as e:
                return e

        def done(res):
            if isinstance(res, Exception):
                return
            if res < 0.012:
                if QtWidgets.QMessageBox.question(
                        self, self.L("Khaali page", "Blank page"),
                        self.L("Ye page lagbhag khaali lagta hai. Hata dein?",
                               "This page looks almost blank. Delete it?"),
                        QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                        QtWidgets.QMessageBox.No) == QtWidgets.QMessageBox.Yes:
                    self.delete_page()
            else:
                self.status.showMessage(self.L("Ye page khaali nahi hai", "This page is not blank"), 3000)
        self._run_bg(job, done, self.L("Jaanch rahe hain…", "Checking…"))

    def _rebuild_jobs_bar(self):
        while self._jobs_lay.count():
            it = self._jobs_lay.takeAt(0)
            w = it.widget()
            if w:
                w.deleteLater()
        lbl = QtWidgets.QLabel(self.L("JOBS:", "JOBS:"))
        lbl.setStyleSheet("color:#92400e;font-weight:700;")
        self._jobs_lay.addWidget(lbl)
        for job in (self._opts.get("jobs") or []):
            b = QtWidgets.QPushButton(job.get("icon", "📋") + " " + job.get("name", "Job"))
            b.setStyleSheet("padding:3px 10px;border:1px solid #fbbf24;border-radius:99px;"
                            "background:#fff;")
            b.clicked.connect(lambda _c=False, j=job: self._apply_job(j))
            self._jobs_lay.addWidget(b)
        add = QtWidgets.QPushButton("➕")
        add.setToolTip(self.L("Naya job banao", "Create a job"))
        add.setStyleSheet("padding:3px 8px;border:1px dashed #fbbf24;border-radius:99px;background:#fff;")
        add.clicked.connect(self._new_job)
        self._jobs_lay.addWidget(add)
        self._jobs_lay.addStretch(1)

    def _apply_job(self, job):
        """Job-chip: ek click me profile + folder + naming set."""
        prof = job.get("profile")
        if prof:
            i = self.cmb_profile.findText(prof)
            if i >= 0:
                self.cmb_profile.setCurrentIndex(i)
        if job.get("folder"):
            self._opts["save_folder"] = job["folder"]
        if job.get("template"):
            self._opts["filename_template"] = job["template"]
        self._opts["auto_name"] = bool(job.get("auto_name", True))
        self._save_opts()
        self._refresh_files_root()
        self.status.showMessage(
            self.L("Job '%s' set ho gaya — ab scan karo." % job.get("name"),
                   "Job '%s' applied — scan now." % job.get("name")), 5000)

    def _new_job(self):
        name, ok = QtWidgets.QInputDialog.getText(
            self, self.L("Naya job", "New job"),
            self.L("Job ka naam (jaise Claim, ID-Cards):", "Job name (e.g. Claim, IDs):"))
        if not ok or not name.strip():
            return
        folder = QtWidgets.QFileDialog.getExistingDirectory(
            self, self.L("Is job ki files kis folder me jaayen?", "Save this job's files to which folder?"),
            self._opts.get("save_folder", ""))
        icon, _ = QtWidgets.QInputDialog.getText(
            self, "Icon", self.L("Ek emoji (optional):", "One emoji (optional):"), text="📋")
        prof = self.cmb_profile.currentText()
        jobs = self._opts.setdefault("jobs", [])
        jobs.append({"name": name.strip(), "icon": (icon.strip() or "📋")[:2],
                     "profile": prof, "folder": folder or "",
                     "template": "{name}_{date}", "auto_name": True})
        self._save_opts()
        self._rebuild_jobs_bar()

    def toggle_kiosk(self):
        on = not self.kiosk.isVisible()
        if on:
            self.kiosk.setGeometry(self.centralWidget().rect())
            self.kiosk.show(); self.kiosk.raise_()
        else:
            self.kiosk.hide()
        self._opts["ui_kiosk"] = on
        self._save_opts()

    # Accent (primary button) rang — customize me chuno.
    ACCENTS = {
        "teal": "#0f766e", "blue": "#1d4ed8", "green": "#15803d",
        "purple": "#7e22ce", "orange": "#c2410c", "rose": "#be123c",
    }

    def _apply_ui_live(self):
        """SAARE UI-customize options ko chalti app par turant laga do.
        Yahi ek jagah hai jahan har UI-setting apply hoti hai — isliye
        live-preview (tick karte hi dikhe), Cancel (wapas pehle jaisa) aur
        OK (save) sab isi ko call karte hai. Har hissa apne try/except me
        hai taaki ek me error aaye to baaki UI theek lage."""
        o = self._opts
        # 1) Theme
        try:
            self._apply_style()
        except Exception:
            pass
        # 2) Accent — primary button ka rang (theme ke upar override)
        try:
            acc = self.ACCENTS.get(o.get("accent", "teal"), "#0f766e")
            self.setStyleSheet(self.styleSheet() +
                "QPushButton#primary{background:%s;border:1px solid %s;color:#fff;"
                "font-weight:700;}QPushButton#primary:hover{background:%s;}"
                % (acc, acc, acc))
        except Exception:
            pass
        # 3) Font scale — poori app ka font (chhota/bada)
        try:
            app = QtWidgets.QApplication.instance()
            if app is not None:
                sc = max(70, min(180, int(o.get("font_scale", 100) or 100)))
                f = app.font(); f.setPointSizeF(9.0 * sc / 100.0); app.setFont(f)
        except Exception:
            pass
        # 4) Ribbon vs classic toolbar
        try:
            rib = bool(o.get("ui_ribbon", False))
            self.ribbon.setVisible(rib)
            self._classic_toolbar.setVisible(not rib)
        except Exception:
            pass
        # 5) Toolbar button style (sirf icon / sirf text / dono) — sirf ASLI
        #    toolbar/ribbon buttons par (baaki ⬅ 📄 jaise glyph-buttons chhedo mat)
        try:
            st = o.get("toolbar_style", "icon_text")
            mode = {"icon_only": QtCore.Qt.ToolButtonIconOnly,
                    "text_only": QtCore.Qt.ToolButtonTextOnly,
                    "icon_text": QtCore.Qt.ToolButtonTextUnderIcon}.get(
                        st, QtCore.Qt.ToolButtonTextUnderIcon)
            _tbtns = []
            for _cont in (getattr(self, "_classic_toolbar", None), getattr(self, "ribbon", None)):
                if _cont is not None:
                    _tbtns += _cont.findChildren(QtWidgets.QToolButton)
            for tb in _tbtns:
                tb.setToolButtonStyle(mode)
        except Exception:
            pass
        # 6) Panel / header / dashboard / graph / analytics dikhao-chhupao
        for attr, key in (("ui_header", "ui_header"), ("side_graph", "ui_graph"),
                          ("preview_panel", "ui_preview"), ("jobs_bar", "ui_jobs"),
                          ("an_box", "ui_analytics"), ("left_panel", "show_left_panel"),
                          ("files_panel", "show_files_panel")):
            try:
                getattr(self, attr).setVisible(bool(o.get(key, True)))
            except Exception:
                pass
        # 7) Panel widths
        for attr, key, dflt in (("left_panel", "left_panel_w", 252),
                                ("files_panel", "files_panel_w", 250),
                                ("preview_panel", "preview_panel_w", 310)):
            try:
                getattr(self, attr).setFixedWidth(
                    max(160, min(520, int(o.get(key, dflt) or dflt))))
            except Exception:
                pass
        # 8) Thumbnail size + page-grid spacing
        try:
            self._apply_thumb_zoom(int(o.get("thumb_size", self.THUMB_W) or self.THUMB_W))
        except Exception:
            pass
        try:
            sp = {"compact": 4, "normal": 10, "roomy": 18}.get(o.get("ui_spacing", "normal"), 10)
            self.list.setSpacing(sp)
        except Exception:
            pass
        # 9) Preview filmstrip
        try:
            self.pv_strip.setVisible(bool(o.get("ui_filmstrip", True)))
        except Exception:
            pass
        # 10) Status-bar footer ke hisse
        for attr, key in (("foot_folder", "footer_folder"), ("foot_pages", "footer_pages"),
                          ("foot_last", "footer_last"), ("foot_disk", "footer_disk"),
                          ("foot_ver", "footer_version"), ("lbl_busy", "footer_scanner"),
                          ("foot_clock", "footer_clock"), ("foot_today", "footer_today"),
                          ("foot_online", "footer_online")):
            try:
                getattr(self, attr).setVisible(bool(o.get(key, True) if key in
                        ("footer_folder", "footer_pages", "footer_last", "footer_disk",
                         "footer_version", "footer_scanner") else o.get(key, False)))
            except Exception:
                pass
        try:
            self.foot_msg.setVisible(bool(str(o.get("footer_msg", "")).strip()))
        except Exception:
            pass
        # 11) Window paardarshita (transparency)
        try:
            op = max(60, min(100, int(o.get("window_opacity", 100) or 100)))
            self.setWindowOpacity(op / 100.0)
        except Exception:
            pass
        # 12) Rounded/sharp corners + high-contrast (theme ke upar)
        try:
            extra = ""
            if o.get("ui_corners") == "sharp":
                extra += ("QPushButton,QToolButton,QLineEdit,QComboBox,QSpinBox,"
                          "#panel,#statsbox{border-radius:0px;}")
            if o.get("high_contrast"):
                extra += ("QLabel{color:#000;} QMainWindow,QWidget{background:#ffffff;}"
                          "QPushButton,QLineEdit,QComboBox{border:2px solid #000;color:#000;}"
                          "QListWidget::item:selected{background:#000;color:#fff;}")
            if extra:
                self.setStyleSheet(self.styleSheet() + extra)
        except Exception:
            pass
        # 13) 'Meri Files' panel ki side (right/left)
        try:
            self._apply_files_side()
        except Exception:
            pass
        # 14) Branding (hospital naam + logo)
        try:
            self._apply_branding()
        except Exception:
            pass
        # 15) dependent views refresh
        for m in ("_update_preview_panel", "_update_empty_state", "_update_sidebar_stats",
                  "_update_footer"):
            try:
                getattr(self, m)()
            except Exception:
                pass

    def _apply_branding(self):
        """Header me hospital/clinic ka naam + logo lagao (customize se)."""
        o = self._opts
        try:
            name = str(o.get("brand_name", "") or "").strip()
            self.hdr_brand.setText(name)
            self.hdr_brand.setVisible(bool(name))
        except Exception:
            pass
        try:
            logo = str(o.get("brand_logo", "") or "").strip()
            if logo and os.path.exists(logo):
                pm = QtGui.QPixmap(logo)
                if not pm.isNull():
                    self.hdr_logo.setPixmap(pm.scaledToHeight(
                        22, QtCore.Qt.SmoothTransformation))
                    self.hdr_logo.setVisible(True)
                else:
                    self.hdr_logo.clear(); self.hdr_logo.setVisible(False)
            else:
                self.hdr_logo.clear(); self.hdr_logo.setVisible(False)
        except Exception:
            pass
        # window title me bhi naam
        try:
            base = "%s v%s" % (APP_NAME, VERSION)
            nm = str(o.get("brand_name", "") or "").strip()
            self.setWindowTitle(("%s — %s" % (nm, base)) if nm else base)
        except Exception:
            pass

    def _apply_files_side(self):
        """'Meri Files' panel ko dayein ya bayein rakho (customize se)."""
        body = getattr(self, "_body_layout", None)
        if body is None:
            return
        try:
            want_left = (self._opts.get("files_panel_side") == "left")
            idx = body.indexOf(self.files_panel)
            if idx < 0:
                return
            is_left = (idx == 0)
            if want_left == is_left:
                return   # pehle se sahi jagah — kuch mat karo
            body.removeWidget(self.files_panel)
            if want_left:
                body.insertWidget(0, self.files_panel)
            else:
                body.addWidget(self.files_panel)
        except Exception:
            pass

    def _on_thumb_dblclick(self, item):
        """Kisi bhi scan kiye page par double-click = poora Document Editor khule.
        (Customize me dbl_action='preview' ho to sirf preview panel dikhe.)"""
        try:
            if self._opts.get("dbl_action", "edit") == "preview":
                if not self.preview_panel.isVisible():
                    self._opts["ui_preview"] = True
                    self.preview_panel.setVisible(True)
                self.list.setCurrentItem(item)
                try:
                    self._update_preview_panel()
                except Exception:
                    pass
                return
        except Exception:
            pass
        # 'edit' → naya Document Editor (crop/erase/pen/sliders/page-nav…)
        if item is not None:
            self.list.setCurrentItem(item)
        self._pv_open_image_editor()

    # ---- Presets (ek-click poora look) ----
    UI_PRESETS = {
        "reception": {"theme": "light", "accent": "teal", "font_scale": 115,
                      "thumb_size": 175, "ui_spacing": "roomy", "ui_dashboard": True,
                      "ui_header": True, "touch_mode": True, "toolbar_style": "icon_text"},
        "doctor": {"theme": "dark", "accent": "blue", "font_scale": 100,
                   "ui_preview": True, "preview_panel_w": 360, "ui_filmstrip": True,
                   "ui_dashboard": False, "thumb_size": 150, "ui_spacing": "normal"},
        "night": {"theme": "darkpro", "accent": "purple", "font_scale": 100,
                  "window_opacity": 96, "high_contrast": False, "ui_spacing": "normal"},
    }

    def _apply_preset(self, key, dlg=None):
        p = self.UI_PRESETS.get(key)
        if not p:
            return
        self._opts.update(p)
        self._apply_ui_live()
        self._save_opts()
        self.status.showMessage(
            self.L("Preset laga diya — pasand aaye to rehne do, warna Cancel/Reset.",
                   "Preset applied — keep it, or Cancel/Reset."), 4000)
        if dlg is not None:
            dlg.accept()

    def _reset_ui_defaults(self, dlg=None):
        r = QtWidgets.QMessageBox.question(
            self, APP_NAME,
            self.L("Saari UI settings wapas shuruaati (default) par le aayein?",
                   "Reset all UI settings back to default?"),
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No, QtWidgets.QMessageBox.No)
        if r != QtWidgets.QMessageBox.Yes:
            return
        # sirf UI-wale keys reset karo (baaki scanner/profile settings chhedo mat)
        for k in self._UI_KEYS:
            if k in DEFAULT_OPTIONS:
                self._opts[k] = DEFAULT_OPTIONS[k]
        self._lang = self._opts.get("language", "en")
        self._apply_ui_live()
        self._save_opts()
        if dlg is not None:
            dlg.accept()

    def _export_ui_settings(self):
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, self.L("Settings save karo", "Save settings"),
            os.path.join(os.path.expanduser("~"), "ApneScan_UI.json"),
            "JSON (*.json)")
        if not path:
            return
        try:
            data = {k: self._opts.get(k) for k in self._UI_KEYS if k in self._opts}
            with open(path, "w", encoding="utf-8") as fh:
                json.dump(data, fh, ensure_ascii=False, indent=2)
            self.status.showMessage(self.L("Settings save ho gayi.", "Settings saved."), 4000)
        except Exception as e:
            self._warn(str(e))

    def _import_ui_settings(self, dlg=None):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, self.L("Settings file chuno", "Choose settings file"),
            os.path.expanduser("~"), "JSON (*.json)")
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            for k in self._UI_KEYS:
                if k in data:
                    self._opts[k] = data[k]
            self._lang = self._opts.get("language", "en")
            self._apply_ui_live()
            self._save_opts()
            self.status.showMessage(self.L("Settings aayi — laga di.", "Settings imported."), 4000)
            if dlg is not None:
                dlg.accept()
        except Exception as e:
            self._warn(str(e))

    # customize me shaamil UI-keys (reset/export/import in par kaam karte hain)
    _UI_KEYS = (
        "theme", "accent", "font_scale", "touch_mode", "ui_animations", "ui_tooltips",
        "start_maximized", "ui_ribbon", "toolbar_style", "ui_header", "ui_dashboard",
        "ui_jobs", "thumb_size", "ui_spacing", "show_page_numbers", "show_left_panel",
        "left_panel_w", "show_files_panel", "files_panel_w", "ui_preview",
        "preview_panel_w", "ui_filmstrip", "ui_graph", "ui_analytics", "footer_folder",
        "footer_pages", "footer_last", "footer_disk", "footer_version", "footer_scanner",
        "after_save", "ui_confirm_delete", "language", "ui_corners", "high_contrast",
        "window_opacity", "brand_name", "brand_logo", "remember_window",
        "auto_update_check", "dbl_action", "files_panel_side", "sound_on_done",
        "confirm_exit", "footer_clock", "footer_today", "footer_online", "footer_msg",
    )

    def customize_ui(self):
        """UI customize — 40+ settings tabs me, LIVE-PREVIEW ke saath (v73).
        Koi bhi option badalte hi bina OK kiye software par turant dikhta hai;
        OK = save, Cancel = sab wapas pehle jaisa."""
        L = self.L
        # tick karne se pehle ki poori state — Cancel par yahi wapas aayegi
        snapshot = dict(self._opts)

        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle(L("🎨 UI customize (live)", "🎨 Customize UI (live)"))
        dlg.setMinimumWidth(440)
        outer = QtWidgets.QVBoxLayout(dlg)
        hint = QtWidgets.QLabel(L(
            "<span style='color:#0f766e;font-size:12px;'>💡 Koi bhi cheez tick/badalte "
            "hi software par TURANT dikh jayegi. Pasand aaye to <b>OK</b>, warna "
            "<b>Cancel</b> — sab wapas pehle jaisa.</span>",
            "<span style='color:#0f766e;font-size:12px;'>💡 Every change shows on the app "
            "<b>instantly</b> (no OK needed). Keep it with <b>OK</b>, or "
            "<b>Cancel</b> to revert everything.</span>"))
        hint.setTextFormat(QtCore.Qt.RichText); hint.setWordWrap(True)
        outer.addWidget(hint)
        tabs = QtWidgets.QTabWidget()
        outer.addWidget(tabs, 1)

        def _set(key, val):
            self._opts[key] = val
            self._apply_ui_live()

        def _tab(title):
            w = QtWidgets.QWidget(); f = QtWidgets.QFormLayout(w)
            f.setLabelAlignment(QtCore.Qt.AlignLeft)
            tabs.addTab(w, title)
            return f

        def _check(f, key, hi, en, default=True):
            c = QtWidgets.QCheckBox(L(hi, en))
            c.setChecked(bool(self._opts.get(key, default)))
            c.toggled.connect(lambda v, k=key: _set(k, bool(v)))
            f.addRow(c)
            return c

        def _combo(f, label, key, items, default):
            # items: list of (value, shown-text)
            cmb = QtWidgets.QComboBox()
            for _v, _t in items:
                cmb.addItem(_t, _v)
            cur = self._opts.get(key, default)
            idx = max(0, [v for v, _ in items].index(cur) if cur in [v for v, _ in items] else 0)
            cmb.setCurrentIndex(idx)
            cmb.currentIndexChanged.connect(
                lambda i, k=key, c=cmb: _set(k, c.itemData(i)))
            f.addRow(label, cmb)
            return cmb

        def _slider(f, label, key, lo, hi, default, suffix=""):
            row = QtWidgets.QWidget(); h = QtWidgets.QHBoxLayout(row)
            h.setContentsMargins(0, 0, 0, 0)
            s = QtWidgets.QSlider(QtCore.Qt.Horizontal)
            s.setMinimum(lo); s.setMaximum(hi)
            s.setValue(int(self._opts.get(key, default) or default))
            val = QtWidgets.QLabel(str(s.value()) + suffix)
            val.setMinimumWidth(46)
            s.valueChanged.connect(lambda v, k=key, lb=val: (lb.setText(str(v) + suffix), _set(k, v)))
            h.addWidget(s, 1); h.addWidget(val)
            f.addRow(label, row)
            return s

        def _line(f, label, key, placeholder=""):
            e = QtWidgets.QLineEdit(str(self._opts.get(key, "") or ""))
            e.setPlaceholderText(placeholder)
            e.textChanged.connect(lambda t, k=key: _set(k, t))
            f.addRow(label, e)
            return e

        # ---------- TAB: Look ----------
        f = _tab(L("🎨 Roop", "🎨 Look"))
        _combo(f, L("Theme:", "Theme:"), "theme",
               [("light", "Light"), ("dark", "Dark"), ("darkpro", "Dark Pro")], "light")
        _combo(f, L("Rang (accent):", "Accent colour:"), "accent",
               [("teal", "🟢 Teal"), ("blue", "🔵 Blue"), ("green", "🟩 Green"),
                ("purple", "🟣 Purple"), ("orange", "🟧 Orange"), ("rose", "🌹 Rose")], "teal")
        _slider(f, L("Font size:", "Font size:"), "font_scale", 70, 180, 100, "%")
        _combo(f, L("Kone (corners):", "Corners:"), "ui_corners",
               [("rounded", L("Gol (rounded)", "Rounded")),
                ("sharp", L("Seedhe (sharp)", "Sharp"))], "rounded")
        _slider(f, L("Window paardarshita:", "Window opacity:"),
                "window_opacity", 60, 100, 100, "%")
        _check(f, "high_contrast", 'High contrast (bold clear colours, for low vision)',
               "High-contrast (bold clear colours, low vision)", False)
        _check(f, "touch_mode", "Touch / buzurg mode (bada sab kuch)",
               "Touch / large mode (bigger everything)", False)
        _check(f, "ui_animations", "Halke animations", "Smooth animations", True)
        _check(f, "ui_tooltips", "Hover par madad-tips", "Hover help-tips", True)
        _check(f, "start_maximized", "App khulte hi poori screen",
               "Open maximised (full screen)", False)

        # ---------- TAB: Toolbar / Header ----------
        f = _tab(L("🧰 Toolbar", "🧰 Toolbar"))
        _check(f, "ui_ribbon", "Ribbon toolbar (Office jaisa — tabs me)",
               "Ribbon toolbar (Office-style tabs)", False)
        _combo(f, L("Button style:", "Button style:"), "toolbar_style",
               [("icon_text", L("Icon + naam", "Icon + text")),
                ("icon_only", L("Sirf icon", "Icon only")),
                ("text_only", L("Sirf naam", "Text only"))], "icon_text")
        _check(f, "ui_header", "Status-patti (scanner/profile/aaj)",
               "Status header (scanner/profile/today)", True)
        _check(f, "ui_dashboard", 'Start dashboard (big buttons on an empty screen)',
               "Start dashboard (big buttons on empty screen)", True)
        _check(f, "ui_jobs", "Job-chips patti (1 click me profile+folder)",
               "Job chips bar (1-click profile+folder)", False)

        # ---------- TAB: Pages ----------
        f = _tab(L("📄 Pages", "📄 Pages"))
        _slider(f, L("Thumbnail size:", "Thumbnail size:"), "thumb_size", 90, 320, 150, "px")
        _combo(f, L("Beech ki jagah:", "Grid spacing:"), "ui_spacing",
               [("compact", L("Compact (paas-paas)", "Compact")),
                ("normal", L("Normal", "Normal")),
                ("roomy", L("Khula (roomy)", "Roomy"))], "normal")
        _check(f, "show_page_numbers", 'Show a name/number under each page',
               "Show name/number under each page", True)
        _combo(f, L("Double-click par:", "Double-click opens:"), "dbl_action",
               [("edit", L("Editor window (rotate/crop)", "Editor window (rotate/crop)")),
                ("preview", L("Sirf preview panel me jhalak", "Just show in preview panel"))],
               "edit")

        # ---------- TAB: Panels ----------
        f = _tab(L("🗂 Panels", "🗂 Panels"))
        _check(f, "show_left_panel", "Baayan scan-settings panel",
               "Left scan-settings panel", True)
        _slider(f, L("  ↳ chaudai:", "  ↳ width:"), "left_panel_w", 180, 420, 252, "px")
        _check(f, "show_files_panel", "Daayan 'Meri Files' panel",
               "Right 'My Files' panel", True)
        _slider(f, L("  ↳ chaudai:", "  ↳ width:"), "files_panel_w", 180, 420, 250, "px")
        _check(f, "ui_preview", "Preview panel (badi jhalak + quick-edit)",
               "Preview panel (big preview + quick edit)", False)
        _slider(f, L("  ↳ chaudai:", "  ↳ width:"), "preview_panel_w", 220, 480, 310, "px")
        _check(f, "ui_filmstrip", "Preview me niche filmstrip",
               "Filmstrip in preview panel", True)
        _check(f, "ui_graph", "Sidebar me 7-din ka graph",
               "7-day graph in the sidebar", True)
        _check(f, "ui_analytics", "Sidebar me analytics card",
               "Analytics card in the sidebar", True)
        _combo(f, L("'Meri Files' panel ki side:", "'My Files' panel side:"),
               "files_panel_side",
               [("right", L("Daayein (right)", "Right")),
                ("left", L("Baayein (left)", "Left"))], "right")

        # ---------- TAB: Status bar ----------
        f = _tab(L("📊 Status", "📊 Status bar"))
        _check(f, "footer_folder", "Folder", "Folder", True)
        _check(f, "footer_pages", "Pages / selection", "Pages / selection", True)
        _check(f, "footer_last", "Aakhri save ki file", "Last saved file", True)
        _check(f, "footer_disk", 'Free disk', "Free disk", True)
        _check(f, "footer_version", "Version", "Version", True)
        _check(f, "footer_scanner", "Scanner / busy", "Scanner / busy", True)
        _check(f, "footer_clock", "Ghadi / tareekh", "Clock / date", False)
        _check(f, "footer_today", "Aaj ke scan", "Today's scans", False)
        _check(f, "footer_online", "Online log (worldwide)", "People online (worldwide)", False)
        _line(f, L("Apna sandesh:", "Custom message:"), "footer_msg",
              L("jaise: OPD 9 se 2", "e.g. OPD 9am-2pm"))

        # ---------- TAB: Behaviour ----------
        f = _tab(L("⚙ Vyavhaar", "⚙ Behaviour"))
        _combo(f, L("Save ke baad:", "After save:"), "after_save",
               [("nothing", L("Kuch nahi", "Do nothing")),
                ("folder", L("Folder kholo", "Open folder")),
                ("open", L("File kholo", "Open file"))], "nothing")
        _check(f, "ui_confirm_delete", 'Ask before deleting',
               "Ask before delete", True)
        _check(f, "confirm_exit", 'Ask when closing',
               "Ask before closing the app", False)
        _check(f, "sound_on_done", "Scan poora hote hi 'ting' awaaz",
               "Beep when a scan finishes", False)
        _check(f, "remember_window", 'Remember the window size/position',
               "Remember window size/position", True)
        _check(f, "auto_update_check", 'Check for updates automatically',
               "Auto-check for updates", True)
        _combo(f, L("Bhasha / Language:", "Language:"), "language",
               [("en", "English"), ("hi", "हिन्दी (Hindi)")], "en")
        bkiosk = QtWidgets.QPushButton(L("🖥 Kiosk mode ab chalu karo (bade buttons)",
                                         "🖥 Enter Kiosk mode now (big buttons)"))
        bkiosk.clicked.connect(lambda: (dlg.accept(), self.toggle_kiosk()))
        f.addRow(bkiosk)
        note = QtWidgets.QLabel(L(
            "<span style='color:#64748b;font-size:11px;'>Bhasha badalne par kuch text "
            "restart ke baad poora badlega.</span>",
            "<span style='color:#64748b;font-size:11px;'>Some text fully switches "
            "language after a restart.</span>"))
        note.setTextFormat(QtCore.Qt.RichText); note.setWordWrap(True)
        f.addRow(note)

        # ---------- TAB: Presets + Branding ----------
        f = _tab(L("🎁 Presets", "🎁 Presets"))
        f.addRow(QtWidgets.QLabel(L(
            "<b>Ek-click look</b> — pasand aaye to rehne do, warna Cancel/Reset.",
            "<b>One-click looks</b> — keep it, or Cancel/Reset.")))
        _prow = QtWidgets.QHBoxLayout()
        for _pk, _pt in (("reception", L("🏥 Reception", "🏥 Reception")),
                         ("doctor", L("🩺 Doctor", "🩺 Doctor")),
                         ("night", L("🌙 Night", "🌙 Night"))):
            _pb = QtWidgets.QPushButton(_pt)
            _pb.clicked.connect(lambda _c, k=_pk: self._apply_preset(k))
            _prow.addWidget(_pb)
        _pw = QtWidgets.QWidget(); _pw.setLayout(_prow); f.addRow(_pw)
        # Branding
        f.addRow(QtWidgets.QLabel(L("<b>Branding</b> (header me dikhega):",
                                    "<b>Branding</b> (shows in the header):")))
        _line(f, L("Hospital/clinic naam:", "Hospital/clinic name:"), "brand_name",
              L("jaise: Noble Hospital", "e.g. Noble Hospital"))
        _logo_row = QtWidgets.QHBoxLayout()
        _logo_lbl = QtWidgets.QLabel(os.path.basename(self._opts.get("brand_logo", "")) or
                                     L("(koi logo nahi)", "(no logo)"))
        _logo_lbl.setStyleSheet("color:#64748b;font-size:11px;")

        def _pick_logo():
            p, _ = QtWidgets.QFileDialog.getOpenFileName(
                self, L("Logo image chuno", "Choose logo image"),
                os.path.expanduser("~"), "Images (*.png *.jpg *.jpeg *.bmp)")
            if p:
                _set("brand_logo", p)
                _logo_lbl.setText(os.path.basename(p))

        def _clr_logo():
            _set("brand_logo", "")
            _logo_lbl.setText(L("(koi logo nahi)", "(no logo)"))
        _blogo = QtWidgets.QPushButton(L("📁 Logo chuno", "📁 Choose logo"))
        _blogo.clicked.connect(_pick_logo)
        _bclr = QtWidgets.QPushButton("✖"); _bclr.setFixedWidth(30); _bclr.clicked.connect(_clr_logo)
        _logo_row.addWidget(_blogo); _logo_row.addWidget(_bclr); _logo_row.addWidget(_logo_lbl, 1)
        _lw = QtWidgets.QWidget(); _lw.setLayout(_logo_row); f.addRow(_lw)
        # Reset / Export / Import
        f.addRow(QtWidgets.QLabel(L("<b>Settings sambhaalo</b>:", "<b>Manage settings</b>:")))
        _mrow = QtWidgets.QHBoxLayout()
        _breset = QtWidgets.QPushButton(L("↺ Reset (default)", "↺ Reset to default"))
        _breset.clicked.connect(lambda: self._reset_ui_defaults(dlg))
        _bexp = QtWidgets.QPushButton(L("⬇ Export", "⬇ Export"))
        _bexp.clicked.connect(self._export_ui_settings)
        _bimp = QtWidgets.QPushButton(L("⬆ Import", "⬆ Import"))
        _bimp.clicked.connect(lambda: self._import_ui_settings(dlg))
        _mrow.addWidget(_breset); _mrow.addWidget(_bexp); _mrow.addWidget(_bimp)
        _mw = QtWidgets.QWidget(); _mw.setLayout(_mrow); f.addRow(_mw)

        # ---------- OK / Cancel ----------
        bb = QtWidgets.QDialogButtonBox(
            QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        bb.button(QtWidgets.QDialogButtonBox.Ok).setText(L("✔ OK (save karo)", "✔ OK (save)"))
        bb.button(QtWidgets.QDialogButtonBox.Cancel).setText(
            L("✖ Cancel (wapas pehle jaisa)", "✖ Cancel (revert)"))
        bb.accepted.connect(dlg.accept); bb.rejected.connect(dlg.reject)
        outer.addWidget(bb)

        # live-preview pehle se chalu — jaisa abhi opts hai waisa laga do
        self._apply_ui_live()
        self._lang = self._opts.get("language", "en")

        if dlg.exec_() == QtWidgets.QDialog.Accepted:
            # OK — jo dikh raha hai wahi save
            self._lang = self._opts.get("language", "en")
            self._save_opts()
        else:
            # Cancel — sab wapas snapshot par
            self._opts.clear(); self._opts.update(snapshot)
            self._lang = self._opts.get("language", "en")
            self._apply_ui_live()

    def toggle_left_panel(self):
        vis = not self.left_panel.isVisible()
        self.left_panel.setVisible(vis)
        self._opts["show_left_panel"] = vis
        self._save_opts()
        self.status.showMessage(
            self.L("Left sidebar %s (F9 se wapas)" % ("ON" if vis else "OFF"),
                   "Left sidebar %s (press F9 to toggle)" % ("ON" if vis else "OFF")), 4000)

    def L(self, hi, en):
        """Chhota helper: language ke hisaab se Hindi/English text."""
        return en if self._lang == "en" else hi

    # ---- Tags + OCR search index ----
    INDEX_PATH = os.path.join(os.path.expanduser("~"), ".apnescan_index.json")

    def _load_index(self):
        try:
            with open(self.INDEX_PATH, "r", encoding="utf-8") as fh:
                return json.load(fh)
        except Exception:
            return {}

    def _save_index(self, idx):
        try:
            with open(self.INDEX_PATH, "w", encoding="utf-8") as fh:
                json.dump(idx, fh)
        except Exception:
            pass

    def build_search_index(self):
        """Saari saved PDFs ka text EK BAAR padh kar index bana lo — uske baad
        'andar ke text' wali search turant hoti hai."""
        if not HAS_OCR_LIBS:
            self._warn("pypdf is not installed."); return
        root = self._opts.get("save_folder", os.path.expanduser("~"))

        def job():
            idx = self._load_index()
            pdfs = []
            for dp, _dn, fn in os.walk(root):
                for f in fn:
                    if f.lower().endswith(".pdf"):
                        pdfs.append(os.path.join(dp, f))
            changed = 0
            for p in pdfs:
                try:
                    m = os.path.getmtime(p)
                except Exception:
                    continue
                e = idx.get(p)
                if e and e.get("m") == m:
                    continue
                text = ""
                try:
                    reader = PdfReader(p)
                    for pg in reader.pages[:10]:
                        text += (pg.extract_text() or "")
                except Exception:
                    pass
                idx[p] = {"m": m, "t": (text or "").lower()[:20000]}
                changed += 1
            for p in list(idx):
                if not os.path.exists(p):
                    idx.pop(p, None)
            self._save_index(idx)
            return (len(pdfs), changed)

        def done(res):
            if isinstance(res, Exception):
                self._warn("Index failed:\n%s" % res); return
            total, changed = res
            QtWidgets.QMessageBox.information(
                self, "Index ready",
                "%d PDFs are indexed (%d new/changed read).\n\n"
                "Now the Ctrl+F search finds 'text inside' instantly.\n"
                "(Only PDFs made with 'OCR searchable' ticked have text inside.)"
                % (total, changed))
        self._run_bg(job, done, "Building search index… (the app keeps running)")

    def tag_pdf(self, src=None):
        """Kisi bhi saved PDF par tags lagao (jaise: Aadhaar, School, Bijli-bill)."""
        if not isinstance(src, str) or not src:
            src = self._pick_pdf("Kis PDF par tag lagana hai?")
        if not src:
            return
        tags = self._opts.setdefault("tags", {})
        cur = ", ".join(tags.get(src, []))
        text, ok = QtWidgets.QInputDialog.getText(
            self, "Tags", "Enter tags (comma-separated, e.g.: Aadhaar, Home):", text=cur)
        if not ok:
            return
        lst = [t.strip() for t in text.split(",") if t.strip()]
        if lst:
            tags[src] = lst
        else:
            tags.pop(src, None)
        self._save_opts()
        self.status.showMessage("Tags saved: %s" % (", ".join(lst) or "(removed)"), 5000)

    def search_by_tag(self):
        tags = self._opts.get("tags", {}) or {}
        all_tags = sorted({t for lst in tags.values() for t in lst})
        if not all_tags:
            self._warn("No PDF has any tags yet.\n(Add them via Tools → Add tag….)")
            return
        pick, ok = QtWidgets.QInputDialog.getItem(
            self, "Find by tag", "Choose a tag:", all_tags, 0, False)
        if not ok or not pick:
            return
        matches = [p for p, lst in tags.items() if pick in lst and os.path.exists(p)]
        if not matches:
            self._warn("No files exist for this tag anymore."); return
        item, ok = QtWidgets.QInputDialog.getItem(
            self, "Files tagged '%s'" % pick, "Found %d file(s) — choose one to open:" % len(matches),
            matches, 0, False)
        if ok and item:
            self._open_path(item)

    # ---- PDF tools (naye) ----
    def _pick_pdf(self, title):
        start = self._opts.get("save_folder", os.path.expanduser("~"))
        f, _ = QtWidgets.QFileDialog.getOpenFileName(self, title, start, "PDF (*.pdf)")
        return f or None

    def _run_bg(self, fn, on_done, busy_msg):
        """Bhaari kaam BACKGROUND me chalao; UI 1 second ke liye bhi nahi rukti.
        done hone par on_done(result) (main thread me). Ek non-blocking
        '⏳ kaam ho raha hai' indicator dikhta hai; kai kaam ek saath bhi
        chal sakte hain (counter se)."""
        self._bg_workers = getattr(self, "_bg_workers", [])
        self._bg_count = getattr(self, "_bg_count", 0) + 1
        self._show_busy_indicator(busy_msg)
        w = FuncWorker(fn)

        def _fin(res, w=w):
            self._bg_count = max(0, getattr(self, "_bg_count", 1) - 1)
            self._show_busy_indicator(None)
            try:
                self._bg_workers.remove(w)
            except ValueError:
                pass
            on_done(res)
        w.done.connect(_fin)
        self._bg_workers.append(w)
        w.start()

    def _show_busy_indicator(self, msg):
        """Status bar me ghoomta hua '⏳' — kaam chal raha hai par app chalu."""
        if not hasattr(self, "_busy_lbl"):
            self._busy_lbl = QtWidgets.QLabel("")
            self._busy_lbl.setStyleSheet(
                "QLabel{background:#0f766e;color:#fff;border-radius:9px;"
                "padding:1px 10px;font-weight:700;}")
            try:
                self.status.addPermanentWidget(self._busy_lbl)
            except Exception:
                pass
            self._busy_spin = ["⏳", "⌛"]
            self._busy_i = 0
            self._busy_timer = QtCore.QTimer(self)
            self._busy_timer.setInterval(500)

            def _tick():
                self._busy_i = (self._busy_i + 1) % len(self._busy_spin)
                base = getattr(self, "_busy_msg", 'Working…')
                n = getattr(self, "_bg_count", 0)
                extra = (" (%d)" % n) if n > 1 else ""
                self._busy_lbl.setText("%s %s%s" %
                                       (self._busy_spin[self._busy_i], base, extra))
            self._busy_timer.timeout.connect(_tick)
        if msg:
            self._busy_msg = msg
        if getattr(self, "_bg_count", 0) > 0:
            self._busy_lbl.setText("⏳ " + getattr(self, "_busy_msg", "Working…"))
            self._busy_lbl.show()
            self._busy_timer.start()
        else:
            self._busy_timer.stop()
            self._busy_lbl.hide()

    def pdf_page_editor(self):
        """Kisi bhi PDF ke pages ka kram badlo / ghumao / hatao — bina quality
        kharaab kiye (lossless, pypdf se)."""
        if not HAS_OCR_LIBS:
            self._warn("pypdf is not installed."); return
        src = self._pick_pdf('Which PDF to edit?')
        if not src:
            return
        try:
            reader = PdfReader(src)
            if reader.is_encrypted:
                pw, ok = QtWidgets.QInputDialog.getText(
                    self, "Password", "Password for this PDF:", QtWidgets.QLineEdit.Password)
                if not ok or not reader.decrypt(pw or ""):
                    self._warn("Wrong password."); return
            n = len(reader.pages)
        except Exception as exc:
            self._warn("Could not open PDF:\n%s" % exc); return
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("PDF page editor — %s (%d pages)" % (os.path.basename(src), n))
        dlg.resize(420, 460)
        v = QtWidgets.QVBoxLayout(dlg)
        lw = QtWidgets.QListWidget()
        for i in range(n):
            it = QtWidgets.QListWidgetItem("Page %d" % (i + 1))
            it.setData(QtCore.Qt.UserRole, [i, 0])   # [original index, extra rotation]
            lw.addItem(it)
        v.addWidget(lw, 1)
        row = QtWidgets.QHBoxLayout()

        def _move(d):
            r = lw.currentRow()
            if r < 0 or not (0 <= r + d < lw.count()):
                return
            it = lw.takeItem(r); lw.insertItem(r + d, it); lw.setCurrentItem(it)

        def _rot():
            it = lw.currentItem()
            if it:
                data = it.data(QtCore.Qt.UserRole)
                data[1] = (data[1] + 90) % 360
                it.setData(QtCore.Qt.UserRole, data)
                it.setText("Page %d  (rotated: %d°)" % (data[0] + 1, data[1]))

        def _del():
            r = lw.currentRow()
            if r >= 0:
                lw.takeItem(r)
        for t, s in [('⬆ Up', lambda: _move(-1)), ('⬇ Down', lambda: _move(1)),
                     ('↻ Rotate', _rot), ('🗑 Delete', _del)]:
            b = QtWidgets.QPushButton(t); b.clicked.connect(s); row.addWidget(b)
        v.addLayout(row)
        bb = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Save | QtWidgets.QDialogButtonBox.Cancel)
        bb.accepted.connect(dlg.accept); bb.rejected.connect(dlg.reject)
        v.addWidget(bb)
        if dlg.exec_() != QtWidgets.QDialog.Accepted or lw.count() == 0:
            return
        out, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, 'Save the new PDF', src[:-4] + "_edit.pdf", "PDF (*.pdf)")
        if not out:
            return
        try:
            writer = PdfWriter()
            for i in range(lw.count()):
                idx, rot = lw.item(i).data(QtCore.Qt.UserRole)
                pg = reader.pages[idx]
                if rot:
                    pg.rotate(rot)
                writer.add_page(pg)
            with open(out, "wb") as fh:
                writer.write(fh)
        except Exception:
            self._warn("Save failed:\n%s" % traceback.format_exc()); return
        self._remember_save_dir(out)
        QtWidgets.QMessageBox.information(self, "Done", "PDF created:\n%s" % out)

    def place_sign(self):
        """Apne sign/mohar ki image current page par lagao (white background
        apne aap transparent ho jata hai)."""
        item = self._current_item_or_warn()
        if not item:
            return
        sp = self._opts.get("sign_image") or ""
        if not sp or not os.path.exists(sp):
            f, _ = QtWidgets.QFileDialog.getOpenFileName(
                self, 'Choose your signature/stamp image (with a white background)',
                "", "Images (*.png *.jpg *.jpeg *.bmp)")
            if not f:
                return
            self._opts["sign_image"] = f
            self._save_opts()
            sp = f
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Place Sign/Stamp")
        form = QtWidgets.QFormLayout(dlg)
        POS = ['Bottom-right', 'Bottom-centre', 'Bottom-left',
               'Top-right', 'Top-centre', 'Top-left', "Beech me"]
        cmb = QtWidgets.QComboBox(); cmb.addItems(POS)
        spn = QtWidgets.QSpinBox(); spn.setRange(8, 60); spn.setValue(22); spn.setSuffix(" % chaudai")
        btn = QtWidgets.QPushButton("Change sign image…")

        def _chg():
            f, _ = QtWidgets.QFileDialog.getOpenFileName(
                self, "Sign/stamp image", "", "Images (*.png *.jpg *.jpeg *.bmp)")
            if f:
                self._opts["sign_image"] = f; self._save_opts()
        btn.clicked.connect(_chg)
        form.addRow("Jagah:", cmb)
        form.addRow("Size:", spn)
        form.addRow(btn)
        bb = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        bb.accepted.connect(dlg.accept); bb.rejected.connect(dlg.reject)
        form.addRow(bb)
        if dlg.exec_() != QtWidgets.QDialog.Accepted:
            return
        path = item.data(QtCore.Qt.UserRole)
        try:
            with Image.open(path) as im:
                page = im.convert("RGB").copy()
            with Image.open(self._opts["sign_image"]) as s:
                sign = s.convert("RGBA")
            g = sign.convert("L")
            sign.putalpha(g.point(lambda v: 0 if v > 230 else 255))
            w = max(30, int(page.width * spn.value() / 100.0))
            h = max(1, int(sign.height * w / sign.width))
            sign = sign.resize((w, h), Image.LANCZOS)
            m = max(20, page.width // 40)
            x_r, x_c, x_l = page.width - w - m, (page.width - w) // 2, m
            y_b, y_c, y_t = page.height - h - m, (page.height - h) // 2, m
            pos = [(x_r, y_b), (x_c, y_b), (x_l, y_b),
                   (x_r, y_t), (x_c, y_t), (x_l, y_t), (x_c, y_c)][cmb.currentIndex()]
            page.paste(sign, pos, sign)
            save_image_keep_ext(page, path)
            self._refresh_item(item)
            self._dirty = True
        except Exception:
            self._warn("Could not place sign:\n%s" % traceback.format_exc())

    def add_page_numbers(self):
        """Sab pages par 'Page X / N' (aur chaaho to upar apna text) chhapo."""
        paths = self._ordered_paths()
        if not paths:
            self._warn(tr("scan_first", self._lang)); return
        header, ok = QtWidgets.QInputDialog.getText(
            self, "Header (optional)", "What to write at the top? (can leave empty):")
        if not ok:
            return
        n = len(paths)
        for i, p in enumerate(paths, 1):
            try:
                with Image.open(p) as im:
                    img = im.convert("RGB").copy()
                d = ImageDraw.Draw(img)
                size = max(16, img.width // 50)
                try:
                    font = ImageFont.truetype("arial.ttf", size)
                except Exception:
                    font = ImageFont.load_default()
                foot = "Page %d / %d" % (i, n)
                try:
                    tw = d.textbbox((0, 0), foot, font=font)[2]
                except Exception:
                    tw = len(foot) * size // 2
                d.text(((img.width - tw) // 2, img.height - size - 12), foot,
                       fill=(70, 70, 70), font=font)
                if header.strip():
                    d.text((16, 10), header.strip()[:80], fill=(70, 70, 70), font=font)
                save_image_keep_ext(img, p)
            except Exception:
                pass
        for r in range(self.list.count()):
            self._refresh_item(self.list.item(r))
        self._dirty = True
        self.status.showMessage("Page numbers added.", 4000)

    def watermark_pdf_tool(self):
        """Kisi bhi purani PDF par watermark/stamp chhapo."""
        src = self._pick_pdf("Kis PDF par watermark lagana hai?")
        if not src:
            return
        text, ok = QtWidgets.QInputDialog.getText(
            self, "Watermark", "What to write?",
            text=self._opts.get("watermark_text", "Noble Care Hospital"))
        if not ok or not text.strip():
            return
        pages = pdf_to_images(src, self._tmpdir)
        if not pages:
            self._warn("Could not extract pages from the PDF (install PyMuPDF)."); return
        out, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, 'Save', src[:-4] + "_wm.pdf", "PDF (*.pdf)")
        if not out:
            return
        try:
            imgs = [apply_watermark(Image.open(p).convert("RGB"), text.strip()) for p in pages]
            imgs[0].save(out, "PDF", save_all=True, append_images=imgs[1:], resolution=200.0)
            for im in imgs:
                im.close()
        except Exception:
            self._warn("Failed:\n%s" % traceback.format_exc()); return
        QtWidgets.QMessageBox.information(self, "Done", "Watermarked PDF:\n%s" % out)

    def remove_pdf_password(self):
        """Password pata ho to PDF ki bina-password copy banao."""
        if not HAS_OCR_LIBS:
            self._warn("pypdf is not installed."); return
        src = self._pick_pdf('Choose a password-protected PDF')
        if not src:
            return
        pw, ok = QtWidgets.QInputDialog.getText(
            self, "Password", "Password for this PDF:", QtWidgets.QLineEdit.Password)
        if not ok:
            return
        try:
            reader = PdfReader(src)
            if reader.is_encrypted and not reader.decrypt(pw or ""):
                self._warn("Wrong password."); return
            writer = PdfWriter()
            for pg in reader.pages:
                writer.add_page(pg)
            out = src[:-4] + "_unlocked.pdf"
            with open(out, "wb") as fh:
                writer.write(fh)
        except Exception:
            self._warn("Failed:\n%s" % traceback.format_exc()); return
        QtWidgets.QMessageBox.information(self, "Done", "Copy without password:\n%s" % out)

    def pdf_to_jpgs(self):
        """Kisi bhi PDF ke pages JPG images me nikaalo."""
        src = self._pick_pdf('Which PDF to extract images from?')
        if not src:
            return
        folder = QtWidgets.QFileDialog.getExistingDirectory(
            self, 'Where to save the images?', os.path.dirname(src))
        if not folder:
            return
        pages = pdf_to_images(src, self._tmpdir)
        if not pages:
            self._warn("Could not extract pages from the PDF (install PyMuPDF)."); return
        base = os.path.splitext(os.path.basename(src))[0]
        cnt = 0
        for i, p in enumerate(pages, 1):
            try:
                with Image.open(p) as im:
                    im.convert("RGB").save(
                        os.path.join(folder, "%s_p%02d.jpg" % (base, i)), "JPEG", quality=90)
                cnt += 1
            except Exception:
                pass
        QtWidgets.QMessageBox.information(self, "Done", "Created %d images:\n%s" % (cnt, folder))

    def folder_to_pdf(self):
        """Ek folder ki SAARI images (naam ke kram me) ek PDF me."""
        folder = QtWidgets.QFileDialog.getExistingDirectory(
            self, 'Choose the images folder', self._opts.get("save_folder", ""))
        if not folder:
            return
        exts = (".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".webp")
        files = sorted(os.path.join(folder, f) for f in os.listdir(folder)
                       if f.lower().endswith(exts))
        if not files:
            self._warn("No images found in this folder."); return
        out, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, 'Save PDF',
            os.path.join(folder, os.path.basename(folder) + ".pdf"), "PDF (*.pdf)")
        if not out:
            return
        try:
            imgs = [Image.open(f).convert("RGB") for f in files]
            imgs[0].save(out, "PDF", save_all=True, append_images=imgs[1:], resolution=200.0)
            for im in imgs:
                im.close()
        except Exception:
            self._warn("Failed:\n%s" % traceback.format_exc()); return
        QtWidgets.QMessageBox.information(
            self, "Done", "PDF created from %d images:\n%s" % (len(files), out))

    def _collect_page_texts(self, src_pdf=None):
        """OCR se har page ka text (current pages ya kisi PDF ka). Bhaari kaam —
        FuncWorker ke andar hi bulayein."""
        if src_pdf:
            pages = pdf_to_images(src_pdf, self._tmpdir) or []
        else:
            pages = self._ordered_paths()
        texts = []
        for p in pages:
            try:
                with Image.open(p) as im:
                    texts.append(pytesseract.image_to_string(im, lang="eng+hin"))
            except Exception:
                texts.append("")
        return texts

    def pdf_to_word(self):
        """Pages/PDF ka text OCR karke Word (.docx) file banao."""
        if not tesseract_available():
            self._warn("This needs Tesseract OCR."); return
        src = None
        if not self._ordered_paths():
            src = self._pick_pdf('Which PDF to convert to Word?')
            if not src:
                return
        default = (src[:-4] + ".docx") if src else self._build_filename(".docx")
        out, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, 'Save Word file', default, "Word (*.docx)")
        if not out:
            return

        def job():
            texts = self._collect_page_texts(src)
            try:
                import docx
                doc = docx.Document()
                for i, t in enumerate(texts):
                    if i:
                        doc.add_page_break()
                    doc.add_paragraph(t)
                doc.save(out)
                return out
            except ImportError:
                alt = os.path.splitext(out)[0] + ".txt"
                with open(alt, "w", encoding="utf-8") as fh:
                    fh.write("\n\n----\n\n".join(texts))
                return alt

        def done(res):
            if isinstance(res, Exception):
                self._warn("Failed to create Word:\n%s" % res); return
            QtWidgets.QMessageBox.information(self, "Done", "File created:\n%s" % res)
        self._run_bg(job, done, "Creating Word… (OCR running)")

    def pdf_to_excel(self):
        """Bill/table wale pages ko OCR karke Excel me nikaalo (best-effort)."""
        if not tesseract_available() or not HAS_XLSX:
            self._warn("This needs Tesseract OCR + openpyxl."); return
        src = None
        if not self._ordered_paths():
            src = self._pick_pdf('Which PDF to convert to Excel?')
            if not src:
                return
        default = (src[:-4] + ".xlsx") if src else self._build_filename(".xlsx")
        out, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, 'Save Excel', default, "Excel (*.xlsx)")
        if not out:
            return

        def job():
            pages = (pdf_to_images(src, self._tmpdir) or []) if src else self._ordered_paths()
            wb = openpyxl.Workbook()
            ws = wb.active
            for p in pages:
                with Image.open(p) as im:
                    data = pytesseract.image_to_data(
                        im.convert("RGB"), lang="eng+hin",
                        output_type=pytesseract.Output.DICT)
                lines = {}
                for i in range(len(data["text"])):
                    w = (data["text"][i] or "").strip()
                    if not w:
                        continue
                    key = (data["block_num"][i], data["par_num"][i], data["line_num"][i])
                    lines.setdefault(key, []).append((data["left"][i], data["width"][i], w))
                for key in sorted(lines):
                    words = sorted(lines[key])
                    cells, cur, last_end = [], [], None
                    for left, width, w in words:
                        if last_end is not None and left - last_end > 45:
                            cells.append(" ".join(cur)); cur = []
                        cur.append(w)
                        last_end = left + width
                    if cur:
                        cells.append(" ".join(cur))
                    ws.append(cells)
                ws.append([])
            wb.save(out)
            return out

        def done(res):
            if isinstance(res, Exception):
                self._warn("Failed to create Excel:\n%s" % res); return
            QtWidgets.QMessageBox.information(self, "Done", "Excel created:\n%s" % res)
        self._run_bg(job, done, "Creating Excel… (OCR running)")

    def save_archival_pdf(self):
        """High-quality PDF + poora metadata (title/date/producer) — lambe samay
        tak sambhal kar rakhne ke liye."""
        paths = self._ordered_paths()
        if not paths:
            self._warn(tr("scan_first", self._lang)); return
        if not HAS_OCR_LIBS:
            self._warn("pypdf is not installed."); return
        default = self._build_filename(".pdf")
        default = (default[:-4] if default.lower().endswith(".pdf") else default) + "_archive.pdf"
        out, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, 'Save archival PDF', default, "PDF (*.pdf)")
        if not out:
            return
        try:
            imgs = [Image.open(p).convert("RGB") for p in paths]
            tmp = out + ".tmp.pdf"
            imgs[0].save(tmp, "PDF", save_all=True, append_images=imgs[1:],
                         resolution=300.0, quality=95)
            for im in imgs:
                im.close()
            reader = PdfReader(tmp)
            writer = PdfWriter()
            for pg in reader.pages:
                writer.add_page(pg)
            writer.add_metadata({
                "/Title": os.path.splitext(os.path.basename(out))[0],
                "/Producer": "ApneScan v%s" % VERSION,
                "/CreationDate": datetime.datetime.now().strftime("D:%Y%m%d%H%M%S"),
            })
            with open(out, "wb") as fh:
                writer.write(fh)
            os.remove(tmp)
        except Exception:
            self._warn("Failed:\n%s" % traceback.format_exc()); return
        self._remember_save_dir(out)
        self._record_save(out, len(paths))
        QtWidgets.QMessageBox.information(
            self, "Done", "Archival PDF (300dpi, with metadata):\n%s" % out)

    def copy_page_text(self):
        """Selected page ka poora text OCR karke clipboard par."""
        item = self._current_item_or_warn()
        if not item:
            return
        if not tesseract_available():
            self._warn("This needs Tesseract OCR."); return
        path = item.data(QtCore.Qt.UserRole)

        def job():
            with Image.open(path) as im:
                return pytesseract.image_to_string(im, lang="eng+hin")

        def done(res):
            if isinstance(res, Exception):
                self._warn("OCR failed:\n%s" % res); return
            QtWidgets.QApplication.clipboard().setText(res or "")
            self.status.showMessage("Text copied (paste anywhere with Ctrl+V).", 5000)
        self._run_bg(job, done, "Reading text…")

    def translate_page(self):
        """Page ka text padh kar Hindi ↔ English translate karo (internet chahiye)."""
        item = self._current_item_or_warn()
        if not item:
            return
        if not tesseract_available():
            self._warn("This needs Tesseract OCR."); return
        opts = ["English → Hindi", "Hindi → English"]
        pick, ok = QtWidgets.QInputDialog.getItem(
            self, "Translate", "Translate in which direction?", opts, 0, False)
        if not ok:
            return
        tl = "hi" if pick.startswith("English") else "en"
        path = item.data(QtCore.Qt.UserRole)

        def job():
            with Image.open(path) as im:
                text = pytesseract.image_to_string(im, lang="eng+hin")
            text = (text or "").strip()[:4500]
            if not text:
                raise RuntimeError('No readable text found on the page.')
            import urllib.parse as P
            import urllib.request as U
            import json as J
            url = ("https://translate.googleapis.com/translate_a/single"
                   "?client=gtx&sl=auto&tl=%s&dt=t&q=%s" % (tl, P.quote(text)))
            req = U.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            data = J.loads(U.urlopen(req, timeout=20).read().decode("utf-8", "ignore"))
            return "".join(seg[0] for seg in data[0] if seg and seg[0])

        def done(res):
            if isinstance(res, Exception):
                self._warn("Translation failed (is the internet working?):\n%s" % res)
                return
            dlg = QtWidgets.QDialog(self)
            dlg.setWindowTitle("Translation (%s)" % pick)
            dlg.resize(560, 420)
            v = QtWidgets.QVBoxLayout(dlg)
            ed = QtWidgets.QPlainTextEdit(res)
            v.addWidget(ed, 1)
            row = QtWidgets.QHBoxLayout()
            bcopy = QtWidgets.QPushButton("Copy")
            bcopy.clicked.connect(
                lambda: QtWidgets.QApplication.clipboard().setText(ed.toPlainText()))
            bcl = QtWidgets.QPushButton("Close"); bcl.clicked.connect(dlg.accept)
            row.addWidget(bcopy); row.addStretch(1); row.addWidget(bcl)
            v.addLayout(row)
            dlg.exec_()
        self._run_bg(job, done, "Translating…")

    # ---- page edit ----
    def _current_item_or_warn(self):
        item = self.list.currentItem()
        if item is None:
            self._warn("Select a page first.")
        return item

    def _rotate(self, angle):
        # ab _edit_current_bg se — undo + 'sab pages par' dono chalte hain
        self._edit_current_bg(lambda im: im.rotate(angle, expand=True),
                              self.L("Ghuma rahe hain…", "Rotating…"))

    def rotate_left(self):
        self._rotate(90)

    def rotate_right(self):
        self._rotate(-90)

    def _enhance_current(self, brightness, contrast):
        from PIL import ImageEnhance
        def _t(im):
            im = ImageEnhance.Brightness(im).enhance(brightness)
            im = ImageEnhance.Contrast(im).enhance(contrast)
            return im
        self._edit_current_bg(_t, self.L("Badlav ho raha hai…", "Applying…"))

    def autocrop_current(self):
        self._edit_current_bg(lambda im: autocrop(im),
                              self.L("Crop ho raha hai…", "Cropping…"))

    def _list_context_menu(self, pos):
        if self.list.count() == 0:
            return
        item = self.list.itemAt(pos)
        menu = QtWidgets.QMenu(self)
        act_rename = menu.addAction("\u270f Rename")
        act_del = menu.addAction("\U0001f5d1 Delete")
        act_rescan = act_insert = None
        if item is not None:
            menu.addSeparator()
            act_rescan = menu.addAction(self.L("\ud83d\udd04 Is page ko dobara scan karo",
                                               "\ud83d\udd04 Rescan this page"))
            act_insert = menu.addAction(self.L("\u2795 Iske baad naya scan jodo",
                                               "\u2795 Scan & insert after this"))
        chosen = menu.exec_(self.list.viewport().mapToGlobal(pos))
        if chosen == act_rename:
            self.rename_current_page()
        elif chosen == act_del:
            self.delete_page()
        elif act_rescan is not None and chosen == act_rescan:
            self._rescan_page(self.list.row(item))
        elif act_insert is not None and chosen == act_insert:
            self._insert_scan_after(self.list.row(item))

    def _rescan_page(self, row):
        """Us page ko hata kar us HI jagah naya scan lagao (Glass/flatbed ke
        liye best \u2014 scanner par sahi page rakho, phir scan)."""
        if row < 0 or row >= self.list.count():
            return
        if getattr(self, "_scanning", False):
            self._warn(self.L("Scan chal raha hai \u2014 ruko.", "A scan is already running."))
            return
        self._pending_place = ("replace", row)
        self.status.showMessage(self.L(
            "\ud83d\udd04 Scanner par sahi page rakho \u2014 naya page usi jagah lag jayega.",
            "\ud83d\udd04 Put the correct page on the scanner \u2014 it will replace this one."), 5000)
        self.do_scan()

    def _insert_scan_after(self, row):
        """Is page ke BAAD naya scan jodo (cram bigade bina beech me page)."""
        if getattr(self, "_scanning", False):
            self._warn(self.L("Scan chal raha hai \u2014 ruko.", "A scan is already running."))
            return
        self._pending_place = ("insert", row + 1)
        self.status.showMessage(self.L(
            "\u2795 Naya scan is page ke baad jud jayega.",
            "\u2795 The new scan will be inserted right after this page."), 5000)
        self.do_scan()

    def rename_current_page(self):
        # If SEVERAL pages are selected, rename them ALL to the same name at once.
        sel = self.list.selectedItems()
        if len(sel) > 1:
            cur = sel[0].data(TITLE_ROLE) or ""
            name, ok = QtWidgets.QInputDialog.getText(
                self, "Rename %d pages" % len(sel),
                "One name for all %d selected pages:" % len(sel), text=cur)
            if not ok:
                return
            name = underscore_name(name)
            if not name:
                return
            for it in sel:
                it.setData(TITLE_ROLE, name)
                it.setText(name)
            self._learn_name(sel[0].data(QtCore.Qt.UserRole), name)
            self.status.showMessage("Renamed %d pages to '%s'" % (len(sel), name), 4000)
            return
        it = self.list.currentItem() or (sel or [None])[0]
        if it is None:
            self._warn("Select a page first."); return
        cur = it.data(TITLE_ROLE) or ""
        name, ok = QtWidgets.QInputDialog.getText(self, "Rename", "Name for this page:", text=cur)
        if not ok:
            return
        name = underscore_name(name)
        if not name:
            return
        it.setData(TITLE_ROLE, name)
        it.setText(name)
        # LEARN: remember this document's content -> this name, so next time a
        # similar page is scanned it gets named automatically.
        self._learn_name(it.data(QtCore.Qt.UserRole), name)

    def _learn_name(self, path, name):
        """Rename ka OCR-learning ab BACKGROUND me hota hai — pehle yahi OCR
        UI thread par chal kar har rename ko 3-10 second ke liye jama deta tha."""
        if not path or not tesseract_available():
            return
        self._learn_workers = getattr(self, "_learn_workers", [])
        w = LearnWorker(path, name)
        w.got.connect(self._store_learned_name)
        w.finished.connect(lambda w=w: self._learn_workers.remove(w)
                           if w in self._learn_workers else None)
        self._learn_workers.append(w)
        w.start()

    def _store_learned_name(self, name, words):
        if len(words) < 3:
            self.status.showMessage(
                self.L("Is page par itna chhapa hua text nahi hai ki naam yaad "
                       "rakha ja sake (jaise haath se likha page). Naam laga diya "
                       "hai par 'auto' nahi aayega.",
                       "Not enough printed text on this page to remember the name "
                       "(e.g. handwritten). Name applied, but won't auto-fill."), 7000)
            return
        learned = self._config.setdefault("learned_names", [])
        wset = set(words)
        # Bahut milta-julta signature pehle se ho to usi ka naam update karo,
        # warna naya jodo. (containment se — same form ki thodi alag scan bhi
        # ek hi entry rahe, bekaar duplicate na banein.)
        for e in learned:
            ex = set(e.get("words", []))
            if max(_jaccard(wset, ex), _containment(wset, ex), _containment(ex, wset)) >= 0.55:
                e["words"] = words
                e["name"] = name
                break
        else:
            learned.append({"words": words, "name": name})
            if len(learned) > 400:
                del learned[0:len(learned) - 400]
        try:
            save_config(self._config)
        except Exception:
            pass
        cnt = len(learned)
        self.status.showMessage(
            self.L("\u2714 Naam yaad rakh liya \u2014 agli baar aisa hi document apne aap "
                   "'%s' ho jayega. (%d naam seekhe hue)" % (name, cnt),
                   "\u2714 Name learned \u2014 next time a similar document auto-names as "
                   "'%s'. (%d names learned)" % (name, cnt)), 6000)

    def manage_learned_names(self):
        """Seekhe hue naam dekho / hatao \u2014 kaunsa document kis naam se yaad hai."""
        learned = self._config.get("learned_names", []) or []
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle(self.L("Seekhe hue naam", "Learned names"))
        dlg.resize(460, 460)
        v = QtWidgets.QVBoxLayout(dlg)
        v.addWidget(QtWidgets.QLabel(self.L(
            "%d document-naam yaad hain. Kisi ko chun kar hata/badal sakte ho:" % len(learned),
            "%d document names are remembered. Select one to remove/rename:" % len(learned))))
        lw = QtWidgets.QListWidget()

        def _row_text(e):
            fol = ("  \u2192  \ud83d\udcc1 " + os.path.basename(e["folder"])) if e.get("folder") else ""
            return "\ud83d\udcc4 %s   (%d shabd)%s" % (e.get("name", "-"), len(e.get("words", [])), fol)
        for e in learned:
            it = QtWidgets.QListWidgetItem(_row_text(e))
            it.setData(QtCore.Qt.UserRole, e)
            lw.addItem(it)
        v.addWidget(lw, 1)
        row = QtWidgets.QHBoxLayout()

        def _set_folder():
            it = lw.currentItem()
            if not it:
                return
            e = it.data(QtCore.Qt.UserRole)
            f = QtWidgets.QFileDialog.getExistingDirectory(
                dlg, self.L("Is naam ke documents kis folder me save hon?",
                            "Save documents with this name to which folder?"),
                e.get("folder") or self._opts.get("save_folder", ""))
            if f:
                e["folder"] = f
                it.setText(_row_text(e))
                save_config(self._config)

        def _rename():
            it = lw.currentItem()
            if not it:
                return
            e = it.data(QtCore.Qt.UserRole)
            nn, ok = QtWidgets.QInputDialog.getText(
                dlg, self.L("Naam badlo", "Rename"), self.L("Naya naam:", "New name:"),
                text=e.get("name", ""))
            if ok and nn.strip():
                e["name"] = underscore_name(nn.strip())
                it.setText("\ud83d\udcc4 %s   (%d words)" % (e["name"], len(e.get("words", []))))
                save_config(self._config)

        def _del():
            it = lw.currentItem()
            if not it:
                return
            e = it.data(QtCore.Qt.UserRole)
            try:
                learned.remove(e)
            except ValueError:
                pass
            lw.takeItem(lw.row(it))
            save_config(self._config)

        def _clear():
            if QtWidgets.QMessageBox.question(
                    dlg, "Confirm", self.L("Saare seekhe naam hata dein?",
                                           "Remove all learned names?"),
                    QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                    QtWidgets.QMessageBox.No) == QtWidgets.QMessageBox.Yes:
                learned.clear(); lw.clear(); save_config(self._config)
        def _export():
            out, _ = QtWidgets.QFileDialog.getSaveFileName(
                dlg, self.L("Seekhe naam export", "Export learned names"),
                os.path.join(os.path.expanduser("~"), "apnescan_names.json"),
                "Names (*.json)")
            if not out:
                return
            try:
                with open(out, "w", encoding="utf-8") as fh:
                    json.dump(learned, fh, ensure_ascii=False)
                QtWidgets.QMessageBox.information(dlg, "OK", self.L(
                    "Export ho gaya:\n%s\n\nNaye PC par Import se le aana." % out,
                    "Exported:\n%s\n\nUse Import on another PC." % out))
            except Exception as exc:
                self._warn("Export failed: %s" % exc)

        def _import():
            f, _ = QtWidgets.QFileDialog.getOpenFileName(
                dlg, self.L("Naam file chuno", "Choose names file"),
                os.path.expanduser("~"), "Names (*.json)")
            if not f:
                return
            try:
                with open(f, "r", encoding="utf-8") as fh:
                    incoming = json.load(fh)
                added = 0
                for e in incoming:
                    if not isinstance(e, dict) or not e.get("name"):
                        continue
                    ex = set(e.get("words", []))
                    dup = any(_containment(ex, set(o.get("words", []))) >= 0.7
                              for o in learned)
                    if not dup:
                        learned.append(e); added += 1
                save_config(self._config)
                lw.clear()
                for e in learned:
                    it = QtWidgets.QListWidgetItem(_row_text(e))
                    it.setData(QtCore.Qt.UserRole, e); lw.addItem(it)
                QtWidgets.QMessageBox.information(dlg, "OK", self.L(
                    "%d naye naam aa gaye." % added, "%d new names imported." % added))
            except Exception as exc:
                self._warn("Import failed: %s" % exc)
        for t, fn in ((self.L("\u270f Naam badlo", "\u270f Rename"), _rename),
                      (self.L("\ud83d\udcc1 Folder set", "\ud83d\udcc1 Set folder"), _set_folder),
                      (self.L("\ud83d\uddd1 Hatao", "\ud83d\uddd1 Remove"), _del),
                      (self.L("Sab hatao", "Clear all"), _clear)):
            b = QtWidgets.QPushButton(t); b.clicked.connect(fn); row.addWidget(b)
        row.addStretch(1)
        v.addLayout(row)
        row2 = QtWidgets.QHBoxLayout()
        be = QtWidgets.QPushButton(self.L("\ud83d\udce4 Export", "\ud83d\udce4 Export")); be.clicked.connect(_export)
        bi = QtWidgets.QPushButton(self.L("\ud83d\udce5 Import", "\ud83d\udce5 Import")); bi.clicked.connect(_import)
        row2.addWidget(be); row2.addWidget(bi); row2.addStretch(1)
        bc = QtWidgets.QPushButton(self.L("Band karo", "Close")); bc.clicked.connect(dlg.accept)
        row2.addWidget(bc)
        v.addLayout(row2)
        dlg.exec_()

    def _list_context_menu_end(self):
        pass

    def delete_page(self):
        items = self.list.selectedItems()
        if not items:
            item = self._current_item_or_warn()
            if not item:
                return
            items = [item]
        # delete from the highest row down so earlier indices stay valid
        rows = sorted((self.list.row(it) for it in items), reverse=True)
        for row in rows:
            it = self.list.item(row)
            if it is None:
                continue
            path = it.data(QtCore.Qt.UserRole)
            self.list.takeItem(row)
            self._undo_stack.append((path, row))     # keep file for undo
        self._undo_stack = self._undo_stack[-15:]
        self._dirty = True
        self._renumber_pages()
        self._update_status(); self._update_empty_state()

    def move_up(self):
        row = self.list.currentRow()
        if row > 0:
            it = self.list.takeItem(row); self.list.insertItem(row - 1, it); self.list.setCurrentItem(it)

    def move_down(self):
        row = self.list.currentRow()
        if 0 <= row < self.list.count() - 1:
            it = self.list.takeItem(row); self.list.insertItem(row + 1, it); self.list.setCurrentItem(it)

    def clear_all(self):
        if self.list.count() == 0:
            return
        if QtWidgets.QMessageBox.question(self, "Confirm", "Delete all pages?") != QtWidgets.QMessageBox.Yes:
            return
        for p in self._ordered_paths():
            try:
                os.remove(p)
            except Exception:
                pass
        self.list.clear()
        for p, _r in self._undo_stack:
            try:
                os.remove(p)
            except Exception:
                pass
        self._undo_stack = []
        self._dirty = False
        self._update_status(); self._update_empty_state()

    # ---- filenames + saving ----
    def _remember_save_dir(self, path):
        try:
            d = os.path.dirname(path)
            if d and os.path.isdir(d):
                self._config["last_save_dir"] = d
                save_config(self._config)
        except Exception:
            pass

    def _target_folder(self):
        # #2: is document ke naam ke liye folder yaad ho to wahi (sabse pehle)
        hint = getattr(self, "_doc_folder_hint", None)
        if hint and os.path.isdir(hint):
            return hint
        # If the user isn't using auto-organizing folders, reopen wherever they
        # saved last time.
        last = self._config.get("last_save_dir")
        if (last and os.path.isdir(last)
                and not self._opts.get("year_month_folders")
                and not self._opts.get("make_claim_folder")):
            return last
        folder = self._opts.get("save_folder") or os.path.join(os.path.expanduser("~"), "Documents")
        now = datetime.datetime.now()
        if self._opts.get("year_month_folders"):
            folder = os.path.join(folder, now.strftime("%Y"), now.strftime("%m"))
        claim = sanitize(self.claim_edit.text().strip(), "")
        if self._opts.get("make_claim_folder") and claim:
            folder = os.path.join(folder, claim)
        try:
            os.makedirs(folder, exist_ok=True)
        except Exception:
            pass
        return folder

    def _suggested_save_name(self, paths=None):
        """Name to default the save filename to. If specific paths are being saved
        (e.g. Save Selected), use the FIRST of THOSE pages' names — so each document
        saves with its own name instead of always the first page's."""
        pathset = set(paths) if paths else None
        for i in range(self.list.count()):
            it = self.list.item(i)
            if pathset is not None and it.data(QtCore.Qt.UserRole) not in pathset:
                continue
            t = it.data(TITLE_ROLE)
            if t:
                return underscore_name(t)
        return None

    def _remember_doc_name(self, out):
        """After saving, remember this saved name against the document title so the
        SAME document (next scan) shows this exact name again."""
        base = os.path.splitext(os.path.basename(out))[0]
        keys = []
        for i in range(self.list.count()):
            k = self.list.item(i).data(NAMEKEY_ROLE)
            if k:
                keys.append(k)
        if not keys:
            return
        dn = self._config.setdefault("doc_names", {})
        # map the first page's title (the document's identity) to this saved name
        dn[keys[0]] = base
        try:
            save_config(self._config)
        except Exception:
            pass

    def _build_filename(self, ext=".pdf", seq=1, paths=None):
        # If pages were auto-named, default the save filename to that document name
        # (spaces already underscored). Otherwise use the claim/date template.
        suggested = self._suggested_save_name(paths)
        if suggested:
            return os.path.join(self._target_folder(), suggested + ext)
        now = datetime.datetime.now()
        claim = sanitize(self.claim_edit.text().strip(), "scan")
        # Pehle page ka OCR-naam (ho to) — {name} placeholder ke liye
        docname = ""
        try:
            it0 = self.list.item(0)
            if it0:
                docname = underscore_name(it0.data(TITLE_ROLE) or "")
        except Exception:
            pass
        name = self._opts.get("filename_template", "{claim}_{date}_{seq}")
        name = (name.replace("{claim}", claim)
                    .replace("{date}", now.strftime("%Y-%m-%d"))
                    .replace("{year}", now.strftime("%Y"))
                    .replace("{month}", now.strftime("%m"))
                    .replace("{day}", now.strftime("%d"))
                    .replace("{time}", now.strftime("%H%M%S"))
                    .replace("{name}", docname or "scan")
                    .replace("{seq}", "%03d" % seq))
        return os.path.join(self._target_folder(), underscore_name(name) + ext)

    def _validate_claim_ok(self):
        if not self._opts.get("validate_claim"):
            return True
        claim = self.claim_edit.text().strip()
        pat = self._opts.get("claim_pattern") or r"^[A-Za-z0-9\-]{4,}$"
        try:
            if re.match(pat, claim):
                return True
        except Exception:
            return True
        self._warn("The claim number does not look valid (pattern did not match).\n"
                   "Please correct it or turn off validation in Settings.")
        return False

    def _duplicate_ok(self):
        if not self._opts.get("duplicate_check"):
            return True
        claim = self.claim_edit.text().strip()
        if claim and claim in self._used_claims:
            return QtWidgets.QMessageBox.question(
                self, "Duplicate", "This claim number has already been saved.\nSave anyway?",
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.No) == QtWidgets.QMessageBox.Yes
        return True

    def _pages_as_pdf(self, paths, out, password=None):
        compress = self._opts.get("compress")
        q = int(self._opts.get("jpeg_quality", 60))
        wm = self._opts.get("watermark")
        wt = self._opts.get("watermark_text", "")
        imgs = []
        for p in paths:
            im = Image.open(p).convert("RGB")
            if wm and wt:
                im = apply_watermark(im, wt)
            if compress:
                buf = io.BytesIO(); im.save(buf, "JPEG", quality=q); buf.seek(0)
                im2 = Image.open(buf); im2.load(); im = im2
            imgs.append(im)
        if password:
            tmp = out + ".tmp.pdf"
            imgs[0].save(tmp, "PDF", save_all=True, append_images=imgs[1:], resolution=200.0)
            reader = PdfReader(tmp); writer = PdfWriter()
            for pg in reader.pages:
                writer.add_page(pg)
            writer.encrypt(password)
            with open(out, "wb") as fh:
                writer.write(fh)
            try:
                os.remove(tmp)
            except Exception:
                pass
        else:
            imgs[0].save(out, "PDF", save_all=True, append_images=imgs[1:], resolution=200.0)
        for im in imgs:
            try:
                im.close()
            except Exception:
                pass

    def _record_save(self, out, num_pages):
        claim = self.claim_edit.text().strip()
        # personal stats (sirf aapke PC par)
        try:
            it0 = self.list.item(0)
            dt = (it0.data(TITLE_ROLE) if it0 else "") or ""
            dt = dt.split("_")[0] if dt else None
        except Exception:
            dt = None
        self._pstats_bump(pages=num_pages, pdfs=1, doc_type=dt)
        # recent
        self._add_recent(out)
        # used claims (duplicate detection)
        if claim:
            self._used_claims.add(claim)
            self._config["used_claims"] = list(self._used_claims)[-2000:]
            save_config(self._config)
        # activity log
        if self._opts.get("activity_log"):
            try:
                logp = os.path.join(self._opts.get("save_folder", "."), "activity_log.txt")
                os.makedirs(os.path.dirname(logp), exist_ok=True)
                with open(logp, "a", encoding="utf-8") as fh:
                    fh.write("%s\tclaim=%s\tpages=%d\t%s\n"
                             % (datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                claim, num_pages, out))
            except Exception:
                pass
        # excel register
        if self._opts.get("excel_log") and HAS_XLSX:
            self._append_excel(claim, num_pages, out)
        # backup
        if self._opts.get("backup"):
            self._backup_file(out)
        # kabhi-kabhi (bahut halke se) "app share karo" ki yaad (30 din me 1 baar)
        self._maybe_growth_nudge()

    def _append_excel(self, claim, num_pages, out):
        try:
            xp = os.path.join(self._opts.get("save_folder", "."), "register.xlsx")
            os.makedirs(os.path.dirname(xp), exist_ok=True)
            if os.path.exists(xp):
                wb = openpyxl.load_workbook(xp)
                ws = wb.active
            else:
                wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Register"
                ws.append(["Date", "Time", "Claim No.", "Pages", "File"])
            now = datetime.datetime.now()
            ws.append([now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S"),
                       claim, num_pages, out])
            wb.save(xp)
        except Exception:
            pass

    def _backup_file(self, out):
        try:
            day = datetime.datetime.now().strftime("%Y-%m-%d")
            bdir = os.path.join(self._opts.get("backup_folder", "."), day)
            os.makedirs(bdir, exist_ok=True)
            shutil.copy2(out, os.path.join(bdir, os.path.basename(out)))
        except Exception:
            pass

    def _auto_save_pdf(self):
        paths = self._ordered_paths()
        if not paths:
            return False
        if not self._validate_claim_ok() or not self._duplicate_ok():
            return False
        out = self._build_filename(".pdf")
        try:
            if HAS_OCR_LIBS and self._opts.get("searchable_pdf"):
                self._save_ocr_pdf(paths, out)   # PDF ke andar dhoondhne layak text
            else:
                self._pages_as_pdf(paths, out)
        except Exception:
            self._warn("Auto-save failed:\n%s" % traceback.format_exc()); return False
        self._record_save(out, len(paths)); self._dirty = False
        if self._opts.get("save_images_too"):
            base = os.path.splitext(out)[0]
            for i, p in enumerate(paths):
                try:
                    with Image.open(p) as im:
                        im.convert("RGB").save("%s_p%d.jpg" % (base, i + 1), "JPEG", quality=90)
                except Exception:
                    pass
        self._after_save_action(out)
        self.status.showMessage("Auto-save: %s" % out, 8000)
        return True

    def save_pdf_all(self):
        self.save_pdf(self._ordered_paths())

    def save_pdf_selected(self):
        paths = self._selected_paths()
        if not paths:
            self._warn(tr("no_selection", self._lang)); return
        self.save_pdf(paths)

    def save_pdf(self, paths=None):
        if not isinstance(paths, (list, tuple)):
            paths = self._ordered_paths()
        if not paths:
            self._warn(tr("scan_first", self._lang)); return
        if not self._validate_claim_ok() or not self._duplicate_ok():
            return
        default = self._build_filename(".pdf", paths=paths)
        out, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Save PDF', default, "PDF (*.pdf)")
        if not out:
            return
        if not out.lower().endswith(".pdf"):
            out += ".pdf"
        ocr = HAS_OCR_LIBS and (self.chk_ocr.isChecked()
                                or self._opts.get("searchable_pdf"))
        npages = len(paths)

        def job():
            if ocr:
                self._save_ocr_pdf(paths, out)   # OCR = sabse bhaari — ab background me
            else:
                self._pages_as_pdf(paths, out)
            return out

        def done(res):
            if isinstance(res, Exception):
                if HAS_OCR_LIBS and isinstance(res, pytesseract.TesseractNotFoundError):
                    self._warn("Tesseract OCR not found. Try without OCR.")
                else:
                    self._warn("PDF save failed:\n%s" % res)
                return
            self._remember_save_dir(out); self._remember_doc_name(out)
            self._record_save(out, npages); self._dirty = False; self._after_save_action(out)
            self.status.showMessage("✔ PDF saved: %s" % out, 8000)
            self._open_saved_in_panel(out)   # us file ka folder Meri Files panel me khol do
        self._run_bg(job, done,
                     "Saving PDF…" if not ocr else "Creating OCR PDF…")

    def _open_saved_in_panel(self, out):
        """Document save hote hi uska folder Meri Files panel me khol do aur
        file ko select kar do (panel chhupa ho to dikha bhi do)."""
        try:
            folder = os.path.dirname(out)
            if not folder or not os.path.isdir(folder):
                return
            if hasattr(self, "files_panel") and not self.files_panel.isVisible():
                self.files_panel.setVisible(True)
                self._opts["show_files_panel"] = True
            self._jump_to_folder(folder)
            idx = self.files_model.index(out)
            self.files_tree.setCurrentIndex(idx)
            self.files_tree.scrollTo(idx)
            self._invalidate_files_index()   # nayi file bani — index refresh
        except Exception:
            pass

    def save_pdf_password(self):
        paths = self._ordered_paths()
        if not paths:
            self._warn("Scan or import a page first."); return
        if not HAS_OCR_LIBS:
            self._warn("pypdf is not installed (required for password)."); return
        pw, ok = QtWidgets.QInputDialog.getText(self, "Password", "PDF password:", QtWidgets.QLineEdit.Password)
        if not ok or not pw:
            return
        out, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Save PDF', self._build_filename(".pdf"), "PDF (*.pdf)")
        if not out:
            return
        if not out.lower().endswith(".pdf"):
            out += ".pdf"
        npages = len(paths)

        def job():
            self._pages_as_pdf(paths, out, password=pw)
            return out

        def done(res):
            if isinstance(res, Exception):
                self._warn("PDF save failed:\n%s" % res); return
            self._remember_save_dir(out); self._remember_doc_name(out)
            self._record_save(out, npages); self._dirty = False; self._after_save_action(out)
            self.status.showMessage("✔ Password PDF saved: %s" % out, 8000)
            self._open_saved_in_panel(out)
        self._run_bg(job, done, "Creating password PDF…")

    def save_images(self):
        paths = self._ordered_paths()
        if not paths:
            self._warn("Scan or import a page first."); return
        out, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, 'Save images', self._build_filename(".jpg"),
            "JPEG (*.jpg);;PNG (*.png);;TIFF (*.tif)")
        if not out:
            return
        base, ext = os.path.splitext(out)
        if not ext:
            ext = ".jpg"
        fmt = {"jpg": "JPEG", "jpeg": "JPEG", "png": "PNG", "tif": "TIFF", "tiff": "TIFF"}\
            .get(ext.lower().lstrip("."), "JPEG")
        npages = len(paths)

        def job():
            for i, p in enumerate(paths):
                target = (base + ext) if i == 0 else ("%s_%d%s" % (base, i + 1, ext))
                with Image.open(p) as im:
                    im.convert("RGB").save(target, fmt)
            return npages

        def done(res):
            if isinstance(res, Exception):
                self._warn("Images save failed:\n%s" % res); return
            self._remember_save_dir(out)
            self.status.showMessage("✔ Saved %d image(s)." % res, 8000)
        self._run_bg(job, done, "Saving images…")

    def export_ocr_text(self):
        if not HAS_OCR_LIBS:
            self._warn("pytesseract is not installed."); return
        paths = self._ordered_paths()
        if not paths:
            self._warn("Scan or import a page first."); return
        out, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save text", self._build_filename(".txt"), "Text (*.txt)")
        if not out:
            return
        def job():
            chunks = []
            for p in paths:
                with Image.open(p) as im:
                    chunks.append(pytesseract.image_to_string(im, lang="hin+eng"))
            with open(out, "w", encoding="utf-8") as fh:
                fh.write("\n\n----\n\n".join(chunks))
            return out

        def done(res):
            if isinstance(res, Exception):
                if isinstance(res, pytesseract.TesseractNotFoundError):
                    self._warn("Tesseract OCR is not installed.")
                else:
                    self._warn("Text export failed:\n%s" % res)
                return
            self._remember_save_dir(out)
            self.status.showMessage("✔ Text saved: %s" % out, 8000)
        self._run_bg(job, done, "Extracting OCR text…")

    def _save_ocr_pdf(self, paths, out):
        writer = PdfWriter()
        for p in paths:
            with Image.open(p) as im:
                pdf_bytes = pytesseract.image_to_pdf_or_hocr(im, extension="pdf", lang="hin+eng")
            reader = PdfReader(io.BytesIO(pdf_bytes))
            for pg in reader.pages:
                writer.add_page(pg)
        with open(out, "wb") as fh:
            writer.write(fh)

    # ---- Tools ----
    def split_pdfs(self):
        paths = self._ordered_paths()
        if not paths:
            self._warn("Scan or import a page first."); return
        text, ok = QtWidgets.QInputDialog.getText(
            self, "Split PDFs", "Which pages start a new document? (e.g. 1,5,9):", text="1")
        if not ok or not text.strip():
            return
        try:
            starts = sorted(set(int(x) for x in text.replace(" ", "").split(",") if x))
        except Exception:
            self._warn("Invalid input."); return
        starts = [s for s in starts if 1 <= s <= len(paths)]
        if not starts or starts[0] != 1:
            starts = [1] + starts
        folder = QtWidgets.QFileDialog.getExistingDirectory(self, "Folder", self._opts.get("save_folder", ""))
        if not folder:
            return
        os.makedirs(folder, exist_ok=True)
        base = sanitize(self.claim_edit.text().strip(), "scan") + "_" + datetime.datetime.now().strftime("%Y-%m-%d")
        ranges = []
        for i, s in enumerate(starts):
            end = (starts[i + 1] - 1) if i + 1 < len(starts) else len(paths)
            ranges.append((s, end))

        def job():
            for idx, (s, e) in enumerate(ranges, 1):
                self._pages_as_pdf(paths[s - 1:e], os.path.join(folder, "%s_part%d.pdf" % (base, idx)))
            return len(ranges)

        def done(res):
            if isinstance(res, Exception):
                self._warn("Split failed:\n%s" % res); return
            QtWidgets.QMessageBox.information(self, "Done", "Created %d separate PDF(s):\n%s" % (res, folder))
        self._run_bg(job, done, "Splitting PDF…")

    def merge_pdfs(self):
        if not HAS_OCR_LIBS:
            self._warn("pypdf is not installed."); return
        files, _ = QtWidgets.QFileDialog.getOpenFileNames(self, 'Choose PDFs to merge', "", "PDF (*.pdf)")
        if not files:
            return
        out, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Merged PDF save", self._build_filename(".pdf"), "PDF (*.pdf)")
        if not out:
            return
        if not out.lower().endswith(".pdf"):
            out += ".pdf"

        def job():
            writer = PdfWriter()
            for f in files:
                r = PdfReader(f)
                for pg in r.pages:
                    writer.add_page(pg)
            with open(out, "wb") as fh:
                writer.write(fh)
            return out

        def done(res):
            if isinstance(res, Exception):
                self._warn("Merge failed:\n%s" % res); return
            self._add_recent(out)
            self.status.showMessage("✔ Merged PDF created: %s" % out, 8000)
        self._run_bg(job, done, "Merging PDFs…")

    def search_pdfs(self):
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Search past PDFs")
        v = QtWidgets.QVBoxLayout(dlg)
        v.addWidget(QtWidgets.QLabel("Claim number / name / any word:"))
        ed = QtWidgets.QLineEdit()
        v.addWidget(ed)
        chk = QtWidgets.QCheckBox("Also search text INSIDE the PDF (slow — only works on OCR'd PDFs)")
        v.addWidget(chk)
        bb = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        bb.accepted.connect(dlg.accept)
        bb.rejected.connect(dlg.reject)
        v.addWidget(bb)
        ed.setFocus()
        if dlg.exec_() != QtWidgets.QDialog.Accepted or not ed.text().strip():
            return
        q = ed.text().strip().lower()
        root = self._opts.get("save_folder", os.path.expanduser("~"))
        pdfs = []
        try:
            for dp, _dn, fn in os.walk(root):
                for f in fn:
                    if f.lower().endswith(".pdf"):
                        pdfs.append(os.path.join(dp, f))
        except Exception:
            pass
        matches = [p for p in pdfs if q in os.path.basename(p).lower()]
        # tag match bhi dikhao
        for p, lst in (self._opts.get("tags") or {}).items():
            if p not in matches and any(q in t.lower() for t in lst) and os.path.exists(p):
                matches.append(p)
        if chk.isChecked() and HAS_OCR_LIBS:
            # pehle INDEX se turant dhoondo; sirf un-indexed files live padho
            idx = self._load_index()
            todo = []
            for p in pdfs:
                if p in matches:
                    continue
                e = idx.get(p)
                if e is not None:
                    if q in (e.get("t") or ""):
                        matches.append(p)
                else:
                    todo.append(p)
            prog = QtWidgets.QProgressDialog(
                'Searching inside PDFs… (%d files)' % len(todo), "Cancel", 0, len(todo), self)
            prog.setWindowModality(QtCore.Qt.WindowModal)
            prog.setMinimumDuration(400)
            for i, p in enumerate(todo):
                prog.setValue(i)
                QtWidgets.QApplication.processEvents()
                if prog.wasCanceled():
                    break
                try:
                    reader = PdfReader(p)
                    text = ""
                    for pg in reader.pages[:5]:
                        text += (pg.extract_text() or "").lower()
                        if q in text:
                            matches.append(p)
                            break
                except Exception:
                    pass
            prog.setValue(len(todo))
        if not matches:
            QtWidgets.QMessageBox.information(
                self, "Search",
                "Nothing found.\n\nTip: text-inside-PDF search only finds PDFs that were saved with 'OCR (searchable)' ticked.")
            return
        item, ok = QtWidgets.QInputDialog.getItem(
            self, "Search results", "Found %d file(s) — choose one to open:" % len(matches),
            matches, 0, False)
        if ok and item:
            self._open_path(item)

    def monthly_report(self):
        month = datetime.datetime.now().strftime("%Y-%m")
        total_files, total_pages = 0, 0
        # Prefer the Excel register if present
        xp = os.path.join(self._opts.get("save_folder", "."), "register.xlsx")
        counted = False
        if HAS_XLSX and os.path.exists(xp):
            try:
                wb = openpyxl.load_workbook(xp); ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and row[0] and str(row[0]).startswith(month):
                        total_files += 1
                        try:
                            total_pages += int(row[3])
                        except Exception:
                            pass
                counted = True
            except Exception:
                counted = False
        if not counted:
            root = self._opts.get("save_folder", os.path.expanduser("~"))
            try:
                for dp, _dn, fn in os.walk(root):
                    for f in fn:
                        if f.lower().endswith(".pdf"):
                            p = os.path.join(dp, f)
                            m = datetime.datetime.fromtimestamp(os.path.getmtime(p)).strftime("%Y-%m")
                            if m == month:
                                total_files += 1
            except Exception:
                pass
        QtWidgets.QMessageBox.information(
            self, "Monthly report (%s)" % month,
            "Is mahine:\n\nKul PDF: %d\nKul pages: %s"
            % (total_files, total_pages if total_pages else "—"))

    def _draw_fit(self, painter, path, target):
        """Draw the image at `path` scaled to fit inside `target` rect, centered,
        keeping aspect ratio."""
        img = QtGui.QImage(path)
        if img.isNull():
            return
        size = img.size()
        size.scale(target.size(), QtCore.Qt.KeepAspectRatio)
        x = target.x() + (target.width() - size.width()) // 2
        y = target.y() + (target.height() - size.height()) // 2
        painter.drawImage(QtCore.QRect(x, y, size.width(), size.height()), img)

    def _do_print(self, paths, per_page=1):
        if not paths:
            self._warn("Scan or import a page first (or select one)."); return
        printer = QPrinter(QPrinter.HighResolution)
        if QPrintDialog(printer, self).exec_() != QtWidgets.QDialog.Accepted:
            return
        painter = QtGui.QPainter()
        if not painter.begin(printer):
            self._warn("Printer did not start."); return
        try:
            page = painter.viewport()
            if per_page <= 1:
                for i, p in enumerate(paths):
                    if i > 0:
                        printer.newPage()
                    self._draw_fit(painter, p, page)
            else:
                # per_page images stacked on ONE sheet (top/bottom halves for 2-up).
                slot_h = page.height() // per_page
                for i, p in enumerate(paths):
                    slot = i % per_page
                    if i > 0 and slot == 0:
                        printer.newPage()
                    # small inset so the two IDs aren't glued together
                    pad = int(slot_h * 0.04)
                    target = QtCore.QRect(page.x(), page.y() + slot * slot_h + pad,
                                          page.width(), slot_h - 2 * pad)
                    self._draw_fit(painter, p, target)
        finally:
            painter.end()
        self._pstats_bump(prints=1)          # personal
        self._an_report("event", prt=1)      # worldwide + refresh
        self._an_update_box()
        self.status.showMessage("Sent to printer.", 4000)

    def print_all(self):
        self._do_print(self._ordered_paths(), per_page=1)

    def print_selected(self):
        paths = self._selected_paths()
        if not paths:
            self._warn("Select some pages from the thumbnails first."); return
        self._do_print(paths, per_page=1)

    def print_ids(self):
        # ID print: 2 IDs/pages per A4 sheet. Uses selected pages if any, else all.
        paths = self._selected_paths() or self._ordered_paths()
        self._do_print(paths, per_page=2)

    def print_ids_selected(self):
        # ID print of ONLY the selected pages (2 per A4 sheet).
        paths = self._selected_paths()
        if not paths:
            self._warn("Select ID pages from the thumbnails first."); return
        self._do_print(paths, per_page=2)

    # kept for the Ctrl+P shortcut / File menu (prints all)
    def print_pages(self):
        self.print_all()

    def create_shortcut(self):
        prof = self._selected_profile()
        if prof is None:
            self._warn("Select a profile first."); return
        script = os.path.abspath(sys.argv[0])
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        bat = os.path.join(desktop, "Scan - %s.bat" % prof.get("name", "Profile"))
        try:
            with open(bat, "w", encoding="utf-8") as fh:
                fh.write("@echo off\r\n")
                fh.write('py -3.12-32 "%s" --scan --profile "%s"\r\n' % (script, prof.get("name", "Profile")))
            QtWidgets.QMessageBox.information(self, "Done",
                "Desktop shortcut created:\n%s\n\nDouble-click it to scan directly." % bat)
        except Exception as exc:
            self._warn("Shortcut failed:\n%s" % exc)

    def _update_status(self):
        self._update_footer()

    def _update_footer(self):
        """Smart footer — sab ek nazar me. (kisi bhi hisse me error aaye to
        baaki footer chalta rahe.)"""
        L = self.L
        # folder
        try:
            root = self._files_root()
            self.foot_folder.setText("📁 " + (os.path.basename(root.rstrip("/\\")) or root))
        except Exception:
            pass
        # pages + selection + estimated size
        try:
            n = self.list.count()
            sel = len(self.list.selectedItems())
            total = 0
            for i in range(n):
                try:
                    total += os.path.getsize(self.list.item(i).data(QtCore.Qt.UserRole))
                except Exception:
                    pass
            mb = total / 1048576.0
            szt = ("~%.0f KB" % (total / 1024.0)) if mb < 1 else ("~%.1f MB" % mb)
            seltxt = (" · <b>%d selected</b>" % sel) if sel else ""
            self.foot_pages.setText("📄 %d pages%s · %s" % (n, seltxt, szt) if n else
                                    L("📄 koi page nahi", "📄 no pages"))
        except Exception:
            pass
        # last saved
        try:
            if self._recent:
                self.foot_last.setText(" · ✔ " + os.path.basename(self._recent[0]))
            else:
                self.foot_last.setText("")
        except Exception:
            pass
        # (Analytics hata diya gaya — 'aaj/streak' footer nahi)
        # disk space
        try:
            root = self._files_root()
            free = shutil.disk_usage(root).free / (1024.0 ** 3)
            self.foot_disk.setText("💽 %.0f GB" % free)
        except Exception:
            pass
        # profile flags -> transient message (hover-tip friendly)
        try:
            prof = self._selected_profile()
            pname = prof.get("name") if prof else tr("none_profile", self._lang)
            self.foot_folder.setToolTip(
                L("Save folder: %s\nProfile: %s — click karke folder kholo",
                  "Save folder: %s\nProfile: %s — click to open the folder")
                % (root if 'root' in dir() else "", pname))
        except Exception:
            pass
        # v74 footer extra: ghadi/date · aaj ke scan · online · apna sandesh
        try:
            if self.foot_clock.isVisible():
                now = QtCore.QDateTime.currentDateTime()
                self.foot_clock.setText("🕒 " + now.toString("ddd d MMM  hh:mm"))
        except Exception:
            pass
        try:
            if self.foot_today.isVisible():
                t = 0
                try:
                    t = int((self._an_world or {}).get("today", 0))
                except Exception:
                    t = 0
                self.foot_today.setText(L("📈 Aaj: %d", "📈 Today: %d") % t)
        except Exception:
            pass
        try:
            if self.foot_online.isVisible():
                on = 0
                try:
                    on = int((self._an_world or {}).get("online", 0))
                except Exception:
                    on = 0
                self.foot_online.setText(L("🌍 %d online", "🌍 %d online") % on)
        except Exception:
            pass
        try:
            m = str(self._opts.get("footer_msg", "") or "").strip()
            self.foot_msg.setText(("📌 " + m) if m else "")
            self.foot_msg.setVisible(bool(m))
        except Exception:
            pass

    def _warn(self, msg):
        QtWidgets.QMessageBox.warning(self, APP_NAME, msg)

    def closeEvent(self, event):
        if self.list.count() > 0 and getattr(self, "_dirty", False):
            r = QtWidgets.QMessageBox.question(
                self, APP_NAME,
                'There are unsaved pages. Close anyway?' if self._lang == "hi"
                else "You have unsaved pages. Close anyway?",
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No, QtWidgets.QMessageBox.No)
            if r != QtWidgets.QMessageBox.Yes:
                event.ignore(); return
        elif self._opts.get("confirm_exit"):
            r = QtWidgets.QMessageBox.question(
                self, APP_NAME,
                self.L("Kya pakka software band karna hai?",
                       "Are you sure you want to close?"),
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No, QtWidgets.QMessageBox.No)
            if r != QtWidgets.QMessageBox.Yes:
                event.ignore(); return
        # Window ka size/jagah yaad rakho (agli baar wahi khule)
        if self._opts.get("remember_window", True):
            try:
                self._opts["win_geometry"] = bytes(
                    self.saveGeometry().toBase64()).decode("ascii")
                self._save_opts()
            except Exception:
                pass
        try:
            self._conn_timer.stop()
        except Exception:
            pass
        try:
            for f in os.listdir(self._tmpdir):
                try:
                    os.remove(os.path.join(self._tmpdir, f))
                except Exception:
                    pass
            os.rmdir(self._tmpdir)
        except Exception:
            pass
        super().closeEvent(event)


def main():
    auto_profile = None
    if "--scan" in sys.argv and "--profile" in sys.argv:
        try:
            auto_profile = sys.argv[sys.argv.index("--profile") + 1]
        except Exception:
            auto_profile = None
    try:
        import ctypes
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("NobleCare.ApneScan")
    except Exception:
        pass
    # CRASH REPORTER: koi bhi anhoni error file me save ho jaati hai taaki
    # report bhej kar agli update me fix ho sake.
    def _crash_hook(t, v, tb):
        try:
            with open(CRASH_PATH, "a", encoding="utf-8") as fh:
                fh.write("\n\n==== %s | ApneScan v%s ====\n" % (datetime.datetime.now(), VERSION))
                fh.write("".join(traceback.format_exception(t, v, tb)))
        except Exception:
            pass
        try:
            QtWidgets.QMessageBox.critical(
                None, "ApneScan — error",
                'An error occurred. A report was saved here:\n%s\n\nPlease send this file as feedback — it will be fixed in the next update.' % CRASH_PATH)
        except Exception:
            pass
    sys.excepthook = _crash_hook

    app = QtWidgets.QApplication(sys.argv)
    app.setApplicationName(APP_NAME)
    app.setWindowIcon(app_icon())
    win = ScannerWindow(auto_scan_profile=auto_profile)
    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
