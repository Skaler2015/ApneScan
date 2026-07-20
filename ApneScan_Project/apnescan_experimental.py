# -*- coding: utf-8 -*-
"""
ApneScan - v15  [experimental: TWAIN file-transfer for continuous ADF feed; native fallback]
=========================================================
Run with 32-bit Python (HP TWAIN driver is 32-bit):
    py -3.12-32 scanner_app_v10.py
One-click scan shortcut:
    py -3.12-32 scanner_app_v10.py --scan --profile "Documents"
"""

import io
import os
import re
import sys
import json
import shutil
import socket
import tempfile
import traceback
import datetime

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog

try:
    import twain
    HAS_TWAIN = True
except Exception:
    HAS_TWAIN = False

from PIL import Image, ImageChops, ImageEnhance, ImageOps, ImageDraw, ImageFont, ImageFilter

try:
    import numpy as np
    HAS_NUMPY = True
except Exception:
    HAS_NUMPY = False

try:
    import pytesseract
    from pypdf import PdfReader, PdfWriter
    HAS_OCR_LIBS = True
except Exception:
    HAS_OCR_LIBS = False

try:
    import fitz  # PyMuPDF — best PDF page rendering (optional)
    HAS_FITZ = True
except Exception:
    HAS_FITZ = False


def pdf_to_images(pdf_path, tmpdir, dpi=200):
    """Turn each page of a PDF into a PNG (for import). Uses PyMuPDF if installed
    (renders any PDF), else falls back to extracting embedded images with pypdf
    (works for scanned PDFs). Returns a list of PNG paths."""
    out = []
    if HAS_FITZ:
        try:
            doc = fitz.open(pdf_path)
            zoom = dpi / 72.0
            mat = fitz.Matrix(zoom, zoom)
            for page in doc:
                pix = page.get_pixmap(matrix=mat)
                fd, png = tempfile.mkstemp(suffix=".png", dir=tmpdir); os.close(fd)
                pix.save(png)
                out.append(png)
            doc.close()
            if out:
                return out
        except Exception:
            out = []
    # fallback: pull the biggest embedded image from each page (scanned PDFs)
    try:
        from pypdf import PdfReader as _PR
        reader = _PR(pdf_path)
        for page in reader.pages:
            try:
                imgs = list(getattr(page, "images", []))
            except Exception:
                imgs = []
            if not imgs:
                continue
            big = max(imgs, key=lambda im: len(getattr(im, "data", b"")))
            fd, png = tempfile.mkstemp(suffix=".png", dir=tmpdir); os.close(fd)
            with open(png, "wb") as fh:
                fh.write(big.data)
            try:
                with Image.open(png) as im2:
                    im2.convert("RGB").save(png, "PNG")
            except Exception:
                pass
            out.append(png)
    except Exception:
        pass
    return out


def _configure_tesseract():
    """Point pytesseract at a Tesseract install even if it's not on PATH."""
    if not HAS_OCR_LIBS:
        return
    try:
        home = os.path.expanduser("~")
        for c in (r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                  r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
                  os.path.join(home, "AppData", "Local", "Tesseract-OCR", "tesseract.exe"),
                  os.path.join(home, "AppData", "Local", "Programs", "Tesseract-OCR", "tesseract.exe")):
            if os.path.exists(c):
                pytesseract.pytesseract.tesseract_cmd = c
                break
    except Exception:
        pass


def tesseract_available():
    if not HAS_OCR_LIBS:
        return False
    try:
        pytesseract.get_tesseract_version()
        return True
    except Exception:
        return False


_configure_tesseract()

try:
    import openpyxl
    HAS_XLSX = True
except Exception:
    HAS_XLSX = False

try:
    from pyzbar.pyzbar import decode as _zbar_decode
    HAS_ZBAR = True
except Exception:
    HAS_ZBAR = False

try:
    import win32com.client as _w32
    import pythoncom as _pythoncom
    HAS_W32 = True
except Exception:
    HAS_W32 = False


APP_NAME = "ApneScan"
VERSION = "1.0"
CONFIG_PATH = os.path.join(os.path.expanduser("~"), ".apnescan.json")
_OLD_CONFIG = os.path.join(os.path.expanduser("~"), ".noble_doc_scanner.json")
if not os.path.exists(CONFIG_PATH) and os.path.exists(_OLD_CONFIG):
    try:
        import shutil as _sh
        _sh.copy2(_OLD_CONFIG, CONFIG_PATH)
    except Exception:
        pass
SCANNER_PORTS = (80, 443, 8080, 9100)

# Built-in ApneScan worldwide-stats endpoint (Google Apps Script). Every install
# reports scan COUNTS here (never any document/patient data).
DEFAULT_STATS_URL = "https://script.google.com/macros/s/AKfycbzwh2HQHXKe_09wgXpSh-NZQu-GbR7N-NVmtUEyRpl8oOa-MdoJ9ShqHB_i0QCcuqe2_w/exec"

_TESS_GUESSES = [
    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
]
if HAS_OCR_LIBS:
    for _p in _TESS_GUESSES:
        if os.path.exists(_p):
            pytesseract.pytesseract.tesseract_cmd = _p
            break

def resource_path(name):
    """Path to a bundled data file, works both from source and PyInstaller .exe."""
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(sys.argv[0])))
    p = os.path.join(base, name)
    if os.path.exists(p):
        return p
    return os.path.join(os.path.dirname(os.path.abspath(sys.argv[0])), name)


# Embedded app icon (base64 PNG) so the ApneScan icon always shows in the title
# bar + taskbar even if the icon file wasn't bundled with the build.
_EMBEDDED_ICON_B64 = """\
iVBORw0KGgoAAAANSUhEUgAAAgAAAAIACAYAAAD0eNT6AAAQoklEQVR4nO3dS48c5d3G4Rqrl0gzlpA/iXdss3bWwQokMZFickAG
EnbhsCOBOOTEJuEYJ+v4O7DjkyAktyX2k4XduMeemT5V1XO4r0tqvYr0ClU/7q7/r57q6T4aqMrx735zWvoYAKbw8A9/OSp9DDzh
H2Nmx7/9tQEPcI6Hf/yrmTQjiz0hwx7gMKJgOhZ2RAY+wLQEwXgs5AEMfICyBMH+LNyODH2AOomB3VisLRj6AG0RA5tZoEsY/ABt
EwIXszBPMfQB+iQGzrIYjxn8ABmEwCPxi2DwA2RKD4HYJ2/wAzAMuSEQ96QNfgDOkxYCMU/W4AdgGykh0P2TNPgB2EfvIdDtkzP4
ARhDryFwpfQBTMHwB2Asvc6Urqqm138kAOrQ025AF0/E4AdgTj2EQPO3AAx/AObWw+xpOgB6+AcAoE2tz6AmtzBaX3QA+tLiLYHm
Dvj4TcMfgPo8/KCtCGjqFoDhD0CtWptRTdRKa4sKQLYWdgOqP8DjN39l+APQnIcf/K3qGVv1LQDDH4BW1T7Dqg2A2hcOADapeZZV
GQA1LxgA7KLWmVZdANS6UACwrxpnWzUfUKhxcQBgbLV8OLCKHQDDH4AUtcy84gFQy0IAwFxqmH1FA6CGBQCAEkrPwGIBUPqJA0Bp
JWdh8VsAAMD8igSAq38AeKTUTJw9AAx/ADirxGycNQAMfwA439wzcrYAMPwB4HJzzspZAsDwB4DtzDUzJw8Awx8AdjPH7PRngAAQ
aNIfJDh+w9U/AOzr4YfT/XDQZDsAhj8AHGbKWTpJABj+ADCOqWaqzwAAQKDRA8DVPwCMa4rZOmoAGP4AMI2xZ+xony48fuOXhj8A
TOzhh38fZXb7DAAABBolAFz9A8A8xpq5dgAAINDBAeDqHwDmNcbsPSgADH8AKOPQGewWAAAE2jsAXP0DQFmHzGI7AAAQaK8AcPUP
AHXYdybbAQCAQDsHgKt/AKjLPrPZDgAABNopAFz9A0Cddp3RdgAAINDWAeDqHwDqtsusXmz9XzX+AaAbW+0AHL/u6h8AWrDtzPYZ
AAAItDEAXP0DQFu2md12AAAgkAAAgECXBoDtfwBo06YZbgcAAAJdGACu/gGgbZfN8ku+CMj8B4BeuQUAAIHODYDj1191+Q8AHbho
ptsBAIBAAgAAAj0TALb/AaAv5812OwAAEEgAAEAgAQAAgc4EgPv/ANCnp2e8HQAACCQAACDQ2d8CcAMAACJ8vwNwfMf9fwDo2fqs
dwsAAAIJAAAIJAAAIJAAAIBAV4bBBwABIMVq5tsBAIBAAgAAAgkAAAgkAAAgkAAAgECPfwvAHwEAQBI7AAAQ6Oj4zm2X/wAQxg4A
AAQSAAAQSAAAQCABAACBBAAABBIAABBo4TuAACCPHQAACCQAACCQAACAQAIAAAIJAAAIJAAAIJAAAIBAAgAAAgkAAAgkAAAgkAAA
gEACAAACLUofAKxb3v1H6UOgIV8/+Kb0IczmB++9U/oQ6IwAoDhDH2B+i8HvAVPI8u7HpQ8BGuJczbh8BoAiDH/GcP3qtdKHAM0S
AMzO8GdMIgD2IwCYleHPFEQA7E4AMBvDnymJANiNAGAWhj9zEAGwPQHA5Ax/5iQCYDsLf1nClJZ/NvyZ3/Wr1/r7kiDnakZmBwDo
kp0AuJwAYDKu/ilNBMDFBADQNREA5xMAQPdEADxLADAJ2//URgTAWQIAiCEC4AkBAEQRAfCIAADiiAAQAEAoEUA6AQDEEgEkEwBA
NBFAKgEAxBMBJBIAAIMIII8AAHhMBJBEAACsEQGkEAAATxEBJBAAAOcQAfRuMQynpY8BoErXr14bvn7wTenDeMy5mnHZAQC4hJ0A
eiUAADYQAfRoYVcJYLPitwOcqxmZHQCALdkJoCcCAGAHIoBeCACAHYkAeiAAAPYgAmidAADYkwigZQIA4AAigFYJAIADiQBaJAAA
RiACaI0AABiJCKAlAgBgRCKAVggAgJGJAFogAAAmIAKonQAAmIgIoGYCAGBCIoBaCQCAiYkAaiQAAGYgAqiNAACYiQigJgIAYEYi
gFoIAICZiQBqIAAAChABlLYYTk9LHwNApOtXrw1fP/hmu/9n52pGZgcAoCA7AZQiAAAKEwGUIAAAKiACmJsAAKiECGBOAgCgIiKA
uQgAgMqIAOYgAAAqJAKY2qL0AQBwPhHAlOwAAEAgAQAAgQQAAAQSAAAQSAAAQCABAACBBAAABBIAABBIAABAIAEAAIEEAAAEEgAA
EEgAAECgxXBa+hAA2Mi5mpHZAQCAQAIAAAIt7CsBtMC5mnHZAQCAQAIAAAIJAAAItCh9AFDKyVuv3Sx9DLRh+f5H90ofA4xNABDH
4GdXq9eMEKAnbgEQxfDnEF4/9EQAEMPJmzF4HdELAUAEJ23G5PVEDwQAAAQSAHTP1RpT8LqidQIAAAIJAAAIJAAAIJAAAIBAAoDu
+fY2puB1ResEAAAEEgBEcLXGmLye6MGV4XQYPDxGf1TISZsxFHsdlX5Pe3T3sANAFBHAIbx+6ImfAybO6iTum9zYlsFPjwQAsZzU
gWRuAQBAIAEAAIEWjz4OCEDdnKsZlx0AAAgkAAAgkAAAgEACAAACCQAACCQAACCQAACAQAIAAAIJAAAIJAAAIJAAAIBAAgAAAgkA
AAgkAAAg0MIvTAI0wLmakdkBAIBAAgAAAgkAAAgkAAAg0KL0AUApJ2+9drP0MdCG5fsf3St9DDA2AUAcg59drV4zQoCeuAVAFMOf
Q3j90BMBQAwnb8bgdUQvrjz6dgkPj7EfdXHSZkxlXk+l39MevT3sAABAIAFA91z9MwWvK1onAAAgkAAAgEACAAACHR2/euu09EFQ
hx++c+fb0scATO9/79x9vvQxUJ4ACGfoQzYxkMstgGCGP+A8kMsOQCBveOA8dgOy2AEIY/gDF3F+yCIAgnhzA5s4T+RYHJU+AgCq
Yi5kODq57TMACW68q+qB7d1/2+cBeucWQADDH9iV80b/FqUPgPm88NyJogc2+uq7peEfwA5ACMMf2JbzRQYB0Lkb79751psZ2NUL
z5087zZA3wRA5wx/YF/OH30TAAAQSAAAQCABAACBBAAABBIAnfP3vMC+nD/6JgA6d//tu897EwO7+uq75be+Drhvi2HwUwAJvvpu
6fsAgK08uWgwH3rmq4CD2AkAYMUtgAD33/6TK39gJ84b/RMAIbyZgW05X2QQAAAQSAAEUfXAJs4TOQRAGG9u4CLOD1mOTm7/zN95
hLrx7uv+KgAw+EPZAQjmTQ84D+Q6OvmFHQAeufGeHQFIcP/3hj4CgIk8+PhfpQ8BunL19q3Sh0BnfBMgsU7eeu1m6WPoyfL9j+6V
PgZgewKAOAb/NFbrKgSgDT4ESBTDf3rWGNogAIhhMM3HWkP9BAARDKT5WXOomwAAgEACgO65Ei3H2kO9BAAABBIAABBIAABAIAEA
AIEEAN3zzXTlWHuolwAAgEBXhuF08PAY/1EXV6Lzs+ZjK/2e9ujtYQeAGAbSfKw11E8AEMVgmp41hjb4OWDirAaUb6kbl8EPbREA
xDKwgGRuAQBAoMVwWvoQANjIuZqR2QEAgEACAAACCQAACCQAACCQAACAQAIAAAIJAAAIJAAAIJAAAIBAAgAAAgkAAAgkAAAgkAAA
gEACAAACCQAACCQAACCQAACAQIthOC19DABs5FzNuOwAAEAgAQAAgQQAAARauK0E0ADnakZmBwAAAgkAAAgkAAAgkAAAgEACAAAC
CQAACCQAACCQAACAQAIAAAIJAAAIJAAAIJAAAIBAAgAAAgkAAAgkAAAgkAAAgEACAAACCQAACCQAACCQAACAQIvh9LT0MQCwiXM1
I7MDAACBBAAABBIAABBIAABAIAEAAIEEAAAEEgAAEEgAAEAgAQAAgQQAAAQSAAAQSAAAQCABAACBBAAABBIAABBIAABAIAEAAIEE
AAAEEgAAEEgAAECgxXBa+hAA2Mi5mpHZAQCAQAIAAAIt7CsBtMC5mnHZAQCAQAIAAAIJAAAIJAAAIJAAAIBAAgAAAgkAAAgkAAAg
kAAAgEACAAACCQAACCQAACCQAGASV3/+culDgG54PzEFAQAAgQQAAAQSAEzGtiUczvuIqQgAAAgkAJiUqxfYn/cPUzo6ufXSaemD
oH8P/vl56UOAplx9xfBnWnYAmIWTGWzP+4U5CABm46QGm3mfMBcBwKyc3OBi3h/MSQAwOyc5eJb3BXMTABThZAdPeD9QwtHJrR/7
KwCKevDPL0ofAhRx9ZWXSh8CwQQAVRED9M7QpxYCAAAC+QwAAAQSAAAQSAAAQCABAACBBAAABBIAABBIAABAIAEAAIEEAAAEEgAA
EEgAAEAgAQAAgRaDnwICgDh2AAAgkAAAgEACAAACCQAACCQAACDQleUnXx6VPggAYD7LT748sgMAAIEEAAAEWjz6P74NCACS2AEA
gEACAAACCQAACCQAACDQlWEYhuUn//ZdAAAQYDXz7QAAQCABAACBBAAABBIAABDo+wDwQUAA6Nv6rLcDAACBFmf+l58EAIAIdgAA
INCZAFh+6nMAANCjp2e8HQAACCQAACCQAACAQM8EgM8BAEBfzpvtdgAAIJAAAIBA5waA2wAA0IeLZrodAAAIJAAAINClW/0nP73p
1wEAoFHLT+9dOOftAABAoEsD4LJyAADqtWmG2wEAgEACAAACbQwAtwEAoC3bzG47AAAQaKsAsAsAAG3YdmbbAQCAQFsHgF0AAKjb
LrN6sdN/2fcCAkAXdroFsPzMLgAA1GjXGe0zAAAQaOcAsAsAAHXZZzbbAQCAQHsFgF0AAKjDvjPZDgAABNo7AOwCAEBZh8xiOwAA
EOigALALAABlHDqDD94BEAEAMK8xZq9bAAAQaJQAsAsAAPMYa+baAQCAQKMFgF0AAJjWmLN29KF98pMX/WgwAIxs+dl/Rp3Zo98C
GPsAASDdFLPVZwAAINAkAWAXAADGMdVMnWwHQAQAwGGmnKWT3gIQAQCwn6ln6GLK//gwDMPgbwIAoDqTfwhw+bldAADYxRyzc5a/
AhABALCduWbmbH8GKAIA4HJzzspZvwdABADA+eaekbN/EZAIAICzSszGIt8EKAIA4JFSM9FXAQNAoGIBYBcAgHQlZ2HRHQARAECq
0jOw+C2A0gsAAHOrYfYVD4BhqGMhAGAOtcy8Kg5i3cnLL/r1AAC6U8vgX6liB2BdbQsEAIeqcbZVFwDDUOdCAcA+ap1pVQbAMNS7
YACwrZpnWbUBMAx1LxwAXKb2GVZ1AAxD/QsIAE9rYXZVf4DrTl7+kb8QAKBay8//28xcrX4HYF1LCwtAltZmVFMBMAztLTAA/Wtx
NjV3wOtOXnJLAIByll+0N/hXmtsBWNfywgPQttZnUNMBMAzt/wMA0J4eZk/zT2CdWwIATKmHwb/SzRNZJwQAGFNPg3+l+VsA5+nx
HwqAMnqdKV0+qXV2AwDYR6+Df6XrJ7dOCACwjd4H/0rEk1wnBAA4T8rgX4l6suuEAADDkDf4VyKf9DohAJApdfCvRD/5dUIAIEP6
4F+xCE8RAgB9MvjPshiXEAMAbTP0L2ZhtiAEANpi8G9mgXYkBgDqZOjvxmIdQAwAlGXo78/CjUgQAEzLwB+PhZyQIAA4jIE/HQs7
M1EAcD7Dfl4WuzICAeiVAV+X/wMZzgP3ZyE7mwAAAABJRU5ErkJggg==
"""


def app_icon():
    for n in ("apnescan.ico", "apnescan_icon.png"):
        p = resource_path(n)
        if os.path.exists(p):
            return QtGui.QIcon(p)
    # fallback: decode the embedded icon
    try:
        import base64 as _b64
        pm = QtGui.QPixmap()
        pm.loadFromData(_b64.b64decode(_EMBEDDED_ICON_B64))
        if not pm.isNull():
            return QtGui.QIcon(pm)
    except Exception:
        pass
    return QtGui.QIcon()


COLOUR_MODES = {"Colour (rang)": "color", "Grayscale": "gray", "Black & White": "bw"}
RESOLUTIONS = ["150", "200", "300", "600"]

# (id, label, default key) — all reassignable via Settings -> Keyboard Shortcuts
SHORTCUTS = [
    ("scan", "Scan", "Return"),
    ("import", "Import images / PDF", "Ctrl+O"),
    ("rename", "Rename page", "F2"),
    ("save_all", "Save all as PDF", "Ctrl+S"),
    ("save_sel", "Save selected as PDF", "Space"),
    ("save_pw", "Save PDF (password)", ""),
    ("save_img", "Save images", "Ctrl+Shift+S"),
    ("ocr_text", "Export OCR text", "Ctrl+Alt+O"),
    ("print", "Print", "Ctrl+P"),
    ("rotate_left", "Rotate left", "Ctrl+Left"),
    ("rotate_right", "Rotate right", "Ctrl+Right"),
    ("bright_up", "Brightness +", ""),
    ("bright_dn", "Brightness -", ""),
    ("contrast_up", "Contrast +", ""),
    ("contrast_dn", "Contrast -", ""),
    ("autocrop", "Auto-crop page", ""),
    ("autoname", "Auto-name pages (OCR)", "Ctrl+Alt+N"),
    ("undo", "Undo delete", "Ctrl+Z"),
    ("delete", "Delete page", "Delete"),
    ("clear", "Clear all", "Ctrl+Shift+Delete"),
    ("move_up", "Move page up", "Ctrl+Up"),
    ("move_down", "Move page down", "Ctrl+Down"),
    ("search", "Search past PDFs", "Ctrl+F"),
    ("merge", "Merge PDFs", ""),
    ("split", "Split into PDFs", ""),
    ("report", "Monthly report", ""),
    ("zoom_in", "Zoom in (thumbnails)", "Ctrl+="),
    ("zoom_out", "Zoom out (thumbnails)", "Ctrl+-"),
    ("options", "Options / Settings", "Ctrl+,"),
    ("profiles", "Profiles", "Ctrl+L"),
]

DEFAULT_OPTIONS = {
    "auto_save": False,
    "save_folder": os.path.join(os.path.expanduser("~"), "Documents", "NobleScans"),
    "filename_template": "{claim}_{date}_{seq}",
    "make_claim_folder": False,
    "year_month_folders": False,
    "remove_blank": False,
    "blank_sensitivity": "normal",   # kam / normal / zyada
    "auto_crop": False,
    "deskew": False,
    "quality_enhance": False,
    "compress": False,
    "jpeg_quality": 60,
    "watermark": False,
    "watermark_text": "Noble Care Hospital",
    "batch_mode": False,
    "validate_claim": False,
    "claim_pattern": r"^[A-Za-z0-9\-]{4,}$",
    "barcode_autofill": False,
    "duplicate_check": False,
    "excel_log": False,
    "activity_log": False,
    "backup": False,
    "backup_folder": os.path.join(os.path.expanduser("~"), "Documents", "NobleScans_Backup"),
    "scanner_method": "twain",   # "twain" ya "wia"
    "wia_device_id": None,
    "language": "hi",            # "hi" ya "en"
    "simple_mode": False,
    "feedback_email": "",
    "fast_mode": False,
    "after_save": "nothing",   # nothing / open / folder
    "naps2_path": "",
    "naps2_profile": "",
    "scanner_ip": "",
    "theme": "light",             # light / dark
    "show_page_numbers": True,
    "save_images_too": False,     # also save each page as an image
    "auto_name": False,           # OCR the document title as the page label
    "shortcuts": {},              # id -> custom key (overrides default)
}


def load_config():
    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_config(cfg):
    try:
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(cfg, f)
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Image helpers
# ---------------------------------------------------------------------------

def is_blank_page(img, dark_fraction_threshold=0.0008, dark_level=75):
    # Count only VERY DARK ink (< dark_level, default 75). Real printed text / stamps
    # are near-black; fold-line shadows and faint bleed-through from the front are
    # medium/light grey (>75) and are therefore IGNORED. So a blank back side (only
    # creases + bleed-through) reads ~0% and gets dropped, while a page with real
    # (even sparse) dark text is kept. No erosion (that was thinning real text).
    try:
        g = img.convert("L")
        ink = g.point(lambda p: 255 if p < dark_level else 0)
        hist = ink.histogram()
        total = (g.width * g.height) or 1
        return (hist[-1] / total) < dark_fraction_threshold
    except Exception:
        return False


def whiten_dark_background(img, dark_pix=90, dark_frac=0.70):
    """Paint the scanner's DARK backing bands (black/navy) at the EDGES of a scan
    (top/bottom/left/right — e.g. around an ECG strip) to WHITE. A row/column at the
    edge is treated as backing if MORE THAN dark_frac of its pixels are darker than
    dark_pix (so uniform navy/black backing qualifies, but grey/white paper and text
    do not). Only edge-contiguous bands are cleared; interior content is never
    touched, and grey/low-contrast documents are left intact."""
    try:
        import numpy as _np
        g = _np.asarray(img.convert("L"))
        # fast early-out: if all four edges are already bright, there's no dark
        # backing to remove — skip the work (keeps normal pages superfast).
        if (g[0].mean() > 150 and g[-1].mean() > 150
                and g[:, 0].mean() > 150 and g[:, -1].mean() > 150):
            return img
        rgb = img.convert("RGB")
        arr = _np.asarray(rgb).copy()
        darkish = g < dark_pix
        col_back = darkish.mean(axis=0) > dark_frac
        row_back = darkish.mean(axis=1) > dark_frac
        H, W = g.shape
        l = 0
        while l < W and col_back[l]:
            l += 1
        r = W - 1
        while r >= 0 and col_back[r]:
            r -= 1
        t = 0
        while t < H and row_back[t]:
            t += 1
        b = H - 1
        while b >= 0 and row_back[b]:
            b -= 1
        if l > 0:
            arr[:, :l] = 255
        if r < W - 1:
            arr[:, r + 1:] = 255
        if t > 0:
            arr[:t, :] = 255
        if b < H - 1:
            arr[b + 1:, :] = 255
        return Image.fromarray(arr.astype("uint8"))
    except Exception:
        return img


def trim_dark_borders(img, bright_thresh=120, min_paper_frac=0.04, pad=6):
    """Remove wide DARK margins (the scanner's backing — black OR dark-blue —
    showing around a narrow sheet like an ECG strip) so the page comes out to its
    real paper. A row/column is "paper" if enough of its pixels are BRIGHT
    (paper background), so dark backing is trimmed. Safe: only crops when a real
    dark border exists; leaves normal white pages and fully-dark pages untouched."""
    try:
        import numpy as _np
        g = img.convert("L")
        a = _np.asarray(g)
        bright = a > bright_thresh                       # paper background is bright
        row_has = bright.mean(axis=1) > min_paper_frac
        col_has = bright.mean(axis=0) > min_paper_frac
        rows = _np.where(row_has)[0]
        cols = _np.where(col_has)[0]
        if len(rows) == 0 or len(cols) == 0:
            return img                                   # no bright paper found
        t, b = int(rows[0]), int(rows[-1])
        l, r = int(cols[0]), int(cols[-1])
        t = max(0, t - pad); l = max(0, l - pad)
        b = min(a.shape[0] - 1, b + pad); r = min(a.shape[1] - 1, r + pad)
        # only crop if it removes a meaningful border (>1.5% on some side)
        if (r - l) < img.width * 0.985 or (b - t) < img.height * 0.985:
            return img.crop((l, t, r + 1, b + 1))
        return img
    except Exception:
        return img


def autocrop(img, border=12):
    try:
        rgb = img.convert("RGB")
        bg = Image.new("RGB", rgb.size, (255, 255, 255))
        bbox = ImageChops.difference(rgb, bg).getbbox()
        if not bbox:
            return img
        l, t, r, b = bbox
        l = max(0, l - border); t = max(0, t - border)
        r = min(img.width, r + border); b = min(img.height, b + border)
        return img.crop((l, t, r, b))
    except Exception:
        return img


def deskew(img):
    if not HAS_NUMPY:
        return img
    try:
        g = img.convert("L")
        small = g.resize((max(1, g.width // 4), max(1, g.height // 4)))
        best_angle, best_score = 0.0, -1.0
        for i in range(-10, 11):
            angle = i * 0.5
            rot = small.rotate(angle, resample=Image.BILINEAR, fillcolor=255)
            arr = np.asarray(rot, dtype=np.float32)
            score = float(np.var(np.diff(arr.sum(axis=1))))
            if score > best_score:
                best_score, best_angle = score, angle
        if abs(best_angle) < 0.25:
            return img
        return img.rotate(best_angle, resample=Image.BICUBIC, fillcolor=(255, 255, 255), expand=True)
    except Exception:
        return img


def auto_enhance(img):
    """Clean up faded / dull documents automatically."""
    try:
        rgb = img.convert("RGB")
        rgb = ImageOps.autocontrast(rgb, cutoff=1)
        rgb = ImageEnhance.Sharpness(rgb).enhance(1.4)
        rgb = ImageEnhance.Contrast(rgb).enhance(1.06)
        return rgb
    except Exception:
        return img


def apply_watermark(img, text):
    try:
        base = img.convert("RGBA")
        overlay = Image.new("RGBA", base.size, (255, 255, 255, 0))
        draw = ImageDraw.Draw(overlay)
        size = max(14, base.width // 45)
        try:
            font = ImageFont.truetype("arial.ttf", size)
        except Exception:
            font = ImageFont.load_default()
        stamp = "%s  |  %s" % (text, datetime.datetime.now().strftime("%d-%m-%Y"))
        try:
            bbox = draw.textbbox((0, 0), stamp, font=font)
            tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
        except Exception:
            tw, th = draw.textsize(stamp, font=font)
        x = (base.width - tw) // 2
        y = base.height - th - max(10, base.height // 60)
        draw.rectangle([x - 8, y - 6, x + tw + 8, y + th + 6], fill=(255, 255, 255, 160))
        draw.text((x, y), stamp, fill=(120, 120, 120, 200), font=font)
        return Image.alpha_composite(base, overlay).convert("RGB")
    except Exception:
        return img


def enhance_image(path, brightness=1.0, contrast=1.0):
    try:
        with Image.open(path) as im:
            im = im.convert("RGB")
            if brightness != 1.0:
                im = ImageEnhance.Brightness(im).enhance(brightness)
            if contrast != 1.0:
                im = ImageEnhance.Contrast(im).enhance(contrast)
            im.save(path, "PNG")
        return True
    except Exception:
        return False


def read_barcode(path):
    """Return first barcode/QR text found in the image, else None."""
    if not HAS_ZBAR:
        return None
    try:
        with Image.open(path) as im:
            results = _zbar_decode(im.convert("RGB"))
        for r in results:
            try:
                val = r.data.decode("utf-8").strip()
            except Exception:
                val = str(r.data).strip()
            if val:
                return val
        return None
    except Exception:
        return None


def sanitize(text, fallback="scan"):
    safe = "".join(c for c in (text or "") if c.isalnum() or c in "-_")
    return safe or fallback


# ---------------------------------------------------------------------------
# Network + TWAIN backend
# ---------------------------------------------------------------------------

def tcp_reachable(ip, ports=SCANNER_PORTS, timeout=1.2):
    for p in ports:
        try:
            with socket.create_connection((ip, p), timeout=timeout):
                return True
        except Exception:
            continue
    return False


class ConnectionChecker(QtCore.QThread):
    result = QtCore.pyqtSignal(bool, str)

    def __init__(self, ip):
        super().__init__()
        self.ip = (ip or "").strip()

    def run(self):
        if not self.ip:
            self.result.emit(False, "Scanner ka IP address set nahi hai")
            return
        if tcp_reachable(self.ip):
            self.result.emit(True, "Connected — scanner network par mil gaya (%s)" % self.ip)
        else:
            self.result.emit(False, "Not connected — %s reachable nahi. Scanner ON hai? "
                                    "Same WiFi/LAN par hai?" % self.ip)


class ScannerStateChecker(QtCore.QThread):
    """Ask an eSCL scanner whether it's Idle (free) or Processing (busy)."""
    state = QtCore.pyqtSignal(str)   # "free" / "busy" / "unknown"

    def __init__(self, ip):
        super().__init__()
        self.ip = (ip or "").strip()

    def run(self):
        if not self.ip:
            self.state.emit("unknown"); return
        try:
            import urllib.request as _u
            r = _u.urlopen("http://%s/eSCL/ScannerStatus" % self.ip, timeout=5)
            xml = r.read().decode("utf-8", "ignore")
            m = re.search(r"State>\s*([A-Za-z]+)", xml)
            st = (m.group(1).lower() if m else "")
            if st == "idle":
                self.state.emit("free")
            elif st:
                self.state.emit("busy")
            else:
                self.state.emit("unknown")
        except Exception:
            self.state.emit("unknown")


class EsclTestWorker(QtCore.QThread):
    """Step-by-step eSCL probe that reports the ACTUAL response at each stage,
    so the real cause of a scan failure (busy job / bad settings / offline) is
    visible instead of just a generic error."""
    done = QtCore.pyqtSignal(str)

    def __init__(self, ip):
        super().__init__()
        self.ip = (ip or "").strip()

    def run(self):
        import urllib.request as U
        import urllib.error as UE
        lines = []
        def add(t=""):
            lines.append(t)
        add("===== ApneScan eSCL Test =====")
        add("Time: %s" % datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        add("Scanner IP: %s" % (self.ip or "(not set)"))
        base = "http://%s/eSCL" % self.ip

        add("")
        add("[1] TCP connect (port 80):")
        try:
            add("    %s" % ("OK - reachable" if tcp_reachable(self.ip) else "FAIL - not reachable (scanner ON? same WiFi/LAN?)"))
        except Exception as e:
            add("    ERROR: %s" % e)

        add("")
        add("[2] GET /eSCL/ScannerCapabilities:")
        try:
            r = U.urlopen(base + "/ScannerCapabilities", timeout=8)
            body = r.read().decode("utf-8", "ignore")
            add("    HTTP %s, %d bytes" % (getattr(r, "status", 200), len(body)))
            mw = re.findall(r"MaxWidth>\s*(\d+)", body)
            mh = re.findall(r"MaxHeight>\s*(\d+)", body)
            if mw and mh:
                add("    Max scan area: W=%s H=%s (1/300 inch)" % (max(mw), max(mh)))
            if "Duplex" in body:
                add("    Duplex: supported (eSCL)")
        except UE.HTTPError as e:
            add("    HTTP ERROR %s (eSCL shayad support nahi)" % e.code)
        except Exception as e:
            add("    ERROR: %s" % e)

        add("")
        add("[3] GET /eSCL/ScannerStatus (free/busy + open jobs):")
        try:
            r = U.urlopen(base + "/ScannerStatus", timeout=8)
            body = r.read().decode("utf-8", "ignore")
            st = re.search(r"State>\s*([A-Za-z]+)", body)
            add("    HTTP %s, State=%s" % (getattr(r, "status", 200), st.group(1) if st else "?"))
            jobs = re.findall(r"JobState>\s*([A-Za-z]+)", body)
            if jobs:
                add("    Open/last jobs: %s" % ", ".join(jobs))
                add("    (Agar koi 'Processing'/'Pending' job hai to scanner busy rahega.)")
            else:
                add("    No open jobs listed.")
        except UE.HTTPError as e:
            add("    HTTP ERROR %s" % e.code)
        except Exception as e:
            add("    ERROR: %s" % e)

        add("")
        add("[4] POST /eSCL/ScanJobs  (THE REAL TEST - minimal A4 simplex gray):")
        settings = (
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            '<scan:ScanSettings xmlns:scan="http://schemas.hp.com/imaging/escl/2011/05/03" '
            'xmlns:pwg="http://www.pwg.org/schemas/2010/12/sm">\n'
            '<pwg:Version>2.6</pwg:Version>\n'
            '<pwg:ScanRegions pwg:MustHonor="false"><pwg:ScanRegion>\n'
            '<pwg:Height>3550</pwg:Height><pwg:Width>2480</pwg:Width>\n'
            '<pwg:XOffset>0</pwg:XOffset><pwg:YOffset>0</pwg:YOffset>\n'
            '</pwg:ScanRegion></pwg:ScanRegions>\n'
            '<pwg:InputSource>Feeder</pwg:InputSource>\n'
            '<scan:ColorMode>Grayscale8</scan:ColorMode>\n'
            '<scan:XResolution>200</scan:XResolution>\n'
            '<scan:YResolution>200</scan:YResolution>\n'
            '<pwg:DocumentFormat>image/jpeg</pwg:DocumentFormat>\n'
            '</scan:ScanSettings>'
        )
        job = None
        try:
            req = U.Request(base + "/ScanJobs", data=settings.encode("utf-8"),
                            headers={"Content-Type": "text/xml"}, method="POST")
            r = U.urlopen(req, timeout=15)
            job = r.headers.get("Location")
            add("    HTTP %s  ->  JOB CREATED" % getattr(r, "status", 201))
            add("    Location: %s" % job)
        except UE.HTTPError as e:
            add("    HTTP ERROR %s" % e.code)
            try:
                body = e.read().decode("utf-8", "ignore")
                if body.strip():
                    add("    Response body: %s" % body.replace("\n", " ")[:300])
            except Exception:
                pass
            if e.code == 503:
                add("    => 503 = scanner BUSY / ek purana job khula hai.")
                add("       Fix: scanner off/on karein; koi aur scan app band karein.")
            elif e.code in (400, 415, 409):
                add("    => %s = scan settings/format ka problem (is scanner ke liye tweak chahiye)." % e.code)
        except Exception as e:
            add("    ERROR: %s" % e)

        if job:
            if job.startswith("/"):
                job = "http://%s%s" % (self.ip, job)
            add("")
            add("[5] GET job/NextDocument (1 page fetch):")
            try:
                r = U.urlopen(job + "/NextDocument", timeout=30)
                d = r.read()
                add("    HTTP %s, %d bytes  ->  PAGE RECEIVED (scan works!)" % (getattr(r, "status", 200), len(d)))
            except UE.HTTPError as e:
                add("    HTTP %s (koi page nahi / done)" % e.code)
            except Exception as e:
                add("    ERROR: %s" % e)
            add("")
            add("[6] DELETE job (release):")
            try:
                U.urlopen(U.Request(job, method="DELETE"), timeout=8)
                add("    OK - job released (scanner free).")
            except Exception as e:
                add("    ERROR: %s" % e)

        add("")
        add("===== RESULT =====")
        add("Agar [4] me 'JOB CREATED' aaya -> scan theek hona chahiye.")
        add("Agar [4] me 503 aaya -> scanner busy: off/on karein + doosri scan app band karein.")
        add("Agar [1] FAIL -> scanner network par nahi mila (IP/WiFi check).")
        add("==================")
        self.done.emit("\n".join(lines))


class StatsWorker(QtCore.QThread):
    """Talk to the ApneScan stats server (Google Apps Script). action:
    'ping' (mark online + fetch), 'scan' (add scans + fetch), 'stats' (fetch)."""
    got = QtCore.pyqtSignal(int, int, int)   # total, today, online
    failed = QtCore.pyqtSignal()

    def __init__(self, url, client, action="ping", n=0):
        super().__init__()
        self.url = url; self.client = client; self.action = action; self.n = n

    def run(self):
        try:
            import urllib.request as U
            import urllib.parse as P
            import json as J
            q = {"action": self.action, "client": self.client}
            if self.action == "scan":
                q["n"] = str(self.n)
            full = self.url + ("&" if "?" in self.url else "?") + P.urlencode(q)
            r = U.urlopen(full, timeout=12)
            data = J.loads(r.read().decode("utf-8", "ignore"))
            if data.get("ok"):
                self.got.emit(int(data.get("total", 0)),
                              int(data.get("today", 0)),
                              int(data.get("online", 0)))
            else:
                self.failed.emit()
        except Exception:
            self.failed.emit()


class ScannerError(Exception):
    pass


def list_sources(hwnd):
    if not HAS_TWAIN:
        raise ScannerError("TWAIN (pytwain) install nahi hai.")
    sm = twain.SourceManager(hwnd)
    try:
        try:
            return list(sm.source_list)
        except Exception:
            return list(sm.GetSourceList())
    finally:
        try:
            sm.close()
        except Exception:
            try:
                sm.destroy()
            except Exception:
                pass


class _FileXferUnsupported(Exception):
    pass


def _open_twain_source(sm, source_name):
    try:
        src = sm.open_source(source_name) if source_name else sm.open_source()
    except Exception:
        try:
            src = sm.OpenSource(source_name) if source_name else sm.OpenSource()
        except Exception:
            src = None
    if src is None:
        try:
            src = sm.open_source()
        except Exception:
            try:
                src = sm.OpenSource()
            except Exception:
                src = None
    return src


def _common_caps(src, dpi, pixel_type, duplex):
    def _try(fn):
        try:
            fn()
        except Exception:
            pass
    _try(lambda: src.set_capability(twain.ICAP_XRESOLUTION, twain.TWTY_FIX32, float(dpi)))
    _try(lambda: src.set_capability(twain.ICAP_YRESOLUTION, twain.TWTY_FIX32, float(dpi)))
    pt = {"color": twain.TWPT_RGB, "gray": twain.TWPT_GRAY, "bw": twain.TWPT_BW}[pixel_type]
    _try(lambda: src.set_capability(twain.ICAP_PIXELTYPE, twain.TWTY_UINT16, pt))
    _try(lambda: src.set_capability(twain.CAP_FEEDERENABLED, twain.TWTY_BOOL, True))
    _try(lambda: src.set_capability(twain.CAP_AUTOFEED, twain.TWTY_BOOL, True))
    _try(lambda: src.set_capability(twain.CAP_AUTOSCAN, twain.TWTY_BOOL, True))
    _try(lambda: src.set_capability(twain.CAP_XFERCOUNT, twain.TWTY_INT16, -1))
    _try(lambda: src.set_capability(twain.CAP_DUPLEXENABLED, twain.TWTY_BOOL, bool(duplex)))


def _scan_file(hwnd, source_name, dpi, pixel_type, duplex, on_page=None, should_stop=None):
    """File-transfer acquire: the driver writes each page to a file and (on most
    ADF scanners) keeps the feeder moving continuously. If the source doesn't
    support it, raises _FileXferUnsupported so the caller falls back to native.
    Only raises _FileXferUnsupported when ZERO pages were produced (so we never
    double-feed paper)."""
    count = 0
    MAX_PAGES = 1000   # hard guard: never loop forever
    sm = twain.SourceManager(hwnd)
    src = None
    try:
        src = _open_twain_source(sm, source_name)
        if src is None:
            raise _FileXferUnsupported()
        _common_caps(src, dpi, pixel_type, duplex)
        # Switch to FILE transfer mechanism. If unsupported -> fall back.
        try:
            src.set_capability(twain.ICAP_XFERMECH, twain.TWTY_UINT16, twain.TWSX_FILE)
        except Exception:
            raise _FileXferUnsupported()
        try:
            src.request_acquire(show_ui=False, modal_ui=False)
        except Exception:
            raise _FileXferUnsupported()

        while count < MAX_PAGES:
            if should_stop and should_stop():
                break
            fd, bmp = tempfile.mkstemp(suffix=".bmp")
            os.close(fd)
            try:
                os.remove(bmp)
            except Exception:
                pass
            # point the driver at our file
            try:
                src.file_xfer_params(bmp, twain.TWFF_BMP)
            except Exception:
                if count == 0:
                    raise _FileXferUnsupported()
                break
            # transfer one page to the file
            try:
                more = src.xfer_image_by_file()
            except Exception:
                # feeder empty / done (or unsupported on first try)
                if count == 0:
                    raise _FileXferUnsupported()
                break
            if not os.path.exists(bmp):
                if count == 0:
                    raise _FileXferUnsupported()
                break
            try:
                with Image.open(bmp) as im:
                    img = im.convert("RGB").copy()
            finally:
                try:
                    os.remove(bmp)
                except Exception:
                    pass
            count += 1
            if on_page:
                on_page(img)
            # 'more' may be a bool/int/pending-count depending on pytwain; stop when falsy
            if more in (None, 0, False):
                break
        return count
    finally:
        try:
            if src is not None:
                src.close()
        except Exception:
            pass
        try:
            sm.close()
        except Exception:
            try:
                sm.destroy()
            except Exception:
                pass


def scan_pages(hwnd, source_name, dpi, pixel_type, duplex, on_page=None, should_stop=None):
    if not HAS_TWAIN:
        raise ScannerError("TWAIN (pytwain) install nahi hai.")
    # Try the continuous file-transfer path first; fall back to the proven
    # native path if the driver doesn't support it (only when 0 pages scanned).
    try:
        n = _scan_file(hwnd, source_name, dpi, pixel_type, duplex, on_page, should_stop)
        if n > 0:
            return n
        raise _FileXferUnsupported()
    except _FileXferUnsupported:
        pass
    except ScannerError:
        raise
    except Exception:
        pass
    return _scan_native(hwnd, source_name, dpi, pixel_type, duplex, on_page, should_stop)


def _scan_native(hwnd, source_name, dpi, pixel_type, duplex, on_page=None, should_stop=None):
    count = 0
    sm = twain.SourceManager(hwnd)
    src = None
    try:
        try:
            src = sm.open_source(source_name) if source_name else sm.open_source()
        except Exception:
            try:
                src = sm.OpenSource(source_name) if source_name else sm.OpenSource()
            except Exception:
                src = None
        if src is None:
            # Stored name might be a WIA name (when switching method) -> open default TWAIN source.
            try:
                src = sm.open_source()
            except Exception:
                try:
                    src = sm.OpenSource()
                except Exception:
                    src = None
        if src is None:
            raise ScannerError("Scanner open nahi hua. Profile me device chuno.")

        def _try(fn):
            try:
                fn()
            except Exception:
                pass

        _try(lambda: src.set_capability(twain.ICAP_XRESOLUTION, twain.TWTY_FIX32, float(dpi)))
        _try(lambda: src.set_capability(twain.ICAP_YRESOLUTION, twain.TWTY_FIX32, float(dpi)))
        pt = {"color": twain.TWPT_RGB, "gray": twain.TWPT_GRAY, "bw": twain.TWPT_BW}[pixel_type]
        _try(lambda: src.set_capability(twain.ICAP_PIXELTYPE, twain.TWTY_UINT16, pt))
        # Continuous ADF feed: pull EVERY page in a single acquire (no per-page gap).
        _try(lambda: src.set_capability(twain.CAP_FEEDERENABLED, twain.TWTY_BOOL, True))
        _try(lambda: src.set_capability(twain.CAP_AUTOFEED, twain.TWTY_BOOL, True))
        _try(lambda: src.set_capability(twain.CAP_AUTOSCAN, twain.TWTY_BOOL, True))
        _try(lambda: src.set_capability(twain.CAP_XFERCOUNT, twain.TWTY_INT16, -1))
        _try(lambda: src.set_capability(twain.CAP_DUPLEXENABLED, twain.TWTY_BOOL, bool(duplex)))
        _try(lambda: src.request_acquire(show_ui=False, modal_ui=False))

        while True:
            if should_stop and should_stop():
                break
            try:
                rv = src.xfer_image_natively()
            except Exception as exc:
                if count:
                    break
                raise ScannerError("Scan transfer fail: %s" % exc)
            if rv is None:
                break
            handle, more = rv
            fd, bmp = tempfile.mkstemp(suffix=".bmp")
            os.close(fd)
            try:
                twain.dib_to_bm_file(handle, bmp)
            finally:
                try:
                    twain.global_handle_free(handle)
                except Exception:
                    pass
            with Image.open(bmp) as im:
                img = im.convert("RGB").copy()
            try:
                os.remove(bmp)
            except Exception:
                pass
            count += 1
            if on_page:
                on_page(img)
            if not more:
                break
    finally:
        try:
            if src is not None:
                src.close()
        except Exception:
            pass
        try:
            sm.close()
        except Exception:
            try:
                sm.destroy()
            except Exception:
                pass

    if count == 0:
        raise ScannerError("Koi page scan nahi hua. Feeder me document rakha hai? Scanner ON hai?")
    return count


# ---------------------------------------------------------------------------
# WIA backend (alternative to TWAIN, for other users' scanners)
# ---------------------------------------------------------------------------

WIA_ERROR_PAPER_EMPTY = -2145320957   # 0x80210003


def list_wia_sources():
    if not HAS_W32:
        raise ScannerError("WIA (pywin32) install nahi hai.")
    dm = _w32.Dispatch("WIA.DeviceManager")
    out = []
    for i in range(1, dm.DeviceInfos.Count + 1):
        info = dm.DeviceInfos.Item(i)
        name = None
        try:
            for p in info.Properties:
                if p.Name == "Name":
                    name = p.Value
                    break
        except Exception:
            pass
        out.append((info.DeviceID, name or ("Scanner %d" % i)))
    return out


def wia_scan_pages(device_id, dpi, pixel_type, duplex, on_page=None, should_stop=None):
    wia_scan_pages.last_duplex_value = getattr(wia_scan_pages, "last_duplex_value", None)
    """Scan via Windows WIA. Kept simple: no forced properties (some drivers
    throw otherwise). Runs inside a thread that has called CoInitialize."""
    if not HAS_W32:
        raise ScannerError("WIA (pywin32) install nahi hai.")
    dm = _w32.Dispatch("WIA.DeviceManager")
    device = None
    for i in range(1, dm.DeviceInfos.Count + 1):
        info = dm.DeviceInfos.Item(i)
        if device_id is None or info.DeviceID == device_id:
            device = info.Connect()
            break
    if device is None:
        raise ScannerError("WIA scanner nahi mila. Settings me scanner chuno.")

    def _wia_set(props, pid, value):
        # Set one WIA property; True if it took, False if the driver rejected it.
        try:
            for p in props:
                if int(p.PropertyID) == pid:
                    p.Value = value
                    return True
        except Exception:
            return False
        return False

    def _wia_valid_dpi(props, want):
        # Pick a resolution the device actually supports (avoids "parameter incorrect").
        try:
            for p in props:
                if int(p.PropertyID) == 6147:
                    vals = None
                    try:
                        vals = [int(x) for x in p.SubType.Values]   # WIA_PROP_LIST
                    except Exception:
                        vals = None
                    if vals:
                        return min(vals, key=lambda v: abs(v - want))
        except Exception:
            pass
        return want

    def _apply_feeder(dev, want_duplex=None):
        # Feeder + (optional) DUPLEX + all pages. HP network drivers use different
        # duplex "select" values, so try a few and keep the first the driver accepts.
        wd = duplex if want_duplex is None else want_duplex
        try:
            dprops = dev.Properties
            if wd:
                for val in (5, 7, 0x8005, 4, 3):   # FEEDER|DUPLEX variants (0x8005=32773 HP)
                    if _wia_set(dprops, 3088, val):
                        wia_scan_pages.last_duplex_value = val
                        break
                else:
                    wia_scan_pages.last_duplex_value = None
            else:
                _wia_set(dprops, 3088, 1)          # feeder only (single side)
            _wia_set(dprops, 3096, 0)              # all pages
        except Exception:
            pass

    def _apply_quality(it):
        try:
            props = it.Properties
            _wia_set(props, 4104, {"bw": 0, "gray": 2, "color": 3}.get(pixel_type, 3))
            dv = _wia_valid_dpi(props, int(dpi))
            _wia_set(props, 6147, dv)
            _wia_set(props, 6148, dv)
        except Exception:
            pass

    _apply_feeder(device)          # set FEEDER|DUPLEX on the DEVICE first
    item = device.Items[1]         # enumerate the page item AFTER duplex is active
    _apply_quality(item)
    _props_ok = True
    # BMP format id (widely accepted by HP WIA drivers).
    WIA_FMT_BMP = "{B96B3CAB-0728-11D3-9D7B-0000F81EF32E}"

    def _transfer_one(it):
        # Try Transfer(format) first (robust on HP), then plain Transfer().
        try:
            return it.Transfer(WIA_FMT_BMP)
        except Exception:
            return it.Transfer()

    count = 0
    retry_stage = 0
    while True:
        if should_stop and should_stop():
            break
        try:
            image = _transfer_one(item)
        except Exception as exc:
            hres = exc.args[0] if getattr(exc, "args", None) else None
            # End-of-feeder conditions -> stop cleanly.
            if hres in (WIA_ERROR_PAPER_EMPTY, -2145320957) or count > 0:
                break
            # First page failed: retry once with a fresh item and NO custom props.
            if retry_stage < 2:
                stage = retry_stage
                retry_stage += 1
                try:
                    device2 = None
                    dm2 = _w32.Dispatch("WIA.DeviceManager")
                    for i in range(1, dm2.DeviceInfos.Count + 1):
                        info = dm2.DeviceInfos.Item(i)
                        if device_id is None or info.DeviceID == device_id:
                            device2 = info.Connect(); break
                    if device2 is not None:
                        if stage == 0:
                            _apply_feeder(device2)          # retry: duplex, no quality props
                        # stage 1: apply nothing -> guaranteed to scan (single side)
                        item = device2.Items[1]
                except Exception:
                    pass
                continue
            raise ScannerError("WIA scan fail: %s" % exc)
        fd, raw = tempfile.mkstemp(suffix=".bmp")
        os.close(fd)
        try:
            os.remove(raw)
        except Exception:
            pass
        try:
            image.SaveFile(raw)
        except Exception as exc:
            raise ScannerError("WIA save fail: %s" % exc)
        with Image.open(raw) as im:
            img = im.convert("RGB").copy()
        try:
            os.remove(raw)
        except Exception:
            pass
        count += 1
        if on_page:
            on_page(img)
    if count == 0:
        raise ScannerError("Koi page scan nahi hua (WIA).")
    return count


# ---------------------------------------------------------------------------
# Friendly error messages (technical -> simple)
# ---------------------------------------------------------------------------

def friendly_error(text, lang="hi"):
    t = str(text)
    low = t.lower()
    hi = {
        "not_found": "Scanner nahi mila. Scanner ON karein aur PC se juda hai ya nahi check karein.",
        "busy": "Scanner abhi busy hai. Thodi der baad dobara try karein.",
        "empty": "Feeder me koi document nahi hai. Page rakh kar dobara scan karein.",
        "jam": "Paper jam hua lagta hai. Feeder check karein.",
        "driver": "Scanner driver ki dikkat. Doosri scan-method (Settings me TWAIN/WIA) try karein.",
        "twain": "TWAIN scanner nahi khula. Settings me WIA method try karein.",
        "generic": "Scan nahi ho paya. Scanner ON hai aur juda hai? Dobara try karein.",
    }
    en = {
        "not_found": "Scanner not found. Turn it on and check it's connected.",
        "busy": "Scanner is busy. Please try again in a moment.",
        "empty": "No document in the feeder. Add a page and scan again.",
        "jam": "Looks like a paper jam. Check the feeder.",
        "driver": "Scanner driver issue. Try the other method (TWAIN/WIA) in Settings.",
        "twain": "TWAIN scanner didn't open. Try WIA method in Settings.",
        "generic": "Scanning failed. Is the scanner on and connected? Try again.",
    }
    m = hi if lang == "hi" else en
    if "paper_empty" in low or "0x80210003" in low or "-2145320957" in low or "feeder me" in low:
        return m["empty"]
    if "busy" in low or "0x80210006" in low:
        return m["busy"]
    if "jam" in low:
        return m["jam"]
    if "device driver threw" in low or "exception_in_driver" in low or "-2145320946" in low:
        return m["driver"]
    if "nahi mila" in low or "not found" in low or "no such" in low or "device" in low and "nahi" in low:
        return m["not_found"]
    if "twain" in low and ("open" in low or "khula" in low or "install" in low):
        return m["twain"]
    return m["generic"] + "\n\n(" + t[:200] + ")"


# ---------------------------------------------------------------------------
# Small translation layer (English / Hindi) for the main interface
# ---------------------------------------------------------------------------

T = {
    "scan": {"hi": "Scan", "en": "Scan"},
    "profile": {"hi": "Profile:", "en": "Profile:"},
    "profiles": {"hi": "Profiles…", "en": "Profiles…"},
    "claim_no": {"hi": "Claim No.:", "en": "Claim No.:"},
    "import": {"hi": "Import", "en": "Import"},
    "save_pdf": {"hi": "Save PDF", "en": "Save PDF"},
    "check": {"hi": "Check", "en": "Check"},
    "up": {"hi": "Upar", "en": "Up"},
    "down": {"hi": "Neeche", "en": "Down"},
    "delete": {"hi": "Delete", "en": "Delete"},
    "clear": {"hi": "Clear", "en": "Clear"},
    "rotate_l": {"hi": "Rotate ⟲", "en": "Rotate ⟲"},
    "rotate_r": {"hi": "Rotate ⟳", "en": "Rotate ⟳"},
    "menu_file": {"hi": "File", "en": "File"},
    "menu_edit": {"hi": "Edit", "en": "Edit"},
    "menu_tools": {"hi": "Tools", "en": "Tools"},
    "menu_settings": {"hi": "Settings", "en": "Settings"},
    "menu_help": {"hi": "Help", "en": "Help"},
    "help_guide": {"hi": "Guide / Madad", "en": "Guide / Help"},
    "about": {"hi": "About", "en": "About"},
    "whatsnew": {"hi": "Naya kya hai", "en": "What's new"},
    "feedback": {"hi": "Feedback / Sujhav", "en": "Feedback"},
    "language": {"hi": "Language / Bhasha", "en": "Language"},
    "scan_method": {"hi": "Scan method (TWAIN/WIA)", "en": "Scan method (TWAIN/WIA)"},
    "options": {"hi": "Options…", "en": "Options…"},
    "pages_hint": {"hi": "Pages (drag se order badlein)", "en": "Pages (drag to reorder)"},
    "simple_on": {"hi": "Simple mode (aasan)", "en": "Simple mode"},
    "scanner_ip": {"hi": "Scanner IP:", "en": "Scanner IP:"},
    "st_pages": {"hi": "Pages", "en": "Pages"},
    "st_profile": {"hi": "Profile", "en": "Profile"},
    "none_profile": {"hi": "koi profile nahi", "en": "no profile"},
    "checking": {"hi": "Connection check ho raha hai...", "en": "Checking connection..."},
    "select_first": {"hi": "Pehle koi page select karein.", "en": "Select a page first."},
    "scan_first": {"hi": "Pehle koi page scan/import karein.", "en": "Scan or import a page first."},
    "fast": {"hi": "Fast", "en": "Fast"},
    "save_all": {"hi": "Sab pages ki PDF", "en": "Save all pages (PDF)"},
    "save_sel": {"hi": "Selected pages ki PDF", "en": "Save selected pages (PDF)"},
    "no_selection": {"hi": "Koi page select nahi hai. Thumbnails me Ctrl/Shift se pages chuno.", "en": "No pages selected. Use Ctrl/Shift to select pages."},
}


def tr(key, lang="hi"):
    e = T.get(key)
    if not e:
        return key
    return e.get(lang, e.get("hi", key))


HELP_TEXT_HI = """<h3>ApneScan — Madad</h3>
<b>Shuru kaise karein</b><br>
1. <b>Settings → Profiles…</b> me ek profile banayein: <i>Choose device</i> se apna
scanner chuno, DPI/Colour/Duplex set karo.<br>
2. Upar profile chuno aur <b>Scan</b> dabao (ya F5).<br>
3. Pages left me dikhenge — drag se order badlo, rotate/delete karo.<br>
4. <b>Save PDF</b> se PDF banao (OCR ka option milega).<br><br>
<b>Scanner nahi mil raha?</b><br>
Settings → <i>Scan method</i> me TWAIN aur WIA dono try karo. Scanner ka driver
install hona chahiye, aur woh ON + connected ho.<br><br>
<b>Fast scan ke liye:</b> Profile me DPI 200, Mode Black & White.<br><br>
<b>Auto-save, register, backup, watermark, split, merge, password PDF</b> —
sab <b>Settings → Options</b> aur <b>Tools</b> menu me hain.<br><br>
Keyboard: F5=Scan, Ctrl+S=Save PDF, Ctrl+P=Print, Ctrl+F=Search, Delete=Delete page.
"""

CHANGELOG_HTML = """<h3>ApneScan — Naya kya hai / What's new</h3>
<b>Version 1.0</b><br>
- TWAIN + WIA dono scanners support<br>
- Setup wizard (pehli baar aasan setup)<br>
- Profiles, auto-save PDF, auto file-naming, claim folders<br>
- OCR searchable PDF (Hindi+English), OCR text export<br>
- Blank-page removal, auto-crop, deskew, quality enhance<br>
- PDF compress, password PDF, split, merge, print<br>
- Excel register, activity log, duplicate check, daily backup<br>
- Barcode/QR se claim number, monthly report<br>
- Hindi / English interface, Simple mode, Feedback, in-app Help<br>
"""


HELP_TEXT_EN = """<h3>ApneScan — Help</h3>
<b>Getting started</b><br>
1. Go to <b>Settings → Profiles…</b> and create a profile: click <i>Choose device</i>
to pick your scanner, then set DPI/Colour/Duplex.<br>
2. Select the profile at the top and click <b>Scan</b> (or press F5).<br>
3. Pages appear on the left — drag to reorder, rotate/delete.<br>
4. Click <b>Save PDF</b> (you'll be asked about OCR).<br><br>
<b>Scanner not detected?</b><br>
In Settings → <i>Scan method</i>, try both TWAIN and WIA. The scanner's driver
must be installed, and the device must be on and connected.<br><br>
<b>For faster scans:</b> set DPI 200 and Black & White in the profile.<br><br>
<b>Auto-save, register, backup, watermark, split, merge, password PDF</b> are all
under <b>Settings → Options</b> and the <b>Tools</b> menu.<br><br>
Shortcuts: F5=Scan, Ctrl+S=Save PDF, Ctrl+P=Print, Ctrl+F=Search, Delete=Delete page.
"""



# ---------------------------------------------------------------------------
# NAPS2 engine backend (uses NAPS2's console to scan — reliable fast + duplex)
# ---------------------------------------------------------------------------

NAPS2_GUESSES = [
    r"C:\\Program Files\\NAPS2\\NAPS2.Console.exe",
    r"C:\\Program Files (x86)\\NAPS2\\NAPS2.Console.exe",
    r"C:\\Program Files\\NAPS2\\NAPS2.exe",
    r"C:\\Program Files (x86)\\NAPS2\\NAPS2.exe",
]


def find_naps2():
    for p in NAPS2_GUESSES:
        if os.path.exists(p):
            return p
    return ""


def scan_via_naps2(naps2_exe, profile, tmpdir, duplex, on_page=None, should_stop=None):
    import subprocess
    import glob as _glob
    import re as _re
    if not naps2_exe or not os.path.exists(naps2_exe):
        raise ScannerError("NAPS2 nahi mila. Settings me NAPS2 ka path set karo.")
    # clean old temp scans
    for f in _glob.glob(os.path.join(tmpdir, "naps2_*.jpg")):
        try:
            os.remove(f)
        except Exception:
            pass
    out_pattern = os.path.join(tmpdir, "naps2_$(n).jpg")
    cmd = [naps2_exe, "-o", out_pattern, "--split", "-f", "-v"]
    if profile:
        cmd += ["-p", profile]
    if duplex:
        cmd += ["--source", "duplex"]
    try:
        proc = subprocess.run(cmd, capture_output=True, text=True, timeout=600,
                              creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0))
    except subprocess.TimeoutExpired:
        raise ScannerError("NAPS2 scan me bahut time lag gaya (timeout).")
    except Exception as exc:
        raise ScannerError("NAPS2 chalane me dikkat: %s" % exc)

    files = _glob.glob(os.path.join(tmpdir, "naps2_*.jpg"))

    def _num(p):
        m = _re.search(r"naps2_(\d+)\.jpg$", os.path.basename(p))
        return int(m.group(1)) if m else 0
    files.sort(key=_num)

    if not files:
        msg = (proc.stderr or proc.stdout or "").strip()
        raise ScannerError("NAPS2 se koi page nahi aaya.\nProfile naam sahi hai? Scanner ON hai?\n\n%s"
                           % msg[:300])
    count = 0
    for f in files:
        if should_stop and should_stop():
            break
        try:
            with Image.open(f) as im:
                img = im.convert("RGB").copy()
        except Exception:
            continue
        try:
            os.remove(f)
        except Exception:
            pass
        count += 1
        if on_page:
            on_page(img)
    if count == 0:
        raise ScannerError("NAPS2 se koi page nahi aaya.")
    return count


# ---------------------------------------------------------------------------
# eSCL (AirScan) network backend — pure Python, talks to the scanner's IP over
# HTTP. Supports DUPLEX over the network with NO other software needed.
# ---------------------------------------------------------------------------

def scan_via_escl(ip, dpi, color, duplex, on_page=None, should_stop=None, page_size="auto"):
    import urllib.request as _u
    import urllib.error as _ue
    import io as _io
    import time as _t

    if not ip:
        raise ScannerError("Scanner IP set nahi hai. Settings \u2192 Scanner IP me IP daalo (jaise 192.168.1.8).")
    ip = ip.strip()
    base = "http://%s/eSCL" % ip
    cmode = {"color": "RGB24", "gray": "Grayscale8", "bw": "BlackAndWhite1"}.get(color, "RGB24")

    # eSCL region units are 1/300 inch. Fixed sizes request that exact sheet.
    # "auto" requests the FULL BED WIDTH (so sides are never cut) with a Legal-length
    # height, and MustHonor="false" so the ADF stops at each sheet's real trailing
    # edge (no forced long scan = no feeder jam). The dark backing that appears
    # around a narrow sheet is painted white by whiten_dark_background.
    ps = (page_size or "auto").lower()
    if ps.startswith("a4"):
        w, h = 2480, 3550
    elif ps.startswith("letter"):
        w, h = 2550, 3300
    elif ps.startswith("legal"):
        w, h = 2550, 4200
    elif ps.startswith("a5"):
        w, h = 1748, 2480
    else:  # auto -> full width, Legal length, let ADF stop at paper end
        w, h = 2550, 4200
        try:
            _rr = _u.urlopen(base + "/ScannerCapabilities", timeout=8)
            _cap = _rr.read().decode("utf-8", "ignore")
            _mw = re.findall(r"MaxWidth>\s*(\d+)", _cap)
            if _mw:
                w = max(int(x) for x in _mw)   # full bed width (no side cut)
        except Exception:
            pass

    settings = (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<scan:ScanSettings xmlns:scan="http://schemas.hp.com/imaging/escl/2011/05/03"'
        ' xmlns:pwg="http://www.pwg.org/schemas/2010/12/sm">\n'
        '<pwg:Version>2.6</pwg:Version>\n'
        '<scan:Intent>Document</scan:Intent>\n'
        '<pwg:ScanRegions pwg:MustHonor="false"><pwg:ScanRegion>\n'
        '<pwg:Height>%d</pwg:Height><pwg:Width>%d</pwg:Width>\n'
        '<pwg:XOffset>0</pwg:XOffset><pwg:YOffset>0</pwg:YOffset>\n'
        '</pwg:ScanRegion></pwg:ScanRegions>\n'
        '<pwg:InputSource>Feeder</pwg:InputSource>\n'
        '<scan:Duplex>%s</scan:Duplex>\n'
        '<scan:ColorMode>%s</scan:ColorMode>\n'
        '<scan:XResolution>%d</scan:XResolution>\n'
        '<scan:YResolution>%d</scan:YResolution>\n'
        '<pwg:DocumentFormat>image/jpeg</pwg:DocumentFormat>\n'
        '</scan:ScanSettings>'
    ) % (h, w, "true" if duplex else "false", cmode, dpi, dpi)

    # Before starting, clear any STUCK job (e.g. a "Processing" job left over from
    # an earlier interrupted scan). Such a job makes every new ScanJobs POST return
    # 503 even though ScannerStatus shows Idle. We read ScannerStatus, find job
    # URIs/UUIDs whose state is not finished, and DELETE them.
    def _clear_stuck_jobs():
        try:
            rr = _u.urlopen(base + "/ScannerStatus", timeout=8)
            xml = rr.read().decode("utf-8", "ignore")
        except Exception:
            return
        blocks = re.findall(r"<[a-zA-Z]*:?JobInfo>.*?</[a-zA-Z]*:?JobInfo>", xml, re.S)
        for blk in blocks:
            stm = re.search(r"JobState>\s*([A-Za-z]+)", blk)
            state = (stm.group(1).lower() if stm else "")
            if state in ("completed", "canceled", "aborted", ""):
                continue
            uri = re.search(r"JobUri>\s*([^<\s]+)", blk)
            uuid = re.search(r"JobUuid>\s*([^<\s]+)", blk)
            target = None
            if uri:
                target = uri.group(1)
                if target.startswith("/"):
                    target = "http://%s%s" % (ip, target)
            elif uuid:
                target = "%s/ScanJobs/%s" % (base, uuid.group(1))
            if target:
                try:
                    _u.urlopen(_u.Request(target, method="DELETE"), timeout=8)
                    _t.sleep(0.5)
                except Exception:
                    pass

    _clear_stuck_jobs()

    # Create the scan job. Try immediately first (this worked before); only if the
    # scanner reports busy (503/409) do we wait briefly and retry a couple of times.
    def _post_job():
        req = _u.Request(base + "/ScanJobs", data=settings.encode("utf-8"),
                         headers={"Content-Type": "text/xml"}, method="POST")
        return _u.urlopen(req, timeout=45)

    resp = None
    last_code = None
    for attempt in range(10):                    # be patient: scanner is briefly busy at start
        if should_stop and should_stop():
            break
        try:
            resp = _post_job()
            break
        except _ue.HTTPError as e:
            last_code = e.code
            if e.code in (503, 409, 429):        # busy -> wait and keep trying
                _t.sleep(1.5)
                continue
            raise ScannerError("eSCL job error: HTTP %s. Scanner network-scan (eSCL) support karta hai?" % e.code)
        except Exception as e:
            last_code = str(e)
            _t.sleep(1.2)
            continue
    if resp is None:
        raise ScannerError(
            "Scanner kaafi der se busy bata raha hai (HTTP %s). Agar koi AUR scan app "
            "(NAPS2 / HP Scan / purani ApneScan window) khuli hai to use band karein, "
            "ya scanner ko ek baar off/on kar dein." % last_code)

    job = resp.headers.get("Location")
    if not job:
        raise ScannerError("eSCL job location nahi mila (scanner ne job start nahi kiya).")
    if job.startswith("/"):
        job = "http://%s%s" % (ip, job)

    count = 0
    empties = 0
    try:
        while True:
            if should_stop and should_stop():
                break
            try:
                r = _u.urlopen(job + "/NextDocument", timeout=90)
            except _ue.HTTPError as e:
                if e.code in (404, 410):
                    break                       # no more pages -> done
                if e.code == 503:
                    _t.sleep(1); empties += 1
                    if empties > 30:
                        break
                    continue
                break
            except Exception:
                break
            data = r.read()
            if not data:
                break
            try:
                with Image.open(_io.BytesIO(data)) as im:
                    img = im.convert("RGB").copy()
            except Exception:
                break
            count += 1
            if on_page:
                on_page(img)
    finally:
        try:
            _u.urlopen(_u.Request(job, method="DELETE"), timeout=10)
        except Exception:
            pass

    if count == 0:
        raise ScannerError("eSCL se koi page nahi aaya. Feeder me kaagaz hai? Scanner ON hai?")
    return count


# Role for storing a custom (document-name) label on a thumbnail item.
TITLE_ROLE = int(QtCore.Qt.UserRole) + 1
NAMEKEY_ROLE = int(QtCore.Qt.UserRole) + 2   # normalized key for remembering saved names


def underscore_name(s):
    """Filename-friendly: spaces -> underscore, drop odd chars, collapse repeats."""
    s = re.sub(r"[^\w\s.\-]", "", s or "")
    s = re.sub(r"\s+", "_", s.strip())
    s = re.sub(r"_+", "_", s)
    return s.strip("_.")


def name_key(s):
    """Normalized key to match the 'same' document title across scans."""
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


# Document TYPE classifier — recognise WHAT the document is (by tell-tale keywords)
# instead of naming it after whatever text is printed on top (which is usually just
# the hospital name and would be the same for every page). Tuned for ECHS/RGHS
# hospital paperwork. Order = priority when tie-broken.
DOC_TYPES = [
    # Only STRONG, distinctive phrases (avoid generic words that appear on every
    # hospital form like self/spouse/patient/hospital). If nothing here matches
    # clearly, the page stays "Page N" (better than a wrong name).
    ("Discharge_Summary", ["discharge summary"]),
    ("Bill",              ["cash receipt", "final bill", "bill of supply", "tax invoice",
                           "bill no.", "receipt no.", "total amount", "net payable"]),
    ("Referral",          ["referral form", "referral slip", "individual referred",
                           "referred hospital", "empanelled hospital"]),
    ("ECHS_Card",         ["semi-pvt", "echs card", "echs smart card", "ex-serviceman card",
                           "esm :"]),
    ("Aadhaar",           ["unique identification authority", "aadhaar", "आधार",
                           "government of india"]),
    ("Prescription",      ["provisional diagnosis", "chief complaint", "on examination",
                           "treatment advised", "diagnosis :", "complaints :"]),
    ("Lab_Report",        ["laboratory report", "pathology report", "haematology",
                           "biochemistry", "specimen", "sample collected", "test result"]),
    ("Xray_Radiology",    ["x-ray", "radiology report", "mri scan", "ct scan",
                           "ultrasound report", "sonography", "impression :"]),
    ("Registration",      ["registration no.", "uhid no", "op registration"]),
    ("Consent_Form",      ["consent form", "i hereby give my consent", "informed consent"]),
    ("Discharge_Card",    ["discharge card"]),
]


def page_ocr_text(path, frac=0.55):
    """OCR the top `frac` of a page and return the raw text (used by both the
    document-type classifier and the learned-name matcher)."""
    if not HAS_OCR_LIBS:
        return ""
    try:
        with Image.open(path) as im:
            w, h = im.size
            crop = im.crop((0, 0, w, max(1, int(h * frac)))).convert("L")
            if crop.width > 1500:
                r = 1500.0 / crop.width
                crop = crop.resize((1500, max(1, int(crop.height * r))))
        try:
            return pytesseract.image_to_string(crop, lang="eng+hin")
        except Exception:
            return pytesseract.image_to_string(crop)
    except Exception:
        return ""


def sig_words(text):
    """Distinctive words that identify a form's FIXED printed layout (letterhead,
    field labels, form title). Patient-specific data (names, numbers) mostly drops
    out, so two scans of the SAME form share most of these words."""
    words = re.findall(r"[a-z]{4,}", (text or "").lower())
    stop = {"name", "date", "time", "age", "years", "year", "male", "female",
            "address", "sign", "signature", "page", "self", "spouse", "mobile",
            "phone", "number", "the", "and", "for", "with", "this", "that"}
    return set(w for w in words if w not in stop)


def _jaccard(a, b):
    if not a or not b:
        return 0.0
    inter = len(a & b)
    union = len(a | b)
    return inter / union if union else 0.0


def learned_name_for(text, learned, thr=0.45):
    """learned = list of (wordset, name). Return the name whose stored signature
    best overlaps this page's words (>= thr), else None."""
    sig = sig_words(text)
    if not sig:
        return None
    best_name, best_score = None, 0.0
    for words, name in learned:
        sc = _jaccard(sig, words)
        if sc > best_score:
            best_score, best_name = sc, name
    return best_name if best_score >= thr else None


def classify_from_text(text):
    t = " " + re.sub(r"\s+", " ", (text or "").lower()) + " "
    scores = []
    for name, kws in DOC_TYPES:
        n = sum(1 for kw in kws if kw in t)
        if n > 0:
            scores.append((n, name))
    if not scores:
        return None
    scores.sort(reverse=True)
    # Accept only a CLEAR winner (top strictly beats runner-up); ties -> None.
    if len(scores) >= 2 and scores[0][0] == scores[1][0]:
        return None
    return scores[0][1]


def classify_document(path):
    """Return a clean document-TYPE name (e.g. Discharge_Summary, Bill, ECHS_Card)
    by matching keywords in the page text. None if no type is confidently found."""
    if not HAS_OCR_LIBS:
        return None
    return classify_from_text(page_ocr_text(path, 0.6))


def ocr_page_title(path):
    """Read the top of a scanned page and return a likely PRINTED document title
    (e.g. 'DISCHARGE SUMMARY'). Uses OCR word-confidence so low-confidence
    handwriting/garbage is rejected. None if nothing reliable is found."""
    if not HAS_OCR_LIBS:
        return None

    def _clean(ln):
        ln = re.sub(r"[^A-Za-z0-9 &.\-/,]", "", ln)
        return re.sub(r"\s+", " ", ln).strip()

    def _looks_garbage(ln):
        # a single alphabetic run longer than 15 chars with no space = OCR garbage
        for tok in ln.split():
            if tok.isalpha() and len(tok) > 15:
                return True
        return False

    def _title_from(fraction):
        try:
            with Image.open(path) as im:
                w, h = im.size
                strip = im.crop((0, 0, w, max(1, int(h * fraction)))).convert("L")
                if strip.width > 1400:
                    r = 1400.0 / strip.width
                    strip = strip.resize((1400, max(1, int(strip.height * r))))
            # Confidence-aware: build lines only from words OCR is confident about.
            try:
                try:
                    data = pytesseract.image_to_data(strip, lang="eng+hin",
                                                     output_type=pytesseract.Output.DICT)
                except Exception:
                    data = pytesseract.image_to_data(strip,
                                                     output_type=pytesseract.Output.DICT)
                lines = {}
                for i in range(len(data.get("text", []))):
                    txt = (data["text"][i] or "").strip()
                    try:
                        conf = float(data["conf"][i])
                    except Exception:
                        conf = -1
                    if not txt or conf < 55:
                        continue
                    if sum(c.isalnum() for c in txt) < 2:
                        continue
                    key = (data["block_num"][i], data["par_num"][i], data["line_num"][i])
                    lines.setdefault(key, []).append((txt, conf))
                cands = []
                for words in lines.values():
                    text = _clean(" ".join(wtc[0] for wtc in words))
                    if len(text) < 4 or _looks_garbage(text):
                        continue
                    avg = sum(wtc[1] for wtc in words) / len(words)
                    cands.append((text, avg))
                if cands:
                    def score(tc):
                        ln, cf = tc
                        letters = sum(c.isalpha() for c in ln)
                        caps = sum(1 for c in ln if c.isupper())
                        return cf * 0.5 + letters + caps * 0.4 - abs(len(ln) - 22) * 0.05
                    best = max(cands, key=score)[0]
                    if sum(c.isalpha() for c in best) >= 3:
                        return best[:38]
            except Exception:
                pass
            return None
        except Exception:
            return None

    return _title_from(0.22) or _title_from(0.45)


class NameWorker(QtCore.QThread):
    named = QtCore.pyqtSignal(int, str)   # (row, title)
    finished_all = QtCore.pyqtSignal()

    def __init__(self, items, learned=None):
        super().__init__()
        self.items = items                # list of (row, path)
        self.learned = learned or []      # list of (wordset, name)

    def run(self):
        for row, path in self.items:
            if self.isInterruptionRequested():
                break
            text = page_ocr_text(path, 0.6)
            # priority: name YOU taught it > document-type > printed heading
            title = (learned_name_for(text, self.learned)
                     or classify_from_text(text)
                     or ocr_page_title(path))
            if title:
                self.named.emit(row, title)
        self.finished_all.emit()


class ScanWorker(QtCore.QThread):
    page_done = QtCore.pyqtSignal(str)
    done = QtCore.pyqtSignal(int, int)
    failed = QtCore.pyqtSignal(str)

    def __init__(self, hwnd, source_name, dpi, pixel_type, duplex, tmpdir, opts):
        super().__init__()
        self.hwnd = hwnd
        self.source_name = source_name
        self.dpi = dpi
        self.pixel_type = pixel_type
        self.duplex = duplex
        self.tmpdir = tmpdir
        self.opts = opts or {}
        self.kept = 0
        self.skipped = 0

    def run(self):
        import queue as _q, threading as _th
        pageq = _q.Queue()

        def _consumer():
            # Runs in parallel: saves/processes each page while the scanner
            # is already pulling the NEXT one (this removes the per-page pause).
            while True:
                img = pageq.get()
                if img is None:
                    break
                try:
                    if self.opts.get("remove_blank"):
                        _sens = {"kam": 0.0003, "normal": 0.0008, "zyada": 0.004}
                        _thr = _sens.get(self.opts.get("blank_sensitivity", "normal"), 0.0008)
                        if is_blank_page(img, _thr):
                            self.skipped += 1
                            continue
                    if self.opts.get("page_size", "auto").startswith("auto"):
                        img = whiten_dark_background(img)
                    if self.opts.get("auto_crop"):
                        img = autocrop(img)
                    if self.opts.get("deskew"):
                        img = deskew(img)
                    if self.opts.get("quality_enhance"):
                        img = auto_enhance(img)
                    if self.pixel_type == "bw":
                        fd, out = tempfile.mkstemp(suffix=".png", dir=self.tmpdir)
                        os.close(fd)
                        try:
                            img.save(out, "PNG", compress_level=1)
                        except Exception:
                            img.save(out, "PNG")
                    else:
                        fd, out = tempfile.mkstemp(suffix=".jpg", dir=self.tmpdir)
                        os.close(fd)
                        try:
                            if img.mode == "L":
                                img.save(out, "JPEG", quality=88)
                            else:
                                img.convert("RGB").save(out, "JPEG", quality=88)
                        except Exception:
                            img.convert("RGB").save(out, "JPEG", quality=85)
                    self.kept += 1
                    self.page_done.emit(out)
                except Exception:
                    pass

        consumer = _th.Thread(target=_consumer, daemon=True)
        consumer.start()

        def _on_page(img):
            # Producer: hand the page off instantly and let the scanner keep going.
            pageq.put(img)

        err = None
        method = self.opts.get("scanner_method", "twain")
        try:
            if method == "escl":
                scan_via_escl(self.opts.get("scanner_ip"), self.dpi, self.pixel_type,
                              self.duplex, _on_page, should_stop=self.isInterruptionRequested,
                              page_size=self.opts.get("page_size", "auto"))
            elif method == "naps2":
                scan_via_naps2(self.opts.get("naps2_path") or find_naps2(),
                               self.opts.get("naps2_profile"), self.tmpdir,
                               self.duplex, _on_page,
                               should_stop=self.isInterruptionRequested)
            elif method == "wia":
                try:
                    _pythoncom.CoInitialize()
                except Exception:
                    pass
                try:
                    wia_scan_pages(self.opts.get("wia_device_id"), self.dpi,
                                   self.pixel_type, self.duplex, _on_page,
                                   should_stop=self.isInterruptionRequested)
                finally:
                    try:
                        _pythoncom.CoUninitialize()
                    except Exception:
                        pass
            else:
                scan_pages(self.hwnd, self.source_name, self.dpi,
                           self.pixel_type, self.duplex, _on_page,
                           should_stop=self.isInterruptionRequested)
        except ScannerError as exc:
            err = str(exc)
        except Exception:
            err = "Scan error:\n%s" % traceback.format_exc()

        pageq.put(None)            # tell the saver thread to finish
        consumer.join(timeout=60)  # wait until all pages are saved

        if err is not None and self.kept == 0:
            self.failed.emit(err)
            return
        if self.kept == 0 and self.skipped == 0:
            self.failed.emit(err or "Koi page scan nahi hua.")
            return
        self.done.emit(self.kept, self.skipped)


# ---------------------------------------------------------------------------
# Dialogs
# ---------------------------------------------------------------------------

class EditProfileDialog(QtWidgets.QDialog):
    def __init__(self, parent, profile=None):
        super().__init__(parent)
        self.setWindowTitle("Profile settings")
        self.setMinimumWidth(420)
        self.profile = dict(profile) if profile else {
            "name": "", "source_name": None, "dpi": 200, "color": "gray", "duplex": False}
        form = QtWidgets.QFormLayout(self)
        self.name_edit = QtWidgets.QLineEdit(self.profile.get("name", ""))
        self.name_edit.setPlaceholderText("jaise: Documents fast")
        form.addRow("Display Name:", self.name_edit)
        dev_row = QtWidgets.QHBoxLayout()
        self.device_label = QtWidgets.QLabel(self.profile.get("source_name") or "(koi device nahi)")
        btn_dev = QtWidgets.QPushButton("Choose device")
        btn_dev.clicked.connect(self._choose_device)
        dev_row.addWidget(self.device_label, 1); dev_row.addWidget(btn_dev)
        dw = QtWidgets.QWidget(); dw.setLayout(dev_row)
        form.addRow("Device:", dw)
        self.cmb_dpi = QtWidgets.QComboBox(); self.cmb_dpi.addItems(RESOLUTIONS)
        self.cmb_dpi.setCurrentText(str(self.profile.get("dpi", 200)))
        form.addRow("Resolution (DPI):", self.cmb_dpi)
        self.cmb_color = QtWidgets.QComboBox(); self.cmb_color.addItems(list(COLOUR_MODES.keys()))
        for label, code in COLOUR_MODES.items():
            if code == self.profile.get("color", "gray"):
                self.cmb_color.setCurrentText(label)
        form.addRow("Colour mode:", self.cmb_color)
        self.chk_duplex = QtWidgets.QCheckBox("Duplex (dono taraf)")
        self.chk_duplex.setChecked(bool(self.profile.get("duplex")))
        form.addRow("", self.chk_duplex)
        self.cmb_psize = QtWidgets.QComboBox()
        self._PSIZES = ["Auto (alag-alag size khud pakde)", "A4 (210x297 mm)", "Letter", "Legal", "A5"]
        self.cmb_psize.addItems(self._PSIZES)
        _ps = (self.profile.get("page_size") or "auto").lower()
        _idx = next((k for k, s in enumerate(self._PSIZES) if s.lower().startswith(_ps[:4])), 0)
        self.cmb_psize.setCurrentIndex(_idx)
        form.addRow("Page size:", self.cmb_psize)
        btns = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        btns.accepted.connect(self._ok); btns.rejected.connect(self.reject)
        form.addRow(btns)

    def _choose_device(self):
        if not HAS_TWAIN:
            QtWidgets.QMessageBox.warning(self, "Error", "TWAIN install nahi hai."); return
        try:
            names = list_sources(int(self.winId()))
        except Exception as exc:
            QtWidgets.QMessageBox.warning(self, "Error", "Scanner list nahi mili:\n%s" % exc); return
        if not names:
            QtWidgets.QMessageBox.warning(self, "Error", "Koi scanner nahi mila."); return
        name, ok = QtWidgets.QInputDialog.getItem(self, "Device chuno", "Scanner:", names, 0, False)
        if ok and name:
            self.profile["source_name"] = name; self.device_label.setText(name)

    def _ok(self):
        if not self.name_edit.text().strip():
            QtWidgets.QMessageBox.warning(self, "Error", "Display Name daalein."); return
        self.accept()

    def get_profile(self):
        self.profile["name"] = self.name_edit.text().strip() or "Profile"
        self.profile["dpi"] = int(self.cmb_dpi.currentText())
        self.profile["color"] = COLOUR_MODES[self.cmb_color.currentText()]
        self.profile["duplex"] = self.chk_duplex.isChecked()
        self.profile["page_size"] = self.cmb_psize.currentText().strip().lower()
        return self.profile


class ProfileManagerDialog(QtWidgets.QDialog):
    def __init__(self, parent, profiles):
        super().__init__(parent)
        self.setWindowTitle("Profiles"); self.resize(420, 320)
        self.profiles = [dict(p) for p in profiles]
        lay = QtWidgets.QVBoxLayout(self)
        lay.addWidget(QtWidgets.QLabel("Scan profiles:"))
        self.listw = QtWidgets.QListWidget()
        self.listw.itemDoubleClicked.connect(lambda *_: self._edit())
        lay.addWidget(self.listw, 1); self._refresh()
        row = QtWidgets.QHBoxLayout()
        for t, s in [("New", self._new), ("Edit", self._edit), ("Delete", self._delete)]:
            b = QtWidgets.QPushButton(t); b.clicked.connect(s); row.addWidget(b)
        row.addStretch(1)
        done = QtWidgets.QPushButton("Done"); done.setObjectName("primary")
        done.clicked.connect(self.accept); row.addWidget(done)
        lay.addLayout(row)

    def _refresh(self):
        self.listw.clear()
        for p in self.profiles:
            self.listw.addItem(p.get("name", "Profile"))

    def _new(self):
        dlg = EditProfileDialog(self)
        if dlg.exec_() == QtWidgets.QDialog.Accepted:
            self.profiles.append(dlg.get_profile()); self._refresh()
            self.listw.setCurrentRow(len(self.profiles) - 1)

    def _edit(self):
        i = self.listw.currentRow()
        if i < 0:
            return
        dlg = EditProfileDialog(self, self.profiles[i])
        if dlg.exec_() == QtWidgets.QDialog.Accepted:
            self.profiles[i] = dlg.get_profile(); self._refresh(); self.listw.setCurrentRow(i)

    def _delete(self):
        i = self.listw.currentRow()
        if i < 0:
            return
        del self.profiles[i]; self._refresh()


class OptionsDialog(QtWidgets.QDialog):
    def __init__(self, parent, opts):
        super().__init__(parent)
        self.setWindowTitle("Options / Settings")
        self.setMinimumWidth(560)
        self.opts = dict(DEFAULT_OPTIONS); self.opts.update(opts or {})

        outer = QtWidgets.QVBoxLayout(self)
        scroll = QtWidgets.QScrollArea(); scroll.setWidgetResizable(True)
        inner = QtWidgets.QWidget(); form = QtWidgets.QFormLayout(inner)
        scroll.setWidget(inner); outer.addWidget(scroll, 1)

        def header(t):
            lbl = QtWidgets.QLabel("<b>%s</b>" % t); form.addRow(lbl)

        def qhelp(tip):
            h = QtWidgets.QLabel("?")
            h.setToolTip(tip); h.setCursor(QtCore.Qt.WhatsThisCursor)
            h.setStyleSheet("QLabel{color:#0f766e; border:1px solid #0f766e; border-radius:9px;"
                            "min-width:18px; max-width:18px; min-height:18px; max-height:18px;"
                            "font-weight:bold; qproperty-alignment:AlignCenter;}")
            return h

        def chkrow(chk, tip):
            chk.setToolTip(tip)
            row = QtWidgets.QHBoxLayout(); row.setContentsMargins(0, 0, 0, 0); row.setSpacing(6)
            row.addWidget(chk); row.addWidget(qhelp(tip)); row.addStretch(1)
            w = QtWidgets.QWidget(); w.setLayout(row); return w

        def lblhelp(text, tip):
            row = QtWidgets.QHBoxLayout(); row.setContentsMargins(0, 0, 0, 0); row.setSpacing(6)
            row.addWidget(QtWidgets.QLabel(text)); row.addWidget(qhelp(tip)); row.addStretch(1)
            w = QtWidgets.QWidget(); w.setLayout(row); return w

        header("Interface")
        self.cmb_theme = QtWidgets.QComboBox(); self.cmb_theme.addItems(["Light", "Dark"])
        self.cmb_theme.setCurrentIndex(1 if self.opts.get("theme") == "dark" else 0)
        form.addRow(lblhelp("Theme:", 'हिन्दी: App ka rang-roop. Dark = kaala theme (aankhon ko aaram, raat me achha). Light = safed.\nEnglish: App look. Dark = dark theme (easy on eyes at night). Light = white.'), self.cmb_theme)
        self.chk_pagenum = QtWidgets.QCheckBox("Page number thumbnails ke neeche dikhao")
        self.chk_pagenum.setChecked(self.opts.get("show_page_numbers", True)); form.addRow(chkrow(self.chk_pagenum, "हिन्दी: Har thumbnail ke neeche 'Page 1, 2..' dikhana. Band karne par number nahi dikhenge.\nEnglish: Show 'Page 1, 2..' under each thumbnail. Off = no numbers."))
        self.chk_autoname = QtWidgets.QCheckBox("Document ka naam apne aap padho (OCR)")
        self.chk_autoname.setChecked(self.opts.get("auto_name", False))
        form.addRow(chkrow(self.chk_autoname, "हिन्दी: ON: scan ke baad har page ka naam (document ka title, jaise DISCHARGE SUMMARY / RECEIPT) apne aap padh kar likh dega — 'Page 1,2' ke bajay. (OCR/Tesseract chahiye.)\nEnglish: ON: after scanning, auto-labels each page with its document title (needs OCR)."))

        header("Save")
        self.chk_autosave = QtWidgets.QCheckBox("Scan ke baad PDF apne aap save karo")
        self.chk_autosave.setChecked(self.opts["auto_save"]); form.addRow(chkrow(self.chk_autosave, 'हिन्दी: ON: scan khatam hote hi PDF apne aap save ho jayegi (har baar save dabana nahi padega). Roz bahut scan karte ho to ON rakho.\nEnglish: ON: PDF saves automatically after each scan. Handy if you scan a lot.'))
        fr = QtWidgets.QHBoxLayout()
        self.folder_edit = QtWidgets.QLineEdit(self.opts["save_folder"])
        b = QtWidgets.QPushButton("…"); b.setFixedWidth(36); b.clicked.connect(self._pick_folder)
        fr.addWidget(self.folder_edit, 1); fr.addWidget(b)
        fw = QtWidgets.QWidget(); fw.setLayout(fr); form.addRow(lblhelp("Save folder:", 'हिन्दी: PDF kahaan save hon wo folder.\nEnglish: Folder where PDFs get saved.'), fw)
        self.tmpl_edit = QtWidgets.QLineEdit(self.opts["filename_template"])
        form.addRow(lblhelp("Filename template:", 'हिन्दी: File ka naam kaise bane. {claim}=claim number, {date}=tareekh, {time}=samay, {seq}=kram sankhya.\nEnglish: How filenames are built: {claim} {date} {time} {seq}.'), self.tmpl_edit)
        form.addRow("", QtWidgets.QLabel("Tags: {claim} {date} {time} {seq}"))
        self.chk_claimfolder = QtWidgets.QCheckBox("Claim number ka alag folder")
        self.chk_claimfolder.setChecked(self.opts["make_claim_folder"]); form.addRow(chkrow(self.chk_claimfolder, 'हिन्दी: ON: har claim number ka alag folder banega (ek claim ke saare pages ek jagah).\nEnglish: ON: a separate folder per claim number.'))
        self.chk_ymfolder = QtWidgets.QCheckBox("Saal/Mahine ke folder (2026/07/...)")
        self.chk_ymfolder.setChecked(self.opts["year_month_folders"]); form.addRow(chkrow(self.chk_ymfolder, 'हिन्दी: ON: saal/mahine ke folder (2026/07/...) — purane scans aasani se milen.\nEnglish: ON: year/month folders (2026/07/...) for easy filing.'))
        self.cmb_after = QtWidgets.QComboBox(); self.cmb_after.addItems(["Kuch nahi", "PDF kholo", "Folder kholo"])
        self.cmb_after.setCurrentIndex({"nothing": 0, "open": 1, "folder": 2}.get(self.opts.get("after_save", "nothing"), 0))
        form.addRow(lblhelp("Save ke baad:", 'हिन्दी: Save ke baad kya ho — kuch nahi / PDF khule / folder khule.\nEnglish: After save — do nothing / open the PDF / open the folder.'), self.cmb_after)
        self.chk_imgtoo = QtWidgets.QCheckBox("Har page ki alag image (JPG) bhi save karo")
        self.chk_imgtoo.setChecked(self.opts.get("save_images_too", False)); form.addRow(chkrow(self.chk_imgtoo, 'हिन्दी: ON: PDF ke saath har page ki alag JPG image bhi banegi.\nEnglish: ON: also save each page as a separate JPG image.'))

        header("Image cleanup")
        self.chk_blank = QtWidgets.QCheckBox("Khaali (blank) pages hatao")
        self.chk_blank.setChecked(self.opts["remove_blank"]); form.addRow(chkrow(self.chk_blank, 'हिन्दी: ON: khaali (blank) pages apne aap hat jayenge — jaise duplex me peeche ka khaali side. (NAPS2 jaisa.)\nEnglish: ON: blank pages are removed automatically (e.g. blank back side in duplex).'))
        self.cmb_blank_sens = QtWidgets.QComboBox(); self.cmb_blank_sens.addItems(["Kam (safe)", "Normal", "Zyada (aggressive)"])
        self.cmb_blank_sens.setCurrentIndex({"kam": 0, "normal": 1, "zyada": 2}.get(self.opts.get("blank_sensitivity", "normal"), 1))
        form.addRow(lblhelp("Blank hatane ki sensitivity:", 'हिन्दी: Kitni sakhti se blank hataye. "Zyada" = fold-line/halke stamp wale peeche ke khaali page bhi hat jayenge. "Kam" = sirf bilkul khaali. Normal beech ka.\nEnglish: How aggressively to drop blanks. Zyada = also removes back sides with fold lines/faint marks; Kam = only truly empty; Normal = balanced.'), self.cmb_blank_sens)
        self.chk_crop = QtWidgets.QCheckBox("Border auto-crop")
        self.chk_crop.setChecked(self.opts["auto_crop"]); form.addRow(chkrow(self.chk_crop, 'हिन्दी: ON: page ke aas-paas ki khaali safed border apne aap kat jayegi.\nEnglish: ON: auto-trims the empty white border around the page.'))
        self.chk_deskew = QtWidgets.QCheckBox("Auto-deskew (tedha seedha)")
        self.chk_deskew.setChecked(self.opts["deskew"])
        if not HAS_NUMPY:
            self.chk_deskew.setEnabled(False)
            self.chk_deskew.setText("Auto-deskew (numpy install karein)")
        form.addRow(chkrow(self.chk_deskew, 'हिन्दी: ON: tedha scan hua page apne aap seedha ho jayega.\nEnglish: ON: straightens a tilted/skewed page automatically.'))
        self.chk_enhance = QtWidgets.QCheckBox("Quality auto-sudhar (faded documents saaf)")
        self.chk_enhance.setChecked(self.opts["quality_enhance"]); form.addRow(chkrow(self.chk_enhance, 'हिन्दी: ON: feeke/halke documents saaf aur gehre dikhenge.\nEnglish: ON: brightens & sharpens faded documents.'))

        header("Output")
        self.chk_compress = QtWidgets.QCheckBox("PDF compress (chhoti file)")
        self.chk_compress.setChecked(self.opts["compress"]); form.addRow(chkrow(self.chk_compress, 'हिन्दी: ON: PDF ki file chhoti banegi (email/upload aasan); quality thodi kam.\nEnglish: ON: smaller PDF (easier to email/upload); slightly lower quality.'))
        self.q_spin = QtWidgets.QSpinBox(); self.q_spin.setRange(20, 95)
        self.q_spin.setValue(int(self.opts["jpeg_quality"]))
        form.addRow(lblhelp("Compress quality:", 'हिन्दी: Compress ki quality (zyada = behtar pic par badi file). 60-80 theek hai.\nEnglish: Compression quality (higher = better image, larger file). 60-80 is fine.'), self.q_spin)
        self.chk_wm = QtWidgets.QCheckBox("Watermark/stamp har page par")
        self.chk_wm.setChecked(self.opts["watermark"]); form.addRow(chkrow(self.chk_wm, 'हिन्दी: ON: har page par aapka text/stamp (jaise hospital ka naam) chhapega.\nEnglish: ON: stamps your text (e.g. hospital name) on every page.'))
        self.wm_edit = QtWidgets.QLineEdit(self.opts["watermark_text"])
        form.addRow(lblhelp("Watermark text:", 'हिन्दी: Watermark me kya likha ho.\nEnglish: The watermark text.'), self.wm_edit)

        header("Workflow")
        self.chk_batch = QtWidgets.QCheckBox("Batch mode (save ke baad next claim ready)")
        self.chk_batch.setChecked(self.opts["batch_mode"]); form.addRow(chkrow(self.chk_batch, 'हिन्दी: ON: ek claim save hote hi agla claim scan karne ke liye screen saaf ho jayegi (tezi se ek ke baad ek).\nEnglish: ON: after saving, screen clears for the next claim (fast back-to-back).'))
        self.chk_validate = QtWidgets.QCheckBox("Claim number validate karo")
        self.chk_validate.setChecked(self.opts["validate_claim"]); form.addRow(chkrow(self.chk_validate, 'हिन्दी: ON: galat/adhoora claim number daalne par chetavni dega.\nEnglish: ON: warns if the claim number looks wrong/incomplete.'))
        self.pat_edit = QtWidgets.QLineEdit(self.opts["claim_pattern"])
        form.addRow(lblhelp("Claim pattern (regex):", 'हिन्दी: Claim number ka sahi roop (regex). Aam taur par badalne ki zaroorat nahi.\nEnglish: The valid claim-number pattern (regex). Usually leave as is.'), self.pat_edit)
        self.chk_barcode = QtWidgets.QCheckBox("Barcode/QR se claim number khud bharo")
        self.chk_barcode.setChecked(self.opts["barcode_autofill"])
        if not HAS_ZBAR:
            self.chk_barcode.setEnabled(False)
            self.chk_barcode.setText("Barcode/QR (pyzbar install karein)")
        form.addRow(chkrow(self.chk_barcode, 'हिन्दी: ON: page par barcode/QR ho to claim number apne aap bhar jayega.\nEnglish: ON: auto-fills the claim number from a barcode/QR on the page.'))
        self.chk_dup = QtWidgets.QCheckBox("Duplicate claim number par chetavni")
        self.chk_dup.setChecked(self.opts["duplicate_check"]); form.addRow(chkrow(self.chk_dup, 'हिन्दी: ON: wahi claim number dobara ho to chetavni (double entry se bachav).\nEnglish: ON: warns on a duplicate claim number.'))

        header("Records & safety")
        self.chk_excel = QtWidgets.QCheckBox("Excel register me entry (register.xlsx)")
        self.chk_excel.setChecked(self.opts["excel_log"])
        if not HAS_XLSX:
            self.chk_excel.setEnabled(False)
            self.chk_excel.setText("Excel register (openpyxl install karein)")
        form.addRow(chkrow(self.chk_excel, 'हिन्दी: ON: har scan ki entry ek Excel register (register.xlsx) me judegi — record ke liye.\nEnglish: ON: logs each scan into an Excel register (register.xlsx).'))
        self.chk_log = QtWidgets.QCheckBox("Activity log (activity_log.txt)")
        self.chk_log.setChecked(self.opts["activity_log"]); form.addRow(chkrow(self.chk_log, 'हिन्दी: ON: kab kya scan/save hua iska text log rakhega.\nEnglish: ON: keeps a text activity log.'))
        self.chk_backup = QtWidgets.QCheckBox("Har save ka backup copy")
        self.chk_backup.setChecked(self.opts["backup"]); form.addRow(chkrow(self.chk_backup, 'हिन्दी: ON: har save ki ek backup copy alag folder me bhi rakhega (surakhsha).\nEnglish: ON: keeps a backup copy of every save in another folder.'))
        br = QtWidgets.QHBoxLayout()
        self.backup_edit = QtWidgets.QLineEdit(self.opts["backup_folder"])
        bb = QtWidgets.QPushButton("…"); bb.setFixedWidth(36); bb.clicked.connect(self._pick_backup)
        br.addWidget(self.backup_edit, 1); br.addWidget(bb)
        bw = QtWidgets.QWidget(); bw.setLayout(br); form.addRow(lblhelp("Backup folder:", 'हिन्दी: Backup kahaan rakhe wo folder.\nEnglish: Folder for backups.'), bw)

        btns = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept); btns.rejected.connect(self.reject)
        outer.addWidget(btns)

    def _pick_folder(self):
        d = QtWidgets.QFileDialog.getExistingDirectory(self, "Folder", self.folder_edit.text())
        if d:
            self.folder_edit.setText(d)

    def _pick_backup(self):
        d = QtWidgets.QFileDialog.getExistingDirectory(self, "Backup folder", self.backup_edit.text())
        if d:
            self.backup_edit.setText(d)

    def get_opts(self):
        o = self.opts
        o["theme"] = "dark" if self.cmb_theme.currentIndex() == 1 else "light"
        o["show_page_numbers"] = self.chk_pagenum.isChecked()
        o["auto_name"] = self.chk_autoname.isChecked()
        o["save_images_too"] = self.chk_imgtoo.isChecked()
        o["auto_save"] = self.chk_autosave.isChecked()
        o["save_folder"] = self.folder_edit.text().strip()
        o["filename_template"] = self.tmpl_edit.text().strip() or "{date}_{seq}"
        o["make_claim_folder"] = self.chk_claimfolder.isChecked()
        o["after_save"] = ["nothing", "open", "folder"][self.cmb_after.currentIndex()]
        o["year_month_folders"] = self.chk_ymfolder.isChecked()
        o["remove_blank"] = self.chk_blank.isChecked()
        o["blank_sensitivity"] = {0: "kam", 1: "normal", 2: "zyada"}.get(self.cmb_blank_sens.currentIndex(), "normal")
        o["auto_crop"] = self.chk_crop.isChecked()
        o["deskew"] = self.chk_deskew.isChecked()
        o["quality_enhance"] = self.chk_enhance.isChecked()
        o["compress"] = self.chk_compress.isChecked()
        o["jpeg_quality"] = int(self.q_spin.value())
        o["watermark"] = self.chk_wm.isChecked()
        o["watermark_text"] = self.wm_edit.text().strip() or "Noble Care Hospital"
        o["batch_mode"] = self.chk_batch.isChecked()
        o["validate_claim"] = self.chk_validate.isChecked()
        o["claim_pattern"] = self.pat_edit.text().strip() or r"^[A-Za-z0-9\-]{4,}$"
        o["barcode_autofill"] = self.chk_barcode.isChecked()
        o["duplicate_check"] = self.chk_dup.isChecked()
        o["excel_log"] = self.chk_excel.isChecked()
        o["activity_log"] = self.chk_log.isChecked()
        o["backup"] = self.chk_backup.isChecked()
        o["backup_folder"] = self.backup_edit.text().strip()
        return o


# ---------------------------------------------------------------------------
# Main Window
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Setup wizard + Help (public-friendly)
# ---------------------------------------------------------------------------

class ScanProgressDialog(QtWidgets.QDialog):
    """NAPS2-style scanning box: page counter + Run in Background + Cancel."""
    cancelled = QtCore.pyqtSignal()

    def __init__(self, parent, title, lang="hi"):
        super().__init__(parent)
        self._lang = lang
        self.setWindowTitle(title or "Scanning")
        self.setModal(False)
        self.setMinimumWidth(430)
        lay = QtWidgets.QVBoxLayout(self)
        self.lbl = QtWidgets.QLabel(self._page_text(1))
        lay.addWidget(self.lbl)
        self.bar = QtWidgets.QProgressBar(); self.bar.setRange(0, 0)
        lay.addWidget(self.bar)
        row = QtWidgets.QHBoxLayout(); row.addStretch(1)
        self.btn_bg = QtWidgets.QPushButton("Run in Background" if lang == "en" else "Background me chalao")
        self.btn_bg.clicked.connect(self.hide)
        self.btn_cancel = QtWidgets.QPushButton("Cancel")
        self.btn_cancel.clicked.connect(lambda: self.cancelled.emit())
        row.addWidget(self.btn_bg); row.addWidget(self.btn_cancel)
        lay.addLayout(row)
        self.show()

    def _page_text(self, n):
        return ("Scanning page %d..." % n) if self._lang == "en" else ("Page %d scan ho raha hai..." % n)

    def set_page(self, n):
        self.lbl.setText(self._page_text(n))

    def closeEvent(self, e):
        e.accept()


class SetupWizard(QtWidgets.QWizard):
    def __init__(self, parent, lang="hi"):
        super().__init__(parent)
        self.lang = lang
        self.result_method = "twain"
        self.result_device_name = None
        self.result_wia_id = None
        self.setWindowTitle("Setup — ApneScan")
        self.resize(580, 440)
        self.addPage(self._welcome())
        self.addPage(self._device_page())
        self.addPage(self._finish())

    def _welcome(self):
        p = QtWidgets.QWizardPage()
        p.setTitle("Swagat hai" if self.lang == "hi" else "Welcome")
        lay = QtWidgets.QVBoxLayout(p)
        txt = ("Ye chhota setup aapke scanner ko jodega.\nNext dabao."
               if self.lang == "hi"
               else "This quick setup will connect your scanner.\nClick Next.")
        lbl = QtWidgets.QLabel(txt); lbl.setWordWrap(True); lay.addWidget(lbl)
        return p

    def _device_page(self):
        p = QtWidgets.QWizardPage()
        p.setTitle("Scanner chuno" if self.lang == "hi" else "Choose scanner")
        lay = QtWidgets.QVBoxLayout(p)
        self.rb_twain = QtWidgets.QRadioButton("TWAIN (zyadatar scanners)")
        self.rb_wia = QtWidgets.QRadioButton("WIA (Windows built-in)")
        self.rb_twain.setChecked(True)
        lay.addWidget(self.rb_twain); lay.addWidget(self.rb_wia)
        row = QtWidgets.QHBoxLayout()
        self.dev_lbl = QtWidgets.QLabel("(koi scanner nahi chuna)")
        btn = QtWidgets.QPushButton("Choose scanner"); btn.clicked.connect(self._choose)
        row.addWidget(self.dev_lbl, 1); row.addWidget(btn)
        w = QtWidgets.QWidget(); w.setLayout(row); lay.addWidget(w)
        hint = QtWidgets.QLabel(
            "Scanner list me na dikhe to doosra option (TWAIN/WIA) chuno."
            if self.lang == "hi"
            else "If your scanner isn't listed, try the other option (TWAIN/WIA).")
        hint.setWordWrap(True); lay.addWidget(hint)
        return p

    def _choose(self):
        if self.rb_wia.isChecked():
            self.result_method = "wia"
            try:
                devs = list_wia_sources()
            except Exception as exc:
                QtWidgets.QMessageBox.warning(self, "Error", friendly_error(exc, self.lang)); return
            if not devs:
                QtWidgets.QMessageBox.warning(self, "Error", "Koi WIA scanner nahi mila."); return
            names = [n for _i, n in devs]
            name, ok = QtWidgets.QInputDialog.getItem(self, "Scanner", "Scanner:", names, 0, False)
            if ok and name:
                for _id, n in devs:
                    if n == name:
                        self.result_wia_id = _id; self.result_device_name = n; break
                self.dev_lbl.setText(name)
        else:
            self.result_method = "twain"
            try:
                names = list_sources(int(self.winId()))
            except Exception as exc:
                QtWidgets.QMessageBox.warning(self, "Error", friendly_error(exc, self.lang)); return
            if not names:
                QtWidgets.QMessageBox.warning(self, "Error", "Koi TWAIN scanner nahi mila."); return
            name, ok = QtWidgets.QInputDialog.getItem(self, "Scanner", "Scanner:", names, 0, False)
            if ok and name:
                self.result_device_name = name; self.dev_lbl.setText(name)

    def _finish(self):
        p = QtWidgets.QWizardPage()
        p.setTitle("Ho gaya!" if self.lang == "hi" else "Done!")
        lay = QtWidgets.QVBoxLayout(p)
        lbl = QtWidgets.QLabel(
            "Finish dabao, phir 'Scan' dabao. Bas!"
            if self.lang == "hi" else "Click Finish, then press 'Scan'. That's it!")
        lbl.setWordWrap(True); lay.addWidget(lbl)
        return p


class HelpDialog(QtWidgets.QDialog):
    def __init__(self, parent, lang="hi"):
        super().__init__(parent)
        self.setWindowTitle("Guide / Help"); self.resize(620, 520)
        lay = QtWidgets.QVBoxLayout(self)
        br = QtWidgets.QTextBrowser()
        br.setHtml(HELP_TEXT_HI if lang == "hi" else HELP_TEXT_EN)
        lay.addWidget(br)
        b = QtWidgets.QPushButton("Band karo" if lang == "hi" else "Close")
        b.clicked.connect(self.accept); lay.addWidget(b)



def _make_icon(kind, color="#0f766e"):
    """Draw a small flat icon (28x28) for the toolbar. No external files needed."""
    from PyQt5 import QtGui, QtCore
    pm = QtGui.QPixmap(30, 30); pm.fill(QtCore.Qt.transparent)
    p = QtGui.QPainter(pm); p.setRenderHint(QtGui.QPainter.Antialiasing)
    pen = QtGui.QPen(QtGui.QColor(color)); pen.setWidth(2); pen.setJoinStyle(QtCore.Qt.RoundJoin)
    pen.setCapStyle(QtCore.Qt.RoundCap); p.setPen(pen)
    br = QtGui.QColor(color)
    def doc():
        p.drawRoundedRect(8, 5, 14, 20, 2, 2)
        p.drawLine(11, 11, 19, 11); p.drawLine(11, 15, 19, 15); p.drawLine(11, 19, 16, 19)
    if kind == "scan":
        doc(); p.setBrush(QtGui.QColor("#5eead4")); p.drawRoundedRect(5, 13, 20, 4, 2, 2)
    elif kind == "profiles":
        p.drawRoundedRect(6, 7, 18, 5, 2, 2); p.drawRoundedRect(6, 14, 18, 5, 2, 2); p.drawRoundedRect(6, 21, 12, 5, 2, 2)
    elif kind == "fast":
        p.setBrush(QtGui.QColor("#f59e0b")); p.setPen(QtGui.QColor("#f59e0b"))
        p.drawPolygon(QtGui.QPolygon([QtCore.QPoint(17,4),QtCore.QPoint(9,17),QtCore.QPoint(14,17),QtCore.QPoint(12,26),QtCore.QPoint(21,12),QtCore.QPoint(15,12)]))
    elif kind == "ocr":
        f = p.font(); f.setBold(True); f.setPixelSize(11); p.setFont(f); p.drawText(pm.rect(), QtCore.Qt.AlignCenter, "OCR")
    elif kind == "import":
        doc(); p.drawLine(15, 26, 15, 19); p.drawLine(12, 22, 15, 19); p.drawLine(18, 22, 15, 19)
    elif kind == "savepdf":
        doc(); f=p.font(); f.setBold(True); f.setPixelSize(7); p.setFont(f); p.drawText(QtCore.QRect(8,17,14,8), QtCore.Qt.AlignCenter, "PDF")
    elif kind == "images":
        p.drawRoundedRect(6, 8, 18, 14, 2, 2); p.setBrush(br); p.drawEllipse(10, 12, 3, 3)
        p.drawLine(8, 20, 13, 15); p.drawLine(13, 20, 18, 14)
    elif kind == "print":
        p.drawRoundedRect(8, 5, 14, 7, 1, 1); p.drawRoundedRect(5, 12, 20, 9, 2, 2); p.drawRoundedRect(9, 19, 12, 7, 1, 1)
    elif kind == "rotate":
        p.drawArc(7, 7, 16, 16, 30*16, 280*16); p.drawLine(22, 8, 22, 13); p.drawLine(22, 13, 17, 13)
    elif kind == "up":
        p.drawLine(15, 22, 15, 9); p.drawLine(15, 9, 10, 14); p.drawLine(15, 9, 20, 14)
    elif kind == "down":
        p.drawLine(15, 8, 15, 21); p.drawLine(15, 21, 10, 16); p.drawLine(15, 21, 20, 16)
    elif kind == "delete":
        p.drawLine(9, 9, 21, 21); p.drawLine(21, 9, 9, 21)
    elif kind == "clear":
        p.drawRoundedRect(8, 8, 14, 16, 2, 2); p.drawLine(6, 8, 24, 8); p.drawLine(13, 5, 17, 5)
    elif kind == "language":
        p.drawEllipse(6, 6, 18, 18); p.drawLine(6, 15, 24, 15); p.drawArc(11, 6, 8, 18, 0, 360*16)
    elif kind == "about":
        p.drawEllipse(6, 6, 18, 18); f=p.font(); f.setBold(True); f.setPixelSize(13); p.setFont(f); p.drawText(pm.rect(), QtCore.Qt.AlignCenter, "i")
    else:
        doc()
    p.end()
    return QtGui.QIcon(pm)


class PreviewDialog(QtWidgets.QDialog):
    """NAPS2-style full preview: navigate pages, zoom, rotate, rename, delete,
    save — all from one window."""
    def __init__(self, win, row):
        super().__init__(win)
        self.win = win
        self.row = row
        self.zoom = 1.0
        self.fit = True
        self.setWindowTitle("Preview")
        self.resize(940, 900)
        v = QtWidgets.QVBoxLayout(self)
        tb = QtWidgets.QHBoxLayout()

        def b(txt, fn, tip=""):
            btn = QtWidgets.QToolButton()
            btn.setText(txt); btn.setToolTip(tip)
            btn.setToolButtonStyle(QtCore.Qt.ToolButtonTextOnly)
            btn.clicked.connect(fn)
            btn.setStyleSheet("QToolButton{padding:5px 9px; font-size:13px;}")
            tb.addWidget(btn)
            return btn

        b("\u25c0", self.prev, "Pichhla page (Left arrow)")
        self.lbl_count = QtWidgets.QLabel(); self.lbl_count.setStyleSheet("font-weight:bold;")
        tb.addWidget(self.lbl_count)
        b("\u25b6", self.next, "Agla page (Right arrow)")
        tb.addSpacing(14)
        b("\u2796", self.zoom_out, "Zoom out ( - )")
        b("Fit", self.fit_view, "Poora page dikhao")
        b("\u2795", self.zoom_in, "Zoom in ( + )")
        b("100%", self.actual_size, "Asli size")
        tb.addSpacing(14)
        b("\u21ba", lambda: self.rotate(-90), "Baayein ghumao")
        b("\u21bb", lambda: self.rotate(90), "Dayein ghumao")
        b("\u2712 Rename", self.rename, "Naam badlo (F2)")
        b("\U0001f5d1 Delete", self.delete, "Is page ko hatao")
        b("\U0001f4be Save", self.save_one, "Sirf is page ko PDF me save karo")
        tb.addStretch(1)
        b("\u2715 Close", self.accept, "Band karo (Esc)")
        v.addLayout(tb)

        self.lbl = QtWidgets.QLabel(); self.lbl.setAlignment(QtCore.Qt.AlignCenter)
        self.scr = QtWidgets.QScrollArea(); self.scr.setWidgetResizable(False)
        self.scr.setAlignment(QtCore.Qt.AlignCenter); self.scr.setWidget(self.lbl)
        v.addWidget(self.scr, 1)
        QtCore.QTimer.singleShot(0, self._load)

    def _path(self):
        it = self.win.list.item(self.row)
        return it.data(QtCore.Qt.UserRole) if it else None

    def _load(self):
        n = self.win.list.count()
        if n == 0:
            self.accept(); return
        self.row = max(0, min(self.row, n - 1))
        it = self.win.list.item(self.row)
        title = it.data(TITLE_ROLE) or it.text() or ("Page %d" % (self.row + 1))
        self.lbl_count.setText("  %d / %d  \u2014 %s  " % (self.row + 1, n, title))
        self.win.list.setCurrentRow(self.row)
        pix = QtGui.QPixmap(self._path())
        if pix.isNull():
            self.lbl.clear(); return
        if self.fit:
            area = self.scr.viewport().size()
            pix = pix.scaled(max(50, area.width() - 6), max(50, area.height() - 6),
                             QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)
        else:
            pix = pix.scaled(int(pix.width() * self.zoom), int(pix.height() * self.zoom),
                             QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)
        self.lbl.setPixmap(pix); self.lbl.resize(pix.size())

    def prev(self):
        self.row -= 1; self.fit = True; self._load()

    def next(self):
        self.row += 1; self.fit = True; self._load()

    def zoom_in(self):
        self.fit = False; self.zoom = min(6.0, self.zoom * 1.25); self._load()

    def zoom_out(self):
        self.fit = False; self.zoom = max(0.1, self.zoom / 1.25); self._load()

    def actual_size(self):
        self.fit = False; self.zoom = 1.0; self._load()

    def fit_view(self):
        self.fit = True; self._load()

    def rotate(self, ang):
        p = self._path()
        try:
            with Image.open(p) as im:
                im.rotate(-ang, expand=True).save(p)
            self.win._refresh_item(self.win.list.item(self.row))
            self.win._dirty = True
        except Exception:
            pass
        self.fit = True; self._load()

    def rename(self):
        self.win.list.setCurrentRow(self.row)
        self.win.rename_current_page()
        self._load()

    def delete(self):
        it = self.win.list.item(self.row)
        if it is None:
            return
        self.win.list.clearSelection(); it.setSelected(True)
        self.win.list.setCurrentRow(self.row)
        self.win.delete_page()
        self._load()

    def save_one(self):
        p = self._path()
        if p:
            self.win.save_pdf([p])

    def keyPressEvent(self, e):
        k = e.key()
        if k == QtCore.Qt.Key_Left:
            self.prev()
        elif k == QtCore.Qt.Key_Right:
            self.next()
        elif k in (QtCore.Qt.Key_Plus, QtCore.Qt.Key_Equal):
            self.zoom_in()
        elif k == QtCore.Qt.Key_Minus:
            self.zoom_out()
        elif k == QtCore.Qt.Key_F2:
            self.rename()
        else:
            super().keyPressEvent(e)

class ScannerWindow(QtWidgets.QMainWindow):
    THUMB_W = 150
    THUMB_H = 200

    def __init__(self, auto_scan_profile=None):
        super().__init__()
        self.setWindowTitle(APP_NAME)
        try:
            self.setWindowIcon(app_icon())
        except Exception:
            pass
        self.resize(1220, 820)
        self._tmpdir = tempfile.mkdtemp(prefix="nobledoc_")
        self._config = load_config()
        self._profiles = self._config.get("profiles", [])
        self._opts = dict(DEFAULT_OPTIONS)
        self._opts.update(self._config.get("options", {}))
        self._recent = self._config.get("recent", [])
        self._used_claims = set(self._config.get("used_claims", []))
        self._barcode_tried = False
        self._dirty = False
        self._undo_stack = []
        self._lang = self._opts.get("language", "hi")

        self._build_menu()
        self._build_ui()
        self._apply_style()
        self._refresh_profile_combo()
        self._load_profile_to_panel(self._selected_profile())
        self._refresh_method_label()
        self._refresh_recent_menu()
        self._update_status()

        QtCore.QTimer.singleShot(400, self.check_connection)
        self._conn_timer = QtCore.QTimer(self)
        self._conn_timer.setInterval(30000)
        self._conn_timer.timeout.connect(self.check_connection)
        self._conn_timer.start()

        if not self._config.get("setup_done"):
            QtCore.QTimer.singleShot(600, self._run_wizard)

        self._apply_simple_mode()

        if auto_scan_profile:
            i = self.cmb_profile.findText(auto_scan_profile)
            if i >= 0:
                self.cmb_profile.setCurrentIndex(i)
            QtCore.QTimer.singleShot(800, self.do_scan)

    # ---- config helpers ----
    def _save_opts(self):
        self._config["options"] = self._opts; save_config(self._config)

    def _save_profiles(self):
        self._config["profiles"] = self._profiles; save_config(self._config)

    def _add_recent(self, path):
        self._recent = [path] + [r for r in self._recent if r != path]
        self._recent = self._recent[:12]
        self._config["recent"] = self._recent; save_config(self._config)
        self._refresh_recent_menu()

    # ---- menu + shortcuts ----
    def _ma(self, menu, text, slot, tip="", sc=None):
        # add a menu action with a "?"-prefixed label + a Hindi/English hover tip
        a = menu.addAction(("\u2753 " + text) if tip else text, slot)
        if tip:
            a.setToolTip(tip); a.setStatusTip(tip)
        return a

    def _build_menu(self):
        mb = self.menuBar()
        mf = mb.addMenu(tr("menu_file", self._lang)); mf.setToolTipsVisible(True)
        self._ma(mf, tr("save_all", self._lang), self.save_pdf_all,
                 "हिन्दी: Sabhi pages ki ek PDF banao.\nEnglish: Save all pages as one PDF.", "Ctrl+S")
        self._ma(mf, tr("save_sel", self._lang), self.save_pdf_selected,
                 "हिन्दी: Sirf chune hue (Ctrl/Shift se) pages ki PDF.\nEnglish: PDF of only the selected pages.")
        self._ma(mf, "Save PDF (password)…", self.save_pdf_password,
                 "हिन्दी: Password se surakshit PDF (khole ke liye password lagega).\nEnglish: Password-protected PDF.")
        self._ma(mf, "Save Images…", self.save_images,
                 "हिन्दी: Pages ko JPG/PNG image me save karo (PDF ke bajay).\nEnglish: Save pages as JPG/PNG images.", "Ctrl+Shift+S")
        self._ma(mf, "Export OCR text…", self.export_ocr_text,
                 "हिन्दी: Page ke text ko padh kar .txt me nikaalo (OCR).\nEnglish: Extract page text to .txt via OCR.")
        pmenu = mf.addMenu("Print")
        pmenu.setToolTipsVisible(True)
        pmenu.addAction("All print (sabhi pages)", self.print_all)
        pmenu.addAction("Selected print (chune hue)", self.print_selected)
        pmenu.addAction("ID print (2 ID ek page par)", self.print_ids)
        pmenu.addAction("ID print - sirf selected", self.print_ids_selected)
        self.recent_menu = mf.addMenu("Recent PDFs")
        mf.addSeparator(); mf.addAction("Exit", self.close)

        me = mb.addMenu(tr("menu_edit", self._lang)); me.setToolTipsVisible(True)
        self._ma(me, "Rotate left", self.rotate_left, "हिन्दी: Selected page ko baayein ghumao.\nEnglish: Rotate the selected page left.")
        self._ma(me, "Rotate right", self.rotate_right, "हिन्दी: Selected page ko daayein ghumao.\nEnglish: Rotate the selected page right.")
        self._ma(me, "Brightness +", lambda: self._enhance_current(1.12, 1.0), "हिन्दी: Page ko halka (bright) karo.\nEnglish: Make the page brighter.")
        self._ma(me, "Brightness -", lambda: self._enhance_current(0.9, 1.0), "हिन्दी: Page ko gehra (dim) karo.\nEnglish: Make the page darker.")
        self._ma(me, "Contrast +", lambda: self._enhance_current(1.0, 1.15), "हिन्दी: Text aur saaf/gehra dikhe.\nEnglish: Increase contrast (sharper text).")
        self._ma(me, "Contrast -", lambda: self._enhance_current(1.0, 0.88), "हिन्दी: Contrast kam karo.\nEnglish: Decrease contrast.")
        self._ma(me, "Auto-crop page", self.autocrop_current, "हिन्दी: Page ke aas-paas ki khaali border kaato.\nEnglish: Trim the empty border around the page.")
        self._ma(me, "Undo delete", self.undo_delete, "हिन्दी: Galti se delete hua page wapas laao.\nEnglish: Restore a deleted page.", "Ctrl+Z")
        me.addSeparator()
        self._ma(me, "Delete page", self.delete_page, "हिन्दी: Selected page hatao.\nEnglish: Delete the selected page.", "Delete")
        self._ma(me, "Clear all", self.clear_all, "हिन्दी: Saare pages hatao (khaali karo).\nEnglish: Remove all pages.")

        mt = mb.addMenu(tr("menu_tools", self._lang)); mt.setToolTipsVisible(True)
        self._ma(mt, "Search past PDFs…", self.search_pdfs, "हिन्दी: Purani save ki hui PDF dhoondo (claim/naam se).\nEnglish: Search your saved PDFs.", "Ctrl+F")
        self._ma(mt, "Merge PDFs…", self.merge_pdfs, "हिन्दी: Kai PDF ko jodkar ek PDF banao.\nEnglish: Merge several PDFs into one.")
        self._ma(mt, "Split into multiple PDFs…", self.split_pdfs, "हिन्दी: Ek scan ko kai alag PDF me baanto.\nEnglish: Split into multiple PDFs.")
        self._ma(mt, "Monthly report…", self.monthly_report, "हिन्दी: Mahine ka scan/claim report banao.\nEnglish: Generate a monthly report.")
        self._ma(mt, "Create desktop shortcut…", self.create_shortcut, "हिन्दी: Desktop par ek-click scan ka shortcut banao.\nEnglish: Make a one-click desktop scan shortcut.")
        self._ma(mt, "Auto-name pages (document ka naam)", self.auto_name_pages, "हिन्दी: Har page ko padh kar uska naam (jaise DISCHARGE SUMMARY, RECEIPT) thumbnail ke neeche likhe. 'Page 1,2' ke bajay asli naam.\nEnglish: Read each page and label it with its document title instead of 'Page 1,2'.")

        ms = mb.addMenu(tr("menu_settings", self._lang)); ms.setToolTipsVisible(True)
        self._ma(ms, tr("options", self._lang), self.open_options, "हिन्दी: App ki saari settings (auto-save, blank hatao, backup, waghera).\nEnglish: All app settings.")
        self._ma(ms, tr("profiles", self._lang), self.open_profiles, "हिन्दी: Scan profiles banao/badlo (device, dpi, colour, duplex).\nEnglish: Create/edit scan profiles.")
        self._ma(ms, tr("scan_method", self._lang) + "…", self.choose_scan_method, "हिन्दी: Scan ka tareeka: escl (network duplex), twain (USB duplex), ya wia.\nEnglish: Scan method: escl (network duplex), twain (USB), or wia.")
        self._ma(ms, tr("language", self._lang) + "…", self.choose_language, "हिन्दी: App ki bhasha badlo (Hindi/English).\nEnglish: Change the app language.")
        self._ma(ms, "Stats server URL…", self.set_stats_url, "हिन्दी: Worldwide stats ke liye Google Apps Script ka URL daalein (kitne scan hue, kitne online).\nEnglish: Set the stats server URL (worldwide scan counts + online users).")
        self._ma(ms, "Scanner IP…", self.set_scanner_ip, "हिन्दी: Network scanner ka IP set karo (jaise 192.168.1.8).\nEnglish: Set the network scanner IP.")
        self._ma(ms, "Keyboard Shortcuts…", self.show_shortcuts, "हिन्दी: Keyboard ke shortcuts ki list dekho.\nEnglish: View keyboard shortcuts.")
        self.act_simple = self._ma(ms, tr("simple_on", self._lang), self.toggle_simple_mode, "हिन्दी: Simple mode: sirf zaroori buttons dikhein (naye users ke liye aasan).\nEnglish: Simple mode: show only the essential buttons.")
        self.act_simple.setCheckable(True)
        self.act_simple.setChecked(bool(self._opts.get("simple_mode")))

        mh = mb.addMenu(tr("menu_help", self._lang)); mh.setToolTipsVisible(True)
        self._ma(mh, tr("help_guide", self._lang), self.show_help, "हिन्दी: App istemal karne ki guide.\nEnglish: How-to guide.")
        self._ma(mh, "Setup wizard", self._run_wizard, "हिन्दी: Pehli baar wala setup dobara chalao.\nEnglish: Re-run the first-time setup.")
        self._ma(mh, tr("whatsnew", self._lang), self.show_whatsnew, "हिन्दी: Naye badlav/features.\nEnglish: What's new.")
        self._ma(mh, "Test / Diagnostics", self.run_diagnostics, "हिन्दी: Scanner/app ki jaankari + error report (share karne ke liye).\nEnglish: Scanner/app info + error report.")
        self._ma(mh, "Duplex Test (both-side)", self.run_duplex_test, "हिन्दी: Jaancho ki dono taraf (duplex) scan ho raha hai ya nahi.\nEnglish: Test whether both-side (duplex) scanning works.")
        self._ma(mh, "eSCL Test (network scan jaanch)", self.run_escl_test, "हिन्दी: Network scan (eSCL) ko step-by-step jaanch kar asli problem batata hai (connect / status / job).\nEnglish: Step-by-step eSCL network-scan test that shows the real problem.")
        self._ma(mh, tr("feedback", self._lang), self.send_feedback, "हिन्दी: Sujhav/shikayat bhejo.\nEnglish: Send feedback.")
        self._ma(mh, tr("about", self._lang), self.show_about, "हिन्दी: App ke baare me.\nEnglish: About this app.")

        self._build_shortcuts()

    def _refresh_recent_menu(self):
        self.recent_menu.clear()
        if not self._recent:
            a = self.recent_menu.addAction("(koi nahi)"); a.setEnabled(False); return
        for path in self._recent:
            self.recent_menu.addAction(os.path.basename(path),
                                       lambda checked=False, p=path: self._open_path(p))

    def _open_path(self, path):
        try:
            os.startfile(path)
        except Exception as exc:
            self._warn("File nahi khuli:\n%s" % exc)

    # ---- profiles / options ----
    def _refresh_profile_combo(self):
        self.cmb_profile.blockSignals(True); self.cmb_profile.clear()
        for p in self._profiles:
            self.cmb_profile.addItem(p.get("name", "Profile"))
        sel = self._config.get("selected_profile")
        if sel:
            i = self.cmb_profile.findText(sel)
            if i >= 0:
                self.cmb_profile.setCurrentIndex(i)
        self.cmb_profile.blockSignals(False)

    def _selected_profile(self):
        name = self.cmb_profile.currentText()
        for p in self._profiles:
            if p.get("name") == name:
                return p
        return None

    def _on_fast_toggled(self, on):
        self._opts["fast_mode"] = bool(on); self._save_opts()
        if on:
            # reflect fast settings in the panel so the user sees them
            self.cmb_dpi.setCurrentText("200 dpi")
            self.cmb_depth.setCurrentText("Black & White")
        self.status.showMessage("Fast mode ON (200 dpi, B&W)" if on else "Fast mode OFF", 3000)

    def _on_profile_changed(self, name):
        self._config["selected_profile"] = name; save_config(self._config)
        self._load_profile_to_panel(self._selected_profile())


    def open_profiles(self):
        dlg = ProfileManagerDialog(self, self._profiles)
        if dlg.exec_() == QtWidgets.QDialog.Accepted:
            self._profiles = dlg.profiles; self._save_profiles(); self._refresh_profile_combo()

    def open_options(self):
        dlg = OptionsDialog(self, self._opts)
        if dlg.exec_() == QtWidgets.QDialog.Accepted:
            self._opts = dlg.get_opts(); self._save_opts(); self._update_status()
            self._apply_style(); self._refresh_page_numbers()

    def _refresh_page_numbers(self):
        show = self._opts.get("show_page_numbers", True)
        for i in range(self.list.count()):
            it = self.list.item(i)
            title = it.data(TITLE_ROLE)
            if title:
                it.setText(title)
            else:
                it.setText(("Page %d" % (i + 1)) if show else "")

    def _vsep(self):
        f = QtWidgets.QFrame(); f.setObjectName("hr"); f.setFrameShape(QtWidgets.QFrame.VLine); return f

    def _panel_scan_params(self, prof):
        try:
            dpi = int(self.cmb_dpi.currentText().split()[0])
        except Exception:
            dpi = int(prof.get("dpi", 200))
        d = self.cmb_depth.currentText()
        color = "color" if d.startswith("24") else ("gray" if d.startswith("Gray") else "bw")
        duplex = "Both" in self.cmb_sides.currentText()
        return dpi, color, duplex

    def _refresh_method_label(self):
        if not hasattr(self, "method_lbl"):
            return
        m = self._opts.get("scanner_method", "twain")
        color = "#0f766e" if m in ("wia", "naps2", "escl") else "#b45309"
        self.method_lbl.setText('<b>Connected via:</b> <span style="color:%s">%s</span>'
                                % (color, m.upper()))

    def _load_profile_to_panel(self, prof):
        if prof is None or not hasattr(self, "cmb_dpi"):
            return
        self.dev_lbl.setText(prof.get("source_name") or "(koi device nahi)")
        self.cmb_dpi.setCurrentText(str(prof.get("dpi", 200)) + " dpi")
        c = prof.get("color", "gray")
        self.cmb_depth.setCurrentText("24-bit Colour" if c == "color"
                                      else ("Grayscale" if c == "gray" else "Black & White"))
        self.cmb_source.setCurrentText("Feeder (ADF)")
        self.cmb_sides.setCurrentText("Both side (dono taraf)" if prof.get("duplex") else "Single side (ek taraf)")
        if hasattr(self, "cmb_pagesize"):
            _ps = (prof.get("page_size") or "auto").lower()
            for k in range(self.cmb_pagesize.count()):
                if self.cmb_pagesize.itemText(k).lower().startswith(_ps[:4]):
                    self.cmb_pagesize.setCurrentIndex(k); break

    def _quick_new_profile(self):
        dlg = EditProfileDialog(self)
        if dlg.exec_() == QtWidgets.QDialog.Accepted:
            self._profiles.append(dlg.get_profile()); self._save_profiles(); self._refresh_profile_combo()
            self.cmb_profile.setCurrentIndex(self.cmb_profile.count() - 1)

    def _quick_edit_profile(self):
        name = self.cmb_profile.currentText()
        idx = next((k for k, p in enumerate(self._profiles) if p.get("name") == name), -1)
        if idx < 0:
            self._quick_new_profile(); return
        dlg = EditProfileDialog(self, self._profiles[idx])
        if dlg.exec_() == QtWidgets.QDialog.Accepted:
            self._profiles[idx] = dlg.get_profile(); self._save_profiles(); self._refresh_profile_combo()
            i = self.cmb_profile.findText(self._profiles[idx].get("name"))
            if i >= 0:
                self.cmb_profile.setCurrentIndex(i)
            self._load_profile_to_panel(self._selected_profile())

    def _open_preview_dialog(self, item):
        if item is None:
            return
        row = self.list.row(item)
        PreviewDialog(self, row).exec_()

    def auto_name_pages(self):
        if not tesseract_available():
            self._warn(
                "Document ka naam padhne ke liye 'Tesseract OCR' program install karna zaroori hai.\n\n"
                "Download (free): https://github.com/UB-Mannheim/tesseract/wiki\n"
                "Install karke ApneScan dobara kholein. Phir Tools \u2192 Auto-name pages chalega."); return
        items = [(i, self.list.item(i).data(QtCore.Qt.UserRole)) for i in range(self.list.count())]
        if not items:
            self._warn("Pehle koi page scan/import karein."); return
        self._named_count = 0
        self._named_total = len(items)
        self.status.showMessage("Document ke naam padhe ja rahe hain... (thoda ruko)", 0)
        self._namer = NameWorker(items, learned=self._learned_names())
        self._namer.named.connect(self._apply_page_title)
        self._namer.finished_all.connect(self._on_naming_done)
        self._namer.start()

    def _learned_names(self):
        out = []
        for e in (self._config.get("learned_names", []) or []):
            try:
                out.append((set(e.get("words", [])), e.get("name", "")))
            except Exception:
                pass
        return out

    def _apply_page_title(self, row, title):
        if 0 <= row < self.list.count():
            it = self.list.item(row)
            key = name_key(title)
            # if this same document was saved before, reuse THAT saved name
            remembered = (self._config.get("doc_names", {}) or {}).get(key)
            label = remembered or underscore_name(title)
            it.setData(TITLE_ROLE, label)
            it.setData(NAMEKEY_ROLE, key)
            it.setText(label)
            self._named_count = getattr(self, "_named_count", 0) + 1

    def _on_naming_done(self):
        got = getattr(self, "_named_count", 0)
        total = getattr(self, "_named_total", 0)
        if got == 0:
            self.status.clearMessage()
            self._warn(
                "OCR chala, par kisi bhi page ka naam nahi mil paaya.\n\n"
                "Iske 2 kaaran ho sakte hain:\n"
                "1) Tesseract theek se nahi mila \u2014 Help \u2192 Test/Diagnostics kholein aur "
                "'Tesseract version' line dekhein (WORKS likha hai ya NOT FOUND).\n"
                "2) Page ke UPAR saaf chhapa hua heading nahi hai (jaise haath se likhe page) \u2014 "
                "aise pages ka naam nahi aata, wo normal hai.")
        else:
            self.status.showMessage("%d / %d pages ke naam mil gaye." % (got, total), 6000)

    def _build_shortcuts(self):
        methods = {
            "scan": self.do_scan,
            "import": self.import_images,
            "rename": self.rename_current_page,
            "save_all": self.save_pdf_all,
            "save_sel": self.save_pdf_selected,
            "save_pw": self.save_pdf_password,
            "save_img": self.save_images,
            "ocr_text": self.export_ocr_text,
            "print": self.print_pages,
            "rotate_left": self.rotate_left,
            "rotate_right": self.rotate_right,
            "bright_up": lambda: self._enhance_current(1.12, 1.0),
            "bright_dn": lambda: self._enhance_current(0.9, 1.0),
            "contrast_up": lambda: self._enhance_current(1.0, 1.15),
            "contrast_dn": lambda: self._enhance_current(1.0, 0.88),
            "autocrop": self.autocrop_current,
            "autoname": self.auto_name_pages,
            "undo": self.undo_delete,
            "delete": self.delete_page,
            "clear": self.clear_all,
            "move_up": self.move_up,
            "move_down": self.move_down,
            "search": self.search_pdfs,
            "merge": self.merge_pdfs,
            "split": self.split_pdfs,
            "report": self.monthly_report,
            "zoom_in": lambda: self._zoom_thumbs(1.15),
            "zoom_out": lambda: self._zoom_thumbs(0.87),
            "options": self.open_options,
            "profiles": self.open_profiles,
        }
        self._sc = {}
        custom = self._opts.get("shortcuts", {}) or {}
        for sid, label, default in SHORTCUTS:
            fn = methods.get(sid)
            if fn is None:
                continue
            scut = QtWidgets.QShortcut(self)
            scut.activated.connect(fn)
            key = custom.get(sid, default)
            if key:
                scut.setKey(QtGui.QKeySequence(key))
            self._sc[sid] = scut

    def _apply_shortcut(self, sid, key):
        scut = self._sc.get(sid)
        if scut is not None:
            scut.setKey(QtGui.QKeySequence(key) if key else QtGui.QKeySequence())

    def _cur_shortcut(self, sid, default):
        return (self._opts.get("shortcuts", {}) or {}).get(sid, default)

    def show_shortcuts(self):
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Keyboard Shortcuts"); dlg.resize(560, 560)
        lay = QtWidgets.QVBoxLayout(dlg)
        lay.addWidget(QtWidgets.QLabel(
            "Shortcut badalne ke liye: row chuno \u2192 'Naya' box me key dabao \u2192 Assign.\n"
            "To change: pick a row, press the new key in 'New', then Assign."))
        tbl = QtWidgets.QTableWidget(len(SHORTCUTS), 2)
        tbl.setHorizontalHeaderLabels(["Action", "Shortcut"])
        tbl.verticalHeader().setVisible(False)
        tbl.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        tbl.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        tbl.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        tbl.horizontalHeader().setStretchLastSection(True); tbl.setColumnWidth(0, 300)

        def refill():
            for r, (sid, label, default) in enumerate(SHORTCUTS):
                tbl.setItem(r, 0, QtWidgets.QTableWidgetItem(label))
                tbl.setItem(r, 1, QtWidgets.QTableWidgetItem(self._cur_shortcut(sid, default) or "—"))
        refill()
        lay.addWidget(tbl, 1)

        row = QtWidgets.QHBoxLayout()
        kseq = QtWidgets.QKeySequenceEdit()
        row.addWidget(QtWidgets.QLabel("Naya / New:")); row.addWidget(kseq, 1)
        b_assign = QtWidgets.QPushButton("Assign"); b_unassign = QtWidgets.QPushButton("Hatao / Unassign")
        row.addWidget(b_assign); row.addWidget(b_unassign)
        lay.addLayout(row)

        def sel():
            r = tbl.currentRow()
            return (SHORTCUTS[r][0], r) if r >= 0 else (None, -1)

        def do_assign():
            sid, r = sel()
            if sid is None:
                return
            key = kseq.keySequence().toString()
            if not key:
                return
            sm = self._opts.setdefault("shortcuts", {})
            for osid, olabel, odef in SHORTCUTS:      # clear conflicts
                if osid != sid and self._cur_shortcut(osid, odef) == key:
                    sm[osid] = ""; self._apply_shortcut(osid, "")
            sm[sid] = key; self._apply_shortcut(sid, key); self._save_opts()
            kseq.clear(); refill()

        def do_unassign():
            sid, r = sel()
            if sid is None:
                return
            self._opts.setdefault("shortcuts", {})[sid] = ""
            self._apply_shortcut(sid, ""); self._save_opts(); refill()

        b_assign.clicked.connect(do_assign); b_unassign.clicked.connect(do_unassign)

        bottom = QtWidgets.QHBoxLayout()
        b_reset = QtWidgets.QPushButton("Restore Defaults")

        def do_reset():
            self._opts["shortcuts"] = {}
            for sid, label, default in SHORTCUTS:
                self._apply_shortcut(sid, default)
            self._save_opts(); refill()
        b_reset.clicked.connect(do_reset)
        b_ok = QtWidgets.QPushButton("OK"); b_ok.clicked.connect(dlg.accept)
        bottom.addWidget(b_reset); bottom.addStretch(1); bottom.addWidget(b_ok)
        lay.addLayout(bottom)
        dlg.exec_()

    def _stats_url(self):
        if "stats_url" in self._config:
            return self._config["stats_url"]      # user override (may be "" to disable)
        return DEFAULT_STATS_URL

    def _get_client_id(self):
        cid = self._config.get("client_id")
        if not cid:
            import uuid
            cid = uuid.uuid4().hex[:16]
            self._config["client_id"] = cid
            try:
                save_config(self._config)
            except Exception:
                pass
        return cid

    def _set_stats_display(self, stats):
        if not self._stats_url():
            self.stats_box.setText(
                '<b>\U0001f30d ApneScan</b><br>'
                '<span style="color:#94a3b8;">Worldwide stats band hai<br>'
                '(Settings \u2192 Stats server URL)</span>')
            return
        if stats is None:
            self.stats_box.setText('<b>\U0001f30d Worldwide</b><br>'
                                   '<span style="color:#94a3b8;">Loading\u2026</span>')
            return
        total, today, online = stats
        self.stats_box.setText(
            '<b>\U0001f30d ApneScan Worldwide</b><br>'
            '\U0001f4c4 Total scans: <b>%s</b><br>'
            '\U0001f4c5 Aaj (today): <b>%s</b><br>'
            '\U0001f7e2 Abhi online: <b>%s</b>'
            % ("{:,}".format(total), "{:,}".format(today), online))

    def _stats_failed(self):
        if self._stats_url():
            self.stats_box.setText('<b>\U0001f30d Worldwide</b><br>'
                                   '<span style="color:#94a3b8;">stats abhi nahi mile</span>')

    def _refresh_stats(self, action="ping"):
        url = self._stats_url()
        if not url:
            self._set_stats_display(None)
            return
        self._stats_worker = StatsWorker(url, self._get_client_id(), action=action)
        self._stats_worker.got.connect(lambda t, d, o: self._set_stats_display((t, d, o)))
        self._stats_worker.failed.connect(self._stats_failed)
        self._stats_worker.start()

    def _report_scan_stat(self, n):
        url = self._stats_url()
        if not url or n <= 0:
            return
        self._scan_reporter = StatsWorker(url, self._get_client_id(), action="scan", n=n)
        self._scan_reporter.got.connect(lambda t, d, o: self._set_stats_display((t, d, o)))
        self._scan_reporter.start()

    def set_stats_url(self):
        cur = self._stats_url()
        url, ok = QtWidgets.QInputDialog.getText(
            self, "Stats server URL",
            "Google Apps Script Web app URL (…/exec):\n(khaali karke stats band ho jayenge)",
            text=cur)
        if not ok:
            return
        self._config["stats_url"] = url.strip()
        try:
            save_config(self._config)
        except Exception:
            pass
        self._set_stats_display(None)
        self._refresh_stats()

    def set_scanner_ip(self):
        ip, ok = QtWidgets.QInputDialog.getText(self, "Scanner IP", "Scanner IP:", text=self.ip_field.text())
        if ok:
            self.ip_field.setText(ip.strip()); self.check_connection()

    def undo_delete(self):
        if not self._undo_stack:
            self.status.showMessage("Undo ke liye kuch nahi", 3000); return
        path, row = self._undo_stack.pop()
        if not os.path.exists(path):
            return
        icon = QtGui.QIcon(self._make_thumb(path))
        item = QtWidgets.QListWidgetItem(icon, "Page")
        item.setData(QtCore.Qt.UserRole, path); item.setTextAlignment(QtCore.Qt.AlignHCenter)
        self.list.insertItem(min(row, self.list.count()), item)
        self.list.setCurrentItem(item); self.list.clearSelection()
        self._renumber_pages(); self._dirty = True
        self._update_status(); self._update_empty_state()

    def _renumber_pages(self):
        show = self._opts.get("show_page_numbers", True)
        for i in range(self.list.count()):
            it = self.list.item(i)
            title = it.data(TITLE_ROLE)
            if title:
                it.setText(title)
            else:
                it.setText(("Page %d" % (i + 1)) if show else "")

    def _update_empty_state(self):
        empty = self.list.count() == 0
        self._empty_lbl.setVisible(empty)
        if empty:
            self._empty_lbl.setGeometry(self.list.viewport().rect())

    def _apply_thumb_zoom(self, w):
        w = max(80, min(340, int(w)))
        h = int(w * 4 / 3)
        self._thumb_w, self._thumb_h = w, h
        self.list.setIconSize(QtCore.QSize(w, h))
        self.list.setGridSize(QtCore.QSize(w + 24, h + 34))

    def _zoom_thumbs(self, factor):
        self._apply_thumb_zoom(self._thumb_w * factor)

    def eventFilter(self, obj, ev):
        if obj is self.list.viewport():
            if ev.type() == QtCore.QEvent.Resize:
                self._empty_lbl.setGeometry(self.list.viewport().rect())
            elif ev.type() == QtCore.QEvent.Wheel and (ev.modifiers() & QtCore.Qt.ControlModifier):
                # Ctrl + mouse scroll -> zoom thumbnails in/out (like NAPS2)
                self._apply_thumb_zoom(self._thumb_w * (1.15 if ev.angleDelta().y() > 0 else 0.87))
                return True
        return super().eventFilter(obj, ev)

    def dragEnterEvent(self, e):
        if e.mimeData().hasUrls():
            e.acceptProposedAction()

    def dropEvent(self, e):
        exts = (".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".pdf")
        count = 0
        for url in e.mimeData().urls():
            f = url.toLocalFile()
            if f.lower().endswith(exts):
                count += self._import_one(f)
        if count:
            self.status.showMessage("%d page add hue" % count, 4000)
            if tesseract_available():
                self.auto_name_pages()

    def _after_save_action(self, out):
        act = self._opts.get("after_save", "nothing")
        try:
            if act == "open":
                os.startfile(out)
            elif act == "folder":
                os.startfile(os.path.dirname(out))
        except Exception:
            pass

    def _run_wizard(self):
        wiz = SetupWizard(self, self._opts.get("language", "hi"))
        if wiz.exec_() == QtWidgets.QDialog.Accepted:
            self._opts["scanner_method"] = wiz.result_method
            if wiz.result_method == "wia":
                self._opts["wia_device_id"] = wiz.result_wia_id
            self._save_opts()
            prof = self._selected_profile()
            if prof is None:
                prof = {"name": "Default", "source_name": wiz.result_device_name,
                        "dpi": 200, "color": "gray", "duplex": False}
                self._profiles.append(prof)
            elif wiz.result_method == "twain" and wiz.result_device_name:
                prof["source_name"] = wiz.result_device_name
            self._save_profiles(); self._refresh_profile_combo()
            i = self.cmb_profile.findText(prof.get("name"))
            if i >= 0:
                self.cmb_profile.setCurrentIndex(i)
        self._config["setup_done"] = True
        save_config(self._config)
        self._update_status()

    def choose_scan_method(self):
        cur = self._opts.get("scanner_method", "twain")
        labels = ["escl (network duplex - no extra software)", "twain (USB duplex)", "wia", "naps2 (needs NAPS2)"]
        keys = ["escl", "twain", "wia", "naps2"]
        idx = keys.index(cur) if cur in keys else 0
        choice, ok = QtWidgets.QInputDialog.getItem(
            self, "Scan method", "Method:", labels, idx, False)
        if not ok:
            return
        method = keys[labels.index(choice)]
        self._opts["scanner_method"] = method
        if method == "escl":
            ip = self.ip_field.text().strip()
            if not ip:
                ip, oki = QtWidgets.QInputDialog.getText(
                    self, "Scanner IP", "Scanner ka network IP daalo (jaise 192.168.1.8):")
                if oki:
                    ip = ip.strip(); self.ip_field.setText(ip)
            self._opts["scanner_ip"] = ip
        elif method == "naps2":
            path = self._opts.get("naps2_path") or find_naps2()
            if not path or not os.path.exists(path):
                path, okp = QtWidgets.QInputDialog.getText(
                    self, "NAPS2",
                    "NAPS2.Console.exe ka poora path daalein\n"
                    "(jaise C:\\Program Files\\NAPS2\\NAPS2.Console.exe):")
                path = path.strip() if okp else ""
            self._opts["naps2_path"] = path
            pname, okn = QtWidgets.QInputDialog.getText(
                self, "NAPS2 profile",
                "NAPS2 ka profile naam (bilkul waisa hi jaisa NAPS2 me hai,\n"
                "jaise: 150dpi Double Side):",
                text=self._opts.get("naps2_profile", ""))
            if okn:
                self._opts["naps2_profile"] = pname.strip()
            if not path:
                self._warn("NAPS2 ka path nahi mila. Pehle NAPS2 install karo.")
        elif method == "wia":
            try:
                devs = list_wia_sources()
            except Exception as exc:
                self._warn(friendly_error(exc, self._opts.get("language", "hi"))); devs = []
            if devs:
                names = [n for _i, n in devs]
                name, ok2 = QtWidgets.QInputDialog.getItem(self, "WIA scanner", "Scanner:", names, 0, False)
                if ok2 and name:
                    for _id, n in devs:
                        if n == name:
                            self._opts["wia_device_id"] = _id; break
        else:
            # TWAIN: let the user pick the real driver source (e.g. "HP TWAIN USB"),
            # which supports duplex — unlike the WIA-bridge sources.
            try:
                names = list_sources(int(self.winId()))
            except Exception as exc:
                self._warn(friendly_error(exc, self._opts.get("language", "hi"))); names = []
            if names:
                name, ok2 = QtWidgets.QInputDialog.getItem(
                    self, "TWAIN scanner",
                    "Scanner chuno (duplex ke liye 'HP TWAIN USB' jaisa asli driver chuno,\nWIA-... wala nahi):",
                    names, 0, False)
                if ok2 and name:
                    prof = self._selected_profile()
                    if prof is not None:
                        prof["source_name"] = name
                        self._save_profiles()
                        self._load_profile_to_panel(prof)
        self._save_opts(); self._update_status(); self._refresh_method_label()

    def choose_language(self):
        cur = self._opts.get("language", "hi")
        lang, ok = QtWidgets.QInputDialog.getItem(
            self, "Language", "Bhasha / Language:", ["Hindi", "English"],
            0 if cur == "hi" else 1, False)
        if not ok:
            return
        self._opts["language"] = "hi" if lang == "Hindi" else "en"
        self._save_opts()
        QtWidgets.QMessageBox.information(
            self, "Language",
            "Language badal gayi. App band karke dobara kholein taaki poora asar dikhe."
            if self._opts["language"] == "hi"
            else "Language changed. Close and reopen the app to fully apply.")

    def show_help(self):
        HelpDialog(self, self._opts.get("language", "hi")).exec_()

    def run_duplex_test(self):
        """Scan with Both-side ON, then report whether the scanner really
        returned 2 sides — with the duplex value used and any error."""
        if self.list.count() > 0:
            if QtWidgets.QMessageBox.question(
                    self, "Duplex Test",
                    "Test se pehle maujuda pages hat jayenge. Aage badhein?"
                    if self._lang == "hi" else "This clears current pages. Continue?",
                    QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                    QtWidgets.QMessageBox.No) != QtWidgets.QMessageBox.Yes:
                return
            for p in self._ordered_paths():
                try:
                    os.remove(p)
                except Exception:
                    pass
            self.list.clear(); self._update_empty_state()

        QtWidgets.QMessageBox.information(
            self, "Duplex Test",
            "Feeder me 1 kaagaz (dono taraf likha hua) rakho, phir OK dabao.\n"
            "Test us kaagaz ko scan karke batayega ki dono taraf aaya ya nahi."
            if self._lang == "hi" else
            "Put ONE double-sided sheet in the feeder, then press OK.")

        prof = self._selected_profile()
        if prof is None:
            self._warn("Pehle profile/scanner chuno."); return
        method = self._opts.get("scanner_method", "twain")
        dpi = 200
        pixel = "gray"
        got = {"pages": 0, "err": None, "trace": []}
        WIA_FMT_BMP = "{B96B3CAB-0728-11D3-9D7B-0000F81EF32E}"

        QtWidgets.QApplication.setOverrideCursor(QtCore.Qt.WaitCursor)
        try:
            if method == "wia":
                try:
                    _pythoncom.CoInitialize()
                except Exception:
                    pass
                try:
                    tr = got["trace"]
                    wid = self._opts.get("wia_device_id")

                    def _connect():
                        dm = _w32.Dispatch("WIA.DeviceManager")
                        for i in range(1, dm.DeviceInfos.Count + 1):
                            info = dm.DeviceInfos.Item(i)
                            if wid is None or info.DeviceID == wid:
                                return info.Connect()
                        return None

                    def _setp(obj, pid, val):
                        try:
                            for p in obj.Properties:
                                if int(p.PropertyID) == pid:
                                    p.Value = val
                                    return True
                        except Exception:
                            pass
                        return False

                    def _valid_dpi(it, want):
                        try:
                            for p in it.Properties:
                                if int(p.PropertyID) == 6147:
                                    vals = [int(x) for x in p.SubType.Values]
                                    if vals:
                                        return min(vals, key=lambda v: abs(v - want))
                        except Exception:
                            pass
                        return want

                    def _try_transfers(item, n=3):
                        pages = 0; err = None
                        for k in range(n):
                            try:
                                try:
                                    item.Transfer(WIA_FMT_BMP)
                                except Exception:
                                    item.Transfer()
                                pages += 1
                            except Exception as e:
                                err = str(e); break
                        return pages, err

                    # Each strategy: (label, set-3088?, set-3096?, set item props?, plain-only?)
                    strategies = [
                        ("A duplex=5, no pages, no item-props", 5, False, False),
                        ("B duplex=5 + pages=0", 5, True, False),
                        ("C duplex=5 + item props(res/depth)", 5, False, True),
                        ("D duplex=7", 7, False, False),
                        ("E duplex=0x8005(32773)", 0x8005, False, False),
                    ]
                    best = 0
                    for label, dupval, setpages, setitem in strategies:
                        dev = None
                        try:
                            dev = _connect()
                            if dev is None:
                                tr.append("%s -> connect fail" % label); continue
                            ok88 = _setp(dev, 3088, dupval)
                            if setpages:
                                _setp(dev, 3096, 0)
                            it = dev.Items[1]
                            if setitem:
                                _setp(it, 4104, 2)  # grayscale
                                dv = _valid_dpi(it, 200)
                                _setp(it, 6147, dv); _setp(it, 6148, dv)
                            pages, err = _try_transfers(it, 3)
                            tr.append("%s -> set88=%s pages=%d err=%s"
                                      % (label, ok88, pages, err or "none"))
                            if pages > best:
                                best = pages
                            if pages >= 1:
                                tr.append(">>> Ye tarika chala! (pages=%d). Baaki skip." % pages)
                                break
                        except Exception as e:
                            tr.append("%s -> EXC %s" % (label, e))
                        finally:
                            try:
                                if dev is not None:
                                    dev = None
                            except Exception:
                                pass
                    got["pages"] = best
                finally:
                    try:
                        _pythoncom.CoUninitialize()
                    except Exception:
                        pass
            else:
                def _count(img):
                    got["pages"] += 1
                try:
                    scan_pages(int(self.winId()), prof.get("source_name"), dpi, pixel, True, _count)
                except Exception as exc:
                    got["err"] = str(exc)
        except Exception as exc:
            got["err"] = str(exc)
        finally:
            QtWidgets.QApplication.restoreOverrideCursor()

        lines = []
        lines.append("===== ApneScan Duplex Test =====")
        lines.append("Time: %s" % datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        lines.append("Scan method: %s" % method)
        lines.append("Device: %s" % (prof.get("source_name") or "(none)"))
        lines.append("Test settings: 1 sheet, Both-side ON, 200 dpi")
        lines.append("")
        lines.append("Pages returned: %d" % got["pages"])
        lines.append("Error: %s" % (got["err"] or "(none)"))
        if got["trace"]:
            lines.append("")
            lines.append("--- Detail (step by step) ---")
            for t in got["trace"]:
                lines.append("  " + t)
        lines.append("")
        if got["pages"] >= 2:
            lines.append("RESULT: DUPLEX WORKING \u2705  (scanner ne dono taraf diye)")
        elif got["pages"] == 1:
            lines.append("RESULT: SINGLE SIDE ONLY \u274c  (scanner ne sirf ek taraf diya)")
            lines.append("Matlab: is method/driver par duplex nahi mil raha.")
        else:
            lines.append("RESULT: koi page nahi aaya (scanner/feeder check karo)")
        lines.append("================================")
        report = "\n".join(lines)

        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Duplex Test Result"); dlg.resize(560, 420)
        lay = QtWidgets.QVBoxLayout(dlg)
        lay.addWidget(QtWidgets.QLabel(
            "Ye result copy karke share karein:" if self._lang == "hi" else "Copy & share this result:"))
        box = QtWidgets.QPlainTextEdit(); box.setPlainText(report); box.setReadOnly(True)
        box.setStyleSheet("font-family: Consolas, monospace; font-size: 12px;")
        lay.addWidget(box, 1)
        row = QtWidgets.QHBoxLayout()
        b_copy = QtWidgets.QPushButton("Copy karo" if self._lang == "hi" else "Copy")
        b_copy.clicked.connect(lambda: (QtWidgets.QApplication.clipboard().setText(report),
                                        self.status.showMessage("Copy ho gaya", 3000)))
        b_close = QtWidgets.QPushButton("Band karo" if self._lang == "hi" else "Close")
        b_close.clicked.connect(dlg.accept)
        row.addWidget(b_copy); row.addStretch(1); row.addWidget(b_close)
        lay.addLayout(row)
        dlg.exec_()

    def _show_report(self, title, report):
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle(title); dlg.resize(640, 560)
        lay = QtWidgets.QVBoxLayout(dlg)
        lay.addWidget(QtWidgets.QLabel("Ye report copy karke share karein / Copy and share this report:"))
        box = QtWidgets.QPlainTextEdit(); box.setPlainText(report); box.setReadOnly(True)
        box.setStyleSheet("font-family: Consolas, monospace; font-size: 12px;")
        lay.addWidget(box, 1)
        row = QtWidgets.QHBoxLayout()
        b_copy = QtWidgets.QPushButton("Copy")
        def _copy():
            QtWidgets.QApplication.clipboard().setText(report)
            self.status.showMessage("Report copy ho gaya", 3000)
        b_copy.clicked.connect(_copy)
        b_save = QtWidgets.QPushButton("Save to file")
        def _save():
            p, _ = QtWidgets.QFileDialog.getSaveFileName(
                dlg, "Save report", os.path.join(os.path.expanduser("~"), "ApneScan_escl_test.txt"), "Text (*.txt)")
            if p:
                try:
                    with open(p, "w", encoding="utf-8") as fh:
                        fh.write(report)
                    self.status.showMessage("Saved: %s" % p, 5000)
                except Exception as exc:
                    self._warn("Save fail: %s" % exc)
        b_save.clicked.connect(_save)
        b_close = QtWidgets.QPushButton("Close"); b_close.clicked.connect(dlg.accept)
        row.addWidget(b_copy); row.addWidget(b_save); row.addStretch(1); row.addWidget(b_close)
        lay.addLayout(row)
        dlg.exec_()

    def run_escl_test(self):
        ip = self.ip_field.text().strip() or self._config.get("scanner_ip", "")
        if not ip:
            self._warn("Scanner IP set nahi hai. Settings \u2192 Scanner IP me daalein (jaise 192.168.1.8).")
            return
        try:
            self._state_timer.stop()
        except Exception:
            pass
        self.status.showMessage("eSCL test chal raha hai... (kuch second, scanner se baat kar rahe hain)", 0)
        self._escl_tester = EsclTestWorker(ip)
        self._escl_tester.done.connect(self._on_escl_test_done)
        self._escl_tester.start()

    def _on_escl_test_done(self, report):
        try:
            self._state_timer.start()
        except Exception:
            pass
        self.status.clearMessage()
        self._show_report("eSCL Test (network scan jaanch)", report)

    def run_diagnostics(self):
        """Collect scanner + app info and any errors into one report the user
        can copy and share. Runs a safe non-scanning probe of TWAIN and WIA."""
        import platform
        lines = []
        def add(t=""):
            lines.append(t)

        add("===== ApneScan Diagnostics =====")
        add("Version: %s" % VERSION)
        add("Time: %s" % datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        try:
            add("Windows: %s" % platform.platform())
            add("Python: %s (%d-bit)" % (platform.python_version(),
                                         64 if sys.maxsize > 2**32 else 32))
        except Exception:
            pass
        add("Libs: TWAIN=%s  WIA(pywin32)=%s  OCR=%s  Excel=%s  Barcode=%s"
            % (HAS_TWAIN, HAS_W32, HAS_OCR_LIBS, HAS_XLSX, HAS_ZBAR))
        add("")
        add("Scan method: %s" % self._opts.get("scanner_method", "twain"))
        prof = self._selected_profile()
        add("Profile: %s" % (prof.get("name") if prof else "(none)"))
        add("Device (stored): %s" % (prof.get("source_name") if prof else "(none)"))
        try:
            add("Panel -> Source: %s | Sides: %s | DPI: %s | Depth: %s"
                % (self.cmb_source.currentText(), self.cmb_sides.currentText(),
                   self.cmb_dpi.currentText(), self.cmb_depth.currentText()))
            add("Fast mode: %s" % self.chk_fast.isChecked())
        except Exception:
            pass
        add("Scanner IP: %s" % self.ip_field.text())
        add("")

        # ---- TWAIN probe ----
        add("----- TWAIN sources -----")
        if HAS_TWAIN:
            try:
                names = list_sources(int(self.winId()))
                if names:
                    for n in names:
                        add("  * %s" % n)
                else:
                    add("  (koi TWAIN source nahi mila)")
            except Exception as exc:
                add("  TWAIN error: %s" % exc)
        else:
            add("  TWAIN (pytwain) install nahi hai")
        add("")

        # ---- WIA probe + duplex capability ----
        add("----- WIA devices -----")
        if HAS_W32:
            try:
                try:
                    _pythoncom.CoInitialize()
                except Exception:
                    pass
                devs = list_wia_sources()
                if not devs:
                    add("  (koi WIA device nahi mila)")
                for dev_id, dev_name in devs:
                    add("  * %s" % dev_name)
                    try:
                        dm = _w32.Dispatch("WIA.DeviceManager")
                        device = None
                        for i in range(1, dm.DeviceInfos.Count + 1):
                            info = dm.DeviceInfos.Item(i)
                            if info.DeviceID == dev_id:
                                device = info.Connect(); break
                        if device is not None:
                            for p in device.Properties:
                                try:
                                    pid = int(p.PropertyID)
                                except Exception:
                                    continue
                                if pid == 3086:
                                    add("      Duplex capable (handling caps value): %s" % p.Value)
                                if pid == 3088:
                                    add("      Current handling select: %s" % p.Value)
                    except Exception as exc:
                        add("      (device probe error: %s)" % exc)
            except Exception as exc:
                add("  WIA error: %s" % exc)
            finally:
                try:
                    _pythoncom.CoUninitialize()
                except Exception:
                    pass
        else:
            add("  WIA (pywin32) install nahi hai")
        add("")
        add("Last scan error: %s" % getattr(self, "_last_error", "(none)"))
        add("================================")

        report = "\n".join(lines)
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Test / Diagnostics"); dlg.resize(640, 560)
        lay = QtWidgets.QVBoxLayout(dlg)
        lay.addWidget(QtWidgets.QLabel(
            "Ye report copy karke share karein (scanner + error info):"
            if self._lang == "hi" else "Copy and share this report:"))
        box = QtWidgets.QPlainTextEdit(); box.setPlainText(report); box.setReadOnly(True)
        box.setStyleSheet("font-family: Consolas, monospace; font-size: 12px;")
        lay.addWidget(box, 1)
        row = QtWidgets.QHBoxLayout()
        b_copy = QtWidgets.QPushButton("Copy karo" if self._lang == "hi" else "Copy")
        def _copy():
            QtWidgets.QApplication.clipboard().setText(report)
            self.status.showMessage("Report copy ho gaya", 3000)
        b_copy.clicked.connect(_copy)
        b_save = QtWidgets.QPushButton("File me save karo" if self._lang == "hi" else "Save to file")
        def _save():
            p, _ = QtWidgets.QFileDialog.getSaveFileName(
                dlg, "Save report", os.path.join(os.path.expanduser("~"), "ApneScan_diagnostics.txt"),
                "Text (*.txt)")
            if p:
                try:
                    with open(p, "w", encoding="utf-8") as fh:
                        fh.write(report)
                    self.status.showMessage("Saved: %s" % p, 5000)
                except Exception as exc:
                    self._warn("Save fail: %s" % exc)
        b_save.clicked.connect(_save)
        b_close = QtWidgets.QPushButton("Band karo" if self._lang == "hi" else "Close")
        b_close.clicked.connect(dlg.accept)
        row.addWidget(b_copy); row.addWidget(b_save); row.addStretch(1); row.addWidget(b_close)
        lay.addLayout(row)
        dlg.exec_()

    def show_about(self):
        QtWidgets.QMessageBox.information(
            self, "About ApneScan",
            "ApneScan\nFree document scanning + PDF tool.\n"
            "TWAIN + WIA scanners support.\nVersion " + VERSION)

    def toggle_simple_mode(self):
        self._opts["simple_mode"] = bool(self.act_simple.isChecked())
        self._save_opts()
        self._apply_simple_mode()

    def _apply_simple_mode(self):
        simple = bool(self._opts.get("simple_mode"))
        for w in getattr(self, "_adv_btns", []):
            try:
                w.setVisible(not simple)
            except Exception:
                pass


    def show_whatsnew(self):
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle(tr("whatsnew", self._lang)); dlg.resize(560, 460)
        lay = QtWidgets.QVBoxLayout(dlg)
        br = QtWidgets.QTextBrowser(); br.setHtml(CHANGELOG_HTML); lay.addWidget(br)
        b = QtWidgets.QPushButton("Close" if self._lang == "en" else "Band karo")
        b.clicked.connect(dlg.accept); lay.addWidget(b)
        dlg.exec_()

    def send_feedback(self):
        title = tr("feedback", self._lang)
        text, ok = QtWidgets.QInputDialog.getMultiLineText(
            self, title,
            ("Aapka sujhav ya problem likhein:" if self._lang == "hi"
             else "Write your feedback or problem:"))
        if not ok or not text.strip():
            return
        email = self._opts.get("feedback_email", "").strip()
        if not email:
            email, ok2 = QtWidgets.QInputDialog.getText(
                self, title,
                ("Feedback kis email par bheju? (ek baar set ho jayega)"
                 if self._lang == "hi" else "Which email should feedback go to?"))
            if not ok2 or not email.strip():
                QtWidgets.QMessageBox.information(self, title, text)
                return
            self._opts["feedback_email"] = email.strip(); self._save_opts()
            email = email.strip()
        try:
            from PyQt5.QtGui import QDesktopServices
            body = QtCore.QUrl.toPercentEncoding(text).data().decode()
            url = QtCore.QUrl("mailto:%s?subject=ApneScan%%20Feedback&body=%s" % (email, body))
            QDesktopServices.openUrl(url)
        except Exception:
            QtWidgets.QMessageBox.information(self, title,
                                              "Email: %s\n\n%s" % (email, text))

    # ---- UI ----
    def _build_ui(self):
        central = QtWidgets.QWidget(); self.setCentralWidget(central)
        outer = QtWidgets.QVBoxLayout(central); outer.setContentsMargins(0, 0, 0, 0); outer.setSpacing(0)

        # ---------- Top toolbar ----------
        tbwrap = QtWidgets.QWidget(); tbwrap.setObjectName("toolbar")
        tb = QtWidgets.QHBoxLayout(tbwrap); tb.setContentsMargins(8, 4, 8, 4); tb.setSpacing(2)
        self._adv_btns = []

        tips = {
            "scan": "Scan shuru karo (F5)", "profiles": "Scan profiles banao/badlo",
            "ocr": "OCR: searchable PDF (dabakar on karo)", "fast": "Fast: 200 dpi + B&W, sabse tez",
            "import": "Computer se image daalo", "savepdf": "PDF save karo (tir se: sab/selected)",
            "images": "Pages ko JPG/PNG me save karo", "print": "Print karo (Ctrl+P)",
            "rotate": "Selected page ghumao", "up": "Page upar karo", "down": "Page neeche karo",
            "delete": "Selected page hatao (Ctrl+Z se wapas)", "clear": "Sab pages hatao",
            "language": "Bhasha badlo (Hindi/English)", "about": "App ke baare me",
        }
        def tbtn(kind, label, fn, advanced=False, checkable=False):
            b = QtWidgets.QToolButton()
            b.setToolTip(tips.get(kind, label))
            b.setToolButtonStyle(QtCore.Qt.ToolButtonTextUnderIcon)
            b.setIcon(_make_icon(kind)); b.setIconSize(QtCore.QSize(28, 28))
            b.setText(label); b.setAutoRaise(True); b.setMinimumWidth(62)
            if checkable:
                b.setCheckable(True)
                b.toggled.connect(fn) if fn else None
            else:
                b.clicked.connect(fn)
            tb.addWidget(b)
            if advanced:
                self._adv_btns.append(b)
            return b

        self.btn_scan_top = tbtn("scan", tr("scan", self._lang), self.do_scan)
        self.btn_profiles = tbtn("profiles", tr("profiles", self._lang).replace("…", ""), self.open_profiles, advanced=True)
        self.chk_ocr = tbtn("ocr", "OCR", None, checkable=True)
        if not HAS_OCR_LIBS:
            self.chk_ocr.setEnabled(False)
        self.chk_fast = tbtn("fast", tr("fast", self._lang), None, checkable=True)
        self.chk_fast.setToolTip("Fast: 200 dpi + Black & White, bina extra processing")
        self.chk_fast.setChecked(bool(self._opts.get("fast_mode")))
        self.chk_fast.toggled.connect(self._on_fast_toggled)
        tb.addWidget(self._vsep())
        self.btn_import = tbtn("import", tr("import", self._lang), self.import_images, advanced=True)
        self.btn_save_pdf = tbtn("savepdf", tr("save_pdf", self._lang), self.save_pdf)
        tbtn("images", "Save Images", self.save_images, advanced=True)
        self.btn_print = tbtn("print", "Print", self.print_all, advanced=True)
        tb.addWidget(self._vsep())
        tbtn("rotate", "Rotate", self.rotate_right, advanced=True)
        tbtn("up", tr("up", self._lang), self.move_up, advanced=True)
        tbtn("down", tr("down", self._lang), self.move_down, advanced=True)
        tbtn("delete", tr("delete", self._lang), self.delete_page, advanced=True)
        tbtn("clear", tr("clear", self._lang), self.clear_all, advanced=True)
        tb.addStretch(1)
        tbtn("language", "Language", self.choose_language)
        tbtn("about", tr("about", self._lang), self.show_about)
        pdfmenu = QtWidgets.QMenu(self.btn_save_pdf)
        pdfmenu.addAction(tr("save_all", self._lang), self.save_pdf_all)
        pdfmenu.addAction(tr("save_sel", self._lang), self.save_pdf_selected)
        self.btn_save_pdf.setMenu(pdfmenu)
        self.btn_save_pdf.setPopupMode(QtWidgets.QToolButton.MenuButtonPopup)
        printmenu = QtWidgets.QMenu(self.btn_print)
        printmenu.addAction("All print (sabhi pages)", self.print_all)
        printmenu.addAction("Selected print (chune hue)", self.print_selected)
        printmenu.addAction("ID print (2 ID ek page par)", self.print_ids)
        printmenu.addAction("ID print - sirf selected", self.print_ids_selected)
        self.btn_print.setMenu(printmenu)
        self.btn_print.setPopupMode(QtWidgets.QToolButton.MenuButtonPopup)
        outer.addWidget(tbwrap)
        hr = QtWidgets.QFrame(); hr.setObjectName("hr"); hr.setFrameShape(QtWidgets.QFrame.HLine); outer.addWidget(hr)

        # ---------- Body: left settings panel | thumbnails ----------
        body = QtWidgets.QHBoxLayout(); body.setContentsMargins(0, 0, 0, 0); body.setSpacing(0)
        panel = QtWidgets.QWidget(); panel.setObjectName("panel"); panel.setFixedWidth(252)
        pl = QtWidgets.QVBoxLayout(panel); pl.setContentsMargins(14, 14, 14, 14); pl.setSpacing(5)

        pl.addWidget(QtWidgets.QLabel(tr("profile", self._lang)))
        prow = QtWidgets.QHBoxLayout(); prow.setSpacing(4)
        self.cmb_profile = QtWidgets.QComboBox(); self.cmb_profile.currentTextChanged.connect(self._on_profile_changed)
        prow.addWidget(self.cmb_profile, 1)
        bnew = QtWidgets.QPushButton("+"); bnew.setFixedWidth(30); bnew.clicked.connect(self._quick_new_profile); prow.addWidget(bnew)
        bedit = QtWidgets.QPushButton("✎"); bedit.setFixedWidth(30); bedit.clicked.connect(self._quick_edit_profile); prow.addWidget(bedit)
        pw = QtWidgets.QWidget(); pw.setLayout(prow); pl.addWidget(pw)

        pl.addSpacing(4); pl.addWidget(QtWidgets.QLabel("Device:"))
        self.dev_lbl = QtWidgets.QLabel("(koi device nahi)"); self.dev_lbl.setObjectName("dev"); self.dev_lbl.setWordWrap(True); pl.addWidget(self.dev_lbl)
        self.method_lbl = QtWidgets.QLabel(""); self.method_lbl.setObjectName("dev")
        pl.addWidget(self.method_lbl)

        pl.addSpacing(4); pl.addWidget(QtWidgets.QLabel("Claim No.:"))
        self.claim_edit = QtWidgets.QLineEdit(); self.claim_edit.setPlaceholderText("optional"); pl.addWidget(self.claim_edit)

        pl.addWidget(QtWidgets.QLabel("Paper source:"))
        self.cmb_source = QtWidgets.QComboBox(); self.cmb_source.addItems(["Feeder (ADF)", "Glass (Flatbed)"]); pl.addWidget(self.cmb_source)
        pl.addWidget(QtWidgets.QLabel("Scan sides:"))
        self.cmb_sides = QtWidgets.QComboBox()
        self.cmb_sides.addItems(["Single side (ek taraf)", "Both side (dono taraf)"])
        self.cmb_sides.setToolTip("Both side = kaagaz ke dono taraf scan (duplex)")
        pl.addWidget(self.cmb_sides)
        pl.addWidget(QtWidgets.QLabel("Page size:"))
        self.cmb_pagesize = QtWidgets.QComboBox(); self.cmb_pagesize.addItems(["Auto (alag-alag size khud pakde)", "A4 (210x297 mm)", "Letter", "Legal", "A5"]); pl.addWidget(self.cmb_pagesize)
        self.cmb_pagesize.setToolTip("Auto = har page ki asli size khud detect (mixed size / ID card / aadha page bhi poora). A4/Letter/Legal = fixed size.")
        pl.addWidget(QtWidgets.QLabel("Resolution:"))
        self.cmb_dpi = QtWidgets.QComboBox(); self.cmb_dpi.addItems([d + " dpi" for d in RESOLUTIONS]); self.cmb_dpi.setCurrentText("200 dpi"); pl.addWidget(self.cmb_dpi)
        pl.addWidget(QtWidgets.QLabel("Bit depth:"))
        self.cmb_depth = QtWidgets.QComboBox(); self.cmb_depth.addItems(["24-bit Colour", "Grayscale", "Black & White"]); self.cmb_depth.setCurrentText("Grayscale"); pl.addWidget(self.cmb_depth)

        pl.addStretch(1)
        self.btn_scan = QtWidgets.QPushButton("▶  " + tr("scan", self._lang)); self.btn_scan.setObjectName("primary")
        self.btn_scan.setMinimumHeight(38); self.btn_scan.clicked.connect(self.do_scan); pl.addWidget(self.btn_scan)
        self.btn_scan.setToolTip("Scan shuru karo (F5)")
        self.claim_edit.setToolTip("Claim/Patient number (file ke naam me aayega)")
        self.cmb_dpi.setToolTip("Resolution: kam dpi = tez scan")
        self.cmb_depth.setToolTip("Black & White sabse tez, Colour dhima")
        self.setAcceptDrops(True)
        body.addWidget(panel)

        vline = QtWidgets.QFrame(); vline.setObjectName("hr"); vline.setFrameShape(QtWidgets.QFrame.VLine); body.addWidget(vline)

        self.list = QtWidgets.QListWidget()
        self.list.setViewMode(QtWidgets.QListView.IconMode)
        self.list.setIconSize(QtCore.QSize(self.THUMB_W, self.THUMB_H))
        self.list.setResizeMode(QtWidgets.QListView.Adjust)
        self.list.setMovement(QtWidgets.QListView.Snap)
        self.list.setDragDropMode(QtWidgets.QAbstractItemView.InternalMove)
        self.list.setSpacing(10); self.list.setUniformItemSizes(True)
        self.list.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self._thumb_w, self._thumb_h = self.THUMB_W, self.THUMB_H   # zoomable display size
        self._apply_thumb_zoom(self.THUMB_W)
        # empty-state hint (shown when no pages)
        self._empty_lbl = QtWidgets.QLabel(
            "\u2b07  Scan dabao ya Import se image daalo\n\n(ya kisi image ko yahan drag karo)"
            if self._lang == "hi" else
            "\u2b07  Press Scan, or Import an image\n\n(or drag an image here)",
            self.list.viewport())
        self._empty_lbl.setAlignment(QtCore.Qt.AlignCenter)
        self._empty_lbl.setStyleSheet("color:#94a3b8; font-size:15px;")
        self.list.viewport().installEventFilter(self)
        self.list.itemDoubleClicked.connect(self._open_preview_dialog)
        self.list.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.list.customContextMenuRequested.connect(self._list_context_menu)
        body.addWidget(self.list, 1)
        outer.addLayout(body, 1)

        # ---------- Status bar: connection ----------
        self.status = self.statusBar()
        self.lbl_conn = QtWidgets.QLabel(); self.lbl_conn.setTextFormat(QtCore.Qt.RichText)
        self._set_conn_display(None, tr("checking", self._lang))
        self.status.addWidget(self.lbl_conn)
        self.lbl_busy = QtWidgets.QLabel(); self.lbl_busy.setTextFormat(QtCore.Qt.RichText)
        self.status.addPermanentWidget(self.lbl_busy)
        self._set_busy_display("unknown")
        self._state_timer = QtCore.QTimer(self)
        self._state_timer.setInterval(6000)
        self._state_timer.timeout.connect(self._tick_scanner_state)
        self._state_timer.start()
        QtCore.QTimer.singleShot(800, self._tick_scanner_state)
        self._stats_timer = QtCore.QTimer(self)
        self._stats_timer.setInterval(90000)
        self._stats_timer.timeout.connect(lambda: self._refresh_stats("ping"))
        self._stats_timer.start()
        QtCore.QTimer.singleShot(1500, lambda: self._refresh_stats("ping"))
        # kept (hidden) for the connection feature; IP set via Settings menu
        self.ip_field = QtWidgets.QLineEdit(self._config.get("scanner_ip", "")); self.ip_field.hide()
        self.btn_check = QtWidgets.QPushButton(); self.btn_check.hide()
        self._pb_holder = None


    def _apply_style(self):
        if self._opts.get("theme") == "dark":
            self.setStyleSheet("""
                QMainWindow, QWidget { background:#0f172a; color:#e2e8f0; }
                QMenuBar { background:#0f172a; color:#e2e8f0; }
                QMenuBar::item:selected { background:#1e293b; }
                QMenu { background:#1e293b; color:#e2e8f0; border:1px solid #334155; }
                QMenu::item:selected { background:#334155; }
                #toolbar { background:#111827; }
                QToolButton { border:1px solid transparent; border-radius:8px; padding:4px 6px; color:#cbd5e1; font-size:11px; }
                QToolButton:hover { background:#1e293b; border-color:#334155; }
                QToolButton:checked { background:#134e4a; border-color:#2dd4bf; color:#5eead4; }
                #panel { background:#111827; }
                #dev { color:#94a3b8; font-size:12px; }
                #hr { color:#334155; }
                QLabel { color:#cbd5e1; }
                QLineEdit, QComboBox, QSpinBox, QPlainTextEdit { background:#1e293b; border:1px solid #475569; border-radius:6px; padding:4px 8px; color:#e2e8f0; }
                QPushButton { background:#1e293b; border:1px solid #475569; border-radius:6px; padding:6px 12px; color:#e2e8f0; }
                QPushButton:hover { background:#334155; }
                QPushButton#primary { background:#0f766e; border:1px solid #0f766e; color:#fff; font-weight:700; }
                QPushButton#primary:hover { background:#115e59; }
                QListWidget { background:#0b1220; border:none; }
                QListWidget::item:selected { background:#134e4a; color:#e2e8f0; }
                QScrollArea { background:#111827; }
            """)
        else:
            self.setStyleSheet("""
                QMainWindow, QWidget { background:#f4f6f8; color:#1f2933; }
                QMenuBar { background:#f4f6f8; }
                #toolbar { background:#ffffff; }
                QToolButton { border:1px solid transparent; border-radius:8px; padding:4px 6px; color:#334155; font-size:11px; }
                QToolButton:hover { background:#eef2f6; border-color:#e2e8f0; }
                QToolButton:checked { background:#d1eae7; border-color:#0f766e; color:#0f766e; }
                #panel { background:#ffffff; }
                #dev { color:#475569; font-size:12px; }
                #hr { color:#e2e8f0; }
                QLabel { color:#334155; }
                QLineEdit, QComboBox, QSpinBox { background:#fff; border:1px solid #cbd5e1; border-radius:6px; padding:4px 8px; }
                QPushButton { background:#fff; border:1px solid #cbd5e1; border-radius:6px; padding:6px 12px; color:#1f2933; }
                QPushButton:hover { background:#eef2f6; }
                QPushButton#primary { background:#0f766e; border:1px solid #0f766e; color:#fff; font-weight:700; }
                QPushButton#primary:hover { background:#115e59; }
                QListWidget { background:#fbfcfd; border:none; }
                QListWidget::item:selected { background:#d1eae7; color:#0f172a; }
            """)


    def _set_conn_display(self, state, message):
        colour = {True: "#16a34a", False: "#dc2626", None: "#9ca3af"}[state]
        dot = '<span style="color:%s; font-size:16px;">&#9679;</span>' % colour
        self.lbl_conn.setText("%s&nbsp; %s" % (dot, message))

    def check_connection(self):
        ip = self.ip_field.text().strip()
        self._config["scanner_ip"] = ip; save_config(self._config)
        if not ip:
            self._set_conn_display(False, "Scanner IP daalein"); return
        self._set_conn_display(None, "Check ho raha hai... (%s)" % ip)
        self.btn_check.setEnabled(False)
        self._checker = ConnectionChecker(ip)
        self._checker.result.connect(self._on_conn_result)
        self._checker.finished.connect(lambda: self.btn_check.setEnabled(True))
        self._checker.start()

    def _on_conn_result(self, ok, message):
        self._set_conn_display(bool(ok), message)

    def _set_busy_display(self, kind):
        if kind == "free":
            txt, col = "Scanner FREE (taiyar)", "#16a34a"
        elif kind == "busy":
            txt, col = "Scanner BUSY (vyast)", "#dc2626"
        else:
            txt, col = "Scanner: --", "#9ca3af"
        self.lbl_busy.setText(
            '<span style="color:%s; font-size:15px;">&#9679;</span>&nbsp;<b>%s</b>' % (col, txt))

    def _tick_scanner_state(self):
        # IMPORTANT: do NOT poll the scanner over the network here. This HP scanner
        # allows only one eSCL connection at a time, so a background status poll
        # collides with the real scan and makes it fail with HTTP 503. The indicator
        # simply reflects OUR own activity: busy only while we are scanning.
        if getattr(self, "_worker", None) is not None and self._worker.isRunning():
            self._set_busy_display("busy")
        elif getattr(self, "_scanning", False):
            self._set_busy_display("busy")
        else:
            self._set_busy_display("free")

    # ---- list ----
    def _add_item_for_path(self, path):
        icon = QtGui.QIcon(self._make_thumb(path))
        _lbl = ("Page %d" % (self.list.count() + 1)) if self._opts.get("show_page_numbers", True) else ""
        item = QtWidgets.QListWidgetItem(icon, _lbl)
        item.setData(QtCore.Qt.UserRole, path); item.setTextAlignment(QtCore.Qt.AlignHCenter)
        self.list.addItem(item); self.list.setCurrentItem(item)
        self.list.clearSelection()  # nothing "selected" by default; user picks with Ctrl/Shift
        self._dirty = True
        self._update_status(); self._update_empty_state()

    THUMB_HI_W = 360
    THUMB_HI_H = 480

    def _make_thumb(self, path):
        pix = QtGui.QPixmap(path)
        if pix.isNull():
            return QtGui.QPixmap(self.THUMB_HI_W, self.THUMB_HI_H)
        # high-res source; the list's iconSize scales it for display (crisp zoom)
        return pix.scaled(self.THUMB_HI_W, self.THUMB_HI_H,
                          QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)

    def _refresh_item(self, item):
        path = item.data(QtCore.Qt.UserRole)
        item.setIcon(QtGui.QIcon(self._make_thumb(path)))
        if item is self.list.currentItem():
            self._show_preview(item, None)

    def _ordered_paths(self):
        return [self.list.item(i).data(QtCore.Qt.UserRole) for i in range(self.list.count())]

    def _selected_paths(self):
        # selected items, kept in the on-screen (list) order
        sel = set(self.list.row(it) for it in self.list.selectedItems())
        return [self.list.item(i).data(QtCore.Qt.UserRole) for i in range(self.list.count()) if i in sel]

    def _show_preview(self, current, _prev=None):
        # NAPS2-style: no side preview pane; double-click opens a viewer.
        return


    def resizeEvent(self, event):
        super().resizeEvent(event)


    def do_scan(self):
        method = self._opts.get("scanner_method", "twain")
        prof = self._selected_profile()
        if method == "escl":
            ip = self.ip_field.text().strip()
            self._opts["scanner_ip"] = ip
            if not ip:
                self._warn("Scanner IP set nahi hai. Settings \u2192 Scanner IP me IP daalo (jaise 192.168.1.8)."); return
        elif method == "naps2":
            if not (self._opts.get("naps2_path") or find_naps2()):
                self._warn("NAPS2 nahi mila. Settings \u2192 Scan method \u2192 naps2 me path set karo."); return
            if not self._opts.get("naps2_profile"):
                self._warn("Settings \u2192 Scan method \u2192 naps2 me NAPS2 profile ka naam set karo."); return
        else:
            if not HAS_TWAIN and method == "twain":
                self._warn("TWAIN (pytwain) install nahi hai."); return
            if prof is None:
                QtWidgets.QMessageBox.information(self, APP_NAME, "Pehli baar ek profile banayein.\n'Profiles…' → New.")
                self.open_profiles(); prof = self._selected_profile()
                if prof is None:
                    return
            if method == "twain" and not prof.get("source_name"):
                self._warn("Is profile me scanner set nahi hai. 'Profiles…' → Edit → Choose device."); return
        self._barcode_tried = False
        self._scan_count = 0
        self._progress = ScanProgressDialog(self, prof.get("source_name") or APP_NAME, self._lang)
        self._progress.cancelled.connect(self._cancel_scan)
        dpi, color, duplex = self._panel_scan_params(prof or {})
        opts = self._opts
        if self.chk_fast.isChecked():
            dpi, color = 200, "bw"     # keep the user's duplex (both-side) choice
            opts = dict(self._opts)    # lean copy: skip heavy per-page processing
            for k in ("remove_blank", "auto_crop", "deskew", "quality_enhance"):
                opts[k] = False
        opts = dict(opts)
        opts["page_size"] = self.cmb_pagesize.currentText().strip().lower()
        self._worker = ScanWorker(int(self.winId()), (prof or {}).get("source_name"),
                                  dpi, color, duplex, self._tmpdir, opts)
        self._worker.page_done.connect(self._on_page_scanned)
        self._worker.done.connect(self._on_scan_done)
        self._worker.failed.connect(self._on_scan_failed)
        self.btn_scan.setEnabled(False)
        try:
            self._conn_timer.stop()
        except Exception:
            pass
        self._scanning = True
        try:
            self._state_timer.stop()
        except Exception:
            pass
        self._worker.start(); self._progress.show()
        self._set_busy_display("busy")

    def _on_page_scanned(self, path):
        self._add_item_for_path(path)
        self._scan_count += 1
        if self._progress:
            self._progress.set_page(self._scan_count)
        if (self._opts.get("barcode_autofill") and not self._barcode_tried
                and not self.claim_edit.text().strip()):
            self._barcode_tried = True
            code = read_barcode(path)
            if code:
                self.claim_edit.setText(code)
                self.status.showMessage("Barcode se claim number mila: %s" % code, 5000)

    def _on_scan_done(self, kept, skipped):
        if self._progress:
            self._progress.close(); self._progress = None
        self._scanning = False
        self.btn_scan.setEnabled(True)
        try:
            self._state_timer.start()
        except Exception:
            pass
        QtCore.QTimer.singleShot(2500, self._tick_scanner_state)
        try:
            self._conn_timer.start()
        except Exception:
            pass
        msg = "%d page scan ho gaye." % kept
        if skipped:
            msg += " (%d blank hataye)" % skipped
        self.status.showMessage(msg, 5000)
        self._report_scan_stat(kept)
        if self._opts.get("auto_name") and kept:
            self.auto_name_pages()
        if self._opts.get("auto_save") and kept:
            saved = self._auto_save_pdf()
            if saved and self._opts.get("batch_mode"):
                self._start_next_batch()

    def _on_scan_failed(self, msg):
        self._last_error = msg
        if self._progress:
            self._progress.close(); self._progress = None
        self._scanning = False
        self.btn_scan.setEnabled(True)
        try:
            self._state_timer.start()
        except Exception:
            pass
        QtCore.QTimer.singleShot(2500, self._tick_scanner_state)
        try:
            self._conn_timer.start()
        except Exception:
            pass
        self._warn(friendly_error(msg, self._opts.get("language", "hi")))

    def _cancel_scan(self):
        try:
            if getattr(self, "_worker", None):
                self._worker.requestInterruption()
        except Exception:
            pass
        if self._progress:
            self._progress.close(); self._progress = None
        self.btn_scan.setEnabled(True)
        try:
            self._conn_timer.start()
        except Exception:
            pass

    def _start_next_batch(self):
        for p in self._ordered_paths():
            try:
                os.remove(p)
            except Exception:
                pass
        self.list.clear(); self._show_preview(None)
        text, ok = QtWidgets.QInputDialog.getText(self, "Batch mode", "Agle document ka Claim No.:")
        if ok:
            self.claim_edit.setText(text.strip())
        self._update_status()

    def import_images(self):
        start = self._config.get("last_import_dir") or self._config.get("last_save_dir") or ""
        files, _ = QtWidgets.QFileDialog.getOpenFileNames(
            self, "Images / PDF chuno", start,
            "Images & PDF (*.png *.jpg *.jpeg *.bmp *.tif *.tiff *.pdf);;PDF (*.pdf);;Images (*.png *.jpg *.jpeg *.bmp *.tif *.tiff)")
        count = 0
        for f in files:
            count += self._import_one(f)
        if files:
            try:
                self._config["last_import_dir"] = os.path.dirname(files[0])
                save_config(self._config)
            except Exception:
                pass
        if count:
            self.status.showMessage("%d page import ho gaye." % count, 4000)
            if tesseract_available():
                self.auto_name_pages()

    def _import_one(self, f):
        """Import a single image OR pdf file. Returns number of pages added."""
        added = 0
        try:
            if f.lower().endswith(".pdf"):
                pages = pdf_to_images(f, self._tmpdir)
                if not pages:
                    self._warn("Is PDF se pages nahi nikle.\nBehtar result ke liye 'PyMuPDF' install karein:\n  py -3.12-32 -m pip install PyMuPDF")
                for png in pages:
                    self._add_item_for_path(png); added += 1
            else:
                with Image.open(f) as im:
                    fd, png = tempfile.mkstemp(suffix=".png", dir=self._tmpdir)
                    os.close(fd); im.convert("RGB").save(png, "PNG")
                    self._add_item_for_path(png); added += 1
        except Exception:
            pass
        return added

    # ---- page edit ----
    def _current_item_or_warn(self):
        item = self.list.currentItem()
        if item is None:
            self._warn("Pehle koi page select karein.")
        return item

    def _rotate(self, angle):
        item = self._current_item_or_warn()
        if not item:
            return
        path = item.data(QtCore.Qt.UserRole)
        try:
            with Image.open(path) as im:
                im.rotate(angle, expand=True).save(path)
        except Exception as exc:
            self._warn("Rotate fail:\n%s" % exc); return
        self._dirty = True
        self._refresh_item(item)

    def rotate_left(self):
        self._rotate(90)

    def rotate_right(self):
        self._rotate(-90)

    def _enhance_current(self, brightness, contrast):
        item = self._current_item_or_warn()
        if not item:
            return
        if enhance_image(item.data(QtCore.Qt.UserRole), brightness, contrast):
            self._refresh_item(item)

    def autocrop_current(self):
        item = self._current_item_or_warn()
        if not item:
            return
        path = item.data(QtCore.Qt.UserRole)
        try:
            with Image.open(path) as im:
                autocrop(im.convert("RGB")).save(path, "PNG")
        except Exception as exc:
            self._warn("Crop fail:\n%s" % exc); return
        self._refresh_item(item)

    def _list_context_menu(self, pos):
        if self.list.count() == 0:
            return
        menu = QtWidgets.QMenu(self)
        act_rename = menu.addAction("\u270f Naam badlo / Rename")
        act_del = menu.addAction("\U0001f5d1 Delete")
        chosen = menu.exec_(self.list.viewport().mapToGlobal(pos))
        if chosen == act_rename:
            self.rename_current_page()
        elif chosen == act_del:
            self.delete_page()

    def rename_current_page(self):
        it = self.list.currentItem() or (self.list.selectedItems() or [None])[0]
        if it is None:
            self._warn("Pehle koi page select karein."); return
        cur = it.data(TITLE_ROLE) or ""
        name, ok = QtWidgets.QInputDialog.getText(self, "Naam badlo", "Is page ka naam:", text=cur)
        if not ok:
            return
        name = underscore_name(name)
        if not name:
            return
        it.setData(TITLE_ROLE, name)
        it.setText(name)
        # LEARN: remember this document's content -> this name, so next time a
        # similar page is scanned it gets named automatically.
        self._learn_name(it.data(QtCore.Qt.UserRole), name)

    def _learn_name(self, path, name):
        if not path or not tesseract_available():
            return
        try:
            words = sorted(sig_words(page_ocr_text(path, 0.6)))
        except Exception:
            words = []
        if len(words) < 4:
            return   # too little text to identify the form reliably
        learned = self._config.setdefault("learned_names", [])
        wset = set(words)
        # if a very similar signature already exists, update its name; else add new
        for e in learned:
            if _jaccard(wset, set(e.get("words", []))) >= 0.6:
                e["words"] = words
                e["name"] = name
                break
        else:
            learned.append({"words": words, "name": name})
            # keep the list from growing without bound
            if len(learned) > 300:
                del learned[0:len(learned) - 300]
        try:
            save_config(self._config)
        except Exception:
            pass
        self.status.showMessage("Naam yaad rakh liya \u2014 agli baar aisa document apne aap '%s' ho jayega." % name, 5000)

    def _list_context_menu_end(self):
        pass

    def delete_page(self):
        items = self.list.selectedItems()
        if not items:
            item = self._current_item_or_warn()
            if not item:
                return
            items = [item]
        rows = sorted((self.list.row(it) for it in items), reverse=True)
        for row in rows:
            it = self.list.item(row)
            if it is None:
                continue
            path = it.data(QtCore.Qt.UserRole)
            self.list.takeItem(row)
            self._undo_stack.append((path, row))
        self._undo_stack = self._undo_stack[-15:]
        self._dirty = True
        self._renumber_pages()
        self._update_status(); self._update_empty_state()

    def move_up(self):
        row = self.list.currentRow()
        if row > 0:
            it = self.list.takeItem(row); self.list.insertItem(row - 1, it); self.list.setCurrentItem(it)

    def move_down(self):
        row = self.list.currentRow()
        if 0 <= row < self.list.count() - 1:
            it = self.list.takeItem(row); self.list.insertItem(row + 1, it); self.list.setCurrentItem(it)

    def clear_all(self):
        if self.list.count() == 0:
            return
        if QtWidgets.QMessageBox.question(self, "Confirm", "Saare pages hata dein?") != QtWidgets.QMessageBox.Yes:
            return
        for p in self._ordered_paths():
            try:
                os.remove(p)
            except Exception:
                pass
        self.list.clear()
        for p, _r in self._undo_stack:
            try:
                os.remove(p)
            except Exception:
                pass
        self._undo_stack = []
        self._dirty = False
        self._update_status(); self._update_empty_state()

    # ---- filenames + saving ----
    def _remember_save_dir(self, path):
        try:
            d = os.path.dirname(path)
            if d and os.path.isdir(d):
                self._config["last_save_dir"] = d
                save_config(self._config)
        except Exception:
            pass

    def _target_folder(self):
        # If the user isn't using auto-organizing folders, reopen wherever they
        # saved last time.
        last = self._config.get("last_save_dir")
        if (last and os.path.isdir(last)
                and not self._opts.get("year_month_folders")
                and not self._opts.get("make_claim_folder")):
            return last
        folder = self._opts.get("save_folder") or os.path.join(os.path.expanduser("~"), "Documents")
        now = datetime.datetime.now()
        if self._opts.get("year_month_folders"):
            folder = os.path.join(folder, now.strftime("%Y"), now.strftime("%m"))
        claim = sanitize(self.claim_edit.text().strip(), "")
        if self._opts.get("make_claim_folder") and claim:
            folder = os.path.join(folder, claim)
        try:
            os.makedirs(folder, exist_ok=True)
        except Exception:
            pass
        return folder

    def _suggested_save_name(self, paths=None):
        pathset = set(paths) if paths else None
        for i in range(self.list.count()):
            it = self.list.item(i)
            if pathset is not None and it.data(QtCore.Qt.UserRole) not in pathset:
                continue
            t = it.data(TITLE_ROLE)
            if t:
                return underscore_name(t)
        return None

    def _remember_doc_name(self, out):
        base = os.path.splitext(os.path.basename(out))[0]
        keys = []
        for i in range(self.list.count()):
            k = self.list.item(i).data(NAMEKEY_ROLE)
            if k:
                keys.append(k)
        if not keys:
            return
        dn = self._config.setdefault("doc_names", {})
        dn[keys[0]] = base
        try:
            save_config(self._config)
        except Exception:
            pass

    def _build_filename(self, ext=".pdf", seq=1, paths=None):
        suggested = self._suggested_save_name(paths)
        if suggested:
            return os.path.join(self._target_folder(), suggested + ext)
        now = datetime.datetime.now()
        claim = sanitize(self.claim_edit.text().strip(), "scan")
        name = self._opts.get("filename_template", "{claim}_{date}_{seq}")
        name = (name.replace("{claim}", claim).replace("{date}", now.strftime("%Y-%m-%d"))
                    .replace("{time}", now.strftime("%H%M%S")).replace("{seq}", "%03d" % seq))
        return os.path.join(self._target_folder(), underscore_name(name) + ext)

    def _validate_claim_ok(self):
        if not self._opts.get("validate_claim"):
            return True
        claim = self.claim_edit.text().strip()
        pat = self._opts.get("claim_pattern") or r"^[A-Za-z0-9\-]{4,}$"
        try:
            if re.match(pat, claim):
                return True
        except Exception:
            return True
        self._warn("Claim number sahi nahi lag raha (pattern match nahi hua).\n"
                   "Sudhaar karein ya Settings me validation band karein.")
        return False

    def _duplicate_ok(self):
        if not self._opts.get("duplicate_check"):
            return True
        claim = self.claim_edit.text().strip()
        if claim and claim in self._used_claims:
            return QtWidgets.QMessageBox.question(
                self, "Duplicate", "Ye claim number pehle bhi save ho chuka hai.\nPhir bhi save karein?",
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.No) == QtWidgets.QMessageBox.Yes
        return True

    def _pages_as_pdf(self, paths, out, password=None):
        compress = self._opts.get("compress")
        q = int(self._opts.get("jpeg_quality", 60))
        wm = self._opts.get("watermark")
        wt = self._opts.get("watermark_text", "")
        imgs = []
        for p in paths:
            im = Image.open(p).convert("RGB")
            if wm and wt:
                im = apply_watermark(im, wt)
            if compress:
                buf = io.BytesIO(); im.save(buf, "JPEG", quality=q); buf.seek(0)
                im2 = Image.open(buf); im2.load(); im = im2
            imgs.append(im)
        if password:
            tmp = out + ".tmp.pdf"
            imgs[0].save(tmp, "PDF", save_all=True, append_images=imgs[1:], resolution=200.0)
            reader = PdfReader(tmp); writer = PdfWriter()
            for pg in reader.pages:
                writer.add_page(pg)
            writer.encrypt(password)
            with open(out, "wb") as fh:
                writer.write(fh)
            try:
                os.remove(tmp)
            except Exception:
                pass
        else:
            imgs[0].save(out, "PDF", save_all=True, append_images=imgs[1:], resolution=200.0)
        for im in imgs:
            try:
                im.close()
            except Exception:
                pass

    def _record_save(self, out, num_pages):
        claim = self.claim_edit.text().strip()
        # recent
        self._add_recent(out)
        # used claims (duplicate detection)
        if claim:
            self._used_claims.add(claim)
            self._config["used_claims"] = list(self._used_claims)[-2000:]
            save_config(self._config)
        # activity log
        if self._opts.get("activity_log"):
            try:
                logp = os.path.join(self._opts.get("save_folder", "."), "activity_log.txt")
                os.makedirs(os.path.dirname(logp), exist_ok=True)
                with open(logp, "a", encoding="utf-8") as fh:
                    fh.write("%s\tclaim=%s\tpages=%d\t%s\n"
                             % (datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                claim, num_pages, out))
            except Exception:
                pass
        # excel register
        if self._opts.get("excel_log") and HAS_XLSX:
            self._append_excel(claim, num_pages, out)
        # backup
        if self._opts.get("backup"):
            self._backup_file(out)

    def _append_excel(self, claim, num_pages, out):
        try:
            xp = os.path.join(self._opts.get("save_folder", "."), "register.xlsx")
            os.makedirs(os.path.dirname(xp), exist_ok=True)
            if os.path.exists(xp):
                wb = openpyxl.load_workbook(xp)
                ws = wb.active
            else:
                wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Register"
                ws.append(["Date", "Time", "Claim No.", "Pages", "File"])
            now = datetime.datetime.now()
            ws.append([now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S"),
                       claim, num_pages, out])
            wb.save(xp)
        except Exception:
            pass

    def _backup_file(self, out):
        try:
            day = datetime.datetime.now().strftime("%Y-%m-%d")
            bdir = os.path.join(self._opts.get("backup_folder", "."), day)
            os.makedirs(bdir, exist_ok=True)
            shutil.copy2(out, os.path.join(bdir, os.path.basename(out)))
        except Exception:
            pass

    def _auto_save_pdf(self):
        paths = self._ordered_paths()
        if not paths:
            return False
        if not self._validate_claim_ok() or not self._duplicate_ok():
            return False
        out = self._build_filename(".pdf")
        try:
            self._pages_as_pdf(paths, out)
        except Exception:
            self._warn("Auto-save fail:\n%s" % traceback.format_exc()); return False
        self._record_save(out, len(paths)); self._dirty = False
        if self._opts.get("save_images_too"):
            base = os.path.splitext(out)[0]
            for i, p in enumerate(paths):
                try:
                    with Image.open(p) as im:
                        im.convert("RGB").save("%s_p%d.jpg" % (base, i + 1), "JPEG", quality=90)
                except Exception:
                    pass
        self._after_save_action(out)
        self.status.showMessage("Auto-save: %s" % out, 8000)
        return True

    def save_pdf_all(self):
        self.save_pdf(self._ordered_paths())

    def save_pdf_selected(self):
        paths = self._selected_paths()
        if not paths:
            self._warn(tr("no_selection", self._lang)); return
        self.save_pdf(paths)

    def save_pdf(self, paths=None):
        if not isinstance(paths, (list, tuple)):
            paths = self._ordered_paths()
        if not paths:
            self._warn(tr("scan_first", self._lang)); return
        if not self._validate_claim_ok() or not self._duplicate_ok():
            return
        default = self._build_filename(".pdf", paths=paths)
        out, _ = QtWidgets.QFileDialog.getSaveFileName(self, "PDF save karein", default, "PDF (*.pdf)")
        if not out:
            return
        if not out.lower().endswith(".pdf"):
            out += ".pdf"
        ocr = HAS_OCR_LIBS and self.chk_ocr.isChecked()
        try:
            if ocr:
                self._save_ocr_pdf(paths, out)
            else:
                self._pages_as_pdf(paths, out)
        except Exception as exc:
            if HAS_OCR_LIBS and isinstance(exc, pytesseract.TesseractNotFoundError):
                self._warn("Tesseract OCR nahi mila. OCR ke bina try karein.")
            else:
                self._warn("PDF save fail:\n%s" % traceback.format_exc())
            return
        self._remember_save_dir(out); self._remember_doc_name(out)
        self._record_save(out, len(paths)); self._dirty = False; self._after_save_action(out)
        QtWidgets.QMessageBox.information(self, "Ho gaya", "PDF save ho gaya:\n%s" % out)

    def save_pdf_password(self):
        paths = self._ordered_paths()
        if not paths:
            self._warn("Pehle koi page scan/import karein."); return
        if not HAS_OCR_LIBS:
            self._warn("pypdf install nahi hai (password ke liye zaroori)."); return
        pw, ok = QtWidgets.QInputDialog.getText(self, "Password", "PDF ka password:", QtWidgets.QLineEdit.Password)
        if not ok or not pw:
            return
        out, _ = QtWidgets.QFileDialog.getSaveFileName(self, "PDF save karein", self._build_filename(".pdf"), "PDF (*.pdf)")
        if not out:
            return
        if not out.lower().endswith(".pdf"):
            out += ".pdf"
        try:
            self._pages_as_pdf(paths, out, password=pw)
        except Exception:
            self._warn("PDF save fail:\n%s" % traceback.format_exc()); return
        self._remember_save_dir(out); self._remember_doc_name(out)
        self._record_save(out, len(paths)); self._dirty = False; self._after_save_action(out)
        QtWidgets.QMessageBox.information(self, "Ho gaya", "Password-PDF save ho gaya:\n%s" % out)

    def save_images(self):
        paths = self._ordered_paths()
        if not paths:
            self._warn("Pehle koi page scan/import karein."); return
        out, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Images save karein", self._build_filename(".jpg"),
            "JPEG (*.jpg);;PNG (*.png);;TIFF (*.tif)")
        if not out:
            return
        base, ext = os.path.splitext(out)
        if not ext:
            ext = ".jpg"
        fmt = {"jpg": "JPEG", "jpeg": "JPEG", "png": "PNG", "tif": "TIFF", "tiff": "TIFF"}\
            .get(ext.lower().lstrip("."), "JPEG")
        try:
            for i, p in enumerate(paths):
                target = (base + ext) if i == 0 else ("%s_%d%s" % (base, i + 1, ext))
                with Image.open(p) as im:
                    im.convert("RGB").save(target, fmt)
        except Exception:
            self._warn("Images save fail:\n%s" % traceback.format_exc()); return
        self._remember_save_dir(out)
        QtWidgets.QMessageBox.information(self, "Ho gaya", "%d image save ho gayi." % len(paths))

    def export_ocr_text(self):
        if not HAS_OCR_LIBS:
            self._warn("pytesseract install nahi hai."); return
        paths = self._ordered_paths()
        if not paths:
            self._warn("Pehle koi page scan/import karein."); return
        out, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Text save karein", self._build_filename(".txt"), "Text (*.txt)")
        if not out:
            return
        try:
            chunks = []
            for p in paths:
                with Image.open(p) as im:
                    chunks.append(pytesseract.image_to_string(im, lang="hin+eng"))
            with open(out, "w", encoding="utf-8") as fh:
                fh.write("\n\n----\n\n".join(chunks))
        except Exception as exc:
            if isinstance(exc, pytesseract.TesseractNotFoundError):
                self._warn("Tesseract OCR install nahi hai.")
            else:
                self._warn("Text export fail:\n%s" % traceback.format_exc())
            return
        self._remember_save_dir(out)
        QtWidgets.QMessageBox.information(self, "Ho gaya", "Text save ho gaya:\n%s" % out)

    def _save_ocr_pdf(self, paths, out):
        writer = PdfWriter()
        for p in paths:
            with Image.open(p) as im:
                pdf_bytes = pytesseract.image_to_pdf_or_hocr(im, extension="pdf", lang="hin+eng")
            reader = PdfReader(io.BytesIO(pdf_bytes))
            for pg in reader.pages:
                writer.add_page(pg)
        with open(out, "wb") as fh:
            writer.write(fh)

    # ---- Tools ----
    def split_pdfs(self):
        paths = self._ordered_paths()
        if not paths:
            self._warn("Pehle koi page scan/import karein."); return
        text, ok = QtWidgets.QInputDialog.getText(
            self, "Split PDFs", "Naye document kaun se page se shuru ho? (jaise 1,5,9):", text="1")
        if not ok or not text.strip():
            return
        try:
            starts = sorted(set(int(x) for x in text.replace(" ", "").split(",") if x))
        except Exception:
            self._warn("Galat input."); return
        starts = [s for s in starts if 1 <= s <= len(paths)]
        if not starts or starts[0] != 1:
            starts = [1] + starts
        folder = QtWidgets.QFileDialog.getExistingDirectory(self, "Folder", self._opts.get("save_folder", ""))
        if not folder:
            return
        os.makedirs(folder, exist_ok=True)
        base = sanitize(self.claim_edit.text().strip(), "scan") + "_" + datetime.datetime.now().strftime("%Y-%m-%d")
        ranges = []
        for i, s in enumerate(starts):
            end = (starts[i + 1] - 1) if i + 1 < len(starts) else len(paths)
            ranges.append((s, end))
        try:
            for idx, (s, e) in enumerate(ranges, 1):
                self._pages_as_pdf(paths[s - 1:e], os.path.join(folder, "%s_part%d.pdf" % (base, idx)))
        except Exception:
            self._warn("Split fail:\n%s" % traceback.format_exc()); return
        QtWidgets.QMessageBox.information(self, "Ho gaya", "%d alag PDF ban gayi:\n%s" % (len(ranges), folder))

    def merge_pdfs(self):
        if not HAS_OCR_LIBS:
            self._warn("pypdf install nahi hai."); return
        files, _ = QtWidgets.QFileDialog.getOpenFileNames(self, "Merge ke liye PDF chuno", "", "PDF (*.pdf)")
        if not files:
            return
        out, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Merged PDF save", self._build_filename(".pdf"), "PDF (*.pdf)")
        if not out:
            return
        if not out.lower().endswith(".pdf"):
            out += ".pdf"
        try:
            writer = PdfWriter()
            for f in files:
                r = PdfReader(f)
                for pg in r.pages:
                    writer.add_page(pg)
            with open(out, "wb") as fh:
                writer.write(fh)
        except Exception:
            self._warn("Merge fail:\n%s" % traceback.format_exc()); return
        self._add_recent(out)
        QtWidgets.QMessageBox.information(self, "Ho gaya", "Merged PDF ban gayi:\n%s" % out)

    def search_pdfs(self):
        q, ok = QtWidgets.QInputDialog.getText(self, "Search", "Claim number / naam ka koi hissa:")
        if not ok or not q.strip():
            return
        root = self._opts.get("save_folder", os.path.expanduser("~"))
        matches = []
        try:
            for dp, _dn, fn in os.walk(root):
                for f in fn:
                    if f.lower().endswith(".pdf") and q.lower() in f.lower():
                        matches.append(os.path.join(dp, f))
        except Exception:
            pass
        if not matches:
            QtWidgets.QMessageBox.information(self, "Search", "Kuch nahi mila."); return
        item, ok = QtWidgets.QInputDialog.getItem(
            self, "Search results", "%d file mili — kholne ke liye chuno:" % len(matches),
            matches, 0, False)
        if ok and item:
            self._open_path(item)

    def monthly_report(self):
        month = datetime.datetime.now().strftime("%Y-%m")
        total_files, total_pages = 0, 0
        # Prefer the Excel register if present
        xp = os.path.join(self._opts.get("save_folder", "."), "register.xlsx")
        counted = False
        if HAS_XLSX and os.path.exists(xp):
            try:
                wb = openpyxl.load_workbook(xp); ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and row[0] and str(row[0]).startswith(month):
                        total_files += 1
                        try:
                            total_pages += int(row[3])
                        except Exception:
                            pass
                counted = True
            except Exception:
                counted = False
        if not counted:
            root = self._opts.get("save_folder", os.path.expanduser("~"))
            try:
                for dp, _dn, fn in os.walk(root):
                    for f in fn:
                        if f.lower().endswith(".pdf"):
                            p = os.path.join(dp, f)
                            m = datetime.datetime.fromtimestamp(os.path.getmtime(p)).strftime("%Y-%m")
                            if m == month:
                                total_files += 1
            except Exception:
                pass
        QtWidgets.QMessageBox.information(
            self, "Monthly report (%s)" % month,
            "Is mahine:\n\nKul PDF: %d\nKul pages: %s"
            % (total_files, total_pages if total_pages else "—"))

    def _draw_fit(self, painter, path, target):
        """Draw the image at `path` scaled to fit inside `target` rect, centered,
        keeping aspect ratio."""
        img = QtGui.QImage(path)
        if img.isNull():
            return
        size = img.size()
        size.scale(target.size(), QtCore.Qt.KeepAspectRatio)
        x = target.x() + (target.width() - size.width()) // 2
        y = target.y() + (target.height() - size.height()) // 2
        painter.drawImage(QtCore.QRect(x, y, size.width(), size.height()), img)

    def _do_print(self, paths, per_page=1):
        if not paths:
            self._warn("Pehle koi page scan/import karein (ya select karein)."); return
        printer = QPrinter(QPrinter.HighResolution)
        if QPrintDialog(printer, self).exec_() != QtWidgets.QDialog.Accepted:
            return
        painter = QtGui.QPainter()
        if not painter.begin(printer):
            self._warn("Printer start nahi hua."); return
        try:
            page = painter.viewport()
            if per_page <= 1:
                for i, p in enumerate(paths):
                    if i > 0:
                        printer.newPage()
                    self._draw_fit(painter, p, page)
            else:
                # per_page images stacked on ONE sheet (top/bottom halves for 2-up).
                slot_h = page.height() // per_page
                for i, p in enumerate(paths):
                    slot = i % per_page
                    if i > 0 and slot == 0:
                        printer.newPage()
                    # small inset so the two IDs aren't glued together
                    pad = int(slot_h * 0.04)
                    target = QtCore.QRect(page.x(), page.y() + slot * slot_h + pad,
                                          page.width(), slot_h - 2 * pad)
                    self._draw_fit(painter, p, target)
        finally:
            painter.end()
        self.status.showMessage("Print bhej diya.", 4000)

    def print_all(self):
        self._do_print(self._ordered_paths(), per_page=1)

    def print_selected(self):
        paths = self._selected_paths()
        if not paths:
            self._warn("Pehle thumbnails me se kuch pages select karein."); return
        self._do_print(paths, per_page=1)

    def print_ids(self):
        # ID print: 2 IDs/pages per A4 sheet. Uses selected pages if any, else all.
        paths = self._selected_paths() or self._ordered_paths()
        self._do_print(paths, per_page=2)

    def print_ids_selected(self):
        paths = self._selected_paths()
        if not paths:
            self._warn("Pehle thumbnails me se ID pages select karein."); return
        self._do_print(paths, per_page=2)

    # kept for the Ctrl+P shortcut / File menu (prints all)
    def print_pages(self):
        self.print_all()

    def create_shortcut(self):
        prof = self._selected_profile()
        if prof is None:
            self._warn("Pehle ek profile chuno."); return
        script = os.path.abspath(sys.argv[0])
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        bat = os.path.join(desktop, "Scan - %s.bat" % prof.get("name", "Profile"))
        try:
            with open(bat, "w", encoding="utf-8") as fh:
                fh.write("@echo off\r\n")
                fh.write('py -3.12-32 "%s" --scan --profile "%s"\r\n' % (script, prof.get("name", "Profile")))
            QtWidgets.QMessageBox.information(self, "Ho gaya",
                "Desktop par shortcut ban gaya:\n%s\n\nDouble-click karte hi seedhe scan." % bat)
        except Exception as exc:
            self._warn("Shortcut fail:\n%s" % exc)

    def _update_status(self):
        prof = self._selected_profile()
        pname = prof.get("name") if prof else tr("none_profile", self._lang)
        flags = [k for k in ("auto_save", "batch_mode", "remove_blank", "auto_crop",
                             "deskew", "quality_enhance", "watermark", "compress",
                             "excel_log", "backup") if self._opts.get(k)]
        tail = ("   |   " + ", ".join(flags)) if flags else ""
        self.status.showMessage("%s: %d   |   %s: %s%s" % (tr("st_pages", self._lang), self.list.count(), tr("st_profile", self._lang), pname, tail))

    def _warn(self, msg):
        QtWidgets.QMessageBox.warning(self, APP_NAME, msg)

    def closeEvent(self, event):
        if self.list.count() > 0 and getattr(self, "_dirty", False):
            r = QtWidgets.QMessageBox.question(
                self, APP_NAME,
                "Bina save kiye pages hain. Band karein?" if self._lang == "hi"
                else "You have unsaved pages. Close anyway?",
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No, QtWidgets.QMessageBox.No)
            if r != QtWidgets.QMessageBox.Yes:
                event.ignore(); return
        try:
            self._conn_timer.stop()
        except Exception:
            pass
        try:
            for f in os.listdir(self._tmpdir):
                try:
                    os.remove(os.path.join(self._tmpdir, f))
                except Exception:
                    pass
            os.rmdir(self._tmpdir)
        except Exception:
            pass
        super().closeEvent(event)


def main():
    auto_profile = None
    if "--scan" in sys.argv and "--profile" in sys.argv:
        try:
            auto_profile = sys.argv[sys.argv.index("--profile") + 1]
        except Exception:
            auto_profile = None
    try:
        import ctypes
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("NobleCare.ApneScan")
    except Exception:
        pass
    app = QtWidgets.QApplication(sys.argv)
    app.setApplicationName(APP_NAME)
    app.setWindowIcon(app_icon())
    win = ScannerWindow(auto_scan_profile=auto_profile)
    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
