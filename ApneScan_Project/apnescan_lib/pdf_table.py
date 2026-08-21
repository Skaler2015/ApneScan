"""
pdf_table.py — PDF ki table/sheet ko HUBAHU Excel me nikaalne ke liye.

Do raaste:
  * Digital PDF (asli text) -> PyMuPDF find_tables() se poori grid (rows x cols)
    seedhi milti hai (sabse sahi). Table na mile to word-position se grid banti.
  * Scanned PDF (photo) -> OCR words (caller deta hai) ko GLOBAL column
    clustering se grid me baithate hain, taaki har row ke column aapas me
    seedhe align rahein (per-line gap-splitting se behtar).

openpyxl se .xlsx likhte waqt: borders (table-jaisa), header row bold,
column-width content ke hisaab se, aur alag-alag page/table alag sheet me.
"""

import fitz

try:
    import openpyxl
    from openpyxl.styles import Border, Side, Font, Alignment
    from openpyxl.utils import get_column_letter
    HAS_XLSX = True
except Exception:
    HAS_XLSX = False


# --------------------------------------------------------------------------
# 1) DIGITAL PDF — find_tables()
# --------------------------------------------------------------------------
def find_tables_digital(doc):
    """Har page par tables dhoondo. Lauta: list of dict
    {page, rows: [[cell_or_'', ...], ...]}.  cell None -> "" ."""
    out = []
    for pno in range(doc.page_count):
        page = doc[pno]
        try:
            tf = page.find_tables()
        except Exception:
            tf = None
        tables = getattr(tf, "tables", []) if tf else []
        for tb in tables:
            try:
                grid = tb.extract()          # list of rows, har cell str/None
            except Exception:
                continue
            rows = [[("" if c is None else str(c)).strip() for c in r]
                    for r in grid]
            rows = [r for r in rows if any(c for c in r)]
            if rows:
                out.append({"page": pno, "rows": rows})
    return out


def page_words_digital(doc, pno):
    """Ek page ke saare words position ke saath: (x0,y0,x1,y1,text)."""
    page = doc[pno]
    words = page.get_text("words")           # (x0,y0,x1,y1,word,block,line,wn)
    return [(w[0], w[1], w[2], w[3], w[4]) for w in words if str(w[4]).strip()]


# --------------------------------------------------------------------------
# 2) WORDS -> GRID  (scanned OCR ke liye bhi, digital fallback ke liye bhi)
# --------------------------------------------------------------------------
def _cluster_rows(words, row_tol=None):
    """Words ko y (top) ke hisaab se rows me baanto."""
    if not words:
        return []
    heights = [w[3] - w[1] for w in words]
    med_h = sorted(heights)[len(heights) // 2] if heights else 10.0
    if row_tol is None:
        row_tol = max(4.0, med_h * 0.6)
    ws = sorted(words, key=lambda w: (w[1] + w[3]) / 2.0)
    rows, cur, cur_y = [], [], None
    for w in ws:
        cy = (w[1] + w[3]) / 2.0
        if cur_y is None or abs(cy - cur_y) <= row_tol:
            cur.append(w)
            cur_y = cy if cur_y is None else (cur_y * (len(cur) - 1) + cy) / len(cur)
        else:
            rows.append(cur); cur = [w]; cur_y = cy
    if cur:
        rows.append(cur)
    return rows


def _median(xs, default=0.0):
    xs = sorted(xs)
    return xs[len(xs) // 2] if xs else default


def _column_cuts(rows_w):
    """GLOBAL column separators — vertical-whitespace projection.
    Ek x tabhi separator maana jaata hai jab wo (lagbhag) SAARI rows me
    khaali ho. Isse (a) ek column ka chaura cell agle column ko chhoo raha
    ho to bhi columns alag rehte hain, aur (b) ek cell ke andar do shabdon
    ke beech ki jagah galti se column nahi banti (kyunki dusri rows me wahan
    text hota hai)."""
    words = [w for r in rows_w for w in r]
    if not words:
        return []
    minx = min(w[0] for w in words)
    maxx = max(w[2] for w in words)
    if maxx - minx < 4:
        return []
    med_w = _median([w[2] - w[0] for w in words], 20.0)
    # ek "character" ki motai ka andaaza (text-length / width se)
    char_w = _median(
        [(w[2] - w[0]) / max(1, len(str(w[4]))) for w in words], 8.0)
    pad = max(0.5, char_w * 0.35)            # word ke aas-paas halka buffer
    step = max(1.0, char_w * 0.5)
    row_ivs = [[(w[0] - pad, w[2] + pad) for w in r] for r in rows_w]
    nrows = max(1, len(rows_w))
    # ek x par kitni rows me text hai
    allow = max(0, int(nrows * 0.12))        # itni rows tak "bleed" allow
    # column-gap: ek cell ke andar ki word-spacing se bada hona chahiye
    min_sep = max(4.0, char_w * 1.3)

    def cover_count(x):
        c = 0
        for ivs in row_ivs:
            for a, b in ivs:
                if a <= x <= b:
                    c += 1
                    break
        return c

    cuts = []
    run = []
    x = minx
    while x <= maxx:
        if cover_count(x) <= allow:
            run.append(x)
        else:
            if run and run[0] > minx and run[-1] < maxx and \
               (run[-1] - run[0] + step) >= min_sep:
                cuts.append((run[0] + run[-1]) / 2.0)
            run = []
        x += step
    # aakhri run (agar maxx tak khaali gaya) ko cut na maano
    return cuts


def _col_index(x_center, cuts):
    i = 0
    for c in cuts:
        if x_center > c:
            i += 1
        else:
            break
    return i


def words_to_grid(words, col_gap=None, row_tol=None):
    """words = list of (x0,y0,x1,y1,text) -> 2D grid (list of rows of str).
    GLOBAL columns (poori table ke liye ek hi column-boundary set) taaki
    har row ke cell seedhe align rahein."""
    if not words:
        return []
    rows_w = _cluster_rows(words, row_tol=row_tol)
    cuts = _column_cuts(rows_w)
    ncols = len(cuts) + 1
    grid = []
    for rw in rows_w:
        cells = [[] for _ in range(ncols)]
        for w in sorted(rw, key=lambda t: t[0]):
            xc = (w[0] + w[2]) / 2.0
            ci = _col_index(xc, cuts)
            cells[ci].append((w[0], w[4]))
        row = [" ".join(t[1] for t in sorted(c)) for c in cells]
        if any(v.strip() for v in row):
            grid.append(row)
    return grid


# --------------------------------------------------------------------------
# 3) GRID(s) -> .xlsx
# --------------------------------------------------------------------------
_THIN = None


def _border():
    global _THIN
    if _THIN is None:
        s = Side(style="thin", color="B0B0B0")
        _THIN = Border(left=s, right=s, top=s, bottom=s)
    return _THIN


def _autofit(ws, grid, col_off=1):
    widths = {}
    for row in grid:
        for ci, val in enumerate(row):
            ln = max((len(x) for x in str(val).split("\n")), default=0)
            widths[ci] = max(widths.get(ci, 0), ln)
    for ci, wd in widths.items():
        ws.column_dimensions[get_column_letter(ci + col_off)].width = \
            min(60, max(8, wd + 2))


def write_tables_xlsx(tables, out, header_bold=True, borders=True):
    """tables = list of {page, rows} (ya sirf {rows}).  Har table ek sheet.
    Ek hi structure PDF jaisa: rows/cols hubahu, borders, header bold."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    if not tables:
        wb.create_sheet("Sheet1")
        wb.save(out); return out
    used = set()
    for ti, tb in enumerate(tables, 1):
        rows = tb.get("rows", [])
        pg = tb.get("page")
        name = ("Page %d" % (pg + 1)) if pg is not None else ("Table %d" % ti)
        if len(tables) > 1 and pg is not None:
            name = "P%d-T%d" % (pg + 1, ti)
        name = name[:31] or ("Sheet%d" % ti)
        base = name; k = 2
        while name in used:
            name = ("%s_%d" % (base, k))[:31]; k += 1
        used.add(name)
        ws = wb.create_sheet(name)
        ncols = max((len(r) for r in rows), default=1)
        for ri, row in enumerate(rows, 1):
            for ci in range(ncols):
                val = row[ci] if ci < len(row) else ""
                cell = ws.cell(row=ri, column=ci + 1, value=val)
                if borders:
                    cell.border = _border()
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                if header_bold and ri == 1:
                    cell.font = Font(bold=True)
        _autofit(ws, rows)
        if rows:
            ws.freeze_panes = "A2"
    wb.save(out)
    return out
